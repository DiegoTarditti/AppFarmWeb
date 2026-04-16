import os
import re
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_cors import CORS
from werkzeug.utils import secure_filename
import database
from database import init_db, Claim, BarcodeMapping, InvoiceBatch, DescuentoCampana, DescuentoModulo, DescuentoModuloItem, Pedido, PedidoItem, Producto, Laboratorio, ModuloPack, Modulo, ErpStock
from data_extract import (extract_provider_name_from_pdf, parse_invoice_pdf,
                          parse_erp_excel, compare_invoice_vs_erp,
                          save_invoice_to_db, save_erp_to_db, save_differences,
                          get_saved_differences, get_erp_items_with_issues,
                          save_barcode_mapping, create_claim, complete_claim)

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
PARSERS_FOLDER = os.path.join(os.path.dirname(__file__), 'parsers')
ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls'}

app = Flask(__name__)
CORS(app)
app.secret_key = os.environ.get('SECRET_KEY', 'supersecretkey')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['TEMPLATES_AUTO_RELOAD'] = True

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.before_request
def bloquear_descuentos():
    from flask import request, abort
    if request.path.startswith('/descuentos'):
        abort(404)

DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:///farmacia.db')
init_db(DATABASE_URL)


@app.template_filter('abs')
def abs_filter(value):
    return abs(value)

@app.template_filter('arg_currency')
def arg_currency(value):
    """Formatea un número como moneda argentina: 1234567.89 → 1.234.567,89"""
    try:
        value = float(value)
    except (TypeError, ValueError):
        return '—'
    int_part, dec_part = f'{value:.2f}'.split('.')
    # Separar miles con punto
    int_formatted = ''
    for i, ch in enumerate(reversed(int_part)):
        if i and i % 3 == 0:
            int_formatted = '.' + int_formatted
        int_formatted = ch + int_formatted
    return f'{int_formatted},{dec_part}'


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_providers():
    session = database.SessionLocal()
    providers = session.query(database.Provider).order_by(database.Provider.razon_social).all()
    result = [{'id': p.id, 'razon_social': p.razon_social, 'cuit': p.cuit or '',
               'parser_file': p.parser_file or '',
               'ruta_facturas': p.ruta_facturas or '',
               'grabar_productos': p.grabar_productos if p.grabar_productos is not None else 1} for p in providers]
    session.close()
    return result


def _make_parser_slug(name):
    """'DROGUERÍA EJEMPLO S.A.' → 'droguer_a_ejemplo_s_a'"""
    return re.sub(r'[^a-z0-9]+', '_', name.lower()).strip('_')


def _ensure_parser_file(parser_name, razon_social, cuit=''):
    """Crea el archivo parser desde la plantilla si no existe."""
    parser_path = os.path.join(PARSERS_FOLDER, f'{parser_name}.py')
    if not os.path.exists(parser_path):
        with open(os.path.join(PARSERS_FOLDER, '_template.py'), encoding='utf-8') as f:
            template = f.read()
        content = (template
                   .replace('{{RAZON_SOCIAL}}', razon_social)
                   .replace('{{CUIT}}', cuit))
        with open(parser_path, 'w', encoding='utf-8') as f:
            f.write(content)


def _get_or_create_provider_by_name(razon_social, cuit='', parser_name=''):
    session = database.SessionLocal()
    provider = None
    if cuit:
        provider = session.query(database.Provider).filter_by(cuit=cuit).first()
    if not provider:
        from sqlalchemy import func
        provider = session.query(database.Provider).filter(
            func.lower(database.Provider.razon_social) == razon_social.lower()
        ).first()
    if not provider:
        provider = database.Provider(razon_social=razon_social,
                                     cuit=cuit or None,
                                     parser_file=parser_name)
        session.add(provider)
        session.commit()
    elif not provider.parser_file and parser_name:
        provider.parser_file = parser_name
        session.commit()
    provider_id = provider.id
    parser_file = provider.parser_file
    session.close()
    return provider_id, parser_file


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

def get_config():
    session = database.SessionLocal()
    cfg = session.get(database.Config, 1)
    if not cfg:
        cfg = database.Config(id=1, farmacia_nombre='Farmacia', ruta_facturas='')
        session.add(cfg)
        session.commit()
    result = {
        'farmacia_nombre': cfg.farmacia_nombre,
        'ruta_facturas': cfg.ruta_facturas or '',
        'umbral_pico': float(cfg.umbral_pico or 1.30),
        'umbral_baja': float(cfg.umbral_baja or 0.70),
        'umbral_tendencia': float(cfg.umbral_tendencia or 0.20),
        'rot_alta_min': float(cfg.rot_alta_min or 20.0),
        'rot_alta_tol': float(cfg.rot_alta_tol or 0.0),
        'rot_media_min': float(cfg.rot_media_min or 5.0),
        'rot_media_tol': float(cfg.rot_media_tol or 0.0),
        'rot_baja_tol': float(cfg.rot_baja_tol or 0.0),
    }
    session.close()
    return result


# ─── HELPERS PRODUCTOS ───────────────────────────────────────────────────────

def _find_producto(session, codigo_barra):
    """Busca un producto por código principal o cualquier alternativo (alt1/2/3)."""
    from sqlalchemy import or_
    bc = str(codigo_barra).strip()
    return session.query(Producto).filter(
        or_(
            Producto.codigo_barra == bc,
            Producto.codigo_barra_alt1 == bc,
            Producto.codigo_barra_alt2 == bc,
            Producto.codigo_barra_alt3 == bc,
        )
    ).first()


def _upsert_producto(session, codigo_barra, descripcion, precio_pvp=None, laboratorio_id=None, fecha_compra=None):
    """Crea o actualiza un producto en la tabla productos."""
    if not codigo_barra:
        return
    codigo_barra = str(codigo_barra).strip()
    prod = _find_producto(session, codigo_barra)
    if prod:
        if descripcion and not prod.descripcion:
            prod.descripcion = str(descripcion).strip()
        if precio_pvp and float(precio_pvp) > 0:
            prod.precio_pvp = precio_pvp
        if laboratorio_id and not prod.laboratorio_id:
            prod.laboratorio_id = laboratorio_id
        if fecha_compra and (not prod.ultima_compra or fecha_compra > prod.ultima_compra):
            prod.ultima_compra = fecha_compra
        from datetime import datetime as _dt; prod.actualizado_en = _dt.utcnow()
    else:
        session.add(Producto(
            codigo_barra=codigo_barra,
            descripcion=str(descripcion).strip() if descripcion else '',
            precio_pvp=precio_pvp,
            laboratorio_id=laboratorio_id,
            ultima_compra=fecha_compra,
        ))


def _add_alt_barcode(session, codigo_barra_erp, codigo_barra_alt):
    """Agrega un código alternativo al producto ERP si no está ya registrado."""
    if not codigo_barra_erp or not codigo_barra_alt:
        return
    codigo_barra_erp = str(codigo_barra_erp).strip()
    codigo_barra_alt = str(codigo_barra_alt).strip()
    if codigo_barra_erp == codigo_barra_alt:
        return
    prod = session.query(Producto).filter_by(codigo_barra=codigo_barra_erp).first()
    if not prod:
        return
    existing = {prod.codigo_barra_alt1, prod.codigo_barra_alt2, prod.codigo_barra_alt3}
    if codigo_barra_alt in existing:
        return
    if not prod.codigo_barra_alt1:
        prod.codigo_barra_alt1 = codigo_barra_alt
    elif not prod.codigo_barra_alt2:
        prod.codigo_barra_alt2 = codigo_barra_alt
    elif not prod.codigo_barra_alt3:
        prod.codigo_barra_alt3 = codigo_barra_alt


# ─────────────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html', config=get_config())


@app.route('/ingresos')
def ingresos():
    return render_template('ingresos.html', providers=get_providers(), config=get_config())


@app.route('/settings')
def settings():
    return render_template('settings.html', config=get_config())


@app.route('/admin')
def admin_console():
    session = database.SessionLocal()
    try:
        stats = {
            'proveedores': session.query(database.Provider).count(),
            'facturas': session.query(database.Invoice).count(),
            'factura_items': session.query(database.InvoiceItem).count(),
            'reclamos': session.query(database.Claim).count(),
            'productos': session.query(database.Producto).count(),
            'pedidos': session.query(database.Pedido).count(),
            'erp_stock': session.query(database.ErpStock).count(),
            'barcode_mappings': session.query(database.BarcodeMapping).count(),
            'modulos': session.query(database.Modulo).count(),
            'modulo_packs': session.query(database.ModuloPack).count(),
        }
    finally:
        session.close()

    deploy = {
        'commit':  os.environ.get('RENDER_GIT_COMMIT', '')[:7] or 'local',
        'branch':  os.environ.get('RENDER_GIT_BRANCH', 'local'),
        'service': os.environ.get('RENDER_SERVICE_NAME', 'local'),
        'url':     os.environ.get('RENDER_EXTERNAL_URL', ''),
    }
    return render_template('admin.html', stats=stats, deploy=deploy)


@app.route('/admin/backup', methods=['POST'])
def admin_backup():
    import subprocess
    from datetime import datetime as _dt

    db_url = (request.form.get('db_url') or '').strip()
    if not db_url:
        flash('Falta la URL de la base de datos.')
        return redirect(url_for('admin_console'))
    if not db_url.startswith(('postgres://', 'postgresql://')):
        flash('URL inválida (debe empezar con postgres:// o postgresql://).')
        return redirect(url_for('admin_console'))

    cmd = ['pg_dump', '--no-owner', '--no-privileges', '--clean',
           '--if-exists', db_url]
    result = subprocess.run(cmd, capture_output=True)

    if result.returncode != 0:
        err = result.stderr.decode('utf-8', errors='replace').strip()
        flash(f'pg_dump falló (exit {result.returncode}): {err[:500]}')
        return redirect(url_for('admin_console'))

    ts = _dt.now().strftime('%Y%m%d_%H%M%S')
    filename = f'farmacia_backup_{ts}.sql'
    resp = make_response(result.stdout)
    resp.headers['Content-Type'] = 'application/sql'
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@app.route('/settings', methods=['POST'])
def settings_save():
    nombre = request.form.get('farmacia_nombre', '').strip() or 'Farmacia'
    ruta = request.form.get('ruta_facturas', '').strip()
    session = database.SessionLocal()
    cfg = session.get(database.Config, 1)
    if not cfg:
        cfg = database.Config(id=1)
        session.add(cfg)
    cfg.farmacia_nombre = nombre
    cfg.ruta_facturas = ruta or None
    try:
        cfg.umbral_pico = max(1.01, min(3.0, float(request.form.get('umbral_pico', 1.30))))
        cfg.umbral_baja = max(0.01, min(0.99, float(request.form.get('umbral_baja', 0.70))))
        cfg.umbral_tendencia = max(0.0, min(5.0, float(request.form.get('umbral_tendencia', 0.20))))
        cfg.rot_alta_min = max(0.0, float(request.form.get('rot_alta_min', 20.0)))
        cfg.rot_alta_tol = max(0.0, float(request.form.get('rot_alta_tol', 0.0)))
        cfg.rot_media_min = max(0.0, float(request.form.get('rot_media_min', 5.0)))
        cfg.rot_media_tol = max(0.0, float(request.form.get('rot_media_tol', 0.0)))
        cfg.rot_baja_tol = max(0.0, float(request.form.get('rot_baja_tol', 0.0)))
    except (ValueError, TypeError):
        pass
    session.commit()
    session.close()
    flash('Configuración guardada.')
    return redirect(url_for('settings'))


@app.route('/api/provider/<int:provider_id>/invoices')
def api_provider_invoices(provider_id):
    session = database.SessionLocal()
    provider = session.get(database.Provider, provider_id)
    if not provider:
        session.close()
        return jsonify([])
    invoices = (session.query(database.Invoice)
                .filter(
                    (database.Invoice.proveedor_cuit == provider.cuit) |
                    (database.Invoice.proveedor_razon == provider.razon_social)
                )
                .order_by(database.Invoice.fecha.desc())
                .limit(50).all())
    result = []
    for inv in invoices:
        result.append({
            'id': inv.id,
            'numero_factura': inv.numero_factura,
            'fecha': inv.fecha.strftime('%d/%m/%Y') if inv.fecha else '—',
            'tipo_comprobante': inv.tipo_comprobante,
            'total_articulos': inv.total_articulos or 0,
            'total': float(inv.total or 0),
        })
    session.close()
    return jsonify(result)


@app.route('/provider/peek', methods=['POST'])
def provider_peek():
    """Recibe el PDF, lee el encabezado y devuelve nombre propuesto + provider_id si ya existe."""
    pdf_file = request.files.get('invoice_pdf')
    if not pdf_file or not allowed_file(pdf_file.filename):
        return jsonify({'error': 'Archivo PDF inválido.'}), 400

    filename = secure_filename(pdf_file.filename)
    path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    pdf_file.save(path)

    proposed_name = extract_provider_name_from_pdf(path)

    # Buscar si ya existe un proveedor con ese nombre
    provider_id = None
    if proposed_name:
        session = database.SessionLocal()
        existing = session.query(database.Provider).filter(
            database.Provider.razon_social.ilike(f'%{proposed_name}%')
        ).first()
        if existing and existing.parser_file:
            provider_id = existing.id
        session.close()

    return jsonify({'proposed_name': proposed_name, 'pdf_filename': filename,
                    'provider_id': provider_id})


@app.route('/api/invoice/probe-create', methods=['POST'])
def invoice_probe_create():
    """Crea una factura mínima (sin ítems, sin ERP) para abrir el asistente de parsing."""
    data = request.get_json(silent=True) or {}
    provider_id = data.get('provider_id')
    pdf_filename = data.get('pdf_filename', '').strip()
    if not pdf_filename:
        return jsonify({'error': 'pdf_filename requerido'}), 400

    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
    if not os.path.exists(pdf_path):
        return jsonify({'error': 'PDF no encontrado en el servidor'}), 404

    # Intentar obtener parser del proveedor
    parser_file = None
    if provider_id:
        _s = database.SessionLocal()
        prov = _s.get(database.Provider, int(provider_id))
        _s.close()
        if prov:
            parser_file = prov.parser_file

    # Ejecutar parser para obtener datos del encabezado (aunque no haya ítems)
    invoice_data = {
        'numero_factura': 'SIN_NUMERO', 'fecha': __import__('datetime').date.today(),
        'proveedor_razon': 'NUEVO PROVEEDOR', 'proveedor_cuit': None,
        'proveedor_domicilio': None, 'total': 0.0, 'total_articulos': 0, 'items': []
    }
    if parser_file:
        try:
            parsed = parse_invoice_pdf(pdf_path, parser_file)
            invoice_data = {**parsed, 'items': []}
        except Exception:
            pass

    _s = database.SessionLocal()
    try:
        inv = save_invoice_to_db(_s, invoice_data,
                                 pdf_filename=pdf_filename, tipo_comprobante='FAC')
        _s.commit()
        inv_id = inv.id
    except Exception as e:
        _s.close()
        return jsonify({'error': str(e)}), 500
    _s.close()
    return jsonify({'invoice_id': inv_id})


@app.route('/provider/create-from-peek', methods=['POST'])
def provider_create_from_peek():
    """Crea o recupera un proveedor desde el flujo peek/batch."""
    data = request.get_json(silent=True) or {}
    name = (data.get('provider_name') or '').strip()
    peek_id = data.get('peek_provider_id')
    if not name:
        return jsonify({'error': 'Nombre requerido.'}), 400

    session = database.SessionLocal()
    try:
        # 1) Si el peek ya identificó un proveedor existente, usarlo directamente
        if peek_id:
            prov = session.get(database.Provider, int(peek_id))
            if prov:
                return jsonify({'provider_id': prov.id})

        # 2) Buscar por nombre con wildcard (por si el usuario editó levemente el nombre)
        existing = session.query(database.Provider).filter(
            database.Provider.razon_social.ilike(f'%{name}%')
        ).first()
        if existing:
            return jsonify({'provider_id': existing.id})

        # 3) Crear nuevo
        prov = database.Provider(razon_social=name)
        session.add(prov)
        session.commit()
        session.refresh(prov)
        pid = prov.id
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

    return jsonify({'provider_id': pid})


@app.route('/provider/<int:provider_id>/parser-preview/export', methods=['POST'])
def provider_parser_preview_export(provider_id):
    """Recibe JSON con datos del preview y devuelve XLS."""
    import io
    import openpyxl
    data = request.get_json(silent=True) or {}
    items = data.get('items', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Parser preview'
    ws.append(['Código de barra', 'Descripción', 'Cantidad', 'Precio unitario',
               'Dto %', 'Importe', 'Lote', 'Vencimiento'])
    for it in items:
        ws.append([it.get('codigo_barra'), it.get('descripcion'), it.get('cantidad'),
                   it.get('precio_unitario'), it.get('dto'), it.get('importe'),
                   it.get('lote'), it.get('vencimiento')])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="parser_preview_{provider_id}.xlsx"'
    return resp


@app.route('/provider/<int:provider_id>/parser-preview-saved', methods=['POST'])
def provider_parser_preview_saved(provider_id):
    """Preview del parser usando un PDF ya guardado en uploads (por nombre de archivo)."""
    data = request.get_json(silent=True) or {}
    pdf_filename = data.get('pdf_filename', '').strip()
    if not pdf_filename:
        return jsonify({'error': 'Falta pdf_filename.'}), 400

    session = database.SessionLocal()
    provider = session.get(database.Provider, provider_id)
    session.close()
    if not provider or not provider.parser_file:
        return jsonify({'error': 'El proveedor no tiene parser configurado.'}), 400

    path = os.path.join(UPLOAD_FOLDER, secure_filename(pdf_filename))
    if not os.path.exists(path):
        return jsonify({'error': 'El archivo PDF ya no está disponible. Volvé a cargarlo.'}), 404

    try:
        result = parse_invoice_pdf(path, provider.parser_file)
    except Exception as e:
        return jsonify({'error': f'Error en el parser: {e}'}), 500

    items = [{'codigo_barra': it.get('codigo_barra') or '',
              'descripcion': it.get('descripcion') or '',
              'cantidad': it.get('cantidad') or '',
              'precio_unitario': it.get('precio_unitario') or '',
              'dto': it.get('dto') or '',
              'importe': it.get('importe') or '',
              'lote': it.get('lote') or '',
              'vencimiento': it.get('vencimiento') or ''}
             for it in (result.get('items') or [])]

    return jsonify({'numero_factura': result.get('numero_factura'),
                    'fecha': str(result.get('fecha') or ''),
                    'proveedor': result.get('proveedor_razon'),
                    'total': str(result.get('total') or ''),
                    'items': items})


def process_upload():
    is_new = request.form.get('is_new_provider') == '1'
    erp_file = request.files.get('erp_excel')

    if not erp_file or not allowed_file(erp_file.filename):
        return {'error': 'Por favor cargue un archivo de informe ERP Excel válido.'}, 400

    # --- Resolver proveedor y ruta del PDF ---
    if is_new:
        razon_social = request.form.get('provider_name_new', '').strip()
        pdf_filename = request.form.get('pdf_temp_filename', '').strip()
        if not razon_social:
            return {'error': 'El nombre del proveedor es obligatorio.'}, 400
        if not pdf_filename:
            return {'error': 'El archivo PDF no está disponible. Intentá de nuevo.'}, 400

        invoice_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
        if not os.path.exists(invoice_path):
            return {'error': 'El archivo PDF ya no está disponible. Intentá de nuevo.'}, 400

        parser_name = _make_parser_slug(razon_social)
        _ensure_parser_file(parser_name, razon_social)
        proveedor_id, parser_file = _get_or_create_provider_by_name(razon_social,
                                                                     parser_name=parser_name)
    else:
        proveedor_id = request.form.get('proveedor_id')
        if not proveedor_id:
            return {'error': 'Seleccioná un proveedor antes de cargar los archivos.'}, 400

        invoice_file = request.files.get('invoice_pdf')
        if not invoice_file or not allowed_file(invoice_file.filename):
            return {'error': 'Por favor cargue un archivo de factura PDF válido.'}, 400

        session = database.SessionLocal()
        provider = session.get(database.Provider, int(proveedor_id))
        session.close()

        if not provider:
            return {'error': 'Proveedor no encontrado.'}, 400
        if not provider.parser_file:
            return {'error': f'El proveedor "{provider.razon_social}" no tiene parser configurado.'}, 400

        parser_file = provider.parser_file
        invoice_filename = secure_filename(invoice_file.filename)
        invoice_path = os.path.join(app.config['UPLOAD_FOLDER'], invoice_filename)
        invoice_file.save(invoice_path)

    # --- Guardar ERP ---
    erp_filename = secure_filename(erp_file.filename)
    erp_path = os.path.join(app.config['UPLOAD_FOLDER'], erp_filename)
    erp_file.save(erp_path)

    # --- Procesar ---
    try:
        invoice_data = parse_invoice_pdf(invoice_path, parser_file)
    except Exception as e:
        return {'error': f'Error al leer el PDF de factura: {e}'}, 400

    # ERP siempre se parsea (incluso si hay 0 ítems en la factura)
    try:
        erp_data = parse_erp_excel(erp_path)
    except Exception as e:
        return {'error': f'Error al leer el Excel ERP: {e}. Asegurate de subir un archivo .xlsx válido.'}, 400

    # 0 ítems → guardar encabezado + ERP y redirigir al asistente de parsing
    if not invoice_data.get('items'):
        _session = database.SessionLocal()
        try:
            _tipo = request.form.get('tipo_comprobante', 'FAC').upper()
            if _tipo not in ('FAC', 'NCR'):
                _tipo = 'FAC'
            _inv = save_invoice_to_db(_session, {**invoice_data, 'items': []},
                                      pdf_filename=os.path.basename(invoice_path),
                                      tipo_comprobante=_tipo)
            _inv.erp_filename = erp_filename
            _session.commit()
            save_erp_to_db(_session, erp_data)
            _invoice_id = _inv.id
        except Exception as e:
            _session.close()
            return {'error': f'Error al guardar encabezado: {e}'}, 400
        _session.close()
        return {'parse_failed': True, 'invoice_id': _invoice_id}, 202

    session = database.SessionLocal()
    try:
        tipo_comprobante = request.form.get('tipo_comprobante', 'FAC').upper()
        if tipo_comprobante not in ('FAC', 'NCR'):
            tipo_comprobante = 'FAC'
        invoice = save_invoice_to_db(session, invoice_data,
                                     pdf_filename=os.path.basename(invoice_path),
                                     tipo_comprobante=tipo_comprobante)
        invoice.erp_filename = erp_filename
        session.commit()
        save_erp_to_db(session, erp_data)
        differences = compare_invoice_vs_erp(session, invoice.id)
        save_differences(session, invoice.id, differences)
        # Poblar tabla productos desde ERP e ítems de factura
        try:
            for erp in session.query(database.ErpStock).all():
                _upsert_producto(session, erp.codigo_barra, erp.descripcion,
                                 float(erp.precio_unitario) if erp.precio_unitario else None)
            _prov = session.get(database.Provider, int(proveedor_id)) if proveedor_id else None
            if not _prov or _prov.grabar_productos != 0:
                for it in session.query(database.InvoiceItem).filter_by(factura_id=invoice.id).all():
                    _upsert_producto(session, it.codigo_barra, it.descripcion,
                                     fecha_compra=invoice.fecha)
            session.commit()
        except Exception:
            session.rollback()
        saved_differences = get_saved_differences(session, invoice.id)
    except Exception as e:
        session.close()
        return {'error': f'Error al procesar los datos: {e}'}, 400
    session.close()

    return {
        'invoice': invoice,
        'differences': [
            {
                'id': d.id,
                'codigo_barra': d.codigo_barra,
                'descripcion': d.descripcion,
                'cantidad_factura': d.cantidad_factura,
                'cantidad_erp': d.cantidad_erp,
                'diferencia': d.diferencia,
                'observaciones': d.observaciones,
            }
            for d in saved_differences
        ]
    }, 200


@app.route('/upload', methods=['POST'])
def upload_files():
    result, status = process_upload()
    if status == 202 and result.get('parse_failed'):
        flash('El parser no detectó artículos. Usá el asistente para extraerlos.', 'warning')
        return redirect(url_for('parse_helper', invoice_id=result['invoice_id']))
    if status != 200:
        flash(result['error'])
        return redirect(url_for('index'))
    return redirect(url_for('compare_view', invoice_id=result['invoice'].id))


@app.route('/api/upload', methods=['POST'])
def upload_files_api():
    result, status = process_upload()
    if status != 200:
        return jsonify(result), status
    return jsonify({
        'invoice': {
            'id': result['invoice'].id,
            'numero_factura': result['invoice'].numero_factura,
            'fecha': str(result['invoice'].fecha),
            'proveedor_razon': result['invoice'].proveedor_razon,
            'total': float(result['invoice'].total or 0),
            'total_articulos': result['invoice'].total_articulos,
        },
        'differences': result['differences']
    }), 200


# ── Asistente de parsing (0 ítems detectados) ────────────────────────────────

@app.route('/invoice/<int:invoice_id>/parse-helper')
def parse_helper(invoice_id):
    import pdfplumber as _plumber
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    session.close()
    if not invoice:
        flash('Factura no encontrada.')
        return redirect(url_for('index'))
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], invoice.pdf_filename or '')
    pdf_text = ''
    if os.path.exists(pdf_path):
        with _plumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                pdf_text += (page.extract_text() or '') + '\n\n'
    return render_template('invoice_parse_helper.html', invoice=invoice, pdf_text=pdf_text)


@app.route('/invoice/<int:invoice_id>/auto-table', methods=['POST'])
def auto_table(invoice_id):
    import pdfplumber as _plumber
    from collections import defaultdict as _dd
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    session.close()
    if not invoice:
        return jsonify({'error': 'Factura no encontrada'}), 404
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], invoice.pdf_filename or '')
    if not os.path.exists(pdf_path):
        return jsonify({'error': 'PDF no encontrado'}), 404

    tables = []
    with _plumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for tbl in (page.extract_tables() or []):
                if tbl and len(tbl) > 2:
                    tables.append(tbl)

    if tables:
        best = max(tables, key=lambda t: len(t))
        return jsonify({'source': 'table', 'rows': best})

    # Fallback: word-based Y-grouping
    all_words = []
    with _plumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            all_words.extend(page.extract_words(x_tolerance=3, y_tolerance=3) or [])

    rows_dict = _dd(list)
    for w in all_words:
        y_key = round(w['top'] / 4) * 4
        rows_dict[y_key].append(w)

    word_rows = []
    for y in sorted(rows_dict.keys()):
        row = sorted(rows_dict[y], key=lambda w: w['x0'])
        word_rows.append([w['text'] for w in row])

    return jsonify({'source': 'words', 'rows': word_rows[:80]})


@app.route('/invoice/<int:invoice_id>/map-columns', methods=['GET', 'POST'])
def map_columns(invoice_id):
    import pdfplumber as _plumber
    import json as _json
    from collections import defaultdict as _dd

    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)

    if request.method == 'POST':
        rows_json = request.form.get('rows_json', '[]')
        mapping = {
            'codigo_barra':    int(request.form.get('col_codigo',  -1)),
            'descripcion':     int(request.form.get('col_desc',    -1)),
            'cantidad':        int(request.form.get('col_cant',    -1)),
            'precio_unitario': int(request.form.get('col_precio',  -1)),
            'dto':             int(request.form.get('col_dto',     -1)),
            'importe':         int(request.form.get('col_importe', -1)),
            'lote':            int(request.form.get('col_lote',    -1)),
        }
        header_row = int(request.form.get('header_row', 0))
        rows = _json.loads(rows_json)

        def _f(s):
            if not s:
                return None
            try:
                return float(str(s).replace('.', '').replace(',', '.'))
            except Exception:
                return None

        def _col(row, idx):
            if idx < 0 or idx >= len(row):
                return None
            v = row[idx]
            return str(v).strip() if v else None

        tipo = invoice.tipo_comprobante or 'FAC'
        sign = -1 if tipo == 'NCR' else 1
        saved = 0
        for i, row in enumerate(rows):
            if i <= header_row:
                continue
            if not any(row):
                continue
            desc    = _col(row, mapping['descripcion'])
            codigo  = _col(row, mapping['codigo_barra'])
            if not desc and not codigo:
                continue
            cant_s  = _col(row, mapping['cantidad'])
            precio  = _f(_col(row, mapping['precio_unitario']))
            dto     = _f(_col(row, mapping['dto']))
            importe = _f(_col(row, mapping['importe']))
            lote    = _col(row, mapping['lote'])
            try:
                cant = int(float(cant_s)) if cant_s else 0
            except Exception:
                cant = 0
            session.add(database.InvoiceItem(
                factura_id=invoice_id, codigo_barra=codigo, descripcion=desc,
                cantidad=cant,
                precio_unitario=sign * precio if precio is not None else None,
                dto=dto,
                importe=sign * importe if importe is not None else None,
                lote=lote,
            ))
            saved += 1

        if saved > 0:
            invoice.total_articulos = saved
            session.commit()
            differences = compare_invoice_vs_erp(session, invoice_id)
            save_differences(session, invoice_id, differences)
            session.close()
            flash(f'{saved} artículos guardados desde el mapeo de columnas.')
            return redirect(url_for('compare_view', invoice_id=invoice_id))
        session.close()
        flash('No se pudieron extraer artículos con esa configuración.')
        return redirect(url_for('map_columns', invoice_id=invoice_id))

    # GET — extraer filas del PDF
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], invoice.pdf_filename or '')
    rows_preview = []
    all_rows = []
    source = 'none'

    if os.path.exists(pdf_path):
        tables = []
        with _plumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                for tbl in (page.extract_tables() or []):
                    if tbl and len(tbl) > 2:
                        tables.append(tbl)
        if tables:
            all_rows = max(tables, key=lambda t: len(t))
            source = 'table'
        else:
            all_words = []
            with _plumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    all_words.extend(page.extract_words(x_tolerance=3, y_tolerance=3) or [])
            rows_dict = _dd(list)
            for w in all_words:
                y_key = round(w['top'] / 4) * 4
                rows_dict[y_key].append(w)
            for y in sorted(rows_dict.keys()):
                row = sorted(rows_dict[y], key=lambda w: w['x0'])
                all_rows.append([w['text'] for w in row])
            source = 'words'
        rows_preview = all_rows[:60]

    session.close()
    num_cols = max((len(r) for r in rows_preview[:10] if r), default=0)
    import json as _j
    return render_template('invoice_map_columns.html', invoice=invoice,
                           rows=rows_preview, rows_json=_j.dumps(all_rows),
                           source=source, num_cols=num_cols)


# ── Convertidor de documentos (herramienta genérica) ─────────────────────────

CONVERTER_DIR = os.path.join(UPLOAD_FOLDER, 'converter')
os.makedirs(CONVERTER_DIR, exist_ok=True)


def _build_item_pattern(example_line, selections):
    """Build regex pattern from example_line + selections.
    Content after the first newline in example_line is wrapped in an optional group,
    so rows that fit on a single line still match."""
    import re as _re
    sel = sorted(selections, key=lambda s: s.get('start', 0))

    def _kind(txt):
        t = (txt or '').strip()
        if _re.fullmatch(r'-?\d[\d.,]*', t):
            return r'[\d.,]+'
        if _re.fullmatch(r'\d+', t):
            return r'\d+'
        if _re.fullmatch(r'\S+', t):
            return r'\S+'
        return r'.+?'

    def _norm_literal(s):
        result = ''
        for tok in _re.findall(r'\s+|\S+', s):
            if tok.isspace():
                result += r'\s+'
            else:
                result += _re.escape(tok)
        return result

    first_nl = example_line.find('\n')
    pattern_main = '^'
    pattern_opt = ''
    cursor = 0
    fields = []
    prev_cap = None
    opened_opt = False

    for s in sel:
        start = int(s.get('start', 0))
        end   = int(s.get('end', start))
        literal = example_line[cursor:start]
        cap = _kind(example_line[start:end])

        # Decide if this capture belongs to optional continuation
        in_opt_zone = first_nl >= 0 and start >= first_nl

        if in_opt_zone and not opened_opt:
            # Split literal: pre-newline goes to main, rest opens optional group
            if first_nl > cursor:
                pre = example_line[cursor:first_nl]
                if pre:
                    pattern_main += _norm_literal(pre)
            pattern_opt = r'(?:'
            # Remaining literal after newline
            rest = example_line[max(cursor, first_nl):start]
            if rest:
                pattern_opt += _norm_literal(rest)
            opened_opt = True
        else:
            target = pattern_opt if opened_opt else pattern_main
            if literal:
                norm = _norm_literal(literal)
                if literal.strip() == '' and prev_cap == r'.+?' and cap in (r'[\d.,]+', r'\d+', r'\S+'):
                    norm = r'\s*'
                if opened_opt:
                    pattern_opt += norm
                else:
                    pattern_main += norm

        if opened_opt:
            pattern_opt += '(' + cap + ')'
        else:
            pattern_main += '(' + cap + ')'
        fields.append(s.get('field'))
        prev_cap = cap
        cursor = end

    tail = example_line[cursor:]
    if tail.strip():
        # Tail with text: part before newline goes to main, part after goes to optional
        if first_nl >= cursor and not opened_opt:
            pre = example_line[cursor:first_nl]
            if pre:
                pattern_main += _norm_literal(pre)
            rest = example_line[first_nl:]
            if rest.strip():
                pattern_opt = r'(?:' + _norm_literal(rest)
                opened_opt = True
        else:
            if opened_opt:
                pattern_opt += _norm_literal(tail)
            else:
                pattern_main += _norm_literal(tail)

    pattern = pattern_main + (pattern_opt + r')?' if opened_opt else '')

    def _base(name):
        return _re.sub(r'_\d+$', '', name or '')

    base_fields = []
    for f in fields:
        b = _base(f)
        if b not in base_fields:
            base_fields.append(b)

    return pattern, fields, base_fields, _base


def _converter_meta_path(fname):
    return os.path.join(CONVERTER_DIR, fname + '.meta.json')

def _converter_read_meta(fname):
    import json as _json
    p = _converter_meta_path(fname)
    if os.path.exists(p):
        try:
            with open(p, 'r', encoding='utf-8') as fh:
                return _json.load(fh) or {}
        except Exception:
            return {}
    return {}

def _converter_write_meta(fname, data):
    import json as _json
    with open(_converter_meta_path(fname), 'w', encoding='utf-8') as fh:
        _json.dump(data, fh)


@app.route('/converter', methods=['GET'])
def converter_index():
    files = []
    try:
        for fn in sorted(os.listdir(CONVERTER_DIR), reverse=True):
            if fn.endswith('.meta.json'):
                continue
            p = os.path.join(CONVERTER_DIR, fn)
            if os.path.isfile(p):
                meta = _converter_read_meta(fn)
                files.append({'name': fn, 'size': os.path.getsize(p), 'tipo': meta.get('tipo_doc', '')})
    except Exception:
        pass
    return render_template('converter_index.html', files=files[:20])


@app.route('/converter/<token>/meta', methods=['POST'])
def converter_meta(token):
    safe = secure_filename(token)
    path = os.path.join(CONVERTER_DIR, safe)
    if not os.path.exists(path):
        return jsonify({'error': 'Documento no encontrado'}), 404
    body = request.get_json(silent=True) or {}
    tipo = (body.get('tipo_doc') or '').strip()
    meta = _converter_read_meta(safe)
    meta['tipo_doc'] = tipo
    _converter_write_meta(safe, meta)
    return jsonify({'ok': True, 'tipo_doc': tipo})


@app.route('/converter/upload', methods=['POST'])
def converter_upload():
    import uuid as _uuid
    f = request.files.get('document')
    if not f or not f.filename:
        flash('Elegí un archivo.')
        return redirect(url_for('converter_index'))
    ext = os.path.splitext(f.filename)[1].lower() or '.pdf'
    if ext not in ('.pdf',):
        flash('Por ahora sólo PDF.')
        return redirect(url_for('converter_index'))
    token = _uuid.uuid4().hex[:12]
    fname = token + '_' + secure_filename(f.filename)
    f.save(os.path.join(CONVERTER_DIR, fname))
    return redirect(url_for('converter_helper', token=fname))


@app.route('/converter/<token>/auto', methods=['GET'])
def converter_auto(token):
    import pdfplumber as _plumber
    from collections import defaultdict as _dd
    safe = secure_filename(token)
    path = os.path.join(CONVERTER_DIR, safe)
    if not os.path.exists(path):
        flash('Documento no encontrado.')
        return redirect(url_for('converter_index'))

    tables = []
    with _plumber.open(path) as pdf:
        for page in pdf.pages:
            for tbl in (page.extract_tables() or []):
                if tbl and len(tbl) > 2:
                    tables.append(tbl)

    if tables:
        best = max(tables, key=lambda t: len(t))
        rows = best
        source = 'Tabla detectada automáticamente'
    else:
        all_words = []
        with _plumber.open(path) as pdf:
            for page in pdf.pages:
                all_words.extend(page.extract_words(x_tolerance=3, y_tolerance=3) or [])
        rows_dict = _dd(list)
        for w in all_words:
            y_key = round(w['top'] / 4) * 4
            rows_dict[y_key].append(w)
        rows = []
        for y in sorted(rows_dict.keys()):
            row = sorted(rows_dict[y], key=lambda w: w['x0'])
            rows.append([w['text'] for w in row])
        source = 'Palabras agrupadas por posición Y'

    return render_template('converter_auto.html', token=safe, filename=safe, rows=rows, source=source)


@app.route('/converter/<token>/delete', methods=['POST'])
def converter_delete(token):
    safe = secure_filename(token)
    path = os.path.join(CONVERTER_DIR, safe)
    if os.path.exists(path):
        try:
            os.remove(path)
            mp = _converter_meta_path(safe)
            if os.path.exists(mp):
                os.remove(mp)
            flash('Documento eliminado.')
        except Exception as e:
            flash(f'Error al eliminar: {e}')
    return redirect(url_for('converter_index'))


@app.route('/converter/<token>/helper', methods=['GET'])
def converter_helper(token):
    safe = secure_filename(token)
    path = os.path.join(CONVERTER_DIR, safe)
    if not os.path.exists(path):
        flash('Documento no encontrado.')
        return redirect(url_for('converter_index'))
    import pdfplumber as _plumber
    pdf_text = ''
    with _plumber.open(path) as pdf:
        for page in pdf.pages:
            pdf_text += (page.extract_text() or '') + '\n\n'
    meta = _converter_read_meta(safe)
    return render_template('converter_helper.html', token=safe, filename=safe, pdf_text=pdf_text, tipo_doc=meta.get('tipo_doc', ''))


@app.route('/converter/<token>/pick', methods=['GET'])
def converter_pick(token):
    import pdfplumber as _plumber
    safe = secure_filename(token)
    path = os.path.join(CONVERTER_DIR, safe)
    if not os.path.exists(path):
        flash('Documento no encontrado.')
        return redirect(url_for('converter_index'))
    pdf_text = ''
    with _plumber.open(path) as pdf:
        for page in pdf.pages:
            pdf_text += (page.extract_text() or '') + '\n\n'
    return render_template('converter_pick.html', token=safe, pdf_text=pdf_text, filename=safe)


@app.route('/converter/<token>/infer', methods=['POST'])
def converter_infer(token):
    import re as _re
    import pdfplumber as _plumber

    body = request.get_json(silent=True) or {}
    example_line = body.get('example_line', '')
    selections = body.get('selections', [])

    safe = secure_filename(token)
    path = os.path.join(CONVERTER_DIR, safe)
    if not os.path.exists(path):
        return jsonify({'error': 'Documento no encontrado'}), 404
    if not example_line or not selections:
        return jsonify({'error': 'Faltan datos'}), 400

    pattern, fields, base_fields, _base = _build_item_pattern(example_line, selections)

    pdf_text = ''
    with _plumber.open(path) as pdf:
        for page in pdf.pages:
            pdf_text += (page.extract_text() or '') + '\n'

    rx = _re.compile(pattern, _re.MULTILINE)

    rows = []
    for m in rx.finditer(pdf_text):
        row = {b: [] for b in base_fields}
        for i, f in enumerate(fields):
            val = m.group(i+1)
            if val:
                row[_base(f)].append(val)
        rows.append({b: _re.sub(r'\s+', ' ', ' '.join(row[b]).strip()) for b in base_fields})

    return jsonify({'pattern': pattern, 'fields': base_fields, 'rows': rows})


@app.route('/converter/<token>/export', methods=['POST'])
def converter_export(token):
    import io as _io
    import openpyxl as _ox
    body = request.get_json(silent=True) or {}
    rows = body.get('rows', [])
    header = body.get('header', {}) or {}
    fields = body.get('fields', [])
    if not rows:
        return jsonify({'error': 'No hay filas para exportar'}), 400

    wb = _ox.Workbook()
    ws = wb.active
    ws.title = 'Datos'

    r = 1
    if header:
        for k, v in header.items():
            ws.cell(row=r, column=1, value=k)
            ws.cell(row=r, column=2, value=v)
            r += 1
        r += 1

    # items
    cols = fields or (list(rows[0].keys()) if rows else [])
    for ci, c in enumerate(cols, start=1):
        ws.cell(row=r, column=ci, value=c)
    r += 1
    for row in rows:
        for ci, c in enumerate(cols, start=1):
            ws.cell(row=r, column=ci, value=row.get(c, ''))
        r += 1

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file as _sf
    out_name = os.path.splitext(token)[0] + '.xlsx'
    return _sf(buf, as_attachment=True, download_name=out_name,
               mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/invoice/<int:invoice_id>/pick-items', methods=['GET'])
def pick_items(invoice_id):
    import pdfplumber as _plumber
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    session.close()
    if not invoice:
        flash('Factura no encontrada.')
        return redirect(url_for('index'))
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], invoice.pdf_filename or '')
    pdf_text = ''
    if os.path.exists(pdf_path):
        with _plumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                pdf_text += (page.extract_text() or '') + '\n\n'
    return render_template('invoice_pick_items.html', invoice=invoice, pdf_text=pdf_text)


@app.route('/invoice/<int:invoice_id>/pick-items/infer', methods=['POST'])
def pick_items_infer(invoice_id):
    import re as _re
    import pdfplumber as _plumber

    body = request.get_json(silent=True) or {}
    example_line = body.get('example_line', '')
    selections = body.get('selections', [])  # [{field, text, start, end}]

    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    session.close()
    if not invoice:
        return jsonify({'error': 'Factura no encontrada'}), 404

    if not example_line or not selections:
        return jsonify({'error': 'Faltan datos'}), 400

    pattern, fields, base_fields, _base = _build_item_pattern(example_line, selections)

    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], invoice.pdf_filename or '')
    pdf_text = ''
    if os.path.exists(pdf_path):
        with _plumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                pdf_text += (page.extract_text() or '') + '\n'

    rx = _re.compile(pattern, _re.MULTILINE)

    rows = []
    for m in rx.finditer(pdf_text):
        row = {b: [] for b in base_fields}
        for i, f in enumerate(fields):
            val = m.group(i+1)
            if val:
                row[_base(f)].append(val)
        rows.append({b: _re.sub(r'\s+', ' ', ' '.join(row[b]).strip()) for b in base_fields})

    return jsonify({'pattern': pattern, 'fields': base_fields, 'rows': rows})


@app.route('/invoice/<int:invoice_id>/pick-items/save', methods=['POST'])
def pick_items_save(invoice_id):
    import re as _re
    import datetime as _dt
    body = request.get_json(silent=True) or {}
    rows = body.get('rows', [])
    header = body.get('header', {}) or {}

    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    if not invoice:
        session.close()
        return jsonify({'error': 'Factura no encontrada'}), 404

    tipo = invoice.tipo_comprobante or 'FAC'
    sign = -1 if tipo == 'NCR' else 1

    def _f(s):
        if s is None or s == '':
            return None
        try:
            return float(str(s).replace('.', '').replace(',', '.'))
        except Exception:
            return None

    # Header update
    if header.get('razon_social'):
        invoice.proveedor_razon = header['razon_social'].strip()
    if header.get('numero_factura'):
        invoice.numero_factura = header['numero_factura'].strip()
    if header.get('total'):
        t = _f(header['total'])
        if t is not None:
            invoice.total = sign * t
    if header.get('fecha'):
        raw = header['fecha'].strip()
        m = _re.search(r'(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})', raw)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y < 100:
                y += 2000
            try:
                invoice.fecha = _dt.date(y, mo, d)
            except Exception:
                pass

    # Alias map: accept common variations for standard fields
    aliases = {
        'descripcion': ['descripcion', 'concepto', 'detalle', 'producto', 'articulo', 'descripción'],
        'codigo_barra': ['codigo_barra', 'codigo', 'código', 'ean', 'cod'],
        'cantidad': ['cantidad', 'cant', 'qty'],
        'precio_unitario': ['precio_unitario', 'pcio_unit', 'pcio', 'precio', 'unitario'],
        'importe': ['importe', 'total', 'subtotal', 'monto'],
        'dto': ['dto', 'descto', 'descuento'],
        'lote': ['lote'],
    }
    def _pick(r, std):
        for k in aliases.get(std, [std]):
            if r.get(k):
                return r.get(k)
        return None

    saved = 0
    for r in rows:
        desc = str(_pick(r, 'descripcion') or '').strip()[:150]
        if not desc:
            continue
        precio  = _f(_pick(r, 'precio_unitario'))
        importe = _f(_pick(r, 'importe'))
        dto     = _f(_pick(r, 'dto'))
        try:
            cant = int(float(str(_pick(r, 'cantidad') or 0).replace(',', '.')))
        except Exception:
            cant = 0
        session.add(database.InvoiceItem(
            factura_id=invoice_id,
            codigo_barra=(_pick(r, 'codigo_barra') or None),
            descripcion=desc, cantidad=cant,
            precio_unitario=sign * precio if precio is not None else None,
            dto=dto,
            importe=sign * importe if importe is not None else None,
            lote=(_pick(r, 'lote') or None),
        ))
        saved += 1

    try:
        if saved > 0:
            invoice.total_articulos = saved
        session.commit()
        if saved > 0:
            differences = compare_invoice_vs_erp(session, invoice_id)
            save_differences(session, invoice_id, differences)
    except Exception as e:
        session.rollback()
        session.close()
        import traceback as _tb
        return jsonify({'error': str(e), 'trace': _tb.format_exc()}), 500
    session.close()
    return jsonify({'saved': saved, 'redirect': url_for('compare_view', invoice_id=invoice_id)})


@app.route('/invoice/<int:invoice_id>/manual-items', methods=['GET', 'POST'])
def manual_items(invoice_id):
    import pdfplumber as _plumber
    import json as _json

    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)

    if request.method == 'POST':
        items_data = _json.loads(request.form.get('items_json', '[]'))
        tipo = invoice.tipo_comprobante or 'FAC'
        sign = -1 if tipo == 'NCR' else 1

        def _f(s):
            if not s:
                return None
            try:
                return float(str(s).replace('.', '').replace(',', '.'))
            except Exception:
                return None

        saved = 0
        for item in items_data:
            desc = str(item.get('descripcion', '')).strip()
            if not desc:
                continue
            precio  = _f(item.get('precio_unitario'))
            importe = _f(item.get('importe'))
            dto     = _f(item.get('dto'))
            try:
                cant = int(float(item.get('cantidad') or 0))
            except Exception:
                cant = 0
            session.add(database.InvoiceItem(
                factura_id=invoice_id,
                codigo_barra=item.get('codigo_barra') or None,
                descripcion=desc, cantidad=cant,
                precio_unitario=sign * precio if precio is not None else None,
                dto=dto,
                importe=sign * importe if importe is not None else None,
                lote=item.get('lote') or None,
            ))
            saved += 1

        if saved > 0:
            invoice.total_articulos = saved
            session.commit()
            differences = compare_invoice_vs_erp(session, invoice_id)
            save_differences(session, invoice_id, differences)
            session.close()
            flash(f'{saved} artículos guardados manualmente.')
            return redirect(url_for('compare_view', invoice_id=invoice_id))
        session.close()
        flash('No se ingresaron artículos.')
        return redirect(url_for('manual_items', invoice_id=invoice_id))

    # GET
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], invoice.pdf_filename or '')
    pdf_text = ''
    if os.path.exists(pdf_path):
        with _plumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                pdf_text += (page.extract_text() or '') + '\n\n'
    session.close()
    return render_template('invoice_manual_items.html', invoice=invoice, pdf_text=pdf_text)


@app.route('/api/invoice/<int:invoice_id>/differences', methods=['GET'])
def invoice_differences(invoice_id):
    session = database.SessionLocal()
    differences = session.query(database.StockDifference).filter_by(factura_id=invoice_id).all()
    session.close()
    return jsonify([
        {
            'id': d.id, 'codigo_barra': d.codigo_barra, 'descripcion': d.descripcion,
            'cantidad_factura': d.cantidad_factura, 'cantidad_erp': d.cantidad_erp,
            'diferencia': d.diferencia, 'observaciones': d.observaciones,
        }
        for d in differences
    ])


@app.route('/claim/<int:claim_id>')
def view_claim(claim_id):
    session = database.SessionLocal()
    claim = session.get(Claim, claim_id)
    if not claim:
        session.close()
        return 'Reclamo no encontrado', 404
    _ = claim.provider
    _ = claim.items
    session.close()
    return render_template('claim.html', claim=claim)


@app.route('/claim/create', methods=['POST'])
def create_claim_route():
    try:
        invoice_id = int(request.form.get('invoice_id'))
    except (TypeError, ValueError):
        flash('Factura inválida para crear el reclamo.')
        return redirect(url_for('index'))

    selected_ids = request.form.getlist('selected_differences')
    if not selected_ids:
        flash('Seleccione al menos un registro para reclamo.')
        return redirect(request.referrer or url_for('index'))

    session = database.SessionLocal()
    claim = create_claim(session, invoice_id, [int(i) for i in selected_ids])
    _ = claim.provider
    _ = claim.items
    session.close()
    return render_template('claim.html', claim=claim, auto_download=True)


@app.route('/claim/<int:claim_id>/complete', methods=['POST'])
def complete_claim_route(claim_id):
    session = database.SessionLocal()
    claim = complete_claim(session, claim_id)
    session.close()
    if not claim:
        return 'Reclamo no encontrado', 404
    return redirect(url_for('view_claim', claim_id=claim.id))


@app.route('/api/claims', methods=['POST'])
def api_create_claim():
    payload = request.get_json() or {}
    factura_id = payload.get('factura_id')
    difference_ids = payload.get('difference_ids', [])
    if not factura_id or not difference_ids:
        return jsonify({'error': 'factura_id y difference_ids son obligatorios'}), 400

    session = database.SessionLocal()
    try:
        claim = create_claim(session, int(factura_id), [int(i) for i in difference_ids])
    except Exception as exc:
        session.close()
        return jsonify({'error': str(exc)}), 400
    session.close()
    return jsonify({'claim_id': claim.id, 'estado': claim.estado}), 201


@app.route('/invoice/<int:invoice_id>/header', methods=['POST'])
def update_invoice_header(invoice_id):
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    if not invoice:
        session.close()
        flash('Factura no encontrada.')
        return redirect(url_for('index'))
    tipo = (request.form.get('tipo_comprobante') or '').upper()
    ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or (
        request.form.keys() and set(request.form.keys()) == {'tipo_comprobante'}
    )
    if tipo in ('FAC', 'NCR'):
        invoice.tipo_comprobante = tipo
    if 'numero_factura' in request.form:
        invoice.numero_factura = request.form.get('numero_factura', invoice.numero_factura).strip() or invoice.numero_factura
    if 'proveedor_razon' in request.form:
        invoice.proveedor_razon = request.form.get('proveedor_razon', invoice.proveedor_razon).strip() or invoice.proveedor_razon
    session.commit()
    session.close()
    if ajax:
        return jsonify({'ok': True})
    return redirect(url_for('show_results', invoice_id=invoice_id))


@app.route('/results/<int:invoice_id>')
def show_results(invoice_id):
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    saved_differences = get_saved_differences(session, invoice_id)
    differences = [
        {
            'id': d.id, 'codigo_barra': d.codigo_barra, 'descripcion': d.descripcion,
            'cantidad_factura': d.cantidad_factura, 'cantidad_erp': d.cantidad_erp,
            'diferencia': d.diferencia, 'observaciones': d.observaciones,
        }
        for d in saved_differences
    ]
    total_unidades_calc = sum(
        item.cantidad for item in invoice.items if item.cantidad
    ) if invoice else 0
    session.close()
    return render_template('results.html', invoice=invoice, differences=differences,
                           total_unidades_calc=total_unidades_calc)


@app.route('/invoice/<int:invoice_id>/pick-fields', methods=['GET'])
def pick_fields(invoice_id):
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    session.close()
    if not invoice:
        flash('Factura no encontrada.')
        return redirect(url_for('index'))

    pdf_text = ''
    if invoice.pdf_filename:
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], invoice.pdf_filename)
        if os.path.exists(pdf_path):
            import pdfplumber
            with pdfplumber.open(pdf_path) as pdf:
                pdf_text = pdf.pages[0].extract_text() or ''

    return render_template('pick_fields.html', invoice=invoice, pdf_text=pdf_text)


@app.route('/invoice/<int:invoice_id>/pick-fields', methods=['POST'])
def pick_fields_save(invoice_id):
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    if not invoice:
        session.close()
        flash('Factura no encontrada.')
        return redirect(url_for('index'))

    fields = ('numero_factura', 'proveedor_razon', 'proveedor_cuit', 'fecha', 'total')
    for field in fields:
        val = request.form.get(field, '').strip()
        if val:
            setattr(invoice, field, val)
    session.commit()
    session.close()
    return redirect(url_for('show_results', invoice_id=invoice_id))


@app.route('/providers')
def providers_list():
    tipo_filter = (request.args.get('tipo') or '').strip().lower()
    session = database.SessionLocal()
    q = session.query(database.Provider)
    if tipo_filter in ('drogueria', 'laboratorio', 'otro'):
        q = q.filter(database.Provider.tipo == tipo_filter)
    providers = q.order_by(database.Provider.razon_social).all()
    provider_data = []
    for p in providers:
        q = session.query(database.Invoice)
        if p.cuit:
            q = q.filter(
                (database.Invoice.proveedor_cuit == p.cuit) |
                (database.Invoice.proveedor_razon == p.razon_social)
            )
        else:
            q = q.filter(database.Invoice.proveedor_razon == p.razon_social)
        invoice_count = q.count()
        claim_count = session.query(database.Claim).filter_by(proveedor_id=p.id).count()
        provider_data.append({
            'id': p.id,
            'razon_social': p.razon_social,
            'cuit': p.cuit or '',
            'parser_file': p.parser_file or '',
            'ruta_facturas': p.ruta_facturas or '',
            'match_strategy': p.match_strategy,
            'grabar_productos': p.grabar_productos if p.grabar_productos is not None else 1,
            'tipo': p.tipo or 'drogueria',
            'invoice_count': invoice_count,
            'claim_count': claim_count,
        })
    session.close()
    return render_template('providers.html', providers=provider_data, tipo_filter=tipo_filter)


# ─── LABORATORIOS ─────────────────────────────────────────────────────────────

@app.route('/laboratorios')
def laboratorios_list():
    from sqlalchemy import func as _func
    session = database.SessionLocal()
    labs = session.query(Laboratorio).order_by(Laboratorio.nombre).all()
    data = []
    for l in labs:
        prod_count = session.query(_func.count(Producto.id)).filter_by(laboratorio_id=l.id).scalar() or 0
        ped_count = session.query(_func.count(database.Pedido.id)).filter_by(laboratorio=l.nombre).scalar() or 0
        analytics_count = session.query(_func.count(database.ProductAnalytics.codigo_barra))\
            .filter_by(laboratorio=l.nombre).scalar() or 0
        data.append({
            'id': l.id, 'nombre': l.nombre,
            'prod_count': prod_count,
            'ped_count': ped_count, 'analytics_count': analytics_count,
        })
    session.close()
    return render_template('laboratorios.html', laboratorios=data)


@app.route('/laboratorio/create', methods=['POST'])
def laboratorio_create():
    nombre = request.form.get('nombre', '').strip()
    if not nombre:
        flash('El nombre es obligatorio.')
        return redirect(url_for('laboratorios_list'))
    session = database.SessionLocal()
    existing = session.query(Laboratorio).filter(Laboratorio.nombre.ilike(nombre)).first()
    if existing:
        flash(f'Ya existe un laboratorio con ese nombre.')
    else:
        session.add(Laboratorio(nombre=nombre))
        session.commit()
    session.close()
    return redirect(url_for('laboratorios_list'))


@app.route('/laboratorio/<int:lab_id>/edit', methods=['POST'])
def laboratorio_edit(lab_id):
    nombre = request.form.get('nombre', '').strip()
    if not nombre:
        flash('El nombre es obligatorio.')
        return redirect(url_for('laboratorios_list'))
    session = database.SessionLocal()
    lab = session.get(Laboratorio, lab_id)
    if lab:
        lab.nombre = nombre
        session.commit()
    session.close()
    return redirect(url_for('laboratorios_list'))


@app.route('/laboratorio/<int:lab_id>/delete', methods=['POST'])
def laboratorio_delete(lab_id):
    session = database.SessionLocal()
    lab = session.get(Laboratorio, lab_id)
    if lab:
        session.query(Producto).filter_by(laboratorio_id=lab_id).update({'laboratorio_id': None})
        session.delete(lab)
        session.commit()
    session.close()
    return redirect(url_for('laboratorios_list'))


@app.route('/producto/<int:prod_id>/laboratorio', methods=['POST'])
def producto_set_laboratorio(prod_id):
    lab_id = request.form.get('laboratorio_id') or None
    session = database.SessionLocal()
    prod = session.get(Producto, prod_id)
    if prod:
        prod.laboratorio_id = int(lab_id) if lab_id else None
        session.commit()
    session.close()
    return ('', 204)


@app.route('/producto/<int:prod_id>/edit', methods=['POST'])
def producto_edit(prod_id):
    data = request.get_json(silent=True) or {}
    field = data.get('field')
    value = (data.get('value') or '').strip()
    allowed = {'descripcion', 'codigo_barra', 'codigo_barra_alt1', 'codigo_barra_alt2', 'codigo_barra_alt3', 'precio_pvp', 'es_pack'}
    if field not in allowed:
        return {'error': 'Campo no permitido'}, 400
    session = database.SessionLocal()
    try:
        prod = session.get(Producto, prod_id)
        if not prod:
            return {'error': 'No encontrado'}, 404
        if field == 'precio_pvp':
            try:
                setattr(prod, field, float(value.replace(',', '.')) if value else None)
            except ValueError:
                return {'error': 'Precio inválido'}, 400
        elif field == 'es_pack':
            prod.es_pack = 1 if value in ('1', 'true', 'True') else 0
        else:
            setattr(prod, field, value or None)
        from datetime import datetime as _dt
        prod.actualizado_en = _dt.now().date()
        session.commit()
        return {'ok': True}
    finally:
        session.close()


@app.route('/producto/edit-by-barcode', methods=['POST'])
def producto_edit_by_barcode():
    data = request.get_json(silent=True) or {}
    cb    = (data.get('codigo_barra') or '').strip()
    field = data.get('field')
    value = (data.get('value') or '').strip()
    if not cb or field not in {'descripcion', 'precio_pvp'}:
        return {'error': 'Parámetros inválidos'}, 400
    session = database.SessionLocal()
    try:
        prod = _find_producto(session, cb)
        if not prod:
            # Crear producto si no existe
            prod = Producto(codigo_barra=cb)
            session.add(prod)
            session.flush()
        if field == 'precio_pvp':
            prod.precio_pvp = float(value.replace(',', '.')) if value else None
        else:
            setattr(prod, field, value or None)
        from datetime import datetime as _dt
        prod.actualizado_en = _dt.now().date()
        session.commit()
        return {'ok': True, 'id': prod.id}
    except Exception as e:
        session.rollback()
        return {'error': str(e)}, 500
    finally:
        session.close()


@app.route('/producto/create', methods=['POST'])
def producto_create():
    data = request.get_json(silent=True) or {}
    cb = (data.get('codigo_barra') or '').strip()
    if not cb:
        return {'error': 'Código de barra requerido'}, 400
    session = database.SessionLocal()
    try:
        if session.query(Producto).filter_by(codigo_barra=cb).first():
            return {'error': 'Ya existe un producto con ese código'}, 409
        prod = Producto(
            codigo_barra=cb,
            descripcion=(data.get('descripcion') or '').strip() or None,
            precio_pvp=float(data['precio_pvp']) if data.get('precio_pvp') else None,
            es_pack=1 if data.get('es_pack') else 0,
        )
        session.add(prod)
        session.commit()
        return {'ok': True, 'id': prod.id}
    except Exception as e:
        session.rollback()
        return {'error': str(e)}, 500
    finally:
        session.close()


@app.route('/producto/<int:prod_id>/delete', methods=['POST'])
def producto_delete(prod_id):
    session = database.SessionLocal()
    try:
        prod = session.get(Producto, prod_id)
        if not prod:
            return {'error': 'No encontrado'}, 404
        session.delete(prod)
        session.commit()
        return {'ok': True}
    finally:
        session.close()


@app.route('/provider/<int:provider_id>/parser-preview', methods=['POST'])
def provider_parser_preview(provider_id):
    session = database.SessionLocal()
    provider = session.get(database.Provider, provider_id)
    session.close()
    if not provider or not provider.parser_file:
        return {'error': 'El proveedor no tiene parser configurado.'}, 400

    f = request.files.get('pdf')
    if not f or not f.filename.lower().endswith('.pdf'):
        return {'error': 'Seleccioná un archivo PDF.'}, 400

    tmp_path = os.path.join(UPLOAD_FOLDER, f'preview_{secure_filename(f.filename)}')
    f.save(tmp_path)
    try:
        data = parse_invoice_pdf(tmp_path, provider.parser_file)
    except Exception as e:
        return {'error': f'Error en el parser: {e}'}, 500
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass

    items = []
    for it in (data.get('items') or []):
        items.append({
            'codigo_barra': it.get('codigo_barra') or '',
            'descripcion': it.get('descripcion') or '',
            'cantidad': it.get('cantidad') or '',
            'precio_unitario': it.get('precio_unitario') or '',
            'dto': it.get('dto') or '',
            'importe': it.get('importe') or '',
            'lote': it.get('lote') or '',
            'vencimiento': it.get('vencimiento') or '',
        })
    return {
        'parser': provider.parser_file,
        'numero_factura': data.get('numero_factura') or '',
        'fecha': str(data.get('fecha') or ''),
        'proveedor': data.get('proveedor_razon') or '',
        'total': str(data.get('total') or ''),
        'items': items,
    }


@app.route('/provider/<int:provider_id>/edit', methods=['POST'])
def provider_edit(provider_id):
    session = database.SessionLocal()
    provider = session.get(database.Provider, provider_id)
    if not provider:
        session.close()
        flash('Proveedor no encontrado.')
        return redirect(url_for('providers_list'))

    provider.razon_social = request.form.get('razon_social', provider.razon_social).strip() or provider.razon_social
    provider.cuit = request.form.get('cuit', provider.cuit or '').strip() or None
    provider.parser_file = request.form.get('parser_file', provider.parser_file or '').strip() or None
    provider.ruta_facturas = request.form.get('ruta_facturas', '').strip() or None
    ms = request.form.get('match_strategy', 'barcode')
    provider.match_strategy = ms if ms in ('barcode', 'descripcion') else 'barcode'
    provider.grabar_productos = 1 if request.form.get('grabar_productos') == '1' else 0
    tipo = (request.form.get('tipo') or '').strip().lower()
    if tipo in ('drogueria', 'laboratorio', 'otro'):
        provider.tipo = tipo
    session.commit()
    session.close()
    return redirect(url_for('providers_list', tipo=request.form.get('tipo_filter') or None))


@app.route('/provider/<int:provider_id>/delete', methods=['POST'])
def provider_delete(provider_id):
    session = database.SessionLocal()
    provider = session.get(database.Provider, provider_id)
    if provider:
        # Borrar claim_items de los reclamos de este proveedor
        claim_ids = [c.id for c in session.query(database.Claim).filter_by(proveedor_id=provider_id).all()]
        if claim_ids:
            session.query(database.ClaimItem).filter(
                database.ClaimItem.reclamo_id.in_(claim_ids)
            ).delete(synchronize_session=False)
        session.query(database.Claim).filter_by(proveedor_id=provider_id).delete()
        session.query(database.BarcodeMapping).filter_by(proveedor_id=provider_id).delete()
        # Borrar invoice_batches del proveedor (nullificar FK en facturas primero)
        batch_ids = [b.id for b in session.query(database.InvoiceBatch).filter_by(proveedor_id=provider_id).all()]
        if batch_ids:
            session.query(database.Invoice).filter(
                database.Invoice.batch_id.in_(batch_ids)
            ).update({'batch_id': None}, synchronize_session=False)
            session.query(database.InvoiceBatch).filter(
                database.InvoiceBatch.id.in_(batch_ids)
            ).delete(synchronize_session=False)
        session.delete(provider)
        session.commit()
    session.close()
    return redirect(url_for('providers_list'))


@app.route('/provider/<int:provider_id>/invoices')
def provider_invoices(provider_id):
    session = database.SessionLocal()
    provider = session.get(database.Provider, provider_id)
    if not provider:
        session.close()
        flash('Proveedor no encontrado.')
        return redirect(url_for('providers_list'))
    invoices = session.query(database.Invoice).filter(
        (database.Invoice.proveedor_cuit == provider.cuit) |
        (database.Invoice.proveedor_razon == provider.razon_social)
    ).order_by(database.Invoice.fecha.desc()).all()
    session.close()
    return render_template('provider_invoices.html', provider=provider, invoices=invoices)


@app.route('/invoice/<int:invoice_id>/delete', methods=['POST'])
def delete_invoice(invoice_id):
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    if not invoice:
        session.close()
        flash('Factura no encontrada.')
        return redirect(url_for('providers_list'))

    # Resolver proveedor para redirigir después
    provider = None
    if invoice.proveedor_cuit:
        provider = session.query(database.Provider).filter_by(cuit=invoice.proveedor_cuit).first()
    if not provider and invoice.proveedor_razon:
        provider = session.query(database.Provider).filter_by(razon_social=invoice.proveedor_razon).first()
    provider_id = provider.id if provider else None

    # Borrar en cascada manualmente
    diff_ids = [d.id for d in session.query(database.StockDifference).filter_by(factura_id=invoice_id).all()]
    if diff_ids:
        session.query(database.ClaimItem).filter(
            database.ClaimItem.diferencia_id.in_(diff_ids)
        ).delete(synchronize_session=False)
    session.query(database.StockDifference).filter_by(factura_id=invoice_id).delete()
    session.query(database.ClaimItem).filter(
        database.ClaimItem.reclamo_id.in_(
            session.query(database.Claim.id).filter_by(factura_id=invoice_id)
        )
    ).delete(synchronize_session=False)
    session.query(database.Claim).filter_by(factura_id=invoice_id).delete()
    session.query(database.InvoiceItem).filter_by(factura_id=invoice_id).delete()
    session.delete(invoice)
    session.commit()
    session.close()

    if provider_id:
        return redirect(url_for('provider_invoices', provider_id=provider_id))
    return redirect(url_for('providers_list'))


@app.route('/claims')
def claims_list():
    session = database.SessionLocal()
    claims = (session.query(database.Claim)
              .order_by(database.Claim.creado_en.desc()).all())
    # Cargar relaciones antes de cerrar
    for c in claims:
        _ = c.provider
    session.close()
    return render_template('claims_list.html', claims=claims)


@app.route('/provider/<int:provider_id>/mappings')
def provider_mappings(provider_id):
    session = database.SessionLocal()
    provider = session.get(database.Provider, provider_id)
    if not provider:
        session.close()
        flash('Proveedor no encontrado.')
        return redirect(url_for('providers_list'))
    mappings = (session.query(database.BarcodeMapping)
                .filter_by(proveedor_id=provider_id)
                .order_by(database.BarcodeMapping.creado_en.desc()).all())
    session.close()
    return render_template('provider_mappings.html', provider=provider, mappings=mappings)


@app.route('/provider/<int:provider_id>/mappings/<int:mapping_id>/delete', methods=['POST'])
def delete_mapping(provider_id, mapping_id):
    session = database.SessionLocal()
    mapping = session.get(database.BarcodeMapping, mapping_id)
    if mapping and mapping.proveedor_id == provider_id:
        session.delete(mapping)
        session.commit()
    session.close()
    return redirect(url_for('provider_mappings', provider_id=provider_id))


@app.route('/provider/<int:provider_id>/mappings/delete-all', methods=['POST'])
def delete_all_mappings(provider_id):
    session = database.SessionLocal()
    session.query(database.BarcodeMapping).filter_by(proveedor_id=provider_id).delete()
    session.commit()
    session.close()
    flash('Todas las equivalencias fueron eliminadas.')
    return redirect(url_for('provider_mappings', provider_id=provider_id))


@app.route('/invoice/<int:invoice_id>/items')
def invoice_items(invoice_id):
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    if not invoice:
        session.close()
        flash('Factura no encontrada.')
        return redirect(url_for('index'))
    items = session.query(database.InvoiceItem).filter_by(factura_id=invoice_id).all()
    session.close()
    return render_template('invoice_items.html', invoice=invoice, items=items)


@app.route('/invoice/<int:invoice_id>/items/export')
def invoice_items_export(invoice_id):
    """Descarga los ítems de la factura como XLS."""
    import io
    import openpyxl
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    items = session.query(database.InvoiceItem).filter_by(factura_id=invoice_id).all()
    session.close()
    if not invoice:
        return 'Factura no encontrada', 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ítems'
    headers = ['Código de barra', 'Descripción', 'Cantidad', 'Precio unitario',
               'Dto %', 'Importe', 'Lote', 'Vencimiento']
    ws.append(headers)
    for it in items:
        ws.append([it.codigo_barra, it.descripcion, it.cantidad,
                   float(it.precio_unitario or 0), float(it.dto or 0),
                   float(it.importe or 0), it.lote, it.vencimiento])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'items_{invoice.numero_factura}.xlsx'
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@app.route('/invoice/<int:invoice_id>/differences/export')
def invoice_differences_export(invoice_id):
    """Descarga las diferencias de stock de la factura como XLS."""
    import io
    import openpyxl
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    diffs = session.query(database.StockDifference).filter_by(factura_id=invoice_id).all()
    session.close()
    if not invoice:
        return 'Factura no encontrada', 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Diferencias'
    ws.append(['Código de barra', 'Descripción', 'Cant. factura', 'Cant. ERP',
               'Diferencia', 'Observaciones'])
    for d in diffs:
        ws.append([d.codigo_barra, d.descripcion, d.cantidad_factura,
                   d.cantidad_erp, d.diferencia, d.observaciones])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'diferencias_{invoice.numero_factura}.xlsx'
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@app.route('/invoice/<int:invoice_id>/compare')
def compare_view(invoice_id):
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    if not invoice:
        session.close()
        flash('Factura no encontrada.')
        return redirect(url_for('index'))
    invoice_diffs = (session.query(database.StockDifference)
                     .filter_by(factura_id=invoice_id)
                     .order_by(database.StockDifference.descripcion).all())
    erp_items = get_erp_items_with_issues(session, invoice_id)
    # Lookup de precio unitario por código de barra (desde InvoiceItem)
    inv_items = session.query(database.InvoiceItem).filter_by(factura_id=invoice_id).all()
    inv_prices = {
        item.codigo_barra: float(item.precio_unitario or 0)
        for item in inv_items if item.codigo_barra
    }
    session.close()
    return render_template('compare.html', invoice=invoice,
                           invoice_diffs=invoice_diffs, erp_items=erp_items,
                           inv_prices=inv_prices)


@app.route('/invoice/<int:invoice_id>/apply-mapping', methods=['POST'])
def apply_mapping(invoice_id):
    session = database.SessionLocal()
    invoice = session.get(database.Invoice, invoice_id)
    diffs = (session.query(database.StockDifference)
             .filter_by(factura_id=invoice_id)
             .order_by(database.StockDifference.descripcion).all())

    # Resolver proveedor_id para guardar mappings
    proveedor_id = None
    if invoice and invoice.proveedor_cuit:
        prov = session.query(database.Provider).filter_by(cuit=invoice.proveedor_cuit).first()
        if prov:
            proveedor_id = prov.id
    if proveedor_id is None and invoice and invoice.proveedor_razon:
        prov = session.query(database.Provider).filter_by(razon_social=invoice.proveedor_razon).first()
        if prov:
            proveedor_id = prov.id

    to_delete = []
    for key, value in request.form.items():
        if not key.startswith('mapping_') or not value.strip():
            continue
        try:
            erp_id = int(key.replace('mapping_', ''))
            inv_num = int(value.strip())
        except ValueError:
            continue
        if inv_num < 1 or inv_num > len(diffs):
            continue

        target_diff = diffs[inv_num - 1]
        erp_item = session.get(database.ErpStock, erp_id)
        if not erp_item:
            continue

        # Guardar correspondencia para próximas facturas
        if proveedor_id and target_diff.codigo_barra and erp_item.codigo_barra:
            save_barcode_mapping(
                session,
                proveedor_id=proveedor_id,
                codigo_barra_factura=target_diff.codigo_barra,
                codigo_barra_erp=erp_item.codigo_barra,
                descripcion_factura=target_diff.descripcion,
                descripcion_erp=erp_item.descripcion,
            )
            _upsert_producto(session, erp_item.codigo_barra, erp_item.descripcion,
                             fecha_compra=invoice.fecha if invoice else None)
            _add_alt_barcode(session, erp_item.codigo_barra, target_diff.codigo_barra)

        target_diff.cantidad_erp = erp_item.cantidad
        target_diff.diferencia = target_diff.cantidad_factura - erp_item.cantidad
        target_diff.observaciones = (
            f'Cruce manual con ERP: {erp_item.descripcion} ({erp_item.codigo_barra})'
        )
        if target_diff.diferencia == 0:
            to_delete.append(target_diff)

    for diff in to_delete:
        session.delete(diff)

    session.commit()
    session.close()
    return redirect(url_for('show_results', invoice_id=invoice_id))


@app.route('/claim/<int:claim_id>/pdf')
def claim_pdf(claim_id):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    session = database.SessionLocal()
    claim = session.get(Claim, claim_id)
    if not claim:
        session.close()
        return 'Reclamo no encontrado', 404
    _ = claim.provider
    _ = claim.items
    session.close()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    DARK   = colors.HexColor('#1a1a1a')
    BRAND  = colors.HexColor('#EAB308')
    GRAY   = colors.HexColor('#555555')
    LGRAY  = colors.HexColor('#f5f5f5')
    HBG    = colors.HexColor('#2c2c2e')

    title_style = ParagraphStyle('title', fontSize=18, textColor=DARK,
                                 fontName='Helvetica-Bold', spaceAfter=4)
    sub_style   = ParagraphStyle('sub',   fontSize=10, textColor=GRAY,
                                 fontName='Helvetica', spaceAfter=2)
    label_style = ParagraphStyle('lbl',   fontSize=8,  textColor=GRAY,
                                 fontName='Helvetica-Bold', spaceBefore=10, spaceAfter=1)
    value_style = ParagraphStyle('val',   fontSize=10, textColor=DARK,
                                 fontName='Helvetica')

    numero_factura = claim.numero_factura or '—'
    proveedor_razon = claim.provider.razon_social if claim.provider else '—'
    proveedor_cuit  = claim.provider.cuit if claim.provider else '—'

    cfg = get_config()
    story = []

    # ── Encabezado ──
    story.append(Paragraph('Reclamo de Faltantes', title_style))
    story.append(Paragraph(f'N° de reclamo: <b>#{claim.id}</b> · {cfg["farmacia_nombre"]}', sub_style))
    story.append(Spacer(1, 0.4*cm))

    # ── Datos en tabla de dos columnas ──
    info_data = [
        [Paragraph('<b>Proveedor</b>', label_style), Paragraph('<b>Factura</b>', label_style)],
        [Paragraph(proveedor_razon, value_style),    Paragraph(numero_factura, value_style)],
        [Paragraph('<b>CUIT</b>', label_style),       Paragraph('<b>Fecha factura</b>', label_style)],
        [Paragraph(proveedor_cuit or '—', value_style), Paragraph(str(claim.fecha), value_style)],
        [Paragraph('<b>Fecha reclamo</b>', label_style), Paragraph('<b>Estado</b>', label_style)],
        [Paragraph(claim.creado_en.strftime('%d/%m/%Y') if claim.creado_en else '—', value_style),
         Paragraph(claim.estado, value_style)],
    ]
    info_table = Table(info_data, colWidths=[8.5*cm, 8.5*cm])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('LINEBELOW', (0,1), (-1,1), 0.5, colors.HexColor('#dddddd')),
        ('LINEBELOW', (0,3), (-1,3), 0.5, colors.HexColor('#dddddd')),
        ('LINEBELOW', (0,5), (-1,5), 0.5, colors.HexColor('#dddddd')),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.6*cm))

    # ── Tabla de ítems ──
    story.append(Paragraph(f'Detalle de ítems ({len(claim.items)} producto{"s" if len(claim.items) != 1 else ""})', label_style))
    story.append(Spacer(1, 0.2*cm))

    headers = ['#', 'Código', 'Descripción', 'Fact.', 'ERP', 'Dif.']
    col_w   = [0.8*cm, 3*cm, 8.2*cm, 1.5*cm, 1.5*cm, 1.5*cm]

    hdr_style = ParagraphStyle('hdr', fontSize=8, textColor=colors.white,
                               fontName='Helvetica-Bold', alignment=TA_CENTER)
    cell_style = ParagraphStyle('cell', fontSize=8, textColor=DARK, fontName='Helvetica')
    num_style  = ParagraphStyle('num',  fontSize=8, textColor=DARK, fontName='Helvetica',
                                alignment=TA_CENTER)

    rows = [[Paragraph(h, hdr_style) for h in headers]]
    for i, item in enumerate(claim.items, 1):
        dif = item.diferencia or 0
        dif_str = f'+{dif}' if dif > 0 else str(dif)
        rows.append([
            Paragraph(str(i), num_style),
            Paragraph(item.codigo_barra or '—', num_style),
            Paragraph(item.descripcion or '—', cell_style),
            Paragraph(str(item.cantidad_factura or 0), num_style),
            Paragraph(str(item.cantidad_erp or 0), num_style),
            Paragraph(dif_str, num_style),
        ])

    items_table = Table(rows, colWidths=col_w, repeatRows=1)
    items_table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,0), HBG),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#fafafa')]),
        ('GRID',         (0,0), (-1,-1), 0.4, colors.HexColor('#e0e0e0')),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
        ('LEFTPADDING',  (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(items_table)

    doc.build(story)
    buf.seek(0)

    # Nombre de archivo: Reclamo_N{id}_{numero_factura}.pdf
    safe_factura = re.sub(r'[^a-zA-Z0-9_-]', '_', numero_factura)
    filename = f'Reclamo_N{claim.id}_{safe_factura}.pdf'

    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@app.route('/batch/new')
def batch_new():
    return render_template('batch.html', providers=get_providers())


@app.route('/batch/add-pdf', methods=['POST'])
def batch_add_pdf():
    proveedor_id = request.form.get('proveedor_id')
    batch_id = request.form.get('batch_id') or None
    tipo_comprobante = request.form.get('tipo_comprobante', 'FAC').upper()
    invoice_file = request.files.get('invoice_pdf')

    if not proveedor_id:
        return jsonify({'error': 'Seleccioná un proveedor.'}), 400
    if not invoice_file or not allowed_file(invoice_file.filename):
        return jsonify({'error': 'PDF inválido.'}), 400

    session = database.SessionLocal()
    provider = session.get(database.Provider, int(proveedor_id))
    if not provider:
        session.close()
        return jsonify({'error': 'Proveedor no encontrado.'}), 400
    if not provider.parser_file:
        session.close()
        return jsonify({'error': f'El proveedor "{provider.razon_social}" no tiene parser configurado.'}), 400

    filename = secure_filename(invoice_file.filename)
    invoice_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    invoice_file.save(invoice_path)

    try:
        invoice_data = parse_invoice_pdf(invoice_path, provider.parser_file)
    except Exception as e:
        session.close()
        return jsonify({'error': f'Error al leer PDF: {e}'}), 400

    if not invoice_data.get('items'):
        session.close()
        return jsonify({'error': f'El parser no detectó artículos en este PDF.'}), 400

    if tipo_comprobante not in ('FAC', 'NCR'):
        tipo_comprobante = 'FAC'

    invoice = save_invoice_to_db(session, invoice_data,
                                  pdf_filename=os.path.basename(invoice_path),
                                  tipo_comprobante=tipo_comprobante)

    # Crear o recuperar batch
    batch = None
    if batch_id:
        batch = session.get(InvoiceBatch, int(batch_id))
    if not batch:
        batch = InvoiceBatch(proveedor_id=int(proveedor_id))
        session.add(batch)
        session.flush()

    invoice.batch_id = batch.id
    session.commit()

    result = {
        'batch_id': batch.id,
        'invoice_id': invoice.id,
        'numero_factura': invoice.numero_factura,
        'total_articulos': invoice.total_articulos or 0,
        'proveedor_razon': invoice.proveedor_razon,
        'fecha': str(invoice.fecha),
        'tipo_comprobante': invoice.tipo_comprobante,
    }
    session.close()
    return jsonify(result), 200


@app.route('/batch/process', methods=['POST'])
def batch_process():
    batch_id = request.form.get('batch_id')
    erp_file = request.files.get('erp_excel')

    if not batch_id:
        flash('Batch no encontrado.')
        return redirect(url_for('index'))
    if not erp_file or not allowed_file(erp_file.filename):
        flash('ERP Excel inválido.')
        return redirect(url_for('batch_new'))

    erp_filename = secure_filename(erp_file.filename)
    erp_path = os.path.join(app.config['UPLOAD_FOLDER'], erp_filename)
    erp_file.save(erp_path)

    try:
        erp_data = parse_erp_excel(erp_path)
    except Exception as e:
        flash(f'Error al leer el ERP: {e}')
        return redirect(url_for('batch_new'))

    session = database.SessionLocal()
    batch = session.get(InvoiceBatch, int(batch_id))
    if not batch:
        session.close()
        flash('Batch no encontrado.')
        return redirect(url_for('index'))

    batch.erp_filename = erp_filename
    batch.estado = 'PROCESADO'
    session.commit()

    save_erp_to_db(session, erp_data)

    invoices = session.query(database.Invoice).filter_by(batch_id=batch.id).all()
    for invoice in invoices:
        invoice.erp_filename = erp_filename
        differences = compare_invoice_vs_erp(session, invoice.id)
        save_differences(session, invoice.id, differences)
    session.commit()
    session.close()

    return redirect(url_for('batch_results', batch_id=batch.id))


@app.route('/batch/<int:batch_id>/results')
def batch_results(batch_id):
    session = database.SessionLocal()
    batch = session.get(InvoiceBatch, batch_id)
    if not batch:
        session.close()
        flash('Batch no encontrado.')
        return redirect(url_for('index'))

    provider = session.get(database.Provider, batch.proveedor_id)
    invoices = session.query(database.Invoice).filter_by(batch_id=batch_id).all()

    invoice_data = []
    for inv in invoices:
        diff_count = session.query(database.StockDifference).filter_by(factura_id=inv.id).count()
        invoice_data.append({
            'invoice': inv,
            'diff_count': diff_count,
        })
    session.close()
    return render_template('batch_results.html', batch=batch, provider=provider,
                           invoices=invoice_data)


@app.route('/health')
def health():
    return 'OK', 200


# ── Análisis de compras ────────────────────────────────────────────────────────

import uuid
import json
from parsers.sales_history import parse_sales_history_pdf
from parsers.sales_history_xls import parse_sales_history_xls
from parsers.sales_history_html import parse_sales_history_html
from purchase_engine import analyze_purchase

PURCHASE_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads', 'purchase')
os.makedirs(PURCHASE_FOLDER, exist_ok=True)


@app.route('/purchase')
def purchase_index():
    return render_template('purchase_analysis.html')


def _analyze_sales_file(tmp_path, ext, n_days):
    """Procesa un único archivo de estadística de ventas.
    Devuelve dict con uid/laboratorio/productos/periodo o {'error': str}."""
    try:
        if ext == 'pdf':
            parsed = parse_sales_history_pdf(tmp_path)
        elif ext in ('html', 'htm'):
            parsed = parse_sales_history_html(tmp_path)
        else:
            parsed = parse_sales_history_xls(tmp_path)
    except Exception as e:
        return {'error': f'Error al parsear: {e}'}

    if not parsed.get('products'):
        return {'error': 'No se encontraron productos en el archivo.'}

    cfg = get_config()
    results = analyze_purchase(
        parsed['products'], n_days,
        parsed['start_month'], parsed['end_month'],
        umbral_pico=cfg['umbral_pico'],
        umbral_baja=cfg['umbral_baja'],
        umbral_tendencia=cfg['umbral_tendencia'],
        rot_alta_min=cfg['rot_alta_min'],
        rot_media_min=cfg['rot_media_min'],
    )

    uid = str(uuid.uuid4())
    data = {
        'uid': uid,
        'farmacia': parsed['farmacia'],
        'laboratorio': parsed['laboratorio'],
        'periodo': parsed['periodo'],
        'start_month': parsed.get('start_month', 4),
        'n_days': n_days,
        'umbral_tendencia': cfg['umbral_tendencia'],
        'rot_alta_min': cfg['rot_alta_min'],
        'rot_alta_tol': cfg['rot_alta_tol'],
        'rot_media_min': cfg['rot_media_min'],
        'rot_media_tol': cfg['rot_media_tol'],
        'rot_baja_tol': cfg['rot_baja_tol'],
        'products': results,
    }
    json_path = os.path.join(PURCHASE_FOLDER, f'{uid}.json')
    with open(json_path, 'w', encoding='utf-8') as jf:
        json.dump(data, jf, ensure_ascii=False)

    _snapshot_product_analytics(results, parsed.get('laboratorio'))

    return {
        'uid': uid,
        'laboratorio': parsed.get('laboratorio') or '(sin laboratorio)',
        'periodo': parsed.get('periodo') or '',
        'count': len(results),
    }


@app.route('/purchase/analyze', methods=['POST'])
def purchase_analyze():
    f = request.files.get('sales_pdf')
    try:
        n_days = max(1, min(365, int(request.form.get('n_days', 35))))
    except (ValueError, TypeError):
        n_days = 35

    if not f or not f.filename:
        flash('Seleccioná un archivo PDF o Excel.')
        return redirect(url_for('purchase_index'))

    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in ('pdf', 'xlsx', 'xls', 'html', 'htm'):
        flash('Formato no soportado. Usá PDF, Excel (.xlsx / .xls) o HTML.')
        return redirect(url_for('purchase_index'))

    filename = secure_filename(f.filename)
    tmp_path = os.path.join(UPLOAD_FOLDER, f'purchase_{filename}')
    f.save(tmp_path)

    try:
        res = _analyze_sales_file(tmp_path, ext, n_days)
    finally:
        try: os.remove(tmp_path)
        except Exception: pass

    if 'error' in res:
        flash(res['error'])
        return redirect(url_for('purchase_index'))

    return redirect(url_for('purchase_results', uid=res['uid']))


@app.route('/purchase/processed')
def purchase_processed():
    """Lista todos los análisis de ventas ya procesados (JSONs en PURCHASE_FOLDER)."""
    from datetime import datetime as _dt
    items = []
    try:
        for fn in os.listdir(PURCHASE_FOLDER):
            if not fn.endswith('.json'):
                continue
            path = os.path.join(PURCHASE_FOLDER, fn)
            try:
                with open(path, encoding='utf-8') as jf:
                    d = json.load(jf)
                items.append({
                    'uid': d.get('uid') or fn[:-5],
                    'laboratorio': d.get('laboratorio') or '(sin laboratorio)',
                    'periodo': d.get('periodo') or '',
                    'n_days': d.get('n_days') or 0,
                    'count': len(d.get('products') or []),
                    'mtime': _dt.fromtimestamp(os.path.getmtime(path)),
                })
            except Exception:
                continue
    except FileNotFoundError:
        pass
    items.sort(key=lambda x: x['mtime'], reverse=True)
    return render_template('purchase_processed.html', items=items)


@app.route('/purchase/batch', methods=['POST'])
def purchase_batch():
    files = request.files.getlist('sales_files')
    try:
        n_days = max(1, min(365, int(request.form.get('n_days', 35))))
    except (ValueError, TypeError):
        n_days = 35

    files = [f for f in files if f and f.filename]
    if not files:
        flash('Seleccioná al menos un archivo.')
        return redirect(url_for('purchase_index'))

    results = []
    for f in files:
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        if ext not in ('pdf', 'xlsx', 'xls', 'html', 'htm'):
            results.append({'filename': f.filename, 'error': f'Formato no soportado (.{ext})'})
            continue

        filename = secure_filename(f.filename)
        tmp_path = os.path.join(UPLOAD_FOLDER, f'batch_{uuid.uuid4().hex}_{filename}')
        f.save(tmp_path)
        try:
            res = _analyze_sales_file(tmp_path, ext, n_days)
        except Exception as e:
            app.logger.exception('Error procesando %s', f.filename)
            res = {'error': str(e)}
        finally:
            try: os.remove(tmp_path)
            except Exception: pass

        res['filename'] = f.filename
        results.append(res)

    ok = [r for r in results if 'uid' in r]
    fail = [r for r in results if 'error' in r]
    return render_template('purchase_batch.html',
                           results=results, ok=ok, fail=fail, n_days=n_days)


def _snapshot_product_analytics(results, laboratorio):
    """Upsert de ProductAnalytics por codigo_barra desde los resultados de analyze_purchase."""
    from datetime import datetime as _dt
    session = database.SessionLocal()
    try:
        seen = {}
        for p in results:
            cb = (p.get('codigo_barra') or '').strip()
            if not cb:
                continue
            seen[cb] = p  # último gana si hay duplicados en el mismo Excel
        for cb, p in seen.items():
            forecast = p.get('forecast')
            forecast_next = None
            if isinstance(forecast, list) and forecast:
                forecast_next = forecast[0]
            elif isinstance(forecast, (int, float)):
                forecast_next = forecast
            pa = session.get(database.ProductAnalytics, cb)
            if pa is None:
                pa = database.ProductAnalytics(codigo_barra=cb)
                session.add(pa)
            pa.descripcion = (p.get('nombre') or p.get('descripcion') or '')[:200]
            pa.laboratorio = laboratorio
            pa.stock = int(p.get('stock') or 0)
            pa.avg_monthly = float(p.get('avg_monthly') or 0)
            pa.rotacion = p.get('rotacion')
            pa.slope = float(p.get('slope') or 0)
            pa.forecast_next = float(forecast_next) if forecast_next is not None else None
            pa.sin_mov_60d = 1 if p.get('sin_mov_60d') else 0
            pa.precio_pvp = float(p.get('precio_pvp') or 0)
            pa.actualizado_en = _dt.utcnow()
        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


@app.route('/purchase/results/<uid>')
def purchase_results(uid):
    if not re.match(r'^[0-9a-f-]{36}$', uid):
        flash('Sesión inválida.')
        return redirect(url_for('purchase_index'))

    json_path = os.path.join(PURCHASE_FOLDER, f'{uid}.json')
    if not os.path.exists(json_path):
        flash('La sesión expiró o no existe. Analizá el PDF nuevamente.')
        return redirect(url_for('purchase_index'))

    with open(json_path, encoding='utf-8') as jf:
        data = json.load(jf)

    # Defaults para JSONs viejos que no traen config de rotación/tendencia
    cfg = get_config()
    data.setdefault('umbral_tendencia', cfg['umbral_tendencia'])
    data.setdefault('rot_alta_min', cfg['rot_alta_min'])
    data.setdefault('rot_alta_tol', cfg['rot_alta_tol'])
    data.setdefault('rot_media_min', cfg['rot_media_min'])
    data.setdefault('rot_media_tol', cfg['rot_media_tol'])
    data.setdefault('rot_baja_tol', cfg['rot_baja_tol'])

    # Enriquecer con es_pack desde Producto y ModuloPack
    session = database.SessionLocal()
    try:
        barcodes = [p['codigo_barra'] for p in data.get('products', []) if p.get('codigo_barra')]
        pack_eans = {mp.ean_pack for mp in session.query(ModuloPack).all()}
        prods_pack = {
            p.codigo_barra: bool(p.es_pack)
            for p in session.query(Producto).filter(
                Producto.codigo_barra.in_(barcodes)
            ).all()
        }
        for p in data.get('products', []):
            cb = p.get('codigo_barra', '')
            p['es_pack'] = prods_pack.get(cb, False) or (cb in pack_eans)
    finally:
        session.close()

    # Build ordered month labels in Spanish for the chart
    _mes_jan = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    sm = data.get('start_month', 4)
    month_es = [_mes_jan[(sm - 1 + i) % 12] for i in range(12)]
    return render_template('purchase_results.html', month_es=month_es, **data)


@app.route('/purchase/export/<uid>/<fmt>', methods=['POST'])
def purchase_export(uid, fmt):
    if not re.match(r'^[0-9a-f-]{36}$', uid):
        return 'UID inválido', 400

    json_path = os.path.join(PURCHASE_FOLDER, f'{uid}.json')
    if not os.path.exists(json_path):
        flash('La sesión expiró.')
        return redirect(url_for('purchase_index'))

    with open(json_path, encoding='utf-8') as jf:
        data = json.load(jf)

    # Aplicar cantidades editadas por el usuario
    for i, p in enumerate(data['products']):
        edited = request.form.get(f'qty_{i}')
        if edited is not None:
            try:
                qty = int(edited)
                p['order_qty'] = max(0, qty)
                p['subtotal'] = round(p['order_qty'] * p['precio_pvp'], 2)
            except ValueError:
                pass

    lab = data.get('laboratorio', 'Compra')
    n = data.get('n_days', 35)
    periodo = data.get('periodo', '')
    farmacia_nombre = data.get('farmacia') or get_config()['farmacia_nombre']

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Compra {lab}"[:31]

        # Metadata
        ws['A1'] = f"Plan de compra — {lab}"
        ws['A1'].font = Font(bold=True, size=13)
        ws['A2'] = periodo
        ws['A3'] = f"Días proyectados: {n}"
        ws.append([])

        headers = ['Producto', 'Cód. Barras', 'P.PVP ($)', 'Stock',
                   'Prom/mes', f'Pronóstico ({n}d)', 'Pedido', 'Subtotal ($)', 'Comentario']
        ws.append(headers)

        hdr_row = ws.max_row
        hdr_fill = PatternFill('solid', fgColor='1C1C1E')
        hdr_font = Font(bold=True, color='EAB308')
        thin = Side(style='thin', color='3A3A3C')
        border = Border(bottom=thin)
        for cell in ws[hdr_row]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        for p in data['products']:
            if p['order_qty'] == 0 and p['total'] == 0:
                continue
            ws.append([
                p['nombre'], p['codigo_barra'],
                p['precio_pvp'], p['stock'],
                p['avg_monthly'], p['forecast'],
                p['order_qty'], p['subtotal'],
                p['comment'],
            ])

        # Anchos de columna
        widths = [40, 16, 12, 8, 10, 14, 10, 14, 50]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = make_response(buf.read())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename="Compra_{lab}_{n}d.xlsx"'
        return resp

    elif fmt == 'pdf':
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas as rl_canvas
        from io import BytesIO
        from datetime import datetime as _dt

        fecha_emision = _dt.now().strftime('%d/%m/%Y %H:%M')
        page_w, page_h = landscape(A4)

        # Canvas con numeración "Página X de Y"
        class _NumberedCanvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                rl_canvas.Canvas.__init__(self, *args, **kwargs)
                self._pages = []

            def showPage(self):
                self._pages.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                total = len(self._pages)
                for i, state in enumerate(self._pages, 1):
                    self.__dict__.update(state)
                    self._draw_footer(i, total)
                    rl_canvas.Canvas.showPage(self)
                rl_canvas.Canvas.save(self)

            def _draw_footer(self, page_num, total):
                self.saveState()
                self.setFont('Helvetica', 7)
                self.setFillColor(colors.HexColor('#6B7280'))
                self.drawString(1.5*cm, 0.6*cm, f"Emitido: {fecha_emision}")
                self.drawRightString(page_w - 1.5*cm, 0.6*cm,
                                     f"Página {page_num} de {total}")
                self.restoreState()

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)

        styles = getSampleStyleSheet()
        hdr_bg  = colors.HexColor('#2D3748')
        row_a   = colors.white
        row_b   = colors.HexColor('#F3F4F6')
        txt     = colors.HexColor('#1A202C')
        accent  = colors.HexColor('#EAB308')
        grid_c  = colors.HexColor('#D1D5DB')

        title_style = ParagraphStyle('title', parent=styles['Normal'],
                                     fontSize=14, textColor=accent, spaceAfter=4)
        sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                   fontSize=9, textColor=txt, spaceAfter=2)
        cell_style = ParagraphStyle('cell', parent=styles['Normal'],
                                    fontSize=7, textColor=txt, leading=9)

        story = [
            Paragraph(f"Plan de compra — {lab}", title_style),
            Paragraph(f"{farmacia_nombre} · {periodo}", sub_style),
            Paragraph(f"Proyección: {n} días", sub_style),
            Spacer(1, 0.4*cm),
        ]

        headers = ['Producto', 'Barcode', 'P.PVP', 'Stock',
                   'Prom/m', 'Tendencia', 'Baja', 'Pico', 'Pedido', 'Subtotal', 'Nota']
        rows = [headers]

        for p in data['products']:
            if p['order_qty'] == 0 and p['total'] == 0:
                continue
            slope = p.get('slope', 0)
            tend = (f"↑ {slope}" if slope > 0.2 else f"↓ {abs(slope)}" if slope < -0.2 else f"{slope}")
            stock_str = 'agotado' if p['stock'] <= 0 else str(p['stock'])
            rows.append([
                Paragraph(p['nombre'], cell_style),
                p['codigo_barra'],
                f"${p['precio_pvp']:,.0f}",
                stock_str,
                p['avg_monthly'],
                tend,
                p.get('low_month', '') or '—',
                p.get('peak_month', '') or '—',
                p['order_qty'],
                f"${p['subtotal']:,.0f}",
                Paragraph(p.get('comment', ''), cell_style),
            ])

        col_widths = [4.8*cm, 3.0*cm, 2.0*cm, 1.6*cm, 1.5*cm, 1.8*cm, 1.4*cm, 1.4*cm, 1.5*cm, 2.2*cm, 3.5*cm]
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), hdr_bg),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0), 8),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [row_a, row_b]),
            ('TEXTCOLOR',     (0, 1), (-1, -1), txt),
            ('FONTSIZE',      (0, 1), (-1, -1), 7),
            ('ALIGN',         (2, 0), (-1, -1), 'RIGHT'),
            ('ALIGN',         (0, 0), (1, -1), 'LEFT'),
            ('ALIGN',         (6, 0), (7, -1), 'CENTER'),
            ('GRID',          (0, 0), (-1, -1), 0.3, grid_c),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t)

        doc.build(story, canvasmaker=_NumberedCanvas)
        buf.seek(0)

        resp = make_response(buf.read())
        resp.headers['Content-Type'] = 'application/pdf'
        resp.headers['Content-Disposition'] = f'attachment; filename="Compra_{lab}_{n}d.pdf"'
        return resp

    return 'Formato no soportado', 400


# ─── PEDIDOS ──────────────────────────────────────────────────────────────────

@app.route('/purchase/save-order/<uid>', methods=['POST'])
def purchase_save_order(uid):
    if not re.match(r'^[0-9a-f-]{36}$', uid):
        flash('UID inválido.')
        return redirect(url_for('purchase_index'))

    json_path = os.path.join(PURCHASE_FOLDER, f'{uid}.json')
    if not os.path.exists(json_path):
        flash('La sesión expiró. Analizá el PDF nuevamente.')
        return redirect(url_for('purchase_index'))

    session = database.SessionLocal()
    try:
        with open(json_path, encoding='utf-8') as jf:
            data = json.load(jf)

        products = data.get('products', [])
        items = []
        for i, p in enumerate(products):
            try:
                qty = int(request.form.get(f'qty_{i}') or 0)
            except (ValueError, TypeError):
                qty = 0
            if qty > 0:
                precio = float(p.get('precio_pvp') or 0)
                items.append(PedidoItem(
                    codigo_barra=p.get('codigo_barra', ''),
                    nombre=p.get('nombre', ''),
                    cantidad=qty,
                    precio_pvp=precio,
                    subtotal=round(qty * precio, 2),
                    rotacion=p.get('rotacion') or None,
                    avg_monthly=p.get('avg_monthly') or None,
                ))

        if not items:
            flash('No hay productos con cantidad > 0 para guardar.')
            return redirect(url_for('purchase_results', uid=uid))

        pedido = Pedido(
            laboratorio=data.get('laboratorio', ''),
            farmacia=data.get('farmacia', ''),
            periodo=data.get('periodo', ''),
            n_days=data.get('n_days', 0),
            items=items,
        )
        session.add(pedido)
        for it in items:
            _upsert_producto(session, it.codigo_barra, it.nombre, float(it.precio_pvp or 0))
        session.commit()
        flash(f'Pedido guardado: {len(items)} productos.')
        return redirect(url_for('orders_list'))
    except Exception as e:
        session.rollback()
        app.logger.exception('Error en purchase_save_order')
        flash(f'Error al guardar el pedido: {e}')
        return redirect(url_for('purchase_results', uid=uid))
    finally:
        session.close()


@app.route('/purchase/suggest', methods=['GET'])
def purchase_suggest():
    """Sugerencia de pedido consolidado por laboratorio."""
    import math
    from sqlalchemy import func as _func

    try:
        threshold_days = max(1, min(365, int(request.args.get('threshold_days', 10))))
    except (ValueError, TypeError):
        threshold_days = 10
    try:
        target_days = max(1, min(365, int(request.args.get('target_days', 10))))
    except (ValueError, TypeError):
        target_days = 10
    calcular = request.args.get('calcular') == '1'

    groups = []
    total_items = 0
    total_importe = 0.0

    if calcular:
        session = database.SessionLocal()
        try:
            PA = database.ProductAnalytics
            rows = session.query(PA).filter(
                PA.avg_monthly > 0,
                PA.stock * 30.0 / PA.avg_monthly < threshold_days
            ).order_by(PA.laboratorio.asc(), PA.descripcion.asc()).all()

            by_lab = {}
            for p in rows:
                avg = float(p.avg_monthly or 0)
                if avg <= 0:
                    continue
                daily = avg / 30.0
                target_stock = daily * target_days
                suggested = max(0, int(math.floor(target_stock - (p.stock or 0))))
                if suggested <= 0:
                    continue
                pvp = float(p.precio_pvp or 0)
                cov = round((p.stock or 0) * 30.0 / avg, 1) if avg > 0 else None
                item = {
                    'codigo_barra': p.codigo_barra,
                    'descripcion': p.descripcion or '',
                    'stock': p.stock or 0,
                    'avg_monthly': round(avg, 1),
                    'rotacion': p.rotacion,
                    'cobertura': cov,
                    'sugerido': suggested,
                    'precio_pvp': pvp,
                    'subtotal': round(suggested * pvp, 2),
                }
                lab = p.laboratorio or '(sin laboratorio)'
                by_lab.setdefault(lab, []).append(item)

            for lab in sorted(by_lab.keys()):
                lab_items = by_lab[lab]
                lab_total = sum(it['subtotal'] for it in lab_items)
                lab_units = sum(it['sugerido'] for it in lab_items)
                total_items += len(lab_items)
                total_importe += lab_total
                groups.append({
                    'laboratorio': lab,
                    'productos': lab_items,
                    'lab_total': round(lab_total, 2),
                    'lab_units': lab_units,
                })
        finally:
            session.close()

    return render_template('purchase_suggest.html',
                           threshold_days=threshold_days,
                           target_days=target_days,
                           calcular=calcular,
                           groups=groups,
                           total_items=total_items,
                           total_importe=round(total_importe, 2))


@app.route('/purchase/suggest/create-order', methods=['POST'])
def purchase_suggest_create_order():
    """Crea un Pedido para un laboratorio con los ítems seleccionados."""
    laboratorio = (request.form.get('laboratorio') or '').strip()
    if not laboratorio:
        flash('Laboratorio faltante.')
        return redirect(url_for('purchase_suggest'))

    session = database.SessionLocal()
    try:
        selected = request.form.getlist('sel')
        items = []
        for cb in selected:
            try:
                qty = int(request.form.get(f'qty_{cb}') or 0)
            except (ValueError, TypeError):
                qty = 0
            if qty <= 0:
                continue
            nombre = request.form.get(f'nom_{cb}') or ''
            try:
                precio = float(request.form.get(f'pvp_{cb}') or 0)
            except (ValueError, TypeError):
                precio = 0.0
            rotacion = request.form.get(f'rot_{cb}') or None
            try:
                avg = float(request.form.get(f'avg_{cb}') or 0)
            except (ValueError, TypeError):
                avg = 0.0
            items.append(PedidoItem(
                codigo_barra=cb,
                nombre=nombre[:200],
                cantidad=qty,
                precio_pvp=precio,
                subtotal=round(qty * precio, 2),
                rotacion=rotacion,
                avg_monthly=avg or None,
            ))

        if not items:
            flash('No seleccionaste productos con cantidad > 0.')
            return redirect(url_for('purchase_suggest', calcular=1))

        pedido = Pedido(
            laboratorio=laboratorio[:150],
            farmacia='',
            periodo='Sugerido',
            n_days=0,
            items=items,
        )
        session.add(pedido)
        for it in items:
            _upsert_producto(session, it.codigo_barra, it.nombre, float(it.precio_pvp or 0))
        session.commit()
        flash(f'Pedido creado para {laboratorio}: {len(items)} productos.')
        return redirect(url_for('orders_list'))
    except Exception as e:
        session.rollback()
        app.logger.exception('Error en purchase_suggest_create_order')
        flash(f'Error al crear pedido: {e}')
        return redirect(url_for('purchase_suggest', calcular=1))
    finally:
        session.close()


@app.route('/orders')
def orders_list():
    session = database.SessionLocal()
    try:
        pedidos = session.query(Pedido).order_by(Pedido.creado_en.desc()).all()
        result = []
        for p in pedidos:
            total_unidades = sum(it.cantidad for it in p.items)
            total_importe = sum(float(it.subtotal or 0) for it in p.items)
            result.append({
                'id': p.id,
                'laboratorio': p.laboratorio,
                'farmacia': p.farmacia,
                'periodo': p.periodo,
                'n_days': p.n_days,
                'creado_en': p.creado_en.strftime('%d/%m/%Y %H:%M') if p.creado_en else '',
                'analizado_en': p.analizado_en.strftime('%d/%m/%Y') if p.analizado_en else '',
                'estado': p.estado,
                'n_productos': len(p.items),
                'total_unidades': total_unidades,
                'total_importe': total_importe,
                'productos': [
                    {
                        'codigo_barra': it.codigo_barra,
                        'nombre': it.nombre,
                        'cantidad': it.cantidad,
                        'precio_pvp': float(it.precio_pvp or 0),
                        'subtotal': float(it.subtotal or 0),
                    }
                    for it in p.items
                ],
            })
        return render_template('orders_list.html', pedidos=result)
    finally:
        session.close()


@app.route('/order/<int:pedido_id>/delete', methods=['POST'])
def order_delete(pedido_id):
    session = database.SessionLocal()
    try:
        pedido = session.query(Pedido).get(pedido_id)
        if pedido:
            session.delete(pedido)
            session.commit()
            flash('Pedido eliminado.')
    except Exception as e:
        session.rollback()
        flash(f'Error: {e}')
    finally:
        session.close()
    return redirect(url_for('orders_list'))


@app.route('/order/<int:pedido_id>/export/<fmt>')
def order_export_file(pedido_id, fmt):
    """Exporta el pedido guardado a xlsx o pdf."""
    session = database.SessionLocal()
    try:
        pedido = session.query(Pedido).get(pedido_id)
        if not pedido:
            return 'Pedido no encontrado', 404
        items = [{
            'codigo_barra': it.codigo_barra or '',
            'nombre': it.nombre or '',
            'cantidad': it.cantidad or 0,
            'precio_pvp': float(it.precio_pvp or 0),
            'subtotal': float(it.subtotal or 0),
        } for it in pedido.items]
        total_unidades = sum(it['cantidad'] for it in items)
        total_importe = sum(it['subtotal'] for it in items)
        lab = pedido.laboratorio or 'Pedido'
        periodo = pedido.periodo or ''
        n_days = pedido.n_days or 0
        safe_lab = secure_filename(lab) or 'pedido'
    finally:
        session.close()

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO as _BIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pedido'
        ws.append([f'{lab} — {periodo}'])
        ws['A1'].font = Font(bold=True, size=13)
        ws.append([f'{n_days} días'])
        ws.append([])

        headers = ['Cód. Barras', 'Producto', 'P.PVP', 'Cantidad', 'Subtotal']
        ws.append(headers)
        hdr_row = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=hdr_row, column=c)
            cell.fill = PatternFill('solid', fgColor='1C1C1E')
            cell.font = Font(bold=True, color='EAB308')
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 14

        for it in items:
            ws.append([it['codigo_barra'], it['nombre'],
                       it['precio_pvp'], it['cantidad'], it['subtotal']])

        ws.append([])
        ws.append(['', 'Total', '', total_unidades, total_importe])
        tot_row = ws.max_row
        for c in (2, 4, 5):
            ws.cell(row=tot_row, column=c).font = Font(bold=True)

        buf = _BIO()
        wb.save(buf); buf.seek(0)
        resp = make_response(buf.read())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename="Pedido_{safe_lab}.xlsx"'
        return resp

    if fmt == 'pdf':
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfgen import canvas as rl_canvas
        from io import BytesIO as _BIO
        from datetime import datetime as _dt

        fecha_emision = _dt.now().strftime('%d/%m/%Y %H:%M')
        page_w, page_h = A4

        class _NumberedCanvas(rl_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                rl_canvas.Canvas.__init__(self, *args, **kwargs)
                self._pages = []
            def showPage(self):
                self._pages.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                total = len(self._pages)
                for i, state in enumerate(self._pages, 1):
                    self.__dict__.update(state)
                    self.saveState()
                    self.setFont('Helvetica', 7)
                    self.setFillColor(colors.HexColor('#6B7280'))
                    self.drawString(1.5*cm, 0.6*cm, f"Emitido: {fecha_emision}")
                    self.drawRightString(page_w - 1.5*cm, 0.6*cm, f"Página {i} de {total}")
                    self.restoreState()
                    rl_canvas.Canvas.showPage(self)
                rl_canvas.Canvas.save(self)

        buf = _BIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        accent = colors.HexColor('#EAB308')
        txt = colors.HexColor('#1A202C')
        hdr_bg = colors.HexColor('#2D3748')
        row_b = colors.HexColor('#F3F4F6')
        title_s = ParagraphStyle('t', parent=styles['Normal'], fontSize=14, textColor=accent, spaceAfter=4)
        sub_s = ParagraphStyle('s', parent=styles['Normal'], fontSize=9, textColor=txt, spaceAfter=2)
        cell_s = ParagraphStyle('c', parent=styles['Normal'], fontSize=7, textColor=txt, leading=9)

        story = [
            Paragraph(f"Pedido — {lab}", title_s),
            Paragraph(f"{periodo} · {n_days} días", sub_s),
            Paragraph(f"{len(items)} productos · {total_unidades} unidades · ${total_importe:,.0f}".replace(',', '.'), sub_s),
            Spacer(1, 0.4*cm),
        ]
        headers = ['Cód. Barras', 'Producto', 'P.PVP', 'Cantidad', 'Subtotal']
        rows = [headers]
        for it in items:
            rows.append([
                it['codigo_barra'],
                Paragraph(it['nombre'], cell_s),
                f"${it['precio_pvp']:,.0f}".replace(',', '.'),
                it['cantidad'],
                f"${it['subtotal']:,.0f}".replace(',', '.'),
            ])
        rows.append(['', 'TOTAL', '', total_unidades,
                     f"${total_importe:,.0f}".replace(',', '.')])

        t = Table(rows, colWidths=[3.2*cm, 9.5*cm, 2.2*cm, 2.2*cm, 2.4*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, row_b]),
            ('FONTSIZE',   (0, 1), (-1, -1), 7),
            ('ALIGN',      (2, 0), (-1, -1), 'RIGHT'),
            ('ALIGN',      (0, 0), (1, -1), 'LEFT'),
            ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
            ('GRID',       (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(t)
        doc.build(story, canvasmaker=_NumberedCanvas)
        buf.seek(0)
        resp = make_response(buf.read())
        resp.headers['Content-Type'] = 'application/pdf'
        resp.headers['Content-Disposition'] = f'attachment; filename="Pedido_{safe_lab}.pdf"'
        return resp

    return 'Formato inválido', 400


@app.route('/modulo-packs')
def modulo_packs_list():
    session = database.SessionLocal()
    try:
        all_prods = session.query(Producto).order_by(Producto.codigo_barra).all()
        prod_map = {p.codigo_barra: p for p in all_prods}
        labs = session.query(Laboratorio).order_by(Laboratorio.nombre).all()

        def _pack_dict(mp):
            return {'id': mp.id, 'ean_pack': mp.ean_pack, 'ean_unidad': mp.ean_unidad,
                    'cantidad': mp.cantidad,
                    'cant_modulo': mp.cant_modulo,
                    'desc_pct': float(mp.desc_pct) if mp.desc_pct is not None else None,
                    'desc_pack':   mp.descripcion or '',
                    'desc_unidad': (prod_map[mp.ean_unidad].descripcion or '') if mp.ean_unidad in prod_map else '',
                    'prod_unidad_id': prod_map[mp.ean_unidad].id if mp.ean_unidad in prod_map else None,
                    'modulo_id': mp.modulo_id}

        modulos_raw = (session.query(Modulo)
                       .outerjoin(Laboratorio)
                       .order_by(Modulo.lista_nombre, Laboratorio.nombre, Modulo.nombre).all())
        def _lista_nombre(m):
            if m.lista_nombre:
                return m.lista_nombre
            return m.laboratorio.nombre if m.laboratorio else '—'

        modulos = [{'id': m.id, 'nombre': m.nombre,
                    'lab_nombre': m.laboratorio.nombre if m.laboratorio else '—',
                    'lab_id': m.laboratorio_id or 0,
                    'lista_nombre': _lista_nombre(m),
                    'is_lista_marker': bool(m.lista_nombre and m.nombre == m.lista_nombre),
                    'creado_en': m.creado_en.strftime('%d/%m/%Y') if m.creado_en else '',
                    'activo': m.activo,
                    'packs': [_pack_dict(mp) for mp in m.packs]}
                   for m in modulos_raw]

        # Pre-computar activo y toggle_id a nivel de lista
        lista_activo_map = {}    # lista_nombre → bool
        lista_toggle_map = {}    # lista_nombre → modulo_id (preferir is_lista_marker)
        for md in modulos:
            ln = md['lista_nombre']
            if md['activo']:
                lista_activo_map[ln] = True
            if ln not in lista_toggle_map or md['is_lista_marker']:
                lista_toggle_map[ln] = md['id']
        for md in modulos:
            ln = md['lista_nombre']
            md['lista_activo']    = lista_activo_map.get(ln, False)
            md['lista_toggle_id'] = lista_toggle_map.get(ln, md['id'])

        orphan_packs = [_pack_dict(mp) for mp in
                        session.query(ModuloPack).filter(ModuloPack.modulo_id.is_(None))
                        .order_by(ModuloPack.ean_pack).all()]

        prods_pack = [{'ean': p.codigo_barra, 'desc': p.descripcion or ''} for p in all_prods if p.es_pack]
        prods_all  = [{'ean': p.codigo_barra, 'desc': p.descripcion or ''} for p in all_prods]
        return render_template('modulo_packs.html',
                               modulos=modulos, orphan_packs=orphan_packs,
                               labs=[{'id': l.id, 'nombre': l.nombre} for l in labs],
                               prods_pack=prods_pack, prods_all=prods_all)
    finally:
        session.close()


@app.route('/modulo-packs/vista')
def modulo_packs_vista():
    session = database.SessionLocal()
    try:
        prod_map = {p.codigo_barra: p for p in session.query(Producto).all()}
        labs = session.query(Laboratorio).order_by(Laboratorio.nombre).all()
        lab_filter = request.args.get('lab', '').strip()

        q = session.query(Modulo).outerjoin(Laboratorio).order_by(Laboratorio.nombre, Modulo.nombre)
        modulos_raw = q.all()

        modulos = []
        for m in modulos_raw:
            lab_nombre = m.laboratorio.nombre if m.laboratorio else ''
            if lab_filter and lab_nombre != lab_filter:
                continue
            packs = [{'ean_pack': mp.ean_pack,
                      'desc_pack': mp.descripcion or '—',
                      'ean_unidad': mp.ean_unidad,
                      'desc_unidad': (prod_map[mp.ean_unidad].descripcion or '—') if mp.ean_unidad in prod_map else '—',
                      'cantidad': mp.cantidad}
                     for mp in m.packs]
            modulos.append({'id': m.id, 'nombre': m.nombre,
                            'lab_nombre': lab_nombre or '—',
                            'packs': packs})

        return render_template('modulo_packs_vista.html',
                               modulos=modulos,
                               labs=[{'id': l.id, 'nombre': l.nombre} for l in labs],
                               lab_filter=lab_filter)
    finally:
        session.close()


@app.route('/modulo-packs/plantilla')
def modulo_packs_plantilla():
    """Descarga plantilla XLSX para importar módulos."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Módulos'
    hdr_fill = PatternFill('solid', fgColor='1C1C1E')
    hdr_font = Font(bold=True, color='EAB308')
    mod_fill = PatternFill('solid', fgColor='FEF9C3')
    mod_font = Font(bold=True, color='92400E')
    border   = Border(bottom=Side(style='thin', color='D0D0D0'))
    ws.append(['MÓDULOS — plantilla de importación'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    headers = ['NOMBRE MÓDULO', 'EAN PACK', 'DESCRIPCIÓN PACK', 'EAN UNIDAD', 'UNID./PACK']
    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci)
        c.fill = hdr_fill; c.font = hdr_font; c.border = border
    ws.append(['MOD. EJEMPLO'])
    c = ws.cell(row=4, column=1); c.fill = mod_fill; c.font = mod_font
    for ean, desc, ean_u, cant in [('7790001000001','PRODUCTO EJEMPLO 1','7790001000002',10),
                                    ('7790001000003','PRODUCTO EJEMPLO 2','7790001000004',6)]:
        ws.append(['', ean, desc, ean_u, cant])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename="plantilla_modulos.xlsx"'
    return resp


@app.route('/modulo-packs/importar', methods=['POST'])
def modulo_packs_importar():
    """Importa módulos desde un XLSX (formato Roemmers o plantilla propia)."""
    from parsers.modulos_xlsx import parse_modulos_xlsx
    f = request.files.get('file')
    lab_id = request.form.get('lab_id') or None
    lista_nombre = (request.form.get('lista_nombre') or '').strip() or None
    if not f:
        return jsonify({'error': 'No se recibió archivo'}), 400
    tmp = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
    f.save(tmp)
    session = database.SessionLocal()
    try:
        modules = parse_modulos_xlsx(tmp)
        if not modules:
            return jsonify({'error': 'No se encontraron módulos en el archivo'}), 400
        creados = 0
        packs_agregados = 0
        for mod in modules:
            nombre_mod = mod['nombre']
            modulo_actual = session.query(Modulo).filter_by(nombre=nombre_mod, lista_nombre=lista_nombre).first()
            if not modulo_actual:
                modulo_actual = Modulo(nombre=nombre_mod,
                                       laboratorio_id=int(lab_id) if lab_id else None,
                                       lista_nombre=lista_nombre)
                session.add(modulo_actual)
                session.flush()
                creados += 1
            for item in mod['items']:
                ean_pack = item['ean']
                if not ean_pack:
                    continue
                existe = session.query(ModuloPack).filter_by(ean_pack=ean_pack).first()
                if not existe:
                    session.add(ModuloPack(
                        ean_pack=ean_pack,
                        ean_unidad=ean_pack,   # default: mismo EAN; el usuario lo vincula luego
                        cantidad=1,
                        descripcion=item.get('descripcion', ''),
                        cant_modulo=item.get('cant'),
                        desc_pct=item.get('desc_pct'),
                        modulo_id=modulo_actual.id,
                    ))
                    packs_agregados += 1
        session.commit()
        return jsonify({'ok': True, 'modulos_creados': creados, 'packs_agregados': packs_agregados})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()
        try: os.remove(tmp)
        except: pass


@app.route('/modulos/delete-by-lista', methods=['POST'])
def modulos_delete_by_lista():
    """Elimina todos los módulos (y sus packs en cascada) de una lista/importación."""
    data = request.get_json(silent=True) or {}
    lista_nombre = data.get('lista_nombre', '').strip()
    if not lista_nombre:
        return jsonify({'error': 'lista_nombre requerido'}), 400
    session = database.SessionLocal()
    try:
        # Buscar por lista_nombre exacto, o por lab_nombre (fallback para datos viejos)
        modulos = session.query(Modulo).filter(
            (Modulo.lista_nombre == lista_nombre) |
            ((Modulo.lista_nombre.is_(None)) &
             (Modulo.laboratorio.has(database.Laboratorio.nombre == lista_nombre)))
        ).all()
        count = len(modulos)
        for m in modulos:
            session.delete(m)
        session.commit()
        return jsonify({'ok': True, 'eliminados': count})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/modulo-packs/activos')
def modulo_packs_activos():
    """Devuelve las listas activas agrupadas por lista_nombre. ?lab=Nombre filtra por laboratorio."""
    lab_nombre = request.args.get('lab', '').strip()
    session = database.SessionLocal()
    try:
        q = session.query(Modulo).filter_by(activo=True).outerjoin(Laboratorio)
        if lab_nombre:
            q = q.filter(Laboratorio.nombre == lab_nombre)
        raw = q.order_by(Modulo.lista_nombre, Modulo.nombre).all()
        prod_map = {p.codigo_barra: p for p in session.query(Producto).all()}

        # Agrupar por lista_nombre → lista de módulos (packs) con sus productos
        from collections import OrderedDict
        listas = OrderedDict()
        for m in raw:
            ln = m.lista_nombre or m.nombre
            if ln not in listas:
                listas[ln] = {'lista_nombre': ln,
                               'lab_nombre': m.laboratorio.nombre if m.laboratorio else '',
                               'modulos': []}
            # Omitir el marcador de lista (es_lista_marker)
            if m.lista_nombre and m.nombre == m.lista_nombre:
                continue
            packs = [{'ean_pack':    mp.ean_pack,
                      'desc_pack':   mp.descripcion or '',
                      'ean_unidad':  mp.ean_unidad,
                      'desc_unidad': (prod_map[mp.ean_unidad].descripcion or '') if mp.ean_unidad in prod_map else '',
                      'cant_modulo': mp.cant_modulo if mp.cant_modulo is not None else mp.cantidad,
                      'desc_pct':    float(mp.desc_pct) if mp.desc_pct is not None else 0.0}
                     for mp in m.packs]
            listas[ln]['modulos'].append({'id': m.id, 'nombre': m.nombre, 'packs': packs})

        return jsonify({'listas': list(listas.values())})
    finally:
        session.close()


@app.route('/modulo/<int:modulo_id>/toggle-activo', methods=['POST'])
def modulo_toggle_activo(modulo_id):
    session = database.SessionLocal()
    try:
        m = session.get(Modulo, modulo_id)
        if not m:
            return jsonify({'error': 'No encontrado'}), 404
        nuevo_estado = not bool(m.activo)
        if nuevo_estado:
            # Desactivar todos los del mismo laboratorio
            session.query(Modulo).filter(
                Modulo.laboratorio_id == m.laboratorio_id
            ).update({'activo': False})
            session.flush()
        # Activar/desactivar todos los de la misma lista
        if m.lista_nombre:
            session.query(Modulo).filter(
                Modulo.lista_nombre == m.lista_nombre
            ).update({'activo': nuevo_estado})
        else:
            m.activo = nuevo_estado
        session.commit()
        return jsonify({'activo': nuevo_estado})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/modulo/add', methods=['POST'])
def modulo_add():
    data = request.get_json(silent=True) or {}
    nombre = (data.get('nombre') or '').strip()
    lab_id = data.get('laboratorio_id')
    lista_nombre = (data.get('lista_nombre') or nombre or '').strip() or None
    if not nombre:
        return jsonify({'error': 'Nombre requerido'}), 400
    session = database.SessionLocal()
    try:
        m = Modulo(nombre=nombre,
                   laboratorio_id=int(lab_id) if lab_id else None,
                   lista_nombre=lista_nombre)
        session.add(m)
        session.commit()
        lab_nombre = m.laboratorio.nombre if m.laboratorio else '—'
        return jsonify({'ok': True, 'id': m.id, 'nombre': m.nombre,
                        'lab_nombre': lab_nombre,
                        'creado_en': m.creado_en.strftime('%d/%m/%Y') if m.creado_en else ''})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/modulo/<int:modulo_id>/delete', methods=['POST'])
def modulo_delete(modulo_id):
    session = database.SessionLocal()
    try:
        m = session.get(Modulo, modulo_id)
        if m:
            session.delete(m)
            session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/modulo-pack/<int:pack_id>/assign', methods=['POST'])
def modulo_pack_assign(pack_id):
    data = request.get_json(silent=True) or {}
    modulo_id = data.get('modulo_id')
    session = database.SessionLocal()
    try:
        mp = session.get(ModuloPack, pack_id)
        if not mp:
            return jsonify({'error': 'Pack no encontrado'}), 404
        mp.modulo_id = int(modulo_id) if modulo_id else None
        session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/modulo-pack/add', methods=['POST'])
def modulo_pack_add():
    data = request.get_json(silent=True) or {}
    ean_pack   = (data.get('ean_pack') or '').strip()
    ean_unidad = (data.get('ean_unidad') or '').strip()
    cantidad   = int(data.get('cantidad') or 1)
    descripcion = (data.get('descripcion') or '').strip()
    modulo_id  = data.get('modulo_id')
    if not ean_pack or not ean_unidad or cantidad < 1:
        return {'error': 'Datos incompletos'}, 400
    session = database.SessionLocal()
    try:
        existing = session.query(ModuloPack).filter_by(ean_pack=ean_pack).first()
        if existing:
            existing.ean_unidad = ean_unidad
            existing.cantidad = cantidad
            existing.descripcion = descripcion or existing.descripcion
            if modulo_id is not None:
                existing.modulo_id = int(modulo_id) if modulo_id else None
        else:
            session.add(ModuloPack(ean_pack=ean_pack, ean_unidad=ean_unidad,
                                   cantidad=cantidad, descripcion=descripcion,
                                   modulo_id=int(modulo_id) if modulo_id else None))
        session.commit()
        return {'ok': True}
    except Exception as e:
        session.rollback()
        return {'error': str(e)}, 500
    finally:
        session.close()


@app.route('/modulo-pack/<int:pack_id>/update', methods=['POST'])
def modulo_pack_update(pack_id):
    session = database.SessionLocal()
    try:
        data = request.get_json(silent=True) or {}
        mp = session.get(ModuloPack, pack_id)
        if not mp:
            return jsonify({'error': 'No encontrado'}), 404
        if 'ean_pack' in data:
            mp.ean_pack = str(data['ean_pack']).strip()
        if 'descripcion' in data:
            mp.descripcion = str(data['descripcion']).strip() or None
        if 'ean_unidad' in data:
            mp.ean_unidad = str(data['ean_unidad']).strip()
        if 'cantidad' in data:
            mp.cantidad = int(data['cantidad'])
        if 'cant_modulo' in data:
            mp.cant_modulo = int(data['cant_modulo']) if data['cant_modulo'] is not None else None
        if 'desc_pct' in data:
            mp.desc_pct = float(data['desc_pct']) if data['desc_pct'] is not None else None
        session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/modulo-pack/<int:pack_id>/delete', methods=['POST'])
def modulo_pack_delete(pack_id):
    session = database.SessionLocal()
    try:
        mp = session.get(ModuloPack, pack_id)
        if mp:
            session.delete(mp)
            session.commit()
        return {'ok': True}
    finally:
        session.close()


@app.route('/productos')
def productos_list():
    session = database.SessionLocal()
    try:
        labs = [{'id': l.id, 'nombre': l.nombre}
                for l in session.query(Laboratorio).order_by(Laboratorio.nombre).all()]
        return render_template('productos.html', laboratorios=labs)
    finally:
        session.close()


@app.route('/api/productos')
def api_productos():
    session = database.SessionLocal()
    try:
        from sqlalchemy.orm import joinedload
        prods = (session.query(Producto)
                 .options(joinedload(Producto.laboratorio))
                 .order_by(Producto.descripcion).all())
        data = [
            {
                'id': p.id,
                'codigo_barra': p.codigo_barra,
                'descripcion': p.descripcion or '',
                'alt1': p.codigo_barra_alt1 or '',
                'alt2': p.codigo_barra_alt2 or '',
                'alt3': p.codigo_barra_alt3 or '',
                'precio_pvp': float(p.precio_pvp) if p.precio_pvp else None,
                'laboratorio_id': p.laboratorio_id or '',
                'laboratorio_nombre': p.laboratorio.nombre if p.laboratorio else '',
                'actualizado_en': p.actualizado_en.strftime('%d/%m/%Y') if p.actualizado_en else '',
                'es_pack': p.es_pack or 0,
            }
            for p in prods
        ]
        return jsonify(data)
    finally:
        session.close()


@app.route('/order/<int:pedido_id>')
def order_detail(pedido_id):
    session = database.SessionLocal()
    try:
        pedido = session.query(Pedido).get(pedido_id)
        if not pedido:
            flash('Pedido no encontrado.')
            return redirect(url_for('orders_list'))
        data = {
            'id': pedido.id,
            'laboratorio': pedido.laboratorio,
            'farmacia': pedido.farmacia,
            'periodo': pedido.periodo,
            'n_days': pedido.n_days,
            'creado_en': pedido.creado_en.strftime('%d/%m/%Y %H:%M') if pedido.creado_en else '',
        }
        # ErpStock lookup para mostrar stock actual en resumen final
        erp_stock_map = {
            row.codigo_barra: int(row.cantidad or 0)
            for row in session.query(ErpStock).all()
        }
        data['productos'] = [
            {
                'codigo_barra': it.codigo_barra,
                'nombre': it.nombre,
                'cantidad': it.cantidad,
                'precio_pvp': float(it.precio_pvp or 0),
                'subtotal': float(it.subtotal or 0),
                'rotacion': it.rotacion or '',
                'avg_monthly': float(it.avg_monthly) if it.avg_monthly else None,
                'erp_qty': erp_stock_map.get(it.codigo_barra),
            }
            for it in pedido.items
        ]
        # Tabla de equivalencias: lista de grupos de barcodes del mismo producto
        equiv = [
            {
                'barcodes': [b for b in [
                    p.codigo_barra,
                    p.codigo_barra_alt1,
                    p.codigo_barra_alt2,
                    p.codigo_barra_alt3,
                ] if b],
            }
            for p in session.query(Producto).all()
        ]
        cfg = session.query(database.Config).get(1)
        tol_config = {
            'A': float(cfg.rot_alta_tol)  if cfg else 0.0,
            'M': float(cfg.rot_media_tol) if cfg else 0.0,
            'B': float(cfg.rot_baja_tol)  if cfg else 0.0,
        }
        packs = [{'id': mp.id, 'ean_pack': mp.ean_pack, 'ean_unidad': mp.ean_unidad,
                  'cantidad': mp.cantidad, 'descripcion': mp.descripcion or ''}
                 for mp in session.query(ModuloPack).order_by(ModuloPack.ean_pack).all()]
        from datetime import datetime as _dt
        if not pedido.analizado_en:
            pedido.analizado_en = _dt.utcnow()
            session.commit()
            data['analizado_en'] = pedido.analizado_en.strftime('%d/%m/%Y')
        else:
            data['analizado_en'] = pedido.analizado_en.strftime('%d/%m/%Y')
        # Mapa de precios desde tabla Producto (todos los EANs incluyendo alts)
        product_prices = {}
        for p in session.query(Producto).filter(Producto.precio_pvp.isnot(None)).all():
            price = float(p.precio_pvp)
            for bc in [p.codigo_barra, p.codigo_barra_alt1,
                       p.codigo_barra_alt2, p.codigo_barra_alt3]:
                if bc:
                    product_prices[bc] = price
        return render_template('order_detail.html', pedido=data, productos_equiv=equiv,
                               tol_config=tol_config, modulo_packs=packs,
                               product_prices=product_prices)
    finally:
        session.close()


@app.route('/order/<int:pedido_id>/save-module-matches', methods=['POST'])
def order_save_module_matches(pedido_id):
    """Guarda equivalencias EAN-módulo → barcode-pedido en tabla productos."""
    body    = request.get_json(silent=True) or {}
    matches = body.get('matches', []) if isinstance(body, dict) else body
    session = database.SessionLocal()
    try:
        # Determinar laboratorio del pedido para asociar a productos
        pedido = session.get(Pedido, pedido_id)
        lab_id = None
        if pedido and pedido.laboratorio:
            lab_name = pedido.laboratorio.strip()
            lab = session.query(Laboratorio).filter(
                Laboratorio.nombre.ilike(lab_name)
            ).first()
            if not lab:
                lab = Laboratorio(nombre=lab_name)
                session.add(lab)
                session.flush()
            lab_id = lab.id

        saved = 0
        for m in matches:
            module_ean  = str(m.get('module_ean', '')).strip()
            pedido_bc   = str(m.get('pedido_barcode', '')).strip()
            pedido_nom  = m.get('pedido_nombre', '')
            if not module_ean or not pedido_bc or module_ean == pedido_bc:
                continue
            _upsert_producto(session, pedido_bc, pedido_nom, laboratorio_id=lab_id)
            _add_alt_barcode(session, pedido_bc, module_ean)
            saved += 1
        session.commit()
        # Devolver equivalencias actualizadas
        equiv = [
            {'barcodes': [b for b in [
                p.codigo_barra, p.codigo_barra_alt1,
                p.codigo_barra_alt2, p.codigo_barra_alt3,
            ] if b]}
            for p in session.query(Producto).all()
        ]
        return jsonify({'ok': True, 'saved': saved, 'equiv': equiv})
    except Exception as e:
        session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        session.close()


@app.route('/order/<int:pedido_id>/modules-template', methods=['GET'])
def order_modules_template(pedido_id):
    """Descarga una plantilla XLSX lista para completar con módulos."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    session = database.SessionLocal()
    try:
        pedido = session.query(database.Pedido).get(pedido_id)
        lab = pedido.laboratorio if pedido else 'Laboratorio'
    finally:
        session.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Módulos'

    # Estilos
    hdr_fill = PatternFill('solid', fgColor='1C1C1E')
    hdr_font = Font(bold=True, color='EAB308')
    mod_fill = PatternFill('solid', fgColor='FEF9C3')
    mod_font = Font(bold=True, color='92400E')
    thin     = Side(style='thin', color='D0D0D0')
    border   = Border(bottom=Side(style='thin', color='D0D0D0'))
    gray     = Font(color='999999', italic=True)

    # Fila 1: título
    ws.append([f'MÓDULOS {lab.upper()}'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])  # fila vacía

    # Fila 3: encabezados
    headers = ['NOMBRE MÓDULO', 'CÓDIGO EAN', 'DESCRIPCIÓN', 'CANT.', 'DESC. %']
    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci)
        c.fill = hdr_fill
        c.font = hdr_font
        c.border = border

    # Fila 4: ejemplo módulo (cabecera)
    ws.append(['MOD. EJEMPLO A'])
    c = ws.cell(row=4, column=1)
    c.fill = mod_fill; c.font = mod_font

    # Filas 5-6: ítems de ejemplo
    for ean, desc, cant, pct in [
        ('7793450000001', 'PRODUCTO EJEMPLO 1', 2, 7),
        ('7793450000002', 'PRODUCTO EJEMPLO 2', 1, 7),
    ]:
        ws.append(['MOD. EJEMPLO A', ean, desc, cant, pct])

    ws.append([])

    # Segundo módulo de ejemplo
    ws.append(['MOD. EJEMPLO B'])
    c = ws.cell(row=ws.max_row, column=1)
    c.fill = mod_fill; c.font = mod_font
    ws.append(['MOD. EJEMPLO B', '7793450000003', 'PRODUCTO EJEMPLO 3', 3, 10])

    # Anchos de columna
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'plantilla_modulos_{lab.lower().replace(" ", "_")}.xlsx'
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@app.route('/order/<int:pedido_id>/parse-modules', methods=['POST'])
def order_parse_modules(pedido_id):
    from parsers.modulos_xlsx import parse_modulos_xlsx
    f = request.files.get('modules_file')
    if not f or not f.filename:
        return jsonify({'error': 'No se recibió archivo'}), 400
    tmp = os.path.join(UPLOAD_FOLDER, f'mod_{pedido_id}_{secure_filename(f.filename)}')
    f.save(tmp)
    try:
        modules = parse_modulos_xlsx(tmp)
        return jsonify({'modules': modules})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try: os.remove(tmp)
        except: pass


@app.route('/order/<int:pedido_id>/parse-offers', methods=['POST'])
def order_parse_offers(pedido_id):
    from parsers.ofertas_xlsx import parse_ofertas_xlsx
    f = request.files.get('offers_file')
    if not f or not f.filename:
        return jsonify({'error': 'No se recibió archivo'}), 400
    tmp = os.path.join(UPLOAD_FOLDER, f'off_{pedido_id}_{secure_filename(f.filename)}')
    f.save(tmp)
    try:
        ofertas = parse_ofertas_xlsx(tmp)
        return jsonify({'ofertas': ofertas})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try: os.remove(tmp)
        except: pass


@app.route('/order/<int:pedido_id>/export/<step>/<fmt>', methods=['POST'])
def order_export(pedido_id, step, fmt):
    """step: modules | offers | nodeal | summary.  fmt: xlsx | pdf"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    raw = request.form.get('data')
    if not raw:
        return 'Sin datos', 400
    data = json.loads(raw)
    lab  = request.form.get('laboratorio', 'Pedido')
    periodo = request.form.get('periodo', '')

    wb = openpyxl.Workbook()
    ws = wb.active

    # ── Estilos ──────────────────────────────────────────────────────
    hdr_fill  = PatternFill('solid', fgColor='1C1C1E')
    hdr_font  = Font(bold=True, color='EAB308')
    mod_fill  = PatternFill('solid', fgColor='FEF9C3')   # amarillo suave
    mod_font  = Font(bold=True, color='92400E')
    bold      = Font(bold=True)
    thin      = Side(style='thin', color='D0D0D0')
    border    = Border(bottom=thin)
    center    = Alignment(horizontal='center')
    right_al  = Alignment(horizontal='right')

    def hrow(ws, values, fill=hdr_fill, font=hdr_font):
        r = ws.max_row + 1
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = fill; c.font = font; c.border = border

    # ── Título ────────────────────────────────────────────────────────
    ws.append([f'{lab} — {periodo}'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])

    if step == 'modules':
        hrow(ws, ['Módulo', 'EAN', 'Descripción', 'Cant/Módulo', 'Cant.Pedida',
                  'Cant.Calculada', 'Propuesta', 'Saldo', 'Desc%'])
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['C'].width = 38
        for mod in data:
            # Fila de módulo
            r = ws.max_row + 1
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
            c = ws.cell(row=r, column=1, value=f"{mod['nombre']}  ·  {mod.get('modulos_sugeridos', '')} módulos sugeridos")
            c.fill = mod_fill; c.font = mod_font
            # Ítems
            for it in mod.get('items', []):
                ws.append([
                    mod['nombre'],
                    it.get('ean', ''),
                    it.get('descripcion', ''),
                    it.get('cant', ''),
                    it.get('cant_pedida', ''),
                    it.get('cant_calculada', ''),
                    it.get('propuesta', ''),
                    it.get('saldo', ''),
                    it.get('desc_pct', ''),
                ])

    elif step == 'offers':
        hrow(ws, ['EAN', 'Descripción', 'Cant. a pedir'])
        ws.column_dimensions['B'].width = 42
        for it in data:
            ws.append([it.get('ean', ''), it.get('nombre', ''), it.get('cantidad', '')])

    elif step == 'nodeal':
        hrow(ws, ['EAN', 'Descripción', 'Cant. a pedir'])
        ws.column_dimensions['B'].width = 42
        for it in data:
            ws.append([it.get('ean', ''), it.get('nombre', ''), it.get('cantidad', '')])

    elif step == 'summary':
        rows = data if isinstance(data, list) else []

        hrow(ws, ['EAN', 'Producto', 'Stock ERP', 'Rot.', 'Prom.mes',
                  'Precio PVP', 'Cant. módulo', 'Cant. oferta', 'Sin Deal',
                  'Total', 'Cant. pedida', 'Saldo'])
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 6
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 10

        for row in rows:
            saldo = row.get('saldo', '')
            ws.append([
                row.get('ean', ''),
                row.get('nombre', ''),
                row.get('erp_qty', '') if row.get('erp_qty') is not None else '',
                row.get('rotacion', ''),
                row.get('avg_monthly', '') if row.get('avg_monthly') is not None else '',
                row.get('precio_pvp', '') if row.get('precio_pvp') else '',
                row.get('cant_modulo', '') if row.get('cant_modulo') else '',
                row.get('cant_oferta', '') if row.get('cant_oferta') else '',
                row.get('cant_nodeal', '') if row.get('cant_nodeal') else '',
                row.get('total', ''),
                row.get('cant_pedida', ''),
                saldo if saldo != '' else '',
            ])
            # Color saldo: rojo si positivo (falta), verde si negativo (exceso)
            saldo_val = row.get('saldo')
            if saldo_val is not None:
                from openpyxl.styles import PatternFill as _PF
                if saldo_val > 0:
                    ws.cell(row=ws.max_row, column=12).fill = _PF(fill_type='solid', fgColor='FEE2E2')
                elif saldo_val < 0:
                    ws.cell(row=ws.max_row, column=12).fill = _PF(fill_type='solid', fgColor='D1FAE5')

    if fmt == 'xlsx':
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        step_names = {'modules': 'Modulos', 'offers': 'Ofertas',
                      'nodeal': 'SinDeal', 'summary': 'Resumen'}
        fname = f"{lab}_{step_names.get(step, step)}.xlsx"
        resp = make_response(buf.read())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    return 'Formato no soportado', 400


# ─── DESCUENTOS ──────────────────────────────────────────────────────────────

@app.route('/descuentos')
def descuentos_list():
    session = database.SessionLocal()
    campanas = session.query(DescuentoCampana).order_by(DescuentoCampana.creado_en.desc()).all()
    data = [{
        'id': c.id,
        'laboratorio_nombre': c.laboratorio_nombre,
        'proveedor_id': c.proveedor_id,
        'fecha': c.fecha.strftime('%d/%m/%Y') if c.fecha else '',
        'observacion': c.observacion or '',
        'n_modulos': len(c.modulos),
        'n_items': sum(len(m.items) for m in c.modulos),
        'creado_en': c.creado_en.strftime('%d/%m/%Y') if c.creado_en else '',
    } for c in campanas]
    session.close()
    return render_template('descuentos.html', campanas=data)


@app.route('/descuentos/upload', methods=['GET', 'POST'])
def descuento_upload():
    """GET: formulario de carga. POST: parsea xlsx y muestra preview."""
    if request.method == 'GET':
        session = database.SessionLocal()
        proveedores = session.query(database.Provider).order_by(database.Provider.razon_social).all()
        provs = [{'id': p.id, 'razon_social': p.razon_social} for p in proveedores]
        session.close()
        return render_template('descuento_upload.html', proveedores=provs, preview=None)

    # POST: parsear archivo
    f = request.files.get('archivo')
    if not f or not f.filename:
        flash('Seleccioná un archivo.')
        return redirect(url_for('descuento_upload'))

    ext = f.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx', 'xls', 'pdf'):
        flash('Solo se aceptan archivos Excel (.xlsx / .xls) o PDF.')
        return redirect(url_for('descuento_upload'))

    import tempfile, json
    from parsers.descuento_xlsx_parser import parse_descuento

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
    f.save(tmp.name)
    tmp.close()

    try:
        modulos = parse_descuento(tmp.name)
    except Exception as e:
        flash(f'Error al leer el archivo: {e}')
        return redirect(url_for('descuento_upload'))
    finally:
        import os as _os; _os.unlink(tmp.name)

    if not modulos:
        flash('El archivo no contiene módulos reconocibles.')
        return redirect(url_for('descuento_upload'))

    session = database.SessionLocal()
    proveedores = session.query(database.Provider).order_by(database.Provider.razon_social).all()
    provs = [{'id': p.id, 'razon_social': p.razon_social} for p in proveedores]
    session.close()

    return render_template('descuento_upload.html',
                           proveedores=provs,
                           preview=modulos,
                           preview_json=json.dumps(modulos))


@app.route('/descuentos/campana/guardar', methods=['POST'])
def descuento_campana_guardar():
    """Guarda la campaña completa desde el preview."""
    import json
    from datetime import date as _date

    proveedor_id = request.form.get('proveedor_id', '').strip()
    lab_nombre = request.form.get('laboratorio_nombre', '').strip()
    fecha_str = request.form.get('fecha', '').strip()
    observacion = request.form.get('observacion', '').strip()
    preview_json = request.form.get('preview_json', '[]')

    if not lab_nombre:
        flash('Ingresá el nombre del laboratorio.')
        return redirect(url_for('descuento_upload'))

    try:
        modulos_data = json.loads(preview_json)
    except (ValueError, TypeError):
        flash('Datos inválidos, volvé a subir el archivo.')
        return redirect(url_for('descuento_upload'))

    fecha = None
    if fecha_str:
        try:
            from datetime import datetime as _dt
            fecha = _dt.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    session = database.SessionLocal()
    campana = DescuentoCampana(
        proveedor_id=int(proveedor_id) if proveedor_id else None,
        laboratorio_nombre=lab_nombre,
        fecha=fecha or _date.today(),
        observacion=observacion or None,
    )
    session.add(campana)
    session.flush()

    for mod_data in modulos_data:
        modulo = DescuentoModulo(
            campana_id=campana.id,
            nombre=mod_data.get('nombre', 'SIN NOMBRE'),
            activo=1,
        )
        session.add(modulo)
        session.flush()
        for it in mod_data.get('items', []):
            session.add(DescuentoModuloItem(
                modulo_id=modulo.id,
                codigo_ean=str(it.get('ean', '')),
                descripcion=it.get('descripcion', '') or None,
                cantidad=int(it.get('cantidad', 1)),
                descuento=float(it.get('descuento', 0)),
                es_principal=1 if it.get('es_principal') else 0,
            ))

    session.commit()
    campana_id = campana.id
    session.close()
    flash(f'Campaña guardada con {len(modulos_data)} módulos.')
    return redirect(url_for('descuento_campana', campana_id=campana_id))


@app.route('/descuentos/campana/<int:campana_id>')
def descuento_campana(campana_id):
    session = database.SessionLocal()
    c = session.get(DescuentoCampana, campana_id)
    if not c:
        session.close()
        flash('Campaña no encontrada.')
        return redirect(url_for('descuentos_list'))
    campana = {
        'id': c.id,
        'laboratorio_nombre': c.laboratorio_nombre,
        'proveedor_id': c.proveedor_id,
        'proveedor_nombre': c.proveedor.razon_social if c.proveedor else None,
        'fecha': c.fecha.strftime('%d/%m/%Y') if c.fecha else '',
        'fecha_iso': c.fecha.strftime('%Y-%m-%d') if c.fecha else '',
        'observacion': c.observacion or '',
        'creado_en': c.creado_en.strftime('%d/%m/%Y') if c.creado_en else '',
    }
    modulos = [{
        'id': m.id,
        'nombre': m.nombre or '—',
        'activo': m.activo,
        'n_items': len(m.items),
    } for m in c.modulos]
    proveedores = session.query(database.Provider).order_by(database.Provider.razon_social).all()
    provs = [{'id': p.id, 'razon_social': p.razon_social} for p in proveedores]
    session.close()
    return render_template('descuento_campana.html',
                           campana=campana, modulos=modulos, proveedores=provs)


@app.route('/descuentos/campana/<int:campana_id>/edit', methods=['POST'])
def descuento_campana_edit(campana_id):
    session = database.SessionLocal()
    c = session.get(DescuentoCampana, campana_id)
    if c:
        proveedor_id = request.form.get('proveedor_id', '').strip()
        lab_nombre = request.form.get('laboratorio_nombre', '').strip()
        fecha_str = request.form.get('fecha', '').strip()
        c.proveedor_id = int(proveedor_id) if proveedor_id else None
        if lab_nombre:
            c.laboratorio_nombre = lab_nombre
        c.observacion = request.form.get('observacion', '').strip() or None
        if fecha_str:
            try:
                from datetime import datetime as _dt
                c.fecha = _dt.strptime(fecha_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        session.commit()
    session.close()
    return redirect(url_for('descuento_campana', campana_id=campana_id))


@app.route('/descuentos/campana/<int:campana_id>/delete', methods=['POST'])
def descuento_campana_delete(campana_id):
    session = database.SessionLocal()
    c = session.get(DescuentoCampana, campana_id)
    if c:
        session.delete(c)
        session.commit()
    session.close()
    flash('Campaña eliminada.')
    return redirect(url_for('descuentos_list'))


@app.route('/descuentos/modulo/<int:modulo_id>')
def descuento_detalle(modulo_id):
    session = database.SessionLocal()
    m = session.get(DescuentoModulo, modulo_id)
    if not m:
        session.close()
        flash('Módulo no encontrado.')
        return redirect(url_for('descuentos_list'))
    items = [{
        'id': it.id,
        'codigo_ean': it.codigo_ean,
        'descripcion': it.descripcion or '',
        'cantidad': it.cantidad,
        'descuento': float(it.descuento),
        'es_principal': bool(it.es_principal),
    } for it in m.items]
    modulo = {
        'id': m.id,
        'nombre': m.nombre or '—',
        'activo': m.activo,
        'campana_id': m.campana_id,
    }
    session.close()
    return render_template('descuento_detalle.html', modulo=modulo, items=items)


@app.route('/descuentos/modulo/<int:modulo_id>/item', methods=['POST'])
def descuento_add_item(modulo_id):
    session = database.SessionLocal()
    m = session.get(DescuentoModulo, modulo_id)
    if not m:
        session.close()
        flash('Módulo no encontrado.')
        return redirect(url_for('descuentos_list'))
    try:
        ean = request.form.get('codigo_ean', '').strip()
        if not ean:
            flash('El código EAN es obligatorio.')
            return redirect(url_for('descuento_detalle', modulo_id=modulo_id))
        session.add(DescuentoModuloItem(
            modulo_id=modulo_id,
            codigo_ean=ean,
            descripcion=request.form.get('descripcion', '').strip() or None,
            cantidad=max(1, int(request.form.get('cantidad', 1))),
            descuento=float(request.form.get('descuento', 0)),
            es_principal=1 if request.form.get('es_principal') else 0,
        ))
        session.commit()
    except (ValueError, TypeError):
        flash('Datos inválidos.')
    session.close()
    return redirect(url_for('descuento_detalle', modulo_id=modulo_id))


@app.route('/descuentos/modulo/<int:modulo_id>/item/<int:item_id>/delete', methods=['POST'])
def descuento_delete_item(modulo_id, item_id):
    session = database.SessionLocal()
    item = session.get(DescuentoModuloItem, item_id)
    if item and item.modulo_id == modulo_id:
        session.delete(item)
        session.commit()
    session.close()
    return redirect(url_for('descuento_detalle', modulo_id=modulo_id))


@app.route('/descuentos/modulo/<int:modulo_id>/toggle', methods=['POST'])
def descuento_toggle(modulo_id):
    session = database.SessionLocal()
    m = session.get(DescuentoModulo, modulo_id)
    if m:
        m.activo = 0 if m.activo else 1
        session.commit()
        campana_id = m.campana_id
    session.close()
    if campana_id:
        return redirect(url_for('descuento_campana', campana_id=campana_id))
    return redirect(url_for('descuentos_list'))


@app.route('/descuentos/modulo/<int:modulo_id>/delete', methods=['POST'])
def descuento_delete(modulo_id):
    session = database.SessionLocal()
    m = session.get(DescuentoModulo, modulo_id)
    campana_id = m.campana_id if m else None
    if m:
        session.delete(m)
        session.commit()
    session.close()
    flash('Módulo eliminado.')
    if campana_id:
        return redirect(url_for('descuento_campana', campana_id=campana_id))
    return redirect(url_for('descuentos_list'))


@app.route('/descuentos/upload-libre', methods=['GET', 'POST'])
def descuento_upload_libre():
    """Importación libre: tabla plana LAB | EAN | DESC | CANT | DTO%"""
    session = database.SessionLocal()
    proveedores = session.query(database.Provider).order_by(database.Provider.razon_social).all()
    provs = [{'id': p.id, 'razon_social': p.razon_social} for p in proveedores]
    session.close()

    if request.method == 'GET':
        return render_template('descuento_upload_libre.html', proveedores=provs, preview=None)

    f = request.files.get('archivo')
    if not f or not f.filename:
        flash('Seleccioná un archivo.')
        return redirect(url_for('descuento_upload_libre'))

    ext = f.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx', 'xls'):
        flash('Solo se aceptan archivos Excel (.xlsx / .xls).')
        return redirect(url_for('descuento_upload_libre'))

    import tempfile, json
    from parsers.descuento_libre_parser import parse_descuento_libre

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
    f.save(tmp.name)
    tmp.close()

    try:
        items = parse_descuento_libre(tmp.name)
    except Exception as e:
        flash(f'Error al leer el archivo: {e}')
        return redirect(url_for('descuento_upload_libre'))
    finally:
        import os as _os; _os.unlink(tmp.name)

    if not items:
        flash('El archivo no contiene artículos reconocibles.')
        return redirect(url_for('descuento_upload_libre'))

    return render_template('descuento_upload_libre.html',
                           proveedores=provs,
                           preview=items,
                           preview_json=json.dumps(items))


@app.route('/descuentos/libre/guardar', methods=['POST'])
def descuento_libre_guardar():
    """Guarda la importación libre como una campaña con un módulo por lab."""
    import json
    from datetime import datetime as _dt

    proveedor_id = request.form.get('proveedor_id') or None
    observacion  = request.form.get('observacion', '').strip()
    fecha_str    = request.form.get('fecha', '').strip()
    preview_json = request.form.get('preview_json', '[]')

    try:
        items = json.loads(preview_json)
    except (ValueError, TypeError):
        flash('Datos inválidos, volvé a subir el archivo.')
        return redirect(url_for('descuento_upload_libre'))

    fecha = None
    if fecha_str:
        try:
            fecha = _dt.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Agrupar por lab
    from collections import defaultdict
    grupos = defaultdict(list)
    for it in items:
        grupos[it['lab'] or 'SIN LAB'].append(it)

    session = database.SessionLocal()
    try:
        # Una campaña por cada lab distinto
        campanas_creadas = 0
        for lab_nombre, lab_items in grupos.items():
            prov_id = int(proveedor_id) if proveedor_id else None
            campana = database.DescuentoCampana(
                proveedor_id=prov_id,
                laboratorio_nombre=lab_nombre,
                fecha=fecha,
                observacion=observacion,
            )
            session.add(campana)
            session.flush()

            modulo = database.DescuentoModulo(
                campana_id=campana.id,
                nombre=f'Importación {lab_nombre}',
                codigo=None,
                laboratorio=lab_nombre,
                descuento_default=0,
            )
            session.add(modulo)
            session.flush()

            for it in lab_items:
                item = database.DescuentoModuloItem(
                    modulo_id=modulo.id,
                    codigo_barra=it['ean'],
                    descripcion=it['descripcion'],
                    cantidad=it['cantidad'] or 0,
                    descuento=it['descuento'] or 0,
                    es_principal=0,
                )
                session.add(item)
            campanas_creadas += 1

        session.commit()
        flash(f'{campanas_creadas} campaña(s) importada(s) correctamente.')
    except Exception as e:
        session.rollback()
        flash(f'Error al guardar: {e}')
    finally:
        session.close()

    return redirect(url_for('descuentos_list'))


@app.route('/dashboard')
def dashboard():
    from sqlalchemy import func as _func, case as _case
    from datetime import date as _date, timedelta as _td
    try:
        n_days = max(1, min(365, int(request.args.get('n_days', 10))))
    except (ValueError, TypeError):
        n_days = 10
    lab_filter = (request.args.get('laboratorio') or '').strip()
    rot_filter = (request.args.get('rotacion') or '').strip().upper()
    if rot_filter not in ('A', 'M', 'B'):
        rot_filter = ''
    q_text = (request.args.get('q') or '').strip()
    only_sin_mov = request.args.get('sin_mov') == '1'

    session = database.SessionLocal()
    try:
        PA = database.ProductAnalytics
        # cobertura = stock / (avg_monthly/30); usamos case para evitar división por cero
        cobertura_expr = _case(
            (PA.avg_monthly == 0, None),
            else_=(PA.stock * 30.0 / PA.avg_monthly)
        )
        q = session.query(PA, cobertura_expr.label('cobertura'))
        if lab_filter:
            q = q.filter(PA.laboratorio == lab_filter)
        if rot_filter:
            q = q.filter(PA.rotacion == rot_filter)
        if q_text:
            like = f'%{q_text}%'
            q = q.filter((PA.descripcion.ilike(like)) | (PA.codigo_barra.ilike(like)))
        if only_sin_mov:
            q_alerts = q.filter(PA.sin_mov_60d == 1).order_by(PA.descripcion.asc())
        else:
            q_alerts = q.filter(PA.avg_monthly > 0, PA.stock * 30.0 / PA.avg_monthly < n_days)
            q_alerts = q_alerts.order_by(cobertura_expr.asc())
        alerts = q_alerts.limit(200).all()

        labs = [row[0] for row in session.query(PA.laboratorio)
                .filter(PA.laboratorio.isnot(None))
                .distinct().order_by(PA.laboratorio).all()]

        total_products = session.query(_func.count(PA.codigo_barra)).scalar() or 0
        alerts_count = session.query(_func.count(PA.codigo_barra)).filter(
            PA.avg_monthly > 0, PA.stock * 30.0 / PA.avg_monthly < n_days
        ).scalar() or 0
        sin_mov_count = session.query(_func.count(PA.codigo_barra)).filter(
            PA.sin_mov_60d == 1
        ).scalar() or 0
        claims_open = session.query(_func.count(database.Claim.id)).filter(
            database.Claim.estado == 'ABIERTO'
        ).scalar() or 0
        first_of_month = _date.today().replace(day=1)
        invoices_month = session.query(_func.count(database.Invoice.id)).filter(
            database.Invoice.fecha >= first_of_month
        ).scalar() or 0

        codigos = [pa.codigo_barra for pa, _ in alerts]
        ultima_compra_map = {}
        if codigos:
            rows = session.query(database.Producto.codigo_barra, database.Producto.ultima_compra)\
                .filter(database.Producto.codigo_barra.in_(codigos)).all()
            ultima_compra_map = {cb: fc for cb, fc in rows if fc}

        alert_rows = [{
            'codigo_barra': pa.codigo_barra,
            'descripcion': pa.descripcion,
            'laboratorio': pa.laboratorio,
            'stock': pa.stock,
            'avg_monthly': float(pa.avg_monthly or 0),
            'rotacion': pa.rotacion,
            'cobertura': round(cov, 1) if cov is not None else None,
            'precio_pvp': float(pa.precio_pvp or 0),
            'ultima_compra': ultima_compra_map.get(pa.codigo_barra),
            'sin_mov_60d': bool(pa.sin_mov_60d),
        } for pa, cov in alerts]

        ultima_act = session.query(_func.max(PA.actualizado_en)).scalar()

        base_q = session.query(PA)
        if lab_filter:
            base_q = base_q.filter(PA.laboratorio == lab_filter)
        top_qty_rows = base_q.order_by(PA.avg_monthly.desc()).limit(10).all()
        top_qty = [{
            'nombre': (p.descripcion or p.codigo_barra or '')[:40],
            'valor': float(p.avg_monthly or 0),
        } for p in top_qty_rows]

        valor_expr = PA.avg_monthly * PA.precio_pvp
        top_val_rows = base_q.order_by(valor_expr.desc()).limit(10).all()
        top_val = [{
            'nombre': (p.descripcion or p.codigo_barra or '')[:40],
            'valor': float(p.avg_monthly or 0) * float(p.precio_pvp or 0),
        } for p in top_val_rows]

        # Pérdida diaria estimada por productos sin stock (stock <= 0)
        # Pérdida = (avg_monthly/30) * precio_pvp
        loss_expr = (PA.avg_monthly / 30.0) * PA.precio_pvp
        loss_rows = base_q.filter(PA.stock <= 0, PA.avg_monthly > 0)\
            .order_by(loss_expr.desc()).limit(10).all()
        top_loss = [{
            'nombre': (p.descripcion or p.codigo_barra or '')[:40],
            'valor': float(p.avg_monthly or 0) / 30.0 * float(p.precio_pvp or 0),
        } for p in loss_rows]

        # Capital inmovilizado (stock * pvp) y stock muerto (sin_mov_60d)
        capital_expr = PA.stock * PA.precio_pvp
        capital_q = session.query(_func.coalesce(_func.sum(capital_expr), 0))
        muerto_q = session.query(_func.coalesce(_func.sum(capital_expr), 0))\
            .filter(PA.sin_mov_60d == 1)
        if lab_filter:
            capital_q = capital_q.filter(PA.laboratorio == lab_filter)
            muerto_q = muerto_q.filter(PA.laboratorio == lab_filter)
        capital_total = float(capital_q.scalar() or 0)
        stock_muerto_total = float(muerto_q.scalar() or 0)

        # Top 10 stock muerto valorizado
        muerto_rows = base_q.filter(PA.sin_mov_60d == 1, PA.stock > 0)\
            .order_by(capital_expr.desc()).limit(10).all()
        top_muerto = [{
            'nombre': (p.descripcion or p.codigo_barra or '')[:40],
            'valor': float(p.stock or 0) * float(p.precio_pvp or 0),
        } for p in muerto_rows]
    finally:
        session.close()

    return render_template('dashboard.html',
                           n_days=n_days,
                           lab_filter=lab_filter,
                           rot_filter=rot_filter,
                           q_text=q_text,
                           only_sin_mov=only_sin_mov,
                           labs=labs,
                           alerts=alert_rows,
                           total_products=total_products,
                           alerts_count=alerts_count,
                           sin_mov_count=sin_mov_count,
                           claims_open=claims_open,
                           invoices_month=invoices_month,
                           ultima_act=ultima_act,
                           top_qty=top_qty,
                           top_val=top_val,
                           top_loss=top_loss,
                           top_muerto=top_muerto,
                           capital_total=capital_total,
                           stock_muerto_total=stock_muerto_total)


@app.route('/dashboard/help')
def dashboard_help():
    return render_template('dashboard_help.html')


if __name__ == '__main__':
    app.run(debug=True)
