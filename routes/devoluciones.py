"""Devoluciones / rendición de recetas: cuando la OS rechaza una receta
presentada o falta info, registramos motivo, observaciones y auditoría.

Flow:
  /rend-recetas                  → listado con banner de alertas según rol
  /rend-recetas/buscar           → form: vendedor + OS + rango fechas + lote
  /rend-recetas/buscar (POST)    → tabla con recetas + form inline
  /rend-recetas/guardar (POST)   → persiste las recetas marcadas
  /rend-recetas/motivos          → ABM motivos (filtrado por uso_rol)
  /rend-recetas/lotes            → ABM lotes de rendición (modelo propio)
  /rend-recetas/filtros-os       → ABM filtros OS por rol (lista negra)
  /rend-recetas/<id>/auditar     → 2da etapa: chequeo del auditor
"""
from datetime import date, datetime, timedelta
from decimal import Decimal

from flask import flash, jsonify, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from sqlalchemy import or_
from sqlalchemy.orm import joinedload

import database
import observer_source

# Carga incremental de recetas (ver devoluciones_buscar):
# - Al rendir traemos TODO desde la última carga del vendedor hasta hoy.
# - MARGEN_REZAGADAS_DIAS: retrocedemos N días de la última op para pescar
#   recetas que ObServer registró tarde (rezagadas). El dedup por
#   id_operacion (badge "ya en rend #N") evita re-cargar las que ya están.
# - DEFAULT_PRIMERA_VEZ_DIAS: cuando el vendedor nunca cargó (sin bookmark),
#   arrancamos N días atrás.
MARGEN_REZAGADAS_DIAS = 20
DEFAULT_PRIMERA_VEZ_DIAS = 30


def _estado_consolidado(n_total, n_audit, n_pend, entregada):
    """Devuelve dict {label, color, prio} con el estado más relevante.
    Lógica jerárquica: gana el más urgente."""
    if n_total == 0:
        return {'label': 'VACÍA', 'color': 'gray', 'prio': 0}
    if entregada and n_pend == 0 and n_audit == n_total:
        return {'label': '✓ COMPLETA', 'color': 'green-dark', 'prio': 1}
    if entregada:
        return {'label': 'ENTREGADA', 'color': 'green', 'prio': 2}
    # No entregada — depende de qué falta
    n_pend_sin_audit = 0   # no podemos saberlo sin SQL aparte
    # Heurística: si hay pendientes y NO todo auditado, prio "SIN AUDITAR";
    # si todo auditado y aún hay pendientes, prio "A RESOLVER";
    # si no hay pendientes (todas resueltas/descartadas), "LISTA P/ENTREGAR".
    if n_pend > 0:
        if n_audit < n_total:
            return {'label': f'SIN AUDITAR ({n_total - n_audit})', 'color': 'purple', 'prio': 5}
        return {'label': f'A RESOLVER ({n_pend})', 'color': 'orange', 'prio': 4}
    # n_pend == 0 → todas resueltas o descartadas
    return {'label': 'LISTA P/ENTREGAR', 'color': 'blue', 'prio': 3}


def _detectar_duplicados():
    """Detecta DevolucionReceta con id_operacion_observer repetido en
    distintos lotes. Retorna lista de grupos con winner + perdedoras.

    Reglas:
      - Si TODAS las copias están en lotes cerrados → skip (no tocar).
      - Si hay copia en lote cerrado + abierto → cerrado gana.
      - Si todas en lotes abiertos → la más vieja gana (primera carga).
      - Filas estado='descartada' se ignoran (ya están out).
    """
    from sqlalchemy import func as _fdd
    grupos = []
    with database.get_db() as session:
        # 1) Encontrar id_operacion_observer con count > 1 (excluyendo descartadas)
        D = database.DevolucionReceta
        dup_ids = (session.query(D.id_operacion_observer)
                   .filter(D.estado != 'descartada')
                   .group_by(D.id_operacion_observer)
                   .having(_fdd.count(D.id) > 1)
                   .all())
        if not dup_ids:
            return []
        ids_set = [r[0] for r in dup_ids]
        # 2) Traer todas las filas de esos ids
        from sqlalchemy.orm import joinedload as _jl
        rows = (session.query(D)
                .options(_jl(D.motivo))
                .filter(D.id_operacion_observer.in_(ids_set))
                .filter(D.estado != 'descartada')
                .order_by(D.id_operacion_observer, D.creado_en)
                .all())
        # Lote map
        lote_ids = list({r.rendicion_lote_id for r in rows if r.rendicion_lote_id})
        lotes_map = {}
        if lote_ids:
            for l in (session.query(database.RendicionLote)
                      .filter(database.RendicionLote.id.in_(lote_ids)).all()):
                lotes_map[l.id] = l
        # Agrupar por id_operacion
        por_op = {}
        for r in rows:
            por_op.setdefault(r.id_operacion_observer, []).append(r)
        for op_id, lista in por_op.items():
            # Anotar lote info
            for r in lista:
                l = lotes_map.get(r.rendicion_lote_id)
                r._lote_estado = (l.estado if l else 'sin-lote')
                r._lote_nro = (l.nro if l else None)
            cerradas = [r for r in lista if r._lote_estado == 'cerrada']
            abiertas = [r for r in lista if r._lote_estado != 'cerrada']
            skip = False
            winner = None
            perdedoras = []
            if len(cerradas) >= 2:
                # Todas cerradas o mix con >1 cerrada → no se tocan
                skip = True
                winner = cerradas[0]
                perdedoras = cerradas[1:] + abiertas
            elif len(cerradas) == 1:
                winner = cerradas[0]
                perdedoras = abiertas
            else:
                # Todas abiertas → más vieja gana
                lista.sort(key=lambda r: r.creado_en)
                winner = lista[0]
                perdedoras = lista[1:]
            def _info(r):
                return {
                    'id': r.id, 'lote_id': r.rendicion_lote_id,
                    'lote_nro': r._lote_nro, 'lote_estado': r._lote_estado,
                    'creado_en': r.creado_en.strftime('%d/%m/%y %H:%M') if r.creado_en else '',
                    'vendedor': r.vendedor_nombre, 'os': r.obra_social_nombre,
                    'motivo': r.motivo.nombre if r.motivo else None,
                    'estado': r.estado,
                }
            grupos.append({
                'op_id': op_id,
                'winner': _info(winner),
                'perdedoras': [_info(r) for r in perdedoras],
                'skip': skip,
            })
        # Detach (las relationships pueden fallar fuera de la sesión)
    return grupos


def init_app(app):

    # ──────────────────────────────────────────────────────────────────
    # Atajo: /rend → /devoluciones/buscar (para shortcut de escritorio)
    # ──────────────────────────────────────────────────────────────────
    @app.route('/rend')
    @login_required
    def rend_alias():
        return redirect(url_for('devoluciones_buscar'))

    # ──────────────────────────────────────────────────────────────────
    # Vista del auditor: agrupado por vendedor → rendiciones (colapsable)
    # ──────────────────────────────────────────────────────────────────
    @app.route('/rend-recetas/por-vendedor')
    @login_required
    def devoluciones_por_vendedor():
        """Vista agrupada por vendedor. Cada vendedor expande para mostrar
        sus rendiciones (lotes) con conteos y estado. Pensada para auditor.
        Click en una rendición → /rend-recetas?presentacion=N (vista plana).
        """
        from sqlalchemy import case
        from sqlalchemy import func as _f
        D = database.DevolucionReceta
        L = database.RendicionLote
        with database.get_db() as session:
            # Para rol rendicion: solo sus propias rendiciones.
            nombre_u = (getattr(current_user, 'nombre_completo', '') or '').strip().upper()
            rol = getattr(current_user, 'rol', None)

            # Agregado por (vendedor, nro_presentacion).
            base = (session.query(
                D.vendedor_nombre.label('vendedor'),
                D.nro_presentacion.label('nro'),
                D.rendicion_lote_id.label('lote_id'),
                _f.count(D.id).label('n_total'),
                _f.sum(case((D.auditor_motivo_id.isnot(None), 1), else_=0)).label('n_audit'),
                _f.sum(case((D.estado == 'pendiente', 1), else_=0)).label('n_pend'),
                _f.sum(case((D.estado == 'resuelta', 1), else_=0)).label('n_res'),
                _f.sum(case((D.estado == 'descartada', 1), else_=0)).label('n_desc'),
                _f.sum(D.importe_a_cargo_os).label('imp_os'),
                _f.max(D.creado_en).label('ult_carga'),
            ).filter(D.nro_presentacion.isnot(None))
             .group_by(D.vendedor_nombre, D.nro_presentacion, D.rendicion_lote_id))
            if rol == 'rendicion' and nombre_u:
                base = base.filter(_f.upper(D.vendedor_nombre) == nombre_u)
            rows = base.order_by(D.vendedor_nombre,
                                 _f.max(D.creado_en).desc()).all()

            # Lote map para sacar estado + entregada.
            lote_ids = list({r.lote_id for r in rows if r.lote_id})
            lotes_map = {}
            if lote_ids:
                for l in (session.query(L).filter(L.id.in_(lote_ids)).all()):
                    lotes_map[l.id] = l

            # Agrupar por vendedor.
            por_vendedor = {}
            for r in rows:
                v = r.vendedor or '—'
                if v not in por_vendedor:
                    por_vendedor[v] = {
                        'vendedor': v,
                        'rendiciones': [],
                        'n_rendiciones': 0,
                        'n_total': 0,
                        'n_audit': 0,
                        'n_pend': 0,
                    }
                lote = lotes_map.get(r.lote_id)
                entregada_flag = bool(lote.entregada) if lote else False
                n_total_int = int(r.n_total or 0)
                n_audit_int = int(r.n_audit or 0)
                n_pend_int = int(r.n_pend or 0)
                rend = {
                    'nro': r.nro,
                    'lote_id': r.lote_id,
                    'lote_estado': (lote.estado if lote else 'sin-lote'),
                    'entregada': entregada_flag,
                    'n_total': n_total_int,
                    'n_audit': n_audit_int,
                    'n_pend': n_pend_int,
                    'n_res': int(r.n_res or 0),
                    'n_desc': int(r.n_desc or 0),
                    'imp_os': float(r.imp_os or 0),
                    'ult_carga': r.ult_carga,
                    'todo_auditado': n_audit_int == n_total_int and n_total_int > 0,
                    'estado': _estado_consolidado(n_total_int, n_audit_int,
                                                   n_pend_int, entregada_flag),
                }
                por_vendedor[v]['rendiciones'].append(rend)
                por_vendedor[v]['n_rendiciones'] += 1
                por_vendedor[v]['n_total'] += rend['n_total']
                por_vendedor[v]['n_audit'] += rend['n_audit']
                por_vendedor[v]['n_pend'] += rend['n_pend']

            # Calcular peor estado por vendedor (rollup) — el de mayor prio
            # entre sus rendiciones. + stock breakdown para auditor.
            for v in por_vendedor.values():
                if v['rendiciones']:
                    peor = max(v['rendiciones'], key=lambda r: r['estado']['prio'])
                    v['peor_estado'] = peor['estado']
                else:
                    v['peor_estado'] = {'label': 'VACÍA', 'color': 'gray', 'prio': 0}
                # Stock por estado del lote (lo ve el auditor).
                # - en_proceso: recetas en lotes abiertos no entregados
                # - pend_entrega: recetas en lotes cerrados pero NO entregados
                # - entregadas: recetas en lotes ya entregados
                v['stock_en_proceso'] = sum(r['n_total'] for r in v['rendiciones']
                                            if r['lote_estado'] == 'abierta' and not r['entregada'])
                v['stock_pend_entrega'] = sum(r['n_total'] for r in v['rendiciones']
                                              if r['lote_estado'] == 'cerrada' and not r['entregada'])
                v['stock_entregadas'] = sum(r['n_total'] for r in v['rendiciones']
                                            if r['entregada'])
            # Orden: peor estado arriba (mayor prio = más urgente).
            vendedores = sorted(por_vendedor.values(),
                                key=lambda v: (-v['peor_estado']['prio'], v['vendedor']))
        return render_template('devoluciones_por_vendedor.html',
                               vendedores=vendedores, rol_actual=rol)

    # ──────────────────────────────────────────────────────────────────
    # ABM Lotes de Rendición (modelo propio)
    # ──────────────────────────────────────────────────────────────────
    def _vendedor_sugerido_actual(vendedores):
        """Retorna (uuid, nombre) del vendedor que matchea con
        current_user.nombre_completo. None si no matchea."""
        nombre_user = (getattr(current_user, 'nombre_completo', '') or '').strip().upper()
        if not nombre_user:
            return None, None
        for v in vendedores:
            v_nombre = (v.get('nombre') or '').strip().upper()
            if v_nombre == nombre_user or v_nombre.startswith(nombre_user) or nombre_user in v_nombre:
                return v['id_usuario'], v['nombre']
        return None, None

    @app.route('/rend-recetas/lotes', methods=['GET'])
    @login_required
    def rendicion_lotes_list():
        """Listado de lotes de rendición. Para rol=rendicion: solo los suyos."""
        rol = getattr(current_user, 'rol', None)
        nombre_user = (getattr(current_user, 'nombre_completo', '') or '').strip().upper()
        with database.get_db() as session:
            from sqlalchemy import func as _f
            D = database.DevolucionReceta
            base = session.query(database.RendicionLote)
            if rol == 'rendicion' and nombre_user:
                base = base.filter(_f.upper(database.RendicionLote.vendedor_nombre) == nombre_user)
            lotes = base.order_by(database.RendicionLote.creado_en.desc()).all()
            # Conteos por lote
            conteos = dict(session.query(
                D.rendicion_lote_id, _f.count(D.id)
            ).filter(D.rendicion_lote_id.isnot(None))
             .group_by(D.rendicion_lote_id).all())
            conteos_audit = dict(session.query(
                D.rendicion_lote_id, _f.count(D.id)
            ).filter(D.rendicion_lote_id.isnot(None),
                     D.auditor_motivo_id.isnot(None))
             .group_by(D.rendicion_lote_id).all())
            items = []
            from datetime import datetime
            now = datetime.utcnow()
            for lote in lotes:
                n_rec = conteos.get(lote.id, 0)
                # Para rol rendicion: esconder lotes vacíos (sin recetas).
                # Admin/auditor sí los ven para poder limpiarlos.
                if rol == 'rendicion' and n_rec == 0:
                    continue
                items.append({
                    'lote': lote,
                    'n_recetas': n_rec,
                    'n_auditadas': conteos_audit.get(lote.id, 0),
                    'dias_sin_actividad': (now - lote.creado_en).days if lote.creado_en else 0,
                })
        return render_template('rendicion_lotes_list.html',
                               items=items, rol_actual=rol)

    @app.route('/rend-recetas/lotes/crear', methods=['POST'])
    @login_required
    def rendicion_lote_crear():
        """Crea un nuevo lote de rendición. Si el operador es rol=rendicion
        el vendedor se fuerza al suyo."""
        from datetime import datetime
        nro = (request.form.get('nro') or '').strip()
        if not nro:
            flash('Nro de rendición obligatorio.', 'error')
            return redirect(url_for('rendicion_lotes_list'))
        # Normalizar: si es numérico puro, sacar leading zeros para evitar
        # que "35227" y "035227" se traten como lotes distintos.
        if nro.isdigit():
            nro = str(int(nro))
        # Sin período manual: las recetas se traen incrementalmente desde la
        # última carga del vendedor (ver devoluciones_buscar). El período del
        # lote se rellena solo cuando se cargan recetas (min/max fecha_op).
        pd = ph = None
        etiqueta = f'Rendición #{nro}'

        # Resolver vendedor: si rol=rendicion, forzar al matcheado por nombre.
        rol = getattr(current_user, 'rol', None)
        vendedor_id = (request.form.get('vendedor_id') or '').strip() or None
        vendedor_nombre = (request.form.get('vendedor_nombre') or '').strip() or None
        if rol == 'rendicion':
            try:
                vendedores = observer_source.listar_vendedores(solo_habilitados=False)
                vendedor_id, vendedor_nombre = _vendedor_sugerido_actual(vendedores)
            except Exception:
                pass

        with database.get_db() as session:
            existe = (session.query(database.RendicionLote)
                      .filter_by(nro=nro, vendedor_observer_id=vendedor_id).first())
            if existe:
                # Idempotente: si ya existe, la abrimos (en vez de dar error y
                # dejar al operador sin saber qué pasó). "Crear y abrir" =
                # "abrir si existe".
                flash(f'Rendición #{nro} ya existía — la abrimos.', 'info')
                return redirect(url_for('devoluciones_buscar', lote_id=existe.id))
            creador = (getattr(current_user, 'nombre_completo', None)
                       or getattr(current_user, 'username', None))
            lote = database.RendicionLote(
                nro=nro,
                vendedor_observer_id=vendedor_id,
                vendedor_nombre=vendedor_nombre,
                periodo_desde=pd,
                periodo_hasta=ph,
                etiqueta=etiqueta,
                estado='abierta',
                creado_por=str(creador) if creador else None,
            )
            session.add(lote)
            session.commit()
            lote_id = lote.id
            flash(f'Rendición #{nro} creada.', 'success')
        # Llevar al buscar pre-cargado con el lote.
        return redirect(url_for('devoluciones_buscar', lote_id=lote_id))

    @app.route('/rend-recetas/lotes/<int:id>/recibo.pdf')
    @login_required
    def rendicion_lote_recibo_pdf(id):
        """PDF imprimible con todas las recetas del lote + espacio para firma.
        Lo usa el operador para entregar físicamente a quien corresponda
        (vendedor, supervisor, OS) y obtener firma de recepción."""
        from io import BytesIO

        from flask import send_file
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        with database.get_db() as session:
            lote = session.get(database.RendicionLote, id)
            if not lote:
                flash('Rendición no encontrada.', 'error')
                return redirect(url_for('rendicion_lotes_list'))
            devs = (session.query(database.DevolucionReceta)
                    .options(joinedload(database.DevolucionReceta.motivo),
                             joinedload(database.DevolucionReceta.auditor_motivo))
                    .filter_by(rendicion_lote_id=id)
                    .order_by(database.DevolucionReceta.obra_social_nombre,
                              database.DevolucionReceta.fecha_operacion)
                    .all())

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        title_st = ParagraphStyle('t', parent=styles['Heading1'],
                                  fontSize=16, textColor=colors.HexColor('#1F2937'))
        meta_st = ParagraphStyle('m', parent=styles['Normal'],
                                 fontSize=10, textColor=colors.HexColor('#4B5563'))
        elems = [
            Paragraph(f'Recibo de Rendición #{lote.nro}', title_st),
            Spacer(1, 6),
            Paragraph(
                f'<b>Vendedor:</b> {lote.vendedor_nombre or "—"} · '
                f'<b>Período:</b> {lote.periodo_desde.strftime("%d/%m/%Y") if lote.periodo_desde else "—"} → '
                f'{lote.periodo_hasta.strftime("%d/%m/%Y") if lote.periodo_hasta else "—"}',
                meta_st),
            Spacer(1, 4),
            Paragraph(
                f'<b>Etiqueta:</b> {lote.etiqueta or "—"} · '
                f'<b>Estado:</b> {lote.estado.upper()} · '
                f'<b>Total recetas:</b> {len(devs)}',
                meta_st),
            Spacer(1, 12),
        ]
        # Tabla de recetas
        hdr = ['#', 'Fecha', 'OS', 'Op#', 'Motivo', 'Auditor', 'A cargo OS']
        rows_data = [hdr]
        total_os = 0.0
        for i, d in enumerate(devs, 1):
            rows_data.append([
                str(i),
                d.fecha_operacion.strftime('%d/%m/%y') if d.fecha_operacion else '—',
                (d.obra_social_nombre or '—')[:28],
                str(d.id_operacion_observer),
                (d.motivo.nombre if d.motivo else '—')[:22],
                (d.auditor_motivo.nombre if d.auditor_motivo else '—')[:22],
                f'${float(d.importe_a_cargo_os or 0):,.0f}'.replace(',', '.'),
            ])
            total_os += float(d.importe_a_cargo_os or 0)
        rows_data.append(['', '', '', '', '', 'TOTAL', f'${total_os:,.0f}'.replace(',', '.')])
        t = Table(rows_data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.HexColor('#F4C430')),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 8),
            ('ALIGN',      (-1, 1), (-1, -1), 'RIGHT'),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F9FAFB')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
            ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
        ]))
        elems.append(t)
        elems.append(Spacer(1, 30))
        # Espacio firmas
        firma_st = ParagraphStyle('f', parent=styles['Normal'],
                                  fontSize=9, alignment=1)
        tbl_firmas = Table([[
            Paragraph('________________________<br/><br/>Entregó<br/>(firma y aclaración)', firma_st),
            Paragraph('________________________<br/><br/>Recibió<br/>(firma y aclaración)', firma_st),
        ]], colWidths=[8*cm, 8*cm])
        elems.append(tbl_firmas)
        doc.build(elems)
        buf.seek(0)
        fname = f'Recibo_Rendicion_{lote.nro}_{lote.vendedor_nombre or "vendedor"}.pdf'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/pdf')

    @app.route('/rend-recetas/lotes/<int:id>/cerrar', methods=['POST'])
    @login_required
    def rendicion_lote_cerrar(id):
        forzar = request.form.get('forzar') == '1'
        with database.get_db() as session:
            lote = session.get(database.RendicionLote, id)
            if not lote:
                flash('Rendición no encontrada.', 'error')
                return redirect(url_for('rendicion_lotes_list'))
            # Validación pre-cierre: si hay devoluciones pendientes o sin
            # auditar, o si está vacío, avisar (a menos que vengan con ?forzar=1).
            from sqlalchemy import func as _fu
            D = database.DevolucionReceta
            n_total = (session.query(_fu.count(D.id))
                       .filter(D.rendicion_lote_id == id).scalar() or 0)
            n_pend = (session.query(_fu.count(D.id))
                      .filter(D.rendicion_lote_id == id,
                              D.estado == 'pendiente').scalar() or 0)
            n_sin_audit = (session.query(_fu.count(D.id))
                           .filter(D.rendicion_lote_id == id,
                                   D.auditor_motivo_id.is_(None)).scalar() or 0)
            if n_total == 0 and not forzar:
                flash('No se puede cerrar un lote vacío (sin recetas cargadas). '
                      'Cargá al menos una o eliminá el lote.', 'error')
                return redirect(url_for('rendicion_lotes_list'))
            if (n_pend or n_sin_audit) and not forzar:
                msg = []
                if n_pend:
                    msg.append(f'{n_pend} pendiente(s)')
                if n_sin_audit:
                    msg.append(f'{n_sin_audit} sin auditar')
                flash('No se puede cerrar: hay ' + ' y '.join(msg)
                      + '. Confirmá el cierre forzado desde el listado si querés cerrar igual.', 'error')
                return redirect(url_for('rendicion_lotes_list'))
            lote.estado = 'cerrada'
            lote.cerrado_en = database.now_ar()
            lote.cerrado_por = (getattr(current_user, 'nombre_completo', None)
                                or getattr(current_user, 'username', None))
            session.commit()
            flash(f'Rendición #{lote.nro} cerrada.', 'success')
        return redirect(url_for('rendicion_lotes_list'))

    @app.route('/rend-recetas/dedup')
    @login_required
    def devoluciones_dedup_vista():
        """Detecta DevolucionReceta con mismo id_operacion_observer en lotes
        distintos. Aplica reglas para identificar qué se mantiene y qué se
        descarta. NO toca nada — solo muestra."""
        if getattr(current_user, 'rol', None) not in ('auditor', 'admin', 'dev'):
            flash('Sin permiso.', 'error')
            return redirect(url_for('devoluciones_list'))
        grupos = _detectar_duplicados()
        return render_template('devoluciones_dedup.html', grupos=grupos)

    @app.route('/rend-recetas/dedup/aplicar', methods=['POST'])
    @login_required
    def devoluciones_dedup_aplicar():
        """Aplica la deduplicación: descarta los perdedores según las reglas
        (mantiene la fila ganadora, marca el resto como estado=descartada
        con nota_cierre indicando el winner)."""
        if getattr(current_user, 'rol', None) not in ('auditor', 'admin', 'dev'):
            flash('Sin permiso.', 'error')
            return redirect(url_for('devoluciones_list'))
        grupos = _detectar_duplicados()
        n_desc = 0
        n_skip = 0
        usuario = (getattr(current_user, 'nombre_completo', None)
                   or getattr(current_user, 'username', None) or 'AUTO')
        with database.get_db() as session:
            for g in grupos:
                if g.get('skip'):
                    n_skip += len(g['perdedoras'])
                    continue
                winner_id = g['winner']['id']
                winner_lote_nro = g['winner']['lote_nro'] or '?'
                for p in g['perdedoras']:
                    dev = session.get(database.DevolucionReceta, p['id'])
                    if dev:
                        dev.estado = 'descartada'
                        dev.nota_cierre = (f'Duplicado de la receta en lote #{winner_lote_nro} '
                                           f'(devolución #{winner_id}). Auto-dedup {usuario}.')
                        dev.cerrada_en = database.now_ar()
                        dev.cerrada_por = usuario
                        n_desc += 1
            session.commit()
        flash(f'Deduplicación aplicada: {n_desc} descartada(s). {n_skip} ignorada(s) por estar en lotes cerrados.', 'success')
        return redirect(url_for('devoluciones_dedup_vista'))

    @app.route('/rend-recetas/asignar-huerfanas', methods=['POST'])
    @login_required
    def devoluciones_asignar_huerfanas():
        """Asigna en bulk un lote a todas las DevolucionReceta sin
        rendicion_lote_id que matcheen vendedor (y opcionalmente rango fecha).
        Solo auditor/admin/dev."""
        if getattr(current_user, 'rol', None) not in ('auditor', 'admin', 'dev'):
            flash('Sin permiso.', 'error')
            return redirect(url_for('devoluciones_list'))
        lote_id = request.form.get('lote_id', type=int)
        vendedor_nombre = (request.form.get('vendedor_nombre') or '').strip().upper()
        if not lote_id or not vendedor_nombre:
            flash('Faltan lote o vendedor.', 'error')
            return redirect(url_for('devoluciones_list'))
        with database.get_db() as session:
            lote = session.get(database.RendicionLote, lote_id)
            if not lote:
                flash('Lote no encontrado.', 'error')
                return redirect(url_for('devoluciones_list'))
            from sqlalchemy import func as _fu
            n = (session.query(database.DevolucionReceta)
                 .filter(database.DevolucionReceta.rendicion_lote_id.is_(None),
                         _fu.upper(database.DevolucionReceta.vendedor_nombre) == vendedor_nombre)
                 .update({database.DevolucionReceta.rendicion_lote_id: lote.id,
                          database.DevolucionReceta.nro_presentacion: lote.nro},
                         synchronize_session=False))
            session.commit()
            flash(f'{n} devolucion(es) huérfanas asignadas al lote #{lote.nro}.', 'success')
        return redirect(url_for('devoluciones_list'))

    @app.route('/rend-recetas/lotes/<int:id>/entregada', methods=['POST'])
    @login_required
    def rendicion_lote_toggle_entregada(id):
        """Marca/desmarca un lote como físicamente entregado a la OS.
        Solo auditor/admin/dev. Si se marca como entregada y el lote está
        abierto, lo cierra automáticamente (la entrega implica cierre)."""
        if getattr(current_user, 'rol', None) not in ('auditor', 'admin', 'dev'):
            flash('Solo auditor o admin puede marcar la entrega.', 'error')
            return redirect(url_for('rendicion_lotes_list'))
        with database.get_db() as session:
            lote = session.get(database.RendicionLote, id)
            if not lote:
                flash('Rendición no encontrada.', 'error')
                return redirect(url_for('rendicion_lotes_list'))
            nueva = not lote.entregada
            usuario = (getattr(current_user, 'nombre_completo', None)
                       or getattr(current_user, 'username', None))
            lote.entregada = nueva
            if nueva:
                lote.entregada_en = database.now_ar()
                lote.entregada_por = usuario
                # Si estaba abierta, cerrar automáticamente — la entrega
                # implica que el lote ya no recibe más recetas.
                if lote.estado == 'abierta':
                    lote.estado = 'cerrada'
                    lote.cerrado_en = database.now_ar()
                    lote.cerrado_por = f'{usuario} (auto al entregar)'
                    flash(f'Lote #{lote.nro} marcado como ENTREGADA y CERRADO automáticamente.', 'success')
                else:
                    flash(f'Lote #{lote.nro} marcado como entregada.', 'success')
            else:
                lote.entregada_en = None
                lote.entregada_por = None
                flash(f'Lote #{lote.nro} ya no figura como entregada.', 'info')
            session.commit()
        return redirect(url_for('rendicion_lotes_list'))

    @app.route('/rend-recetas/lotes/<int:id>/eliminar', methods=['POST'])
    @login_required
    def rendicion_lote_eliminar(id):
        """Elimina un lote — solo si está vacío (sin devoluciones).
        Para limpiar lotes creados por error. Admin/auditor/dev only."""
        if getattr(current_user, 'rol', None) not in ('auditor', 'admin', 'dev'):
            flash('Sin permiso para eliminar lotes.', 'error')
            return redirect(url_for('rendicion_lotes_list'))
        with database.get_db() as session:
            from sqlalchemy import func as _ff
            n = (session.query(_ff.count(database.DevolucionReceta.id))
                 .filter(database.DevolucionReceta.rendicion_lote_id == id).scalar() or 0)
            if n > 0:
                flash(f'No se puede eliminar: el lote tiene {n} receta(s). '
                      'Cerrá las devoluciones primero.', 'error')
                return redirect(url_for('rendicion_lotes_list'))
            lote = session.get(database.RendicionLote, id)
            if lote:
                nro = lote.nro
                session.delete(lote)
                session.commit()
                flash(f'Lote #{nro} eliminado.', 'success')
        return redirect(url_for('rendicion_lotes_list'))

    @app.route('/rend-recetas/lotes/<int:id>/reabrir', methods=['POST'])
    @login_required
    def rendicion_lote_reabrir(id):
        if getattr(current_user, 'rol', None) not in ('admin', 'dev'):
            flash('Solo admin puede reabrir rendiciones.', 'error')
            return redirect(url_for('rendicion_lotes_list'))
        with database.get_db() as session:
            lote = session.get(database.RendicionLote, id)
            if lote:
                lote.estado = 'abierta'
                lote.cerrado_en = None
                lote.cerrado_por = None
                session.commit()
                flash(f'Rendición #{lote.nro} reabierta.', 'success')
        return redirect(url_for('rendicion_lotes_list'))

    # ──────────────────────────────────────────────────────────────────
    # Listado
    # ──────────────────────────────────────────────────────────────────
    @app.route('/rend-recetas')
    @login_required
    def devoluciones_list():
        q_presentacion = (request.args.get('presentacion') or '').strip()
        q_vendedor = (request.args.get('vendedor') or '').strip()
        q_estado = (request.args.get('estado') or '').strip()
        # Filtro extra: 'pend_auditor' = sin auditor_motivo_id seteado
        q_auditoria = (request.args.get('auditoria') or '').strip()
        try:
            page = max(1, int(request.args.get('page', '1')))
        except ValueError:
            page = 1
        per_page = 50
        offset = (page - 1) * per_page

        rol_actual_lista = getattr(current_user, 'rol', None)

        with database.get_db() as session:
            base = session.query(database.DevolucionReceta).options(
                joinedload(database.DevolucionReceta.motivo),
                joinedload(database.DevolucionReceta.auditor_motivo),
            )
            if q_presentacion:
                base = base.filter(
                    database.DevolucionReceta.nro_presentacion.ilike(f'%{q_presentacion}%')
                )
            if q_vendedor:
                base = base.filter(
                    database.DevolucionReceta.vendedor_nombre.ilike(f'%{q_vendedor}%')
                )
            if q_estado:
                base = base.filter(database.DevolucionReceta.estado == q_estado)
            if q_auditoria == 'pend':
                base = base.filter(database.DevolucionReceta.auditor_motivo_id.is_(None))
            elif q_auditoria == 'ok':
                base = base.filter(database.DevolucionReceta.auditor_motivo_id.isnot(None))

            # Rol rendicion solo ve sus propias devoluciones (matchea por
            # vendedor_nombre con nombre_completo del user). Además, deja de ver
            # las que el auditor ya marcó 'ok' (el ciclo del vendedor terminó).
            if rol_actual_lista == 'rendicion':
                nombre_user = (getattr(current_user, 'nombre_completo', '') or '').strip().upper()
                if nombre_user:
                    from sqlalchemy import func as _fu
                    base = base.filter(
                        _fu.upper(database.DevolucionReceta.vendedor_nombre) == nombre_user
                    )
                base = base.filter(database.DevolucionReceta.estado != 'ok')

            total = base.count()
            devoluciones = (base.order_by(database.DevolucionReceta.creado_en.desc())
                            .offset(offset).limit(per_page).all())
            last_page = max(1, (total + per_page - 1) // per_page)

            # Pre-calcular estado consolidado del LOTE para cada fila visible.
            # Una sola query agregada por los lote_ids únicos de las visibles.
            from sqlalchemy import case as _case
            from sqlalchemy import func as _ff
            lote_ids_v = list({d.rendicion_lote_id for d in devoluciones
                               if d.rendicion_lote_id})
            estado_por_lote = {}
            if lote_ids_v:
                agg = (session.query(
                    database.DevolucionReceta.rendicion_lote_id.label('lid'),
                    _ff.count(database.DevolucionReceta.id).label('nt'),
                    _ff.sum(_case((database.DevolucionReceta.auditor_motivo_id.isnot(None), 1), else_=0)).label('na'),
                    _ff.sum(_case((database.DevolucionReceta.estado == 'pendiente', 1), else_=0)).label('np'),
                ).filter(database.DevolucionReceta.rendicion_lote_id.in_(lote_ids_v))
                 .group_by(database.DevolucionReceta.rendicion_lote_id).all())
                lotes_d = {l.id: l for l in
                           session.query(database.RendicionLote)
                           .filter(database.RendicionLote.id.in_(lote_ids_v)).all()}
                for row in agg:
                    lote = lotes_d.get(row.lid)
                    estado_por_lote[row.lid] = _estado_consolidado(
                        int(row.nt or 0), int(row.na or 0), int(row.np or 0),
                        bool(lote.entregada) if lote else False)

            # Conteos por estado para chips. DEBEN respetar el mismo scope que
            # la tabla (rol rendicion = solo sus propias + filtros de vendedor/
            # nro), sino los chips mostraban totales globales mientras la tabla
            # filtrada quedaba vacía. NO aplicamos q_estado/q_auditoria acá
            # porque los chips SON el selector de estado.
            from sqlalchemy import func as _f
            _scope = session.query(database.DevolucionReceta)
            if q_presentacion:
                _scope = _scope.filter(database.DevolucionReceta.nro_presentacion.ilike(f'%{q_presentacion}%'))
            if q_vendedor:
                _scope = _scope.filter(database.DevolucionReceta.vendedor_nombre.ilike(f'%{q_vendedor}%'))
            if rol_actual_lista == 'rendicion':
                _nu = (getattr(current_user, 'nombre_completo', '') or '').strip().upper()
                if _nu:
                    _scope = _scope.filter(_f.upper(database.DevolucionReceta.vendedor_nombre) == _nu)
                _scope = _scope.filter(database.DevolucionReceta.estado != 'ok')
            _scope_sub = _scope.subquery()
            cuentas_estado = dict(session.query(
                _scope_sub.c.estado, _f.count(_scope_sub.c.id)
            ).group_by(_scope_sub.c.estado).all())

            # Conteo pendientes de auditor (mismo scope).
            pend_auditor = (session.query(_f.count(_scope_sub.c.id))
                            .filter(_scope_sub.c.auditor_motivo_id.is_(None))
                            .scalar() or 0)

            # Motivos disponibles para el auditor (uso_rol = auditor o ambos).
            motivos_auditor = []
            if rol_actual_lista in ('auditor', 'admin', 'dev'):
                motivos_auditor = (session.query(database.MotivoDevolucion)
                                   .filter_by(activo=True)
                                   .filter(database.MotivoDevolucion.uso_rol.in_(['auditor', 'ambos']))
                                   .order_by(database.MotivoDevolucion.nombre).all())
            # Motivos para el vendedor (edición inline en su listado).
            motivos_rendicion = (session.query(database.MotivoDevolucion)
                                 .filter_by(activo=True)
                                 .filter(database.MotivoDevolucion.uso_rol.in_(['rendicion', 'ambos']))
                                 .order_by(database.MotivoDevolucion.nombre).all())

            # Vendedores únicos que tienen devoluciones cargadas — para el
            # filtro dropdown del header.
            vendedores_filtro = []
            if rol_actual_lista in ('auditor', 'admin', 'dev'):
                vendedores_filtro = [v[0] for v in
                    session.query(database.DevolucionReceta.vendedor_nombre)
                    .filter(database.DevolucionReceta.vendedor_nombre.isnot(None))
                    .distinct()
                    .order_by(database.DevolucionReceta.vendedor_nombre)
                    .all() if v[0]]

            # Nros de rendición que tienen devoluciones cargadas — para el
            # filtro dropdown. Más recientes primero (por max creado_en).
            from sqlalchemy import func as _fnr
            nros_q = (session.query(database.DevolucionReceta.nro_presentacion,
                                     _fnr.max(database.DevolucionReceta.creado_en).label('ult'))
                      .filter(database.DevolucionReceta.nro_presentacion.isnot(None))
                      .group_by(database.DevolucionReceta.nro_presentacion)
                      .order_by(_fnr.max(database.DevolucionReceta.creado_en).desc())
                      .all())
            # Para rol rendicion: solo los suyos (ya filtrado por base más arriba
            # no aplica acá porque esta query es independiente — replicamos filtro).
            if rol_actual_lista == 'rendicion':
                nombre_u = (getattr(current_user, 'nombre_completo', '') or '').strip().upper()
                if nombre_u:
                    from sqlalchemy import func as _fnr2
                    nros_q = (session.query(database.DevolucionReceta.nro_presentacion,
                                             _fnr2.max(database.DevolucionReceta.creado_en).label('ult'))
                              .filter(database.DevolucionReceta.nro_presentacion.isnot(None))
                              .filter(_fnr2.upper(database.DevolucionReceta.vendedor_nombre) == nombre_u)
                              .group_by(database.DevolucionReceta.nro_presentacion)
                              .order_by(_fnr2.max(database.DevolucionReceta.creado_en).desc())
                              .all())
            nros_filtro = [n[0] for n in nros_q if n[0]]

            # Último lote abierto del operador → pre-rellenar "+ Nueva búsqueda"
            # con ?lote_id=N para preservar contexto.
            lote_activo_id = None
            if rol_actual_lista == 'rendicion':
                nombre_u_ = (getattr(current_user, 'nombre_completo', '') or '').strip().upper()
                if nombre_u_:
                    from sqlalchemy import func as _fuu
                    _ult = (session.query(database.RendicionLote.id)
                            .filter(database.RendicionLote.estado == 'abierta')
                            .filter(_fuu.upper(database.RendicionLote.vendedor_nombre) == nombre_u_)
                            .order_by(database.RendicionLote.creado_en.desc())
                            .first())
                    if _ult:
                        lote_activo_id = _ult[0]

            # Alertas según rol (banner arriba del listado).
            # - rendicion: AGREGAR DATOS sin respuesta de >=3 días (recetas
            #   cargadas con motivo AGREGAR DATOS, todavía en estado pendiente
            #   y > 3 días desde creado_en).
            # - auditor: cantidad de recetas pendientes de auditar.
            from datetime import datetime, timedelta

            from sqlalchemy import or_ as _or
            alertas = {}
            if rol_actual_lista == 'rendicion':
                hace_3d = datetime.utcnow() - timedelta(days=3)
                hace_7d = datetime.utcnow() - timedelta(days=7)
                # Filtro de vendedor ya aplicado a `base`, hacemos sobre todas
                # las devoluciones del operador.
                from sqlalchemy import func as _fu
                _q_atraso = (session.query(_f.count(database.DevolucionReceta.id))
                             .filter(database.DevolucionReceta.estado == 'pendiente'))
                if (getattr(current_user, 'nombre_completo', '') or '').strip():
                    nu = current_user.nombre_completo.strip().upper()
                    _q_atraso = _q_atraso.filter(
                        _fu.upper(database.DevolucionReceta.vendedor_nombre) == nu
                    )
                alertas['pendientes'] = _q_atraso.scalar() or 0
                alertas['atrasadas_3d'] = (_q_atraso
                    .filter(database.DevolucionReceta.creado_en < hace_3d)
                    .scalar() or 0)
                alertas['atrasadas_7d'] = (_q_atraso
                    .filter(database.DevolucionReceta.creado_en < hace_7d)
                    .scalar() or 0)
            elif rol_actual_lista in ('auditor', 'admin', 'dev'):
                alertas['pend_auditar'] = pend_auditor
                # Por vendedor: top 5 vendedores con más pendientes
                top_q = (session.query(database.DevolucionReceta.vendedor_nombre,
                                       _f.count(database.DevolucionReceta.id).label('n'))
                         .filter(database.DevolucionReceta.auditor_motivo_id.is_(None))
                         .filter(database.DevolucionReceta.vendedor_nombre.isnot(None))
                         .group_by(database.DevolucionReceta.vendedor_nombre)
                         .order_by(_f.count(database.DevolucionReceta.id).desc())
                         .limit(5).all())
                alertas['top_vendedores'] = [(v or '—', int(n)) for v, n in top_q]
                # Detectar duplicados (id_operacion en más de un lote no descartado)
                dup_q = (session.query(_f.count(_f.distinct(database.DevolucionReceta.id_operacion_observer)))
                         .select_from(
                            session.query(database.DevolucionReceta.id_operacion_observer)
                            .filter(database.DevolucionReceta.estado != 'descartada')
                            .group_by(database.DevolucionReceta.id_operacion_observer)
                            .having(_f.count(database.DevolucionReceta.id) > 1)
                            .subquery()
                         ).scalar() or 0)
                alertas['duplicados'] = int(dup_q)

            return render_template('devoluciones_list.html',
                                   devoluciones=devoluciones, total=total,
                                   page=page, last_page=last_page,
                                   q_presentacion=q_presentacion, q_vendedor=q_vendedor,
                                   q_estado=q_estado, q_auditoria=q_auditoria,
                                   cuentas_estado=cuentas_estado,
                                   pend_auditor=pend_auditor,
                                   motivos_auditor=motivos_auditor,
                                   motivos_rendicion=motivos_rendicion,
                                   rol_actual=rol_actual_lista,
                                   alertas=alertas,
                                   lote_activo_id=lote_activo_id,
                                   vendedores_filtro=vendedores_filtro,
                                   nros_filtro=nros_filtro,
                                   estado_por_lote=estado_por_lote)

    # ──────────────────────────────────────────────────────────────────
    # Búsqueda + registro
    # ──────────────────────────────────────────────────────────────────
    @app.route('/rend-recetas/buscar', methods=['GET', 'POST'])
    @login_required
    def devoluciones_buscar():
        # Auditor no carga recetas — solo revisa lo ya cargado. Redirigir
        # a la vista por-vendedor que es su flujo natural.
        if getattr(current_user, 'rol', None) == 'auditor':
            flash('El rol auditor no carga recetas — usá la vista por vendedor para revisar lo cargado.', 'info')
            return redirect(url_for('devoluciones_por_vendedor'))
        # Catálogos: obras sociales desde la DB local sincronizada (rápido, no
        # depende de SQL Server estar online en cada búsqueda).
        # Filtramos las OS ocultas para el rol del operador (lista negra
        # configurable desde /devoluciones/filtros-os).
        rol_param = getattr(current_user, 'rol', None)
        with database.get_db() as session:
            os_ocultas = set()
            if rol_param:
                os_ocultas = {f.obra_social_observer_id for f in
                              session.query(database.RolFiltroObraSocial)
                              .filter_by(rol=rol_param).all()}
            obras_sociales = [{
                'id_obra_social': o.observer_id,
                'nombre': o.descripcion,
            } for o in (session.query(database.ObsObraSocial)
                        .filter(database.ObsObraSocial.fecha_baja.is_(None))
                        .order_by(database.ObsObraSocial.descripcion).all())
                if o.observer_id not in os_ocultas]

        # Vendedores: van live a ObServer (no hay tabla local de OperadoresVenta)
        try:
            vendedores = observer_source.listar_vendedores(solo_habilitados=True)
            observer_ok = True
        except Exception as e:
            vendedores = []
            observer_ok = False
            flash(f'ObServer no responde: {e}', 'warning')

        hoy = date.today()

        # Auto-sugerencia: si el user es rol=rendicion, pre-seleccionar el
        # vendedor de ObServer cuyo nombre matchea con nombre_completo.
        rol_actual_get = getattr(current_user, 'rol', None)
        nombre_user = (getattr(current_user, 'nombre_completo', '') or '').strip().upper()
        vendedor_sugerido = ''
        if rol_actual_get == 'rendicion' and nombre_user:
            for v in vendedores:
                v_nombre = (v.get('nombre') or '').strip().upper()
                if v_nombre == nombre_user or v_nombre.startswith(nombre_user) or nombre_user in v_nombre:
                    vendedor_sugerido = v['id_usuario']
                    break

        # Lotes de rendición abiertos para el dropdown. Para rol=rendicion
        # solo los suyos (filtra por vendedor matcheado).
        from sqlalchemy import func as _f
        with database.get_db() as _s:
            _q = _s.query(database.RendicionLote).filter_by(estado='abierta')
            if rol_actual_get == 'rendicion' and nombre_user:
                _q = _q.filter(_f.upper(database.RendicionLote.vendedor_nombre) == nombre_user)
            lotes_abiertos = [{
                'id': l.id, 'nro': l.nro, 'etiqueta': l.etiqueta or '',
                'periodo_desde': l.periodo_desde.strftime('%d/%m') if l.periodo_desde else '',
                'periodo_hasta': l.periodo_hasta.strftime('%d/%m') if l.periodo_hasta else '',
                # ISO para auto-pre-fill de inputs <input type="date">
                'periodo_desde_iso': l.periodo_desde.isoformat() if l.periodo_desde else '',
                'periodo_hasta_iso': l.periodo_hasta.isoformat() if l.periodo_hasta else '',
            } for l in _q.order_by(database.RendicionLote.creado_en.desc()).all()]

        # Lote pre-seleccionado vía ?lote_id=N (cuando vienen del listado de lotes)
        lote_id_preselect = request.args.get('lote_id', type=int)

        # Bookmark del vendedor: si existe, usamos su ultima_fecha_op como
        # 'desde' default (evita re-procesar recetas ya cargadas). Solo
        # aplica para GET inicial sin filtros explícitos en URL.
        # Rango automático: el operador NO elige fechas. Traemos desde la
        # última carga del vendedor (− margen rezagadas) hasta hoy.
        desde_default = (hoy - timedelta(days=DEFAULT_PRIMERA_VEZ_DIAS)).isoformat()
        if vendedor_sugerido:
            with database.get_db() as _s_bm:
                _bm = (_s_bm.query(database.VendedorBookmark)
                       .filter_by(vendedor_observer_id=vendedor_sugerido).first())
                if _bm and _bm.ultima_fecha_op:
                    # Retrocedemos MARGEN_REZAGADAS_DIAS de la última op para
                    # pescar rezagadas; el dedup marca las ya cargadas.
                    desde_default = (_bm.ultima_fecha_op.date()
                                     - timedelta(days=MARGEN_REZAGADAS_DIAS)).isoformat()

        # Última receta procesada del vendedor (para mostrar en el encabezado
        # de dónde venimos). Es la de mayor creado_en (la última que se cargó).
        ultima_procesada = None
        if vendedor_sugerido:
            with database.get_db() as _s_up:
                _u = (_s_up.query(database.DevolucionReceta)
                      .filter_by(vendedor_observer_id=vendedor_sugerido)
                      .order_by(database.DevolucionReceta.creado_en.desc())
                      .first())
                if _u:
                    ultima_procesada = {
                        'fecha_op': _u.fecha_operacion.strftime('%d/%m/%y %H:%M') if _u.fecha_operacion else '—',
                        'op': _u.id_operacion_observer,
                        'os': _u.obra_social_nombre or '—',
                        'importe': float(_u.importe_total or 0),
                        'nro': _u.nro_presentacion or '',
                        'cargada_en': _u.creado_en.strftime('%d/%m/%y %H:%M') if _u.creado_en else '',
                    }

        if request.method == 'GET':
            return render_template('devoluciones_buscar.html',
                                   obras_sociales=obras_sociales,
                                   vendedores=vendedores,
                                   observer_ok=observer_ok,
                                   desde=desde_default,
                                   hasta=hoy.isoformat(),
                                   nro_presentacion='',
                                   vendedor_id=vendedor_sugerido,
                                   obra_social_id='',
                                   solo_a_cargo_os=False,
                                   resultados=None,
                                   motivos=[], destinos=[],
                                   rol_actual=rol_actual_get,
                                   lotes_abiertos=lotes_abiertos,
                                   ultima_procesada=ultima_procesada,
                                   lote_id=lote_id_preselect or '')

        # POST: buscar
        vendedor_id = (request.form.get('vendedor_id') or '').strip() or None
        obra_social_id = request.form.get('obra_social_id', type=int) or None
        nro_presentacion = (request.form.get('nro_presentacion') or '').strip() or None
        desde_str = (request.form.get('desde') or '').strip()
        hasta_str = (request.form.get('hasta') or '').strip()
        solo_a_cargo_os = request.form.get('solo_a_cargo_os') == '1'

        # SEGURIDAD: si el user es rol=rendicion, forzamos su propio vendedor
        # (ignoramos cualquier vendedor_id manipulado en el form para que un
        # operador no pueda mirar las rendiciones de otro).
        if rol_actual_get == 'rendicion':
            if vendedor_sugerido:
                vendedor_id = vendedor_sugerido
            else:
                flash('Tu usuario no está mapeado a un vendedor de ObServer. '
                      'Pedile al admin que ajuste tu nombre completo para que '
                      'matchee con un OperadorVenta.', 'error')
                return redirect(url_for('devoluciones_buscar'))

        if not vendedor_id and not obra_social_id:
            flash('Seleccioná al menos un vendedor o una obra social.', 'error')
            return redirect(url_for('devoluciones_buscar'))
        try:
            desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
            hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Fechas inválidas.', 'error')
            return redirect(url_for('devoluciones_buscar'))
        if desde > hasta:
            flash('"Desde" no puede ser posterior a "hasta".', 'error')
            return redirect(url_for('devoluciones_buscar'))

        # Si el operador pidió una OS específicamente bloqueada, abortar.
        if obra_social_id and obra_social_id in os_ocultas:
            flash('No tenés permiso para ver esa obra social.', 'error')
            return redirect(url_for('devoluciones_buscar'))

        try:
            resultados = observer_source.buscar_recetas(
                vendedor_uuid=vendedor_id,
                obra_social_id=obra_social_id,
                desde=desde, hasta=hasta,
                solo_a_cargo_os=solo_a_cargo_os,
            )
        except Exception as e:
            flash(f'Error consultando ObServer: {e}', 'error')
            return redirect(url_for('devoluciones_buscar'))

        # Filtrar OS ocultas de los resultados (defensa en profundidad: el
        # dropdown ya las esconde, pero si vienen sin filtro de OS en la
        # query, podrían aparecer en el resultado).
        if os_ocultas:
            # buscar_recetas no devuelve observer_id de OS — solo nombre.
            # Construimos un set de nombres a esconder para matchear.
            nombres_ocultas = set()
            for f in (session.query(database.RolFiltroObraSocial)
                      .filter_by(rol=rol_param).all()):
                nombres_ocultas.add(f.nombre_cached.upper())
            if nombres_ocultas:
                resultados = [r for r in resultados
                              if (r.get('obra_social') or '').upper() not in nombres_ocultas]

        # Orden: por fecha+hora de operación ascendente (cronológico). Al traer
        # recetas nuevas conviene verlas en el orden en que se vendieron.
        resultados.sort(key=lambda r: r.get('fecha_operacion') or 0)

        # Snapshots de labels
        vendedor_nombre = (next((v['nombre'] for v in vendedores
                                 if v['id_usuario'] == vendedor_id), None)
                           if vendedor_id else None)
        os_nombre = (next((o['nombre'] for o in obras_sociales
                          if o['id_obra_social'] == obra_social_id), None)
                     if obra_social_id else None)

        # Cargar catálogos para el form inline.
        # Filtramos motivos según el rol del operador: rendicion ve solo los
        # suyos + 'ambos', auditor ve solo los suyos + 'ambos'. Cualquier
        # otro rol (admin, dev) ve todos.
        rol_actual = getattr(current_user, 'rol', None)
        with database.get_db() as session:
            q_motivos = (session.query(database.MotivoDevolucion)
                         .filter_by(activo=True))
            if rol_actual == 'rendicion':
                q_motivos = q_motivos.filter(
                    database.MotivoDevolucion.uso_rol.in_(['rendicion', 'ambos']))
            elif rol_actual == 'auditor':
                q_motivos = q_motivos.filter(
                    database.MotivoDevolucion.uso_rol.in_(['auditor', 'ambos']))
            motivos = q_motivos.order_by(database.MotivoDevolucion.nombre).all()
            destinos = []  # deprecado 2026-05-18, queda para no romper template
            # IDs de operaciones ya devueltas (para mostrar badge).
            # Mapeamos id_operacion → nro_rendicion para indicar EN QUÉ lote
            # ya está cargada (ayuda al operador a no re-cargarla).
            ya_devueltas = {}
            ids = [r['id_operacion'] for r in resultados]
            if ids:
                qd = (session.query(database.DevolucionReceta.id_operacion_observer,
                                     database.DevolucionReceta.nro_presentacion)
                      .filter(database.DevolucionReceta.id_operacion_observer.in_(ids))
                      .filter(database.DevolucionReceta.estado != 'descartada')
                      .all())
                for op_id, nro in qd:
                    ya_devueltas[op_id] = nro or ''

        return render_template('devoluciones_buscar.html',
                               obras_sociales=obras_sociales,
                               vendedores=vendedores,
                               observer_ok=True,
                               desde=desde.isoformat(), hasta=hasta.isoformat(),
                               nro_presentacion=nro_presentacion or '',
                               vendedor_id=vendedor_id or '',
                               vendedor_nombre=vendedor_nombre,
                               obra_social_id=obra_social_id or '',
                               os_nombre=os_nombre,
                               solo_a_cargo_os=solo_a_cargo_os,
                               resultados=resultados,
                               motivos=motivos, destinos=destinos,
                               ya_devueltas=ya_devueltas,
                               rol_actual=rol_actual,
                               lotes_abiertos=lotes_abiertos,
                               lote_id=request.form.get('lote_id', type=int) or '')

    @app.route('/rend-recetas/guardar', methods=['POST'])
    @login_required
    def devoluciones_guardar():
        vendedor_id = (request.form.get('vendedor_id') or '').strip() or None
        vendedor_nombre = (request.form.get('vendedor_nombre') or '').strip() or None
        # Lote de rendición (nuevo modelo). El nro_presentacion se desnormaliza
        # desde el lote para compat con queries viejas.
        lote_id = request.form.get('lote_id', type=int)
        rol_g = getattr(current_user, 'rol', None)

        # Fallback defensivo: si el form NO mandó lote_id (bug viejo del
        # template), intentamos resolverlo. 1ro por nro_presentacion del
        # hidden, 2do por el último lote abierto del vendedor.
        if not lote_id and rol_g == 'rendicion':
            nro_hint = (request.form.get('nro_presentacion') or '').strip()
            with database.get_db() as _s_fb:
                from sqlalchemy import func as _fu_fb
                q_lote = _s_fb.query(database.RendicionLote).filter_by(estado='abierta')
                if vendedor_id:
                    q_lote = q_lote.filter_by(vendedor_observer_id=vendedor_id)
                elif vendedor_nombre:
                    q_lote = q_lote.filter(_fu_fb.upper(database.RendicionLote.vendedor_nombre)
                                            == vendedor_nombre.upper())
                if nro_hint:
                    lote_fb = q_lote.filter_by(nro=nro_hint).first()
                else:
                    lote_fb = q_lote.order_by(database.RendicionLote.creado_en.desc()).first()
                if lote_fb:
                    lote_id = lote_fb.id

        # Si después del fallback sigue sin lote_id y es rendicion, abortar.
        if rol_g == 'rendicion' and not lote_id:
            flash('Tenés que seleccionar una rendición antes de guardar. '
                  'Si no hay ninguna abierta, creala desde 📋 Rendiciones.', 'error')
            return redirect(url_for('devoluciones_buscar'))

        nro_presentacion = None
        if lote_id:
            with database.get_db() as _s:
                _lote = _s.get(database.RendicionLote, lote_id)
                if _lote:
                    nro_presentacion = _lote.nro
        # Modelo nuevo: se guarda TODO el lote (todas las recetas traídas), no
        # solo las marcadas. `op_all` = todas las op de la pantalla; `marcar` =
        # las que el vendedor marcó "Rendida" (en_auditoria=True → pasan al
        # auditor). Las no marcadas quedan del lado del vendedor (en_auditoria=
        # False). Las ya cargadas (en cualquier lote, no descartadas) se saltean.
        todas_op = request.form.getlist('op_all')
        marcados = set(request.form.getlist('marcar'))
        if not todas_op:
            # Fallback compat: si el template viejo no mandó op_all, usar marcar.
            todas_op = request.form.getlist('marcar')
            marcados = set(todas_op)
        if not todas_op:
            flash('No hay recetas para guardar.', 'error')
            return redirect(url_for('devoluciones_buscar'))

        # Validación del primer control: cada receta del lote debe estar
        # "revisada" → marcada Rendida, o con motivo, o con observación. Si
        # alguna queda sin tocar, no se registra (obliga a revisar todas).
        sin_revisar = 0
        for op_id_str in todas_op:
            rendida = op_id_str in marcados
            tiene_obs = bool((request.form.get(f'obs_{op_id_str}') or '').strip())
            tiene_motivo = bool(request.form.get(f'motivo_{op_id_str}'))
            if not (rendida or tiene_obs or tiene_motivo):
                sin_revisar += 1
        if sin_revisar:
            flash(f'{sin_revisar} receta(s) sin revisar. Marcá "Rendida" o '
                  'agregá motivo/observación a TODAS antes de registrar.', 'error')
            _lote = request.form.get('lote_id', type=int)
            return redirect(url_for('devoluciones_buscar', lote_id=_lote) if _lote
                            else url_for('devoluciones_buscar'))

        # Preferimos el nombre completo (display name del operador en pantalla),
        # luego username, después email. NO caemos en id numérico — preferimos
        # '?' a un número sin contexto (data inutilizable en reportes).
        creador = (getattr(current_user, 'nombre_completo', None)
                   or getattr(current_user, 'username', None)
                   or getattr(current_user, 'email', None)
                   or '?')

        # Cache de vendedores para resolver nombre por UUID (1 sola query a ObServer)
        vendedor_name_by_id = {}
        try:
            for v in observer_source.listar_vendedores(solo_habilitados=False):
                vendedor_name_by_id[v['id_usuario']] = v['nombre']
        except Exception:
            pass  # si ObServer no responde, guardamos solo UUID sin nombre

        n_creadas = 0
        n_rendidas = 0
        errores = []
        with database.get_db() as session:
            # Op ya cargadas (en cualquier lote, no descartadas) → no duplicar.
            ya_cargadas = set()
            _ids_int = [int(x) for x in todas_op if x.isdigit()]
            if _ids_int:
                for (oid,) in (session.query(database.DevolucionReceta.id_operacion_observer)
                               .filter(database.DevolucionReceta.id_operacion_observer.in_(_ids_int))
                               .filter(database.DevolucionReceta.estado != 'descartada').all()):
                    ya_cargadas.add(oid)

            for op_id_str in todas_op:
                try:
                    op_id = int(op_id_str)
                except ValueError:
                    continue
                if op_id in ya_cargadas:
                    continue  # ya está en un lote — no re-crear
                es_rendida = op_id_str in marcados
                motivo_id = request.form.get(f'motivo_{op_id}', type=int)
                # destino_vendedor_* quedó deprecado (2026-05-18) — ya no se
                # captura en el form. Lo dejamos como None para mantener
                # compat con la columna en DB.
                destino_vendedor_id = None
                destino_vendedor_nombre = None
                obs = (request.form.get(f'obs_{op_id}') or '').strip() or None
                # Campos nuevos v2: observación exclusiva del rol rendicion +
                # multi-checkbox de "AGREGAR DATOS" (afiliado, fecha, etc.).
                obs_rend = (request.form.get(f'obs_rend_{op_id}') or '').strip() or None
                ad_list = request.form.getlist(f'ad_{op_id}')  # múltiples checkboxes
                import json as _json
                ad_json = _json.dumps(ad_list) if ad_list else None
                # Motivo ya NO es obligatorio: una receta OK se guarda sin motivo.
                # Solo si la marcan "Rendida" con observación conviene motivo,
                # pero no lo forzamos (el auditor puede completar).
                # Snapshot de datos de la receta
                fop = request.form.get(f'fop_{op_id}')
                fop_dt = None
                if fop:
                    try:
                        fop_dt = datetime.fromisoformat(fop)
                    except ValueError:
                        pass
                os_nombre = (request.form.get(f'os_{op_id}') or '').strip() or None
                imp_total = request.form.get(f'imp_{op_id}')
                imp_os = request.form.get(f'imp_os_{op_id}')
                try:
                    imp_total = Decimal(imp_total) if imp_total else None
                except Exception:
                    imp_total = None
                try:
                    imp_os = Decimal(imp_os) if imp_os else None
                except Exception:
                    imp_os = None

                dev = database.DevolucionReceta(
                    nro_presentacion=nro_presentacion,
                    rendicion_lote_id=lote_id,
                    vendedor_observer_id=vendedor_id,
                    vendedor_nombre=vendedor_nombre,
                    id_operacion_observer=op_id,
                    fecha_operacion=fop_dt,
                    obra_social_nombre=os_nombre,
                    importe_total=imp_total,
                    importe_a_cargo_os=imp_os,
                    motivo_id=motivo_id,
                    destino_vendedor_observer_id=destino_vendedor_id,
                    destino_vendedor_nombre=destino_vendedor_nombre,
                    observaciones=obs,
                    observaciones_rendicion=obs_rend,
                    agregar_datos_json=ad_json,
                    en_auditoria=es_rendida,  # marcada "Rendida" → pasa al auditor
                    creado_por=str(creador) if creador else None,
                )
                session.add(dev)
                n_creadas += 1
                if es_rendida:
                    n_rendidas += 1
            if n_creadas:
                session.commit()

            # Actualizar bookmark del vendedor con la op más nueva procesada
            # en este batch. Sirve para auto-rellenar el filtro 'desde' la
            # próxima vez que el operador entre al form (evita re-procesar
            # recetas que ya pasaron por un lote).
            if vendedor_id and n_creadas:
                from sqlalchemy import func as _fbk
                ult_op = (session.query(_fbk.max(database.DevolucionReceta.id_operacion_observer),
                                         _fbk.max(database.DevolucionReceta.fecha_operacion))
                          .filter(database.DevolucionReceta.vendedor_observer_id == vendedor_id)
                          .first())
                if ult_op:
                    bm = (session.query(database.VendedorBookmark)
                          .filter_by(vendedor_observer_id=vendedor_id).first())
                    if not bm:
                        bm = database.VendedorBookmark(
                            vendedor_observer_id=vendedor_id,
                            vendedor_nombre=vendedor_nombre,
                        )
                        session.add(bm)
                    bm.ultima_op_id = ult_op[0]
                    bm.ultima_fecha_op = ult_op[1]
                    bm.ultimo_lote_id = lote_id
                    bm.actualizado_en = database.now_ar()
                    session.commit()

        if errores:
            for e in errores:
                flash(e, 'error')
        if n_creadas:
            msg = f'{n_creadas} receta(s) registrada(s) en el lote'
            if n_rendidas:
                msg += f' · {n_rendidas} marcada(s) "Rendida" (al auditor)'
            flash(msg + '.', 'success')
        else:
            flash('No se registraron recetas nuevas (ya estaban cargadas).', 'info')
        return redirect(url_for('devoluciones_list'))

    @app.route('/rend-recetas/<int:id>/estado', methods=['POST'])
    @login_required
    def devolucion_cambiar_estado(id):
        nuevo = (request.form.get('estado') or '').strip()
        # 'ok' = receta sin observaciones, todo bien (el auditor solo confirma).
        # 'resuelta' = tuvo observación (emisor/auditor) y se corrigió.
        if nuevo not in ('pendiente', 'ok', 'resuelta', 'descartada', 'devuelta'):
            flash('Estado inválido.', 'error')
            return redirect(url_for('devoluciones_list'))
        nota = (request.form.get('nota_cierre') or '').strip() or None
        with database.get_db() as session:
            dev = session.get(database.DevolucionReceta, id)
            if not dev:
                flash('Devolución no encontrada.', 'error')
                return redirect(url_for('devoluciones_list'))
            dev.estado = nuevo
            dev.nota_cierre = nota
            # Posesión según la decisión:
            #  - devuelta → vuelve al vendedor (en_auditoria=False) para recorregir
            #  - ok/resuelta/descartada → queda del lado del auditor (Rendida X)
            #  - pendiente (reabrir) → vuelve al vendedor
            if nuevo == 'devuelta':
                dev.en_auditoria = False
            elif nuevo in ('ok', 'resuelta', 'descartada'):
                dev.en_auditoria = True
            elif nuevo == 'pendiente':
                dev.en_auditoria = False
            if nuevo == 'pendiente':
                dev.cerrada_en = None
                dev.cerrada_por = None
            else:
                dev.cerrada_en = database.now_ar()
                dev.cerrada_por = (getattr(current_user, 'nombre_completo', None)
                                   or getattr(current_user, 'username', None)
                                   or getattr(current_user, 'email', None)
                                   or '?')
            session.commit()
            flash(f'Devolución #{id} → {nuevo}.', 'success')
        return redirect(url_for('devoluciones_list'))


    @app.route('/rend-recetas/<int:id>/set-motivo-vendedor', methods=['POST'])
    @login_required
    def set_motivo_vendedor(id):
        """Autosave del motivo + obs por parte del vendedor, sobre recetas que
        todavía tiene (en_auditoria=False). Si el motivo nuevo bloquea Rendida,
        además des-marca en_auditoria por las dudas."""
        with database.get_db() as session:
            d = session.get(database.DevolucionReceta, id)
            if not d:
                return jsonify({'ok': False, 'error': 'no encontrada'}), 404
            if d.en_auditoria:
                return jsonify({'ok': False, 'error': 'ya está en auditoría'}), 400
            # Solo tocar los campos que vienen en el form (motivo y obs se
            # guardan por separado; no pisar uno al guardar el otro).
            if 'motivo_id' in request.form:
                d.motivo_id = request.form.get('motivo_id', type=int) or None
            if 'obs' in request.form:
                d.observaciones = (request.form.get('obs') or '').strip() or None
            session.commit()
        return jsonify({'ok': True})

    @app.route('/rend-recetas/<int:id>/marcar-rendida', methods=['POST'])
    @login_required
    def marcar_rendida_vendedor(id):
        """El vendedor marca una receta como Rendida → pasa al auditor
        (en_auditoria=True). Bloqueado si el motivo no lo permite."""
        with database.get_db() as session:
            d = (session.query(database.DevolucionReceta)
                 .options(joinedload(database.DevolucionReceta.motivo))
                 .filter_by(id=id).first())
            if not d:
                flash('Receta no encontrada.', 'error')
                return redirect(url_for('devoluciones_list'))
            if d.motivo and d.motivo.bloquea_rendida:
                flash(f'El motivo "{d.motivo.nombre}" no permite rendir '
                      '(receta no disponible).', 'error')
                return redirect(url_for('devoluciones_list'))
            d.en_auditoria = True
            session.commit()
            flash(f'Receta #{d.id_operacion_observer} marcada Rendida → al auditor.', 'success')
        return redirect(url_for('devoluciones_list'))

    @app.route('/rend-recetas/rendir-os')
    @login_required
    def rendir_os():
        """Pantalla "Rendición a Obras Sociales": lista las recetas en estado
        'ok' que todavía no se rindieron a la OS (rendida_os=False), ordenadas
        por fecha+hora. Filtro multitoken por obra social. Al marcarlas como
        rendidas pasan a histórico (salen de esta vista)."""
        q_os = (request.args.get('os') or '').strip()
        # Filtro de estado de rendición a OS: 'sin' (default) | 'rendidas' | 'todas'.
        ver = (request.args.get('ver') or 'sin').strip()
        with database.get_db() as session:
            query = (session.query(database.DevolucionReceta)
                     .filter(database.DevolucionReceta.estado == 'ok'))
            if ver == 'sin':
                query = query.filter(database.DevolucionReceta.rendida_os.is_(False))
            elif ver == 'rendidas':
                query = query.filter(database.DevolucionReceta.rendida_os.is_(True))
            # 'todas' → sin filtro de rendida_os
            # Filtro multitoken AND sobre el nombre de la OS.
            for tok in q_os.split():
                query = query.filter(
                    database.DevolucionReceta.obra_social_nombre.ilike(f'%{tok}%'))
            recetas = (query.order_by(database.DevolucionReceta.fecha_operacion.asc()).all())
            data = [{
                'id': r.id,
                'fecha_op': r.fecha_operacion,
                'op': r.id_operacion_observer,
                'os': r.obra_social_nombre or '—',
                'vendedor': r.vendedor_nombre or '—',
                'nro': r.nro_presentacion or '',
                'importe_os': float(r.importe_a_cargo_os or 0),
                'importe_total': float(r.importe_total or 0),
                'rendida_os': r.rendida_os,
                'rendida_os_en': r.rendida_os_en,
            } for r in recetas]
            total_os = sum(d['importe_os'] for d in data)
            total_100 = sum(d['importe_total'] for d in data)
        return render_template('devoluciones_rendir_os.html',
                               recetas=data, q_os=q_os, ver=ver,
                               total_os=total_os, total_100=total_100)

    @app.route('/rend-recetas/rendir-os/marcar', methods=['POST'])
    @login_required
    def rendir_os_marcar():
        """Marca como rendidas a la OS las recetas seleccionadas → pasan a
        histórico. Solo aplica a recetas en estado 'ok'."""
        ids = request.form.getlist('marcar')
        ids_int = [int(x) for x in ids if x.isdigit()]
        if not ids_int:
            flash('No seleccionaste ninguna receta.', 'error')
            return redirect(url_for('rendir_os'))
        quien = (getattr(current_user, 'nombre_completo', None)
                 or getattr(current_user, 'username', None) or '?')
        n = 0
        with database.get_db() as session:
            for rid in ids_int:
                r = session.get(database.DevolucionReceta, rid)
                if r and r.estado == 'ok' and not r.rendida_os:
                    r.rendida_os = True
                    r.rendida_os_en = database.now_ar()
                    r.rendida_os_por = str(quien)
                    n += 1
            session.commit()
        flash(f'{n} receta(s) rendida(s) a la OS — pasaron a histórico.', 'success')
        return redirect(url_for('rendir_os', os=request.form.get('os', '')))

    def _rendir_os_data(session, q_os):
        """Helper: recetas OK no rendidas a OS, filtro multitoken, orden fecha."""
        query = (session.query(database.DevolucionReceta)
                 .filter(database.DevolucionReceta.estado == 'ok')
                 .filter(database.DevolucionReceta.rendida_os.is_(False)))
        for tok in (q_os or '').split():
            query = query.filter(
                database.DevolucionReceta.obra_social_nombre.ilike(f'%{tok}%'))
        return query.order_by(database.DevolucionReceta.fecha_operacion.asc()).all()

    @app.route('/rend-recetas/rendir-os/export.xlsx')
    @login_required
    def rendir_os_export_xlsx():
        import io as _io
        import openpyxl
        from flask import send_file
        from openpyxl.styles import Font, PatternFill
        q_os = (request.args.get('os') or '').strip()
        with database.get_db() as session:
            recetas = _rendir_os_data(session, q_os)
            rows = [(r.fecha_operacion, r.vendedor_nombre or '', r.nro_presentacion or '',
                     r.obra_social_nombre or '', float(r.importe_a_cargo_os or 0),
                     float(r.importe_total or 0)) for r in recetas]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Rendición a OS'
        hdr = ['Fecha', 'Vendedor', 'Rendición', 'Obra Social', 'Total a cargo O.Social', 'Total al 100']
        ws.append(hdr)
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='1E3A5F')
        for f, vend, nro, os_n, imp_os, imp_t in rows:
            ws.append([f.strftime('%d/%m/%y %H:%M') if f else '', vend, nro, os_n, imp_os, imp_t])
        ws.append(['', '', '', 'TOTALES', sum(r[4] for r in rows), sum(r[5] for r in rows)])
        for col, w in zip('ABCDEF', (16, 20, 12, 32, 14, 14)):
            ws.column_dimensions[col].width = w
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name='rendicion_obras_sociales.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/rend-recetas/rendir-os/export.pdf')
    @login_required
    def rendir_os_export_pdf():
        import io as _io
        from flask import send_file
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                        Table, TableStyle)
        from reportlab.lib.styles import getSampleStyleSheet
        q_os = (request.args.get('os') or '').strip()
        with database.get_db() as session:
            recetas = _rendir_os_data(session, q_os)
            rows = [(r.fecha_operacion, r.vendedor_nombre or '', r.nro_presentacion or '',
                     r.obra_social_nombre or '', float(r.importe_a_cargo_os or 0),
                     float(r.importe_total or 0)) for r in recetas]
        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                topMargin=1*cm, bottomMargin=1*cm,
                                leftMargin=1*cm, rightMargin=1*cm)
        styles = getSampleStyleSheet()
        story = [Paragraph('Rendición a Obras Sociales', styles['Title'])]
        if q_os:
            story.append(Paragraph(f'Filtro OS: {q_os}', styles['Normal']))
        story.append(Spacer(1, 8))
        data = [['Fecha', 'Vendedor', 'Rendición', 'Obra Social', 'Total a cargo O.Social', 'Total al 100']]
        def _m(v):
            return '$' + f'{int(round(v)):,}'.replace(',', '.')
        for f, vend, nro, os_n, imp_os, imp_t in rows:
            data.append([f.strftime('%d/%m/%y %H:%M') if f else '', vend,
                         f'#{nro}' if nro else '', os_n, _m(imp_os), _m(imp_t)])
        data.append(['', '', '', 'TOTALES', _m(sum(r[4] for r in rows)), _m(sum(r[5] for r in rows))])
        t = Table(data, colWidths=[3*cm, 4*cm, 2.5*cm, 9*cm, 3.5*cm, 3.5*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f3f4f6')]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d1d5db')),
        ]))
        story.append(t)
        doc.build(story)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name='rendicion_obras_sociales.pdf',
                         mimetype='application/pdf')

    @app.route('/rend-recetas/<int:id>/timeline')
    @login_required
    def devolucion_timeline(id):
        """Devuelve JSON con los eventos cronológicos de una devolución:
        creación, auditoría, cierre. Para el modal de historial."""
        with database.get_db() as session:
            d = (session.query(database.DevolucionReceta)
                 .options(joinedload(database.DevolucionReceta.motivo),
                          joinedload(database.DevolucionReceta.auditor_motivo))
                 .filter_by(id=id).first())
            if not d:
                return jsonify({'error': 'No encontrada'}), 404
            eventos = []
            if d.creado_en:
                ad_list = []
                if d.agregar_datos_json:
                    import json as _j
                    try:
                        ad_list = _j.loads(d.agregar_datos_json) or []
                    except (ValueError, TypeError):
                        pass
                eventos.append({
                    'tipo': 'creada',
                    'cuando': d.creado_en.strftime('%d/%m/%Y %H:%M'),
                    'quien': d.creado_por or '—',
                    'detalle': {
                        'motivo': d.motivo.nombre if d.motivo else None,
                        'obs': d.observaciones,
                        'obs_rendicion': d.observaciones_rendicion,
                        'agregar_datos': ad_list,
                    },
                })
            if d.auditor_fecha:
                eventos.append({
                    'tipo': 'auditada',
                    'cuando': d.auditor_fecha.strftime('%d/%m/%Y %H:%M'),
                    'quien': d.auditor_user or '—',
                    'detalle': {
                        'motivo': d.auditor_motivo.nombre if d.auditor_motivo else None,
                        'obs': d.auditor_observaciones,
                    },
                })
            if d.cerrada_en:
                eventos.append({
                    'tipo': 'cerrada',
                    'estado': d.estado,
                    'cuando': d.cerrada_en.strftime('%d/%m/%Y %H:%M'),
                    'quien': d.cerrada_por or '—',
                    'detalle': {'nota': d.nota_cierre},
                })
            return jsonify({
                'id': d.id,
                'receta_op': d.id_operacion_observer,
                'os': d.obra_social_nombre,
                'vendedor': d.vendedor_nombre,
                'rendicion': d.nro_presentacion,
                'eventos': eventos,
            })

    @app.route('/rend-recetas/export.xlsx')
    @login_required
    def devoluciones_export_xlsx():
        """Export del listado de devoluciones con los mismos filtros que /rend-recetas.
        Para auditor/admin: todo lo visible. Para rendicion: solo lo del operador."""
        from io import BytesIO

        import openpyxl
        from flask import send_file
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        q_presentacion = (request.args.get('presentacion') or '').strip()
        q_vendedor = (request.args.get('vendedor') or '').strip()
        q_estado = (request.args.get('estado') or '').strip()
        q_auditoria = (request.args.get('auditoria') or '').strip()
        rol = getattr(current_user, 'rol', None)

        with database.get_db() as session:
            D = database.DevolucionReceta
            from sqlalchemy import func as _f2
            base = session.query(D).options(
                joinedload(D.motivo),
                joinedload(D.auditor_motivo),
            )
            if q_presentacion:
                base = base.filter(D.nro_presentacion.ilike(f'%{q_presentacion}%'))
            if q_vendedor:
                base = base.filter(D.vendedor_nombre.ilike(f'%{q_vendedor}%'))
            if q_estado:
                base = base.filter(D.estado == q_estado)
            if q_auditoria == 'pend':
                base = base.filter(D.auditor_motivo_id.is_(None))
            elif q_auditoria == 'ok':
                base = base.filter(D.auditor_motivo_id.isnot(None))
            if rol == 'rendicion':
                nu = (getattr(current_user, 'nombre_completo', '') or '').strip().upper()
                if nu:
                    base = base.filter(_f2.upper(D.vendedor_nombre) == nu)
            devs = base.order_by(D.creado_en.desc()).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Devoluciones'
        headers = ['Estado', 'Fecha registro', 'Cargó', 'Rendición', 'Vendedor',
                   'Op#', 'Fecha receta', 'Obra Social', 'A cargo OS', 'Total',
                   'Motivo', 'Obs.', 'Obs. rendición', 'Agregar datos',
                   'Auditor motivo', 'Auditor obs.', 'Auditor user', 'Auditor fecha',
                   'Cerrada en', 'Cerrada por', 'Nota cierre']
        ws.append(headers)
        hdr_fill = PatternFill('solid', fgColor='1F2937')
        hdr_font = Font(bold=True, color='F4C430')
        thin = Side(style='thin', color='374151')
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=thin)
        for d in devs:
            import json as _json
            ad = []
            if d.agregar_datos_json:
                try:
                    ad = _json.loads(d.agregar_datos_json) or []
                except (ValueError, TypeError):
                    ad = []
            ws.append([
                d.estado,
                d.creado_en.strftime('%d/%m/%Y %H:%M') if d.creado_en else '',
                d.creado_por or '',
                d.nro_presentacion or '',
                d.vendedor_nombre or '',
                d.id_operacion_observer,
                d.fecha_operacion.strftime('%d/%m/%Y %H:%M') if d.fecha_operacion else '',
                d.obra_social_nombre or '',
                float(d.importe_a_cargo_os) if d.importe_a_cargo_os else 0,
                float(d.importe_total) if d.importe_total else 0,
                d.motivo.nombre if d.motivo else '',
                d.observaciones or '',
                d.observaciones_rendicion or '',
                ', '.join(ad),
                d.auditor_motivo.nombre if d.auditor_motivo else '',
                d.auditor_observaciones or '',
                d.auditor_user or '',
                d.auditor_fecha.strftime('%d/%m/%Y %H:%M') if d.auditor_fecha else '',
                d.cerrada_en.strftime('%d/%m/%Y %H:%M') if d.cerrada_en else '',
                d.cerrada_por or '',
                d.nota_cierre or '',
            ])
        widths = [12, 16, 14, 12, 16, 10, 16, 24, 12, 12, 22, 30, 30, 24, 22, 30, 14, 16, 16, 14, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        from datetime import datetime as _dt
        fname = f'Devoluciones_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/rend-recetas/<int:id>/auditar', methods=['POST'])
    @login_required
    def devolucion_auditar(id):
        """Etapa 2: el auditor revisa lo cargado por rendicion y agrega
        motivo + observaciones de auditoría. El motivo debe ser uno con
        uso_rol IN ('auditor', 'ambos')."""
        if getattr(current_user, 'rol', None) not in ('auditor', 'admin', 'dev'):
            flash('No tenés permiso para auditar.', 'error')
            return redirect(url_for('devoluciones_list'))

        motivo_id = request.form.get('auditor_motivo_id', type=int)
        obs = (request.form.get('auditor_observaciones') or '').strip() or None

        with database.get_db() as session:
            dev = session.get(database.DevolucionReceta, id)
            if not dev:
                flash('Devolución no encontrada.', 'error')
                return redirect(url_for('devoluciones_list'))

            # Validar que el motivo elegido sea de uso auditor.
            if motivo_id:
                m = session.get(database.MotivoDevolucion, motivo_id)
                if not m:
                    flash('Motivo inválido.', 'error')
                    return redirect(url_for('devoluciones_list'))
                if m.uso_rol not in ('auditor', 'ambos'):
                    flash('Ese motivo no es de uso auditor.', 'error')
                    return redirect(url_for('devoluciones_list'))

            dev.auditor_motivo_id = motivo_id
            dev.auditor_observaciones = obs
            dev.auditor_user = (getattr(current_user, 'nombre_completo', None)
                                or getattr(current_user, 'username', None)
                                or str(getattr(current_user, 'id', '') or ''))
            dev.auditor_fecha = database.now_ar()
            session.commit()
            flash(f'Devolución #{id} auditada.', 'success')
        return redirect(request.referrer or url_for('devoluciones_list'))

    @app.route('/rend-recetas/<int:id>/eliminar', methods=['POST'])
    @login_required
    def devolucion_eliminar(id):
        with database.get_db() as session:
            dev = session.get(database.DevolucionReceta, id)
            if dev:
                session.delete(dev)
                session.commit()
                flash('Devolución eliminada.', 'success')
        return redirect(url_for('devoluciones_list'))

    # ──────────────────────────────────────────────────────────────────
    # ABM Filtros de Obra Social por rol (bloquear OS para cierto rol)
    # ──────────────────────────────────────────────────────────────────
    @app.route('/rend-recetas/filtros-os', methods=['GET', 'POST'])
    @login_required
    def devoluciones_filtros_os():
        from database import ObsObraSocial, RolFiltroObraSocial
        with database.get_db() as session:
            if request.method == 'POST':
                rol = (request.form.get('rol') or '').strip()
                os_observer_id = request.form.get('obra_social_observer_id', type=int)
                if not rol or not os_observer_id:
                    flash('Rol y obra social son obligatorios.', 'error')
                else:
                    # Cachear el nombre para mostrar en el listado.
                    os_obj = (session.query(ObsObraSocial)
                              .filter_by(observer_id=os_observer_id).first())
                    nombre_cached = (os_obj.descripcion if os_obj else f'OS#{os_observer_id}')
                    existe = (session.query(RolFiltroObraSocial)
                              .filter_by(rol=rol, obra_social_observer_id=os_observer_id)
                              .first())
                    if existe:
                        flash('Ese filtro ya existe para ese rol.', 'error')
                    else:
                        session.add(RolFiltroObraSocial(
                            rol=rol,
                            obra_social_observer_id=os_observer_id,
                            nombre_cached=nombre_cached,
                        ))
                        session.commit()
                        flash(f'OS "{nombre_cached}" ocultada para rol {rol}.', 'success')
                return redirect(url_for('devoluciones_filtros_os'))

            filtros = (session.query(RolFiltroObraSocial)
                       .order_by(RolFiltroObraSocial.rol,
                                 RolFiltroObraSocial.nombre_cached).all())
            obras_sociales = [{
                'id': o.observer_id, 'nombre': o.descripcion,
            } for o in (session.query(ObsObraSocial)
                        .filter(ObsObraSocial.fecha_baja.is_(None))
                        .order_by(ObsObraSocial.descripcion).all())]
        return render_template('devoluciones_filtros_os.html',
                               filtros=filtros, obras_sociales=obras_sociales)

    @app.route('/rend-recetas/filtros-os/<int:id>/eliminar', methods=['POST'])
    @login_required
    def devoluciones_filtro_os_eliminar(id):
        from database import RolFiltroObraSocial
        with database.get_db() as session:
            f = session.get(RolFiltroObraSocial, id)
            if f:
                session.delete(f)
                session.commit()
                flash('Filtro eliminado.', 'success')
        return redirect(url_for('devoluciones_filtros_os'))

    # ──────────────────────────────────────────────────────────────────
    # ABM Motivos
    # ──────────────────────────────────────────────────────────────────
    _USO_ROL_VALIDO = ('rendicion', 'auditor', 'ambos')

    @app.route('/rend-recetas/motivos', methods=['GET', 'POST'])
    @login_required
    def devoluciones_motivos():
        with database.get_db() as session:
            if request.method == 'POST':
                nombre = (request.form.get('nombre') or '').strip()
                uso_rol = (request.form.get('uso_rol') or 'auditor').strip()
                if uso_rol not in _USO_ROL_VALIDO:
                    uso_rol = 'auditor'
                if not nombre:
                    flash('Nombre obligatorio.', 'error')
                else:
                    existe = session.query(database.MotivoDevolucion).filter_by(nombre=nombre).first()
                    if existe:
                        flash('Ya existe un motivo con ese nombre.', 'error')
                    else:
                        session.add(database.MotivoDevolucion(nombre=nombre, uso_rol=uso_rol))
                        session.commit()
                        flash('Motivo creado.', 'success')
                return redirect(url_for('devoluciones_motivos'))
            motivos = (session.query(database.MotivoDevolucion)
                       .order_by(database.MotivoDevolucion.activo.desc(),
                                 database.MotivoDevolucion.uso_rol,
                                 database.MotivoDevolucion.nombre).all())
        return render_template('devoluciones_motivos.html', motivos=motivos)

    @app.route('/rend-recetas/motivos/<int:id>/toggle', methods=['POST'])
    @login_required
    def devoluciones_motivo_toggle(id):
        with database.get_db() as session:
            m = session.get(database.MotivoDevolucion, id)
            if m:
                m.activo = not m.activo
                session.commit()
        return redirect(url_for('devoluciones_motivos'))

    @app.route('/rend-recetas/motivos/<int:id>/uso-rol', methods=['POST'])
    @login_required
    def devoluciones_motivo_uso_rol(id):
        nuevo = (request.form.get('uso_rol') or '').strip()
        if nuevo not in _USO_ROL_VALIDO:
            flash('Rol de uso inválido.', 'error')
            return redirect(url_for('devoluciones_motivos'))
        with database.get_db() as session:
            m = session.get(database.MotivoDevolucion, id)
            if m:
                m.uso_rol = nuevo
                session.commit()
        return redirect(url_for('devoluciones_motivos'))

    @app.route('/rend-recetas/motivos/<int:id>/bloquea-rendida', methods=['POST'])
    @login_required
    def devoluciones_motivo_bloquea_rendida(id):
        """Toggle: si el motivo bloquea el check 'Rendida' (receta no disponible
        para rendir — ej. EXTRAVIADA, la tiene el cadete)."""
        with database.get_db() as session:
            m = session.get(database.MotivoDevolucion, id)
            if m:
                m.bloquea_rendida = not m.bloquea_rendida
                session.commit()
        return redirect(url_for('devoluciones_motivos'))

    @app.route('/rend-recetas/motivos/<int:id>/eliminar', methods=['POST'])
    @login_required
    def devoluciones_motivo_eliminar(id):
        with database.get_db() as session:
            usado = (session.query(database.DevolucionReceta)
                     .filter_by(motivo_id=id).first())
            if usado:
                flash('No se puede eliminar: el motivo está en uso. Desactivalo.', 'error')
                return redirect(url_for('devoluciones_motivos'))
            m = session.get(database.MotivoDevolucion, id)
            if m:
                session.delete(m)
                session.commit()
                flash('Motivo eliminado.', 'success')
        return redirect(url_for('devoluciones_motivos'))

