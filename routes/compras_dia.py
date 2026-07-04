"""Pantalla "Compra del día" — punto de entrada del flujo de pedidos a Kel/20j.

Muestra la matriz semanal de horarios de reparto de cada droguería + countdown
live al próximo cierre. Desde acá se entra al armado del pedido.

Empleados pueden editar la tabla de horarios; descuentos quedan fuera del scope
de este rol (ver decisión de roles).
"""
import math
from datetime import date as _date
from datetime import datetime
from datetime import timedelta as _td

from flask import flash, jsonify, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from sqlalchemy import or_, text

import database
from database import ProveedorHorarioReparto, Provider, get_db
from purchase_engine import AVG_DAYS_PER_MONTH, rotation_index
from purchase_helpers import calcular_min_sugerido, clasificar_min, pvp_reciente
from services.horarios import horarios_por_dia, proximo_cierre, urgencia_cierre

DIAS_LABELS = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']

# Cobertura objetivo default para entrar al armado: si stock cubre menos de N
# días de venta proyectada (basado en u12m), incluirlo aunque esté arriba del
# mín. Se puede sobrescribir por query param `?target=N`.
TARGET_DIAS_COBERTURA_DEFAULT = 7


def _calcular_labs_con_alertas(session):
    """Labs con al menos 1 producto bajo mínimo en ObServer.

    Lo usa la pantalla /compras/laboratorio (selector de lab para armar
    pedido por laboratorio). Devuelve [{lab_id, nombre, n}, ...] orden alfa.
    """
    from sqlalchemy import distinct as _distinct
    from sqlalchemy import func as _func2

    from database import ObsLaboratorio, ObsProducto, ObsStock
    _stock_lab = (
        session.query(
            ObsStock.producto_observer.label('pid'),
            _func2.sum(ObsStock.stock_actual).label('stock'),
            _func2.sum(ObsStock.minimo).label('minimo'),
        )
        .filter(ObsStock.minimo.isnot(None), ObsStock.minimo > 0)
        .group_by(ObsStock.producto_observer)
        .subquery()
    )
    q = (
        session.query(
            ObsLaboratorio.observer_id,
            ObsLaboratorio.descripcion,
            _func2.count(_distinct(ObsProducto.observer_id)).label('n'),
        )
        .join(ObsProducto,
              ObsProducto.laboratorio_observer == ObsLaboratorio.observer_id)
        .join(_stock_lab, _stock_lab.c.pid == ObsProducto.observer_id)
        .filter(ObsProducto.fecha_baja.is_(None))
        .filter(_stock_lab.c.stock < _stock_lab.c.minimo)
        .group_by(ObsLaboratorio.observer_id, ObsLaboratorio.descripcion)
        .order_by(ObsLaboratorio.descripcion.asc())
    )
    return [{'lab_id': r[0], 'nombre': r[1], 'n': int(r[2])} for r in q.all()]


def _sigla_drog(nombre):
    """Genera sigla corta para una droguería: 'Kellerhoff' → 'Kel',
    '20 de Junio' → '20J'. Para mostrar como badge en el armado."""
    if not nombre:
        return '?'
    n = str(nombre).strip()
    for suf in (' S.A.', ' S.R.L.', ' S.A', ' SA', ' SRL', ' S.A.S'):
        if n.endswith(suf):
            n = n[:-len(suf)].strip()
    # Strip prefijos genéricos para que "Drogueria 20 de Junio" → "20J" y no "Dro"
    _low = n.lower()
    for pref in ('drogueria ', 'droguería ', 'distribuidora '):
        if _low.startswith(pref):
            n = n[len(pref):].strip()
            break
    if n and n[0].isdigit():
        partes = n.split()
        digs = ''.join(c for c in partes[0] if c.isdigit())
        ult = ''
        for p in reversed(partes):
            if p and p[0].isalpha():
                ult = p[0].upper()
                break
        return f'{digs}{ult}'
    # Solo letras → primeras 3 caps de la primera palabra significativa
    primera = n.split()[0] if n.split() else n
    return primera[:3].title()


def _recalc_item_canonico(it):
    """Canónico = COALESCE(confirmada_obs, revisada_op, 0). Setea estado."""
    if it.cantidad_confirmada_obs is not None:
        rec = it.cantidad_confirmada_obs
    elif it.cantidad_revisada_op is not None:
        rec = it.cantidad_revisada_op
    else:
        rec = 0
    it.cantidad_recibida = max(0, rec)
    if it.cantidad_revisada_op is None and it.cantidad_confirmada_obs is None:
        it.estado = 'PENDIENTE'
    elif it.cantidad_recibida <= 0 and it.cantidad_pedida > 0:
        it.estado = 'NO_VINO'
    else:
        it.estado = 'RECIBIDO'


def _recalc_pedido(p):
    estados = [i.estado for i in p.items]
    if all(e != 'PENDIENTE' for e in estados):
        p.estado = 'CERRADO'
    elif any(e in ('RECIBIDO', 'NO_VINO') for e in estados):
        p.estado = 'RECIBIDO_PARCIAL'
    else:
        p.estado = 'ABIERTO'


def init_app(app):

    @app.route('/pedidos/dia')
    @login_required
    def compras_dia():
        with get_db() as session:
            # Drogerías candidatas: las que tengan al menos 1 horario cargado.
            prov_ids = [r[0] for r in session.query(ProveedorHorarioReparto.proveedor_id)
                        .distinct().all()]
            from sqlalchemy import case as _case

            from database import PedidoEmitido
            provs = (session.query(Provider)
                     .filter(Provider.id.in_(prov_ids), Provider.activo.is_(True))
                     .order_by(
                         _case((Provider.matriz_orden.isnot(None), Provider.matriz_orden), else_=9999),
                         Provider.razon_social)
                     .all())
            # Conteo de pedidos activos (no CERRADO) por droguería para el badge
            from sqlalchemy import func
            pedidos_activos = dict(
                session.query(PedidoEmitido.drogueria_id, func.count(PedidoEmitido.id))
                .filter(PedidoEmitido.estado != 'CERRADO')
                .group_by(PedidoEmitido.drogueria_id)
                .all()
            )
            # Pre-fetch plantillas de pedido por droguería (default primero, sino la primera)
            import json as _json

            from database import Plantilla as _Plantilla
            _plant_rows = (session.query(_Plantilla)
                           .filter(_Plantilla.entidad_tipo == 'drogueria',
                                   _Plantilla.tipo_doc == 'pedido')
                           .order_by(_Plantilla.entidad_id,
                                     _Plantilla.es_default.desc(),
                                     _Plantilla.id)
                           .all())
            _plant_map = {}  # drogueria_id → {nombre, formato, n_campos}
            for _pl in _plant_rows:
                if _pl.entidad_id in _plant_map:
                    continue
                try:
                    _cfg = _json.loads(_pl.config_json or '{}')
                    _cols = _cfg.get('columnas') or _cfg.get('campos') or []
                    _n = len(_cols)
                except Exception:
                    _n = 0
                _plant_map[_pl.entidad_id] = {
                    'nombre': _pl.nombre,
                    'formato': _pl.formato.upper(),
                    'n_campos': _n,
                    'es_default': bool(_pl.es_default),
                }

            proveedores = []
            for p in provs:
                matriz = horarios_por_dia(session, p.id)  # {0: ['07:10', ...], ...}
                # ordenado por dia_semana
                matriz_ordenada = [matriz.get(d, []) for d in range(7)]
                cierre = proximo_cierre(session, p.id)
                urgencia = urgencia_cierre(cierre['falta_segundos']) if cierre else None
                proveedores.append({
                    'id': p.id,
                    'nombre': p.razon_social,
                    'horarios_por_dia': matriz_ordenada,
                    'proximo_cierre': cierre,
                    'urgencia': urgencia,
                    'pedidos_activos': pedidos_activos.get(p.id, 0),
                    'plantilla': _plant_map.get(p.id),
                })
            # Dropdown "Cargar/editar horarios": las droguerías ACTIVAS EN LA
            # MATRIZ (matriz_visible=True, definidas con el botón "Columnas de
            # droguería" de la matriz), tengan o no horarios. Selecccionar una
            # abre el modal de edición (carga las existentes o arranca vacía).
            # No usamos "las que no tienen horarios": eso mostraba droguerías que
            # no van a la matriz (Ciafarma/PHARMOS) y ocultaba las que sí.
            from sqlalchemy import case as _case_dh
            _con_hor = set(prov_ids)
            _drogs_matriz = (
                session.query(Provider)
                .filter(Provider.tipo == 'drogueria',
                        Provider.activo.is_(True),
                        Provider.matriz_visible.is_(True))
                .order_by(
                    _case_dh((Provider.matriz_orden.isnot(None), Provider.matriz_orden), else_=9999),
                    Provider.razon_social)
                .all()
            )
            sin_horarios = [{'id': p.id, 'nombre': p.razon_social,
                             'tiene_horarios': p.id in _con_hor}
                            for p in _drogs_matriz]

            # Card "Comportamientos activos": resumen de ProductoFlag vigentes,
            # agrupado por slug. Se muestra arriba para que el operador vea al
            # toque qué productos tienen reglas especiales hoy (sin tener que
            # entrar al armado para descubrirlos).
            import json as _json_cb
            from datetime import date as _date_cb

            from sqlalchemy import func as _func2

            from database import ProductoFlag, TipoPedidoConfig
            _hoy = _date_cb.today()
            _flag_rows = (session.query(ProductoFlag.flag_slug,
                                        _func2.count(ProductoFlag.id))
                          .filter(or_(ProductoFlag.vigente_hasta.is_(None),
                                      ProductoFlag.vigente_hasta >= _hoy))
                          .group_by(ProductoFlag.flag_slug)
                          .order_by(_func2.count(ProductoFlag.id).desc())
                          .all())
            comportamientos = []
            comportamientos_total = 0
            if _flag_rows:
                _slugs = [r[0] for r in _flag_rows]
                _cfg_rows = (session.query(TipoPedidoConfig)
                             .filter(TipoPedidoConfig.slug.in_(_slugs),
                                     TipoPedidoConfig.categoria == 'flag').all())
                _cfg_por_slug = {c.slug: c for c in _cfg_rows}
                for slug, cnt in _flag_rows:
                    cfg = _cfg_por_slug.get(slug)
                    cfg_d = {}
                    if cfg and cfg.config_json:
                        try:
                            cfg_d = _json_cb.loads(cfg.config_json)
                        except Exception:
                            cfg_d = {}
                    comportamientos.append({
                        'slug': slug,
                        'nombre': cfg.nombre if cfg else slug,
                        'icono': cfg_d.get('icono', '🚩'),
                        'color': cfg_d.get('color', 'sky'),
                        'efecto_armado': cfg_d.get('efecto_armado', 'ninguno'),
                        'count': int(cnt),
                    })
                    comportamientos_total += int(cnt)

            # ── Tabla única de cierres: ventana deslizante centrada en AHORA ──
            # Todas las droguerías intercaladas en orden cronológico, cruzando
            # días: mostramos los 3 cierres más recientes ya pasados (gris) + los
            # próximos N futuros. Así siempre hay 3-4 cierres "hacia adelante"
            # visibles aunque cambie el día. Dinámico: sale de `proveedores`.
            from datetime import datetime as _dt_h
            from datetime import time as _time_h
            from datetime import timedelta as _td_h

            from services.horarios import _parse_hhmm
            _PASADOS_N = 3
            _FUTUROS_N = 4
            _DIAS_ATRAS = 4    # ventana de búsqueda hacia atrás (cubre fin de semana)
            _DIAS_ADELANTE = 8  # hacia adelante
            _ahora = _dt_h.now()
            _hoy_d = _ahora.date()
            _todos = []
            for _p in proveedores:
                _matriz = _p['horarios_por_dia'] or []
                if not _matriz:
                    continue
                for _delta in range(-_DIAS_ATRAS, _DIAS_ADELANTE + 1):
                    _d = _hoy_d + _td_h(days=_delta)
                    for _hora_str in _matriz[_d.weekday()]:
                        _hm = _parse_hhmm(_hora_str)
                        if not _hm:
                            continue
                        _dt = _dt_h.combine(_d, _time_h(_hm[0], _hm[1]))
                        _falta = int((_dt - _ahora).total_seconds())
                        # Etiqueta de día relativa.
                        _dd = _delta if _delta in (-1, 0, 1) else None
                        if _dd == 0:
                            _dia_lbl = 'hoy'
                        elif _dd == 1:
                            _dia_lbl = 'mañana'
                        elif _dd == -1:
                            _dia_lbl = 'ayer'
                        else:
                            _dia_lbl = ['lun', 'mar', 'mié', 'jue', 'vie', 'sáb', 'dom'][_d.weekday()] + f' {_d.day}'
                        _todos.append({
                            'prov_id': _p['id'],
                            'prov_nombre': _p['nombre'],
                            'hora': _hora_str,
                            'dia_label': _dia_lbl,
                            'es_hoy': _delta == 0,
                            'dt': _dt,
                            'falta_segundos': _falta,
                            'pasado': _falta < 0,
                            'urgencia': urgencia_cierre(_falta) if _falta >= 0 else None,
                            'plantilla': _p['plantilla'],
                            'pedidos_activos': _p['pedidos_activos'],
                        })
            _todos.sort(key=lambda s: s['dt'])
            _pasados = [s for s in _todos if s['falta_segundos'] < 0]
            _futuros = [s for s in _todos if s['falta_segundos'] >= 0]
            slots_hoy = _pasados[-_PASADOS_N:] + _futuros[:_FUTUROS_N]

            # ── ¿Hubo pedido emitido para cada slot YA PASADO? ──
            # Un pedido "cuenta" para un cierre si se emitió a esa droguería en la
            # ventana (cierre anterior de la misma drog, este cierre]. Sirve para
            # ver de un vistazo si se respondió o se dejó pasar un reparto.
            from collections import defaultdict as _dd_h
            _dt_por_drog = _dd_h(list)
            for _t in _todos:
                _dt_por_drog[_t['prov_id']].append(_t['dt'])
            for _lst in _dt_por_drog.values():
                _lst.sort()
            for _s in slots_hoy:
                _s['tuvo_pedido'] = None  # solo aplica a pasados
                if _s['falta_segundos'] >= 0:
                    continue
                _cierre = _s['dt']
                # Cierre anterior de la MISMA droguería (ventana del reparto).
                _previos = [d for d in _dt_por_drog[_s['prov_id']] if d < _cierre]
                _ini = _previos[-1] if _previos else (_cierre - _td_h(days=7))
                _n_ped = (session.query(PedidoEmitido)
                          .filter(PedidoEmitido.drogueria_id == _s['prov_id'],
                                  PedidoEmitido.fecha > _ini,
                                  PedidoEmitido.fecha <= _cierre)
                          .count())
                _s['tuvo_pedido'] = _n_ped > 0

            for _s in slots_hoy:
                _s.pop('dt', None)  # no serializable / no se usa en template

            # ── Frescura de datos para el banner "Sincronizar todo" ──
            # Tomamos la MÁS VIEJA de las max(sync_en) de las tablas críticas para
            # el armado (stock + ventas). Así el "hace Xh" garantiza que TODO lo
            # que alimenta el pedido está al menos así de fresco. /pedidos/dia es
            # local-only → siempre hay ObServer, no hace falta gatear.
            from sqlalchemy import func as _func_sync

            from database import ObsStock, ObsVentaMensual
            _maxes = []
            for _modelo in (ObsStock, ObsVentaMensual):
                _mx = session.query(_func_sync.max(_modelo.sync_en)).scalar()
                if _mx is not None:
                    _maxes.append(_mx)
            sync_age_min = None
            if _maxes:
                _viejo = min(_maxes)  # la tabla menos fresca manda
                sync_age_min = int((datetime.now() - _viejo).total_seconds() // 60)

        return render_template('compras_dia.html',
                               proveedores=proveedores,
                               slots_hoy=slots_hoy,
                               sin_horarios=sin_horarios,
                               dias=DIAS_LABELS,
                               sync_age_min=sync_age_min,
                               comportamientos=comportamientos,
                               comportamientos_total=comportamientos_total)

    @app.route('/compras/laboratorio')
    @login_required
    def compras_laboratorio():
        """Selector de laboratorio para armar pedido por lab (entrada desde el
        card "Compras Laboratorio" del home). Lista labs con productos bajo
        mínimo; al elegir uno → /pedidos/dia/armar?lab_id=N.

        Cada fila se enriquece con: lab local id, si tiene plantilla de export
        configurada, fecha del módulo activo y fecha de la última transfer
        (OfertaMinimo) — para que el operador vea de un vistazo qué tan
        actualizado está cada lab.
        """
        from sqlalchemy import func as _f

        from database import (
            Laboratorio,
            Modulo,
            OfertaMinimo,
            Plantilla,
            Provider,
        )
        with get_db() as session:
            labs = _calcular_labs_con_alertas(session)
            obs_ids = [l['lab_id'] for l in labs]

            # Map observer_id → Laboratorio local id (los módulos/ofertas/
            # plantillas se relacionan por el id LOCAL del lab).
            local_por_obs = {}
            usa_packs_por_local = {}
            if obs_ids:
                for obs_id, loc_id, up in (session.query(
                        Laboratorio.observer_id, Laboratorio.id, Laboratorio.usa_packs)
                        .filter(Laboratorio.observer_id.in_(obs_ids)).all()):
                    local_por_obs[obs_id] = loc_id
                    usa_packs_por_local[loc_id] = bool(up)
            local_ids = list(local_por_obs.values())

            # Plantillas configuradas (tabla unificada Plantilla, entidad lab).
            # Guardamos el nombre de la plantilla por lab (la default primero;
            # si no hay default, la primera alfabética) para mostrarlo en la fila.
            plantilla_nombre = {}
            if local_ids:
                for ent_id, nombre, es_def in (session.query(
                        Plantilla.entidad_id, Plantilla.nombre, Plantilla.es_default)
                        .filter(Plantilla.entidad_tipo == 'laboratorio',
                                Plantilla.entidad_id.in_(local_ids))
                        .order_by(Plantilla.es_default.desc(), Plantilla.nombre).all()):
                    # Primera aparición por lab gana (default > alfabético).
                    if ent_id not in plantilla_nombre:
                        plantilla_nombre[ent_id] = nombre

            # Módulo activo más reciente por lab (Modulo.activo, creado_en).
            modulo_fecha = {}
            if local_ids:
                for lab_local_id, fmax in (session.query(
                        Modulo.laboratorio_id, _f.max(Modulo.creado_en))
                        .filter(Modulo.laboratorio_id.in_(local_ids),
                                Modulo.activo.is_(True))
                        .group_by(Modulo.laboratorio_id).all()):
                    modulo_fecha[lab_local_id] = fmax

            # Última transfer (OfertaMinimo) por lab — fecha de actualización.
            transfer_fecha = {}
            if local_ids:
                for lab_local_id, fmax in (session.query(
                        OfertaMinimo.laboratorio_id, _f.max(OfertaMinimo.actualizado_en))
                        .filter(OfertaMinimo.laboratorio_id.in_(local_ids))
                        .group_by(OfertaMinimo.laboratorio_id).all()):
                    transfer_fecha[lab_local_id] = fmax

            for l in labs:
                lid = local_por_obs.get(l['lab_id'])
                l['lab_local_id'] = lid
                l['usa_packs'] = usa_packs_por_local.get(lid, False)
                l['plantilla_nombre'] = plantilla_nombre.get(lid)
                l['tiene_plantilla'] = lid in plantilla_nombre
                mf = modulo_fecha.get(lid)
                tf = transfer_fecha.get(lid)
                l['modulo_fecha'] = mf.strftime('%d/%m/%Y') if mf else None
                l['transfer_fecha'] = tf.strftime('%d/%m/%Y') if tf else None

            droguerias = [{'id': p.id, 'nombre': p.razon_social}
                          for p in (session.query(Provider)
                                    .filter(Provider.tipo == 'drogueria',
                                            Provider.activo.is_(True))
                                    .order_by(Provider.razon_social).all())]

        return render_template('compras_laboratorio.html', labs_con_alertas=labs,
                               droguerias=droguerias)

    @app.route('/compras/laboratorio/<int:obs_lab_id>/comprar-modulos', methods=['POST'])
    @login_required
    def compras_laboratorio_comprar_modulos(obs_lab_id):
        """Crea un Pedido nuevo con el universo bajo-mínimo del lab y abre el
        análisis de módulos (/order/<id>). Solo para labs usa_packs. El sugerido
        sembrado es el gap (mínimo − stock); el análisis lo refina con módulos."""
        from sqlalchemy import func as _f

        from database import (
            Laboratorio,
            ObsCodigoBarras,
            ObsLaboratorio,
            ObsProducto,
            ObsStock,
            ObsVentaMensual,
            Pedido,
            PedidoItem,
        )
        from helpers import now_ar as _now

        # Días de cobertura + canal (lo elige el operador en el modal, al inicio).
        dias = request.form.get('dias', type=int) or 30
        dias = max(1, min(dias, 180))
        canal = (request.form.get('canal') or 'laboratorio').strip()
        if canal not in ('laboratorio', 'drogueria'):
            canal = 'laboratorio'
        partner_id = request.form.get('partner_id', type=int) if canal == 'drogueria' else None
        with get_db() as session:
            lab_obs = session.get(ObsLaboratorio, obs_lab_id)
            if not lab_obs:
                flash('Laboratorio no encontrado.', 'error')
                return redirect(url_for('compras_laboratorio'))
            lab_local = (session.query(Laboratorio)
                         .filter_by(observer_id=obs_lab_id).first())
            if not (lab_local and lab_local.usa_packs):
                flash('Ese laboratorio no está marcado como "compra por módulos".', 'error')
                return redirect(url_for('compras_laboratorio'))

            # Bajo mínimo del lab (stock < mínimo) desde obs_stock.
            rows = (session.query(
                        ObsStock.producto_observer.label('pid'),
                        _f.sum(ObsStock.stock_actual).label('stock'),
                        _f.sum(ObsStock.minimo).label('minimo'))
                    .join(ObsProducto, ObsProducto.observer_id == ObsStock.producto_observer)
                    .filter(ObsProducto.laboratorio_observer == obs_lab_id,
                            ObsStock.minimo.isnot(None), ObsStock.minimo > 0)
                    .group_by(ObsStock.producto_observer).all())
            bajo = [(r.pid, int(r.stock or 0), int(r.minimo or 0)) for r in rows
                    if int(r.stock or 0) < int(r.minimo or 0)]
            if not bajo:
                flash(f'{lab_obs.descripcion}: no hay productos bajo mínimo.', 'info')
                return redirect(url_for('compras_laboratorio'))

            pids = [b[0] for b in bajo]
            nombres = {o.observer_id: o.descripcion for o in session.query(ObsProducto)
                       .filter(ObsProducto.observer_id.in_(pids))}
            ean_por_pid = {}
            for oid, cb in (session.query(ObsCodigoBarras.producto_observer,
                                          ObsCodigoBarras.codigo_barras)
                            .filter(ObsCodigoBarras.producto_observer.in_(pids),
                                    ObsCodigoBarras.fecha_baja.is_(None),
                                    ObsCodigoBarras.orden == 1)):
                ean_por_pid.setdefault(oid, cb)
            v12 = {}
            for oid, u, m in (session.query(ObsVentaMensual.producto_observer,
                                            _f.sum(ObsVentaMensual.unidades),
                                            _f.sum(ObsVentaMensual.monto))
                              .filter(ObsVentaMensual.producto_observer.in_(pids))
                              .group_by(ObsVentaMensual.producto_observer)):
                v12[oid] = (int(u or 0), float(m or 0))

            items = []
            for pid, stock, minimo in bajo:
                u12, m12 = v12.get(pid, (0, 0.0))
                pvp = round(m12 / u12, 2) if u12 else 0.0
                # Sugerido = cubrir `dias` de venta (venta diaria × días) − stock.
                # Piso: el gap al mínimo. Mínimo 1 (está bajo mínimo).
                diaria = (u12 / 365.0) if u12 else 0.0
                objetivo = math.ceil(diaria * dias)
                sugerido = max(1, minimo - stock, objetivo - stock)
                items.append(PedidoItem(
                    codigo_barra=(ean_por_pid.get(pid) or f'OBS:{pid}'),
                    nombre=nombres.get(pid, ''),
                    cantidad=sugerido,
                    precio_pvp=pvp,
                    subtotal=round(sugerido * pvp, 2),
                    avg_monthly=round(u12 / 12, 1) if u12 else None,
                ))
            _farmacia = getattr(current_user, 'username', None) or 'Administrador'
            _periodo = f"{_now().strftime('%d-%m')} · módulos {dias}d"
            # Reusar el borrador de módulos del lab (no acumula): si ya hay uno,
            # le regeneramos los ítems; si no, lo creamos. Recién es pedido real
            # cuando el operador confirma en el análisis (BORRADOR → PENDIENTE).
            pedido = (session.query(Pedido)
                      .filter(Pedido.laboratorio == lab_obs.descripcion,
                              Pedido.estado == 'BORRADOR',
                              Pedido.origen == 'Modulos').first())
            if pedido:
                for it in list(pedido.items):
                    session.delete(it)
                session.flush()
                pedido.items = items
                pedido.farmacia = _farmacia
                pedido.periodo = _periodo
                pedido.n_days = dias
            else:
                pedido = Pedido(
                    laboratorio=lab_obs.descripcion, farmacia=_farmacia,
                    periodo=_periodo, n_days=dias, items=items, origen='Modulos',
                    estado='BORRADOR',
                )
                session.add(pedido)
            pedido.estado = 'BORRADOR'
            pedido.canal = canal
            pedido.partner_id = partner_id
            pedido.canal_elegido_en = _now()
            session.flush()
            pedido_id = pedido.id
            session.commit()
        return redirect(url_for('order_detail', pedido_id=pedido_id))

    @app.route('/api/drogueria/<int:prov_id>/pedidos-emitidos')
    @login_required
    def api_drogueria_pedidos_emitidos(prov_id):
        from database import PedidoEmitido
        with get_db() as session:
            pedidos = (session.query(PedidoEmitido)
                       .filter_by(drogueria_id=prov_id)
                       .order_by(PedidoEmitido.fecha.desc())
                       .all())
            data = []
            for p in pedidos:
                data.append({
                    'id': p.id,
                    'fecha': p.fecha.strftime('%d/%m/%Y %H:%M') if p.fecha else '—',
                    'estado': p.estado,
                    'total_items': p.total_items,
                    'recibido_por': p.recibido_por,
                    'cargado_por': p.cargado_por,
                    'tiene_factura': False,
                    'drogueria_id': prov_id,
                    'drogueria_nombre': p.drogueria.razon_social if p.drogueria else '—',
                })
        return jsonify({'ok': True, 'pedidos': data})

    @app.route('/api/pedidos-emitidos/todos')
    @login_required
    def api_pedidos_emitidos_todos():
        """Devuelve los pedidos emitidos recientes (últimos N días, por default 30).
        Acepta ?dias=N&limit=M. Sin filtrar por droguería — cada armado multi-drog
        genera ~1 PedidoEmitido por droguería; este endpoint los muestra unificados.
        """
        from datetime import datetime, timedelta

        from database import PedidoEmitido
        try:
            dias = max(1, min(int(request.args.get('dias', 30)), 365))
        except (TypeError, ValueError):
            dias = 30
        try:
            limit = max(1, min(int(request.args.get('limit', 200)), 1000))
        except (TypeError, ValueError):
            limit = 200
        desde = datetime.now() - timedelta(days=dias)
        with get_db() as session:
            pedidos = (session.query(PedidoEmitido)
                       .filter(PedidoEmitido.fecha >= desde)
                       .order_by(PedidoEmitido.fecha.desc())
                       .limit(limit)
                       .all())
            data = []
            for p in pedidos:
                data.append({
                    'id': p.id,
                    'fecha': p.fecha.strftime('%d/%m/%Y %H:%M') if p.fecha else '—',
                    'fecha_grupo': p.fecha.strftime('%Y-%m-%dT%H:%M') if p.fecha else '',
                    'estado': p.estado,
                    'total_items': p.total_items,
                    'recibido_por': p.recibido_por,
                    'cargado_por': p.cargado_por,
                    'tiene_factura': False,
                    'drogueria_id': p.drogueria_id,
                    'drogueria_nombre': p.drogueria.razon_social if p.drogueria else '—',
                    'drogueria_sigla': _sigla_drog(p.drogueria.razon_social) if p.drogueria else '?',
                })
        return jsonify({'ok': True, 'pedidos': data})

    @app.route('/api/pedido-emitido/<int:pedido_id>', methods=['DELETE'])
    @login_required
    def api_pedido_emitido_borrar(pedido_id):
        if getattr(current_user, 'rol', None) not in ('dev', 'admin'):
            return jsonify({'ok': False, 'error': 'Sin permisos'}), 403
        from database import PedidoEmitido
        with get_db() as session:
            p = session.get(PedidoEmitido, pedido_id)
            if not p:
                return jsonify({'ok': False, 'error': 'No encontrado'}), 404
            session.delete(p)
            session.commit()
        return jsonify({'ok': True})

    @app.route('/api/pedidos/dia/countdown')
    @login_required
    def api_compras_dia_countdown():
        """Devuelve el próximo cierre + segundos faltantes para cada drog activa.
        Lo consume el JS del front para refrescar el countdown sin recargar la página.
        """
        ahora = datetime.now()
        out = {}
        with get_db() as session:
            prov_ids = [r[0] for r in session.query(ProveedorHorarioReparto.proveedor_id)
                        .distinct().all()]
            for pid in prov_ids:
                cierre = proximo_cierre(session, pid, ahora=ahora)
                if cierre:
                    out[pid] = {
                        'fecha': cierre['fecha'].isoformat(),
                        'falta_segundos': cierre['falta_segundos'],
                        'hora_str': cierre['hora_str'],
                        'urgencia': urgencia_cierre(cierre['falta_segundos']),
                    }
        return jsonify({'ok': True, 'now': ahora.isoformat(), 'cierres': out})

    @app.route('/api/pedidos/dia/horarios/<int:proveedor_id>', methods=['GET', 'POST', 'DELETE'])
    @login_required
    def api_horarios_crud(proveedor_id):
        """CRUD básico para horarios de un proveedor.

        GET    → lista los horarios.
        POST   → agrega un slot. body: {dia_semana: 0-6, hora: 'HH:MM'}.
        DELETE → borra slot. body: {id: <slot_id>}.
        """
        with get_db() as session:
            p = session.get(Provider, proveedor_id)
            if not p:
                return jsonify({'ok': False, 'error': 'Proveedor no encontrado'}), 404

            if request.method == 'GET':
                rows = (session.query(ProveedorHorarioReparto)
                        .filter_by(proveedor_id=proveedor_id, activo=True)
                        .order_by(ProveedorHorarioReparto.dia_semana,
                                  ProveedorHorarioReparto.hora).all())
                return jsonify({'ok': True, 'horarios': [{
                    'id': r.id,
                    'dia_semana': r.dia_semana,
                    'hora': r.hora,
                } for r in rows]})

            if request.method == 'POST':
                data = request.get_json(silent=True) or {}
                try:
                    dia = int(data.get('dia_semana', -1))
                    hora = (data.get('hora') or '').strip()
                except (ValueError, TypeError):
                    return jsonify({'ok': False, 'error': 'Datos inválidos'}), 400
                if dia < 0 or dia > 6 or len(hora) != 5 or hora[2] != ':':
                    return jsonify({'ok': False, 'error': 'Día (0-6) y hora HH:MM requeridos'}), 400
                # Idempotente: si ya existe ese slot, no rompe.
                ya = (session.query(ProveedorHorarioReparto.id)
                      .filter_by(proveedor_id=proveedor_id, dia_semana=dia, hora=hora)
                      .first())
                if ya:
                    return jsonify({'ok': True, 'id': ya[0], 'duplicado': True})
                row = ProveedorHorarioReparto(
                    proveedor_id=proveedor_id, dia_semana=dia, hora=hora, activo=True
                )
                session.add(row)
                session.commit()
                return jsonify({'ok': True, 'id': row.id})

            # DELETE
            data = request.get_json(silent=True) or {}
            try:
                slot_id = int(data.get('id') or 0)
            except (ValueError, TypeError):
                return jsonify({'ok': False, 'error': 'id inválido'}), 400
            row = session.get(ProveedorHorarioReparto, slot_id)
            if row and row.proveedor_id == proveedor_id:
                session.delete(row)
                session.commit()
            return jsonify({'ok': True})

    @app.route('/pedidos/dia/armar')
    @login_required
    def compras_dia_armar():
        """Armado del pedido para una droguería específica.

        Simplificado: solo bajo mínimo en obs_stock + rubro Medicamentos + venta 12m.
        Descuentos por lab/proveedor se evalúan en una fase posterior.
        """
        from sqlalchemy import func

        from database import (
            Laboratorio,
            LaboratorioDrogueria,
            ObsLaboratorio,
            ObsProducto,
            ObsStock,
            ObsVentaDetalle,
            ObsVentaMensual,
            PedidoEmitido,
            PedidoEmitidoItem,
            Producto,
        )

        # prov ahora es OPCIONAL. Si no viene → modo "todas las drogs activas
        # con horarios cargados". Mostramos badge por droguería en cada fila.
        prov_id = request.args.get('prov', type=int)
        # lab_id: cuando viene del dropdown "🧪 Laboratorio" de /pedidos/dia.
        # Filtra el universo a productos del lab y desactiva la matriz lab×drog
        # (el user eligió un solo lab, no hace falta resolver qué drog cubre).
        lab_id = request.args.get('lab_id', type=int)
        # libres_a: cuando se entra desde la tabla de cierres de hoy (/pedidos/dia).
        # Es el id de la droguería del renglón elegido. Modo matriz (sin prov):
        # los productos libres (sin asignación lab×drog) se preasignan a esta
        # droguería en el front (editable después). Solo aplica en modo matriz.
        libres_a = request.args.get('libres_a', type=int)
        # Cobertura objetivo configurable por query param. Default 7 días.
        target_dias = request.args.get('target', type=int) or TARGET_DIAS_COBERTURA_DEFAULT
        target_dias = max(1, min(target_dias, 90))  # clamp 1-90
        # Modo lab: "cubrir N días" reemplaza min_efectivo como piso de ideal.
        # Solo aplica cuando hay lab_id (compra grande directa al lab); si no,
        # el cálculo REPO sigue usando min_efectivo (matriz lab/drog).
        cubrir_dias = request.args.get('cubrir_dias', type=int) or 30
        cubrir_dias = max(1, min(cubrir_dias, 120))
        # Ventana de meses para calcular tasa de rotación diaria. Default 3.
        # Nota: los 12 meses siguen usándose para estacionalidad/forecast en
        # purchase_engine; este parámetro solo afecta `target_unid` (cuánto pedir).
        meses_rotacion = max(1, min(int(request.args.get('meses_rot') or 3), 12))
        DIAS_PROM_MES = 30.42
        dias_rotacion = int(meses_rotacion * DIAS_PROM_MES)
        # Decisión sobre oferta-drog: '1' aplicar, '0' ignorar, None preguntar.
        usar_oferta = (request.args.get('usar_oferta') or '').strip()
        # Rubros: CSV ej. ?rubros=12,5. Default: '12' (Medicamentos).
        # Pasar ?rubros=all (o vacío) para no filtrar.
        rubros_raw = (request.args.get('rubros') or '12').strip()
        if rubros_raw.lower() == 'all' or not rubros_raw:
            rubros_filtro = None  # None = mostrar todos
        else:
            try:
                rubros_filtro = set(int(x) for x in rubros_raw.split(',') if x.strip())
            except ValueError:
                rubros_filtro = {12}  # fallback Medicamentos si vino mal

        with get_db() as session:
            prov = None
            if prov_id:
                prov = session.get(Provider, prov_id)
                if not prov:
                    return redirect(url_for('compras_dia'))

            # Lab pre-seleccionado (flujo "🧪 Laboratorio"). Resolvemos el
            # observer_id del lab, su nombre y su Laboratorio LOCAL (para
            # poder traer ofertas cargadas por lab_id local).
            lab_obs = None
            lab_nombre = None
            lab_local = None
            if lab_id:
                lab_obs = session.get(ObsLaboratorio, lab_id)
                if not lab_obs:
                    return redirect(url_for('compras_dia'))
                lab_nombre = lab_obs.descripcion
                lab_local = (session.query(Laboratorio)
                             .filter(Laboratorio.observer_id == lab_id).first())

            # Universo: bajo mínimo en obs_stock + rubro Medicamentos (12).
            stock_q = (session.query(
                ObsStock.producto_observer.label('pid'),
                func.sum(ObsStock.stock_actual).label('stock'),
                func.sum(ObsStock.minimo).label('minimo'),
                func.sum(ObsStock.maximo).label('maximo'),
            ).filter(ObsStock.minimo.isnot(None), ObsStock.minimo > 0)
              .group_by(ObsStock.producto_observer).subquery())

            # Ventas 12m por producto (agregado para tabla).
            # Incluye monto para derivar PVP estimado = m12m / u12m.
            v12_q = (session.query(
                ObsVentaMensual.producto_observer.label('pid'),
                func.sum(ObsVentaMensual.unidades).label('u12m'),
                func.sum(ObsVentaMensual.monto).label('m12m'),
            ).group_by(ObsVentaMensual.producto_observer).subquery())

            # Ventas ayer y última semana por producto.
            from helpers import excluir_no_medicamentos_ovd, ventas_periodo_filter
            hoy_d = _date.today()
            _ayer = hoy_d - _td(days=1)
            _semana = hoy_d - _td(days=7)
            _det_rows = session.query(
                ObsVentaDetalle.producto_observer,
                ObsVentaDetalle.fecha_estadistica,
                func.sum(ObsVentaDetalle.cantidad).label('cant'),
            ).filter(
                ventas_periodo_filter(ObsVentaDetalle, _semana, hoy_d),
                excluir_no_medicamentos_ovd(ObsVentaDetalle, ObsProducto, session),
             ).group_by(ObsVentaDetalle.producto_observer,
                        ObsVentaDetalle.fecha_estadistica).all()
            v24h_rows = {}
            v7d_rows  = {}
            for pid_d, fec, cant in _det_rows:
                v7d_rows[pid_d] = v7d_rows.get(pid_d, 0) + int(cant or 0)
                if fec >= _ayer:
                    v24h_rows[pid_d] = v24h_rows.get(pid_d, 0) + int(cant or 0)

            # Ventas detalladas por mes para alimentar purchase_engine.
            # Construimos el array de 12 meses ventas[0..11] terminando en el mes actual.
            end_month = hoy_d.month
            start_month = ((end_month - 11 - 1) % 12) + 1  # mes-11 (1..12)
            start_year = hoy_d.year if start_month <= end_month else hoy_d.year - 1

            # Cleanup temporal de exclusión.
            session.execute(text("""
                UPDATE productos p SET excluido_armado_actual = FALSE
                WHERE p.excluido_armado_actual = TRUE
                  AND p.observer_id IN (
                    SELECT s.producto_observer
                      FROM obs_stock s
                     WHERE s.minimo IS NOT NULL AND s.stock_actual > s.minimo
                  )
            """))
            session.commit()

            # Si hay oferta activa cargada para esta drog, hay 3 estados:
            #   - usar_oferta='1' → filtrar SOLO los del archivo.
            #   - usar_oferta='0' → ignorar la oferta, modo armado normal.
            #   - usar_oferta=None → mostrar pregunta intermedia (sin filtrar).
            from database import ObsCodigoBarras as _OCB
            from database import OfertaMinimo as _OM_drog
            oferta_pids = set()
            oferta_disponible_n = 0
            oferta_nombre_drog = None
            oferta_vigencia_hasta = None
            oferta_observacion = None
            if prov_id:
                hoy_o = _date.today()
                # Traemos ean + vigencia + observacion en una sola pasada para
                # poder mostrar info de la oferta en cabecera (vigencia + nombre).
                rows_oferta = (session.query(_OM_drog.ean,
                                              _OM_drog.vigencia_hasta,
                                              _OM_drog.observacion)
                               .filter(_OM_drog.drogueria_id == prov_id,
                                       _OM_drog.activo.is_(True),
                                       or_(_OM_drog.vigencia_hasta.is_(None),
                                           _OM_drog.vigencia_hasta >= hoy_o))
                               .all())
                eans_oferta = list({r[0] for r in rows_oferta if r[0]})
                if rows_oferta:
                    # Tomar la vigencia mas LEJANA (la oferta vigente "mas larga")
                    # y la primera observacion no vacia.
                    vigs = [r[1] for r in rows_oferta if r[1]]
                    if vigs:
                        oferta_vigencia_hasta = max(vigs)
                    for r in rows_oferta:
                        if r[2] and r[2].strip():
                            oferta_observacion = r[2].strip()[:120]
                            break
                if eans_oferta:
                    pids_oferta_full = {r[0] for r in (session.query(_OCB.producto_observer)
                                        .filter(_OCB.codigo_barras.in_(eans_oferta),
                                                _OCB.fecha_baja.is_(None)).distinct().all())}
                    oferta_disponible_n = len(pids_oferta_full)
                    if prov:
                        oferta_nombre_drog = prov.razon_social
                    # Solo aplicar el filtro si el usuario explícitamente eligió '1'.
                    if usar_oferta == '1':
                        oferta_pids = pids_oferta_full

            from database import ObsNombreDroga
            base_q = (session.query(
                ObsProducto.observer_id.label('pid'),
                ObsProducto.descripcion.label('desc'),
                ObsProducto.id_tipo_venta_control.label('tvc'),
                ObsLaboratorio.observer_id.label('lab_obs_id'),
                ObsLaboratorio.descripcion.label('lab_nombre'),
                ObsNombreDroga.descripcion.label('droga_nombre'),
                stock_q.c.stock,
                stock_q.c.minimo,
                func.coalesce(v12_q.c.u12m, 0).label('u12m'),
                func.coalesce(v12_q.c.m12m, 0).label('m12m'),
            )
            .join(stock_q, stock_q.c.pid == ObsProducto.observer_id)
            .outerjoin(ObsLaboratorio,
                       ObsLaboratorio.observer_id == ObsProducto.laboratorio_observer)
            .outerjoin(ObsNombreDroga,
                       ObsNombreDroga.observer_id == ObsProducto.nombre_droga_observer)
            .outerjoin(v12_q, v12_q.c.pid == ObsProducto.observer_id)
            .filter(ObsProducto.fecha_baja.is_(None))
            .filter(ObsProducto.subrubro_observer.isnot(None)))

            # Excluir items que no son medicamentos (sellado de recetas, cupones, etc.).
            from helpers import filtro_solo_medicamentos
            base_q = filtro_solo_medicamentos(base_q, ObsProducto)

            if oferta_pids:
                # Modo oferta: solo productos del archivo (sin importar stock/lab).
                base_q = base_q.filter(ObsProducto.observer_id.in_(oferta_pids))
            else:
                # Universo normal: bajo mínimo OR cobertura insuficiente.
                # Cobertura insuficiente: stock × 365 < TARGET_DIAS × u12m
                # (equivale a stock/avg_diario < TARGET_DIAS, sin división).
                base_q = base_q.filter(or_(
                    stock_q.c.stock <= stock_q.c.minimo,
                    stock_q.c.stock * 365 < target_dias *
                        func.coalesce(v12_q.c.u12m, 0),
                ))

            # Filtro por lab pre-seleccionado: descartamos la lógica multi-drog
            # y solo mostramos productos de ESE lab. Cubre el flujo
            # /pedidos/dia → 🧪 Laboratorio → elegir lab.
            if lab_id:
                base_q = base_q.filter(ObsProducto.laboratorio_observer == lab_id)
            base = base_q.all()

            # Filtro rubro=Medicamentos (12). Lo aplicamos en Python por simplicidad
            # (los rubros viven en obs_subrubros.rubro_observer).
            from database import ObsSubrubro
            subrubro_a_rubro = dict(session.query(ObsSubrubro.observer_id,
                                                   ObsSubrubro.rubro_observer).all())
            obs_pids = [r.pid for r in base]
            sub_de_prod = {}
            if obs_pids:
                rows = (session.query(ObsProducto.observer_id,
                                       ObsProducto.subrubro_observer)
                        .filter(ObsProducto.observer_id.in_(obs_pids)).all())
                sub_de_prod = dict(rows)

            # Cargar ventas mes-a-mes para los productos en juego. Unidades para
            # rotación/forecast; montos para derivar el PVP del último mes con
            # ventas (más fiel que el promedio 12m bajo inflación).
            ventas_por_pid = {pid: [0]*12 for pid in [r.pid for r in base]}
            montos_por_pid = {pid: [0.0]*12 for pid in ventas_por_pid}
            if ventas_por_pid:
                rows_vm = (session.query(ObsVentaMensual.producto_observer,
                                          ObsVentaMensual.anio,
                                          ObsVentaMensual.mes,
                                          func.sum(ObsVentaMensual.unidades),
                                          func.sum(ObsVentaMensual.monto))
                           .filter(ObsVentaMensual.producto_observer.in_(list(ventas_por_pid.keys())))
                           .group_by(ObsVentaMensual.producto_observer,
                                     ObsVentaMensual.anio, ObsVentaMensual.mes)
                           .all())
                for pid_v, anio, mes, uds, mto in rows_vm:
                    # Slot 0..11: ventas[0]=start_month, ventas[11]=end_month
                    offset = (anio - start_year) * 12 + (mes - start_month)
                    if 0 <= offset <= 11 and pid_v in ventas_por_pid:
                        ventas_por_pid[pid_v][offset] += int(uds or 0)
                        montos_por_pid[pid_v][offset] += float(mto or 0)

            # Resolver Producto local + flags excluido / no_pedir +
            # cantidad_reposicion_fija (override del cálculo dinámico).
            local_por_obs = {}
            if obs_pids:
                # fraccionado (master) + cantidad_envase (ProductoAtributo, lo que
                # se carga en Presentación) → para mostrar la equivalencia en cajas.
                rows = (session.query(Producto.observer_id, Producto.id,
                                       Producto.excluido_armado_actual,
                                       Producto.no_pedir, Producto.laboratorio_id,
                                       Producto.cantidad_reposicion_fija,
                                       Producto.fraccionado,
                                       database.ProductoAtributo.cantidad_envase)
                        .outerjoin(database.ProductoAtributo,
                                   database.ProductoAtributo.producto_id == Producto.id)
                        .filter(Producto.observer_id.in_(obs_pids)).all())
                local_por_obs = {r[0]: {
                    'id': r[1], 'excluido': r[2], 'no_pedir': r[3],
                    'lab_local_id': r[4],
                    'cantidad_reposicion_fija': r[5],
                    'fraccionado': bool(r[6]),
                    'cantidad_envase': int(r[7]) if r[7] else None,
                } for r in rows}

            # Modo multi-drog: necesitamos saber qué drog(s) cubre cada lab.
            # `labs_a_drogs[lab_id]` = lista de prov_ids que cubren ese lab.
            from collections import defaultdict
            labs_a_drogs = defaultdict(list)
            for ld_lab, ld_drog in (session.query(
                LaboratorioDrogueria.laboratorio_id,
                LaboratorioDrogueria.drogueria_id,
            ).all()):
                labs_a_drogs[ld_lab].append(ld_drog)
            # Si vino prov específico, mantenemos el set para legacy `cubre_lab`.
            labs_cubiertos = set(labs_a_drogs.keys()) if not prov_id else set(
                lab for lab, drogs in labs_a_drogs.items() if prov_id in drogs
            )
            # Diccionario de drogs activas Y visibles en la matriz, con sus siglas.
            # El filtro `matriz_visible` evita mostrar siglas como "Pha" o "Via"
            # de droguerías que no aparecen como columna en /compras/labs-drogerias.
            drogs_activas = (session.query(Provider)
                             .filter(Provider.tipo == 'drogueria',
                                     Provider.activo.is_(True),
                                     Provider.matriz_visible.is_(True))
                             .order_by(Provider.razon_social).all())
            drog_label = {d.id: _sigla_drog(d.razon_social) for d in drogs_activas}
            drog_nombre_full = {d.id: d.razon_social for d in drogs_activas}
            # Ponderación por horas hasta el próximo cierre de cada droguería.
            # target_unid = ceil(daily_rate * factor_h). Piso 0.25 (cubre min 6h)
            # para evitar pedir 0u con cierre inminente. Sin techo para que un
            # cierre el lunes desde el viernes pida ~3 días (factor 3.0).
            # Sin matriz de horarios → factor 1.0 (cobertura 1 día por default).
            factor_h_por_drog = {}
            horas_prox_por_drog = {}
            for _d in drogs_activas:
                _cierre = proximo_cierre(session, _d.id)
                if _cierre:
                    _h = _cierre['falta_segundos'] / 3600
                    factor_h_por_drog[_d.id] = max(0.25, _h / 24)
                    horas_prox_por_drog[_d.id] = round(_h, 1)
                else:
                    factor_h_por_drog[_d.id] = 1.0
                    horas_prox_por_drog[_d.id] = None
            # Map lab observer → lab local id (por observer_id si está linkeado).
            lab_obs_to_local = dict(
                session.query(Laboratorio.observer_id, Laboratorio.id)
                .filter(Laboratorio.observer_id.isnot(None)).all()
            )
            # Fallback por nombre normalizado: obs_lab_id → local lab id.
            # Permite usar la matriz aunque los labs no tengan observer_id.
            from database import ObsLaboratorio
            from helpers import _normalizar_nombre_entidad as _norm_lab
            obs_lab_norm = {
                r[0]: _norm_lab(r[1])
                for r in session.query(ObsLaboratorio.observer_id,
                                       ObsLaboratorio.descripcion).all()
            }
            local_lab_por_norm = {
                _norm_lab(l.nombre): l.id
                for l in session.query(Laboratorio).all()
            }

            # EANs desde obs_codigos_barras (orden 1 = principal)
            from database import ObsCodigoBarras, OfertaMinimo
            all_pids = [r.pid for r in base]
            eans_armar = {}  # observer_id → ean principal
            if all_pids:
                for ecb in (session.query(ObsCodigoBarras.producto_observer,
                                          ObsCodigoBarras.codigo_barras)
                            .filter(ObsCodigoBarras.producto_observer.in_(all_pids),
                                    ObsCodigoBarras.fecha_baja.is_(None),
                                    ObsCodigoBarras.orden == 1)
                            .all()):
                    eans_armar[ecb.producto_observer] = ecb.codigo_barras

            # Ofertas activas por EAN (TRF): mayor descuento si hay varias.
            all_eans_set = {v for v in eans_armar.values() if v}
            ofertas_por_ean = {}
            if all_eans_set:
                for of in (session.query(OfertaMinimo)
                           .filter(OfertaMinimo.ean.in_(all_eans_set),
                                   OfertaMinimo.activo.is_(True))
                           .all()):
                    um  = int(of.unidades_minima or 1)
                    dto = float(of.descuento_psl or 0)
                    prev = ofertas_por_ean.get(of.ean)
                    if not prev or dto > prev['oferta_dto']:
                        ofertas_por_ean[of.ean] = {'oferta_dto': dto, 'oferta_min': um}

            # Umbrales de rotación (Settings → Config singleton) para clasificar
            # A/M/B. Se usan para la regla "caro + rotación baja" del motor.
            _rcfg = session.get(database.Config, 1)
            _rot_alta = float(getattr(_rcfg, 'rot_alta_min', 20.0) or 20.0) if _rcfg else 20.0
            _rot_media = float(getattr(_rcfg, 'rot_media_min', 5.0) or 5.0) if _rcfg else 5.0

            items = []
            for r in base:
                # Filtro rubro: aplica el set elegido vía ?rubros=… (default 12 Medicamentos).
                sub_id = sub_de_prod.get(r.pid)
                rub_id = subrubro_a_rubro.get(sub_id)
                if rubros_filtro is not None and rub_id not in rubros_filtro:
                    continue
                local = local_por_obs.get(r.pid)
                # 'excluido' (sacar temporal) → no entra al armado.
                # 'no_pedir' (permanente) → SÍ entra, pero con badge + botón
                # reactivar y a_pedir=0 por default.
                if local and local['excluido']:
                    continue
                lab_local_id = (local['lab_local_id'] if local else None) \
                                or lab_obs_to_local.get(r.lab_obs_id) \
                                or local_lab_por_norm.get(obs_lab_norm.get(r.lab_obs_id, ''))
                cubre_lab = lab_local_id in labs_cubiertos
                u12m_int = int(r.u12m or 0)
                u24h_val = int(v24h_rows.get(r.pid, 0) or 0)
                u7d_val  = int(v7d_rows.get(r.pid, 0) or 0)
                min_actual = int(r.minimo or 0)
                stock_actual = int(r.stock or 0)
                ventas_arr = ventas_por_pid.get(r.pid, [0]*12)
                montos_arr = montos_por_pid.get(r.pid, [0.0]*12)
                # PVP del ÚLTIMO MES con ventas (monto/unidades de ese mes).
                # Más fiel que el promedio 12m bajo inflación. 0 si nunca vendió.
                pvp_est = pvp_reciente(ventas_arr, montos_arr)

                # Forecast a 7 días + promedio mensual con prorrateo.
                min_sugerido, avg_m, sin_mov, tipo = calcular_min_sugerido(
                    ventas_arr, stock_actual, start_month, end_month,
                )
                avg_diario = avg_m / AVG_DAYS_PER_MONTH if avg_m else 0
                rotacion_cls = rotation_index(avg_m, _rot_alta, _rot_media)
                cobertura_d = (round(min_actual / avg_diario)
                               if avg_diario > 0 and min_actual > 0 else None)
                # Sugerencia up/down/ok comparando contra el forecast 7d.
                if u12m_int == 0 or sin_mov:
                    min_sugerencia = None
                else:
                    min_sugerencia = clasificar_min(min_actual, min_sugerido)

                # u_rot = unidades vendidas en los últimos `meses_rotacion`
                # meses COMPLETOS (excluye el mes actual parcial). Es la base
                # de la tasa de rotación diaria: capta mejor demanda reciente
                # que el u12m anualizado (estacionalidad y cambios de hábito).
                # ventas_arr es array 0..11 con [11] = mes actual parcial.
                _i_end = 11  # exclusive (el actual parcial no entra)
                _i_start = max(0, _i_end - meses_rotacion)
                u_rot = sum(ventas_arr[_i_start:_i_end])

                # MÍN EFECTIVO: si Observer tiene el mínimo desactualizado y
                # nuestro forecast (min_sugerido) difiere, usamos NUESTRO valor
                # — no arrastramos el error. Se marca como `min_corregido` para
                # que la UI lo muestre claramente.
                if min_sugerencia in ('up', 'down') and min_sugerido > 0:
                    min_efectivo = min_sugerido
                    min_corregido = True
                else:
                    min_efectivo = min_actual
                    min_corregido = False

                # Multi-drog: prov_ids que cubren este lab. drog_principal =
                # el filtro `prov_id` (modo single drog) o el primero de la lista.
                # Se usa acá para ponderar el target_unid, y abajo se persiste al dict.
                drogs_que_cubren = list(labs_a_drogs.get(lab_local_id, [])) if lab_local_id else []
                drog_principal = (prov_id if prov_id in drogs_que_cubren
                                  else (drogs_que_cubren[0] if drogs_que_cubren else None))
                factor_h = factor_h_por_drog.get(drog_principal, 1.0)
                horas_prox_item = horas_prox_por_drog.get(drog_principal)

                # Override por producto: si el operador seteó cantidad_reposicion_fija
                # en /productos, ese valor manda cuando el stock baja del mínimo —
                # ignora ponderación por horas y forecast.
                cant_fija = (local or {}).get('cantidad_reposicion_fija')

                # Motor unificado de cálculo de cantidad. Selecciona tipo de pedido
                # según contexto (COMPRA_LAB si viene lab_id, sino REPOSICION).
                # La config completa (piso/target/buffer/override) vive en la tabla
                # tipo_pedido_config — ver services/calculo_pedido.py.
                daily_rate = (u_rot / dias_rotacion) if dias_rotacion else 0
                target_unid = math.ceil(daily_rate * factor_h)
                from services.calculo_pedido import calcular_a_pedir, cargar_config
                _tipo_slug = 'COMPRA_LAB' if lab_id else 'REPOSICION'
                _cfg = cargar_config(_tipo_slug) or {}
                _ctx_base = {
                    'daily_rate': daily_rate,
                    'min_efectivo': min_efectivo,
                    'factor_h': factor_h,
                    'cubrir_dias': cubrir_dias,
                    'stock_actual': stock_actual,
                    'cantidad_reposicion_fija': cant_fija,
                    'u12m': u12m_int,
                    'sin_mov': sin_mov,
                    'pvp': pvp_est,
                    'rotacion': rotacion_cls,
                    'ventas_mensuales': ventas_arr,
                }
                _result = calcular_a_pedir(_cfg, _ctx_base)
                a_pedir = _result['a_pedir']
                # Si cant_fija aplicó override, calcular el "sin override" para
                # mostrar tachado en la UI ("habría pedido X, pero override → Y").
                a_pedir_sin_override = None
                if _result.get('override_aplicado') and cant_fija:
                    _ctx_no_ov = dict(_ctx_base, cantidad_reposicion_fija=None)
                    a_pedir_sin_override = calcular_a_pedir(_cfg, _ctx_no_ov).get('a_pedir')

                # Urgente = bajo o igual al mínimo. No urgente = entró sólo por
                # cobertura insuficiente (stock arriba del mín pero rota rápido).
                urgente = stock_actual <= min_actual
                no_pedir_flag = bool(local and local.get('no_pedir'))
                if no_pedir_flag:
                    # Marcado "no pedir" — entra al listado pero default 0.
                    a_pedir = 0
                # Si la sugerencia es 0 (sin rotación / sin mov 60d) y NO
                # está marcado no_pedir, lo escondemos para no llenar la
                # lista de ruido. Los no_pedir SÍ entran (con badge + botón
                # ↻ Reactivar) para que se puedan rehabilitar desde acá.
                if a_pedir <= 0 and not no_pedir_flag:
                    continue
                drogs_siglas = [drog_label.get(d, '?') for d in drogs_que_cubren if d in drog_label]

                items.append({
                    'no_pedir': no_pedir_flag,
                    'pid': r.pid,
                    'producto_id_local': local['id'] if local else None,
                    # Presentación: para mostrar equivalencia en cajas (sin tocar
                    # la cantidad en unidades). Solo si fraccionado + envase>1.
                    'fraccionado': bool(local and local.get('fraccionado')),
                    'cantidad_envase': (local.get('cantidad_envase') if local else None),
                    'desc': r.desc,
                    'droga_nombre': r.droga_nombre or '',
                    'lab_nombre': r.lab_nombre or '—',
                    'lab_local_id': lab_local_id,
                    'urgente': urgente,
                    'tvc': r.tvc,
                    'tipo': tipo,  # 'C' crónico, 'N' normal
                    'stock': stock_actual,
                    'minimo': min_actual,
                    'min_efectivo': min_efectivo,        # el que usamos para a_pedir
                    'min_corregido': min_corregido,      # True si reemplazamos el de Observer
                    'min_sugerido': min_sugerido,
                    'min_sugerencia': min_sugerencia,
                    'cobertura_d': cobertura_d,
                    'u24h': u24h_val,
                    'u7d': u7d_val,
                    'u12m': u12m_int,
                    'pvp': round(pvp_est, 2),
                    'rotacion': rotacion_cls,            # 'A'|'M'|'B' (alta/media/baja)
                    'ventas_arr': ventas_arr,            # 12 meses, [11]=mes actual parcial
                    'sin_mov_60d': bool(sin_mov),
                    'a_pedir': a_pedir,
                    'factor_h': round(factor_h, 2),          # multiplicador aplicado
                    'horas_prox': horas_prox_item,           # horas hasta próximo cierre del drog principal
                    'target_unid': int(target_unid),         # cantidad para cubrir hasta el próximo cierre
                    'daily_rate': round(daily_rate, 2),      # tasa diaria de venta (u_rot / dias_rotacion)
                    'cant_reposicion_fija': int(cant_fija) if cant_fija else None,  # override por producto si seteado
                    'override_aplicado': bool(_result.get('override_aplicado')),
                    'a_pedir_sin_override': int(a_pedir_sin_override) if a_pedir_sin_override is not None else None,
                    'avg_diario': round(avg_diario, 3),
                    'cubre_lab': cubre_lab,
                    'drogs_ids': drogs_que_cubren,    # [prov_id, ...]
                    'drogs_siglas': drogs_siglas,     # ['Kel', '20J']
                    'drog_principal': drog_principal,
                    'ean': eans_armar.get(r.pid, ''),
                    **(ofertas_por_ean.get(eans_armar.get(r.pid, ''), {'oferta_dto': None, 'oferta_min': None})),
                })
            # Orden: alfabético por descripción (modo multi-drog).
            # Si vino prov específico, mantener orden cubre primero.
            if prov_id:
                items.sort(key=lambda x: (not x['cubre_lab'], x['desc'].lower()))
            else:
                items.sort(key=lambda x: x['desc'].lower())

            # Counter de pendientes (NO_VINO de pedidos previos) por producto.
            # Sirve para la columna "Pendientes" del armado.
            obs_ids_items = [it['pid'] for it in items if it.get('pid')]
            pendientes_por_obs = {}
            if obs_ids_items:
                from sqlalchemy import func as _func2
                rows_pend = (session.query(
                                PedidoEmitidoItem.observer_id,
                                _func2.count(PedidoEmitidoItem.id),
                            ).filter(
                                PedidoEmitidoItem.estado == 'NO_VINO',
                                PedidoEmitidoItem.observer_id.in_(obs_ids_items),
                            ).group_by(PedidoEmitidoItem.observer_id).all())
                pendientes_por_obs = {oid: cnt for oid, cnt in rows_pend}
            for it in items:
                it['pendientes_count'] = pendientes_por_obs.get(it['pid'], 0)

            # ── Flags (Comportamientos excepcionales) por EAN ──
            # Cualquier EAN del producto (principal o alt) cuenta. El display del
            # chip lo arma services.flags (dedup); acá aplicamos el efecto sobre la
            # cantidad:
            #   tope_uno       → a_pedir máx 1.
            #   agotar_todo    → bloquear reposición mientras quede stock (>0).
            #   agotar_hasta_1 → bloquear reposición mientras stock > 1.
            # Los "agotar" se apoyan SOLO en el stock real de ObServer (se liberan
            # solos al bajar al umbral; no dependen de marcar recepción).
            if obs_ids_items:
                from services.flags import flags_display_por_producto
                eans_completos_por_pid = {}
                for ecb in (session.query(ObsCodigoBarras.producto_observer,
                                          ObsCodigoBarras.codigo_barras)
                            .filter(ObsCodigoBarras.producto_observer.in_(obs_ids_items))
                            .filter(ObsCodigoBarras.fecha_baja.is_(None))
                            .all()):
                    eans_completos_por_pid.setdefault(ecb.producto_observer, []).append(ecb.codigo_barras)
                flag_por_pid = flags_display_por_producto(session, eans_completos_por_pid)
                for it in items:
                    it['flag'] = flag_por_pid.get(it['pid'])
                    _ef = it['flag'].get('efecto_armado') if it['flag'] else None
                    if _ef == 'tope_uno' and it.get('a_pedir', 0) > 1:
                        it['a_pedir'] = 1
                        it['tope_uno'] = True
                    elif _ef == 'agotar_todo':
                        # Nunca repone (discontinuar / agotar a 0).
                        it['a_pedir'] = 0
                        it['agotar'] = True
                    elif _ef == 'agotar_hasta_1':
                        # No repone mientras tenga stock; repone 1 al llegar a 0
                        # (mantiene 1 unidad). Usa el stock real de ObServer.
                        it['a_pedir'] = 1 if (it.get('stock') or 0) == 0 else 0
                        it['agotar'] = True
            else:
                for it in items:
                    it['flag'] = None

            # Los "agotar" con a_pedir 0 ni se traen al armado (no se van a pedir):
            # "Agotar stock" desaparece siempre; "Agotar hasta 1" solo aparece
            # cuando llega a 0 y repone 1.
            items = [it for it in items if not (it.get('agotar') and not it.get('a_pedir'))]

            cierre = proximo_cierre(session, prov_id) if prov_id else None

            # Pendientes de pedidos anteriores a esta drog: pedida > recibida.
            # Incluye estado=NO_VINO y RECIBIDO_PARCIAL (recibida < pedida).
            pendientes_anteriores = []
            ped_rows = (session.query(PedidoEmitidoItem, PedidoEmitido)
                        .join(PedidoEmitido,
                              PedidoEmitido.id == PedidoEmitidoItem.pedido_id)
                        .filter(PedidoEmitido.drogueria_id == prov_id,
                                PedidoEmitidoItem.estado.in_(('NO_VINO', 'RECIBIDO')),
                                PedidoEmitidoItem.cantidad_recibida < PedidoEmitidoItem.cantidad_pedida)
                        .order_by(PedidoEmitido.fecha.desc())
                        .all())
            for it, ped in ped_rows:
                pendiente = max(0, (it.cantidad_pedida or 0) - (it.cantidad_recibida or 0))
                if pendiente <= 0:
                    continue
                lab_id_pend = (lab_obs_to_local.get(it.observer_id)
                               or local_lab_por_norm.get(_norm_lab(it.lab_nombre or '')))
                pendientes_anteriores.append({
                    'item_id': it.id,
                    'pedido_id': ped.id,
                    'pedido_fecha': ped.fecha,
                    'pid': it.observer_id,
                    'producto_id_local': it.producto_id_local,
                    'desc': it.descripcion,
                    'lab_nombre': it.lab_nombre or '—',
                    'pendiente': pendiente,
                    'cubre_lab': lab_id_pend in labs_cubiertos if lab_id_pend else False,
                })

            # Pedidos guardados con mostrar_hasta vigente → mapa
            # observer_id → [{cantidad, fecha, pedido_id, pedido_lab, mostrar_hasta}].
            # Se attacha a cada item del armado para mostrar inline en la fila.
            from collections import defaultdict

            from database import Pedido
            ped_guardado_por_obs = defaultdict(list)
            hoy_d = _date.today()
            ped_q = (session.query(Pedido)
                     .filter(Pedido.mostrar_hasta.isnot(None),
                             Pedido.mostrar_hasta >= hoy_d)
                     .order_by(Pedido.creado_en.desc()).all())
            if ped_q:
                todos_cb = set()
                for ped in ped_q:
                    for pi in ped.items:
                        if pi.codigo_barra:
                            todos_cb.add(pi.codigo_barra.strip())
                cb_to_obs = {}
                if todos_cb:
                    for prod in (session.query(Producto)
                                 .filter(Producto.codigo_barra.in_(todos_cb)).all()):
                        if prod.observer_id:
                            cb_to_obs[prod.codigo_barra] = prod.observer_id
                    faltantes = todos_cb - set(cb_to_obs.keys())
                    if faltantes:
                        from database import ProductoCodigoBarra
                        rows = (session.query(ProductoCodigoBarra.codigo_barra,
                                              Producto.observer_id)
                                .join(Producto,
                                      Producto.id == ProductoCodigoBarra.producto_id)
                                .filter(ProductoCodigoBarra.codigo_barra.in_(faltantes),
                                        Producto.observer_id.isnot(None)).all())
                        for cb, oid in rows:
                            cb_to_obs.setdefault(cb, oid)
                for ped in ped_q:
                    info_pedido = {
                        'pedido_id': ped.id,
                        'pedido_lab': ped.laboratorio,
                        'fecha': ped.creado_en.strftime('%d/%m') if ped.creado_en else '',
                        'mostrar_hasta': ped.mostrar_hasta.strftime('%d/%m/%y'),
                    }
                    for pi in ped.items:
                        oid = cb_to_obs.get((pi.codigo_barra or '').strip())
                        if not oid:
                            continue
                        ped_guardado_por_obs[oid].append({
                            **info_pedido, 'cantidad': pi.cantidad,
                        })
            # Attach a cada item del armado.
            for it in items:
                it['pedidos_guardados'] = ped_guardado_por_obs.get(it['pid'], [])

            # Lista de rubros para el dropdown (orden por descripción).
            from database import ObsRubro
            rubros_disponibles = [
                {'id': r.observer_id, 'nombre': r.descripcion}
                for r in (session.query(ObsRubro)
                          .order_by(ObsRubro.descripcion).all())
            ]
            rubros_seleccionados = (sorted(rubros_filtro)
                                     if rubros_filtro is not None else None)

            # Último sync exitoso de stock (para mostrar en el encabezado).
            from database import ObsSyncLog, now_ar
            last_sync = (session.query(ObsSyncLog)
                         .filter(ObsSyncLog.entidad == 'stock',
                                 ObsSyncLog.error.is_(None))
                         .order_by(ObsSyncLog.ejecutado_en.desc()).first())
            last_sync_stock = last_sync.ejecutado_en if last_sync else None
            last_sync_min = None
            if last_sync_stock:
                last_sync_min = int((now_ar() - last_sync_stock).total_seconds() // 60)

            # Ofertas del lab pre-seleccionado: OfertaMinimo activas+vigentes
            # del lab_local. Aprovechamos que ya tenemos el lab para mostrar
            # un resumen en el header con: total, por drog/directo, vigencia,
            # mejor descuento.
            lab_ofertas = []
            usa_packs = False
            modulos_lab = []
            if lab_id and lab_local:
                from datetime import date as _date_lo
                hoy_lo = _date_lo.today()
                rows_lo = (session.query(_OM_drog)
                           .filter(_OM_drog.laboratorio_id == lab_local.id,
                                   _OM_drog.activo.is_(True),
                                   or_(_OM_drog.vigencia_hasta.is_(None),
                                       _OM_drog.vigencia_hasta >= hoy_lo))
                           .all())
                # Map drogueria_id → razon_social para etiquetar.
                drog_ids_of = {o.drogueria_id for o in rows_lo if o.drogueria_id}
                drog_nombres_of = {}
                if drog_ids_of:
                    for p_of in (session.query(Provider)
                                 .filter(Provider.id.in_(drog_ids_of)).all()):
                        drog_nombres_of[p_of.id] = p_of.razon_social
                for o in rows_lo:
                    lab_ofertas.append({
                        'id': o.id,
                        'ean': o.ean,
                        'descripcion': o.descripcion or '',
                        'descuento_psl': float(o.descuento_psl or 0),
                        'unidades_minima': int(o.unidades_minima) if o.unidades_minima else None,
                        'tipo_descuento': o.tipo_descuento or 'simple',
                        'drogueria_id': o.drogueria_id,
                        'drogueria_nombre': drog_nombres_of.get(o.drogueria_id) if o.drogueria_id else None,
                        'vigencia_hasta': o.vigencia_hasta,
                        'observacion': (o.observacion or '').strip()[:120] if o.observacion else None,
                    })
                # Orden: descuento DESC, después por droguería.
                lab_ofertas.sort(key=lambda x: (-x['descuento_psl'],
                                                x['drogueria_nombre'] or ''))

            # Módulos activos del lab — SOLO si el lab está marcado usa_packs.
            # Mismo dataset que /modulo-packs/activos pero filtrado al lab, para
            # el paso de módulos single-page en el armar. El cálculo (floor/min)
            # se hace en el front (ver order_detail.html → procesarModulos).
            usa_packs = bool(lab_local and getattr(lab_local, 'usa_packs', False))
            modulos_lab = []
            if usa_packs:
                from database import Modulo
                mods = (session.query(Modulo)
                        .filter(Modulo.laboratorio_id == lab_local.id,
                                Modulo.activo.is_(True))
                        .order_by(Modulo.lista_nombre, Modulo.nombre).all())
                unidad_eans = {mp.ean_unidad for m in mods for mp in m.packs}
                prod_desc = {}
                if unidad_eans:
                    for cb, desc in (session.query(Producto.codigo_barra,
                                                   Producto.descripcion)
                                     .filter(Producto.codigo_barra.in_(
                                         list(unidad_eans)[:1000])).all()):
                        prod_desc[cb] = desc or ''
                for m in mods:
                    # Saltar la fila "header" de la lista (nombre == lista_nombre).
                    if m.lista_nombre and m.nombre == m.lista_nombre:
                        continue
                    packs = [{'ean_pack': mp.ean_pack,
                              'ean_unidad': mp.ean_unidad,
                              'cantidad': mp.cantidad or 1,
                              'cant_modulo': (mp.cant_modulo if mp.cant_modulo is not None
                                              else (mp.cantidad or 1)),
                              'desc_pct': float(mp.desc_pct) if mp.desc_pct is not None else 0.0,
                              'desc_unidad': prod_desc.get(mp.ean_unidad, mp.descripcion or '')}
                             for mp in m.packs]
                    if packs:
                        modulos_lab.append({'id': m.id, 'nombre': m.nombre,
                                            'lista_nombre': m.lista_nombre or m.nombre,
                                            'packs': packs})

            # Modo lab: droguerías activas para el selector de "Canal de compra"
            # (Directo al lab / Vía droguería). Sugerimos la más usada con ese
            # lab según historial de pedidos (igual criterio que el viejo flujo).
            droguerias_canal = []
            if lab_id:
                provs_drog = (session.query(Provider)
                              .filter(Provider.tipo == 'drogueria',
                                      Provider.activo.is_(True))
                              .order_by(Provider.razon_social).all())
                droguerias_canal = [{'id': p.id, 'nombre': p.razon_social}
                                    for p in provs_drog]

            # Valor piso ("productos caros $") desde la config del pedido, para
            # pre-cargar el filtro de PVP en el header (editable por el operador).
            from services.calculo_pedido import cargar_config as _cargar_cfg_vp
            _cfg_vp = _cargar_cfg_vp('COMPRA_LAB' if lab_id else 'REPOSICION') or {}
            valor_piso = float(_cfg_vp.get('valor_piso') or 0)

            return render_template('compras_dia_armar.html',
                                   prov=prov, items=items,
                                   total_items=len(items),
                                   valor_piso=valor_piso,
                                   cubre=sum(1 for i in items if i['cubre_lab']),
                                   pendientes=sum(1 for i in items if not i['cubre_lab']),
                                   cierre=cierre,
                                   target_dias=target_dias,
                                   cubrir_dias=cubrir_dias,
                                   drog_label=drog_label,
                                   drog_nombre_full=drog_nombre_full,
                                   pendientes_anteriores=pendientes_anteriores,
                                   rubros_disponibles=rubros_disponibles,
                                   rubros_seleccionados=rubros_seleccionados,
                                   last_sync_stock=last_sync_stock,
                                   last_sync_min=last_sync_min,
                                   oferta_drog_n=len(oferta_pids) if oferta_pids else 0,
                                   oferta_drog_nombre=oferta_nombre_drog,
                                   oferta_disponible_n=oferta_disponible_n,
                                   oferta_vigencia_hasta=oferta_vigencia_hasta,
                                   oferta_observacion=oferta_observacion,
                                   usar_oferta=usar_oferta,
                                   lab_id=lab_id,
                                   lab_nombre=lab_nombre,
                                   libres_a=libres_a,
                                   droguerias_canal=droguerias_canal,
                                   usa_packs=usa_packs,
                                   modulos_lab=modulos_lab,
                                   lab_ofertas=lab_ofertas if lab_id else [])

    @app.route('/compras/multi-lab')
    @login_required
    def compras_multi_lab():
        """Listado de droguerías que tienen oferta vigente multi-lab cargada.

        Cada fila lleva al armado de pedido filtrado por esa oferta
        (`/compras/armar?prov=<id>&usar_oferta=1`). Útil para el flujo
        Ciafarma: una oferta con productos de varios labs, se elige la drog
        y se arma el pedido directo de los productos de la oferta.
        """
        from datetime import date as _date

        from sqlalchemy import func as _func

        from database import OfertaMinimo, Provider
        hoy = _date.today()
        with get_db() as session:
            # Drogs con al menos 1 OfertaMinimo vigente (vigencia_hasta >= hoy o NULL).
            sub = (session.query(OfertaMinimo.drogueria_id,
                                  _func.count(OfertaMinimo.id).label('n'),
                                  _func.min(OfertaMinimo.vigencia_desde).label('vd'),
                                  _func.max(OfertaMinimo.vigencia_hasta).label('vh'))
                   .filter(OfertaMinimo.drogueria_id.isnot(None),
                           or_(OfertaMinimo.vigencia_hasta.is_(None),
                               OfertaMinimo.vigencia_hasta >= hoy))
                   .group_by(OfertaMinimo.drogueria_id)
                   .all())
            drog_ids = [r[0] for r in sub if r[0]]
            prov_map = {}
            if drog_ids:
                for p in (session.query(Provider)
                          .filter(Provider.id.in_(drog_ids)).all()):
                    prov_map[p.id] = p.razon_social or f'#{p.id}'
            drogs = sorted(
                [{
                    'id':       r[0],
                    'nombre':   prov_map.get(r[0], f'#{r[0]}'),
                    'n_items':  int(r[1]),
                    'vigencia_desde': r[2].strftime('%d/%m/%Y') if r[2] else None,
                    'vigencia_hasta': r[3].strftime('%d/%m/%Y') if r[3] else None,
                } for r in sub if r[0]],
                key=lambda x: -x['n_items'],
            )
        return render_template('compras_multi_lab.html', drogs=drogs)

    @app.route('/compras/armar/exportar-minimos')
    @login_required
    def compras_armar_exportar_minimos():
        """Exporta XLSX con productos que tienen sugerencia de cambio de mínimo
        (subir o bajar). Para enviar a la gente del POS Observer y que actualicen
        manualmente, ya que Observer no recalcula los mínimos automático.

        Query params:
          tipo=up|down|both (default: both)
          rubros=12 (default Medicamentos, igual que /compras/armar)
        """
        from io import BytesIO

        from flask import send_file
        from sqlalchemy import func

        from database import (
            ObsLaboratorio,
            ObsProducto,
            ObsRubro,
            ObsStock,
            ObsSubrubro,
            ObsVentaMensual,
        )

        tipo = (request.args.get('tipo') or 'both').lower()
        if tipo not in ('up', 'down', 'both'):
            tipo = 'both'
        rubros_raw = (request.args.get('rubros') or '12').strip()
        if rubros_raw.lower() == 'all' or not rubros_raw:
            rubros_filtro = None
        else:
            try:
                rubros_filtro = set(int(x) for x in rubros_raw.split(',') if x.strip())
            except ValueError:
                rubros_filtro = {12}

        with get_db() as session:
            # Stock y mín por producto
            stock_q = (session.query(
                ObsStock.producto_observer.label('pid'),
                func.sum(ObsStock.stock_actual).label('stock'),
                func.sum(ObsStock.minimo).label('minimo'),
            ).filter(ObsStock.minimo.isnot(None), ObsStock.minimo > 0)
              .group_by(ObsStock.producto_observer).subquery())
            # u12m total
            v12_q = (session.query(
                ObsVentaMensual.producto_observer.label('pid'),
                func.sum(ObsVentaMensual.unidades).label('u12m'),
            ).group_by(ObsVentaMensual.producto_observer).subquery())

            base = (session.query(
                ObsProducto.observer_id.label('pid'),
                ObsProducto.descripcion.label('desc'),
                ObsProducto.subrubro_observer,
                ObsLaboratorio.descripcion.label('lab_nombre'),
                stock_q.c.stock,
                stock_q.c.minimo,
                func.coalesce(v12_q.c.u12m, 0).label('u12m'),
            )
            .join(stock_q, stock_q.c.pid == ObsProducto.observer_id)
            .outerjoin(ObsLaboratorio,
                       ObsLaboratorio.observer_id == ObsProducto.laboratorio_observer)
            .outerjoin(v12_q, v12_q.c.pid == ObsProducto.observer_id)
            .filter(ObsProducto.fecha_baja.is_(None))
            .filter(ObsProducto.subrubro_observer.isnot(None))
            ).all()

            # Rubro lookup
            subrubro_a_rubro = dict(
                session.query(ObsSubrubro.observer_id, ObsSubrubro.rubro_observer).all())
            rubro_nombres = {r.observer_id: r.descripcion
                             for r in session.query(ObsRubro).all()}

            # EANs (orden 1)
            from database import ObsCodigoBarras
            pids = [r.pid for r in base]
            eans = {}
            if pids:
                for cb in (session.query(ObsCodigoBarras.producto_observer,
                                          ObsCodigoBarras.codigo_barras)
                           .filter(ObsCodigoBarras.producto_observer.in_(pids),
                                   ObsCodigoBarras.fecha_baja.is_(None),
                                   ObsCodigoBarras.orden == 1).all()):
                    eans[cb.producto_observer] = cb.codigo_barras

            # Ventas mes-a-mes para purchase_engine
            hoy_d = _date.today()
            end_month = hoy_d.month
            start_month = ((end_month - 11 - 1) % 12) + 1
            start_year = hoy_d.year if start_month <= end_month else hoy_d.year - 1
            ventas_por_pid = {pid: [0]*12 for pid in pids}
            if pids:
                rows_vm = (session.query(ObsVentaMensual.producto_observer,
                                          ObsVentaMensual.anio,
                                          ObsVentaMensual.mes,
                                          func.sum(ObsVentaMensual.unidades))
                           .filter(ObsVentaMensual.producto_observer.in_(pids))
                           .group_by(ObsVentaMensual.producto_observer,
                                     ObsVentaMensual.anio, ObsVentaMensual.mes)
                           .all())
                for pid_v, anio, mes, uds in rows_vm:
                    offset = (anio - start_year) * 12 + (mes - start_month)
                    if 0 <= offset <= 11 and pid_v in ventas_por_pid:
                        ventas_por_pid[pid_v][offset] += int(uds or 0)

            filas = []
            for r in base:
                # Filtro rubro
                rub_id = subrubro_a_rubro.get(r.subrubro_observer)
                if rubros_filtro is not None and rub_id not in rubros_filtro:
                    continue
                u12m = int(r.u12m or 0)
                if u12m == 0:
                    continue
                ventas_arr = ventas_por_pid.get(r.pid, [0]*12)
                min_sugerido, _avg_m, sin_mov, tipo_p = calcular_min_sugerido(
                    ventas_arr, int(r.stock or 0), start_month, end_month,
                )
                if sin_mov:
                    continue
                min_actual = int(r.minimo or 0)
                sug = clasificar_min(min_actual, min_sugerido)
                if sug == 'ok':
                    continue
                if tipo == 'up' and sug != 'up':
                    continue
                if tipo == 'down' and sug != 'down':
                    continue
                filas.append({
                    'ean': eans.get(r.pid, ''),
                    'descripcion': r.desc,
                    'lab': r.lab_nombre or '—',
                    'rubro': rubro_nombres.get(rub_id, ''),
                    'sugerencia': 'SUBIR' if sug == 'up' else 'BAJAR',
                    'min_actual': min_actual,
                    'min_sugerido': min_sugerido,
                    'diferencia': min_sugerido - min_actual,
                    'stock_actual': int(r.stock or 0),
                    'u12m': u12m,
                    'tipo': tipo_p,
                })

            # Orden: SUBIR primero (más urgente), después BAJAR. Dentro: por
            # diferencia absoluta descendente.
            filas.sort(key=lambda x: (
                0 if x['sugerencia'] == 'SUBIR' else 1,
                -abs(x['diferencia']),
            ))

            # Excel
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Sugerencias'
            headers = ['EAN', 'Producto', 'Laboratorio', 'Rubro', 'Sugerencia',
                       'Mín actual', 'Mín sugerido', 'Diferencia',
                       'Stock actual', 'Ventas 12m', 'Tipo']
            ws.append(headers)
            # Header style
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1c1c1e', end_color='1c1c1e',
                                         fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            # Filas
            for f in filas:
                ws.append([f['ean'], f['descripcion'], f['lab'], f['rubro'],
                            f['sugerencia'], f['min_actual'], f['min_sugerido'],
                            f['diferencia'], f['stock_actual'], f['u12m'], f['tipo']])
            # Color rojo claro las SUBIR, azul claro las BAJAR
            from openpyxl.styles import PatternFill as PF
            row_red = PF(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
            row_blu = PF(start_color='dbeafe', end_color='dbeafe', fill_type='solid')
            for i, f in enumerate(filas, start=2):
                fill = row_red if f['sugerencia'] == 'SUBIR' else row_blu
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = fill
            # Anchos
            anchos = [16, 50, 25, 18, 12, 12, 14, 12, 14, 12, 8]
            for i, w in enumerate(anchos, start=1):
                ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w
            ws.freeze_panes = 'A2'

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            fname = f'Sugerencias_minimos_{hoy_d.strftime("%Y%m%d")}.xlsx'
            return send_file(buf, as_attachment=True, download_name=fname,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/api/pedidos/dia/buscar-producto')
    @login_required
    def api_compras_dia_buscar_producto():
        """Busca producto ObServer por descripción tokenizada (AND).
        Devuelve datos listos para agregar como fila al armado: stock, mínimo,
        u12m, lab, cubre_lab según la droguería pasada por ?prov.
        """
        from sqlalchemy import and_, func

        from database import (
            Laboratorio,
            LaboratorioDrogueria,
            ObsCodigoBarras,
            ObsLaboratorio,
            ObsProducto,
            ObsStock,
            ObsVentaDetalle,
            Producto,
        )
        q = (request.args.get('q') or '').strip()
        prov_id = request.args.get('prov', type=int)
        # En multi-drog (prov_id None/null/empty) la búsqueda igual debe funcionar
        # — solo no calculamos cubre_lab para una drog específica.
        if len(q) < 2:
            return jsonify({'ok': True, 'items': []})

        # Tokenizar y armar filtros AND sobre descripcion (case-insensitive).
        tokens = [t for t in q.split() if t]
        with get_db() as session:
            query = (session.query(ObsProducto.observer_id,
                                    ObsProducto.descripcion,
                                    ObsLaboratorio.observer_id.label('lab_obs_id'),
                                    ObsLaboratorio.descripcion.label('lab_nombre'))
                     .outerjoin(ObsLaboratorio,
                                ObsLaboratorio.observer_id == ObsProducto.laboratorio_observer)
                     .filter(ObsProducto.fecha_baja.is_(None)))
            for tok in tokens:
                query = query.filter(ObsProducto.descripcion.ilike(f'%{tok}%'))
            base = query.order_by(ObsProducto.descripcion).limit(20).all()

            obs_ids = [r.observer_id for r in base]
            if not obs_ids:
                return jsonify({'ok': True, 'items': []})

            stock_rows = dict(session.query(
                ObsStock.producto_observer,
                func.sum(ObsStock.stock_actual)
            ).filter(ObsStock.producto_observer.in_(obs_ids))
             .group_by(ObsStock.producto_observer).all())
            min_rows = dict(session.query(
                ObsStock.producto_observer,
                func.sum(ObsStock.minimo)
            ).filter(ObsStock.producto_observer.in_(obs_ids),
                     ObsStock.minimo.isnot(None))
             .group_by(ObsStock.producto_observer).all())
            hoy2   = _date.today()
            _ayer2 = hoy2 - _td(days=1)
            _sem2  = hoy2 - _td(days=7)
            _det2  = session.query(
                ObsVentaDetalle.producto_observer,
                ObsVentaDetalle.fecha_estadistica,
                func.sum(ObsVentaDetalle.cantidad).label('cant'),
            ).filter(ObsVentaDetalle.producto_observer.in_(obs_ids),
                     ObsVentaDetalle.fecha_estadistica >= _sem2)\
             .group_by(ObsVentaDetalle.producto_observer,
                       ObsVentaDetalle.fecha_estadistica).all()
            v24h_rows2 = {}
            v7d_rows2  = {}
            for pid_d, fec, cant in _det2:
                v7d_rows2[pid_d] = v7d_rows2.get(pid_d, 0) + int(cant or 0)
                if fec >= _ayer2:
                    v24h_rows2[pid_d] = v24h_rows2.get(pid_d, 0) + int(cant or 0)

            local_rows = {r[0]: r for r in (
                session.query(Producto.observer_id, Producto.id, Producto.laboratorio_id)
                .filter(Producto.observer_id.in_(obs_ids)).all())}
            lab_obs_to_local = dict(session.query(
                Laboratorio.observer_id, Laboratorio.id
            ).filter(Laboratorio.observer_id.isnot(None)).all())
            from helpers import _normalizar_nombre_entidad as _norm_lab
            local_lab_por_norm = {
                _norm_lab(l.nombre): l.id
                for l in session.query(Laboratorio).all()
            }
            obs_lab_norm = {
                r[0]: _norm_lab(r[1])
                for r in session.query(ObsLaboratorio.observer_id,
                                       ObsLaboratorio.descripcion).all()
            }
            # Si vino prov específica, calculamos `cubre_lab` para esa drog;
            # si no, en modo multi-drog asumimos cubierto cuando hay matriz.
            if prov_id:
                labs_cubiertos = set(r[0] for r in session.query(
                    LaboratorioDrogueria.laboratorio_id
                ).filter(LaboratorioDrogueria.drogueria_id == prov_id).all())
            else:
                labs_cubiertos = set(r[0] for r in session.query(
                    LaboratorioDrogueria.laboratorio_id
                ).distinct().all())

            eans_buscar = {}
            if obs_ids:
                for ecb in (session.query(ObsCodigoBarras.producto_observer,
                                          ObsCodigoBarras.codigo_barras)
                            .filter(ObsCodigoBarras.producto_observer.in_(obs_ids),
                                    ObsCodigoBarras.fecha_baja.is_(None),
                                    ObsCodigoBarras.orden == 1)
                            .all()):
                    eans_buscar[ecb.producto_observer] = ecb.codigo_barras

            from database import OfertaMinimo as _OM
            eans_buscar_set = {v for v in eans_buscar.values() if v}
            ofertas_buscar = {}
            if eans_buscar_set:
                for of in (session.query(_OM)
                           .filter(_OM.ean.in_(eans_buscar_set), _OM.activo.is_(True))
                           .all()):
                    um  = int(of.unidades_minima or 1)
                    dto = float(of.descuento_psl or 0)
                    prev = ofertas_buscar.get(of.ean)
                    if not prev or dto > prev['oferta_dto']:
                        ofertas_buscar[of.ean] = {'oferta_dto': dto, 'oferta_min': um}

            # Ventas mes-a-mes para forecast (mismo pattern que /compras/armar).
            from database import ObsVentaMensual
            hoy_f = _date.today()
            end_month_f = hoy_f.month
            start_month_f = ((end_month_f - 11 - 1) % 12) + 1
            start_year_f = hoy_f.year if start_month_f <= end_month_f else hoy_f.year - 1
            ventas_por_pid_f = {pid: [0]*12 for pid in obs_ids}
            if obs_ids:
                rows_vm_f = (session.query(ObsVentaMensual.producto_observer,
                                            ObsVentaMensual.anio,
                                            ObsVentaMensual.mes,
                                            func.sum(ObsVentaMensual.unidades))
                             .filter(ObsVentaMensual.producto_observer.in_(obs_ids))
                             .group_by(ObsVentaMensual.producto_observer,
                                       ObsVentaMensual.anio, ObsVentaMensual.mes).all())
                for pid_v, anio, mes, uds in rows_vm_f:
                    offset = (anio - start_year_f) * 12 + (mes - start_month_f)
                    if 0 <= offset <= 11 and pid_v in ventas_por_pid_f:
                        ventas_por_pid_f[pid_v][offset] += int(uds or 0)
            items = []
            for r in base:
                local = local_rows.get(r.observer_id)
                lab_local_id = ((local[2] if local else None)
                                or lab_obs_to_local.get(r.lab_obs_id)
                                or local_lab_por_norm.get(obs_lab_norm.get(r.lab_obs_id, '')))
                stock = int(stock_rows.get(r.observer_id, 0) or 0)
                minimo = int(min_rows.get(r.observer_id, 0) or 0)
                ean_b = eans_buscar.get(r.observer_id, '')

                # Calcular min_sugerido + min_sugerencia + min_corregido (igual
                # que en /compras/armar). Si no hay ventas, no hay sugerencia.
                ventas_arr = ventas_por_pid_f.get(r.observer_id, [0]*12)
                u12m_int = sum(ventas_arr)
                min_sugerido, _avg_m, sin_mov, _tipo_p = calcular_min_sugerido(
                    ventas_arr, stock, start_month_f, end_month_f,
                )
                if u12m_int == 0 or sin_mov:
                    min_sugerencia = None
                else:
                    min_sugerencia = clasificar_min(minimo, min_sugerido)
                # Mín efectivo: si Observer está desfasado, usamos el sugerido.
                if min_sugerencia in ('up', 'down') and min_sugerido > 0:
                    min_efectivo = min_sugerido
                    min_corregido = True
                else:
                    min_efectivo = minimo
                    min_corregido = False
                # a_pedir simple: llegar al mín efectivo desde el stock.
                if u12m_int == 0 or sin_mov:
                    a_pedir = 0
                else:
                    a_pedir = max(0, min_efectivo - stock) if min_efectivo else 1

                items.append({
                    'pid': r.observer_id,
                    'producto_id_local': local[1] if local else None,
                    'desc': r.descripcion,
                    'lab_nombre': r.lab_nombre or '—',
                    'stock': stock,
                    'minimo': minimo,
                    'min_sugerido': min_sugerido,
                    'min_sugerencia': min_sugerencia,
                    'min_corregido': min_corregido,
                    'min_efectivo': min_efectivo,
                    'u24h': int(v24h_rows2.get(r.observer_id, 0) or 0),
                    'u7d':  int(v7d_rows2.get(r.observer_id, 0) or 0),
                    'u12m': u12m_int,
                    'a_pedir': a_pedir,
                    'cubre_lab': lab_local_id in labs_cubiertos,
                    'ean': ean_b,
                    **(ofertas_buscar.get(ean_b, {'oferta_dto': None, 'oferta_min': None})),
                })
            return jsonify({'ok': True, 'items': items})

    @app.route('/compras/labs-drogerias')
    @login_required
    def labs_drogerias_matriz():
        """Matriz lab × droguería: marca por qué drogerías va cada laboratorio.
        Persiste en LaboratorioDrogueria (tabla simple sin descuento).
        """
        from sqlalchemy import func

        from database import Laboratorio, LaboratorioDrogueria, OfertaMinimo
        with get_db() as session:
            labs_q = (session.query(Laboratorio)
                      .filter(Laboratorio.activo.is_(True))
                      .order_by(Laboratorio.nombre).all())
            labs = labs_q
            from sqlalchemy import case
            todas_drogs = (session.query(Provider)
                           .filter(Provider.tipo == 'drogueria')
                           .order_by(
                               case((Provider.matriz_orden.isnot(None), Provider.matriz_orden), else_=9999),
                               Provider.razon_social)
                           .all())
            drogs = [d for d in todas_drogs if d.activo and d.matriz_visible]
            existentes = set(
                (r.laboratorio_id, r.drogueria_id)
                for r in session.query(LaboratorioDrogueria).all()
            )
            # Ofertas importadas por lab: última fecha de actualización
            ofertas_lab = {}
            for r in (session.query(
                    OfertaMinimo.laboratorio_id,
                    func.count(OfertaMinimo.id).label('cnt'),
                    func.max(OfertaMinimo.actualizado_en).label('ultima'))
                    .filter(OfertaMinimo.activo.is_(True))
                    .group_by(OfertaMinimo.laboratorio_id)
                    .all()):
                fecha = r.ultima.date() if r.ultima else None
                ofertas_lab[r.laboratorio_id] = {
                    'count': r.cnt,
                    'fecha': fecha.strftime('%d/%m/%y') if fecha else '',
                }
            # Droguerías que tienen plantilla de pedido configurada (set-once;
            # se muestra como indicador "con/sin plantilla" bajo cada columna).
            from database import Plantilla as _Plant
            _drog_ids = [d.id for d in drogs]
            _con_plantilla = set()
            if _drog_ids:
                _con_plantilla = set(
                    r[0] for r in session.query(_Plant.entidad_id)
                    .filter(_Plant.entidad_tipo == 'drogueria',
                            _Plant.tipo_doc == 'pedido',
                            _Plant.entidad_id.in_(_drog_ids)).distinct().all()
                )
            labs_data = [{'id': l.id, 'nombre': l.nombre} for l in labs]
            drogs_data = [{'id': d.id, 'nombre': d.razon_social,
                           'tiene_plantilla': d.id in _con_plantilla} for d in drogs]
            todas_drogs_data = [{'id': d.id, 'nombre': d.razon_social,
                                  'visible': d.matriz_visible, 'orden': d.matriz_orden,
                                  'activo': d.activo} for d in todas_drogs]
            matriz = {}  # {lab_id: set(drog_ids)}
            for (lab_id, drog_id) in existentes:
                matriz.setdefault(lab_id, set()).add(drog_id)
        return render_template('labs_drogerias.html',
                               labs=labs_data, drogs=drogs_data,
                               todas_drogs=todas_drogs_data,
                               matriz={k: list(v) for k, v in matriz.items()},
                               ofertas_lab=ofertas_lab)

    @app.route('/api/matriz/drog-visible', methods=['POST'])
    @login_required
    def api_matriz_drog_visible():
        """Body: {drog_id, visible}"""
        data = request.get_json(silent=True) or {}
        with get_db() as session:
            drog = session.get(Provider, int(data.get('drog_id') or 0))
            if not drog:
                return jsonify({'ok': False}), 404
            drog.matriz_visible = bool(data.get('visible'))
            session.commit()
        return jsonify({'ok': True})

    @app.route('/api/matriz/drog-config', methods=['POST'])
    @login_required
    def api_matriz_drog_config():
        """Body: [{id, visible, orden}] — actualiza visibilidad y orden de droguerías en la matriz."""
        items = request.get_json(silent=True) or []
        with get_db() as session:
            for item in items:
                try:
                    drog_id = int(item.get('id') or 0)
                except (TypeError, ValueError):
                    continue
                d = session.get(Provider, drog_id)
                if not d:
                    continue
                d.matriz_visible = bool(item.get('visible', True))
                d.matriz_orden   = item.get('orden')
            session.commit()
        return jsonify({'ok': True})

    @app.route('/api/lab-drog/toggle', methods=['POST'])
    @login_required
    def api_lab_drog_toggle():
        """Body: {laboratorio_id, drogueria_id, activo: bool}.
        Crea/borra el match. Idempotente."""
        from database import LaboratorioDrogueria
        data = request.get_json(silent=True) or {}
        try:
            lab_id  = int(data.get('laboratorio_id') or 0)
            drog_id = int(data.get('drogueria_id') or 0)
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'IDs inválidos'}), 400
        if not lab_id or not drog_id:
            return jsonify({'ok': False, 'error': 'lab y drog requeridos'}), 400
        activo = bool(data.get('activo'))
        with get_db() as session:
            row = (session.query(LaboratorioDrogueria)
                   .filter_by(laboratorio_id=lab_id, drogueria_id=drog_id).first())
            if activo and not row:
                session.add(LaboratorioDrogueria(
                    laboratorio_id=lab_id, drogueria_id=drog_id))
                session.commit()
            elif not activo and row:
                session.delete(row)
                session.commit()
        return jsonify({'ok': True})

    @app.route('/api/lab-drog/asignar-bulk', methods=['POST'])
    @login_required
    def api_lab_drog_asignar_bulk():
        """Body: {drogueria_id, laboratorio_ids: [int, ...]}.
        Crea entradas en LaboratorioDrogueria (lab × drog). Idempotente.
        Usado para asignar en bulk los productos 'Libres' del armado a una drog."""
        from database import LaboratorioDrogueria
        data = request.get_json(silent=True) or {}
        try:
            drog_id = int(data.get('drogueria_id') or 0)
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'drogueria_id inválido'}), 400
        if not drog_id:
            return jsonify({'ok': False, 'error': 'drogueria_id requerido'}), 400
        labs_raw = data.get('laboratorio_ids') or []
        try:
            lab_ids = [int(x) for x in labs_raw if x]
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'laboratorio_ids inválidos'}), 400
        if not lab_ids:
            return jsonify({'ok': False, 'error': 'Sin laboratorios para asignar'}), 400
        with get_db() as session:
            existentes = set(r[0] for r in (session.query(LaboratorioDrogueria.laboratorio_id)
                .filter(LaboratorioDrogueria.drogueria_id == drog_id,
                        LaboratorioDrogueria.laboratorio_id.in_(lab_ids))).all())
            creados = 0
            for lid in lab_ids:
                if lid in existentes:
                    continue
                session.add(LaboratorioDrogueria(laboratorio_id=lid, drogueria_id=drog_id))
                creados += 1
            session.commit()
            return jsonify({
                'ok': True,
                'creados': creados,
                'ya_existian': len(lab_ids) - creados,
                'total_solicitados': len(lab_ids),
            })

    @app.route('/api/pedidos/dia/emitir', methods=['POST'])
    @login_required
    def api_compras_dia_emitir():
        """Snapshot del armado a PedidoEmitido + PedidoEmitidoItem.
        Body: {prov_id, items: [{observer_id, producto_id_local, descripcion,
                                  lab_nombre, cantidad}], observacion?}
        """
        from database import PedidoEmitido, PedidoEmitidoItem, Producto, Provider
        data = request.get_json(silent=True) or {}
        try:
            prov_id = int(data.get('prov_id') or 0)
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'prov_id inválido'}), 400
        items = data.get('items') or []
        # Canal "Directo al lab" (modo lab): no viene prov_id sino lab_nombre.
        # Resolvemos/creamos un Provider tipo 'laboratorio' para colgar el pedido.
        lab_nombre_in = (data.get('lab_nombre') or '').strip()
        if not items:
            return jsonify({'ok': False, 'error': 'Faltan datos'}), 400
        if not prov_id and not lab_nombre_in:
            return jsonify({'ok': False, 'error': 'Faltan datos'}), 400
        with get_db() as session:
            if not prov_id and lab_nombre_in:
                lab_prov = (session.query(Provider)
                            .filter(Provider.razon_social == lab_nombre_in,
                                    Provider.tipo == 'laboratorio').first())
                if not lab_prov:
                    lab_prov = Provider(razon_social=lab_nombre_in,
                                        tipo='laboratorio', activo=True)
                    session.add(lab_prov)
                    session.flush()
                prov_id = lab_prov.id
            # Pre-cargar observer_ids de los productos locales para auto-bridging
            prod_ids = [int(it['producto_id_local']) for it in items
                        if it.get('producto_id_local') and not it.get('observer_id')]
            obs_by_prod = {}
            if prod_ids:
                obs_by_prod = {
                    r.id: r.observer_id
                    for r in session.query(Producto.id, Producto.observer_id)
                                    .filter(Producto.id.in_(prod_ids),
                                            Producto.observer_id.isnot(None)).all()
                }
            tipo_prov = session.query(Provider.tipo).filter_by(id=prov_id).scalar() or 'drogueria'
            origen = 'P.Dia.Lab' if tipo_prov == 'laboratorio' else 'P.Dia.Drog'
            ped = PedidoEmitido(
                drogueria_id=prov_id,
                usuario=getattr(current_user, 'username', None),
                emitido_por=(data.get('emitido_por') or '').strip() or None,
                total_items=len(items),
                total_unidades=sum(int(it.get('cantidad') or 0) for it in items),
                observacion=(data.get('observacion') or None),
                origen=origen,
            )
            session.add(ped)
            session.flush()
            for it in items:
                cant = int(it.get('cantidad') or 0)
                if cant <= 0:
                    continue
                prod_id_local = it.get('producto_id_local') or None
                observer_id = (it.get('observer_id')
                               or obs_by_prod.get(int(prod_id_local) if prod_id_local else 0)
                               or None)
                _dto = it.get('oferta_dto')
                _min = it.get('oferta_min')
                session.add(PedidoEmitidoItem(
                    pedido_id=ped.id,
                    observer_id=observer_id,
                    producto_id_local=prod_id_local,
                    descripcion=it.get('descripcion') or '',
                    lab_nombre=it.get('lab_nombre') or None,
                    cantidad_pedida=cant,
                    cantidad_recibida=0,
                    estado='PENDIENTE',
                    oferta_dto=float(_dto) if _dto is not None else None,
                    oferta_min=int(_min) if _min is not None else None,
                ))
            session.commit()
            return jsonify({'ok': True, 'pedido_id': ped.id})

    @app.route('/pedidos-emitidos')
    @login_required
    def pedidos_emitidos_list():
        from datetime import timedelta

        from database import PedidoEmitido
        from helpers import now_ar
        es_pedidos = getattr(current_user, 'rol', None) == 'pedidos'
        with get_db() as session:
            q = session.query(PedidoEmitido)
            if es_pedidos:
                # El rol pedidos solo ve los no cerrados de los últimos 30 días.
                desde = now_ar() - timedelta(days=30)
                q = q.filter(PedidoEmitido.estado != 'CERRADO',
                             PedidoEmitido.fecha >= desde)
            peds = q.order_by(PedidoEmitido.fecha.desc()).all()
            data = []
            for p in peds:
                items = p.items
                pendientes  = sum(1 for i in items if i.estado == 'PENDIENTE')
                no_vino     = sum(1 for i in items if i.estado == 'NO_VINO')
                recibidos   = sum(1 for i in items if i.estado == 'RECIBIDO')
                # Etapa Recepción: un operador firmó la recepción
                etapa_recep  = bool(p.recibido_por)
                # Etapa Carga: un operador firmó la carga del XLS Observer
                etapa_carga  = bool(p.cargado_por)
                # Etapa Factura: TODO — por ahora siempre False
                etapa_factura = False
                data.append({
                    'id': p.id,
                    'fecha': p.fecha,
                    'drog': p.drogueria.razon_social if p.drogueria else '—',
                    'drog_id': p.drogueria_id,
                    'total_items': p.total_items,
                    'total_unidades': p.total_unidades,
                    'estado': p.estado,
                    'origen': p.origen or '—',
                    'usuario': p.usuario or '—',
                    'emitido_por': p.emitido_por or '—',
                    'recibido_por': p.recibido_por or None,
                    'cargado_por': p.cargado_por or None,
                    'pendientes': pendientes,
                    'no_vino': no_vino,
                    'recibidos': recibidos,
                    'etapa_recep': etapa_recep,
                    'etapa_carga': etapa_carga,
                    'etapa_factura': etapa_factura,
                })
        return render_template('pedidos_emitidos_list.html', pedidos=data)

    @app.route('/pedidos-emitidos/<int:pedido_id>')
    @login_required
    def pedido_emitido_detalle(pedido_id):
        from database import PedidoEmitido
        with get_db() as session:
            p = session.get(PedidoEmitido, pedido_id)
            if not p:
                return redirect(url_for('pedidos_emitidos_list'))
            items = sorted(p.items, key=lambda i: (i.descripcion or '').lower())
            from database import ObsCodigoBarras, Producto, ProductoCodigoBarra
            # EANs desde ObServer (fuente principal — obs_codigos_barras)
            obs_ids = [i.observer_id for i in items if i.observer_id]
            obs_eans_map = {}  # observer_id → [ean1, ean2, ...] ordered by orden
            if obs_ids:
                for r in (session.query(ObsCodigoBarras.producto_observer,
                                        ObsCodigoBarras.codigo_barras,
                                        ObsCodigoBarras.orden)
                          .filter(ObsCodigoBarras.producto_observer.in_(obs_ids),
                                  ObsCodigoBarras.fecha_baja.is_(None))
                          .order_by(ObsCodigoBarras.producto_observer,
                                    ObsCodigoBarras.orden)
                          .all()):
                    obs_eans_map.setdefault(r.producto_observer, []).append(r.codigo_barras)
            # EANs desde productos locales (fallback / EANs extra mapeados por scanner)
            prod_ids = [i.producto_id_local for i in items if i.producto_id_local]
            local_eans_map = {}   # producto_id_local → [ean, ...]
            if prod_ids:
                for r in session.query(Producto.id, Producto.codigo_barra)\
                                .filter(Producto.id.in_(prod_ids)).all():
                    if r.codigo_barra:
                        local_eans_map.setdefault(r.id, []).append(r.codigo_barra)
                for r in session.query(ProductoCodigoBarra.producto_id,
                                       ProductoCodigoBarra.codigo_barra)\
                                .filter(ProductoCodigoBarra.producto_id.in_(prod_ids)).all():
                    if r.codigo_barra and r.codigo_barra not in local_eans_map.get(r.producto_id, []):
                        local_eans_map.setdefault(r.producto_id, []).append(r.codigo_barra)
            ped_data = {
                'id': p.id, 'fecha': p.fecha, 'estado': p.estado,
                'drog': p.drogueria.razon_social if p.drogueria else '—',
                'usuario': p.usuario or '—', 'observacion': p.observacion or '',
                'total_items': p.total_items, 'total_unidades': p.total_unidades,
                'recibido_por': p.recibido_por,
                'cargado_por': p.cargado_por,
                'etapa_recep': bool(p.recibido_por),
                'etapa_carga': bool(p.cargado_por),
                'etapa_factura': False,  # TODO: vincular con Invoice
                'items': [{
                    'id': i.id, 'observer_id': i.observer_id,
                    'producto_id_local': i.producto_id_local,
                    'ean': (obs_eans_map.get(i.observer_id) or local_eans_map.get(i.producto_id_local) or [''])[0],
                    'eans': obs_eans_map.get(i.observer_id) or local_eans_map.get(i.producto_id_local) or [],
                    'descripcion': i.descripcion, 'lab': i.lab_nombre or '—',
                    'pedida': i.cantidad_pedida,
                    'revisada': i.cantidad_revisada_op,
                    'confirmada': i.cantidad_confirmada_obs,
                    'factura': None,   # TODO: cruzar con InvoiceItem por droguería + fecha
                    'recibida': i.cantidad_recibida,
                    'estado': i.estado,
                } for i in items],
            }
        es_pedidos = getattr(current_user, 'rol', None) == 'pedidos'
        solo_lectura = es_pedidos and ped_data['estado'] == 'CERRADO'
        return render_template('pedido_emitido_detalle.html',
                               pedido=ped_data, solo_lectura=solo_lectura)

    @app.route('/api/pedido-emitido/<int:pedido_id>/recepcion', methods=['POST'])
    @login_required
    def api_pedido_recepcion(pedido_id):
        """Primera revisión manual del operador.

        Body: {items: [{id, revisada}]}. `revisada` es la cantidad que entró
        según el operador (default = cantidad_pedida; si "no entró" → 0).
        Si la confirmación de Observer ya está cargada, NO la pisa.
        """
        from database import PedidoEmitido, PedidoEmitidoItem
        from helpers import now_ar
        data = request.get_json(silent=True) or {}
        items_in = {int(it['id']): it for it in (data.get('items') or []) if it.get('id')}
        with get_db() as session:
            p = session.get(PedidoEmitido, pedido_id)
            if not p:
                return jsonify({'ok': False, 'error': 'Pedido no encontrado'}), 404
            if (getattr(current_user, 'rol', None) == 'pedidos'
                    and p.estado == 'CERRADO'):
                return jsonify({'ok': False, 'error': 'Pedido cerrado, solo lectura'}), 403
            ahora = now_ar()
            for it in p.items:
                if it.id not in items_in:
                    continue
                payload = items_in[it.id]
                try:
                    rev = int(payload.get('revisada') or 0)
                except (TypeError, ValueError):
                    rev = 0
                it.cantidad_revisada_op = max(0, rev)
                it.revisada_en = ahora
                _recalc_item_canonico(it)
            _recalc_pedido(p)
            recibido_por = (data.get('recibido_por') or '').strip()
            if recibido_por and not p.recibido_por:
                p.recibido_por = recibido_por
            session.commit()
        return jsonify({'ok': True})

    @app.route('/api/pedido-emitido/<int:pedido_id>/mapear-ean', methods=['POST'])
    @login_required
    def api_pedido_mapear_ean(pedido_id):
        """Guarda la equivalencia EAN escaneado → producto local.

        Body: {ean, producto_id_local?N, observer_id?N}
        Si solo viene observer_id, busca/crea el Producto local correspondiente.
        Inserta en producto_codigos_barra con fuente='scanner_recep'.
        También actualiza pedido_emitido_item.producto_id_local para todas las
        filas de este pedido con el mismo observer_id.
        """
        from database import ObsCodigoBarras, ObsProducto, PedidoEmitidoItem, Producto, ProductoCodigoBarra
        data = request.get_json(silent=True) or {}
        ean = (data.get('ean') or '').strip()
        try:
            prod_id = int(data.get('producto_id_local') or 0)
        except (TypeError, ValueError):
            prod_id = 0
        try:
            obs_id = int(data.get('observer_id') or 0)
        except (TypeError, ValueError):
            obs_id = 0
        if not ean:
            return jsonify({'ok': False, 'error': 'Faltan datos'}), 400
        with get_db() as session:
            if not prod_id and obs_id:
                # Buscar o crear Producto local vinculado a este observer_id
                prod = session.query(Producto).filter_by(observer_id=obs_id).first()
                if not prod:
                    obs = session.get(ObsProducto, obs_id)
                    if not obs:
                        return jsonify({'ok': False, 'error': 'Producto observer no encontrado'}), 404
                    # EAN principal desde obs_codigos_barras (orden 1)
                    ean_principal_row = (session.query(ObsCodigoBarras.codigo_barras)
                                         .filter_by(producto_observer=obs_id)
                                         .filter(ObsCodigoBarras.fecha_baja.is_(None))
                                         .order_by(ObsCodigoBarras.orden)
                                         .first())
                    ean_principal = ean_principal_row[0] if ean_principal_row else ean
                    prod = Producto(
                        codigo_barra=ean_principal,
                        descripcion=obs.descripcion,
                        observer_id=obs_id,
                    )
                    session.add(prod)
                    session.flush()
                prod_id = prod.id
                # Vincular todas las filas del pedido con este observer_id
                (session.query(PedidoEmitidoItem)
                 .filter_by(pedido_id=pedido_id, observer_id=obs_id)
                 .update({'producto_id_local': prod_id}))
            elif not prod_id:
                return jsonify({'ok': False, 'error': 'Faltan datos'}), 400
            else:
                prod = session.get(Producto, prod_id)
                if not prod:
                    return jsonify({'ok': False, 'error': 'Producto no encontrado'}), 404
            existe = session.query(ProductoCodigoBarra)\
                            .filter_by(producto_id=prod_id, codigo_barra=ean).first()
            if not existe:
                session.add(ProductoCodigoBarra(
                    producto_id=prod_id,
                    codigo_barra=ean,
                    es_principal=False,
                    fuente='scanner_recep',
                ))
            session.commit()
        return jsonify({'ok': True, 'ean': ean, 'producto_id_local': prod_id})

    @app.route('/api/pedido-emitido/<int:pedido_id>/importar-xls', methods=['POST'])
    @login_required
    def api_pedido_importar_xls(pedido_id):
        """Importa XLS de ingreso Observer (mismo formato que ERP upload).

        Cruza por EAN → descripción normalizada. Llena cantidad_confirmada_obs
        en cada ítem matcheado. Guarda cargado_por desde el form.
        """
        import os
        import tempfile

        from data_extract import parse_erp_excel
        from database import PedidoEmitido
        from helpers import now_ar

        f = request.files.get('xls')
        cargado_por = (request.form.get('cargado_por') or '').strip() or None
        if not f:
            return jsonify({'ok': False, 'error': 'Falta el archivo XLS'}), 400

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name

        try:
            erp_items = parse_erp_excel(tmp_path)
        except Exception as e:
            return jsonify({'ok': False, 'error': f'No se pudo leer el XLS: {e}'}), 400
        finally:
            os.unlink(tmp_path)

        # Índice por EAN y por descripción normalizada
        def _norm(s):
            return ' '.join((s or '').lower().split())

        erp_by_ean  = {}
        erp_by_desc = {}
        for row in erp_items:
            ean = str(row.get('codigo_barra') or '').strip()
            if ean and ean != 'nan':
                erp_by_ean[ean] = row
            desc = _norm(row.get('descripcion', ''))
            if desc:
                erp_by_desc[desc] = row

        with get_db() as session:
            p = session.get(PedidoEmitido, pedido_id)
            if not p:
                return jsonify({'ok': False, 'error': 'Pedido no encontrado'}), 404
            ahora = now_ar()
            matched = 0
            for it in p.items:
                ean = str(it.observer_id or '')
                row = erp_by_ean.get(ean) or erp_by_desc.get(_norm(it.descripcion))
                if row is None:
                    continue
                it.cantidad_confirmada_obs = max(0, int(row.get('cantidad') or 0))
                it.confirmada_en = ahora
                _recalc_item_canonico(it)
                matched += 1
            if cargado_por and not p.cargado_por:
                p.cargado_por = cargado_por
            _recalc_pedido(p)
            session.commit()

        return jsonify({'ok': True, 'matched': matched, 'total': len(erp_items)})

    # ── Export plantilla desde PedidoEmitido ──────────────────────────────
    @app.route('/api/pedido-emitido/<int:pedido_id>/export-plantilla')
    @login_required
    def api_pedido_emitido_export_plantilla(pedido_id):
        """Genera archivo con la plantilla default (tipo_doc=pedido) de la droguería."""
        import json as _json
        from io import BytesIO, StringIO

        from flask import Response, send_file

        from database import PedidoEmitido, Plantilla

        with get_db() as session:
            ped = session.get(PedidoEmitido, pedido_id)
            if not ped:
                return jsonify({'ok': False, 'error': 'Pedido no encontrado'}), 404
            plant = (session.query(Plantilla)
                     .filter_by(entidad_tipo='drogueria', entidad_id=ped.drogueria_id,
                                tipo_doc='pedido', es_default=True)
                     .order_by(Plantilla.formato.desc())   # xlsx (x) antes que txt_fijo (t)
                     .first()
                     or session.query(Plantilla)
                     .filter_by(entidad_tipo='drogueria', entidad_id=ped.drogueria_id,
                                tipo_doc='pedido')
                     .order_by(Plantilla.formato.desc())
                     .first())
            if not plant:
                return jsonify({'ok': False, 'error': 'Sin plantilla de pedido configurada para esta droguería'}), 404

            # Resolver EANs reales desde obs_codigos_barras
            obs_ids = [it.observer_id for it in ped.items if it.observer_id]
            ean_map = {}
            if obs_ids:
                from database import ObsCodigoBarras
                for oid, cb in (session.query(ObsCodigoBarras.producto_observer,
                                              ObsCodigoBarras.codigo_barras)
                                .filter(ObsCodigoBarras.producto_observer.in_(obs_ids),
                                        ObsCodigoBarras.fecha_baja.is_(None))
                                .order_by(ObsCodigoBarras.orden.asc()).all()):
                    if oid not in ean_map and cb:
                        ean_map[oid] = cb

            # Fallback al master local: donde viven los EANs backfilleados
            # (Kellerhoff / farmacia hermana) cuando ObServer no tiene
            # obs_codigos_barras cargado. ObServer tiene precedencia (se cargó arriba).
            faltan = [o for o in obs_ids if o not in ean_map]
            if faltan:
                from database import Producto
                for oid, cb in (session.query(Producto.observer_id, Producto.codigo_barra)
                                .filter(Producto.observer_id.in_(faltan),
                                        Producto.codigo_barra.isnot(None))):
                    if cb and not str(cb).startswith('OBS-') and oid not in ean_map:
                        ean_map[oid] = cb

            rows = [{
                'ean': ean_map.get(it.observer_id, it.observer_id or ''),
                'codigo_barra': ean_map.get(it.observer_id, it.observer_id or ''),
                'nombre': it.descripcion,
                'descripcion': it.descripcion,
                'cantidad': it.cantidad_pedida,
                'total': it.cantidad_pedida,
                'lab': it.lab_nombre or '',
            } for it in ped.items]

            # Columna EAN-Kellerhoff: el EAN que Kellerhoff reconoce. Si el nuestro
            # ya está en su catálogo se manda igual; si no, se corrige al EAN de
            # Kellerhoff vía la equivalencia. Kellerhoff importa por EAN, no por su
            # código interno.
            from routes.kellerhoff import corregir_eans
            _kelmap = corregir_eans(session, [r['ean'] for r in rows])
            for r in rows:
                r['ean_kellerhoff'] = _kelmap.get(r['ean'], r['ean'])

            try:
                cfg = _json.loads(plant.config_json or '{}')
            except Exception:
                cfg = {}

            import unicodedata as _ud
            _drog_raw = ped.drogueria.razon_social if ped.drogueria else 'drog'
            drog = _ud.normalize('NFKD', _drog_raw).encode('ascii', 'ignore').decode().replace(' ', '_').strip('_') or 'drog'
            from helpers import now_ar as _now
            fecha_str = _now().strftime('%Y%m%d')

            # Alias: los nombres de campo del editor de plantillas → keys del dict de fila
            _FIELD_ALIAS = {
                'codigo_barra': 'ean', 'ean': 'ean',
                'descripcion': 'nombre', 'nombre': 'nombre',
                'cantidad': 'cantidad', 'total': 'cantidad',
                'laboratorio': 'lab', 'lab': 'lab',
            }

            from database import CAMPOS_SISTEMA as _CAMPOS
            _CAMPO_LABEL = dict(_CAMPOS)

            # Encabezado libre del usuario (config.advanced.header_*).
            # Se renderiza arriba de los headers de columnas. El texto es por
            # columna: dict {field: valor} en header_por_columna. Las columnas
            # vacías quedan en blanco. Backward compat: si viene header_texto
            # legacy (multilínea libre), se usa esa.
            _adv = cfg.get('advanced') or {}
            _header_on  = bool(_adv.get('header_incluir'))
            _header_por_col = _adv.get('header_por_columna') or {}
            _header_legacy_txt = str(_adv.get('header_texto') or '').strip('\r\n')

            if plant.formato == 'xlsx':
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                # Formato Excel por defecto por campo. Se aplica a cada celda
                # de datos. Los headers + encabezado libre quedan en General.
                _FORMAT_BY_FIELD = {
                    'codigo_barra':     '0',       # EAN: entero sin decimales
                    'ean_kellerhoff':   '0',       # EAN Kellerhoff: idem
                    'precio':           '0.00',    # PVP: 2 decimales
                    'avg_monthly':      '0.0',     # Promedio mensual: 1 decimal
                }
                # Labels cortos para el header del XLSX. Las drogs esperan
                # "CodigoBarra" / "Cantidad", no el label largo de CAMPOS_SISTEMA
                # ("Código de barra (EAN)" / "Cantidad total (mod+oferta+sin deal)").
                _HEADER_LABEL = {
                    'codigo_barra':     'CodigoBarra',
                    'ean_kellerhoff':   'CodigoBarra',
                    'descripcion':      'Descripcion',
                    'cantidad':         'Cantidad',
                    'cant_modulo':      'CantModulo',
                    'cant_oferta_min':  'CantOfertaMin',
                    'cant_nodeal':      'CantSinDeal',
                    'precio':           'Precio',
                    'erp_qty':          'StockERP',
                    'rotacion':         'Rotacion',
                    'avg_monthly':      'PromedioMensual',
                    'espacio':          '',
                }
                cols = [c if isinstance(c, str) else c.get('field', '') for c in cfg.get('columnas', [])]
                if not cols:
                    cols = ['codigo_barra', 'descripcion', 'cantidad']

                # Headers cortos por droguería (CodigoBarra, Cantidad, …).
                headers = [_HEADER_LABEL.get(c, _CAMPO_LABEL.get(c, c)) for c in cols]
                # Encabezado libre opcional ARRIBA de los headers de columna
                # (CUIT, código de cliente, etiqueta libre). Si el usuario tipeó
                # justo los mismos nombres que los headers de columna (caso
                # Kellerhoff: "CodigoBarra"/"Cantidad"), NO lo duplicamos — los
                # headers de columna ya los emiten igual.
                if _header_on:
                    if _header_por_col:
                        hdr_row = [str(_header_por_col.get(c, '') or '') for c in cols]
                        if hdr_row != headers:
                            ws.append(hdr_row)
                    elif _header_legacy_txt:
                        for line in _header_legacy_txt.splitlines():
                            ws.append([line])
                ws.append(headers)
                # Filas de datos + formato de celda por columna.
                for row in rows:
                    row_vals = []
                    for field in cols:
                        v = row.get(_FIELD_ALIAS.get(field, field), '')
                        # EAN: convertir a int para que Excel lo trate numérico.
                        if field in ('codigo_barra', 'ean_kellerhoff') and v not in (None, ''):
                            try:
                                v = int(str(v).strip())
                            except (ValueError, TypeError):
                                pass  # dejar string si no es numérico
                        row_vals.append(v)
                    ws.append(row_vals)
                    # Aplicar formato a la fila recién agregada.
                    for col_idx, field in enumerate(cols, start=1):
                        fmt = _FORMAT_BY_FIELD.get(field)
                        if fmt:
                            ws.cell(row=ws.max_row, column=col_idx).number_format = fmt
                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                fname = f'Pedido_{drog}_{fecha_str}.xlsx'
                return send_file(buf, as_attachment=True, download_name=fname,
                                 mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            # txt_fijo
            campos = sorted(cfg.get('campos', []), key=lambda c: c.get('col_inicio', 0))
            if not campos:
                return jsonify({'ok': False, 'error': 'Plantilla sin campos'}), 400
            line_len = max(c['col_inicio'] + c['longitud'] for c in campos)
            lines = []
            # Encabezado: una línea concatenando los valores por columna en
            # el mismo orden que las columnas activas (TXT/CSV). Si solo hay
            # header_texto legacy, se respeta ese fallback (multilínea).
            if _header_on:
                if _header_por_col:
                    # Para TXT no tenemos columnas (es flat). Lo aproximamos
                    # uniendo los valores con un espacio.
                    vals_hdr = [str(_header_por_col.get(c, '') or '')
                                for c in (cfg.get('columnas') or [])
                                if isinstance(c, str)]
                    line_hdr = ' '.join(v for v in vals_hdr if v).strip()
                    if line_hdr:
                        lines.append(line_hdr)
                elif _header_legacy_txt:
                    for hl in _header_legacy_txt.splitlines():
                        lines.append(hl)
            for row in rows:
                line = bytearray(b' ' * line_len)
                for c in campos:
                    cs = c.get('campo_sistema', '')
                    val = str(row.get('ean', '')) if cs == 'codigo_barra' else \
                          str(row.get('ean_kellerhoff', '')) if cs == 'ean_kellerhoff' else \
                          str(row.get('nombre', '')) if cs == 'descripcion' else \
                          str(int(row.get('cantidad', 0) or 0)) if cs == 'cantidad' else \
                          (c.get('valor_fijo') or '') if cs == 'fijo' else ''
                    pad = (c.get('relleno') or ' ')[0].encode()
                    lng = c['longitud']
                    ali = c.get('alineacion', 'L')
                    val_b = val.encode('latin-1', errors='replace')[:lng]
                    if ali == 'R':
                        val_b = val_b.rjust(lng, pad)
                    else:
                        val_b = val_b.ljust(lng, pad)
                    line[c['col_inicio']:c['col_inicio'] + lng] = val_b
                lines.append(line.decode('latin-1'))
            ext = plant.formato.replace('txt_fijo', 'txt')
            fname = f'Pedido_{drog}_{fecha_str}.{ext}'
            content = '\r\n'.join(lines)
            return Response(content, mimetype='text/plain',
                            headers={'Content-Disposition': f'attachment; filename="{fname}"'})

    @app.route('/api/pedido-emitido/<int:pedido_id>/export-xls')
    @login_required
    def api_pedido_emitido_export_xls(pedido_id):
        """Export XLSX simple del pedido (no depende de plantilla de droguería).

        Columnas fijas: EAN, Descripción, Laboratorio, Cant. pedida,
        Cant. recibida, Estado.
        """
        import unicodedata as _ud
        from io import BytesIO

        from flask import send_file

        from database import ObsCodigoBarras, PedidoEmitido, Producto, ProductoCodigoBarra

        with get_db() as session:
            ped = session.get(PedidoEmitido, pedido_id)
            if not ped:
                return jsonify({'ok': False, 'error': 'Pedido no encontrado'}), 404
            items = sorted(ped.items, key=lambda i: (i.descripcion or '').lower())

            # Resolver EAN: ObServer (obs_codigos_barras) → master local.
            obs_ids = [it.observer_id for it in items if it.observer_id]
            ean_map = {}
            if obs_ids:
                for oid, cb in (session.query(ObsCodigoBarras.producto_observer,
                                              ObsCodigoBarras.codigo_barras)
                                .filter(ObsCodigoBarras.producto_observer.in_(obs_ids),
                                        ObsCodigoBarras.fecha_baja.is_(None))
                                .order_by(ObsCodigoBarras.orden.asc()).all()):
                    if oid not in ean_map and cb:
                        ean_map[oid] = cb
            prod_ids = [it.producto_id_local for it in items if it.producto_id_local]
            local_ean = {}
            if prod_ids:
                for pid, cb in (session.query(Producto.id, Producto.codigo_barra)
                                .filter(Producto.id.in_(prod_ids)).all()):
                    if cb and not str(cb).startswith('OBS-'):
                        local_ean[pid] = cb
                for pid, cb in (session.query(ProductoCodigoBarra.producto_id,
                                              ProductoCodigoBarra.codigo_barra)
                                .filter(ProductoCodigoBarra.producto_id.in_(prod_ids)).all()):
                    if pid not in local_ean and cb:
                        local_ean[pid] = cb

            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f'Pedido {ped.id}'
            ws.append(['EAN', 'Descripción', 'Laboratorio',
                       'Cant. pedida', 'Cant. recibida', 'Estado'])
            for it in items:
                ean = ean_map.get(it.observer_id) or local_ean.get(it.producto_id_local) or ''
                try:
                    ean = int(str(ean).strip()) if ean else ''
                except (ValueError, TypeError):
                    pass
                ws.append([ean, it.descripcion or '', it.lab_nombre or '',
                           it.cantidad_pedida or 0, it.cantidad_recibida or 0,
                           it.estado or ''])
                ws.cell(row=ws.max_row, column=1).number_format = '0'

            _drog_raw = ped.drogueria.razon_social if ped.drogueria else 'drog'
            drog = (_ud.normalize('NFKD', _drog_raw).encode('ascii', 'ignore')
                    .decode().replace(' ', '_').strip('_')) or 'drog'
            fname = f'Pedido_{ped.id}_{drog}.xlsx'
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf, as_attachment=True, download_name=fname,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ── Usuarios de pedidos (operadores) ──────────────────────────────────
    @app.route('/api/usuarios-pedidos', methods=['GET', 'POST'])
    @login_required
    def api_usuarios_pedidos():
        from database import UsuarioPedido
        if request.method == 'POST':
            data = request.get_json(silent=True) or {}
            nombre = (data.get('nombre') or '').strip()
            if not nombre:
                return jsonify({'ok': False, 'error': 'Nombre requerido'}), 400
            with get_db() as session:
                existente = session.query(UsuarioPedido).filter_by(nombre=nombre).first()
                if existente:
                    if existente.activo:
                        return jsonify({'ok': False, 'error': 'Ya existe ese nombre'}), 400
                    existente.activo = True
                    session.commit()
                    return jsonify({'ok': True, 'id': existente.id, 'nombre': existente.nombre})
                u = UsuarioPedido(nombre=nombre)
                session.add(u)
                try:
                    session.commit()
                except Exception:
                    session.rollback()
                    return jsonify({'ok': False, 'error': 'Ya existe ese nombre'}), 400
                return jsonify({'ok': True, 'id': u.id, 'nombre': u.nombre})
        with get_db() as session:
            users = session.query(UsuarioPedido).filter_by(activo=True)\
                           .order_by(UsuarioPedido.nombre).all()
            return jsonify({'ok': True, 'users': [{'id': u.id, 'nombre': u.nombre} for u in users]})

    @app.route('/api/usuarios-pedidos/<int:uid>', methods=['DELETE'])
    @login_required
    def api_usuarios_pedidos_borrar(uid):
        from database import UsuarioPedido
        with get_db() as session:
            u = session.get(UsuarioPedido, uid)
            if u:
                u.activo = False
                session.commit()
        return jsonify({'ok': True})

    @app.route('/api/producto/<int:prod_id>/excluir', methods=['POST'])
    @login_required
    def api_producto_excluir(prod_id):
        """Body: {modo: 'temporal'|'permanente'}.
        - temporal: excluido_armado_actual=TRUE (se autodescactiva cuando stock>min).
        - permanente: no_pedir=TRUE (manual reactivación).
        """
        from database import Producto
        data = request.get_json(silent=True) or {}
        modo = (data.get('modo') or 'temporal').strip()
        with get_db() as session:
            p = session.get(Producto, prod_id)
            if not p:
                return jsonify({'ok': False, 'error': 'Producto no encontrado'}), 404
            if modo == 'permanente':
                p.no_pedir = True
            else:
                p.excluido_armado_actual = True
            session.commit()
        return jsonify({'ok': True})

    @app.route('/api/producto/<int:prod_id>/reactivar', methods=['POST'])
    @login_required
    def api_producto_reactivar(prod_id):
        """Saca el flag 'no_pedir' (y de paso 'excluido_armado_actual')."""
        from database import Producto
        with get_db() as session:
            p = session.get(Producto, prod_id)
            if not p:
                return jsonify({'ok': False, 'error': 'Producto no encontrado'}), 404
            p.no_pedir = False
            p.excluido_armado_actual = False
            session.commit()
        return jsonify({'ok': True})
