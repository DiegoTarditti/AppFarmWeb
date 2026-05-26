"""Procesos de compra: ciclo análisis → pedido → factura → cruce → reclamo → cierre."""

import json
import os
from datetime import datetime

from flask import flash, jsonify, redirect, render_template, request, url_for
from flask_login import current_user

import database
from database import AnalisisSesion, Claim, Invoice, Laboratorio, Pedido, ProcesoCompra, Provider
from helpers import now_ar

# ── Configuración de pasos ──────────────────────────────────────────────────
PASOS = ['analisis', 'pedido', 'factura', 'cruce', 'reclamo']

ESTADO_POR_PASO = {
    'analisis': 'BORRADOR',
    'pedido':   'PEDIDO',
    'factura':  'FACTURADO',
    'cruce':    'INGRESADO',
    'reclamo':  'INGRESADO',   # reclamo no cambia el estado principal
}

ESTADOS_ORDEN = ['BORRADOR', 'PEDIDO', 'FACTURADO', 'INGRESADO', 'CERRADO']


def _estado_index(est):
    try:
        return ESTADOS_ORDEN.index(est or 'BORRADOR')
    except ValueError:
        return 0


def _bump_estado(proc, nuevo):
    if _estado_index(nuevo) > _estado_index(proc.estado):
        proc.estado = nuevo


def _inferir_hecho(proc, session):
    """Decide si cada paso está hecho combinando timestamps + asociaciones reales.

    Esto es lo que hace 'auto-marcado': el usuario no necesita clickear 'Marcar
    hecho' si ya subió factura / creó pedido / hizo cruce.
    """
    from database import StockDifference
    tiene_diff_cruce = False
    if proc.factura_id:
        tiene_diff_cruce = session.query(StockDifference).filter_by(
            factura_id=proc.factura_id).first() is not None
    return {
        'analisis': bool(proc.analisis_hecho_en) or proc.analisis_sesion_id is not None,
        'pedido':   bool(proc.pedido_hecho_en)   or proc.pedido_id is not None,
        'factura':  bool(proc.factura_hecha_en)  or proc.factura_id is not None,
        'cruce':    bool(proc.cruce_hecho_en)    or tiene_diff_cruce,
        'reclamo':  bool(proc.reclamo_hecho_en)  or proc.reclamo_id is not None,
    }


def _serialize_list(proc, session=None):
    hechos_dict = _inferir_hecho(proc, session) if session else None
    pasos = []
    for p in PASOS:
        campo = 'factura_hecha_en' if p == 'factura' else f'{p}_hecho_en'
        ts = getattr(proc, campo, None)
        hecho = hechos_dict[p] if hechos_dict is not None else bool(ts)
        pasos.append({'paso': p, 'hecho': hecho,
                      'fecha': ts.strftime('%d/%m/%Y') if ts else None})
    return {
        'id': proc.id, 'tipo': proc.tipo,
        'partner_nombre': proc.partner_nombre, 'estado': proc.estado,
        'creado_en': proc.creado_en.strftime('%d/%m/%Y') if proc.creado_en else '',
        'actualizado_en': proc.actualizado_en.strftime('%d/%m/%Y') if proc.actualizado_en else '',
        'pedido_id': proc.pedido_id, 'factura_id': proc.factura_id, 'reclamo_id': proc.reclamo_id,
        'analisis_periodo': proc.analisis_periodo or '',
        'pasos': pasos,
    }


def _paso_field(paso):
    return 'factura_hecha_en' if paso == 'factura' else f'{paso}_hecho_en'


def init_app(app):

    @app.route('/procesos')
    def procesos_list():
        estado_filter = (request.args.get('estado') or '').strip().upper()
        tipo_filter = (request.args.get('tipo') or '').strip()
        q_text = (request.args.get('q') or '').strip().lower()

        with database.get_db() as session:
            base = session.query(ProcesoCompra)
            if estado_filter in ESTADOS_ORDEN:
                base = base.filter(ProcesoCompra.estado == estado_filter)
            if tipo_filter in ('laboratorio', 'drogueria'):
                base = base.filter(ProcesoCompra.tipo == tipo_filter)
            procs = base.order_by(ProcesoCompra.actualizado_en.desc()).all()
            if q_text:
                procs = [p for p in procs if q_text in (p.partner_nombre or '').lower()]
            data = [_serialize_list(p, session) for p in procs]

            # Conteos para filtros: GROUP BY en SQL en vez de traer todas las
            # filas y contar en Python. NULL → 'BORRADOR'; el `+= n` acumula el
            # grupo NULL y el grupo 'BORRADOR' explícito en la misma clave.
            from sqlalchemy import func
            counts = {est: 0 for est in ESTADOS_ORDEN}
            for estado, n in (session.query(ProcesoCompra.estado, func.count())
                              .group_by(ProcesoCompra.estado).all()):
                est = estado or 'BORRADOR'
                if est in counts:
                    counts[est] += n

            laboratorios = [{'id': l.id, 'nombre': l.nombre}
                            for l in session.query(Laboratorio).order_by(Laboratorio.nombre).all()]
            proveedores = [{'id': p.id, 'nombre': p.razon_social}
                           for p in session.query(Provider).order_by(Provider.razon_social).all()]

            import observer_source
            estado_ventas = observer_source.estado_ventas_mensuales(session)
            # Para habilitar el flujo "ir directo a análisis" basta con tener datos
            # locales en obs_ventas_mensuales (no hace falta conexión SQL Server).
            observer_disponible = observer_source.observer_analisis_disponible()

        return render_template('procesos_list.html', procesos=data, counts=counts,
                               estado_filter=estado_filter, tipo_filter=tipo_filter,
                               q_text=q_text, total=len(data),
                               laboratorios=laboratorios, proveedores=proveedores,
                               estados=ESTADOS_ORDEN,
                               estado_ventas=estado_ventas,
                               observer_disponible=observer_disponible)

    @app.route('/consulta-stock')
    def consulta_stock():
        """Entry point móvil-first: elegir lab/drog/prov y arrancar análisis.

        El POST va a /consulta-stock/iniciar, que para laboratorio ejecuta el
        análisis con defaults (n_days=35, mes/año actual) y redirige al
        template mobile-friendly /consulta-stock/resultado/<uid>.
        """
        import observer_source
        with database.get_db() as session:
            estado_ventas = observer_source.estado_ventas_mensuales(session)
        observer_disponible = observer_source.observer_analisis_disponible()
        return render_template('consulta_stock.html',
                               estado_ventas=estado_ventas,
                               observer_disponible=observer_disponible)

    @app.route('/consulta-stock/iniciar', methods=['POST'])
    def consulta_stock_iniciar():
        """Pipeline móvil: crea ProcesoCompra + ejecuta análisis con defaults
        + redirige al template resultado móvil. Skip la pantalla intermedia
        observer_analizar (no requiere ajustar n_days/año/mes en el cel).
        """
        import json as _json
        import os as _os
        import uuid as _uuid
        from datetime import datetime as _datetime

        import observer_source
        from helpers import PURCHASE_FOLDER, get_config
        from purchase_engine import analyze_purchase

        tipo = (request.form.get('tipo') or '').strip()
        partner_id = request.form.get('partner_id') or None
        periodo = (request.form.get('periodo') or '').strip()

        if tipo not in ('laboratorio', 'drogueria', 'proveedor'):
            flash('Tipo inválido.', 'error')
            return redirect(url_for('consulta_stock'))

        partner_nombre = ''
        with database.get_db() as session:
            if partner_id:
                if tipo == 'laboratorio':
                    lab = session.get(Laboratorio, int(partner_id))
                    partner_nombre = lab.nombre if lab else ''
                else:
                    prov = session.get(Provider, int(partner_id))
                    partner_nombre = prov.razon_social if prov else ''
        if not partner_nombre:
            flash('Seleccioná un partner.', 'error')
            return redirect(url_for('consulta_stock'))

        # Crear ProcesoCompra (igual que el desktop).
        with database.get_db() as session:
            proc = ProcesoCompra(
                tipo=tipo,
                partner_id=int(partner_id) if partner_id else None,
                partner_nombre=partner_nombre,
                analisis_periodo=periodo or None,
                estado='BORRADOR',
            )
            session.add(proc)
            session.commit()
            proc_id = proc.id

        # Solo laboratorio tiene flujo de análisis. Drog/prov van al detail.
        if tipo != 'laboratorio':
            flash(f'Proceso creado para {partner_nombre}. Análisis móvil de '
                  f'{tipo} todavía no implementado — desktop por ahora.', 'info')
            return redirect(url_for('proceso_detail', proceso_id=proc_id))

        if not observer_source.observer_analisis_disponible():
            flash('Sin datos de ventas. Sync ObServer pendiente.', 'error')
            return redirect(url_for('consulta_stock'))

        # n_days viene del slider mobile (7-120, default 35). mes/año actual.
        try:
            n_days = max(7, min(120, int(request.form.get('n_days', 35))))
        except (ValueError, TypeError):
            n_days = 35
        hoy = _datetime.now()
        anio, mes = hoy.year, hoy.month

        # Fecha del último sync de ventas (para mostrar "Datos al: ..." en la UI).
        with database.get_db() as _s_est:
            est_v = observer_source.estado_ventas_mensuales(_s_est)
        ultimo_sync = est_v.get('ultimo_sync')
        ultimo_sync_str = (ultimo_sync.strftime('%d/%m/%Y') if ultimo_sync else None)

        productos = observer_source.get_ventas_laboratorio(partner_nombre, anio, mes)
        if not productos:
            flash(f'Sin ventas de "{partner_nombre}" en {mes:02d}/{anio}.', 'warning')
            return redirect(url_for('consulta_stock'))

        # Calcular start_month (12 meses atrás)
        start_m = mes - 11
        start_y = anio
        while start_m <= 0:
            start_m += 12
            start_y -= 1

        cfg = get_config()
        results = analyze_purchase(
            productos, n_days, start_m, mes,
            umbral_pico=cfg['umbral_pico'],
            umbral_baja=cfg['umbral_baja'],
            umbral_tendencia=cfg['umbral_tendencia'],
            rot_alta_min=cfg['rot_alta_min'],
            rot_media_min=cfg['rot_media_min'],
        )

        uid = str(_uuid.uuid4())
        periodo_str = f'{start_m:02d}/{start_y} - {mes:02d}/{anio}'
        data = {
            'uid': uid,
            'farmacia': getattr(current_user, 'nombre_completo', None) or 'Farmacia',
            'laboratorio': partner_nombre,
            'periodo': periodo_str,
            'start_month': start_m,
            'n_days': n_days,
            'umbral_tendencia': cfg['umbral_tendencia'],
            'rot_alta_min': cfg['rot_alta_min'],
            'rot_alta_tol': cfg['rot_alta_tol'],
            'rot_media_min': cfg['rot_media_min'],
            'rot_media_tol': cfg['rot_media_tol'],
            'rot_baja_tol': cfg['rot_baja_tol'],
            'products': results,
            'proceso_id': proc_id,
            'datos_al': ultimo_sync_str,
        }

        with database.get_db() as session:
            sesion = AnalisisSesion(
                laboratorio_nombre=partner_nombre,
                periodo=periodo_str,
                farmacia=data['farmacia'],
                n_days=n_days,
                fuente='observer',
                n_productos=len(results),
            )
            session.add(sesion)
            session.commit()
            data['sesion_id'] = sesion.id

        json_path = _os.path.join(PURCHASE_FOLDER, f'{uid}.json')
        with open(json_path, 'w', encoding='utf-8') as jf:
            _json.dump(data, jf, ensure_ascii=False)

        return redirect(url_for('consulta_stock_resultado', uid=uid))

    @app.route('/consulta-stock/sync-stock', methods=['POST'])
    def consulta_stock_sync_stock():
        """Encola un comando 'sync_inteligente' en panel_comandos para que el
        DockerPanel local lo levante en el próximo polling y refresque SOLO lo
        vencido por tolerancia (stock 3h, ventas_mensuales 24h, productos 7d).
        Mucho más liviano que el sync completo — no trae medicos/clientes/
        ventas_detalle (esos solo en el sync completo manual).

        Retorna JSON con el id del comando + ETA típico (~30s polling +
        ~40s-1.5min de sync, según qué esté vencido).
        """
        from database import PanelComando
        username = getattr(current_user, 'username', None) or 'usuario_movil'
        with database.get_db() as session:
            cmd = PanelComando(
                comando='sync_inteligente',
                estado='pendiente',
                solicitado_por=f'{username} (móvil)',
            )
            session.add(cmd)
            session.commit()
            cmd_id = cmd.id
        return jsonify({
            'ok': True,
            'id': cmd_id,
            'mensaje': ('Sync (inteligente) encolado. La PC de la farmacia lo levanta '
                        'en ~30s y refresca solo lo vencido (stock, y ventas/catálogo '
                        'si hace falta) en ~40s-1.5 min. Re-ejecutá la consulta después.'),
        })

    @app.route('/consulta-stock/export-xls', methods=['POST'])
    def consulta_stock_export_xls():
        """Genera un XLS desde la consulta móvil. Si vienen items con qty>0,
        exporta solo esos. Sino, exporta todos los productos del análisis.

        Body JSON: { uid, laboratorio, items: [{idx, qty}] (opcional) }
        Returns: archivo .xlsx con cols EAN, Producto, Stock, Prom/mes,
                 Vendido 3m, Vendido 12m, Cobertura, Rotación, Sugerido, A pedir.
        """
        import io as _io
        import json as _json
        import os as _os
        import re as _re

        import openpyxl
        from flask import send_file
        from openpyxl.styles import Alignment, Font, PatternFill

        from helpers import PURCHASE_FOLDER

        data_req = request.get_json(silent=True) or {}
        uid = (data_req.get('uid') or '').strip()
        if not _re.match(r'^[0-9a-f-]{36}$', uid):
            return jsonify({'ok': False, 'error': 'uid inválido'}), 400
        json_path = _os.path.join(PURCHASE_FOLDER, f'{uid}.json')
        if not _os.path.exists(json_path):
            return jsonify({'ok': False, 'error': 'análisis expirado'}), 404
        with open(json_path, encoding='utf-8') as jf:
            data = _json.load(jf)

        items_qty_map = {}  # idx → qty (si vienen items específicos)
        for it in (data_req.get('items') or []):
            try:
                items_qty_map[int(it['idx'])] = int(it.get('qty') or 0)
            except (ValueError, KeyError, TypeError):
                continue
        # Si vinieron items con qty > 0, filtrar solo esos. Sino, exportar todos.
        only_marked = any(q > 0 for q in items_qty_map.values())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stock'

        # Header
        headers = ['EAN', 'Producto', 'Stock', 'Mín', 'Prom/mes', 'Vendido 3m',
                   'Vendido 12m', 'Cobertura (d)', 'Rotación', 'Tendencia',
                   'Sugerido', 'A pedir', 'Rubro']
        header_fill = PatternFill('solid', fgColor='10B981')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22

        # Filas
        row = 2
        for i, p in enumerate(data.get('products', [])):
            qty_pedida = items_qty_map.get(i, 0)
            if only_marked and qty_pedida <= 0:
                continue
            ventas = p.get('ventas') or []
            v3m = sum(ventas[:3]) if len(ventas) >= 3 else sum(ventas)
            v12m = p.get('total') or sum(ventas)
            avg = p.get('avg_monthly') or 0
            stock = p.get('stock') or 0
            cov = ''
            if avg > 0 and stock > 0:
                cov = round(stock / (avg / 30.4))
            elif stock <= 0:
                cov = 0
            slope = p.get('slope') or 0
            tend = ('↑' if slope > 0 else ('↓' if slope < 0 else '→')) + f' {slope:.1f}/m'
            ws.append([
                p.get('codigo_barra') or '',
                p.get('nombre') or '',
                stock,
                p.get('minimo') or 0,
                int(avg),
                int(v3m),
                int(v12m),
                cov,
                p.get('rotacion') or '',
                tend,
                p.get('order_qty') or 0,
                qty_pedida,
                p.get('rubro') or '',
            ])
            row += 1

        # Ajustar anchos
        widths = [16, 38, 8, 8, 10, 11, 11, 13, 10, 12, 10, 9, 22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Freeze header
        ws.freeze_panes = 'A2'

        bio = _io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        from datetime import datetime as _dt
        lab_name = _re.sub(r'[^a-zA-Z0-9_\- ]', '', data.get('laboratorio', 'lab'))[:40]
        fname = f'consulta_{lab_name}_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(
            bio, as_attachment=True, download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @app.route('/consulta-stock/armar-pedido', methods=['POST'])
    def consulta_stock_armar_pedido():
        """Crea un Pedido borrador desde el cel con los items que el operador
        marcó en /consulta-stock/resultado/<uid>. Devuelve JSON con redirect.
        """
        from database import Pedido, PedidoItem
        data = request.get_json(silent=True) or {}
        uid = (data.get('uid') or '').strip()
        laboratorio = (data.get('laboratorio') or '').strip()
        proceso_id = data.get('proceso_id')
        items = data.get('items') or []

        if not items:
            return jsonify({'ok': False, 'error': 'sin items'}), 400
        if not laboratorio:
            return jsonify({'ok': False, 'error': 'sin laboratorio'}), 400

        from datetime import date as _date_d
        fecha_str = _date_d.today().strftime('%Y-%m-%d')
        with database.get_db() as session:
            pedido = Pedido(
                laboratorio=laboratorio,
                farmacia=getattr(current_user, 'nombre_completo', None) or 'Farmacia',
                # Formato pedido: "📱 móvil · {prov/lab} · YYYY-MM-DD"
                # Aparece en /orders junto a los demás pedidos. El emoji 📱
                # lo distingue visualmente como creado desde móvil.
                periodo=f'📱 móvil · {laboratorio} · {fecha_str}',
                n_days=35,
                canal='laboratorio',
                estado='PENDIENTE',
                analizado_en=now_ar(),
                origen='Movil.Lab',
            )
            session.add(pedido)
            session.flush()
            for it in items:
                cantidad = int(it.get('cantidad') or 0)
                if cantidad <= 0:
                    continue
                precio = float(it.get('precio_pvp') or 0)
                session.add(PedidoItem(
                    pedido_id=pedido.id,
                    codigo_barra=str(it.get('codigo_barra') or '')[:20],
                    nombre=str(it.get('nombre') or '')[:200],
                    cantidad=cantidad,
                    precio_pvp=precio,
                    subtotal=cantidad * precio,
                ))
            # Vincular al ProcesoCompra si vino.
            if proceso_id:
                proc = session.get(ProcesoCompra, int(proceso_id))
                if proc:
                    proc.pedido_id = pedido.id
                    if proc.estado == 'BORRADOR':
                        proc.estado = 'PEDIDO'
                    proc.pedido_hecho_en = now_ar()
            session.commit()
            pedido_id = pedido.id

        return jsonify({
            'ok': True,
            'pedido_id': pedido_id,
            'redirect': url_for('orders_list'),
        })

    @app.route('/consulta-stock/resultado/<uid>')
    def consulta_stock_resultado(uid):
        """Versión mobile-first de purchase_results: cards apiladas + 4 indicadores
        (prom/mes, vendido 3m+12m, mini-chart 12m, cobertura días) + marcado de
        items para armar pedido desde el cel.
        """
        import json as _json
        import os as _os
        import re as _re
        from datetime import datetime as _datetime

        from helpers import PURCHASE_FOLDER, get_config

        if not _re.match(r'^[0-9a-f-]{36}$', uid):
            flash('Sesión inválida.', 'error')
            return redirect(url_for('consulta_stock'))
        json_path = _os.path.join(PURCHASE_FOLDER, f'{uid}.json')
        if not _os.path.exists(json_path):
            flash('El análisis expiró. Hacelo de nuevo.', 'error')
            return redirect(url_for('consulta_stock'))
        with open(json_path, encoding='utf-8') as jf:
            data = _json.load(jf)

        cfg = get_config()
        for k in ('umbral_tendencia', 'rot_alta_min', 'rot_alta_tol',
                  'rot_media_min', 'rot_media_tol', 'rot_baja_tol'):
            data.setdefault(k, cfg[k])

        _mes_jan = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul',
                    'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        sm = data.get('start_month', 4)
        month_es = [_mes_jan[(sm - 1 + i) % 12] for i in range(12)]
        analizado_en = _datetime.fromtimestamp(
            _os.path.getmtime(json_path)).strftime('%d/%m/%Y')

        # Orden alfabético por nombre (case-insensitive). Aplica en los 3
        # modos del template (lista / detalle / tabla) — ambos consumen
        # el mismo array data['products'] y los dicts JS (PRECIOS/NOMBRES/
        # CHART_DATA) se generan con el orden post-sort.
        data['products'] = sorted(
            data.get('products', []) or [],
            key=lambda p: (p.get('nombre') or '').upper(),
        )

        # Pre-cómputo por producto para que el template sea liviano:
        # - vendido 3m (suma de los últimos 3 meses) + 12m (total).
        # - dias de cobertura.
        for p in data.get('products', []):
            ventas = p.get('ventas') or []
            avg = p.get('avg_monthly') or 0
            stock = p.get('stock') or 0
            p['v3m'] = sum(ventas[:3]) if len(ventas) >= 3 else sum(ventas)
            p['v12m'] = p.get('total') or sum(ventas)
            if avg > 0 and stock > 0:
                p['dias_cobertura'] = round(stock / (avg / 30.4))
            elif stock <= 0:
                p['dias_cobertura'] = 0
            else:
                p['dias_cobertura'] = None  # sin ventas → infinito, no mostrar nº

        return render_template('consulta_stock_resultado.html',
                               month_es=month_es,
                               analizado_en=analizado_en,
                               **data)

    @app.route('/procesos/crear', methods=['POST'])
    def proceso_crear():
        tipo = (request.form.get('tipo') or '').strip()
        partner_id = request.form.get('partner_id') or None
        partner_nombre = (request.form.get('partner_nombre') or '').strip()
        periodo = (request.form.get('periodo') or '').strip()

        if tipo not in ('laboratorio', 'drogueria', 'proveedor'):
            flash('Tipo inválido.')
            return redirect(url_for('procesos_list'))
        if not partner_nombre:
            # Resolver desde partner_id si viene
            with database.get_db() as session:
                if partner_id:
                    if tipo == 'laboratorio':
                        lab = session.get(Laboratorio, int(partner_id))
                        partner_nombre = lab.nombre if lab else ''
                    else:
                        # 'drogueria' o 'proveedor' → ambos viven en la tabla Provider
                        prov = session.get(Provider, int(partner_id))
                        partner_nombre = prov.razon_social if prov else ''
        if not partner_nombre:
            flash('Seleccioná un partner.')
            return redirect(url_for('procesos_list'))

        with database.get_db() as session:
            proc = ProcesoCompra(
                tipo=tipo,
                partner_id=int(partner_id) if partner_id else None,
                partner_nombre=partner_nombre,
                analisis_periodo=periodo or None,
                estado='BORRADOR',
            )
            session.add(proc)
            session.commit()
            pid = proc.id

        # Para laboratorio: arrancar directo en Analizar Ventas con el lab precargado.
        # Para droguería/proveedor: ir al detail (el flujo arranca desde factura).
        if tipo == 'laboratorio':
            import observer_source
            if observer_source.observer_analisis_disponible():
                return redirect(url_for('observer_analizar',
                                        lab=partner_nombre, proceso=pid))
        return redirect(url_for('proceso_detail', proceso_id=pid))

    @app.route('/proceso/<int:proceso_id>')
    def proceso_detail(proceso_id):
        with database.get_db() as session:
            proc = session.get(ProcesoCompra, proceso_id)
            if not proc:
                flash('Proceso no encontrado.')
                return redirect(url_for('procesos_list'))

            pedido = session.get(Pedido, proc.pedido_id) if proc.pedido_id else None
            invoice = session.get(Invoice, proc.factura_id) if proc.factura_id else None
            reclamo = session.get(Claim, proc.reclamo_id) if proc.reclamo_id else None
            sesion_id = proc.analisis_sesion_id or (pedido.analisis_sesion_id if pedido else None)
            sesion = session.get(AnalisisSesion, sesion_id) if sesion_id else None

            analisis_pasos = {}
            if proc.analisis_pasos_json:
                try:
                    analisis_pasos = json.loads(proc.analisis_pasos_json)
                except (ValueError, TypeError):
                    analisis_pasos = {}

            # Para sugerir links: pedidos/facturas del mismo partner sin proceso
            pedidos_libres = []
            if not proc.pedido_id:
                query_p = session.query(Pedido).filter(
                    Pedido.laboratorio.ilike(f'%{proc.partner_nombre}%')
                ).order_by(Pedido.creado_en.desc()).limit(10).all()
                usados = {p.pedido_id for p in session.query(ProcesoCompra).filter(
                    ProcesoCompra.pedido_id.isnot(None)).all()}
                pedidos_libres = [{'id': p.id, 'periodo': p.periodo,
                                   'creado_en': p.creado_en.strftime('%d/%m/%Y') if p.creado_en else ''}
                                  for p in query_p if p.id not in usados]

            facturas_libres = []
            if not proc.factura_id and proc.tipo == 'drogueria':
                query_f = session.query(Invoice).filter(
                    Invoice.proveedor_razon.ilike(f'%{proc.partner_nombre}%')
                ).order_by(Invoice.fecha.desc()).limit(10).all()
                usados_f = {p.factura_id for p in session.query(ProcesoCompra).filter(
                    ProcesoCompra.factura_id.isnot(None)).all()}
                facturas_libres = [{'id': f.id, 'numero': f.numero_factura,
                                    'fecha': f.fecha.strftime('%d/%m/%Y') if f.fecha else '',
                                    'total': float(f.total or 0)}
                                   for f in query_f if f.id not in usados_f]

            hechos = _inferir_hecho(proc, session)

            data = {
                'id': proc.id, 'tipo': proc.tipo, 'partner_nombre': proc.partner_nombre,
                'partner_id': proc.partner_id, 'estado': proc.estado,
                'analisis_periodo': proc.analisis_periodo or '',
                'notas': proc.notas or '',
                'creado_en': proc.creado_en.strftime('%d/%m/%Y %H:%M') if proc.creado_en else '',
                'analisis_hecho_en': proc.analisis_hecho_en,
                'pedido_hecho_en': proc.pedido_hecho_en,
                'factura_hecha_en': proc.factura_hecha_en,
                'cruce_hecho_en': proc.cruce_hecho_en,
                'reclamo_hecho_en': proc.reclamo_hecho_en,
                'cerrado_en': proc.cerrado_en,
                # Flags auto-calculados (verdadero si el paso está implícitamente hecho
                # por tener el pedido/factura/etc asociado, aunque no haya timestamp)
                'hechos': hechos,
                'pedido': {
                    'id': pedido.id, 'periodo': pedido.periodo,
                    'laboratorio': pedido.laboratorio,
                    'n_items': len(pedido.items),
                    'total': float(sum(float(i.subtotal or 0) for i in pedido.items)),
                    'creado_en': pedido.creado_en.strftime('%d/%m/%Y') if pedido.creado_en else '',
                } if pedido else None,
                'invoice': {
                    'id': invoice.id, 'numero': invoice.numero_factura,
                    'fecha': invoice.fecha.strftime('%d/%m/%Y') if invoice.fecha else '',
                    'total': float(invoice.total or 0),
                    'tipo_comprobante': invoice.tipo_comprobante or 'FAC',
                    'pdf_filename': invoice.pdf_filename or '',
                    'pdf_disponible': bool(invoice.pdf_filename and os.path.exists(
                        os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads', invoice.pdf_filename)
                    )),
                } if invoice else None,
                'reclamo': {
                    'id': reclamo.id, 'numero_factura': reclamo.numero_factura,
                    'estado': reclamo.estado,
                    'fecha': reclamo.fecha.strftime('%d/%m/%Y') if reclamo.fecha else '',
                } if reclamo else None,
                'analisis_pasos': analisis_pasos,
                'analisis_sesion': {
                    'id': sesion.id,
                    'laboratorio': sesion.laboratorio_nombre,
                    'periodo': sesion.periodo or '',
                    'farmacia': sesion.farmacia or '',
                    'n_days': sesion.n_days,
                    'fuente': sesion.fuente or 'pdf',
                    'n_productos': sesion.n_productos,
                    'creado_en': sesion.creado_en.strftime('%d/%m/%Y %H:%M') if sesion.creado_en else '',
                } if sesion else None,
            }

        return render_template('proceso_detail.html', proc=data,
                               pedidos_libres=pedidos_libres,
                               facturas_libres=facturas_libres)

    @app.route('/proceso/<int:proceso_id>/paso/<paso>', methods=['POST'])
    def proceso_marcar_paso(proceso_id, paso):
        if paso not in PASOS:
            return jsonify({'error': 'Paso inválido'}), 400
        with database.get_db() as session:
            proc = session.get(ProcesoCompra, proceso_id)
            if not proc:
                return jsonify({'error': 'No encontrado'}), 404
            field = _paso_field(paso)
            setattr(proc, field, now_ar())
            _bump_estado(proc, ESTADO_POR_PASO[paso])
            proc.actualizado_en = now_ar()
            session.commit()
        return redirect(url_for('proceso_detail', proceso_id=proceso_id))

    @app.route('/proceso/<int:proceso_id>/paso/<paso>/undo', methods=['POST'])
    def proceso_desmarcar_paso(proceso_id, paso):
        if paso not in PASOS:
            return jsonify({'error': 'Paso inválido'}), 400
        with database.get_db() as session:
            proc = session.get(ProcesoCompra, proceso_id)
            if not proc:
                return jsonify({'error': 'No encontrado'}), 404
            setattr(proc, _paso_field(paso), None)
            # Recalcular estado
            if proc.cruce_hecho_en:
                proc.estado = 'INGRESADO'
            elif proc.factura_hecha_en:
                proc.estado = 'FACTURADO'
            elif proc.pedido_hecho_en:
                proc.estado = 'PEDIDO'
            else:
                proc.estado = 'BORRADOR'
            proc.actualizado_en = now_ar()
            session.commit()
        return redirect(url_for('proceso_detail', proceso_id=proceso_id))

    @app.route('/proceso/<int:proceso_id>/link-pedido', methods=['POST'])
    def proceso_link_pedido(proceso_id):
        pedido_id = request.form.get('pedido_id')
        with database.get_db() as session:
            proc = session.get(ProcesoCompra, proceso_id)
            if not proc:
                flash('Proceso no encontrado.')
                return redirect(url_for('procesos_list'))
            if pedido_id:
                proc.pedido_id = int(pedido_id)
                if not proc.pedido_hecho_en:
                    proc.pedido_hecho_en = now_ar()
                _bump_estado(proc, 'PEDIDO')
                pedido = session.get(Pedido, int(pedido_id))
                if pedido and pedido.analisis_sesion_id and not proc.analisis_sesion_id:
                    proc.analisis_sesion_id = pedido.analisis_sesion_id
            else:
                proc.pedido_id = None
            proc.actualizado_en = now_ar()
            session.commit()
        return redirect(url_for('proceso_detail', proceso_id=proceso_id))

    @app.route('/proceso/<int:proceso_id>/link-factura', methods=['POST'])
    def proceso_link_factura(proceso_id):
        factura_id = request.form.get('factura_id')
        with database.get_db() as session:
            proc = session.get(ProcesoCompra, proceso_id)
            if not proc:
                flash('Proceso no encontrado.')
                return redirect(url_for('procesos_list'))
            if factura_id:
                proc.factura_id = int(factura_id)
                if not proc.factura_hecha_en:
                    proc.factura_hecha_en = now_ar()
                _bump_estado(proc, 'FACTURADO')
            else:
                proc.factura_id = None
            proc.actualizado_en = now_ar()
            session.commit()
        return redirect(url_for('proceso_detail', proceso_id=proceso_id))

    @app.route('/proceso/<int:proceso_id>/snapshot-analisis', methods=['POST'])
    def proceso_snapshot_analisis(proceso_id):
        """Guarda qué sub-pasos del análisis se usaron. Body JSON:
        {modulos:{hecho:true, cant:N}, ofertas:{...}, ofertas_min:{...}, sin_deal:{...}}
        """
        body = request.get_json(silent=True) or {}
        with database.get_db() as session:
            proc = session.get(ProcesoCompra, proceso_id)
            if not proc:
                return jsonify({'error': 'No encontrado'}), 404
            proc.analisis_pasos_json = json.dumps(body)
            if not proc.analisis_hecho_en:
                proc.analisis_hecho_en = now_ar()
            _bump_estado(proc, 'BORRADOR')  # no sube estado, solo marca
            proc.actualizado_en = now_ar()
            session.commit()
            return jsonify({'ok': True})

    @app.route('/proceso/<int:proceso_id>/notas', methods=['POST'])
    def proceso_notas(proceso_id):
        notas = (request.form.get('notas') or '').strip()
        with database.get_db() as session:
            proc = session.get(ProcesoCompra, proceso_id)
            if not proc:
                flash('Proceso no encontrado.')
                return redirect(url_for('procesos_list'))
            proc.notas = notas or None
            proc.actualizado_en = now_ar()
            session.commit()
        return redirect(url_for('proceso_detail', proceso_id=proceso_id))

    @app.route('/proceso/<int:proceso_id>/cerrar', methods=['POST'])
    def proceso_cerrar(proceso_id):
        with database.get_db() as session:
            proc = session.get(ProcesoCompra, proceso_id)
            if not proc:
                flash('Proceso no encontrado.')
                return redirect(url_for('procesos_list'))
            proc.cerrado_en = now_ar()
            proc.estado = 'CERRADO'
            proc.actualizado_en = now_ar()
            session.commit()
        return redirect(url_for('proceso_detail', proceso_id=proceso_id))

    @app.route('/proceso/<int:proceso_id>/reabrir', methods=['POST'])
    def proceso_reabrir(proceso_id):
        with database.get_db() as session:
            proc = session.get(ProcesoCompra, proceso_id)
            if not proc:
                flash('Proceso no encontrado.')
                return redirect(url_for('procesos_list'))
            proc.cerrado_en = None
            if proc.cruce_hecho_en:
                proc.estado = 'INGRESADO'
            elif proc.factura_hecha_en:
                proc.estado = 'FACTURADO'
            elif proc.pedido_hecho_en:
                proc.estado = 'PEDIDO'
            else:
                proc.estado = 'BORRADOR'
            proc.actualizado_en = now_ar()
            session.commit()
        return redirect(url_for('proceso_detail', proceso_id=proceso_id))

    @app.route('/pedido/<int:pedido_id>/enviar-a-proceso', methods=['POST'])
    def pedido_enviar_a_proceso(pedido_id):
        """Crea un ProcesoCompra desde un Pedido guardado.
        Si el pedido ya tiene canal+partner_id decidido (desde el paso 4 del análisis),
        los usa directamente. Si no, acepta canal y drogueria_id del form.
        """
        with database.get_db() as session:
            pedido = session.get(Pedido, pedido_id)
            if not pedido:
                flash('Pedido no encontrado.')
                return redirect(url_for('orders_list'))

            existente = session.query(ProcesoCompra).filter_by(pedido_id=pedido.id).first()
            if existente:
                flash(f'Este pedido ya está en el proceso #{existente.id}.')
                return redirect(url_for('proceso_detail', proceso_id=existente.id))

            # Preferir lo que el pedido ya tiene guardado (paso 4 del análisis)
            if pedido.canal:
                canal = pedido.canal
                drog_id = pedido.partner_id if canal == 'drogueria' else None
            else:
                # Fallback al form tradicional (orders_list)
                canal = (request.form.get('canal') or 'laboratorio').strip()
                drog_id = request.form.get('drogueria_id') or None

            if canal == 'drogueria':
                if not drog_id:
                    flash('Elegí la droguería por la que va a entrar el pedido.')
                    return redirect(url_for('orders_list'))
                prov = session.get(Provider, int(drog_id))
                if not prov:
                    flash('Droguería no encontrada.')
                    return redirect(url_for('orders_list'))
                partner_nombre = prov.razon_social
                partner_id = prov.id
                tipo = 'drogueria'
                notas = f'Pedido generado para laboratorio: {pedido.laboratorio}'
            else:
                lab = session.query(Laboratorio).filter_by(nombre=pedido.laboratorio).first()
                partner_nombre = pedido.laboratorio or '—'
                partner_id = lab.id if lab else None
                tipo = 'laboratorio'
                notas = None

            # Si el pedido no tenía canal persistido, lo guardamos ahora para congruencia
            if not pedido.canal:
                pedido.canal = canal
                pedido.partner_id = partner_id
                pedido.canal_elegido_en = now_ar()

            proc = ProcesoCompra(
                tipo=tipo,
                partner_id=partner_id,
                partner_nombre=partner_nombre,
                analisis_periodo=pedido.periodo or None,
                pedido_id=pedido.id,
                analisis_sesion_id=pedido.analisis_sesion_id,
                analisis_hecho_en=pedido.creado_en or now_ar(),
                estado='BORRADOR',
                notas=notas,
            )
            session.add(proc)
            pedido.estado = 'ENVIADO'
            session.commit()
            pid = proc.id
        flash('Pedido enviado a Procesos.')
        return redirect(url_for('proceso_detail', proceso_id=pid))

    @app.route('/proceso/<int:proceso_id>/delete', methods=['POST'])
    def proceso_delete(proceso_id):
        with database.get_db() as session:
            proc = session.get(ProcesoCompra, proceso_id)
            if proc:
                session.delete(proc)
                session.commit()
        return redirect(url_for('procesos_list'))
