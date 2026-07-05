"""Mis Informes — pantallas de cruce/análisis sobre el catálogo y las ventas.

Cada informe vive en su propia ruta + template. La pantalla `/informes` es
el índice con tarjetas para cada informe. Pensado para crecer agregando
más cruces sin romper la organización.

Informes implementados:
1. Labs por droga — "¿Qué labs fabrican esta droga y cuál vendo más?"

Pendientes (próximas iteraciones):
2. Drogas con un solo proveedor — alerta de dependencia.
4. Presentaciones por droga — qué tamaños venden más.
"""
import json
from datetime import date

from flask import jsonify, render_template, request
from flask_login import login_required
from sqlalchemy import distinct, func

import database
from database import ObsLaboratorio, ObsNombreDroga, ObsProducto, ObsStock, ObsVentaMensual, Producto


def _ventana_12m():
    """Devuelve (desde_ym, hasta_ym) como ints YYYYMM para los últimos 12 meses."""
    hoy = date.today()
    hasta = hoy.year * 100 + hoy.month
    desde_y = hoy.year - 1
    desde_m = hoy.month + 1
    if desde_m > 12:
        desde_m -= 12
        desde_y += 1
    desde = desde_y * 100 + desde_m
    return desde, hasta


def _leer_snapshot_cadencias(session):
    """Lee el snapshot de cadencias (1 fila por lab) ya ordenado por
    monto_mensual desc. Devuelve (filas, totales, meta).

    Una sola fuente de verdad compartida por la pantalla
    `/informes/cadencias-resumen` y el análisis IA `/.../analizar-ia`.
    """
    from database import CadenciaLabSnapshot
    int_keys = ('core', 'ocasional', 'caida', 'dormido', 'alta', 'media_alta',
                'media', 'baja', 'muy_baja', 'con_ventas', 'sin_ventas',
                'dormido_con_stock', 'dormido_stock_u')
    float_keys = ('monto_mensual', 'dormido_valor',
                  'core_monto', 'ocasional_monto', 'caida_monto', 'dormido_monto',
                  'alta_monto', 'media_alta_monto', 'media_monto', 'baja_monto',
                  'muy_baja_monto')
    keys = int_keys + float_keys
    tot = {k: 0 for k in keys}
    filas = []
    meta = None
    rows = (session.query(CadenciaLabSnapshot)
            .order_by(CadenciaLabSnapshot.monto_mensual.desc()).all())
    for r in rows:
        fila = {'lab_id': r.lab_id, 'nombre': r.lab_nombre or str(r.lab_id)}
        for k in keys:
            v = getattr(r, k) or 0
            v = float(v) if k in float_keys else int(v)
            fila[k] = v
            tot[k] += v
        filas.append(fila)
    if rows:
        meta = {
            'actualizado_en': max((r.actualizado_en for r in rows if r.actualizado_en),
                                  default=None),
            'cobertura': rows[0].cobertura,
            'meses_rot': rows[0].meses_rot,
            'n_labs': len(rows),
        }
    return filas, tot, meta


def _guardar_analisis_cache(clave, titulo, texto, usage=None):
    """Upsert del último análisis IA por clave (para re-mostrarlo sin gastar API).

    Defensivo: si el guardado falla, no rompe la respuesta del análisis.
    """
    try:
        ti = getattr(usage, 'input_tokens', None) if usage else None
        to = getattr(usage, 'output_tokens', None) if usage else None
        with database.get_db() as s:
            row = s.get(database.AnalisisIaCache, clave)
            if row:
                row.titulo, row.texto = titulo, texto
                row.tokens_in, row.tokens_out = ti, to
                row.creado_en = database.now_ar()
            else:
                s.add(database.AnalisisIaCache(
                    clave=clave, titulo=titulo, texto=texto,
                    tokens_in=ti, tokens_out=to))
            s.commit()
    except Exception:
        import logging
        logging.getLogger(__name__).warning('no pude cachear análisis %s', clave, exc_info=True)


def _friendly_api_error(e):
    """Mapea una excepción de la API de Claude a un JSON 502 con mensaje claro."""
    msg = str(e); low = msg.lower()
    if 'credit balance' in low or 'insufficient' in low:
        friendly = 'Sin crédito en la cuenta de Anthropic. Cargá crédito y reintentá.'
    elif 'invalid' in low and 'api' in low and 'key' in low:
        friendly = 'La ANTHROPIC_API_KEY es inválida o fue revocada.'
    elif 'rate limit' in low or '429' in low:
        friendly = 'Rate limit alcanzado. Esperá unos segundos y reintentá.'
    else:
        friendly = f'Claude API: {msg}'
    return jsonify({'ok': False, 'error': friendly}), 502


def init_app(app):

    @app.route('/informes')
    @login_required
    def informes_index():
        """Índice con tarjetas para cada informe disponible."""
        return render_template('informes_index.html')

    @app.route('/informes/eventos-sla')
    @login_required
    def informes_eventos_sla():
        """Log de eventos SLA (cadetes, droguería, demoras, sin respuesta).
        Solo lectura. Filtros: tipo, severidad, resuelto/pendiente."""
        from sqlalchemy import desc
        tipo = request.args.get('tipo') or ''
        sev = request.args.get('sev') or ''
        estado = request.args.get('estado') or 'pendientes'  # pendientes | resueltos | todos
        with database.get_db() as s:
            E = database.EventoSLA
            q = s.query(E)
            if tipo:
                q = q.filter(E.tipo == tipo)
            if sev:
                q = q.filter(E.severidad == sev)
            if estado == 'pendientes':
                q = q.filter(E.resuelto_en.is_(None))
            elif estado == 'resueltos':
                q = q.filter(E.resuelto_en.isnot(None))
            eventos = q.order_by(desc(E.id)).limit(500).all()
            # Resumen contadores por tipo (todos, sin filtros)
            from sqlalchemy import func
            por_tipo = (s.query(E.tipo, func.count(E.id))
                        .filter(E.resuelto_en.is_(None))
                        .group_by(E.tipo).all())
        return render_template('eventos_sla.html',
                               eventos=eventos, por_tipo=por_tipo,
                               tipo=tipo, sev=sev, estado=estado)

    @app.route('/informes/ventas-vendedor')
    @login_required
    def informes_ventas_vendedor():
        """Ventas por vendedor/operador (ObServer DW.ProductosVendidos.IdOperador)."""
        import calendar
        from datetime import date as _date

        from services.comparativa_ventas import meses_disponibles
        from services.ventas_vendedor import ventas_por_vendedor
        meses = meses_disponibles()
        mes = request.args.get('mes', type=int) or meses[0]['key']
        anio, m = mes // 100, mes % 100
        if not (1 <= m <= 12):
            mes, anio, m = meses[0]['key'], meses[0]['key'] // 100, meses[0]['key'] % 100
        d1 = _date(anio, m, 1)
        d2 = _date(anio, m, calendar.monthrange(anio, m)[1])
        with database.get_db() as session:
            data = ventas_por_vendedor(session, d1, d2)
        tot_imp = sum(x['importe'] for x in data)
        tot_u = sum(x['unidades'] for x in data)
        return render_template('informes_ventas_vendedor.html',
                               data=data, meses=meses, mes_sel=mes,
                               tot_imp=tot_imp, tot_u=tot_u)

    @app.route('/informes/cadencias-lab')
    @login_required
    def informe_cadencias_lab():
        """Agrupa los productos de un laboratorio según su cadencia natural
        de compra (cuánto dura una reposición típica al ritmo actual).

        Buckets: alta rotación (≤10d) · media (10-20d) · baja (20-30d) ·
        muy baja (30-60d) · ocasional (>60d).

        Sin lab seleccionado: solo renderiza la pantalla con el selector.
        """
        from helpers import analizar_cadencias_lab
        lab_id = request.args.get('lab_id', type=int)
        cobertura = request.args.get('cobertura', type=int) or 30
        cobertura = max(7, min(cobertura, 90))
        meses_rot = request.args.get('meses_rot', type=int) or 3
        meses_rot = max(1, min(meses_rot, 12))
        data = None
        lab_nombre = None
        with database.get_db() as session:
            # Listar labs con al menos 1 producto activo (para el dropdown).
            labs = (session.query(ObsLaboratorio.observer_id,
                                  ObsLaboratorio.descripcion)
                    .order_by(ObsLaboratorio.descripcion).all())
            if lab_id:
                lab = session.get(ObsLaboratorio, lab_id)
                lab_nombre = lab.descripcion if lab else None
                data = analizar_cadencias_lab(
                    session, lab_id,
                    meses_rotacion=meses_rot,
                    cobertura_default=cobertura,
                )
        return render_template('informes_cadencias_lab.html',
                               labs=[{'id': r[0], 'nombre': r[1]} for r in labs],
                               lab_id=lab_id, lab_nombre=lab_nombre,
                               data=data, cobertura=cobertura,
                               meses_rot=meses_rot)

    @app.route('/informes/cadencias-resumen')
    @login_required
    def informe_cadencias_resumen():
        """Plataforma cross-lab: lee el snapshot materializado (1 fila por lab)
        y lo renderiza para filtrar/ordenar client-side. El cálculo pesado se
        hace en /recalcular (todos los labs de una). Drill-down por lab en vivo."""
        with database.get_db() as session:
            filas, tot, meta = _leer_snapshot_cadencias(session)
        return render_template('informes_cadencias_resumen.html',
                               filas=filas, totales=tot, meta=meta)

    @app.route('/informes/cadencias-resumen/recalcular', methods=['POST'])
    @login_required
    def informe_cadencias_resumen_recalcular():
        """Materializa el análisis de cadencias para TODOS los labs con ventas
        y reemplaza el snapshot. ~25s. Params cobertura/meses_rot baked in."""
        import time

        from helpers import recalcular_snapshot_cadencias
        data = request.get_json(silent=True) or {}
        cobertura = max(7, min(int(data.get('cobertura') or 30), 90))
        meses_rot = max(1, min(int(data.get('meses_rot') or 3), 12))
        t0 = time.time()
        with database.get_db() as session:
            n = recalcular_snapshot_cadencias(session, cobertura, meses_rot)
        return jsonify({'ok': True, 'filas': n,
                        'segundos': round(time.time() - t0, 1)})

    @app.route('/informes/cadencias-resumen/analizar', methods=['POST'])
    @login_required
    def informe_cadencias_resumen_analizar():
        """Manda el top N labs del snapshot a Claude y devuelve un análisis en
        prosa (para mostrar en un modal). On-demand: consume crédito de la API."""
        import os
        api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
        if not api_key:
            return jsonify({'ok': False, 'error': 'Falta ANTHROPIC_API_KEY en el servidor.'}), 400
        body = request.get_json(silent=True) or {}
        try:
            top_n = max(1, min(int(body.get('top') or 10), 30))
        except (TypeError, ValueError):
            top_n = 10
        vista = 'monto' if (body.get('vista') == 'monto') else 'cantidad'
        with database.get_db() as session:
            filas, _tot, meta = _leer_snapshot_cadencias(session)
        if not filas:
            return jsonify({'ok': False, 'error': 'No hay snapshot. Tocá "Recalcular" primero.'}), 400
        from services import cadencias_analisis
        try:
            texto, usage = cadencias_analisis.analizar_cadencias(
                filas[:top_n], meta, api_key, vista=vista)
            _guardar_analisis_cache('cadencias', 'Análisis de cadencias por laboratorio', texto, usage)
        except ImportError:
            return jsonify({'ok': False, 'error': 'Paquete anthropic no instalado. Reiniciá el container.'}), 500
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 400
        except Exception as e:
            msg = str(e); low = msg.lower()
            if 'credit balance' in low or 'insufficient' in low:
                friendly = 'Sin crédito en la cuenta de Anthropic. Cargá crédito y reintentá.'
            elif 'invalid' in low and 'api' in low and 'key' in low:
                friendly = 'La ANTHROPIC_API_KEY es inválida o fue revocada.'
            elif 'rate limit' in low or '429' in low:
                friendly = 'Rate limit alcanzado. Esperá unos segundos y reintentá.'
            else:
                friendly = f'Claude API: {msg}'
            return jsonify({'ok': False, 'error': friendly}), 502
        return jsonify({'ok': True, 'analisis': texto, 'top': top_n,
                        'tokens_in': getattr(usage, 'input_tokens', None),
                        'tokens_out': getattr(usage, 'output_tokens', None)})

    @app.route('/informes/analisis-ia/ultimo')
    @login_required
    def informe_analisis_ia_ultimo():
        """Devuelve el último análisis IA cacheado para una clave (sin llamar API).

        Para "Ver último análisis" en demos: cero gasto, cero riesgo de fallo.
        """
        clave = (request.args.get('clave') or '').strip()
        if not clave:
            return jsonify({'ok': False, 'error': 'clave requerida'}), 400
        with database.get_db() as s:
            row = s.get(database.AnalisisIaCache, clave)
        if not row:
            return jsonify({'ok': False, 'error': 'Todavía no generaste un análisis para esto.'}), 404
        return jsonify({'ok': True, 'analisis': row.texto, 'titulo': row.titulo,
                        'tokens_in': row.tokens_in, 'tokens_out': row.tokens_out,
                        'creado_en': row.creado_en.strftime('%d/%m/%Y %H:%M') if row.creado_en else None})

    # ── Comparación portfolio líder de un lab vs ventas propias ──
    # Datasets de referencia (IQVIA/IMS) en referencia_mercado.py. Por ahora
    # solo Roemmers (152). El selector lista los labs con dataset cargado.
    def _ctx_referencia(funcion_analisis, template):
        """Helper común a los 3 informes: resuelve lab, corre el análisis."""
        import referencia_mercado
        lab_id = request.args.get('lab_id', type=int)
        labs_ref = [{'id': lid, 'nombre': nom}
                    for lid, nom in referencia_mercado.labs_con_referencia()]
        # Default: si hay un solo lab con referencia, lo pre-selecciona.
        if not lab_id and len(labs_ref) == 1:
            lab_id = labs_ref[0]['id']
        data = None
        if lab_id:
            with database.get_db() as session:
                data = funcion_analisis(session, lab_id)
        return render_template(template, labs_ref=labs_ref, lab_id=lab_id, data=data)

    # ── Gap de marcas (8 labs, web search) ──
    # Flujo en 2 pasos: (1) /recopilar → Claude+web search trae marcas+fuentes
    # (cacheado por nombre normalizado del lab); el front muestra un modal con
    # esos datos; (2) /analizar → cruza esas marcas vs ventas propias + prosa.
    _GAP_WS_TTL_DIAS = 90  # los datos de mercado cacheados valen 90 días; después se re-buscan

    def _clave_ws_data(nombre_lab):
        from helpers import _normalizar_nombre_entidad
        return f'gap_ws_data:{_normalizar_nombre_entidad(nombre_lab)}'[:80]

    @app.route('/informes/lab-gap-marcas')
    @login_required
    def informe_lab_gap_marcas():
        """Informe — Gap de captura por marca estrella. Dropdown de 8 labs; el
        análisis es on-demand por POST (no se corre en el GET)."""
        import referencia_mercado
        with database.get_db() as session:
            disponibles = referencia_mercado.labs_gap_disponibles(session)
        labs_ref = [{'id': d['observer_id'], 'nombre': d['nombre']} for d in disponibles]
        lab_id = request.args.get('lab_id', type=int)
        return render_template('informes_lab_gap_marcas.html',
                               labs_ref=labs_ref, lab_id=lab_id, data=None)

    @app.route('/informes/lab-gap-marcas/recopilar', methods=['POST'])
    @login_required
    def informe_lab_gap_marcas_recopilar():
        """PASO 1 — recopila marcas estrella del lab vía web search (cacheado)."""
        import os
        api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
        if not api_key:
            return jsonify({'ok': False, 'error': 'Falta ANTHROPIC_API_KEY en el servidor.'}), 400
        body = request.get_json(silent=True) or {}
        try:
            lab_id = int(body.get('lab_id'))
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'lab_id inválido.'}), 400
        forzar = bool(body.get('forzar'))
        with database.get_db() as session:
            lab = session.get(ObsLaboratorio, lab_id)
            if not lab:
                return jsonify({'ok': False, 'error': 'Laboratorio inexistente.'}), 404
            nombre = lab.descripcion
        clave = _clave_ws_data(nombre)
        if not forzar:
            with database.get_db() as s:
                row = s.get(database.AnalisisIaCache, clave)
            if row and row.texto:
                edad = (database.now_ar() - row.creado_en).days if row.creado_en else None
                vencido = edad is not None and edad > _GAP_WS_TTL_DIAS
                if not vencido:
                    try:
                        cached = json.loads(row.texto)
                        return jsonify({'ok': True, 'cacheado': True, 'nombre_lab': nombre,
                                        'marcas': cached.get('marcas', []),
                                        'fuentes': cached.get('fuentes', []),
                                        'fecha': row.creado_en.strftime('%d/%m/%Y') if row.creado_en else None,
                                        'edad_dias': edad})
                    except (ValueError, TypeError):
                        pass
                # vencido → cae al web search de abajo (datos frescos)
        from services import referencia_websearch
        try:
            marcas, fuentes, usage = referencia_websearch.recopilar_marcas_estrella(nombre, api_key)
        except ImportError:
            return jsonify({'ok': False, 'error': 'Paquete anthropic no instalado. Reiniciá el container.'}), 500
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 400
        except Exception as e:
            return _friendly_api_error(e)
        _guardar_analisis_cache(clave, f'Marcas {nombre} (web)',
                                json.dumps({'marcas': marcas, 'fuentes': fuentes}), usage)
        return jsonify({'ok': True, 'cacheado': False, 'nombre_lab': nombre,
                        'marcas': marcas, 'fuentes': fuentes,
                        'fecha': database.now_ar().strftime('%d/%m/%Y'), 'edad_dias': 0,
                        'tokens_in': getattr(usage, 'input_tokens', None),
                        'tokens_out': getattr(usage, 'output_tokens', None)})

    @app.route('/informes/lab-gap-marcas/analizar', methods=['POST'])
    @login_required
    def informe_lab_gap_marcas_analizar():
        """PASO 2 — cruza las marcas recopiladas vs ventas propias + prosa Claude."""
        import os
        api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
        if not api_key:
            return jsonify({'ok': False, 'error': 'Falta ANTHROPIC_API_KEY en el servidor.'}), 400
        body = request.get_json(silent=True) or {}
        try:
            lab_id = int(body.get('lab_id'))
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'lab_id inválido.'}), 400
        from helpers import cruzar_marcas_vs_ventas
        with database.get_db() as session:
            lab = session.get(ObsLaboratorio, lab_id)
            if not lab:
                return jsonify({'ok': False, 'error': 'Laboratorio inexistente.'}), 404
            nombre = lab.descripcion
            row = session.get(database.AnalisisIaCache, _clave_ws_data(nombre))
            if not row or not row.texto:
                return jsonify({'ok': False, 'error': 'Primero recopilá los datos del lab.'}), 400
            try:
                cached = json.loads(row.texto)
            except (ValueError, TypeError):
                return jsonify({'ok': False, 'error': 'Datos recopilados corruptos. Re-buscá.'}), 400
            marcas = cached.get('marcas', [])
            fuentes = cached.get('fuentes', [])
            if not marcas:
                return jsonify({'ok': False, 'error': 'No hay marcas recopiladas para este lab.'}), 400
            data = cruzar_marcas_vs_ventas(session, lab_id, marcas, nombre)
        from services import referencia_ia
        try:
            texto, usage = referencia_ia.analizar_gap_marcas(data, api_key, model='claude-sonnet-4-6')
            _guardar_analisis_cache(f'lab_gap_marcas:{lab_id}',
                                    f'Gap de captura — {nombre}', texto, usage)
        except ImportError:
            return jsonify({'ok': False, 'error': 'Paquete anthropic no instalado. Reiniciá el container.'}), 500
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 400
        except Exception as e:
            return _friendly_api_error(e)
        return jsonify({'ok': True, 'analisis': texto, 'data': data, 'fuentes': fuentes,
                        'tokens_in': getattr(usage, 'input_tokens', None),
                        'tokens_out': getattr(usage, 'output_tokens', None)})

    @app.route('/informes/comparativa-drogas')
    @login_required
    def informe_comparativa_drogas():
        """Comparativa por droga: cruza la inteligencia de mercado (marcas del
        web search cacheadas por lab) con las ventas propias. Panel de labs +
        última consulta arriba; multiselect (≤3) para refrescar vía web search."""
        from services import mercado_drogas
        with database.get_db() as session:
            labs_estado = mercado_drogas.labs_cacheados_estado(session)
            comparativa = mercado_drogas.comparativa_mercado_por_droga(session)
            fuentes_lab = mercado_drogas.fuentes_por_lab(session)
        return render_template('comparativa_drogas.html',
                               labs_estado=labs_estado, comparativa=comparativa,
                               fuentes_lab=fuentes_lab, ttl_dias=_GAP_WS_TTL_DIAS)

    @app.route('/informes/lab-ranking-nacional')
    @login_required
    def informe_lab_ranking_nacional():
        """Informe — Mi ranking del lab vs marcas estrella nacionales."""
        from helpers import analizar_ranking_vs_nacional
        return _ctx_referencia(analizar_ranking_vs_nacional, 'informes_lab_ranking_nacional.html')

    @app.route('/informes/lab-ranking-nacional/analizar', methods=['POST'])
    @login_required
    def informe_lab_ranking_nacional_analizar():
        """Análisis IA del ranking propio del lab vs marcas estrella nacionales."""
        import os

        from helpers import analizar_ranking_vs_nacional
        api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
        if not api_key:
            return jsonify({'ok': False, 'error': 'Falta ANTHROPIC_API_KEY en el servidor.'}), 400
        body = request.get_json(silent=True) or {}
        try:
            lab_id = int(body.get('lab_id'))
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'lab_id inválido.'}), 400
        with database.get_db() as session:
            data = analizar_ranking_vs_nacional(session, lab_id)
        if not data:
            return jsonify({'ok': False, 'error': 'No hay dataset de referencia para ese lab.'}), 400
        from services import referencia_ia
        try:
            texto, usage = referencia_ia.analizar_ranking_vs_nacional(data, api_key)
            _guardar_analisis_cache(f'lab_ranking_nacional:{lab_id}',
                                    f'Ranking vs nacional — {data.get("nombre_lab") or ""}', texto, usage)
        except ImportError:
            return jsonify({'ok': False, 'error': 'Paquete anthropic no instalado. Reiniciá el container.'}), 500
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 400
        except Exception as e:
            msg = str(e); low = msg.lower()
            if 'credit balance' in low or 'insufficient' in low:
                friendly = 'Sin crédito en la cuenta de Anthropic. Cargá crédito y reintentá.'
            elif 'invalid' in low and 'api' in low and 'key' in low:
                friendly = 'La ANTHROPIC_API_KEY es inválida o fue revocada.'
            elif 'rate limit' in low or '429' in low:
                friendly = 'Rate limit alcanzado. Esperá unos segundos y reintentá.'
            else:
                friendly = f'Claude API: {msg}'
            return jsonify({'ok': False, 'error': friendly}), 502
        return jsonify({'ok': True, 'analisis': texto,
                        'tokens_in': getattr(usage, 'input_tokens', None),
                        'tokens_out': getattr(usage, 'output_tokens', None)})

    @app.route('/informes/lab-cobertura-moleculas')
    @login_required
    def informe_lab_cobertura_moleculas():
        """Informe — Cobertura de moléculas líderes nacionales."""
        from helpers import analizar_cobertura_moleculas
        return _ctx_referencia(analizar_cobertura_moleculas, 'informes_lab_cobertura_moleculas.html')

    @app.route('/informes/lab-cobertura-moleculas/analizar', methods=['POST'])
    @login_required
    def informe_lab_cobertura_moleculas_analizar():
        """Análisis IA de la cobertura de moléculas líderes de un lab."""
        import os

        from helpers import analizar_cobertura_moleculas
        api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
        if not api_key:
            return jsonify({'ok': False, 'error': 'Falta ANTHROPIC_API_KEY en el servidor.'}), 400
        body = request.get_json(silent=True) or {}
        try:
            lab_id = int(body.get('lab_id'))
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'lab_id inválido.'}), 400
        with database.get_db() as session:
            data = analizar_cobertura_moleculas(session, lab_id)
        if not data:
            return jsonify({'ok': False, 'error': 'No hay dataset de referencia para ese lab.'}), 400
        from services import referencia_ia
        try:
            texto, usage = referencia_ia.analizar_cobertura_moleculas(data, api_key)
            _guardar_analisis_cache(f'lab_cobertura_moleculas:{lab_id}',
                                    f'Cobertura de moléculas — {data.get("nombre_lab") or ""}', texto, usage)
        except ImportError:
            return jsonify({'ok': False, 'error': 'Paquete anthropic no instalado. Reiniciá el container.'}), 500
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 400
        except Exception as e:
            msg = str(e); low = msg.lower()
            if 'credit balance' in low or 'insufficient' in low:
                friendly = 'Sin crédito en la cuenta de Anthropic. Cargá crédito y reintentá.'
            elif 'invalid' in low and 'api' in low and 'key' in low:
                friendly = 'La ANTHROPIC_API_KEY es inválida o fue revocada.'
            elif 'rate limit' in low or '429' in low:
                friendly = 'Rate limit alcanzado. Esperá unos segundos y reintentá.'
            else:
                friendly = f'Claude API: {msg}'
            return jsonify({'ok': False, 'error': friendly}), 502
        return jsonify({'ok': True, 'analisis': texto,
                        'tokens_in': getattr(usage, 'input_tokens', None),
                        'tokens_out': getattr(usage, 'output_tokens', None)})

    @app.route('/informes/labs-por-droga')
    @login_required
    def informe_labs_por_droga():
        """Informe #1: dada una droga, muestra todos los labs que la fabrican
        con sus productos y ventas 12m.

        Sin droga seleccionada, solo renderiza la pantalla con el buscador.
        """
        droga_id = request.args.get('droga_id', type=int)
        droga_nombre = None
        rows = []
        stats = {}

        if droga_id:
            with database.get_db() as session:
                droga = session.get(ObsNombreDroga, droga_id)
                if droga:
                    droga_nombre = droga.descripcion
                    desde, hasta = _ventana_12m()
                    # Para cada producto de la droga: lab + ventas 12m.
                    q = (session.query(
                            ObsLaboratorio.observer_id.label('lab_id'),
                            ObsLaboratorio.descripcion.label('lab_nombre'),
                            ObsProducto.observer_id.label('prod_id'),
                            ObsProducto.descripcion.label('prod_descripcion'),
                            ObsProducto.codigo_alfabeta,
                            ObsProducto.fecha_baja,
                            func.coalesce(func.sum(
                                ObsVentaMensual.unidades), 0).label('u12m'),
                            func.coalesce(func.sum(
                                ObsVentaMensual.monto), 0).label('m12m'),
                         )
                         .join(ObsLaboratorio,
                               ObsLaboratorio.observer_id == ObsProducto.laboratorio_observer)
                         .outerjoin(ObsVentaMensual,
                                    (ObsVentaMensual.producto_observer == ObsProducto.observer_id) &
                                    (ObsVentaMensual.anio * 100 + ObsVentaMensual.mes >= desde) &
                                    (ObsVentaMensual.anio * 100 + ObsVentaMensual.mes <= hasta))
                         .filter(ObsProducto.nombre_droga_observer == droga_id)
                         .group_by(
                            ObsLaboratorio.observer_id,
                            ObsLaboratorio.descripcion,
                            ObsProducto.observer_id,
                            ObsProducto.descripcion,
                            ObsProducto.codigo_alfabeta,
                            ObsProducto.fecha_baja,
                         )
                         .order_by(func.coalesce(
                            func.sum(ObsVentaMensual.unidades), 0).desc())
                    )
                    obs_prod_ids = []
                    for r in q.all():
                        rows.append({
                            'lab_id': r.lab_id,
                            'lab_nombre': r.lab_nombre,
                            'producto_id': r.prod_id,
                            'descripcion': r.prod_descripcion,
                            'codigo_alfabeta': r.codigo_alfabeta,
                            'baja': r.fecha_baja is not None,
                            'u12m': int(r.u12m or 0),
                            'm12m': float(r.m12m or 0),
                            'ean': None,    # se llena abajo si hay producto local
                        })
                        obs_prod_ids.append(r.prod_id)

                    # Mapear obs_producto → EAN del producto local (si existe).
                    # El endpoint /api/product/<ean>/chart espera EAN, no observer_id.
                    if obs_prod_ids:
                        ean_by_obs = dict(
                            session.query(Producto.observer_id, Producto.codigo_barra)
                            .filter(Producto.observer_id.in_(obs_prod_ids))
                            .all()
                        )
                        for r in rows:
                            r['ean'] = ean_by_obs.get(r['producto_id'])
                    # Agregados
                    total_u = sum(r['u12m'] for r in rows)
                    total_m = sum(r['m12m'] for r in rows)
                    labs = {r['lab_id']: r['lab_nombre'] for r in rows}
                    stats = {
                        'productos': len(rows),
                        'labs': len(labs),
                        'unidades_12m': total_u,
                        'monto_12m': total_m,
                    }
                    # Datos para gráficos
                    # 1. Donut: agregar por lab.
                    por_lab = {}
                    for r in rows:
                        por_lab.setdefault(r['lab_nombre'], 0)
                        por_lab[r['lab_nombre']] += r['u12m']
                    chart_donut = sorted(por_lab.items(), key=lambda kv: -kv[1])
                    # 2. Barras top 10 productos (por unidades).
                    top10 = sorted(rows, key=lambda r: -r['u12m'])[:10]
                    chart_top = [{
                        'label': r['descripcion'],
                        'lab': r['lab_nombre'],
                        'u12m': r['u12m'],
                    } for r in top10 if r['u12m'] > 0]

        return render_template('informes_labs_por_droga.html',
                               droga_id=droga_id,
                               droga_nombre=droga_nombre,
                               rows=rows,
                               stats=stats,
                               chart_donut=chart_donut if droga_nombre else [],
                               chart_top=chart_top if droga_nombre else [])

    @app.route('/informes/presentaciones-por-droga')
    @login_required
    def informe_presentaciones_por_droga():
        """Informe #3: para una droga elegida, agrupa los productos por
        presentación (dosis × cant por envase) y muestra qué se vende más.

        Útil para decidir qué stockear: ej. para AMOXICILINA, ¿se vende
        más x10 o x20? ¿La de 500mg o 750mg?
        """
        import re as _re

        droga_id = request.args.get('droga_id', type=int)
        droga_nombre = None
        presentaciones = []
        chart_data = None

        if droga_id:
            desde, hasta = _ventana_12m()
            with database.get_db() as session:
                droga = session.get(ObsNombreDroga, droga_id)
                if droga:
                    droga_nombre = droga.descripcion

                    # Traer todos los productos de la droga + su lab + ventas 12m.
                    q = (session.query(
                            ObsProducto.descripcion.label('desc'),
                            ObsProducto.cantidad_envase,
                            ObsLaboratorio.descripcion.label('lab'),
                            func.coalesce(func.sum(ObsVentaMensual.unidades), 0).label('u12m'),
                         )
                         .join(ObsLaboratorio,
                               ObsLaboratorio.observer_id == ObsProducto.laboratorio_observer)
                         .outerjoin(ObsVentaMensual,
                                    (ObsVentaMensual.producto_observer == ObsProducto.observer_id) &
                                    (ObsVentaMensual.anio * 100 + ObsVentaMensual.mes >= desde) &
                                    (ObsVentaMensual.anio * 100 + ObsVentaMensual.mes <= hasta))
                         .filter(ObsProducto.nombre_droga_observer == droga_id)
                         .filter(ObsProducto.fecha_baja.is_(None))
                         .group_by(
                            ObsProducto.descripcion,
                            ObsProducto.cantidad_envase,
                            ObsLaboratorio.descripcion,
                         ))

                    # Extraer "dosis mg" de la descripción + cantidad de envase.
                    # Combinarlos en una clave de presentación.
                    re_dosis = _re.compile(r'(\d+(?:[.,]\d+)?)\s*mg', _re.IGNORECASE)
                    by_pres = {}    # clave → {total, por_lab: {lab: unidades}}
                    for r in q.all():
                        m = re_dosis.search(r.desc or '')
                        dosis = m.group(1).replace(',', '.') if m else '?'
                        cant = int(r.cantidad_envase) if r.cantidad_envase else None
                        clave = f'{dosis} mg' + (f' × {cant}' if cant else '')
                        ent = by_pres.setdefault(clave, {
                            'presentacion': clave,
                            'dosis': dosis,
                            'cant_envase': cant,
                            'total_u12m': 0,
                            'por_lab': {},
                        })
                        ent['total_u12m'] += int(r.u12m or 0)
                        ent['por_lab'].setdefault(r.lab, 0)
                        ent['por_lab'][r.lab] += int(r.u12m or 0)

                    presentaciones = sorted(by_pres.values(),
                                             key=lambda p: -p['total_u12m'])

                    # Datos para el chart: barras agrupadas por presentación.
                    # Series = top labs (max 6 para no saturar). El resto se agrupa en "Otros".
                    todos_labs = {}
                    for p in presentaciones:
                        for lab, u in p['por_lab'].items():
                            todos_labs[lab] = todos_labs.get(lab, 0) + u
                    top_labs = sorted(todos_labs.items(), key=lambda kv: -kv[1])
                    series_labs = [l for l, _ in top_labs[:6]]
                    chart_data = {
                        'labels': [p['presentacion'] for p in presentaciones[:15]],
                        'series': [],
                    }
                    for lab in series_labs:
                        chart_data['series'].append({
                            'lab': lab,
                            'data': [p['por_lab'].get(lab, 0) for p in presentaciones[:15]],
                        })
                    if len(top_labs) > 6:
                        chart_data['series'].append({
                            'lab': 'Otros',
                            'data': [
                                sum(u for lab, u in p['por_lab'].items()
                                    if lab not in series_labs)
                                for p in presentaciones[:15]
                            ],
                        })

        return render_template('informes_presentaciones_por_droga.html',
                               droga_id=droga_id,
                               droga_nombre=droga_nombre,
                               presentaciones=presentaciones,
                               chart_data=chart_data)

    @app.route('/informes/drogas-sin-alternativa')
    @login_required
    def informe_drogas_sin_alternativa():
        """Informe #2: drogas críticas (con pocos labs proveedores).

        Filtra solo drogas que tuvieron ventas en los últimos 12 meses
        (las que NO vendo no son urgentes). Ordena por unidades desc para
        que las críticas-y-vendidas estén arriba.

        Query param `max_labs` (default 1): cuántos labs como máximo se
        consideran 'pocos'. Si 1, son monopolios; si 2, también pares.
        """
        try:
            max_labs = max(1, min(5, int(request.args.get('max_labs', 1))))
        except (TypeError, ValueError):
            max_labs = 1

        desde, hasta = _ventana_12m()
        rows = []
        with database.get_db() as session:
            # Subquery: drogas con ventas 12m + total unidades.
            ventas_por_droga = (
                session.query(
                    ObsProducto.nombre_droga_observer.label('droga_id'),
                    func.sum(ObsVentaMensual.unidades).label('u12m'),
                    func.sum(ObsVentaMensual.monto).label('m12m'),
                )
                .join(ObsVentaMensual,
                      ObsVentaMensual.producto_observer == ObsProducto.observer_id)
                .filter(ObsProducto.nombre_droga_observer.isnot(None))
                .filter(ObsVentaMensual.anio * 100 + ObsVentaMensual.mes >= desde)
                .filter(ObsVentaMensual.anio * 100 + ObsVentaMensual.mes <= hasta)
                .group_by(ObsProducto.nombre_droga_observer)
                .subquery()
            )

            # Para cada droga con ventas: count(distinct lab) + nombre droga.
            q = (session.query(
                    ObsNombreDroga.observer_id.label('droga_id'),
                    ObsNombreDroga.descripcion.label('droga_nombre'),
                    func.count(distinct(ObsProducto.laboratorio_observer)).label('n_labs'),
                    func.count(distinct(ObsProducto.observer_id)).label('n_productos'),
                    ventas_por_droga.c.u12m,
                    ventas_por_droga.c.m12m,
                 )
                 .join(ObsProducto,
                       ObsProducto.nombre_droga_observer == ObsNombreDroga.observer_id)
                 .join(ventas_por_droga,
                       ventas_por_droga.c.droga_id == ObsNombreDroga.observer_id)
                 .filter(ObsProducto.fecha_baja.is_(None))   # solo activos
                 .group_by(
                    ObsNombreDroga.observer_id,
                    ObsNombreDroga.descripcion,
                    ventas_por_droga.c.u12m,
                    ventas_por_droga.c.m12m,
                 )
                 .having(func.count(distinct(ObsProducto.laboratorio_observer)) <= max_labs)
                 .order_by(ventas_por_droga.c.u12m.desc())
            )

            for r in q.all():
                # Si es monopolio (1 lab), traer el nombre del único proveedor.
                lab_unico = None
                if r.n_labs == 1:
                    lab = (session.query(ObsLaboratorio.descripcion)
                           .join(ObsProducto,
                                 ObsProducto.laboratorio_observer == ObsLaboratorio.observer_id)
                           .filter(ObsProducto.nombre_droga_observer == r.droga_id)
                           .filter(ObsProducto.fecha_baja.is_(None))
                           .first())
                    if lab:
                        lab_unico = lab[0]
                rows.append({
                    'droga_id': r.droga_id,
                    'droga_nombre': r.droga_nombre,
                    'n_labs': r.n_labs,
                    'n_productos': r.n_productos,
                    'lab_unico': lab_unico,
                    'u12m': int(r.u12m or 0),
                    'm12m': float(r.m12m or 0),
                })

        # Stats agregados
        total_drogas_criticas = len(rows)
        total_u = sum(r['u12m'] for r in rows)
        total_m = sum(r['m12m'] for r in rows)
        monopolios = sum(1 for r in rows if r['n_labs'] == 1)

        return render_template('informes_drogas_sin_alternativa.html',
                               max_labs=max_labs,
                               rows=rows,
                               stats={
                                   'criticas': total_drogas_criticas,
                                   'monopolios': monopolios,
                                   'unidades_12m': total_u,
                                   'monto_12m': total_m,
                               })

    @app.route('/informes/correcciones-minimos')
    @login_required
    def informe_correcciones_minimos():
        """Productos cuyo mínimo en Observer está desfasado del calculado por
        nuestra forecast (subir/bajar). Agrupado por laboratorio. Útil para
        mandar al staff del POS para que lo actualicen manualmente.
        """
        from datetime import date as _date

        from sqlalchemy import func

        from database import (
            ObsCodigoBarras,
            ObsLaboratorio,
            ObsProducto,
            ObsRubro,
            ObsStock,
            ObsSubrubro,
            ObsVentaMensual,
        )
        from purchase_helpers import calcular_min_sugerido, clasificar_min

        lab_id_filter = request.args.get('lab_id', type=int)
        tipo_filter = (request.args.get('tipo') or 'both').lower()
        if tipo_filter not in ('up', 'down', 'both'):
            tipo_filter = 'both'
        rubros_raw = (request.args.get('rubros') or '12').strip()
        if rubros_raw.lower() == 'all' or not rubros_raw:
            rubros_filtro = None
        else:
            try:
                rubros_filtro = set(int(x) for x in rubros_raw.split(',') if x.strip())
            except ValueError:
                rubros_filtro = {12}

        with database.get_db() as session:
            stock_q = (session.query(
                ObsStock.producto_observer.label('pid'),
                func.sum(ObsStock.stock_actual).label('stock'),
                func.sum(ObsStock.minimo).label('minimo'),
            ).filter(ObsStock.minimo.isnot(None), ObsStock.minimo > 0)
              .group_by(ObsStock.producto_observer).subquery())
            v12_q = (session.query(
                ObsVentaMensual.producto_observer.label('pid'),
                func.sum(ObsVentaMensual.unidades).label('u12m'),
            ).group_by(ObsVentaMensual.producto_observer).subquery())

            base = (session.query(
                ObsProducto.observer_id.label('pid'),
                ObsProducto.descripcion.label('desc'),
                ObsProducto.subrubro_observer,
                ObsLaboratorio.observer_id.label('lab_id'),
                ObsLaboratorio.descripcion.label('lab_nombre'),
                stock_q.c.stock,
                stock_q.c.minimo,
                func.coalesce(v12_q.c.u12m, 0).label('u12m'),
            )
            .join(stock_q, stock_q.c.pid == ObsProducto.observer_id)
            .outerjoin(ObsLaboratorio,
                       ObsLaboratorio.observer_id == ObsProducto.laboratorio_observer)
            .outerjoin(v12_q, v12_q.c.pid == ObsProducto.observer_id)
            .filter(ObsProducto.fecha_baja.is_(None))
            .filter(ObsProducto.subrubro_observer.isnot(None))
            )
            # Excluir items no-medicamento (sellado recetas, cupones, etc.)
            from helpers import filtro_solo_medicamentos
            base = filtro_solo_medicamentos(base, ObsProducto).all()

            subrubro_a_rubro = dict(
                session.query(ObsSubrubro.observer_id, ObsSubrubro.rubro_observer).all())

            pids = [r.pid for r in base]
            eans = {}
            if pids:
                for cb in (session.query(ObsCodigoBarras.producto_observer,
                                          ObsCodigoBarras.codigo_barras)
                           .filter(ObsCodigoBarras.producto_observer.in_(pids),
                                   ObsCodigoBarras.fecha_baja.is_(None),
                                   ObsCodigoBarras.orden == 1).all()):
                    eans[cb.producto_observer] = cb.codigo_barras

            hoy_d = _date.today()
            end_month = hoy_d.month
            start_month = ((end_month - 11 - 1) % 12) + 1
            start_year = hoy_d.year if start_month <= end_month else hoy_d.year - 1
            ventas_por_pid = {pid: [0]*12 for pid in pids}
            if pids:
                rows_vm = (session.query(ObsVentaMensual.producto_observer,
                                          ObsVentaMensual.anio,
                                          ObsVentaMensual.mes,
                                          func.sum(ObsVentaMensual.unidades))
                           .filter(ObsVentaMensual.producto_observer.in_(pids))
                           .group_by(ObsVentaMensual.producto_observer,
                                     ObsVentaMensual.anio, ObsVentaMensual.mes)
                           .all())
                for pid_v, anio, mes, uds in rows_vm:
                    offset = (anio - start_year) * 12 + (mes - start_month)
                    if 0 <= offset <= 11 and pid_v in ventas_por_pid:
                        ventas_por_pid[pid_v][offset] += int(uds or 0)

            # Procesar
            grupos = {}  # lab_id -> {nombre, productos: []}
            for r in base:
                rub_id = subrubro_a_rubro.get(r.subrubro_observer)
                if rubros_filtro is not None and rub_id not in rubros_filtro:
                    continue
                if lab_id_filter and r.lab_id != lab_id_filter:
                    continue
                u12m = int(r.u12m or 0)
                if u12m == 0:
                    continue
                ventas_arr = ventas_por_pid.get(r.pid, [0]*12)
                min_sug, _avg_m, sin_mov, tipo_p = calcular_min_sugerido(
                    ventas_arr, int(r.stock or 0), start_month, end_month,
                )
                if sin_mov:
                    continue
                min_act = int(r.minimo or 0)
                sug = clasificar_min(min_act, min_sug)
                if sug == 'ok':
                    continue  # OK, no sugerencia
                if tipo_filter == 'up' and sug != 'up':
                    continue
                if tipo_filter == 'down' and sug != 'down':
                    continue
                lab_key = r.lab_id or 0
                if lab_key not in grupos:
                    grupos[lab_key] = {
                        'lab_id': r.lab_id,
                        'lab_nombre': r.lab_nombre or '— sin lab —',
                        'productos': [],
                    }
                grupos[lab_key]['productos'].append({
                    'pid': r.pid,
                    'desc': r.desc,
                    'ean': eans.get(r.pid, ''),
                    'sugerencia': sug,
                    'min_actual': min_act,
                    'min_sugerido': min_sug,
                    'diferencia': min_sug - min_act,
                    'stock_actual': int(r.stock or 0),
                    'u12m': u12m,
                    'tipo': tipo_p,
                })

            # Sort productos dentro de cada lab por urgencia (subir primero, dif desc)
            for g in grupos.values():
                g['productos'].sort(key=lambda x: (
                    0 if x['sugerencia'] == 'up' else 1,
                    -abs(x['diferencia']),
                ))
                g['n_up'] = sum(1 for p in g['productos'] if p['sugerencia'] == 'up')
                g['n_down'] = sum(1 for p in g['productos'] if p['sugerencia'] == 'down')
            # Ordenar grupos por cantidad de productos descendente
            grupos_list = sorted(grupos.values(),
                                  key=lambda g: -len(g['productos']))

            # Para el dropdown del filtro de lab
            labs_disponibles = (session.query(ObsLaboratorio.observer_id,
                                               ObsLaboratorio.descripcion)
                                .filter(ObsLaboratorio.fecha_baja.is_(None))
                                .order_by(ObsLaboratorio.descripcion).all())

            return render_template(
                'informe_correcciones_minimos.html',
                grupos=grupos_list,
                total=sum(len(g['productos']) for g in grupos_list),
                lab_id_filter=lab_id_filter,
                tipo_filter=tipo_filter,
                labs_disponibles=labs_disponibles,
            )

    @app.route('/informes/bajo-minimo')
    @login_required
    def informe_bajo_minimo():
        """Análisis de mínimos #1: productos con stock_actual < minimo en
        ObServer (sumado por farmacia). Suma las ventas 12m para que el
        usuario priorice los de alta rotación.

        Filtros: lab (opcional), solo activos por defecto.
        """
        lab_id = request.args.get('lab_id', type=int)
        venta_tipo = (request.args.get('venta_tipo') or '').strip()
        rows = []
        labs_disponibles = []
        with database.get_db() as session:
            stock_q = (
                session.query(
                    ObsStock.producto_observer.label('pid'),
                    func.sum(ObsStock.stock_actual).label('stock'),
                    func.sum(ObsStock.minimo).label('minimo'),
                )
                .filter(ObsStock.minimo.isnot(None))
                .filter(ObsStock.minimo > 0)
                .group_by(ObsStock.producto_observer)
                .subquery()
            )

            desde, hasta = _ventana_12m()
            ventas_sub = (
                session.query(
                    ObsVentaMensual.producto_observer.label('pid'),
                    func.sum(ObsVentaMensual.unidades).label('u12m'),
                )
                .filter(ObsVentaMensual.anio * 100 + ObsVentaMensual.mes >= desde)
                .filter(ObsVentaMensual.anio * 100 + ObsVentaMensual.mes <= hasta)
                .group_by(ObsVentaMensual.producto_observer)
                .subquery()
            )

            q = (session.query(
                    ObsProducto.observer_id.label('pid'),
                    ObsProducto.descripcion.label('desc'),
                    ObsProducto.codigo_alfabeta,
                    ObsProducto.id_tipo_venta_control.label('tvc'),
                    ObsProducto.nombre_droga_observer.label('droga_id'),
                    ObsLaboratorio.observer_id.label('lab_id'),
                    ObsLaboratorio.descripcion.label('lab_nombre'),
                    stock_q.c.stock,
                    stock_q.c.minimo,
                    func.coalesce(ventas_sub.c.u12m, 0).label('u12m'),
                 )
                 .join(stock_q, stock_q.c.pid == ObsProducto.observer_id)
                 .outerjoin(ObsLaboratorio,
                            ObsLaboratorio.observer_id == ObsProducto.laboratorio_observer)
                 .outerjoin(ventas_sub, ventas_sub.c.pid == ObsProducto.observer_id)
                 .filter(ObsProducto.fecha_baja.is_(None))
                 .filter(stock_q.c.stock < stock_q.c.minimo)
            )
            if lab_id:
                q = q.filter(ObsProducto.laboratorio_observer == lab_id)
            if venta_tipo == 'libre':
                q = q.filter(ObsProducto.id_tipo_venta_control == 'L')
            elif venta_tipo == 'receta':
                q = q.filter(ObsProducto.id_tipo_venta_control.in_(['R', 'A']))
            elif venta_tipo == 'controlado':
                q = q.filter(ObsProducto.id_tipo_venta_control.in_(['1','2','3','4','5','6','7','8']))
            # Orden: por unidades faltantes (minimo - stock) desc, después por
            # ventas 12m desc para que los de alta rotación queden arriba.
            q = q.order_by(
                (stock_q.c.minimo - stock_q.c.stock).desc(),
                func.coalesce(ventas_sub.c.u12m, 0).desc(),
            )

            obs_ids = []
            for r in q.all():
                stock = int(r.stock or 0)
                minimo = int(r.minimo or 0)
                rows.append({
                    'producto_id': r.pid,
                    'descripcion': r.desc,
                    'codigo_alfabeta': r.codigo_alfabeta,
                    'tvc': (r.tvc or '').strip(),
                    'droga_id': r.droga_id,
                    'lab_id': r.lab_id,
                    'lab_nombre': r.lab_nombre or '—',
                    'stock': stock,
                    'minimo': minimo,
                    'faltan': max(0, minimo - stock),
                    'u12m': int(r.u12m or 0),
                    'ean': None,
                })
                obs_ids.append(r.pid)

            if obs_ids:
                ean_by_obs = dict(
                    session.query(Producto.observer_id, Producto.codigo_barra)
                    .filter(Producto.observer_id.in_(obs_ids)).all()
                )
                for r in rows:
                    r['ean'] = ean_by_obs.get(r['producto_id'])

            # Lista de labs presentes en el resultado (para el filtro).
            labs_set = {(r['lab_id'], r['lab_nombre']) for r in rows
                        if r['lab_id']}
            labs_disponibles = sorted(labs_set, key=lambda kv: kv[1])

        stats = {
            'productos': len(rows),
            'total_faltan': sum(r['faltan'] for r in rows),
            'con_ventas': sum(1 for r in rows if r['u12m'] > 0),
        }
        return render_template('informes_bajo_minimo.html',
                               rows=rows,
                               stats=stats,
                               lab_id=lab_id,
                               venta_tipo=venta_tipo,
                               labs_disponibles=labs_disponibles)

    @app.route('/informes/ventas-multi')
    @login_required
    def informe_ventas_multi():
        """Cruce de ventas por droga / producto / médico / fecha — pivot
        configurable. Filtros opcionales y group_by para agrupar resultados.
        """
        from datetime import date as _date
        from datetime import timedelta as _td

        from sqlalchemy import func as _func

        from database import (
            ObsMedico,
            ObsNombreDroga,
            ObsProducto,
            ObsVentaDetalle,
        )

        def _parse_d(s):
            try:
                return _date.fromisoformat(s) if s else None
            except (ValueError, TypeError):
                return None

        hoy = _date.today()
        desde = _parse_d(request.args.get('desde')) or (hoy - _td(days=30))
        hasta = _parse_d(request.args.get('hasta')) or hoy
        droga_id = request.args.get('droga_id', type=int)
        producto_id = request.args.get('producto_id', type=int)
        medico_id = request.args.get('medico_id', type=int)
        os_id = request.args.get('os_id', type=int)
        lab_id = request.args.get('lab_id', type=int)
        # Rubro: si no viene en URL, default = 12 (Medicamentos). Para "Todos"
        # el user pasa rubro_id=0 explícito.
        if 'rubro_id' in request.args:
            rubro_id = request.args.get('rubro_id', type=int)
        else:
            rubro_id = 12  # Medicamentos por default
        if rubro_id == 0:
            rubro_id = None
        excluir_sin_droga = request.args.get('excluir_sin_droga') == '1'
        # "Solo con receta": filtra por ObsProducto.id_tipo_venta_control != 'L'.
        # Valores: L=Venta Libre, R=Bajo Receta, A=Receta Archivada,
        # 1-4=Psicotrópico, 5-8=Estupefaciente. Todo lo no-libre requiere receta.
        solo_con_receta = request.args.get('solo_con_receta') == '1'
        group_by = (request.args.get('group_by') or 'producto').strip()
        if group_by not in ('producto', 'droga', 'laboratorio', 'medico', 'mes', 'dia', 'os'):
            group_by = 'producto'

        # Etiquetas opcionales para los filtros aplicados (mostrar en UI).
        droga_nombre = producto_desc = medico_nombre = os_nombre = lab_nombre = None

        rows = []
        total_cantidad = 0.0
        total_importe = 0.0

        # Solo ejecutamos el cruce si algún filtro fue aplicado o es rango corto.
        # Sin filtros + 30d puede ser pesado, lo dejamos correr igual capeado a 200.
        with database.get_db() as session:
            from helpers import excluir_no_medicamentos_ovd, ventas_periodo_filter
            base = (session.query(ObsVentaDetalle)
                    .filter(ventas_periodo_filter(ObsVentaDetalle, desde, hasta),
                            excluir_no_medicamentos_ovd(ObsVentaDetalle, ObsProducto, session)))
            if producto_id:
                base = base.filter(ObsVentaDetalle.producto_observer == producto_id)
                op = session.get(ObsProducto, producto_id)
                if op:
                    producto_desc = op.descripcion
            if medico_id:
                from helpers import medicos_observer_ids_compartidos
                ids_med = medicos_observer_ids_compartidos(session, medico_id)
                base = base.filter(ObsVentaDetalle.medico_observer.in_(ids_med))
                m = session.get(ObsMedico, medico_id)
                if m:
                    medico_nombre = m.nombre
            if os_id:
                from database import ObsObraSocial
                base = base.filter(ObsVentaDetalle.obra_social_observer == os_id)
                os_obj = session.get(ObsObraSocial, os_id)
                if os_obj:
                    os_nombre = os_obj.descripcion
            ya_joined_obs = False
            ya_joined_subrubro = False
            if droga_id or excluir_sin_droga or rubro_id or solo_con_receta or lab_id:
                # Cualquiera de estos requiere joinear ObsProducto.
                base = base.join(
                    ObsProducto,
                    ObsProducto.observer_id == ObsVentaDetalle.producto_observer,
                )
                ya_joined_obs = True
            if droga_id:
                base = base.filter(ObsProducto.nombre_droga_observer == droga_id)
                d = session.get(ObsNombreDroga, droga_id)
                if d:
                    droga_nombre = d.descripcion
            if lab_id:
                from database import ObsLaboratorio
                base = base.filter(ObsProducto.laboratorio_observer == lab_id)
                lab_obj = session.get(ObsLaboratorio, lab_id)
                if lab_obj:
                    lab_nombre = lab_obj.descripcion
            if excluir_sin_droga:
                base = base.filter(ObsProducto.nombre_droga_observer.isnot(None))
            if solo_con_receta:
                base = base.filter(ObsProducto.id_tipo_venta_control.isnot(None),
                                   ObsProducto.id_tipo_venta_control != 'L')
            if rubro_id:
                from database import ObsSubrubro
                base = base.join(
                    ObsSubrubro,
                    ObsSubrubro.observer_id == ObsProducto.subrubro_observer,
                ).filter(ObsSubrubro.rubro_observer == rubro_id)
                ya_joined_subrubro = True

            # GROUP BY según el pivot elegido.
            if group_by == 'producto':
                q = (base.with_entities(
                        ObsVentaDetalle.producto_observer.label('key'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                        _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                    ).group_by(ObsVentaDetalle.producto_observer)
                     .order_by(_func.sum(ObsVentaDetalle.cantidad).desc())
                     .limit(200))
                base_rows = q.all()
                pids = [r.key for r in base_rows]
                desc_por_pid = dict(session.query(ObsProducto.observer_id,
                                                   ObsProducto.descripcion)
                                     .filter(ObsProducto.observer_id.in_(pids)).all()) if pids else {}
                for r in base_rows:
                    rows.append({
                        'key_id': r.key, 'key_label': desc_por_pid.get(r.key, f'#{r.key}'),
                        'cantidad': float(r.cant or 0), 'importe': float(r.imp or 0),
                        'operaciones': int(r.ops or 0),
                    })

            elif group_by == 'droga':
                # Si ya joineamos por filtro de droga, no volver a joinear
                # (psycopg2 rompe con DuplicateAlias).
                base_q = base if ya_joined_obs else base.join(
                    ObsProducto,
                    ObsProducto.observer_id == ObsVentaDetalle.producto_observer,
                )
                q = (base_q.with_entities(
                             ObsProducto.nombre_droga_observer.label('key'),
                             _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                             _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                             _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                         ).group_by(ObsProducto.nombre_droga_observer)
                         .order_by(_func.sum(ObsVentaDetalle.cantidad).desc())
                         .limit(200))
                base_rows = q.all()
                ids = [r.key for r in base_rows if r.key]
                desc_por_id = dict(session.query(ObsNombreDroga.observer_id,
                                                  ObsNombreDroga.descripcion)
                                   .filter(ObsNombreDroga.observer_id.in_(ids)).all()) if ids else {}
                for r in base_rows:
                    rows.append({
                        'key_id': r.key, 'key_label': desc_por_id.get(r.key, '— sin droga —' if not r.key else f'#{r.key}'),
                        'cantidad': float(r.cant or 0), 'importe': float(r.imp or 0),
                        'operaciones': int(r.ops or 0),
                    })

            elif group_by == 'laboratorio':
                from database import ObsLaboratorio
                base_q = base if ya_joined_obs else base.join(
                    ObsProducto,
                    ObsProducto.observer_id == ObsVentaDetalle.producto_observer,
                )
                q = (base_q.with_entities(
                             ObsProducto.laboratorio_observer.label('key'),
                             _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                             _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                             _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                         ).group_by(ObsProducto.laboratorio_observer)
                         .order_by(_func.sum(ObsVentaDetalle.cantidad).desc())
                         .limit(200))
                base_rows = q.all()
                ids = [r.key for r in base_rows if r.key]
                lab_por_id = dict(session.query(ObsLaboratorio.observer_id,
                                                 ObsLaboratorio.descripcion)
                                  .filter(ObsLaboratorio.observer_id.in_(ids)).all()) if ids else {}
                for r in base_rows:
                    rows.append({
                        'key_id': r.key,
                        'key_label': lab_por_id.get(r.key, '— sin laboratorio —' if not r.key else f'#{r.key}'),
                        'cantidad': float(r.cant or 0), 'importe': float(r.imp or 0),
                        'operaciones': int(r.ops or 0),
                    })

            elif group_by == 'medico':
                q = (base.with_entities(
                        ObsVentaDetalle.medico_observer.label('key'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                        _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                    ).group_by(ObsVentaDetalle.medico_observer)
                     .order_by(_func.sum(ObsVentaDetalle.cantidad).desc())
                     .limit(200))
                base_rows = q.all()
                ids = [r.key for r in base_rows if r.key]
                med_por_id = dict(session.query(ObsMedico.observer_id, ObsMedico.nombre)
                                  .filter(ObsMedico.observer_id.in_(ids)).all()) if ids else {}
                for r in base_rows:
                    rows.append({
                        'key_id': r.key, 'key_label': med_por_id.get(r.key, '— sin médico (venta libre) —' if not r.key else f'#{r.key}'),
                        'cantidad': float(r.cant or 0), 'importe': float(r.imp or 0),
                        'operaciones': int(r.ops or 0),
                    })

            elif group_by == 'os':
                from database import ObsObraSocial
                q = (base.with_entities(
                        ObsVentaDetalle.obra_social_observer.label('key'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                        _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                    ).group_by(ObsVentaDetalle.obra_social_observer)
                     .order_by(_func.sum(ObsVentaDetalle.cantidad).desc())
                     .limit(200))
                base_rows = q.all()
                ids = [r.key for r in base_rows if r.key]
                os_por_id = dict(session.query(ObsObraSocial.observer_id,
                                                ObsObraSocial.descripcion)
                                 .filter(ObsObraSocial.observer_id.in_(ids)).all()) if ids else {}
                for r in base_rows:
                    rows.append({
                        'key_id': r.key,
                        'key_label': os_por_id.get(r.key, '— particular —' if not r.key else f'#{r.key}'),
                        'cantidad': float(r.cant or 0), 'importe': float(r.imp or 0),
                        'operaciones': int(r.ops or 0),
                    })

            elif group_by == 'mes':
                anio = _func.extract('year', ObsVentaDetalle.fecha_estadistica)
                mes = _func.extract('month', ObsVentaDetalle.fecha_estadistica)
                q = (base.with_entities(
                        anio.label('anio'), mes.label('mes'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                        _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                    ).group_by(anio, mes).order_by(anio, mes))
                for r in q.all():
                    rows.append({
                        'key_id': f'{int(r.anio)}-{int(r.mes):02d}',
                        'key_label': f'{int(r.mes):02d}/{int(r.anio)}',
                        'cantidad': float(r.cant or 0), 'importe': float(r.imp or 0),
                        'operaciones': int(r.ops or 0),
                    })

            else:  # dia
                q = (base.with_entities(
                        ObsVentaDetalle.fecha_estadistica.label('fec'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                        _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                    ).group_by(ObsVentaDetalle.fecha_estadistica)
                     .order_by(ObsVentaDetalle.fecha_estadistica))
                for r in q.all():
                    rows.append({
                        'key_id': r.fec.isoformat() if r.fec else '',
                        'key_label': r.fec.strftime('%d/%m/%Y') if r.fec else '—',
                        'cantidad': float(r.cant or 0), 'importe': float(r.imp or 0),
                        'operaciones': int(r.ops or 0),
                    })

            for r in rows:
                total_cantidad += r['cantidad']
                total_importe += r['importe']

            # KPIs adicionales para el banner.
            from datetime import timedelta as _td2
            dias = max(1, (hasta - desde).days + 1)
            ops_total = (base.with_entities(_func.count(ObsVentaDetalle.id_producto_vendido))
                         .scalar() or 0)
            top_med_row = (base.with_entities(
                            ObsVentaDetalle.medico_observer.label('mid'),
                            _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('c'),
                          ).filter(ObsVentaDetalle.medico_observer.isnot(None))
                           .group_by(ObsVentaDetalle.medico_observer)
                           .order_by(_func.sum(ObsVentaDetalle.cantidad).desc())
                           .first())
            top_med_nombre = None
            if top_med_row:
                m = session.get(ObsMedico, top_med_row.mid)
                top_med_nombre = m.nombre if m else None

            from database import ObsObraSocial as _ObsOS
            top_os_row = (base.with_entities(
                            ObsVentaDetalle.obra_social_observer.label('oid'),
                            _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('c'),
                          ).filter(ObsVentaDetalle.obra_social_observer.isnot(None))
                           .group_by(ObsVentaDetalle.obra_social_observer)
                           .order_by(_func.sum(ObsVentaDetalle.cantidad).desc())
                           .first())
            top_os_nombre = None
            if top_os_row:
                _o = session.get(_ObsOS, top_os_row.oid)
                top_os_nombre = _o.descripcion if _o else None

            # Donut top 10 del grupo principal (los primeros rows ya están ordenados).
            donut_data = [{'label': r['key_label'], 'value': r['cantidad']}
                          for r in rows[:10]]

            # Línea temporal por mes del período (cant + importe).
            anio = _func.extract('year', ObsVentaDetalle.fecha_estadistica)
            mes = _func.extract('month', ObsVentaDetalle.fecha_estadistica)
            tl_q = (base.with_entities(
                        anio.label('a'), mes.label('m'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('c'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('i'),
                    ).group_by(anio, mes).order_by(anio, mes))
            tl_por_mes = {f'{int(r.a)}-{int(r.m):02d}': (float(r.c or 0), float(r.i or 0))
                          for r in tl_q.all()}
            tl_labels = []
            cur = _date(desde.year, desde.month, 1)
            fin_mes = _date(hasta.year, hasta.month, 1)
            while cur <= fin_mes:
                tl_labels.append(f'{cur.year}-{cur.month:02d}')
                if cur.month == 12:
                    cur = _date(cur.year + 1, 1, 1)
                else:
                    cur = _date(cur.year, cur.month + 1, 1)
            timeline = {
                'labels': tl_labels,
                'cantidad': [tl_por_mes.get(lb, (0, 0))[0] for lb in tl_labels],
                'importe':  [tl_por_mes.get(lb, (0, 0))[1] for lb in tl_labels],
            }

        kpis = {
            'ops_total': int(ops_total),
            'ops_por_dia': round(ops_total / dias, 1) if dias else 0,
            'ticket_promedio': (total_importe / ops_total) if ops_total else 0,
            'top_medico': top_med_nombre,
            'top_os': top_os_nombre,
            'dias': dias,
        }

        # Lista de rubros para el dropdown.
        from database import ObsRubro
        with database.get_db() as session:
            rubros = [{'id': r.observer_id, 'nombre': r.descripcion}
                      for r in session.query(ObsRubro).order_by(ObsRubro.descripcion).all()]

        return render_template('informe_ventas_multi.html',
                               desde=desde, hasta=hasta,
                               droga_id=droga_id, droga_nombre=droga_nombre,
                               producto_id=producto_id, producto_desc=producto_desc,
                               medico_id=medico_id, medico_nombre=medico_nombre,
                               os_id=os_id, os_nombre=os_nombre,
                               lab_id=lab_id, lab_nombre=lab_nombre,
                               rubro_id=rubro_id,
                               excluir_sin_droga=excluir_sin_droga,
                               solo_con_receta=solo_con_receta,
                               rubros=rubros,
                               group_by=group_by, rows=rows,
                               total_cantidad=total_cantidad,
                               total_importe=total_importe,
                               kpis=kpis,
                               donut_data=donut_data,
                               timeline=timeline)

    @app.route('/informes/ventas-multi/export.xlsx')
    @login_required
    def informe_ventas_multi_export():
        """Exporta la tabla del informe a XLSX. Acepta los mismos filtros
        que la pantalla. Genera un workbook con headers + filas.
        """
        from datetime import date as _date
        from datetime import timedelta as _td
        from io import BytesIO

        from sqlalchemy import func as _func

        from database import (
            ObsMedico,
            ObsNombreDroga,
            ObsProducto,
            ObsVentaDetalle,
        )

        def _parse_d(s):
            try:
                return _date.fromisoformat(s) if s else None
            except (ValueError, TypeError):
                return None

        hoy = _date.today()
        desde = _parse_d(request.args.get('desde')) or (hoy - _td(days=30))
        hasta = _parse_d(request.args.get('hasta')) or hoy
        droga_id = request.args.get('droga_id', type=int)
        producto_id = request.args.get('producto_id', type=int)
        medico_id = request.args.get('medico_id', type=int)
        lab_id = request.args.get('lab_id', type=int)
        solo_con_receta = request.args.get('solo_con_receta') == '1'
        group_by = (request.args.get('group_by') or 'producto').strip()
        if group_by not in ('producto', 'droga', 'laboratorio', 'medico', 'mes', 'dia'):
            group_by = 'producto'

        # Reusar la misma lógica de query (copia del handler de la pantalla).
        with database.get_db() as session:
            from helpers import excluir_no_medicamentos_ovd, ventas_periodo_filter
            base = (session.query(ObsVentaDetalle)
                    .filter(ventas_periodo_filter(ObsVentaDetalle, desde, hasta),
                            excluir_no_medicamentos_ovd(ObsVentaDetalle, ObsProducto, session)))
            if producto_id:
                base = base.filter(ObsVentaDetalle.producto_observer == producto_id)
            if medico_id:
                from helpers import medicos_observer_ids_compartidos
                ids_med = medicos_observer_ids_compartidos(session, medico_id)
                base = base.filter(ObsVentaDetalle.medico_observer.in_(ids_med))
            ya_joined_obs = False
            if droga_id or solo_con_receta or lab_id:
                base = base.join(
                    ObsProducto,
                    ObsProducto.observer_id == ObsVentaDetalle.producto_observer,
                )
                ya_joined_obs = True
            if droga_id:
                base = base.filter(ObsProducto.nombre_droga_observer == droga_id)
            if lab_id:
                base = base.filter(ObsProducto.laboratorio_observer == lab_id)
            if solo_con_receta:
                base = base.filter(ObsProducto.id_tipo_venta_control.isnot(None),
                                   ObsProducto.id_tipo_venta_control != 'L')

            rows_data = []
            if group_by == 'producto':
                q = (base.with_entities(
                        ObsVentaDetalle.producto_observer.label('key'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                        _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                    ).group_by(ObsVentaDetalle.producto_observer)
                     .order_by(_func.sum(ObsVentaDetalle.cantidad).desc()).limit(2000))
                base_rows = q.all()
                pids = [r.key for r in base_rows]
                desc_por_pid = dict(session.query(ObsProducto.observer_id, ObsProducto.descripcion)
                                    .filter(ObsProducto.observer_id.in_(pids)).all()) if pids else {}
                for r in base_rows:
                    rows_data.append((desc_por_pid.get(r.key, f'#{r.key}'),
                                      int(r.ops or 0), float(r.cant or 0), float(r.imp or 0)))
            elif group_by == 'droga':
                base_q = base if ya_joined_obs else base.join(
                    ObsProducto, ObsProducto.observer_id == ObsVentaDetalle.producto_observer)
                q = (base_q.with_entities(
                        ObsProducto.nombre_droga_observer.label('key'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                        _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                    ).group_by(ObsProducto.nombre_droga_observer)
                     .order_by(_func.sum(ObsVentaDetalle.cantidad).desc()).limit(2000))
                base_rows = q.all()
                ids = [r.key for r in base_rows if r.key]
                desc_por_id = dict(session.query(ObsNombreDroga.observer_id, ObsNombreDroga.descripcion)
                                   .filter(ObsNombreDroga.observer_id.in_(ids)).all()) if ids else {}
                for r in base_rows:
                    rows_data.append((desc_por_id.get(r.key, '— sin droga —'),
                                      int(r.ops or 0), float(r.cant or 0), float(r.imp or 0)))
            elif group_by == 'laboratorio':
                from database import ObsLaboratorio
                base_q = base if ya_joined_obs else base.join(
                    ObsProducto, ObsProducto.observer_id == ObsVentaDetalle.producto_observer)
                q = (base_q.with_entities(
                        ObsProducto.laboratorio_observer.label('key'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                        _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                    ).group_by(ObsProducto.laboratorio_observer)
                     .order_by(_func.sum(ObsVentaDetalle.cantidad).desc()).limit(2000))
                base_rows = q.all()
                ids = [r.key for r in base_rows if r.key]
                lab_por_id = dict(session.query(ObsLaboratorio.observer_id,
                                                 ObsLaboratorio.descripcion)
                                  .filter(ObsLaboratorio.observer_id.in_(ids)).all()) if ids else {}
                for r in base_rows:
                    rows_data.append((lab_por_id.get(r.key, '— sin laboratorio —'),
                                      int(r.ops or 0), float(r.cant or 0), float(r.imp or 0)))
            elif group_by == 'medico':
                q = (base.with_entities(
                        ObsVentaDetalle.medico_observer.label('key'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                        _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                    ).group_by(ObsVentaDetalle.medico_observer)
                     .order_by(_func.sum(ObsVentaDetalle.cantidad).desc()).limit(2000))
                base_rows = q.all()
                ids = [r.key for r in base_rows if r.key]
                med_por_id = dict(session.query(ObsMedico.observer_id, ObsMedico.nombre)
                                  .filter(ObsMedico.observer_id.in_(ids)).all()) if ids else {}
                for r in base_rows:
                    rows_data.append((med_por_id.get(r.key, '— sin médico (venta libre) —'),
                                      int(r.ops or 0), float(r.cant or 0), float(r.imp or 0)))
            elif group_by == 'mes':
                anio = _func.extract('year', ObsVentaDetalle.fecha_estadistica)
                mes = _func.extract('month', ObsVentaDetalle.fecha_estadistica)
                q = (base.with_entities(anio.label('a'), mes.label('m'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                        _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                    ).group_by(anio, mes).order_by(anio, mes))
                for r in q.all():
                    rows_data.append((f'{int(r.m):02d}/{int(r.a)}',
                                      int(r.ops or 0), float(r.cant or 0), float(r.imp or 0)))
            else:  # dia
                q = (base.with_entities(ObsVentaDetalle.fecha_estadistica.label('fec'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                        _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                    ).group_by(ObsVentaDetalle.fecha_estadistica)
                     .order_by(ObsVentaDetalle.fecha_estadistica))
                for r in q.all():
                    rows_data.append((r.fec.strftime('%d/%m/%Y') if r.fec else '—',
                                      int(r.ops or 0), float(r.cant or 0), float(r.imp or 0)))

        # Generar workbook.
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = 'Ventas'

        # Cabecera con resumen de filtros.
        ws.append(['Informe ventas multi-dimensional'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([f'Período: {desde.isoformat()} → {hasta.isoformat()}'])
        filtros_txt = []
        if droga_id: filtros_txt.append(f'droga_id={droga_id}')
        if producto_id: filtros_txt.append(f'producto_id={producto_id}')
        if medico_id: filtros_txt.append(f'medico_id={medico_id}')
        ws.append([f"Filtros: {', '.join(filtros_txt) or '(ninguno)'}"])
        ws.append([f'Agrupado por: {group_by}'])
        ws.append([])

        # Headers de tabla.
        col_label = {'producto': 'Producto', 'droga': 'Droga',
                     'medico': 'Médico', 'mes': 'Mes', 'dia': 'Día'}.get(group_by, 'Grupo')
        headers = [col_label, 'Operaciones', 'Cantidad', 'Importe']
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='6B21A8')
            cell.alignment = Alignment(horizontal='center')

        for grupo, ops, cant, imp in rows_data:
            ws.append([grupo, ops, cant, imp])

        # Anchos.
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 16

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        from flask import send_file
        nombre = f'ventas_multi_{group_by}_{desde.isoformat()}_{hasta.isoformat()}.xlsx'
        return send_file(
            bio,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre,
        )

    @app.route('/api/informes/ventas-multi/detalle')
    @login_required
    def api_ventas_multi_detalle():
        """Detalle drill-down de un grupo del informe ventas-multi.

        Body:
        - desde, hasta: rango de fechas (igual que la tabla principal).
        - droga_id, producto_id, medico_id: filtros heredados de la tabla.
        - drill_dim: dimensión del grupo clickeado (medico/droga/producto/mes/dia).
        - drill_value: valor del grupo (ej. medico_observer=1234, mes='2026-04').

        Devuelve top 50 productos del grupo con cantidad e importe.
        """
        from datetime import date as _date
        from datetime import timedelta as _td

        from sqlalchemy import func as _func

        from database import ObsProducto, ObsVentaDetalle

        def _parse_d(s):
            try:
                return _date.fromisoformat(s) if s else None
            except (ValueError, TypeError):
                return None

        hoy = _date.today()
        desde = _parse_d(request.args.get('desde')) or (hoy - _td(days=30))
        hasta = _parse_d(request.args.get('hasta')) or hoy
        droga_id = request.args.get('droga_id', type=int)
        producto_id = request.args.get('producto_id', type=int)
        medico_id = request.args.get('medico_id', type=int)
        lab_id = request.args.get('lab_id', type=int)
        os_id = request.args.get('os_id', type=int)
        # Mismos filtros de producto que el informe principal, para que el
        # drill-down sea un subconjunto real (si no, mostraba venta libre y
        # otros rubros bajo el grupo filtrado).
        if 'rubro_id' in request.args:
            rubro_id = request.args.get('rubro_id', type=int)
        else:
            rubro_id = 12  # Medicamentos por default (igual que el informe)
        if rubro_id == 0:
            rubro_id = None
        solo_con_receta = request.args.get('solo_con_receta') == '1'
        excluir_sin_droga = request.args.get('excluir_sin_droga') == '1'
        drill_dim = (request.args.get('drill_dim') or '').strip()
        drill_value = (request.args.get('drill_value') or '').strip()

        with database.get_db() as session:
            from helpers import excluir_no_medicamentos_ovd, ventas_periodo_filter
            base = (session.query(ObsVentaDetalle)
                    .filter(ventas_periodo_filter(ObsVentaDetalle, desde, hasta),
                            excluir_no_medicamentos_ovd(ObsVentaDetalle, ObsProducto, session)))
            if producto_id:
                base = base.filter(ObsVentaDetalle.producto_observer == producto_id)
            if medico_id:
                from helpers import medicos_observer_ids_compartidos
                ids_med = medicos_observer_ids_compartidos(session, medico_id)
                base = base.filter(ObsVentaDetalle.medico_observer.in_(ids_med))
            if os_id:
                base = base.filter(ObsVentaDetalle.obra_social_observer == os_id)
            ya_joined_obs = False
            if droga_id or lab_id or rubro_id or solo_con_receta or excluir_sin_droga:
                base = base.join(
                    ObsProducto,
                    ObsProducto.observer_id == ObsVentaDetalle.producto_observer,
                )
                ya_joined_obs = True
            if droga_id:
                base = base.filter(ObsProducto.nombre_droga_observer == droga_id)
            if lab_id:
                base = base.filter(ObsProducto.laboratorio_observer == lab_id)
            if excluir_sin_droga:
                base = base.filter(ObsProducto.nombre_droga_observer.isnot(None))
            if solo_con_receta:
                base = base.filter(ObsProducto.id_tipo_venta_control.isnot(None),
                                   ObsProducto.id_tipo_venta_control != 'L')
            if rubro_id:
                from database import ObsSubrubro
                base = base.join(
                    ObsSubrubro,
                    ObsSubrubro.observer_id == ObsProducto.subrubro_observer,
                ).filter(ObsSubrubro.rubro_observer == rubro_id)

            # Aplicar el drill: filtrar por la dimensión clickeada.
            if drill_dim == 'medico':
                try:
                    mid = int(drill_value) if drill_value else None
                except ValueError:
                    mid = None
                if mid is not None:
                    from helpers import medicos_observer_ids_compartidos
                    ids_med = medicos_observer_ids_compartidos(session, mid)
                    base = base.filter(ObsVentaDetalle.medico_observer.in_(ids_med))
                else:
                    base = base.filter(ObsVentaDetalle.medico_observer.is_(None))
            elif drill_dim == 'producto':
                try:
                    pid = int(drill_value)
                    base = base.filter(ObsVentaDetalle.producto_observer == pid)
                except (ValueError, TypeError):
                    pass
            elif drill_dim == 'droga':
                try:
                    did = int(drill_value) if drill_value else None
                except ValueError:
                    did = None
                if not ya_joined_obs:
                    base = base.join(
                        ObsProducto,
                        ObsProducto.observer_id == ObsVentaDetalle.producto_observer,
                    )
                if did is not None:
                    base = base.filter(ObsProducto.nombre_droga_observer == did)
                else:
                    base = base.filter(ObsProducto.nombre_droga_observer.is_(None))
            elif drill_dim == 'os':
                try:
                    oid = int(drill_value) if drill_value else None
                except ValueError:
                    oid = None
                if oid is not None:
                    base = base.filter(ObsVentaDetalle.obra_social_observer == oid)
                else:
                    base = base.filter(ObsVentaDetalle.obra_social_observer.is_(None))
            elif drill_dim == 'mes':
                try:
                    anio_s, mes_s = drill_value.split('-')
                    base = base.filter(
                        _func.extract('year', ObsVentaDetalle.fecha_estadistica) == int(anio_s),
                        _func.extract('month', ObsVentaDetalle.fecha_estadistica) == int(mes_s),
                    )
                except (ValueError, AttributeError):
                    pass
            elif drill_dim == 'dia':
                fec = _parse_d(drill_value)
                if fec:
                    base = base.filter(ObsVentaDetalle.fecha_estadistica == fec)

            # Agregar por producto.
            q = (base.with_entities(
                    ObsVentaDetalle.producto_observer.label('pid'),
                    _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                    _func.coalesce(_func.sum(ObsVentaDetalle.importe), 0).label('imp'),
                    _func.count(ObsVentaDetalle.id_producto_vendido).label('ops'),
                ).group_by(ObsVentaDetalle.producto_observer)
                 .order_by(_func.sum(ObsVentaDetalle.cantidad).desc())
                 .limit(50))
            rows = q.all()
            pids = [r.pid for r in rows if r.pid]
            desc_por_pid = dict(session.query(ObsProducto.observer_id, ObsProducto.descripcion)
                                .filter(ObsProducto.observer_id.in_(pids)).all()) if pids else {}
            return jsonify({'ok': True, 'items': [{
                'producto_id': r.pid,
                'descripcion': desc_por_pid.get(r.pid, f'#{r.pid}' if r.pid else '—'),
                'cantidad': float(r.cant or 0),
                'importe': float(r.imp or 0),
                'operaciones': int(r.ops or 0),
            } for r in rows]})

    @app.route('/api/informes/ventas-multi/historico-droga-medico')
    @login_required
    def api_ventas_multi_hist_droga_medico():
        """Histórico mensual de prescripción de una droga por médico.

        Dado un droga_id + rango, devuelve top N médicos por cantidad total
        y la serie de cantidad mensual de cada uno. Ideal para chart de líneas.
        """
        from datetime import date as _date
        from datetime import timedelta as _td

        from sqlalchemy import func as _func

        from database import (
            ObsMedico,
            ObsProducto,
            ObsVentaDetalle,
        )
        from helpers import ventas_periodo_filter

        def _parse_d(s):
            try:
                return _date.fromisoformat(s) if s else None
            except (ValueError, TypeError):
                return None

        droga_id = request.args.get('droga_id', type=int)
        if not droga_id:
            return jsonify({'ok': False, 'error': 'droga_id requerido'}), 400
        hoy = _date.today()
        desde = _parse_d(request.args.get('desde')) or (hoy - _td(days=180))
        hasta = _parse_d(request.args.get('hasta')) or hoy
        top_n = max(1, min(request.args.get('top', default=5, type=int), 15))

        with database.get_db() as session:
            from helpers import excluir_no_medicamentos_ovd
            anio = _func.extract('year', ObsVentaDetalle.fecha_estadistica)
            mes = _func.extract('month', ObsVentaDetalle.fecha_estadistica)
            base = (session.query(ObsVentaDetalle)
                    .join(ObsProducto,
                          ObsProducto.observer_id == ObsVentaDetalle.producto_observer)
                    .filter(ObsProducto.nombre_droga_observer == droga_id,
                            ventas_periodo_filter(ObsVentaDetalle, desde, hasta),
                            excluir_no_medicamentos_ovd(ObsVentaDetalle, ObsProducto, session),
                            ObsVentaDetalle.medico_observer.isnot(None)))

            # Top N médicos por cantidad total en el período.
            top_q = (base.with_entities(
                        ObsVentaDetalle.medico_observer.label('mid'),
                        _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                    ).group_by(ObsVentaDetalle.medico_observer)
                     .order_by(_func.sum(ObsVentaDetalle.cantidad).desc())
                     .limit(top_n))
            top_rows = top_q.all()
            top_ids = [r.mid for r in top_rows if r.mid]
            nombres = dict(session.query(ObsMedico.observer_id, ObsMedico.nombre)
                           .filter(ObsMedico.observer_id.in_(top_ids)).all()) if top_ids else {}

            # Serie mensual por médico (solo top N).
            por_medico = {mid: {} for mid in top_ids}
            if top_ids:
                serie_q = (base.with_entities(
                              ObsVentaDetalle.medico_observer.label('mid'),
                              anio.label('anio'), mes.label('mes'),
                              _func.coalesce(_func.sum(ObsVentaDetalle.cantidad), 0).label('cant'),
                          ).filter(ObsVentaDetalle.medico_observer.in_(top_ids))
                           .group_by(ObsVentaDetalle.medico_observer, anio, mes))
                for r in serie_q.all():
                    if not r.mid:
                        continue
                    key = f'{int(r.anio)}-{int(r.mes):02d}'
                    por_medico[r.mid][key] = float(r.cant or 0)

            # Construir labels de meses (incluyendo los vacíos).
            labels = []
            cur = _date(desde.year, desde.month, 1)
            fin = _date(hasta.year, hasta.month, 1)
            while cur <= fin:
                labels.append(f'{cur.year}-{cur.month:02d}')
                # avanzar 1 mes
                if cur.month == 12:
                    cur = _date(cur.year + 1, 1, 1)
                else:
                    cur = _date(cur.year, cur.month + 1, 1)

            series = []
            for mid in top_ids:
                series.append({
                    'medico_id': mid,
                    'nombre': nombres.get(mid, f'Médico #{mid}'),
                    'data': [por_medico[mid].get(lb, 0) for lb in labels],
                    'total': sum(por_medico[mid].values()),
                })
            return jsonify({'ok': True, 'labels': labels, 'series': series})

    # /api/informes/buscar-medico eliminado — el JS de informe_ventas_multi
    # ahora consume /api/consulta-medico/buscar (superset: multi-token AND +
    # matrícula en el response).

    @app.route('/api/informes/buscar-os')
    @login_required
    def api_informes_buscar_os():
        """Autocomplete para filtro Obra Social."""
        from database import ObsObraSocial
        q = (request.args.get('q') or '').strip()
        if len(q) < 2:
            return jsonify({'items': []})
        with database.get_db() as session:
            results = (session.query(ObsObraSocial)
                       .filter(ObsObraSocial.fecha_baja.is_(None),
                               ObsObraSocial.descripcion.ilike(f'%{q}%'))
                       .order_by(ObsObraSocial.descripcion)
                       .limit(20).all())
            return jsonify({'items': [{'id': o.observer_id, 'nombre': o.descripcion}
                                       for o in results]})

    @app.route('/api/informes/buscar-lab')
    @login_required
    def api_informes_buscar_lab():
        """Autocomplete para filtro Laboratorio (catálogo ObServer)."""
        from database import ObsLaboratorio
        q = (request.args.get('q') or '').strip()
        if len(q) < 2:
            return jsonify({'items': []})
        with database.get_db() as session:
            results = (session.query(ObsLaboratorio)
                       .filter(ObsLaboratorio.descripcion.ilike(f'%{q}%'))
                       .order_by(ObsLaboratorio.descripcion)
                       .limit(20).all())
            return jsonify({'items': [{'id': l.observer_id, 'nombre': l.descripcion}
                                       for l in results]})

    @app.route('/api/informes/buscar-producto-obs')
    @login_required
    def api_informes_buscar_producto_obs():
        """Autocomplete para filtro producto (catálogo Observer)."""
        q = (request.args.get('q') or '').strip()
        if len(q) < 2:
            return jsonify({'items': []})
        with database.get_db() as session:
            results = (session.query(ObsProducto)
                       .filter(ObsProducto.fecha_baja.is_(None),
                               ObsProducto.descripcion.ilike(f'%{q}%'))
                       .order_by(ObsProducto.descripcion)
                       .limit(20).all())
            return jsonify({'items': [{'id': p.observer_id, 'nombre': p.descripcion}
                                       for p in results]})


    @app.route('/api/observer-product/<int:observer_id>/chart')
    @login_required
    def api_observer_product_chart(observer_id):
        """Datos para el gráfico histórico, leídos desde obs_ventas_mensuales.

        No requiere que el producto esté en el catálogo local — solo necesita
        el observer_id. Útil para informes/listados que cruzan con ObServer
        directo y no pasan por Producto local.
        """
        from services.producto_metrics import metricas_producto
        with database.get_db() as session:
            obs = session.get(ObsProducto, observer_id)
            if not obs:
                return jsonify({'ok': False, 'error': 'Producto ObServer no encontrado'}), 404

            # Source of truth unico: stock/min de 1 farmacia, avg_3m + avg_12m,
            # rotacion via rotation_index. Ver services/producto_metrics.py.
            m = metricas_producto(session, observer_id)
            return jsonify({
                'ok': True,
                'nombre': obs.descripcion or '',
                'codigo_barra': obs.codigo_alfabeta or str(observer_id),
                'ventas': m['ventas12'],
                'avg_monthly': m['avg_monthly'],   # alias de avg_12m (backcompat)
                'avg_3m': m['avg_3m'],
                'avg_12m': m['avg_12m'],
                'slope': m['slope'],
                'stock': m['stock'],
                'minimo': m['minimo'],
                'rotacion': m['rotacion'],
                'tipo': 'N',
                'start_month': m['start_month'],
                'n_days': 35,
                'sin_historial': m['sin_historial'],
                'analizado_en': None,
            })

    @app.route('/api/observer-product/<int:observer_id>/chart-mes')
    @login_required
    def api_observer_product_chart_mes(observer_id):
        """Gráfico del último mes (30 días) por DÍA, desde obs_ventas_detalle.

        Devuelve labels = ['DD/MM' x 30] y data = unidades vendidas ese día.
        """
        from datetime import date as _date
        from datetime import timedelta

        from database import ObsVentaDetalle
        with database.get_db() as session:
            obs = session.get(ObsProducto, observer_id)
            if not obs:
                return jsonify({'ok': False, 'error': 'Producto ObServer no encontrado'}), 404

            hoy = _date.today()
            desde = hoy - timedelta(days=29)  # ventana de 30 días incluyendo hoy
            rows = (session.query(
                        ObsVentaDetalle.fecha_estadistica,
                        func.coalesce(func.sum(ObsVentaDetalle.cantidad), 0),
                    )
                    .filter(ObsVentaDetalle.producto_observer == observer_id,
                            ObsVentaDetalle.fecha_estadistica.isnot(None),
                            ObsVentaDetalle.fecha_estadistica >= desde,
                            ObsVentaDetalle.fecha_estadistica <= hoy)
                    .group_by(ObsVentaDetalle.fecha_estadistica)
                    .all())
            por_fecha = {r[0]: float(r[1] or 0) for r in rows}
            labels, datos = [], []
            d = desde
            while d <= hoy:
                labels.append(d.strftime('%d/%m'))
                datos.append(round(por_fecha.get(d, 0), 2))
                d += timedelta(days=1)
            total = sum(datos)
            avg = total / 30.0
            return jsonify({
                'ok': True,
                'nombre': obs.descripcion or '',
                'observer_id': observer_id,
                'labels': labels,
                'data': datos,
                'total_30d': round(total, 2),
                'avg_diario': round(avg, 2),
                'desde': desde.isoformat(),
                'hasta': hoy.isoformat(),
            })

    @app.route('/api/observer-product/<int:observer_id>/ingresos-mes')
    @login_required
    def api_observer_product_ingresos_mes(observer_id):
        """Ingresos (factura_items) + Pedidos (pedido_emitido_item) por día,
        últimos 30 días.

        - Ingresos: mapea observer_id → códigos de barra (Producto + obs_codigos_barras).
          NCR resta.
        - Pedidos: usa observer_id directo de pedido_emitido_item.
        """
        from datetime import date as _date
        from datetime import timedelta

        from database import Invoice, InvoiceItem, ObsCodigoBarras, PedidoEmitido, PedidoEmitidoItem, Producto
        with database.get_db() as session:
            obs = session.get(ObsProducto, observer_id)
            if not obs:
                return jsonify({'ok': False, 'error': 'Producto ObServer no encontrado'}), 404

            codigos = set()
            # Vía obs_codigos_barras (todas las orden, sin baja)
            for r in (session.query(ObsCodigoBarras.codigo_barras)
                      .filter(ObsCodigoBarras.producto_observer == observer_id,
                              ObsCodigoBarras.fecha_baja.is_(None)).all()):
                if r[0]: codigos.add(r[0])
            # Vía Producto local: codigo_barra principal + 1-a-N
            # (alt1/2/3 legacy ya no se consultan)
            prod = (session.query(Producto)
                    .filter(Producto.observer_id == observer_id).first())
            if prod:
                if prod.codigo_barra:
                    codigos.add(prod.codigo_barra)
                from database import ProductoCodigoBarra
                for cb, in (session.query(ProductoCodigoBarra.codigo_barra)
                            .filter_by(producto_id=prod.id).all()):
                    if cb:
                        codigos.add(cb)

            hoy = _date.today()
            desde = hoy - timedelta(days=29)
            por_fecha = {}
            if codigos:
                rows = (session.query(Invoice.fecha,
                                      func.coalesce(func.sum(InvoiceItem.cantidad), 0))
                        .join(InvoiceItem, InvoiceItem.factura_id == Invoice.id)
                        .filter(InvoiceItem.codigo_barra.in_(list(codigos)))
                        .filter(Invoice.fecha >= desde,
                                Invoice.fecha <= hoy)
                        .group_by(Invoice.fecha).all())
                por_fecha = {r[0]: float(r[1] or 0) for r in rows}

            # Pedidos por día (cantidad_pedida agregada por fecha del pedido).
            ped_rows = (session.query(
                            func.date(PedidoEmitido.fecha),
                            func.coalesce(func.sum(PedidoEmitidoItem.cantidad_pedida), 0))
                        .join(PedidoEmitido,
                              PedidoEmitido.id == PedidoEmitidoItem.pedido_id)
                        .filter(PedidoEmitidoItem.observer_id == observer_id,
                                func.date(PedidoEmitido.fecha) >= desde,
                                func.date(PedidoEmitido.fecha) <= hoy)
                        .group_by(func.date(PedidoEmitido.fecha))
                        .all())
            por_fecha_ped = {r[0]: float(r[1] or 0) for r in ped_rows}

            # Total pendiente (independiente de fecha): suma de
            # cantidad_pedida - cantidad_recibida en PedidoEmitidoItem con
            # estado=PENDIENTE para este observer_id. Se muestra como barra
            # extra a la derecha del chart de 30 días.
            pend_row = (session.query(
                            func.coalesce(func.sum(
                                PedidoEmitidoItem.cantidad_pedida
                                - PedidoEmitidoItem.cantidad_recibida), 0))
                        .filter(PedidoEmitidoItem.observer_id == observer_id,
                                PedidoEmitidoItem.estado == 'PENDIENTE')
                        .scalar())
            pendiente_total = float(pend_row or 0)

            labels, datos, datos_ped = [], [], []
            d = desde
            while d <= hoy:
                labels.append(d.strftime('%d/%m'))
                datos.append(round(por_fecha.get(d, 0), 2))
                datos_ped.append(round(por_fecha_ped.get(d, 0), 2))
                d += timedelta(days=1)
            total = sum(datos)
            total_ped = sum(datos_ped)
            return jsonify({
                'ok': True,
                'nombre': obs.descripcion or '',
                'observer_id': observer_id,
                'codigos_resueltos': len(codigos),
                'labels': labels,
                'data': datos,
                'total_30d': round(total, 2),
                'pedido_data': datos_ped,
                'pedido_total_30d': round(total_ped, 2),
                'pendiente_total': round(pendiente_total, 2),
                'desde': desde.isoformat(),
                'hasta': hoy.isoformat(),
            })

    @app.route('/api/informes/buscar-droga')
    @login_required
    def api_buscar_droga():
        """Autocomplete para el buscador de drogas. Devuelve top 20 que
        contengan el texto en la descripción, ordenadas alfabéticamente."""
        q = (request.args.get('q') or '').strip()
        if len(q) < 2:
            return jsonify({'items': []})
        with database.get_db() as session:
            results = (session.query(ObsNombreDroga)
                       .filter(ObsNombreDroga.descripcion.ilike(f'%{q}%'))
                       .order_by(ObsNombreDroga.descripcion)
                       .limit(20).all())
            items = [{'id': r.observer_id, 'descripcion': r.descripcion}
                     for r in results]
        return jsonify({'items': items})

    # Una observación mal importada suele ser una presentación de producto
    # ("30 comp", "100 ml", "60 caps") que el wizard mapeó a observacion por
    # error. Sirve para marcar grupos sospechosos y ofrecer limpieza bulk.
    import re as _re_obs
    _OBS_SOSPECHOSA_RE = _re_obs.compile(
        r'^\s*\d+\s*(ml|comp|cap|caps|sob|amp|sup|kg|grm|gr|g|mg)\b', _re_obs.I,
    )

    def _obs_es_sospechosa(s):
        return bool(s and _OBS_SOSPECHOSA_RE.match(s))

    @app.route('/informes/ofertas-activas')
    @login_required
    def informe_ofertas_activas():
        """Gestión global de ofertas cargadas: OfertaMinimo agrupadas + Módulos."""
        from sqlalchemy import func as _func

        from database import Laboratorio, Modulo, ModuloPack, OfertaMinimo, Provider
        with database.get_db() as session:
            # ── OfertaMinimo — agrupar por (lab, tipo, drogueria) ─────────────
            rows_om = (session.query(OfertaMinimo, Laboratorio, Provider)
                       .outerjoin(Laboratorio, Laboratorio.id == OfertaMinimo.laboratorio_id)
                       .outerjoin(Provider, Provider.id == OfertaMinimo.drogueria_id)
                       .filter(OfertaMinimo.activo == True)
                       .order_by(Laboratorio.nombre, OfertaMinimo.observacion,
                                 OfertaMinimo.tipo_descuento, OfertaMinimo.descripcion)
                       .all())

            # Agrupar por (lab_id, observacion, tipo, drogueria_id)
            grupos_dict = {}
            for o, lab, prov in rows_om:
                key = (o.laboratorio_id, o.observacion or '', o.tipo_descuento or 'simple', o.drogueria_id)
                if key not in grupos_dict:
                    grupos_dict[key] = {
                        'lab':        lab.nombre if lab else '—',
                        'lab_id':     o.laboratorio_id,
                        'drog':       prov.razon_social if prov else None,
                        'drog_id':    o.drogueria_id,
                        'tipo':       o.tipo_descuento or 'simple',
                        'observacion': o.observacion or '',
                        'vigencia':   o.vigencia_hasta.strftime('%d/%m/%Y') if o.vigencia_hasta else None,
                        'actualizado': o.actualizado_en.strftime('%d/%m/%Y') if o.actualizado_en else None,
                        '_vh':        o.vigencia_hasta,
                        'prods':      [],
                    }
                g = grupos_dict[key]
                g['prods'].append({
                    'id':              o.id,
                    'ean':             o.ean or '',
                    'descripcion':     o.descripcion or '',
                    'unidades_minima': o.unidades_minima,
                    'descuento':       float(o.descuento_psl) if o.descuento_psl is not None else None,
                    'vigencia':        o.vigencia_hasta.strftime('%d/%m/%Y') if o.vigencia_hasta else None,
                    'observacion':     o.observacion or '',
                })
                if o.vigencia_hasta and (g['_vh'] is None or o.vigencia_hasta > g['_vh']):
                    g['_vh'] = o.vigencia_hasta
                    g['vigencia'] = o.vigencia_hasta.strftime('%d/%m/%Y')

            grupos = [dict(g, _raw_vh=g.get('_vh'), _vh=None,
                           sospechosa=_obs_es_sospechosa(g.get('observacion')))
                      for g in grupos_dict.values()]

            # ── Módulos — agrupar por lab (nivel 1 solamente) ────────────────
            from sqlalchemy import func as _func2
            rows_mod = (session.query(
                            Laboratorio.nombre.label('lab_nombre'),
                            Modulo.laboratorio_id,
                            _func2.count(Modulo.id).label('n_modulos'),
                            _func2.max(Modulo.creado_en).label('ultima_importacion'),
                        )
                        .outerjoin(Laboratorio, Laboratorio.id == Modulo.laboratorio_id)
                        .group_by(Modulo.laboratorio_id, Laboratorio.nombre)
                        .order_by(Laboratorio.nombre)
                        .all())
            modulos = [{
                'lab_id':   r.laboratorio_id,
                'lab':      r.lab_nombre or '—',
                'n_modulos': r.n_modulos,
                'importado': r.ultima_importacion.strftime('%d/%m/%Y') if r.ultima_importacion else '',
            } for r in rows_mod]

        resumen = {
            'con_minimo': sum(1 for g in grupos if g['tipo'] == 'con_minimo' and not g['drog']),
            'simple':     sum(1 for g in grupos if g['tipo'] == 'simple' and not g['drog']),
            'multi_lab':  sum(1 for g in grupos if g['drog']),
            'modulos':    len(modulos),
        }
        return render_template('informe_ofertas_activas.html',
                               grupos=grupos, modulos=modulos, resumen=resumen)

    # ─── Sync/pull bulk de TODAS las ofertas con Render ──────────────────
    # Versiones "bulk" del sync/pull por-lab que están en routes/laboratorios.py.
    # Más cómodo cuando las ofertas entran por droguería (no por lab) y querés
    # mover todo de una.

    @app.route('/informes/ofertas-activas/grupo/toggle-activa', methods=['POST'])
    @login_required
    def informes_ofertas_grupo_toggle_activa():
        """Activa/desactiva (activo=True/False) todas las ofertas que matcheen
        un grupo (lab, observacion, vigencia_hasta, drog, tipo). Para que el
        operador deje 1 batch activo por lab y los anteriores como histórico.
        Body JSON: {lab_id, observacion, vigencia_hasta, drog_id, tipo, set_to}.
        """
        from datetime import datetime as _dt

        from database import OfertaMinimo
        d = request.get_json(silent=True) or {}
        try:
            lab_id = int(d.get('lab_id')) if d.get('lab_id') not in (None, 'null', '') else None
        except (TypeError, ValueError):
            lab_id = None
        try:
            drog_id = int(d.get('drog_id')) if d.get('drog_id') not in (None, 'null', '') else None
        except (TypeError, ValueError):
            drog_id = None
        observacion = (d.get('observacion') or '').strip() or None
        tipo        = (d.get('tipo') or '').strip() or None
        set_to      = bool(d.get('set_to'))
        vh_str      = (d.get('vigencia_hasta') or '').strip()
        vh = None
        if vh_str:
            for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
                try:
                    vh = _dt.strptime(vh_str, fmt).date(); break
                except ValueError:
                    continue
        with database.get_db() as session:
            q = session.query(OfertaMinimo)
            # Lab puede ser None (multi-lab); usamos `is_` cuando corresponde.
            q = q.filter(OfertaMinimo.laboratorio_id == lab_id) if lab_id is not None \
                else q.filter(OfertaMinimo.laboratorio_id.is_(None))
            q = q.filter(OfertaMinimo.drogueria_id == drog_id) if drog_id is not None \
                else q.filter(OfertaMinimo.drogueria_id.is_(None))
            if observacion:
                q = q.filter(OfertaMinimo.observacion == observacion)
            else:
                q = q.filter(OfertaMinimo.observacion.is_(None))
            if tipo:
                q = q.filter(OfertaMinimo.tipo_descuento == tipo)
            if vh:
                q = q.filter(OfertaMinimo.vigencia_hasta == vh)
            n = q.update({OfertaMinimo.activo: set_to},
                         synchronize_session=False)
            session.commit()
        return jsonify({'ok': True, 'actualizadas': int(n), 'set_to': set_to})

    @app.route('/informes/ofertas-activas/sospechosas/preview', methods=['GET'])
    @login_required
    def informes_ofertas_sospechosas_preview():
        """Lista las ofertas con `observacion` tipo presentación ('30 comp',
        '100 ml', etc.) — síntoma típico de un import donde la columna se
        mapeó por error al campo observacion."""
        from database import Laboratorio, OfertaMinimo
        with database.get_db() as session:
            rows = (session.query(OfertaMinimo, Laboratorio.nombre)
                    .outerjoin(Laboratorio, Laboratorio.id == OfertaMinimo.laboratorio_id)
                    .filter(OfertaMinimo.activo.is_(True),
                            OfertaMinimo.observacion.isnot(None))
                    .order_by(Laboratorio.nombre, OfertaMinimo.observacion).all())
            items = []
            for o, lab in rows:
                if _obs_es_sospechosa(o.observacion):
                    items.append({
                        'id':  o.id,
                        'lab': lab or '—',
                        'observacion': o.observacion,
                        'ean': o.ean or '',
                        'descripcion': (o.descripcion or '')[:120],
                    })
        return jsonify({'items': items, 'n': len(items)})

    @app.route('/informes/ofertas-activas/sospechosas/borrar', methods=['POST'])
    @login_required
    def informes_ofertas_sospechosas_borrar():
        """Borra (delete) todas las ofertas con observación tipo presentación.
        El frontend ya confirmó con el user previa visualización."""
        from database import OfertaMinimo
        with database.get_db() as session:
            rows = (session.query(OfertaMinimo)
                    .filter(OfertaMinimo.activo.is_(True),
                            OfertaMinimo.observacion.isnot(None)).all())
            n = 0
            for o in rows:
                if _obs_es_sospechosa(o.observacion):
                    session.delete(o)
                    n += 1
            session.commit()
        return jsonify({'ok': True, 'borradas': n})

    _QUEUE_OBS_PREFIX = 'Aplicada desde queue por '

    @app.route('/informes/ofertas-activas/queue/preview', methods=['GET'])
    @login_required
    def informes_ofertas_queue_preview():
        """Lista las ofertas auto-creadas desde la queue de pendientes
        (observacion empieza con 'Aplicada desde queue por ...')."""
        from database import Laboratorio, OfertaMinimo
        with database.get_db() as session:
            rows = (session.query(OfertaMinimo, Laboratorio.nombre)
                    .outerjoin(Laboratorio, Laboratorio.id == OfertaMinimo.laboratorio_id)
                    .filter(OfertaMinimo.observacion.like(_QUEUE_OBS_PREFIX + '%'))
                    .order_by(Laboratorio.nombre, OfertaMinimo.observacion).all())
        items = [{
            'id':  o.id,
            'lab': lab or '—',
            'observacion': o.observacion,
            'ean': o.ean or '',
            'descripcion': (o.descripcion or '')[:120],
        } for o, lab in rows]
        return jsonify({'items': items, 'n': len(items)})

    @app.route('/informes/ofertas-activas/queue/borrar', methods=['POST'])
    @login_required
    def informes_ofertas_queue_borrar():
        """Borra todas las ofertas auto-creadas desde la queue."""
        from database import OfertaMinimo
        with database.get_db() as session:
            n = (session.query(OfertaMinimo)
                 .filter(OfertaMinimo.observacion.like(_QUEUE_OBS_PREFIX + '%'))
                 .delete(synchronize_session=False))
            session.commit()
        return jsonify({'ok': True, 'borradas': int(n)})

    @app.route('/informes/ofertas-activas/sync-render-bulk', methods=['POST'])
    @login_required
    def informe_sync_render_bulk():
        """Push TODAS las ofertas locales a Render, agrupadas por lab.

        Itera Laboratorio.id, llama a /api/ofertas/sync-from-local de Render
        una vez por lab (Render se encarga de upsertear allá). Devuelve el
        total agregado.
        """
        import os as _os

        import requests

        from database import Laboratorio, OfertaMinimo
        render_url = _os.environ.get('RENDER_BASE_URL', '').rstrip('/')
        token = _os.environ.get('PANEL_REMOTO_TOKEN', '')
        if not render_url or not token:
            return jsonify({
                'ok': False,
                'error': 'Sync no configurado. Setear RENDER_BASE_URL + PANEL_REMOTO_TOKEN en env.'
            }), 400

        with database.get_db() as session:
            # Pre-agrupar ofertas por lab para no hacer N queries.
            from sqlalchemy.orm import joinedload  # noqa: F401
            labs_con_ofertas = (session.query(Laboratorio)
                                .join(OfertaMinimo, OfertaMinimo.laboratorio_id == Laboratorio.id)
                                .distinct().all())
            payloads = []
            for lab in labs_con_ofertas:
                rows = session.query(OfertaMinimo).filter_by(laboratorio_id=lab.id).all()
                payloads.append({
                    'laboratorio_nombre': lab.nombre,
                    'ofertas': [{
                        'ean':             r.ean,
                        'codigo':          r.codigo,
                        'descripcion':     r.descripcion,
                        'unidades_minima': r.unidades_minima,
                        'descuento_psl':   float(r.descuento_psl) if r.descuento_psl is not None else None,
                        'rentabilidad':    float(r.rentabilidad)  if r.rentabilidad  is not None else None,
                        'plazo_pago':      r.plazo_pago,
                        'grupo_id':        r.grupo_id,
                        'tipo_descuento':  r.tipo_descuento,
                        'drogueria_id':    r.drogueria_id,
                        'vigencia_desde':  r.vigencia_desde.isoformat() if r.vigencia_desde else None,
                        'vigencia_hasta':  r.vigencia_hasta.isoformat() if r.vigencia_hasta else None,
                        'observacion':     r.observacion,
                    } for r in rows],
                })

        total_creadas = 0
        total_actualizadas = 0
        errores = []
        for p in payloads:
            try:
                r = requests.post(
                    f'{render_url}/api/ofertas/sync-from-local',
                    json=p,
                    headers={'X-Panel-Token': token},
                    timeout=60,
                )
            except requests.exceptions.RequestException as e:
                errores.append(f'{p["laboratorio_nombre"]}: conexión: {e}')
                continue
            if r.status_code != 200:
                errores.append(f'{p["laboratorio_nombre"]}: HTTP {r.status_code}: {r.text[:120]}')
                continue
            d = r.json()
            total_creadas += d.get('creadas') or 0
            total_actualizadas += d.get('actualizadas') or 0
            for e in d.get('errores') or []:
                errores.append(f'{p["laboratorio_nombre"]}: {e}')

        return jsonify({
            'ok': True,
            'labs': len(payloads),
            'creadas': total_creadas,
            'actualizadas': total_actualizadas,
            'errores': errores,
        })

    @app.route('/informes/ofertas-activas/pull-render-bulk', methods=['POST'])
    @login_required
    def informe_pull_render_bulk():
        """Pull TODAS las ofertas de Render → upsert local.

        GET a /api/ofertas/from-server sin filtro → devuelve todas. Cada
        oferta tiene `laboratorio_id` (id de Render, no necesariamente
        igual al local). Resolvemos lab por nombre o creamos uno nuevo.
        """
        import datetime as _dt
        import os as _os

        import requests

        from database import Laboratorio, OfertaMinimo
        from helpers import _normalizar_nombre_entidad, normalizar_unidades_minima
        render_url = _os.environ.get('RENDER_BASE_URL', '').rstrip('/')
        token = _os.environ.get('PANEL_REMOTO_TOKEN', '')
        if not render_url or not token:
            return jsonify({
                'ok': False,
                'error': 'Pull no configurado. Setear RENDER_BASE_URL + PANEL_REMOTO_TOKEN en env.'
            }), 400

        # 1. Listar todos los labs de Render con ofertas (un GET por lab no
        # escala bien si hay muchos). Como `from-server` sin filtro devuelve
        # TODAS las ofertas (con sus laboratorio_id de Render), igual sirve;
        # pero el endpoint actual no incluye el nombre del lab. Para resolver
        # ese lookup, hacemos primero una lista de labs locales y mappeamos
        # por nombre normalizado. Si un lab de Render NO existe local, lo
        # creamos.
        try:
            r = requests.get(
                f'{render_url}/api/ofertas/from-server',
                headers={'X-Panel-Token': token},
                timeout=120,
            )
        except requests.exceptions.RequestException as e:
            return jsonify({'ok': False, 'error': f'No pude conectar con Render: {e}'}), 502

        if r.status_code != 200:
            return jsonify({
                'ok': False,
                'error': f'Render devolvió {r.status_code}: {r.text[:300]}',
            }), 502
        data = r.json()
        ofertas_remote = data.get('ofertas') or []

        # Sin nombre del lab por fila → necesitamos un segundo lookup por lab_id
        # remoto. Para simplificar: hacemos N GET por nombre. Pero como el
        # endpoint actual no expone "listar labs", optamos por una estrategia
        # diferente: cuando Render devuelve `laboratorio_id` y `laboratorio_nombre`
        # solo en el meta del response (no por fila), trabajamos con esos.
        # Si el endpoint no devuelve un nombre por fila, mejor pedimos por
        # cada lab local.

        # Estrategia alternativa: iteramos los labs LOCALES y para cada uno
        # pullea desde Render (ya tenemos endpoint por lab). Más simple y
        # consistente con el flujo del botón individual.
        creadas = 0
        actualizadas = 0
        fetched = 0
        labs_creados = 0
        errores = []
        with database.get_db() as session:
            labs_locales = session.query(Laboratorio).all()

        for lab in labs_locales:
            try:
                rr = requests.get(
                    f'{render_url}/api/ofertas/from-server',
                    params={'laboratorio_nombre': lab.nombre},
                    headers={'X-Panel-Token': token},
                    timeout=60,
                )
            except requests.exceptions.RequestException as e:
                errores.append(f'{lab.nombre}: conexión: {e}')
                continue
            if rr.status_code == 404:
                continue  # ese lab no existe en Render — saltear
            if rr.status_code != 200:
                errores.append(f'{lab.nombre}: HTTP {rr.status_code}')
                continue
            payload = rr.json()
            for o in payload.get('ofertas') or []:
                ean = (o.get('ean') or '').strip()
                if not ean:
                    continue
                fetched += 1
                key = {
                    'laboratorio_id': lab.id,
                    'ean':            ean,
                    'grupo_id':       o.get('grupo_id'),
                    'drogueria_id':   o.get('drogueria_id'),
                }
                vd = vh = None
                try:
                    if o.get('vigencia_desde'): vd = _dt.date.fromisoformat(o['vigencia_desde'])
                    if o.get('vigencia_hasta'): vh = _dt.date.fromisoformat(o['vigencia_hasta'])
                except (ValueError, TypeError) as e:
                    errores.append(f'{lab.nombre} EAN {ean}: fecha: {e}')
                with database.get_db() as session:
                    existente = session.query(OfertaMinimo).filter_by(**key).first()
                    if existente:
                        existente.descripcion     = o.get('descripcion')
                        existente.codigo          = o.get('codigo')
                        existente.unidades_minima = normalizar_unidades_minima(o.get('unidades_minima'))
                        existente.descuento_psl   = o.get('descuento_psl')
                        existente.rentabilidad    = o.get('rentabilidad')
                        existente.plazo_pago      = o.get('plazo_pago')
                        existente.tipo_descuento  = o.get('tipo_descuento')
                        existente.vigencia_desde  = vd
                        existente.vigencia_hasta  = vh
                        existente.observacion     = o.get('observacion')
                        actualizadas += 1
                    else:
                        session.add(OfertaMinimo(
                            laboratorio_id  = lab.id,
                            ean             = ean,
                            descripcion     = o.get('descripcion'),
                            codigo          = o.get('codigo'),
                            unidades_minima = normalizar_unidades_minima(o.get('unidades_minima')),
                            descuento_psl   = o.get('descuento_psl'),
                            rentabilidad    = o.get('rentabilidad'),
                            plazo_pago      = o.get('plazo_pago'),
                            grupo_id        = o.get('grupo_id'),
                            tipo_descuento  = o.get('tipo_descuento'),
                            drogueria_id    = o.get('drogueria_id'),
                            vigencia_desde  = vd,
                            vigencia_hasta  = vh,
                            observacion     = o.get('observacion'),
                        ))
                        creadas += 1
                    session.commit()

        _ = _normalizar_nombre_entidad  # ruff: usado por endpoints relacionados
        return jsonify({
            'ok': True,
            'fetched': fetched,
            'creadas': creadas,
            'actualizadas': actualizadas,
            'labs_creados': labs_creados,
            'errores': errores,
        })

    @app.route('/informes/ofertas-activas/borrar-grupo', methods=['POST'])
    @login_required
    def informe_grupo_borrar():
        """Elimina todas las OfertaMinimo de un grupo (lab+tipo+drogueria)."""
        from database import OfertaMinimo
        data = request.get_json(silent=True) or {}
        lab_id  = data.get('lab_id')
        tipo    = data.get('tipo')
        obs     = data.get('obs', '')
        drog_id = data.get('drog_id')
        with database.get_db() as session:
            q = session.query(OfertaMinimo)
            if lab_id is not None:
                q = q.filter(OfertaMinimo.laboratorio_id == lab_id)
            else:
                q = q.filter(OfertaMinimo.laboratorio_id.is_(None))
            if tipo:
                q = q.filter(OfertaMinimo.tipo_descuento == tipo)
            if obs:
                q = q.filter(OfertaMinimo.observacion == obs)
            else:
                q = q.filter(OfertaMinimo.observacion.is_(None) | (OfertaMinimo.observacion == ''))
            if drog_id is not None:
                q = q.filter(OfertaMinimo.drogueria_id == drog_id)
            else:
                q = q.filter(OfertaMinimo.drogueria_id.is_(None))
            q.delete(synchronize_session=False)
            session.commit()
        return ('', 204)

    @app.route('/informes/ofertas-activas/borrar-grupos-bulk', methods=['POST'])
    @login_required
    def informe_grupos_borrar_bulk():
        """Borra varios grupos (lab+tipo+obs+drogueria) en una sola transacción."""
        from database import OfertaMinimo
        data = request.get_json(silent=True) or {}
        groups = data.get('groups') or []
        if not isinstance(groups, list) or not groups:
            return jsonify({'ok': False, 'error': 'Sin grupos'}), 400
        borradas = 0
        with database.get_db() as session:
            for g in groups:
                lab_id  = g.get('lab_id')
                tipo    = g.get('tipo')
                obs     = g.get('obs', '')
                drog_id = g.get('drog_id')
                q = session.query(OfertaMinimo)
                if lab_id is not None:
                    q = q.filter(OfertaMinimo.laboratorio_id == lab_id)
                else:
                    q = q.filter(OfertaMinimo.laboratorio_id.is_(None))
                if tipo:
                    q = q.filter(OfertaMinimo.tipo_descuento == tipo)
                if obs:
                    q = q.filter(OfertaMinimo.observacion == obs)
                else:
                    q = q.filter(OfertaMinimo.observacion.is_(None) | (OfertaMinimo.observacion == ''))
                if drog_id is not None:
                    q = q.filter(OfertaMinimo.drogueria_id == drog_id)
                else:
                    q = q.filter(OfertaMinimo.drogueria_id.is_(None))
                borradas += q.delete(synchronize_session=False)
            session.commit()
        return jsonify({'ok': True, 'borradas': borradas, 'grupos': len(groups)})

    @app.route('/informes/ofertas-activas/borrar-modulo/<int:modulo_id>', methods=['POST'])
    @login_required
    def informe_modulo_borrar(modulo_id):
        from database import Modulo
        with database.get_db() as session:
            m = session.get(Modulo, modulo_id)
            if m:
                session.delete(m)
                session.commit()
        return ('', 204)
