"""Consulta de compras por artículo.

Busca sobre TODOS los ítems de factura importados (InvoiceItem + Invoice): un
producto (por descripción o EAN) en un rango de fechas → cada compra con fecha,
nº de factura, proveedor, cantidad, precio unitario, %Dto e importe. Sirve para
ver el histórico de precios/descuentos de un producto entre droguerías.

(El "nº de ingreso" = recepción de ObServer queda pendiente: hoy no se captura,
ver get_recepciones_factura.)
"""
from __future__ import annotations

from datetime import date as _date
from datetime import datetime as _dt

from flask import Response, abort, render_template, request
from flask_login import login_required
from sqlalchemy import or_

import database
from database import Invoice, InvoiceItem, get_db
from helpers import get_providers

_LIMITE = 500


def _parse_fecha(s):
    s = (s or '').strip()
    if not s:
        return None
    try:
        return _dt.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return None


def _consultar(session, q, desde, hasta, prov_cuit):
    """(filas, resumen) de las compras que matchean. filas=[] si no hay término."""
    if not q:
        return [], None
    query = (session.query(InvoiceItem, Invoice)
             .join(Invoice, InvoiceItem.factura_id == Invoice.id))
    # Cada palabra tiene que aparecer en la descripción (o el EAN):
    # "optamox duo" matchea "OPTAMOX DUO 1g…".
    for palabra in q.split():
        like = f'%{palabra}%'
        query = query.filter(or_(InvoiceItem.descripcion.ilike(like),
                                 InvoiceItem.codigo_barra.ilike(like)))
    if desde:
        query = query.filter(Invoice.fecha >= desde)
    if hasta:
        query = query.filter(Invoice.fecha <= hasta)
    if prov_cuit:
        query = query.filter(Invoice.proveedor_cuit == prov_cuit)
    rows = (query.order_by(Invoice.fecha.desc(),
                           InvoiceItem.descripcion).limit(_LIMITE).all())
    filas = []
    tot_u = 0
    tot_imp = 0.0
    dtos = []
    for it, inv in rows:
        dto = float(it.dto) if it.dto is not None else None
        filas.append({
            'invoice_id': inv.id,
            'fecha': inv.fecha,
            'numero': inv.numero_factura,
            'tipo': inv.tipo_comprobante or 'FAC',
            'proveedor': inv.proveedor_razon or '—',
            'codigo_barra': it.codigo_barra or '',
            'descripcion': it.descripcion or '',
            'cantidad': it.cantidad or 0,
            'precio_unitario': float(it.precio_unitario) if it.precio_unitario is not None else None,
            'dto': dto,
            'importe': float(it.importe) if it.importe is not None else None,
        })
        tot_u += it.cantidad or 0
        tot_imp += float(it.importe or 0)
        if dto is not None and dto > 0:
            dtos.append(dto)
    resumen = {
        'n': len(filas), 'unidades': tot_u, 'importe': tot_imp,
        'dto_min': min(dtos) if dtos else None,
        'dto_max': max(dtos) if dtos else None,
        'dto_prom': (sum(dtos) / len(dtos)) if dtos else None,
        'truncado': len(filas) >= _LIMITE,
    }
    return filas, resumen


def _xlsx(filas, q) -> bytes:
    import io

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Compras'
    cols = [('Fecha', 12), ('Factura', 16), ('Tipo', 7), ('Proveedor', 28),
            ('EAN', 16), ('Producto', 42), ('Cant.', 8), ('P. Unitario', 13),
            ('% Dto', 8), ('Importe', 14)]
    ws.append([c[0] for c in cols])
    for i, (_t, w) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1C1C1E')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for f in filas:
        ws.append([
            f['fecha'].strftime('%d/%m/%Y') if f['fecha'] else '',
            f['numero'], f['tipo'], f['proveedor'], f['codigo_barra'],
            f['descripcion'], f['cantidad'],
            f['precio_unitario'], f['dto'], f['importe'],
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def init_app(app):

    @app.route('/compras/consulta')
    @login_required
    def consulta_compras():
        q = (request.args.get('q') or '').strip()
        desde = _parse_fecha(request.args.get('desde'))
        hasta = _parse_fecha(request.args.get('hasta'))
        prov_cuit = (request.args.get('proveedor') or '').strip()
        with get_db() as session:
            proveedores = get_providers()
            filas, resumen = _consultar(session, q, desde, hasta, prov_cuit)
        return render_template('consulta_compras.html', q=q,
                               desde=request.args.get('desde') or '',
                               hasta=request.args.get('hasta') or '',
                               prov_cuit=prov_cuit, proveedores=proveedores,
                               filas=filas, resumen=resumen, hoy=_date.today())

    @app.route('/compras/consulta/export.xlsx')
    @login_required
    def consulta_compras_export():
        q = (request.args.get('q') or '').strip()
        if not q:
            abort(400, description='Falta el término de búsqueda.')
        desde = _parse_fecha(request.args.get('desde'))
        hasta = _parse_fecha(request.args.get('hasta'))
        prov_cuit = (request.args.get('proveedor') or '').strip()
        with get_db() as session:
            filas, _ = _consultar(session, q, desde, hasta, prov_cuit)
        contenido = _xlsx(filas, q)
        slug = ''.join(c if c.isalnum() else '-' for c in q)[:30] or 'compras'
        nombre = f'Compras-{slug}-{_date.today():%Y-%m-%d}.xlsx'
        return Response(
            contenido,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{nombre}"'})
