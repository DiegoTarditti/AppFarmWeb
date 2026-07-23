"""Transferencias entre sucursales: comparador N-way (de a pares) + export XLSX."""
import io

from flask import jsonify, render_template, request, send_file
from flask_login import login_required

from helpers import get_config
from services import transferencias as svc


def _umbrales():
    cfg = get_config()
    exc = cfg.get('transfer_excedente_meses', 6.0)
    nec = cfg.get('transfer_necesita_meses', 2.0)
    try:
        exc = float(request.args.get('exc', exc))
        nec = float(request.args.get('nec', nec))
    except (ValueError, TypeError):
        pass
    return exc, nec


def init_app(app):
    @app.route('/transferencias')
    @login_required
    def transferencias():
        exc, nec = _umbrales()
        otra = (request.args.get('otra') or '').strip() or None
        data = svc.analizar(excedente_meses=exc, necesita_meses=nec, otra=otra)
        return render_template('transferencias.html', data=data, exc=exc, nec=nec)

    @app.route('/transferencias/export')
    @login_required
    def transferencias_export():
        import openpyxl as ox
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        exc, nec = _umbrales()
        otra = (request.args.get('otra') or '').strip() or None
        data = svc.analizar(excedente_meses=exc, necesita_meses=nec, otra=otra)
        if not data.get('ok'):
            return jsonify(data), 400

        loc = data['local']['nombre']
        otr = data['otra']['nombre']
        wb = ox.Workbook()
        ws = wb.active
        ws.title = 'Transferencias'
        cols = ['Alfabeta', 'Producto',
                f'{otr} stock', f'{otr} vta/m', f'{otr} cob(m)',
                f'{loc} stock', f'{loc} vta/m', f'{loc} cob(m)',
                'Direccion', 'Transferir']
        hf = Font(bold=True, color='FFFFFF')
        fill = PatternFill('solid', fgColor='1F2937')
        for ci, h in enumerate(cols, 1):
            cell = ws.cell(1, ci, h)
            cell.font = hf
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
        r = 2
        for f in data['filas']:
            direc = (f'{otr} -> {loc}' if f['direccion'] == 'otra_a_local'
                     else f'{loc} -> {otr}')
            vals = [f['alfabeta'], f['descripcion'],
                    f['o_stock'], f['o_avg'], f['o_cob'],
                    f['l_stock'], f['l_avg'], f['l_cob'], direc, f['qty']]
            for ci, v in enumerate(vals, 1):
                ws.cell(r, ci, v)
            r += 1
        for ci, w in enumerate([10, 38, 11, 10, 11, 11, 10, 11, 18, 11], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='transferencias.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
