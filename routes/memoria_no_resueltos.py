"""Panel: memoria de no-resueltos + analítica de caminos del bot.

Muestra qué pudo y qué no pudo resolver el bot: demanda perdida (productos sin
stock), consultas no entendidas y derivaciones a humano. Lista filtrable +
resumen del día + export a Excel. Solo lectura (la captura la hace el cerebro).

Rutas:
  GET /bot/no-resueltos                  → panel
  GET /bot/no-resueltos/api/lista        → JSON listado filtrable
  GET /bot/no-resueltos/api/resumen      → JSON agregados del día + texto en criollo
  GET /bot/no-resueltos/export.xlsx      → Excel del listado filtrado
"""
from io import BytesIO

from flask import jsonify, render_template, request, send_file
from flask_login import current_user, login_required

from bot import store

_ROLES_OK = ('admin', 'dev', 'farmacia')

# Etiquetas legibles para los motivos (panel + export + resumen).
MOTIVO_LABEL = {
    'sin_stock': 'Sin stock / no encontrado',
    'no_entendido': 'No entendido',
    'derivado': 'Derivado a persona',
    'receta_ilegible': 'Receta ilegible',
    'falta_info': 'Faltó info',
    'rechazado_malicioso': 'Rechazado (indebido)',
}


def _permitido():
    return getattr(current_user, 'rol', None) in _ROLES_OK


def _arg(name):
    v = (request.args.get(name) or '').strip()
    return v or None


def _resumen_criollo(r):
    """Convierte el dict de agregados en frases para el dueño."""
    if not r['total']:
        return ['Sin actividad ese día.']
    out = [f"Procesé {r['total']} consultas: {r['resueltas']} resueltas, "
           f"{r['no_resueltas']} sin resolver."]
    m = r['por_motivo']
    if m.get('sin_stock'):
        prods = ', '.join(f"{p['producto']} ({p['veces']})"
                          for p in r['productos_sin_stock'][:8])
        out.append(f"Me pidieron {m['sin_stock']} veces productos que no tenés o "
                   f"están sin stock{': ' + prods if prods else ''}.")
    if m.get('no_entendido'):
        out.append(f"No entendí {m['no_entendido']} consultas.")
    if m.get('derivado'):
        out.append(f"Derivé {m['derivado']} a una persona del equipo.")
    if m.get('falta_info'):  # tanda 2 (IA)
        temas = ', '.join(f"{t['tema']} ({t['veces']})" for t in r['temas'][:8])
        out.append(f"Me faltó info en {m['falta_info']} casos{': ' + temas if temas else ''}.")
    if m.get('rechazado_malicioso'):  # tanda 2 (IA)
        out.append(f"Rechacé {m['rechazado_malicioso']} consultas indebidas/maliciosas.")
    return out


def init_app(app):

    @app.route('/bot/no-resueltos')
    @login_required
    def memoria_panel():
        if not _permitido():
            return 'Sin permiso', 403
        return render_template('memoria_no_resueltos.html',
                               lineas=store.lineas_interacciones(),
                               motivos=MOTIVO_LABEL)

    @app.route('/bot/no-resueltos/api/lista')
    @login_required
    def memoria_lista():
        if not _permitido():
            return jsonify({'error': 'sin permiso'}), 403
        filas = store.listar_interacciones(
            desde=_arg('desde'), hasta=_arg('hasta'), linea=_arg('linea'),
            motivo=_arg('motivo'),
            solo_no_resueltas=(request.args.get('solo_no_resueltas') == '1'))
        return jsonify({'interacciones': filas, 'motivos': MOTIVO_LABEL})

    @app.route('/bot/no-resueltos/api/resumen')
    @login_required
    def memoria_resumen():
        if not _permitido():
            return jsonify({'error': 'sin permiso'}), 403
        r = store.resumen_del_dia(fecha=_arg('fecha'), linea=_arg('linea'))
        return jsonify({'resumen': r, 'texto': _resumen_criollo(r)})

    @app.route('/bot/no-resueltos/export.xlsx')
    @login_required
    def memoria_export():
        if not _permitido():
            return 'Sin permiso', 403
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        filas = store.listar_interacciones(
            desde=_arg('desde'), hasta=_arg('hasta'), linea=_arg('linea'),
            motivo=_arg('motivo'),
            solo_no_resueltas=(request.args.get('solo_no_resueltas') == '1'),
            limite=5000)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'No resueltos'
        headers = ['Fecha', 'Canal', 'Línea', 'Camino', 'Resuelto',
                   'Motivo', 'Tema', 'Producto', 'Texto']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1c1c1e', end_color='1c1c1e',
                                    fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        for f in filas:
            ws.append([f['fecha'], f['canal'] or '', f['linea'] or '',
                       f['camino'] or '', 'sí' if f['resuelto'] else 'no',
                       MOTIVO_LABEL.get(f['motivo'], f['motivo'] or ''),
                       f['tema'] or '', f['producto'] or '', f['texto'] or ''])
        for i, w in enumerate([14, 10, 12, 14, 9, 22, 18, 26, 60], start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A2'
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name='bot_no_resueltos.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
