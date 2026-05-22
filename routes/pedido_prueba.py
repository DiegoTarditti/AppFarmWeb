"""Pedido prueba — pantalla de planificacion grande con estacionalidad.

A diferencia de /pedido/dia (reposicion tactica con u3m/90), aca usamos
u12m/365 como base "neutra" + indice estacional del mes objetivo +
cobertura en dias. Pensado para pedidos grandes anticipados.

Flujo: usuario elige un laboratorio → pantalla popula con todos los
productos del lab con ventas en 12m → cada producto muestra:
- Sugerido del dia (calculado con la matriz REPOSICION en vivo)
- Sugerido prueba (este modulo, con escenario aplicable)
- Diferencia
- Flag activo si aplica
- Drawer al click con desglose + sliders para editar escenario producto.
"""

import json
import os

from flask import jsonify, render_template, request
from flask_login import current_user, login_required
from sqlalchemy import func

from database import (
    ObsLaboratorio,
    ObsNombreDroga,
    ObsProducto,
    ObsStock,
    ObsVentaMensual,
    ProductoFlag,
    TipoPedidoConfig,
    get_db,
)
from helpers import aplicar_overrides_planificador, calcular_metricas_pedido_auto
from services.pedido_estacional import (
    LIMITES,
    MESES_ES,
    calcular_sugerido_dia_actual,
    calcular_sugerido_estacional,
    obtener_escenarios_bulk,
    obtener_flags_bulk,
    obtener_precios_publicos_bulk,
    obtener_ventas_arr_bulk,
)


def _u12m_por_producto(session, producto_ids, id_farmacia):
    """Suma u12m por producto en un solo query."""
    if not producto_ids:
        return {}
    rows = (session.query(
        ObsVentaMensual.producto_observer,
        func.sum(ObsVentaMensual.unidades).label('u12m'),
    )
    .filter(ObsVentaMensual.producto_observer.in_(producto_ids),
            ObsVentaMensual.id_farmacia == id_farmacia)
    .group_by(ObsVentaMensual.producto_observer)
    .all())
    return {r.producto_observer: float(r.u12m or 0) for r in rows}


def _stock_por_producto(session, producto_ids, id_farmacia):
    if not producto_ids:
        return {}
    rows = (session.query(ObsStock.producto_observer, ObsStock.stock_actual,
                          ObsStock.minimo)
            .filter(ObsStock.producto_observer.in_(producto_ids),
                    ObsStock.id_farmacia == id_farmacia)
            .all())
    return {r.producto_observer: {'stock': int(r.stock_actual or 0),
                                  'minimo': int(r.minimo or 0)} for r in rows}


# _sugerido_dia_actual eliminada — ahora se usa la replica fiel
# calcular_sugerido_dia_actual de services/pedido_estacional.py que usa
# los mismos building blocks que routes/compras_dia.py (purchase_helpers.
# calcular_min_sugerido + services.calculo_pedido.calcular_a_pedir).


def init_app(app):

    @app.route('/pedido/prueba')
    @login_required
    def pedido_prueba():
        """Pantalla principal. Arranca con selector de lab vacio."""
        with get_db() as session:
            labs = (session.query(ObsLaboratorio.observer_id,
                                  ObsLaboratorio.descripcion)
                    .order_by(ObsLaboratorio.descripcion)
                    .all())
            labs_list = [{'id': l.observer_id, 'nombre': l.descripcion}
                         for l in labs]
        return render_template('pedido_prueba.html',
                               labs=labs_list,
                               meses_es=MESES_ES,
                               limites=LIMITES)

    @app.route('/api/pedido-prueba/historico/<int:producto_id>')
    @login_required
    def api_pedido_prueba_historico(producto_id):
        """Serie mensual de ventas del producto por anio, para chart del drawer."""
        id_farmacia = int(os.environ.get('OBSERVER_ID_FARMACIA', '10525'))
        from collections import defaultdict
        with get_db() as session:
            rows = (session.query(
                ObsVentaMensual.anio, ObsVentaMensual.mes,
                func.sum(ObsVentaMensual.unidades).label('u'))
                .filter(ObsVentaMensual.producto_observer == producto_id,
                        ObsVentaMensual.id_farmacia == id_farmacia)
                .group_by(ObsVentaMensual.anio, ObsVentaMensual.mes)
                .order_by(ObsVentaMensual.anio, ObsVentaMensual.mes)
                .all())
        por_anio = defaultdict(lambda: [0.0] * 12)
        for r in rows:
            por_anio[r.anio][r.mes - 1] = float(r.u or 0)
        return jsonify({
            'producto_id': producto_id,
            'series': [{'anio': a, 'unidades': por_anio[a]}
                       for a in sorted(por_anio.keys())],
        })

    @app.route('/api/pedido-prueba/calcular', methods=['POST'])
    @login_required
    def api_pedido_prueba_calcular():
        """Devuelve productos del lab + sugeridos (prueba y dia) + flag."""
        payload = request.get_json(silent=True) or {}
        try:
            lab_id = int(payload.get('lab_id') or 0)
        except (TypeError, ValueError):
            return jsonify({'error': 'lab_id invalido'}), 400
        if not lab_id:
            return jsonify({'error': 'lab_id requerido'}), 400

        min_u12m = max(0, float(payload.get('min_u12m') or 0))
        # Defaults configurables desde la cabecera. Se aplican a productos
        # SIN escenario propio (origen='auto'). Los productos con escenario
        # propio o de droga mantienen sus valores.
        try:
            lead_default = max(
                LIMITES['lead_dias_piso'],
                min(LIMITES['lead_dias_max'],
                    int(payload.get('lead_default') or LIMITES['lead_dias_default'])),
            )
        except (TypeError, ValueError):
            lead_default = LIMITES['lead_dias_default']
        try:
            cob_default = max(
                LIMITES['cob_dias_min'],
                min(LIMITES['cob_dias_max'],
                    int(payload.get('cob_default') or LIMITES['cob_dias_default'])),
            )
        except (TypeError, ValueError):
            cob_default = LIMITES['cob_dias_default']
        id_farmacia = int(os.environ.get('OBSERVER_ID_FARMACIA', '10525'))

        with get_db() as session:
            productos = (session.query(ObsProducto)
                         .filter(ObsProducto.laboratorio_observer == lab_id,
                                 ObsProducto.fecha_baja.is_(None))
                         .all())
            producto_ids = [p.observer_id for p in productos]
            u12m_map = _u12m_por_producto(session, producto_ids, id_farmacia)
            stock_map = _stock_por_producto(session, producto_ids, id_farmacia)

            # Precios PVP (precio publico). Pre-cargar todos los EANs del
            # lab en 2 queries: 1 para EANs activos + 1 para precios.
            from database import ObsCodigoBarras
            ean_rows = (session.query(
                ObsCodigoBarras.producto_observer, ObsCodigoBarras.codigo_barras)
                .filter(ObsCodigoBarras.producto_observer.in_(producto_ids),
                        ObsCodigoBarras.fecha_baja.is_(None))
                .order_by(ObsCodigoBarras.orden)
                .all()) if producto_ids else []
            eans_por_producto = {}
            for r in ean_rows:
                eans_por_producto.setdefault(r.producto_observer, []).append(
                    r.codigo_barras)
            todos_eans = list({e for eans in eans_por_producto.values() for e in eans})
            precios_map = obtener_precios_publicos_bulk(session, todos_eans)

            # Nombres de drogas
            drogas_ids = list({p.nombre_droga_observer for p in productos
                               if p.nombre_droga_observer is not None})
            drogas_map = dict(session.query(
                ObsNombreDroga.observer_id, ObsNombreDroga.descripcion)
                .filter(ObsNombreDroga.observer_id.in_(drogas_ids)).all()) if drogas_ids else {}

            # Bulk precargas para eliminar N+1 dentro del loop.
            escenarios_bulk = obtener_escenarios_bulk(session, drogas_ids, producto_ids)
            flags_bulk = obtener_flags_bulk(session, todos_eans, lab_id=lab_id)
            ventas_arr_bulk = obtener_ventas_arr_bulk(session, producto_ids, id_farmacia)

            # Política de overrides: la define el TipoPedidoConfig 'PRUEBA'
            # (configurable desde /config/tipos-pedido). Si no existe, defaults
            # = comportamiento histórico (cant_fija override, oferta_min piso).
            import json as _json_tp

            from database import TipoPedidoConfig as _TPC
            _tp_prueba = (session.query(_TPC)
                          .filter_by(slug='PRUEBA', categoria='pedido').first())
            _cfg_prueba = {}
            if _tp_prueba and _tp_prueba.config_json:
                try:
                    _cfg_prueba = _json_tp.loads(_tp_prueba.config_json)
                except (ValueError, TypeError):
                    _cfg_prueba = {}
            _cant_fija_efecto = _cfg_prueba.get('cant_fija_efecto', 'override')
            _oferta_min_efecto = _cfg_prueba.get('oferta_min_efecto', 'piso')

            # Overrides operativos del catálogo (cant_fija + oferta_min).
            # Sirven para que el planificador no diverja de /compras/dia/armar
            # cuando el operador ya cargó política puntual por producto/EAN.
            from datetime import date as _date_h

            from database import Laboratorio, OfertaMinimo, Producto
            cant_fija_por_obs = {}  # producto_observer_id → cant_fija (int) | None
            oferta_min_por_obs = {}  # producto_observer_id → unidades_minima (int) | None
            if producto_ids:
                rows_cf = (session.query(Producto.observer_id,
                                          Producto.cantidad_reposicion_fija)
                           .filter(Producto.observer_id.in_(producto_ids),
                                   Producto.cantidad_reposicion_fija.isnot(None))
                           .all())
                cant_fija_por_obs = {r.observer_id: int(r.cantidad_reposicion_fija)
                                     for r in rows_cf}
                # OfertaMinimo: vinculo via EAN + lab LOCAL (no observer).
                # Si no hay Laboratorio local mapeado, no hay ofertas posibles.
                local_lab = (session.query(Laboratorio)
                             .filter(Laboratorio.observer_id == lab_id).first())
                if local_lab and todos_eans:
                    hoy_h = _date_h.today()
                    of_rows = (session.query(OfertaMinimo.ean,
                                             OfertaMinimo.unidades_minima,
                                             OfertaMinimo.vigencia_hasta)
                               .filter(OfertaMinimo.laboratorio_id == local_lab.id,
                                       OfertaMinimo.ean.in_(todos_eans),
                                       OfertaMinimo.unidades_minima.isnot(None),
                                       OfertaMinimo.unidades_minima > 1)
                               .all())
                    # Por EAN, quedarse con el mínimo vigente más alto (la
                    # oferta más exigente). Si una está vencida (vigencia_hasta
                    # < hoy), se descarta.
                    min_por_ean = {}
                    for r in of_rows:
                        if r.vigencia_hasta and r.vigencia_hasta < hoy_h:
                            continue
                        cur = min_por_ean.get(r.ean, 0)
                        if int(r.unidades_minima) > cur:
                            min_por_ean[r.ean] = int(r.unidades_minima)
                    # Mapear de EAN → producto_observer_id (cualquier EAN del producto).
                    for pid, eans_p in eans_por_producto.items():
                        mejor = 0
                        for e in eans_p:
                            v = min_por_ean.get(e, 0)
                            if v > mejor:
                                mejor = v
                        if mejor > 0:
                            oferta_min_por_obs[pid] = mejor

            # Frescura de datos: tomar el sync_en mas reciente de obs_stock
            # y obs_ventas_mensuales (proxies de "stock" y "ventas"
            # actualizadas). Si no hay data, devuelve None.
            stock_sync = (session.query(func.max(ObsStock.sync_en))
                          .filter(ObsStock.id_farmacia == id_farmacia,
                                  ObsStock.producto_observer.in_(producto_ids))
                          .scalar() if producto_ids else None)
            ventas_sync = (session.query(func.max(ObsVentaMensual.sync_en))
                           .filter(ObsVentaMensual.id_farmacia == id_farmacia,
                                   ObsVentaMensual.producto_observer.in_(producto_ids))
                           .scalar() if producto_ids else None)

            resultado = []
            total_dia = 0
            total_prueba = 0
            monto_dia = 0.0
            monto_prueba = 0.0
            for p in productos:
                u12m = u12m_map.get(p.observer_id, 0)
                if u12m < min_u12m:
                    continue
                st = stock_map.get(p.observer_id, {'stock': 0, 'minimo': 0})
                # Precio publico: tomar de cualquiera de los EANs del producto.
                precio_pvp = None
                for ean in eans_por_producto.get(p.observer_id, []):
                    if ean in precios_map:
                        precio_pvp = precios_map[ean]
                        break

                # Sugerido estacional (con bulks precargados → cero N+1)
                est = calcular_sugerido_estacional(
                    session, p, u12m=u12m,
                    stock_actual=st['stock'], minimo=st['minimo'],
                    lead_default=lead_default, cob_default=cob_default,
                    escenarios_bulk=escenarios_bulk,
                    eans_producto=eans_por_producto.get(p.observer_id, []),
                    flags_bulk=flags_bulk,
                    lab_id_hint=lab_id,
                )
                # Sugerido dia (REPOSICION) — replica fiel + bulk
                sug_dia = calcular_sugerido_dia_actual(
                    session, p.observer_id, id_farmacia,
                    stock_actual=st['stock'], min_actual=st['minimo'],
                    ventas_arr_bulk=ventas_arr_bulk)

                # Overrides operativos: aplicar a AMBOS sugeridos para que la
                # comparación dia ↔ prueba siga siendo apples-to-apples y para
                # que el planificador refleje lo que pasaría en /compras/dia.
                cant_fija_p = cant_fija_por_obs.get(p.observer_id)
                oferta_min_p = oferta_min_por_obs.get(p.observer_id)
                sug_prueba_original = est['sugerido_final']
                sug_dia_original = sug_dia
                sug_prueba_final, ov_slug, ov_valor = aplicar_overrides_planificador(
                    sugerido=sug_prueba_original,
                    stock=st['stock'], minimo=st['minimo'],
                    cant_fija=cant_fija_p, oferta_min=oferta_min_p,
                    cant_fija_efecto=_cant_fija_efecto,
                    oferta_min_efecto=_oferta_min_efecto)
                if sug_dia is not None:
                    sug_dia_final, _, _ = aplicar_overrides_planificador(
                        sugerido=sug_dia, stock=st['stock'], minimo=st['minimo'],
                        cant_fija=cant_fija_p, oferta_min=oferta_min_p,
                        cant_fija_efecto=_cant_fija_efecto,
                        oferta_min_efecto=_oferta_min_efecto)
                else:
                    sug_dia_final = None
                # Reemplazo el estacional con el final (overrides aplicados)
                # para que totales/montos/charts usen el valor real.
                est['sugerido_final'] = sug_prueba_final
                sug_dia = sug_dia_final

                delta = (sug_prueba_final - (sug_dia or 0)) if sug_dia is not None else None

                # u3m: ultimos 3 meses (excluyendo mes parcial actual).
                # Aprovecho ventas_arr_bulk que ya tengo cargado.
                _arr = ventas_arr_bulk.get(p.observer_id, [0.0] * 12)
                u3m = int(sum(_arr[8:11]))  # indices 8,9,10 = 3 meses antes del actual
                # Promedios mensuales + rotacion (usando rotation_index de
                # purchase_engine.py que ya define A/M/B: >=20/mes, >=5/mes, resto).
                prom_3m = round(u3m / 3.0, 1) if u3m else 0
                prom_12m = round(float(u12m) / 12.0, 1) if u12m else 0
                from purchase_engine import rotation_index as _rot_idx
                rotacion = _rot_idx(prom_12m)  # 'A' | 'M' | 'B'

                # Diagnóstico de mínimo + pérdida estimada por faltante
                # (heredado de /informes/pedido-auto antes de eliminarlo).
                # m12m aproximado = precio_pvp × u12m si hay precio, sino 0
                # (sin precio no podemos valorizar la pérdida).
                m12m_aprox = (precio_pvp * u12m) if precio_pvp else 0
                met = calcular_metricas_pedido_auto(
                    stock=st['stock'], minimo=st['minimo'], maximo=None,
                    u12m=u12m, m12m=m12m_aprox,
                )
                resultado.append({
                    'producto_id': p.observer_id,
                    'producto_nombre': p.descripcion,
                    'droga_id': p.nombre_droga_observer,
                    'droga_nombre': drogas_map.get(p.nombre_droga_observer) or '',
                    'stock': st['stock'],
                    'minimo': st['minimo'],
                    'u12m': int(u12m),
                    'u3m': u3m,
                    'prom_3m': prom_3m,
                    'prom_12m': prom_12m,
                    'rotacion': rotacion,
                    'sugerido_dia': sug_dia,
                    'sugerido_prueba': est['sugerido_final'],
                    'sugerido_base_prueba': est['sugerido_base'],
                    'delta': delta,
                    'precio_pvp': precio_pvp,
                    'monto_dia': (precio_pvp * sug_dia) if (precio_pvp and sug_dia) else None,
                    'monto_prueba': (precio_pvp * est['sugerido_final']) if precio_pvp else None,
                    'origen_escenario': est['origen_escenario'],
                    'escenario_nombre': est['escenario_nombre'],
                    'indices': est['indices'],
                    'lead_dias': est['lead_dias'],
                    'cobertura_dias': est['cobertura_dias'],
                    'mes_objetivo': est['mes_objetivo'],
                    'mes_objetivo_label': est['mes_objetivo_label'],
                    'indice_aplicado': est['indice_aplicado'],
                    'ritmo_diario': est['ritmo_diario'],
                    'demanda_proyectada': est['demanda_proyectada'],
                    'flag': est['flag'],
                    'razon': est['razon'],
                    'min_diag': met['min_diag'],
                    'min_diag_label': met['min_diag_label'],
                    'perdida_mensual': met['perdida_mensual'],
                    'perdida_pesos': met['perdida_pesos'],
                    # Overrides operativos (chip UI siempre que existan en DB,
                    # incluso si no modificaron el sugerido).
                    'cant_fija': cant_fija_p,
                    'oferta_min': oferta_min_p,
                    'override_aplicado': ov_slug,  # 'cant_fija'|'oferta_min'|None
                    'override_valor': ov_valor,
                    'sugerido_prueba_original': sug_prueba_original,
                    'sugerido_dia_original': sug_dia_original,
                })
                total_prueba += est['sugerido_final']
                if sug_dia is not None:
                    total_dia += sug_dia
                if precio_pvp:
                    if sug_dia:
                        monto_dia += precio_pvp * sug_dia
                    monto_prueba += precio_pvp * est['sugerido_final']

            # Top 10 productos por pérdida $/mes (faltantes que más cuestan).
            con_perdida = [it for it in resultado if it['perdida_pesos'] > 0]
            top_perdida = sorted(con_perdida,
                                 key=lambda it: -it['perdida_pesos'])[:10]
            chart_perdida_pesos = {
                'labels': [it['producto_nombre'][:50] for it in top_perdida],
                'data': [it['perdida_pesos'] for it in top_perdida],
            }
            perdida_pesos_total = round(
                sum(it['perdida_pesos'] for it in resultado), 2)

            return jsonify({
                'lab_id': lab_id,
                'items': resultado,
                'total_dia': total_dia,
                'total_prueba': total_prueba,
                'delta_total': total_prueba - total_dia,
                'monto_dia': round(monto_dia, 2),
                'monto_prueba': round(monto_prueba, 2),
                'monto_delta': round(monto_prueba - monto_dia, 2),
                'perdida_pesos_total': perdida_pesos_total,
                'chart_perdida_pesos': chart_perdida_pesos,
                'lead_default': lead_default,
                'cob_default': cob_default,
                'stock_sync_en': stock_sync.isoformat() if stock_sync else None,
                'ventas_sync_en': ventas_sync.isoformat() if ventas_sync else None,
            })

    # Endpoint /escenario-producto eliminado tras merge con main:
    # la asignación producto → escenario ahora va por la tabla
    # EstacionalidadProducto y los endpoints de routes/estacionalidad.py:
    #   POST /api/estacionalidad/droga/<id>/aplicar   {escenario_id, producto_observer_ids: [...]}
    #   POST /api/estacionalidad/droga/<id>/desvincular  {producto_observer_ids: [...]}
    # El drawer del pedido_prueba solo lee + muestra; las ediciones de
    # estacionalidad se hacen desde /informes/estacionalidad-drogas.

    @app.route('/api/pedido-prueba/flag/<int:producto_id>', methods=['POST'])
    @login_required
    def api_pedido_prueba_flag(producto_id):
        """Aplica o quita un flag a un producto puntual via EAN principal.

        Body: { action: 'apply'|'remove', flag_slug: 'DISCONTINUADO'|... }
        """
        payload = request.get_json(silent=True) or {}
        action = payload.get('action')
        slug = payload.get('flag_slug')
        if action not in ('apply', 'remove') or not slug:
            return jsonify({'error': 'action y flag_slug requeridos'}), 400

        with get_db() as session:
            producto = (session.query(ObsProducto)
                        .filter_by(observer_id=producto_id).first())
            if not producto:
                return jsonify({'error': 'producto inexistente'}), 404

            # Validar tipo de flag
            tipo = (session.query(TipoPedidoConfig)
                    .filter_by(slug=slug, categoria='flag').first())
            if not tipo:
                return jsonify({'error': f'flag {slug} inexistente'}), 404

            # EAN principal
            from services.pedido_estacional import obtener_eans_producto
            eans = obtener_eans_producto(session, producto_id)
            if not eans:
                return jsonify({'error': 'producto sin EAN'}), 400
            ean = eans[0]

            if action == 'apply':
                # Crear si no existe
                existente = (session.query(ProductoFlag)
                             .filter_by(flag_slug=slug, ean=ean).first())
                if not existente:
                    session.add(ProductoFlag(
                        flag_slug=slug, ean=ean, laboratorio_id=None,
                    ))
                    session.commit()
            else:  # remove
                (session.query(ProductoFlag)
                 .filter_by(flag_slug=slug, ean=ean).delete())
                session.commit()

            return jsonify({'ok': True})

    @app.route('/api/pedido-prueba/export-xlsx', methods=['POST'])
    @login_required
    def api_pedido_prueba_export_xlsx():
        """Exporta a XLSX el resultado del calculo actual (lab + filtros).

        Body: mismo formato que /calcular. Reusa la logica para regenerar
        los items y los escribe a un workbook. No persiste nada.
        """
        from io import BytesIO

        import openpyxl
        from flask import send_file
        from openpyxl.styles import Alignment, Font, PatternFill

        # Re-ejecutar el calculo
        with app.test_request_context('/api/pedido-prueba/calcular',
                                       method='POST', json=request.get_json()):
            try:
                resp = api_pedido_prueba_calcular()
            except Exception as e:
                return jsonify({'error': f'calcular fallo: {e}'}), 500
            if isinstance(resp, tuple):
                resp = resp[0]
            data = resp.get_json()
        items = data.get('items', [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'PedidoPrueba'

        headers = ['Producto', 'Droga', 'Stock', 'u3m', 'u12m',
                   'Sug. dia act.', 'Sug. prueba', 'Δ',
                   'PVP', '$ Prueba', 'Origen', 'Flag']
        hdr_fill = PatternFill('solid', fgColor='1e1e1e')
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        for ri, it in enumerate(items, 2):
            ws.cell(row=ri, column=1, value=it.get('producto_nombre'))
            ws.cell(row=ri, column=2, value=it.get('droga_nombre'))
            ws.cell(row=ri, column=3, value=it.get('stock'))
            ws.cell(row=ri, column=4, value=it.get('u3m'))
            ws.cell(row=ri, column=5, value=it.get('u12m'))
            ws.cell(row=ri, column=6, value=it.get('sugerido_dia'))
            ws.cell(row=ri, column=7, value=it.get('sugerido_prueba'))
            ws.cell(row=ri, column=8, value=it.get('delta'))
            ws.cell(row=ri, column=9, value=it.get('precio_pvp'))
            ws.cell(row=ri, column=10, value=it.get('monto_prueba'))
            ws.cell(row=ri, column=11, value=it.get('origen_escenario'))
            ws.cell(row=ri, column=12,
                    value=(it['flag']['slug'] if it.get('flag') else ''))

        widths = [38, 28, 8, 8, 8, 12, 12, 8, 12, 14, 12, 18]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        # Fila final con totales
        totals_row = len(items) + 2
        ws.cell(row=totals_row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=totals_row, column=6, value=data.get('total_dia')).font = Font(bold=True)
        ws.cell(row=totals_row, column=7, value=data.get('total_prueba')).font = Font(bold=True)
        ws.cell(row=totals_row, column=8, value=data.get('delta_total')).font = Font(bold=True)
        ws.cell(row=totals_row, column=10, value=data.get('monto_prueba')).font = Font(bold=True)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Nombre archivo con lab_id + timestamp
        from datetime import datetime as _dt2
        lab_id = data.get('lab_id', 0)
        ts = _dt2.now().strftime('%Y%m%d-%H%M')
        fname = f'PedidoPrueba_lab{lab_id}_{ts}.xlsx'
        return send_file(
            buf, as_attachment=True, download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    # Endpoint /save eliminado en el pivot a "solo configuración".
    # La pantalla ya no arma pedidos; los escenarios y flags se persisten
    # individualmente por el drawer (escenario-producto) y por el endpoint
    # de flag. Si en el futuro vuelve a hacer falta armar pedido desde
    # aca, ver git history en branch feat/pedido-prueba (commit ef68334).
