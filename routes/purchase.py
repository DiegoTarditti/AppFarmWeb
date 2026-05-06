"""Purchase analysis, orders, and module/offer processing routes."""

import json
import os
import re
import uuid
from datetime import datetime

from flask import flash, jsonify, make_response, redirect, render_template, request, url_for
from werkzeug.utils import secure_filename

import database
from database import AnalisisSesion, ErpStock, Laboratorio, ModuloPack, Pedido, PedidoItem, Producto
from helpers import (
    PURCHASE_FOLDER,
    UPLOAD_FOLDER,
    _add_alt_barcode,
    _upsert_pedido_items,
    _upsert_producto,
    get_config,
    now_ar,
)
from parsers.sales_history import parse_sales_history_pdf
from parsers.sales_history_html import parse_sales_history_html
from parsers.sales_history_xls import parse_sales_history_xls
from purchase_engine import analyze_purchase

_PACK_PATTERN = re.compile(r'\bPACK\s*X\s*(\d+)\b', re.IGNORECASE)


def _detectar_packs_en_modulos(modules, session):
    """Clasifica cada ítem del módulo combinando 3 señales independientes:

      1. Destacado en amarillo en el Excel (criterio del vendedor).
      2. Regex 'PACK X N' en la descripción (explícito, aporta cantidad).
      3. Sin ventas históricas por ese EAN (un pack nunca se vende por su
         propio código, se vende la unidad individual).

    Un ítem es pack si tiene ≥1 señal. La confianza es proporcional a la
    suma. Para cada pack detectado busca su unidad equivalente: prioriza
    otro ítem del mismo módulo que SÍ tenga ventas y descripción base similar.

    Cada candidato: {ean_pack, desc_pack, cantidad, ean_unidad_sug,
                     desc_unidad_sug, fuente, modulo, destacado, tiene_regex,
                     tuvo_ventas (del pack), confianza ('alta'|'media'|'baja')}
    """
    from database import ModuloPack, ObsProducto, ObsVentaMensual, Producto
    ya_registrados = {ep for (ep,) in session.query(ModuloPack.ean_pack).all()}

    # 1. Juntar todos los EANs del archivo para bulk lookup
    todos_eans = set()
    for mod in modules or []:
        for it in mod.get('items') or mod.get('productos') or []:
            e = (it.get('ean') or '').strip()
            if e:
                todos_eans.add(e)

    # 2. Map EAN → observer_id via tabla productos local
    ean_a_obs = dict(
        session.query(Producto.codigo_barra, Producto.observer_id)
        .filter(Producto.codigo_barra.in_(todos_eans),
                Producto.observer_id.isnot(None)).all()
    )

    # 3. Set de observer_ids con ventas > 0 en los últimos 12 meses
    from datetime import datetime
    hoy = datetime.now()
    obs_ids = {oid for oid in ean_a_obs.values() if oid}
    con_ventas = set()
    if obs_ids:
        rows = (session.query(ObsVentaMensual.producto_observer)
                .filter(ObsVentaMensual.producto_observer.in_(obs_ids),
                        ObsVentaMensual.unidades > 0)
                .distinct().all())
        con_ventas = {r[0] for r in rows}

    def tuvo_ventas(ean):
        oid = ean_a_obs.get(ean)
        if oid is None:  # Sin registro local: probablemente pack (nunca se vendió)
            return False
        return oid in con_ventas

    candidatos = []
    for mod in modules or []:
        items = mod.get('items') or mod.get('productos') or []
        for it in items:
            desc = (it.get('desc') or it.get('descripcion') or '').strip()
            ean_pack = (it.get('ean') or '').strip()
            if not ean_pack or not desc or ean_pack in ya_registrados:
                continue
            destacado = bool(it.get('destacado'))
            m = _PACK_PATTERN.search(desc)
            sin_ventas = not tuvo_ventas(ean_pack)

            # Requerimos al menos 1 señal fuerte.
            senales = sum([destacado, bool(m), sin_ventas])
            if senales == 0:
                continue

            # Cantidad: del regex si hay, sino None (user la completa)
            try:
                cantidad = int(m.group(1)) if m else None
            except (ValueError, TypeError):
                cantidad = None

            # Confianza:
            # - alta:  2+ señales
            # - media: 1 señal fuerte (destacado o regex)
            # - baja:  solo "sin ventas" (podría ser un producto nuevo que nadie vendió aún)
            if senales >= 2:
                confianza = 'alta'
            elif destacado or m:
                confianza = 'media'
            else:
                confianza = 'baja'

            # Buscar unidad equivalente
            base = re.sub(r'\s*\(?\s*PACK\s*X\s*\d+\s*\)?\s*', ' ', desc, flags=re.I).strip()
            base_toks = {t for t in re.split(r'\s+', base.lower()) if len(t) >= 2}

            unidad_ean = unidad_desc = None
            fuente = 'none'
            # Preferimos items del mismo módulo que SÍ tengan ventas
            candidatos_unidad = []
            for it2 in items:
                d2 = (it2.get('desc') or it2.get('descripcion') or '').strip()
                e2 = (it2.get('ean') or '').strip()
                if not e2 or not d2 or e2 == ean_pack:
                    continue
                toks2 = {t for t in re.split(r'\s+', d2.lower()) if len(t) >= 2}
                inter = base_toks & toks2
                if len(inter) >= max(2, int(len(base_toks) * 0.5)):
                    score = len(inter) / max(len(base_toks | toks2), 1)
                    # Bonus si tuvo ventas (es la unidad que se vende)
                    if tuvo_ventas(e2):
                        score += 0.5
                    candidatos_unidad.append((score, e2, d2))
            if candidatos_unidad:
                candidatos_unidad.sort(key=lambda x: -x[0])
                unidad_ean, unidad_desc = candidatos_unidad[0][1], candidatos_unidad[0][2]
                fuente = 'modulo'

            # Fallback: obs_productos
            if not unidad_ean and base_toks:
                primer = next(iter(sorted(base_toks, key=len, reverse=True)), '')
                if len(primer) >= 3:
                    q = session.query(ObsProducto).filter(
                        ObsProducto.descripcion.ilike(f'%{primer}%'),
                        ObsProducto.fecha_baja.is_(None),
                    ).limit(50).all()
                    best, best_score = None, 0
                    for op in q:
                        toks_op = {t for t in re.split(r'\s+', op.descripcion.lower()) if len(t) >= 2}
                        if not toks_op:
                            continue
                        score = len(base_toks & toks_op) / len(base_toks | toks_op)
                        if score > best_score:
                            best, best_score = op, score
                    if best and best_score >= 0.4:
                        unidad_ean = str(best.observer_id)
                        unidad_desc = best.descripcion
                        fuente = 'catalogo'

            candidatos.append({
                'ean_pack':        ean_pack,
                'desc_pack':       desc,
                'cantidad':        cantidad,
                'ean_unidad_sug':  unidad_ean or '',
                'desc_unidad_sug': unidad_desc or '',
                'fuente':          fuente,
                'modulo':          mod.get('nombre') or '',
                'destacado':       destacado,
                'tiene_regex':     bool(m),
                'sin_ventas':      sin_ventas,
                'confianza':       confianza,
            })
    return candidatos


def _analyze_sales_file(tmp_path, ext, n_days):
    """Procesa un único archivo de estadística de ventas.
    Devuelve dict con uid/laboratorio/productos/periodo o {'error': str}."""
    try:
        if ext == 'pdf':
            parsed = parse_sales_history_pdf(tmp_path)
        elif ext in ('html', 'htm'):
            parsed = parse_sales_history_html(tmp_path)
        else:
            parsed = parse_sales_history_xls(tmp_path)
    except Exception as e:
        return {'error': f'Error al parsear: {e}'}

    if not parsed.get('products'):
        return {'error': 'No se encontraron productos en el archivo.'}

    cfg = get_config()
    results = analyze_purchase(
        parsed['products'], n_days,
        parsed['start_month'], parsed['end_month'],
        umbral_pico=cfg['umbral_pico'],
        umbral_baja=cfg['umbral_baja'],
        umbral_tendencia=cfg['umbral_tendencia'],
        rot_alta_min=cfg['rot_alta_min'],
        rot_media_min=cfg['rot_media_min'],
    )

    uid = str(uuid.uuid4())
    data = {
        'uid': uid,
        'farmacia': parsed['farmacia'],
        'laboratorio': parsed['laboratorio'],
        'periodo': parsed['periodo'],
        'start_month': parsed.get('start_month', 4),
        'n_days': n_days,
        'umbral_tendencia': cfg['umbral_tendencia'],
        'rot_alta_min': cfg['rot_alta_min'],
        'rot_alta_tol': cfg['rot_alta_tol'],
        'rot_media_min': cfg['rot_media_min'],
        'rot_media_tol': cfg['rot_media_tol'],
        'rot_baja_tol': cfg['rot_baja_tol'],
        'products': results,
    }
    json_path = os.path.join(PURCHASE_FOLDER, f'{uid}.json')
    with open(json_path, 'w', encoding='utf-8') as jf:
        json.dump(data, jf, ensure_ascii=False)

    _snapshot_product_analytics(results, parsed.get('laboratorio'),
                                start_month=parsed.get('start_month', 4),
                                n_days=n_days)

    fuente = ext if ext in ('pdf', 'xls', 'xlsx', 'html', 'htm') else 'pdf'
    if fuente in ('xlsx', 'xls'):
        fuente = 'xls'
    sesion_id = _create_analisis_sesion(
        laboratorio=parsed.get('laboratorio'),
        periodo=parsed.get('periodo'),
        farmacia=parsed.get('farmacia'),
        n_days=n_days,
        fuente=fuente,
        n_productos=len(results),
    )
    data['sesion_id'] = sesion_id
    with open(json_path, 'w', encoding='utf-8') as jf:
        json.dump(data, jf, ensure_ascii=False)

    return {
        'uid': uid,
        'laboratorio': parsed.get('laboratorio') or '(sin laboratorio)',
        'periodo': parsed.get('periodo') or '',
        'count': len(results),
        'sesion_id': sesion_id,
    }


def _create_analisis_sesion(laboratorio, periodo, farmacia, n_days, fuente, n_productos):
    """Crea un registro AnalisisSesion y retorna su id."""
    with database.get_db() as session:
        sesion = AnalisisSesion(
            laboratorio_nombre=laboratorio or '',
            periodo=periodo or '',
            farmacia=farmacia or '',
            n_days=n_days,
            fuente=fuente,
            n_productos=n_productos,
        )
        session.add(sesion)
        session.commit()
        return sesion.id


def _snapshot_product_analytics(results, laboratorio, start_month=4, n_days=35):
    """Upsert de ProductAnalytics por codigo_barra desde los resultados de analyze_purchase."""
    from datetime import datetime as _dt
    seen = {}
    for p in results:
        cb = (p.get('codigo_barra') or '').strip()
        if cb:
            seen[cb] = p

    if not seen:
        return

    with database.get_db() as session:
        existing = {
            pa.codigo_barra: pa
            for pa in session.query(database.ProductAnalytics)
            .filter(database.ProductAnalytics.codigo_barra.in_(seen.keys())).all()
        }
        for cb, p in seen.items():
            forecast = p.get('forecast')
            forecast_next = None
            if isinstance(forecast, list) and forecast:
                forecast_next = forecast[0]
            elif isinstance(forecast, (int, float)):
                forecast_next = forecast
            pa = existing.get(cb)
            if pa is None:
                pa = database.ProductAnalytics(codigo_barra=cb)
                session.add(pa)
            pa.descripcion = (p.get('nombre') or p.get('descripcion') or '')[:200]
            pa.laboratorio = laboratorio
            pa.stock = int(p.get('stock') or 0)
            pa.avg_monthly = float(p.get('avg_monthly') or 0)
            pa.rotacion = p.get('rotacion')
            pa.slope = float(p.get('slope') or 0)
            pa.forecast_next = float(forecast_next) if forecast_next is not None else None
            pa.sin_mov_60d = 1 if p.get('sin_mov_60d') else 0
            pa.precio_pvp = float(p.get('precio_pvp') or 0)
            pa.tipo = p.get('tipo') or 'N'
            ventas = p.get('ventas')
            if ventas and isinstance(ventas, list):
                pa.ventas_json = json.dumps(ventas)
                pa.start_month = start_month
                pa.n_days = n_days
            pa.actualizado_en = now_ar()
        session.commit()


def init_app(app):

    @app.route('/purchase')
    def purchase_index():
        import observer_source
        return render_template('purchase_analysis.html',
                               observer_disponible=observer_source.observer_disponible())

    @app.route('/purchase/analyze', methods=['POST'])
    def purchase_analyze():
        f = request.files.get('sales_pdf')
        try:
            n_days = max(1, min(365, int(request.form.get('n_days', 35))))
        except (ValueError, TypeError):
            n_days = 35

        if not f or not f.filename:
            flash('Seleccioná un archivo PDF o Excel.')
            return redirect(url_for('purchase_index'))

        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        if ext not in ('pdf', 'xlsx', 'xls', 'html', 'htm'):
            flash('Formato no soportado. Usá PDF, Excel (.xlsx / .xls) o HTML.')
            return redirect(url_for('purchase_index'))

        filename = secure_filename(f.filename)
        tmp_path = os.path.join(UPLOAD_FOLDER, f'purchase_{filename}')
        f.save(tmp_path)

        try:
            res = _analyze_sales_file(tmp_path, ext, n_days)
        finally:
            try: os.remove(tmp_path)
            except OSError: pass

        if 'error' in res:
            flash(res['error'])
            return redirect(url_for('purchase_index'))

        return redirect(url_for('purchase_results', uid=res['uid']))

    @app.route('/purchase/processed')
    def purchase_processed():
        """Lista todos los análisis de ventas ya procesados (JSONs en PURCHASE_FOLDER)."""
        from datetime import datetime as _dt
        items = []
        try:
            for fn in os.listdir(PURCHASE_FOLDER):
                if not fn.endswith('.json'):
                    continue
                path = os.path.join(PURCHASE_FOLDER, fn)
                try:
                    with open(path, encoding='utf-8') as jf:
                        d = json.load(jf)
                    items.append({
                        'uid': d.get('uid') or fn[:-5],
                        'laboratorio': d.get('laboratorio') or '(sin laboratorio)',
                        'periodo': d.get('periodo') or '',
                        'n_days': d.get('n_days') or 0,
                        'count': len(d.get('products') or []),
                        'mtime': _dt.fromtimestamp(os.path.getmtime(path)),
                    })
                except Exception:
                    continue
        except FileNotFoundError:
            pass
        items.sort(key=lambda x: x['mtime'], reverse=True)
        return render_template('purchase_processed.html', items=items)

    @app.route('/purchase/batch', methods=['POST'])
    def purchase_batch():
        files = request.files.getlist('sales_files')
        try:
            n_days = max(1, min(365, int(request.form.get('n_days', 35))))
        except (ValueError, TypeError):
            n_days = 35

        files = [f for f in files if f and f.filename]
        if not files:
            flash('Seleccioná al menos un archivo.')
            return redirect(url_for('purchase_index'))

        results = []
        for f in files:
            ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
            if ext not in ('pdf', 'xlsx', 'xls', 'html', 'htm'):
                results.append({'filename': f.filename, 'error': f'Formato no soportado (.{ext})'})
                continue

            filename = secure_filename(f.filename)
            tmp_path = os.path.join(UPLOAD_FOLDER, f'batch_{uuid.uuid4().hex}_{filename}')
            f.save(tmp_path)
            try:
                res = _analyze_sales_file(tmp_path, ext, n_days)
            except Exception as e:
                app.logger.exception('Error procesando %s', f.filename)
                res = {'error': str(e)}
            finally:
                try: os.remove(tmp_path)
                except OSError: pass

            res['filename'] = f.filename
            results.append(res)

        ok = [r for r in results if 'uid' in r]
        fail = [r for r in results if 'error' in r]
        return render_template('purchase_batch.html',
                               results=results, ok=ok, fail=fail, n_days=n_days)

    @app.route('/purchase/results/<uid>')
    def purchase_results(uid):
        if not re.match(r'^[0-9a-f-]{36}$', uid):
            flash('Sesión inválida.')
            return redirect(url_for('purchase_index'))

        json_path = os.path.join(PURCHASE_FOLDER, f'{uid}.json')
        if not os.path.exists(json_path):
            flash('La sesión expiró o no existe. Analizá el PDF nuevamente.')
            return redirect(url_for('purchase_index'))

        with open(json_path, encoding='utf-8') as jf:
            data = json.load(jf)

        cfg = get_config()
        data.setdefault('umbral_tendencia', cfg['umbral_tendencia'])
        data.setdefault('rot_alta_min', cfg['rot_alta_min'])
        data.setdefault('rot_alta_tol', cfg['rot_alta_tol'])
        data.setdefault('rot_media_min', cfg['rot_media_min'])
        data.setdefault('rot_media_tol', cfg['rot_media_tol'])
        data.setdefault('rot_baja_tol', cfg['rot_baja_tol'])

        with database.get_db() as session:
            barcodes = [p['codigo_barra'] for p in data.get('products', []) if p.get('codigo_barra')]
            pack_eans = {mp.ean_pack for mp in session.query(ModuloPack).all()}
            prods_pack = {
                p.codigo_barra: bool(p.es_pack)
                for p in session.query(Producto).filter(
                    Producto.codigo_barra.in_(barcodes)
                ).all()
            }
            # Resolver monodroga_id por producto (para botón "Comparar otros labs").
            # Si el item ya trae observer_id (análisis ObServer), lookup directo;
            # sino, vía Producto.observer_id local. Idempotente para análisis viejos.
            obs_ids_a_buscar = set()
            for p in data.get('products', []):
                if p.get('observer_id'):
                    obs_ids_a_buscar.add(int(p['observer_id']))
            cbs_sin_obs = [p['codigo_barra'] for p in data.get('products', [])
                           if p.get('codigo_barra') and not p.get('observer_id')]
            cb_to_obs = {}
            if cbs_sin_obs:
                for prod in (session.query(Producto)
                             .filter(Producto.codigo_barra.in_(cbs_sin_obs),
                                     Producto.observer_id.isnot(None)).all()):
                    cb_to_obs[prod.codigo_barra] = prod.observer_id
                    obs_ids_a_buscar.add(prod.observer_id)
            droga_by_obs = {}
            if obs_ids_a_buscar:
                droga_by_obs = dict(
                    session.query(database.ObsProducto.observer_id,
                                  database.ObsProducto.nombre_droga_observer)
                    .filter(database.ObsProducto.observer_id.in_(list(obs_ids_a_buscar)),
                            database.ObsProducto.nombre_droga_observer.isnot(None)).all()
                )

            for p in data.get('products', []):
                cb = p.get('codigo_barra', '')
                p['es_pack'] = prods_pack.get(cb, False) or (cb in pack_eans)
                obs_id = p.get('observer_id') or cb_to_obs.get(cb)
                p['monodroga_id'] = droga_by_obs.get(obs_id) if obs_id else None

        _mes_jan = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        sm = data.get('start_month', 4)
        month_es = [_mes_jan[(sm - 1 + i) % 12] for i in range(12)]
        analizado_en = datetime.fromtimestamp(os.path.getmtime(json_path)).strftime('%d/%m/%Y')
        return render_template('purchase_results.html', month_es=month_es, analizado_en=analizado_en, **data)

    @app.route('/purchase/export/<uid>/<fmt>', methods=['POST'])
    def purchase_export(uid, fmt):
        if not re.match(r'^[0-9a-f-]{36}$', uid):
            return 'UID inválido', 400

        json_path = os.path.join(PURCHASE_FOLDER, f'{uid}.json')
        if not os.path.exists(json_path):
            flash('La sesión expiró.')
            return redirect(url_for('purchase_index'))

        with open(json_path, encoding='utf-8') as jf:
            data = json.load(jf)

        for i, p in enumerate(data['products']):
            edited = request.form.get(f'qty_{i}')
            if edited is not None:
                try:
                    qty = int(edited)
                    p['order_qty'] = max(0, qty)
                    p['subtotal'] = round(p['order_qty'] * p['precio_pvp'], 2)
                except ValueError:
                    pass

        lab = data.get('laboratorio', 'Compra')
        n = data.get('n_days', 35)
        periodo = data.get('periodo', '')
        farmacia_nombre = data.get('farmacia') or get_config()['farmacia_nombre']

        if fmt == 'xlsx':
            from io import BytesIO

            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Compra {lab}"[:31]

            ws['A1'] = f"Plan de compra — {lab}"
            ws['A1'].font = Font(bold=True, size=13)
            ws['A2'] = periodo
            ws['A3'] = f"Días proyectados: {n}"
            ws.append([])

            headers = ['Producto', 'Cód. Barras', 'P.PVP ($)', 'Stock',
                       'Prom/mes', f'Pronóstico ({n}d)', 'Pedido', 'Subtotal ($)', 'Comentario']
            ws.append(headers)

            hdr_row = ws.max_row
            hdr_fill = PatternFill('solid', fgColor='1C1C1E')
            hdr_font = Font(bold=True, color='EAB308')
            thin = Side(style='thin', color='3A3A3C')
            border = Border(bottom=thin)
            for cell in ws[hdr_row]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            for p in data['products']:
                if p['order_qty'] == 0 and p['total'] == 0:
                    continue
                ws.append([
                    p['nombre'], p['codigo_barra'],
                    p['precio_pvp'], p['stock'],
                    p['avg_monthly'], p['forecast'],
                    p['order_qty'], p['subtotal'],
                    p['comment'],
                ])

            widths = [40, 16, 12, 8, 10, 14, 10, 14, 50]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)

            resp = make_response(buf.read())
            resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            resp.headers['Content-Disposition'] = f'attachment; filename="Compra_{lab}_{n}d.xlsx"'
            return resp

        elif fmt == 'pdf':
            from datetime import datetime as _dt
            from io import BytesIO

            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.pdfgen import canvas as rl_canvas
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            fecha_emision = _dt.now().strftime('%d/%m/%Y %H:%M')
            page_w, page_h = landscape(A4)

            class _NumberedCanvas(rl_canvas.Canvas):
                def __init__(self, *args, **kwargs):
                    rl_canvas.Canvas.__init__(self, *args, **kwargs)
                    self._pages = []
                def showPage(self):
                    self._pages.append(dict(self.__dict__))
                    self._startPage()
                def save(self):
                    total = len(self._pages)
                    for i, state in enumerate(self._pages, 1):
                        self.__dict__.update(state)
                        self._draw_footer(i, total)
                        rl_canvas.Canvas.showPage(self)
                    rl_canvas.Canvas.save(self)
                def _draw_footer(self, page_num, total):
                    self.saveState()
                    self.setFont('Helvetica', 7)
                    self.setFillColor(colors.HexColor('#6B7280'))
                    self.drawString(1.5*cm, 0.6*cm, f"Emitido: {fecha_emision}")
                    self.drawRightString(page_w - 1.5*cm, 0.6*cm,
                                         f"Página {page_num} de {total}")
                    self.restoreState()

            buf = BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                    leftMargin=1.5*cm, rightMargin=1.5*cm,
                                    topMargin=1.5*cm, bottomMargin=1.5*cm)

            styles = getSampleStyleSheet()
            hdr_bg  = colors.HexColor('#2D3748')
            row_a   = colors.white
            row_b   = colors.HexColor('#F3F4F6')
            txt     = colors.HexColor('#1A202C')
            accent  = colors.HexColor('#EAB308')
            grid_c  = colors.HexColor('#D1D5DB')

            title_style = ParagraphStyle('title', parent=styles['Normal'],
                                         fontSize=14, textColor=accent, spaceAfter=4)
            sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                       fontSize=9, textColor=txt, spaceAfter=2)
            cell_style = ParagraphStyle('cell', parent=styles['Normal'],
                                        fontSize=7, textColor=txt, leading=9)

            story = [
                Paragraph(f"Plan de compra — {lab}", title_style),
                Paragraph(f"{farmacia_nombre} · {periodo}", sub_style),
                Paragraph(f"Proyección: {n} días", sub_style),
                Spacer(1, 0.4*cm),
            ]

            headers = ['Producto', 'Barcode', 'P.PVP', 'Stock',
                       'Prom/m', 'Tendencia', 'Baja', 'Pico', 'Pedido', 'Subtotal', 'Nota']
            rows = [headers]

            for p in data['products']:
                if p['order_qty'] == 0 and p['total'] == 0:
                    continue
                slope = p.get('slope', 0)
                tend = (f"↑ {slope}" if slope > 0.2 else f"↓ {abs(slope)}" if slope < -0.2 else f"{slope}")
                stock_str = 'agotado' if p['stock'] <= 0 else str(p['stock'])
                rows.append([
                    Paragraph(p['nombre'], cell_style),
                    p['codigo_barra'],
                    f"${p['precio_pvp']:,.0f}",
                    stock_str,
                    p['avg_monthly'],
                    tend,
                    p.get('low_month', '') or '—',
                    p.get('peak_month', '') or '—',
                    p['order_qty'],
                    f"${p['subtotal']:,.0f}",
                    Paragraph(p.get('comment', ''), cell_style),
                ])

            col_widths = [4.8*cm, 3.0*cm, 2.0*cm, 1.6*cm, 1.5*cm, 1.8*cm, 1.4*cm, 1.4*cm, 1.5*cm, 2.2*cm, 3.5*cm]
            t = Table(rows, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0), (-1, 0), hdr_bg),
                ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
                ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0), (-1, 0), 8),
                ('ROWBACKGROUNDS',(0, 1), (-1, -1), [row_a, row_b]),
                ('TEXTCOLOR',     (0, 1), (-1, -1), txt),
                ('FONTSIZE',      (0, 1), (-1, -1), 7),
                ('ALIGN',         (2, 0), (-1, -1), 'RIGHT'),
                ('ALIGN',         (0, 0), (1, -1), 'LEFT'),
                ('ALIGN',         (6, 0), (7, -1), 'CENTER'),
                ('GRID',          (0, 0), (-1, -1), 0.3, grid_c),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING',    (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.append(t)

            doc.build(story, canvasmaker=_NumberedCanvas)
            buf.seek(0)

            resp = make_response(buf.read())
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Content-Disposition'] = f'attachment; filename="Compra_{lab}_{n}d.pdf"'
            return resp

        return 'Formato no soportado', 400

    # ─── PEDIDOS ──────────────────────────────────────────────────────────────

    @app.route('/purchase/save-order/<uid>', methods=['POST'])
    def purchase_save_order(uid):
        if not re.match(r'^[0-9a-f-]{36}$', uid):
            flash('UID inválido.')
            return redirect(url_for('purchase_index'))

        json_path = os.path.join(PURCHASE_FOLDER, f'{uid}.json')
        if not os.path.exists(json_path):
            flash('La sesión expiró. Analizá el PDF nuevamente.')
            return redirect(url_for('purchase_index'))

        with database.get_db() as session:
            try:
                with open(json_path, encoding='utf-8') as jf:
                    data = json.load(jf)

                products = data.get('products', [])
                items = []
                for i, p in enumerate(products):
                    try:
                        qty = int(request.form.get(f'qty_{i}') or 0)
                    except (ValueError, TypeError):
                        qty = 0
                    if qty > 0:
                        precio = float(p.get('precio_pvp') or 0)
                        cb = (p.get('codigo_barra') or '').strip()
                        obs_id = p.get('observer_id')
                        # Resolver EAN real desde obs_codigos_barras si el item no
                        # lo trae (lookup directo, ya no hace falta pseudo-EAN).
                        if not cb and obs_id:
                            ean_real = (session.query(database.ObsCodigoBarras.codigo_barras)
                                        .filter(database.ObsCodigoBarras.producto_observer == obs_id,
                                                database.ObsCodigoBarras.orden == 1,
                                                database.ObsCodigoBarras.fecha_baja.is_(None))
                                        .first())
                            cb = ean_real[0] if ean_real else ''
                        items.append(PedidoItem(
                            codigo_barra=cb,
                            nombre=p.get('nombre', ''),
                            cantidad=qty,
                            precio_pvp=precio,
                            subtotal=round(qty * precio, 2),
                            rotacion=p.get('rotacion') or None,
                            avg_monthly=p.get('avg_monthly') or None,
                        ))

                if not items:
                    flash('No hay productos con cantidad > 0 para guardar.')
                    return redirect(url_for('purchase_results', uid=uid))

                pedido = Pedido(
                    laboratorio=data.get('laboratorio', ''),
                    farmacia=data.get('farmacia', ''),
                    periodo=data.get('periodo', ''),
                    n_days=data.get('n_days', 0),
                    analisis_sesion_id=data.get('sesion_id'),
                    items=items,
                )
                session.add(pedido)
                _upsert_pedido_items(session, items, observer_bridge=True)
                session.flush()  # para tener pedido.id

                # Si viene de un proceso → asociar + avanzar pasos 1 y 2
                proceso_id = request.form.get('proceso_id', type=int) or data.get('proceso_id')
                if proceso_id:
                    proc = session.get(database.ProcesoCompra, proceso_id)
                    if proc:
                        proc.pedido_id = pedido.id
                        if data.get('sesion_id'):
                            proc.analisis_sesion_id = data.get('sesion_id')
                        if not proc.analisis_hecho_en:
                            proc.analisis_hecho_en = now_ar()
                        if not proc.pedido_hecho_en:
                            proc.pedido_hecho_en = now_ar()
                        if proc.estado in ('BORRADOR', None):
                            proc.estado = 'PEDIDO'
                        proc.actualizado_en = now_ar()
                session.commit()

                if proceso_id:
                    flash(f'Pedido guardado y asociado al proceso #{proceso_id}.', 'success')
                    return redirect(url_for('proceso_detail', proceso_id=proceso_id))
                flash(f'Pedido guardado: {len(items)} productos.')
                return redirect(url_for('orders_list'))
            except Exception as e:
                session.rollback()
                app.logger.exception('Error en purchase_save_order')
                flash(f'Error al guardar el pedido: {e}')
                return redirect(url_for('purchase_results', uid=uid))

    @app.route('/purchase/suggest', methods=['GET'])
    def purchase_suggest():
        """Sugerencia de pedido consolidado por laboratorio."""
        import math

        from sqlalchemy import func as _func

        try:
            threshold_days = max(1, min(365, int(request.args.get('threshold_days', 10))))
        except (ValueError, TypeError):
            threshold_days = 10
        try:
            target_days = max(1, min(365, int(request.args.get('target_days', 10))))
        except (ValueError, TypeError):
            target_days = 10
        calcular = request.args.get('calcular') == '1'

        groups = []
        total_items = 0
        total_importe = 0.0

        if calcular:
            with database.get_db() as session:
                PA = database.ProductAnalytics
                rows = session.query(PA).filter(
                    PA.avg_monthly > 0,
                    PA.stock * 30.0 / PA.avg_monthly < threshold_days
                ).order_by(PA.laboratorio.asc(), PA.descripcion.asc()).all()

                by_lab = {}
                for p in rows:
                    avg = float(p.avg_monthly or 0)
                    if avg <= 0:
                        continue
                    daily = avg / 30.0
                    target_stock = daily * target_days
                    suggested = max(0, int(math.floor(target_stock - (p.stock or 0))))
                    if suggested <= 0:
                        continue
                    pvp = float(p.precio_pvp or 0)
                    cov = round((p.stock or 0) * 30.0 / avg, 1) if avg > 0 else None
                    item = {
                        'codigo_barra': p.codigo_barra,
                        'descripcion': p.descripcion or '',
                        'stock': p.stock or 0,
                        'avg_monthly': round(avg, 1),
                        'rotacion': p.rotacion,
                        'tipo': p.tipo or 'N',
                        'cobertura': cov,
                        'sugerido': suggested,
                        'precio_pvp': pvp,
                        'subtotal': round(suggested * pvp, 2),
                    }
                    lab = p.laboratorio or '(sin laboratorio)'
                    by_lab.setdefault(lab, []).append(item)

                for lab in sorted(by_lab.keys()):
                    lab_items = by_lab[lab]
                    lab_total = sum(it['subtotal'] for it in lab_items)
                    lab_units = sum(it['sugerido'] for it in lab_items)
                    total_items += len(lab_items)
                    total_importe += lab_total
                    groups.append({
                        'laboratorio': lab,
                        'productos': lab_items,
                        'lab_total': round(lab_total, 2),
                        'lab_units': lab_units,
                    })

        return render_template('purchase_suggest.html',
                               threshold_days=threshold_days,
                               target_days=target_days,
                               calcular=calcular,
                               groups=groups,
                               total_items=total_items,
                               total_importe=round(total_importe, 2))

    @app.route('/purchase/suggest/create-order', methods=['POST'])
    def purchase_suggest_create_order():
        """Crea un Pedido para un laboratorio con los ítems seleccionados."""
        laboratorio = (request.form.get('laboratorio') or '').strip()
        if not laboratorio:
            flash('Laboratorio faltante.')
            return redirect(url_for('purchase_suggest'))

        with database.get_db() as session:
            try:
                selected = request.form.getlist('sel')
                items = []
                for cb in selected:
                    try:
                        qty = int(request.form.get(f'qty_{cb}') or 0)
                    except (ValueError, TypeError):
                        qty = 0
                    if qty <= 0:
                        continue
                    nombre = request.form.get(f'nom_{cb}') or ''
                    try:
                        precio = float(request.form.get(f'pvp_{cb}') or 0)
                    except (ValueError, TypeError):
                        precio = 0.0
                    rotacion = request.form.get(f'rot_{cb}') or None
                    try:
                        avg = float(request.form.get(f'avg_{cb}') or 0)
                    except (ValueError, TypeError):
                        avg = 0.0
                    items.append(PedidoItem(
                        codigo_barra=cb,
                        nombre=nombre[:200],
                        cantidad=qty,
                        precio_pvp=precio,
                        subtotal=round(qty * precio, 2),
                        rotacion=rotacion,
                        avg_monthly=avg or None,
                    ))

                if not items:
                    flash('No seleccionaste productos con cantidad > 0.')
                    return redirect(url_for('purchase_suggest', calcular=1))

                pedido = Pedido(
                    laboratorio=laboratorio[:150],
                    farmacia='',
                    periodo='Sugerido',
                    n_days=0,
                    items=items,
                )
                session.add(pedido)
                _upsert_pedido_items(session, items)
                session.commit()
                flash(f'Pedido creado para {laboratorio}: {len(items)} productos.')
                return redirect(url_for('orders_list'))
            except Exception as e:
                session.rollback()
                app.logger.exception('Error en purchase_suggest_create_order')
                flash(f'Error al crear pedido: {e}')
                return redirect(url_for('purchase_suggest', calcular=1))

    @app.route('/orders')
    def orders_list():
        from collections import defaultdict

        from sqlalchemy.orm import joinedload

        from routes.compras_dia import _sigla_drog
        with database.get_db() as session:
            pedidos = (session.query(Pedido)
                       .options(joinedload(Pedido.items))
                       .order_by(Pedido.creado_en.desc()).all())

            # Construir mapas de items emitidos PENDIENTES (no recibidos):
            #   observer_id → [{cantidad, fecha, drog_sigla, drog_nombre, pedido_id}]
            #   producto_id_local → idem
            # Sirve para cruzar contra cada producto de un Pedido guardado y
            # mostrar "ya pediste X u el dd/mm a {drog}".
            pend_q = (session.query(database.PedidoEmitidoItem)
                      .options(joinedload(database.PedidoEmitidoItem.pedido)
                               .joinedload(database.PedidoEmitido.drogueria))
                      .filter(database.PedidoEmitidoItem.estado == 'PENDIENTE')
                      .all())
            pend_by_obs = defaultdict(list)
            pend_by_prod = defaultdict(list)
            for pi in pend_q:
                drog_nombre = pi.pedido.drogueria.razon_social if pi.pedido and pi.pedido.drogueria else ''
                info = {
                    'cantidad': pi.cantidad_pedida,
                    'fecha': pi.pedido.fecha.strftime('%d/%m') if pi.pedido and pi.pedido.fecha else '',
                    'drog_sigla': _sigla_drog(drog_nombre),
                    'drog_nombre': drog_nombre,
                    'pedido_id': pi.pedido_id,
                }
                if pi.observer_id:
                    pend_by_obs[pi.observer_id].append(info)
                if pi.producto_id_local:
                    pend_by_prod[pi.producto_id_local].append(info)

            # Pre-resolver codigo_barra → (producto_id_local, observer_id) en lote
            # para todos los productos de todos los pedidos guardados (1 query).
            todos_codigos = set()
            for p in pedidos:
                for it in p.items:
                    if it.codigo_barra:
                        todos_codigos.add(it.codigo_barra.strip())
            cb_to_prod = {}
            if todos_codigos:
                # Match por codigo_barra principal
                for prod in (session.query(database.Producto)
                             .filter(database.Producto.codigo_barra.in_(todos_codigos)).all()):
                    cb_to_prod[prod.codigo_barra] = (prod.id, prod.observer_id)
                # Match adicional por tabla 1-a-N
                faltantes = todos_codigos - set(cb_to_prod.keys())
                if faltantes:
                    rows = (session.query(database.ProductoCodigoBarra.codigo_barra,
                                           database.Producto.id,
                                           database.Producto.observer_id)
                            .join(database.Producto,
                                  database.Producto.id == database.ProductoCodigoBarra.producto_id)
                            .filter(database.ProductoCodigoBarra.codigo_barra.in_(faltantes)).all())
                    for cb, pid, oid in rows:
                        cb_to_prod.setdefault(cb, (pid, oid))

            def _pendientes_de(cb):
                if not cb:
                    return []
                pid_oid = cb_to_prod.get(cb.strip())
                if not pid_oid:
                    return []
                pid, oid = pid_oid
                # Dedup por pedido_id (puede aparecer doble si matchea por ambos)
                vistos = set()
                out = []
                for src in (pend_by_prod.get(pid, []), pend_by_obs.get(oid, []) if oid else []):
                    for inf in src:
                        if inf['pedido_id'] in vistos:
                            continue
                        vistos.add(inf['pedido_id'])
                        out.append(inf)
                return out

            result = []
            for p in pedidos:
                total_unidades = sum(it.cantidad for it in p.items)
                total_importe = sum(float(it.subtotal or 0) for it in p.items)
                # Resolver nombre de la droguería si el canal ya fue elegido
                canal_partner_nombre = None
                if p.canal == 'drogueria' and p.partner_id:
                    _prov = session.get(database.Provider, p.partner_id)
                    if _prov:
                        canal_partner_nombre = _prov.razon_social
                result.append({
                    'id': p.id,
                    'laboratorio': p.laboratorio,
                    'farmacia': p.farmacia,
                    'periodo': p.periodo,
                    'n_days': p.n_days,
                    'creado_en': p.creado_en.strftime('%d/%m/%Y %H:%M') if p.creado_en else '',
                    'analizado_en': p.analizado_en.strftime('%d/%m/%Y') if p.analizado_en else '',
                    'estado': p.estado,
                    'tiene_analisis_guardado': bool(p.analisis_json),
                    'analisis_guardado_en': p.analisis_guardado_en.strftime('%d/%m/%Y %H:%M') if p.analisis_guardado_en else '',
                    'mostrar_hasta': p.mostrar_hasta.strftime('%Y-%m-%d') if p.mostrar_hasta else '',
                    'mostrar_hasta_label': p.mostrar_hasta.strftime('%d/%m/%y') if p.mostrar_hasta else '',
                    'canal': p.canal,
                    'canal_partner_nombre': canal_partner_nombre,
                    'n_productos': len(p.items),
                    'total_unidades': total_unidades,
                    'total_importe': total_importe,
                    'productos': sorted([
                        {
                            'codigo_barra': it.codigo_barra,
                            'nombre': it.nombre,
                            'cantidad': it.cantidad,
                            'precio_pvp': float(it.precio_pvp or 0),
                            'subtotal': float(it.subtotal or 0),
                            'pendientes': _pendientes_de(it.codigo_barra),
                        }
                        for it in p.items
                    ], key=lambda x: (x['nombre'] or '').lower()),
                })
            proveedores = [{'id': pv.id, 'nombre': pv.razon_social}
                           for pv in session.query(database.Provider).order_by(database.Provider.razon_social).all()]
            return render_template('orders_list.html', pedidos=result, proveedores=proveedores)

    @app.route('/api/pedido/<int:pedido_id>/indicadores')
    def api_pedido_indicadores(pedido_id):
        """Devuelve indicadores agregados del pedido (mini dashboard).

        Para cada item del pedido cruza vía bridge productos.observer_id ↔ ObsProducto
        y trae stock, ventas 3m/12m, monodroga, laboratorio. Calcula:
        - cobertura post-compra (días)
        - momentum (anualizado vs 12m)
        - riesgos (sin movimiento, sobre-pedido, stock dormido)
        - top 10 por unidades 12m
        - mix por monodroga y laboratorio
        - estacionalidad mensual del pedido global
        """
        import os as _os
        from datetime import datetime

        from sqlalchemy import func as _f

        id_farmacia = int(_os.environ.get('OBSERVER_ID_FARMACIA', '10525'))
        hoy = datetime.now()

        # Ventana 12m
        meses = []
        y, m = hoy.year, hoy.month
        for _ in range(12):
            meses.append((y, m))
            m -= 1
            if m == 0:
                m, y = 12, y - 1
        meses.reverse()
        desde_12m = meses[0][0] * 100 + meses[0][1]
        hasta = meses[-1][0] * 100 + meses[-1][1]

        def _ym_hace(n):
            yy, mm = hoy.year, hoy.month - n
            while mm <= 0:
                mm += 12
                yy -= 1
            return yy * 100 + mm
        desde_3m = _ym_hace(2)

        with database.get_db() as session:
            pedido = session.get(Pedido, pedido_id)
            if not pedido:
                return jsonify({'error': 'Pedido no encontrado'}), 404
            items = session.query(database.PedidoItem).filter_by(pedido_id=pedido_id).all()

            # Filtro opcional multi-token AND. Tokens separados por espacios o '+'.
            # Cada token debe matchear nombre o codigo_barra. Ejemplo:
            #   q="400 susp" → producto debe contener "400" Y "susp".
            #   q="actron"   → producto debe contener "actron".
            q_filter = (request.args.get('q') or '').strip().lower()
            items_total_pedido = len(items)
            if q_filter:
                tokens = [t for t in q_filter.replace('+', ' ').split() if t]
                if tokens:
                    def _matches(it):
                        nombre_l = (it.nombre or '').lower()
                        cb_l = (it.codigo_barra or '').lower()
                        return all(t in nombre_l or t in cb_l for t in tokens)
                    items = [i for i in items if _matches(i)]

            # Bridge codigo_barra → observer_id, en 2 pasos:
            # 1) ATAJO: pedidos generados desde ObServer guardan el IdProducto como string
            #    en codigo_barra. Si el código es numérico y existe en obs_productos, lo
            #    usamos directo (sin pasar por la tabla productos local).
            # 2) FALLBACK: tabla productos local (codigo_barra principal o alts) → observer_id.
            codigos = [(i.codigo_barra or '').strip() for i in items if i.codigo_barra]
            cb_to_obs = {}

            # Paso 1: códigos numéricos que matchean directo a obs_productos.observer_id
            cb_numericos = []
            for cb in codigos:
                try:
                    cb_numericos.append((cb, int(cb)))
                except (ValueError, TypeError):
                    continue
            if cb_numericos:
                ids_a_buscar = [n for (_, n) in cb_numericos]
                obs_existentes = {oid for (oid,) in session.query(database.ObsProducto.observer_id)
                                  .filter(database.ObsProducto.observer_id.in_(ids_a_buscar)).all()}
                for cb, n in cb_numericos:
                    if n in obs_existentes:
                        cb_to_obs[cb] = n

            # Paso 2: bridge vía tabla productos. Match al principal +
            # producto_codigos_barra (1-a-N) para alts. Las columnas legacy
            # alt1/2/3 ya no se consultan.
            codigos_pendientes = [c for c in codigos if c not in cb_to_obs]
            if codigos_pendientes:
                Producto = database.Producto
                # 2a. Match al principal
                for p in (session.query(Producto)
                          .filter(Producto.codigo_barra.in_(codigos_pendientes)).all()):
                    if p.observer_id and p.codigo_barra in codigos_pendientes:
                        cb_to_obs[p.codigo_barra] = p.observer_id
                # 2b. Match en producto_codigos_barra
                pendientes2 = [c for c in codigos_pendientes if c not in cb_to_obs]
                if pendientes2:
                    rows = (session.query(database.ProductoCodigoBarra.codigo_barra,
                                          database.ProductoCodigoBarra.producto_id)
                            .filter(database.ProductoCodigoBarra.codigo_barra.in_(pendientes2))
                            .all())
                    if rows:
                        ids = {pid for _, pid in rows}
                        obs_by_pid = {p.id: p.observer_id for p in
                                      session.query(Producto)
                                      .filter(Producto.id.in_(ids))
                                      .filter(Producto.observer_id.isnot(None))
                                      .all()}
                        for ean, pid in rows:
                            if pid in obs_by_pid and ean not in cb_to_obs:
                                cb_to_obs[ean] = obs_by_pid[pid]

            obs_ids = list(set(cb_to_obs.values()))
            obs_data = {}
            stock_map = {}
            ventas_3m = {}
            ventas_12m = {}
            serie_mensual = {}  # {(anio, mes): total_unidades} — agregado de TODOS los items.
            labs_map = {}
            drogas_map = {}

            if obs_ids:
                obs_prods = session.query(database.ObsProducto).filter(
                    database.ObsProducto.observer_id.in_(obs_ids)).all()
                obs_data = {p.observer_id: p for p in obs_prods}

                # Stock + mínimo en una sola query.
                rows_stock = session.query(
                    database.ObsStock.producto_observer,
                    database.ObsStock.stock_actual,
                    database.ObsStock.minimo,
                ).filter(database.ObsStock.id_farmacia == id_farmacia,
                         database.ObsStock.producto_observer.in_(obs_ids)).all()
                stock_map = {r[0]: int(r[1] or 0) for r in rows_stock}
                minimo_map = {r[0]: int(r[2] or 0) for r in rows_stock if r[2] is not None}

                ym_expr = database.ObsVentaMensual.anio * 100 + database.ObsVentaMensual.mes

                # Una sola query para 3m y 12m. El rango 3m está dentro del 12m,
                # así que filtramos por 12m y sumamos condicionalmente con CASE.
                # Antes eran 2 round-trips → 1.
                from sqlalchemy import case
                rows_ventas = session.query(
                    database.ObsVentaMensual.producto_observer,
                    _f.sum(case(
                        (ym_expr.between(desde_3m, hasta),
                         database.ObsVentaMensual.unidades),
                        else_=0)).label('u3m'),
                    _f.sum(database.ObsVentaMensual.unidades).label('u12m'),
                ).filter(
                    database.ObsVentaMensual.id_farmacia == id_farmacia,
                    database.ObsVentaMensual.producto_observer.in_(obs_ids),
                    ym_expr.between(desde_12m, hasta),
                ).group_by(database.ObsVentaMensual.producto_observer).all()
                ventas_3m = {pid: float(u3 or 0) for (pid, u3, _) in rows_ventas}
                ventas_12m = {pid: float(u12 or 0) for (pid, _, u12) in rows_ventas}

                # Estacionalidad agregada: solo necesitamos total por (anio, mes) sumando
                # TODOS los productos. Pre-2026-05 esto traía 1 row por (producto, mes)
                # y se sumaba en Python — con N items × 12m = N×12 rows. Ahora SUMA en
                # SQL devolviendo 12 rows fijas.
                for (anio, mes, u_total) in session.query(
                        database.ObsVentaMensual.anio,
                        database.ObsVentaMensual.mes,
                        _f.sum(database.ObsVentaMensual.unidades))\
                    .filter(database.ObsVentaMensual.id_farmacia == id_farmacia,
                            database.ObsVentaMensual.producto_observer.in_(obs_ids),
                            ym_expr.between(desde_12m, hasta))\
                    .group_by(database.ObsVentaMensual.anio,
                              database.ObsVentaMensual.mes).all():
                    serie_mensual[(anio, mes)] = float(u_total or 0)

                lab_ids = {p.laboratorio_observer for p in obs_prods if p.laboratorio_observer}
                droga_ids = {p.nombre_droga_observer for p in obs_prods if p.nombre_droga_observer}
                if lab_ids:
                    labs_map = dict(session.query(
                        database.ObsLaboratorio.observer_id,
                        database.ObsLaboratorio.descripcion)
                        .filter(database.ObsLaboratorio.observer_id.in_(lab_ids)).all())
                if droga_ids:
                    drogas_map = dict(session.query(
                        database.ObsNombreDroga.observer_id,
                        database.ObsNombreDroga.descripcion)
                        .filter(database.ObsNombreDroga.observer_id.in_(droga_ids)).all())

            # Construir items enriquecidos
            items_enriched = []
            for it in items:
                cb = (it.codigo_barra or '').strip()
                obs_id = cb_to_obs.get(cb)
                op = obs_data.get(obs_id) if obs_id else None
                stock = int(stock_map.get(obs_id, 0) or 0) if obs_id else 0
                minimo = int(minimo_map.get(obs_id, 0) or 0) if obs_id else 0
                u3m = ventas_3m.get(obs_id, 0) if obs_id else 0
                u12m = ventas_12m.get(obs_id, 0) if obs_id else 0
                cantidad = int(it.cantidad or 0)
                # Cobertura post-compra (días)
                if u3m > 0:
                    dias_post = (stock + cantidad) / (u3m / 90)
                    dias_pre = stock / (u3m / 90)
                else:
                    dias_post = None
                    dias_pre = None
                # Momentum
                momentum = None
                if u12m > 0:
                    momentum = (u3m * 4 - u12m) / u12m * 100
                items_enriched.append({
                    'pedido_item_id': it.id,
                    'codigo_barra':   cb,
                    'nombre':         it.nombre or '',
                    'cantidad_pedida': cantidad,
                    'observer_id':    obs_id,
                    'tiene_obs':      obs_id is not None,
                    'baja':           bool(op.fecha_baja) if op else False,
                    'stock':          stock,
                    'minimo':         minimo,
                    'u3m':            u3m,
                    'u12m':           u12m,
                    'avg_mensual':    round(u3m / 3, 2),
                    'dias_pre_compra': dias_pre,
                    'dias_post_compra': dias_post,
                    'momentum_pct':   momentum,
                    'monodroga':      drogas_map.get(op.nombre_droga_observer) if op else '',
                    'monodroga_id':   op.nombre_droga_observer if op else None,
                    'laboratorio':    labs_map.get(op.laboratorio_observer) if op else '',
                })

            # Riesgos
            riesgos = []
            for it in items_enriched:
                tags = []
                if it['cantidad_pedida'] == 0:
                    continue
                if it['u12m'] == 0 and it['tiene_obs']:
                    tags.append('Sin ventas en 12m')
                elif it['u3m'] == 0 and it['u12m'] > 0:
                    tags.append('Sin ventas en 3m')
                if it['avg_mensual'] > 0 and it['cantidad_pedida'] > it['avg_mensual'] * 3:
                    tags.append(f'Sobre-pedido ({it["cantidad_pedida"]} vs {it["avg_mensual"]:.0f}/mes)')
                if it['dias_pre_compra'] is not None and it['dias_pre_compra'] > 180:
                    tags.append(f'Stock previo {round(it["dias_pre_compra"])}d (>180)')
                if it['dias_post_compra'] is not None and it['dias_post_compra'] < 15:
                    tags.append(f'Aún corto post-compra ({round(it["dias_post_compra"])}d)')
                if not it['tiene_obs']:
                    tags.append('Sin link a ObServer')
                if tags:
                    riesgos.append({'item': it, 'tags': tags})

            # Top 10 por uni 12m (filtrado a items con cantidad pedida)
            top10 = sorted([i for i in items_enriched if i['cantidad_pedida'] > 0],
                           key=lambda x: -x['u12m'])[:10]

            # Mix por monodroga (suma cantidad pedida)
            mix_droga = {}
            for it in items_enriched:
                if it['cantidad_pedida'] == 0:
                    continue
                key = it['monodroga'] or '(sin monodroga)'
                mix_droga[key] = mix_droga.get(key, 0) + it['cantidad_pedida']
            mix_droga_list = sorted([{'label': k, 'count': v} for k, v in mix_droga.items()],
                                    key=lambda x: -x['count'])[:10]

            # Mix por laboratorio
            mix_lab = {}
            for it in items_enriched:
                if it['cantidad_pedida'] == 0:
                    continue
                key = it['laboratorio'] or '(sin lab)'
                mix_lab[key] = mix_lab.get(key, 0) + it['cantidad_pedida']
            mix_lab_list = sorted([{'label': k, 'count': v} for k, v in mix_lab.items()],
                                  key=lambda x: -x['count'])

            # Estacionalidad: total mensual ya viene agregado desde SQL en serie_mensual
            # como {(anio, mes): total_unidades}.
            estacionalidad = {
                'labels': [f'{m:02d}/{y}' for (y, m) in meses],
                'unidades': [float(serie_mensual.get((y, m), 0) or 0) for (y, m) in meses],
            }

            # Resumen general
            n_items = len(items_enriched)
            n_con_obs = sum(1 for i in items_enriched if i['tiene_obs'])
            unidades_pedidas = sum(i['cantidad_pedida'] for i in items_enriched)

        return jsonify({
            'pedido': {
                'id': pedido.id,
                'laboratorio': pedido.laboratorio,
                'periodo': pedido.periodo or '',
                'n_items': n_items,
                'n_con_obs': n_con_obs,
                'unidades_pedidas': unidades_pedidas,
            },
            'items': items_enriched,
            'riesgos': riesgos,
            'top10': top10,
            'mix_monodroga': mix_droga_list,
            'mix_laboratorio': mix_lab_list,
            'estacionalidad': estacionalidad,
            'filtro': {
                'q': q_filter,
                'items_filtrados': n_items,
                'items_total': items_total_pedido,
            } if q_filter else None,
        })

    @app.route('/api/pedido/<int:pedido_id>/vincular-observer', methods=['POST'])
    def api_pedido_vincular_observer(pedido_id):
        """Linkea items del pedido contra obs_productos por descripción + laboratorio.
        Reusa la lógica del script scripts/vincular_pedido_observer.py."""
        import os as _os
        import sys as _sys

        import cron_log as _cron_log
        scripts_dir = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), 'scripts')
        if scripts_dir not in _sys.path:
            _sys.path.insert(0, scripts_dir)
        from vincular_pedido_observer import procesar_pedido as _procesar
        with _cron_log.registrar(f'vincular_observer:pedido_{pedido_id}', origen='web') as log:
            with database.get_db() as session:
                pedido = session.get(Pedido, pedido_id)
                if not pedido:
                    return jsonify({'error': 'Pedido no encontrado'}), 404
                stats = _procesar(session, pedido, dry_run=False)
            log.set_mensaje(
                f'{pedido.laboratorio}: linkeados={stats.get("linkeados", 0)} '
                f'ya={stats.get("ya_linkeado", 0)} '
                f'amb={stats.get("ambiguos", 0)} '
                f'sin={stats.get("no_encontrados", 0)}'
            )
            return jsonify({'ok': True, **stats, 'pedido': pedido.laboratorio})

    @app.route('/order/<int:pedido_id>/mostrar-hasta', methods=['POST'])
    def order_mostrar_hasta(pedido_id):
        """Marca un pedido guardado para que aparezca como sugerencia en
        'Pedido Reposición' (compras_dia) hasta la fecha indicada.
        Body JSON: {fecha: 'YYYY-MM-DD' | null | ''}.
        Pasar null/''  → desactiva la marca.
        """
        from datetime import date as _date
        data = request.get_json(silent=True) or {}
        raw = (data.get('fecha') or '').strip()
        with database.get_db() as session:
            pedido = session.get(Pedido, pedido_id)
            if not pedido:
                return jsonify({'ok': False, 'error': 'Pedido no encontrado'}), 404
            if not raw:
                pedido.mostrar_hasta = None
            else:
                try:
                    pedido.mostrar_hasta = _date.fromisoformat(raw)
                except ValueError:
                    return jsonify({'ok': False, 'error': 'Fecha inválida (esperado YYYY-MM-DD)'}), 400
            session.commit()
            return jsonify({
                'ok': True,
                'mostrar_hasta': pedido.mostrar_hasta.isoformat() if pedido.mostrar_hasta else '',
                'mostrar_hasta_label': pedido.mostrar_hasta.strftime('%d/%m/%y') if pedido.mostrar_hasta else '',
            })

    @app.route('/order/<int:pedido_id>/delete', methods=['POST'])
    def order_delete(pedido_id):
        with database.get_db() as session:
            try:
                pedido = session.query(Pedido).get(pedido_id)
                if pedido:
                    session.delete(pedido)
                    session.commit()
                    flash('Pedido eliminado.')
            except Exception as e:
                session.rollback()
                flash(f'Error: {e}')
        return redirect(url_for('orders_list'))

    @app.route('/order/<int:pedido_id>/export/<fmt>')
    def order_export_file(pedido_id, fmt):
        """Exporta el pedido guardado a xlsx o pdf."""
        with database.get_db() as session:
            pedido = session.query(Pedido).get(pedido_id)
            if not pedido:
                return 'Pedido no encontrado', 404
            items = [{
                'codigo_barra': it.codigo_barra or '',
                'nombre': it.nombre or '',
                'cantidad': it.cantidad or 0,
                'precio_pvp': float(it.precio_pvp or 0),
                'subtotal': float(it.subtotal or 0),
            } for it in pedido.items]
            total_unidades = sum(it['cantidad'] for it in items)
            total_importe = sum(it['subtotal'] for it in items)
            lab = pedido.laboratorio or 'Pedido'
            periodo = pedido.periodo or ''
            n_days = pedido.n_days or 0
            safe_lab = secure_filename(lab) or 'pedido'

        if fmt == 'xlsx':
            from io import BytesIO as _BIO

            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Pedido'
            ws.append([f'{lab} — {periodo}'])
            ws['A1'].font = Font(bold=True, size=13)
            ws.append([f'{n_days} días'])
            ws.append([])

            headers = ['Cód. Barras', 'Producto', 'P.PVP', 'Cantidad', 'Subtotal']
            ws.append(headers)
            hdr_row = ws.max_row
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=hdr_row, column=c)
                cell.fill = PatternFill('solid', fgColor='1C1C1E')
                cell.font = Font(bold=True, color='EAB308')
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 42
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 14

            for it in items:
                ws.append([it['codigo_barra'], it['nombre'],
                           it['precio_pvp'], it['cantidad'], it['subtotal']])

            ws.append([])
            ws.append(['', 'Total', '', total_unidades, total_importe])
            tot_row = ws.max_row
            for c in (2, 4, 5):
                ws.cell(row=tot_row, column=c).font = Font(bold=True)

            buf = _BIO()
            wb.save(buf); buf.seek(0)
            resp = make_response(buf.read())
            resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            resp.headers['Content-Disposition'] = f'attachment; filename="Pedido_{safe_lab}.xlsx"'
            return resp

        if fmt == 'pdf':
            from datetime import datetime as _dt
            from io import BytesIO as _BIO

            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.pdfgen import canvas as rl_canvas
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            fecha_emision = _dt.now().strftime('%d/%m/%Y %H:%M')
            page_w, page_h = A4

            class _NumberedCanvas(rl_canvas.Canvas):
                def __init__(self, *args, **kwargs):
                    rl_canvas.Canvas.__init__(self, *args, **kwargs)
                    self._pages = []
                def showPage(self):
                    self._pages.append(dict(self.__dict__))
                    self._startPage()
                def save(self):
                    total = len(self._pages)
                    for i, state in enumerate(self._pages, 1):
                        self.__dict__.update(state)
                        self.saveState()
                        self.setFont('Helvetica', 7)
                        self.setFillColor(colors.HexColor('#6B7280'))
                        self.drawString(1.5*cm, 0.6*cm, f"Emitido: {fecha_emision}")
                        self.drawRightString(page_w - 1.5*cm, 0.6*cm, f"Página {i} de {total}")
                        self.restoreState()
                        rl_canvas.Canvas.showPage(self)
                    rl_canvas.Canvas.save(self)

            buf = _BIO()
            doc = SimpleDocTemplate(buf, pagesize=A4,
                                    leftMargin=1.5*cm, rightMargin=1.5*cm,
                                    topMargin=1.5*cm, bottomMargin=1.5*cm)
            styles = getSampleStyleSheet()
            accent = colors.HexColor('#EAB308')
            txt = colors.HexColor('#1A202C')
            hdr_bg = colors.HexColor('#2D3748')
            row_b = colors.HexColor('#F3F4F6')
            title_s = ParagraphStyle('t', parent=styles['Normal'], fontSize=14, textColor=accent, spaceAfter=4)
            sub_s = ParagraphStyle('s', parent=styles['Normal'], fontSize=9, textColor=txt, spaceAfter=2)
            cell_s = ParagraphStyle('c', parent=styles['Normal'], fontSize=7, textColor=txt, leading=9)

            story = [
                Paragraph(f"Pedido — {lab}", title_s),
                Paragraph(f"{periodo} · {n_days} días", sub_s),
                Paragraph(f"{len(items)} productos · {total_unidades} unidades · ${total_importe:,.0f}".replace(',', '.'), sub_s),
                Spacer(1, 0.4*cm),
            ]
            headers = ['Cód. Barras', 'Producto', 'P.PVP', 'Cantidad', 'Subtotal']
            rows = [headers]
            for it in items:
                rows.append([
                    it['codigo_barra'],
                    Paragraph(it['nombre'], cell_s),
                    f"${it['precio_pvp']:,.0f}".replace(',', '.'),
                    it['cantidad'],
                    f"${it['subtotal']:,.0f}".replace(',', '.'),
                ])
            rows.append(['', 'TOTAL', '', total_unidades,
                         f"${total_importe:,.0f}".replace(',', '.')])

            t = Table(rows, colWidths=[3.2*cm, 9.5*cm, 2.2*cm, 2.2*cm, 2.4*cm], repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), hdr_bg),
                ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
                ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0, 0), (-1, 0), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, row_b]),
                ('FONTSIZE',   (0, 1), (-1, -1), 7),
                ('ALIGN',      (2, 0), (-1, -1), 'RIGHT'),
                ('ALIGN',      (0, 0), (1, -1), 'LEFT'),
                ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
                ('GRID',       (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
                ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(t)
            doc.build(story, canvasmaker=_NumberedCanvas)
            buf.seek(0)
            resp = make_response(buf.read())
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Content-Disposition'] = f'attachment; filename="Pedido_{safe_lab}.pdf"'
            return resp

        return 'Formato inválido', 400

    # ─── ORDER DETAIL & MODULES ──────────────────────────────────────────────

    @app.route('/order/<int:pedido_id>')
    def order_detail(pedido_id):
        with database.get_db() as session:
            pedido = session.query(Pedido).get(pedido_id)
            if not pedido:
                flash('Pedido no encontrado.')
                return redirect(url_for('orders_list'))
            from datetime import datetime as _dt2
            dias_desde = None
            if pedido.creado_en:
                dias_desde = (now_ar() - pedido.creado_en).days
            data = {
                'id': pedido.id,
                'laboratorio': pedido.laboratorio,
                'farmacia': pedido.farmacia,
                'periodo': pedido.periodo,
                'n_days': pedido.n_days,
                'creado_en': pedido.creado_en.strftime('%d/%m/%Y %H:%M') if pedido.creado_en else '',
                'dias_desde_analisis': dias_desde,
                'analisis_json': pedido.analisis_json or '',
                'analisis_guardado_en': pedido.analisis_guardado_en.strftime('%d/%m/%Y %H:%M') if pedido.analisis_guardado_en else '',
                'canal': pedido.canal or '',
                'partner_id': pedido.partner_id,
                'canal_elegido_en': pedido.canal_elegido_en.strftime('%d/%m/%Y %H:%M') if pedido.canal_elegido_en else '',
            }
            erp_stock_map = {
                row.codigo_barra: int(row.cantidad or 0)
                for row in session.query(ErpStock).all()
            }
            all_prods = session.query(Producto).all()
            # Pre-cargar todos los EANs alternativos desde producto_codigos_barra
            # en un solo query — los alts ya no viven en alt1/2/3.
            alts_por_pid = {}
            for pid, ean in (session.query(database.ProductoCodigoBarra.producto_id,
                                            database.ProductoCodigoBarra.codigo_barra)
                              .filter(database.ProductoCodigoBarra.es_principal.is_(False))
                              .all()):
                alts_por_pid.setdefault(pid, []).append(ean)
            def _all_eans(p):
                return [p.codigo_barra] + alts_por_pid.get(p.id, [])
            monodroga_by_bc = {}
            for p in all_prods:
                if not p.monodroga:
                    continue
                for bc in _all_eans(p):
                    if bc:
                        monodroga_by_bc[bc] = p.monodroga

            # Map codigo_barra → tvc (tipo venta y control: 'L'=Libre, 'R'=Receta...).
            # Resolución en 2 pasos: 1) si cb es OBS:N o numérico, lookup directo en
            # obs_productos por observer_id; 2) sino, vía Producto.observer_id local.
            tvc_by_cb = {}
            cbs_pedido = [(it.codigo_barra or '').strip() for it in pedido.items if it.codigo_barra]
            obs_ids_directos = {}
            for cb in cbs_pedido:
                obs_id = None
                if cb.startswith('OBS:'):
                    try: obs_id = int(cb[4:])
                    except (ValueError, TypeError): pass
                elif cb.isdigit() and len(cb) <= 7:
                    try: obs_id = int(cb)
                    except (ValueError, TypeError): pass
                if obs_id is not None:
                    obs_ids_directos[cb] = obs_id
            # Vía Producto local (principal + alts en 1-a-N)
            for p in all_prods:
                if not p.observer_id:
                    continue
                for bc in _all_eans(p):
                    if bc and bc not in obs_ids_directos:
                        obs_ids_directos[bc] = p.observer_id
            todos_obs_ids = list({oid for oid in obs_ids_directos.values()})
            droga_id_by_cb = {}
            if todos_obs_ids:
                obs_rows = session.query(database.ObsProducto.observer_id,
                                          database.ObsProducto.id_tipo_venta_control,
                                          database.ObsProducto.nombre_droga_observer)\
                                  .filter(database.ObsProducto.observer_id.in_(todos_obs_ids)).all()
                tvc_map = {oid: tvc for (oid, tvc, _) in obs_rows}
                droga_map = {oid: drg for (oid, _, drg) in obs_rows}
                for cb, oid in obs_ids_directos.items():
                    if tvc_map.get(oid):
                        tvc_by_cb[cb] = tvc_map[oid]
                    if droga_map.get(oid):
                        droga_id_by_cb[cb] = droga_map[oid]

            data['productos'] = [
                {
                    'codigo_barra': it.codigo_barra,
                    'nombre': it.nombre,
                    'cantidad': it.cantidad,
                    'precio_pvp': float(it.precio_pvp or 0),
                    'subtotal': float(it.subtotal or 0),
                    'rotacion': it.rotacion or '',
                    'avg_monthly': float(it.avg_monthly) if it.avg_monthly else None,
                    'erp_qty': erp_stock_map.get(it.codigo_barra),
                    'monodroga': monodroga_by_bc.get(it.codigo_barra, ''),
                    'tvc': tvc_by_cb.get(it.codigo_barra, ''),
                    'monodroga_id': droga_id_by_cb.get(it.codigo_barra),
                }
                for it in pedido.items
            ]
            equiv = [
                {'barcodes': [b for b in _all_eans(p) if b]}
                for p in all_prods
            ]
            product_prices = {}
            for p in all_prods:
                if p.precio_pvp is not None:
                    price = float(p.precio_pvp)
                    for bc in _all_eans(p):
                        if bc:
                            product_prices[bc] = price

            cfg = session.query(database.Config).get(1)
            tol_config = {
                'A': float(cfg.rot_alta_tol)  if cfg else 0.0,
                'M': float(cfg.rot_media_tol) if cfg else 0.0,
                'B': float(cfg.rot_baja_tol)  if cfg else 0.0,
            }
            packs = [{'id': mp.id, 'ean_pack': mp.ean_pack, 'ean_unidad': mp.ean_unidad,
                      'cantidad': mp.cantidad, 'descripcion': mp.descripcion or ''}
                     for mp in session.query(ModuloPack).order_by(ModuloPack.ean_pack).all()]
            from datetime import datetime as _dt
            if not pedido.analizado_en:
                pedido.analizado_en = now_ar()
                session.commit()
                data['analizado_en'] = pedido.analizado_en.strftime('%d/%m/%Y')
            else:
                data['analizado_en'] = pedido.analizado_en.strftime('%d/%m/%Y')
            lab_obj = session.query(database.Laboratorio).filter_by(nombre=pedido.laboratorio).first()
            data['lab_id'] = lab_obj.id if lab_obj else None
            prov_plantilla = None
            _prov = session.query(database.Provider).filter(
                database.Provider.razon_social.ilike(f'%{pedido.laboratorio or ""}%')
            ).first()
            if _prov:
                _pl = session.query(database.PlantillaExportacion).filter_by(proveedor_id=_prov.id).first()
                if _pl:
                    prov_plantilla = {'proveedor_id': _prov.id, 'nombre': _pl.nombre,
                                      'extension': _pl.extension}
            lab_plantilla = None
            if lab_obj:
                _lt = session.get(database.ExportTemplate, lab_obj.id)
                if _lt and _lt.columns_json:
                    lab_plantilla = {'laboratorio_id': lab_obj.id}

            # Unified Plantillas: se filtra según el canal decidido del pedido.
            # - Si canal='drogueria' + partner_id: solo plantillas de esa droguería.
            # - Si canal='laboratorio': solo plantillas del lab fabricante.
            # - Sin canal decidido: todas las que podrían aplicar (lab + cualquier prov asociado),
            #   como fallback para que el user vea opciones antes de decidir.
            plantillas_entidad = []
            _filters = []
            if pedido.canal == 'drogueria' and pedido.partner_id:
                _prov_canal = session.get(database.Provider, pedido.partner_id)
                if _prov_canal:
                    _filters.append((_prov_canal.tipo or 'drogueria', _prov_canal.id))
            elif pedido.canal == 'laboratorio' and lab_obj:
                _filters.append(('laboratorio', lab_obj.id))
            else:
                # Sin canal decidido: mostrar todas las opciones.
                # Incluye lab + droguerías configuradas en la matriz LaboratorioDrogueria.
                if lab_obj:
                    _filters.append(('laboratorio', lab_obj.id))
                    drogs_lab = (session.query(database.Provider)
                                 .join(database.LaboratorioDrogueria,
                                       database.LaboratorioDrogueria.drogueria_id == database.Provider.id)
                                 .filter(database.LaboratorioDrogueria.laboratorio_id == lab_obj.id)
                                 .all())
                    for _d in drogs_lab:
                        _filters.append((_d.tipo or 'drogueria', _d.id))
                elif _prov:
                    _filters.append((_prov.tipo or 'proveedor', _prov.id))
            for _tipo, _eid in _filters:
                rows = (session.query(database.Plantilla)
                        .filter_by(entidad_tipo=_tipo, entidad_id=_eid)
                        .order_by(database.Plantilla.es_default.desc(),
                                  database.Plantilla.nombre).all())
                for p in rows:
                    plantillas_entidad.append({
                        'id': p.id, 'nombre': p.nombre, 'formato': p.formato,
                        'tipo_doc': p.tipo_doc, 'es_default': bool(p.es_default),
                        'entidad_tipo': _tipo, 'entidad_id': _eid,
                    })

            # Droguerías disponibles para elegir como canal
            droguerias = [{'id': p.id, 'razon_social': p.razon_social}
                          for p in (session.query(database.Provider)
                                    .filter(database.Provider.tipo == 'drogueria')
                                    .order_by(database.Provider.razon_social).all())]

            return render_template('order_detail.html', pedido=data, productos_equiv=equiv,
                                   tol_config=tol_config, modulo_packs=packs,
                                   product_prices=product_prices,
                                   prov_plantilla=prov_plantilla,
                                   lab_plantilla=lab_plantilla,
                                   plantillas_entidad=plantillas_entidad,
                                   droguerias=droguerias)

    @app.route('/order/<int:pedido_id>/save-state', methods=['POST'])
    def order_save_state(pedido_id):
        """Persiste un snapshot JSON del análisis (módulos, ofertas, cantidades, resumen)."""
        body = request.get_json(silent=True) or {}
        with database.get_db() as session:
            try:
                pedido = session.get(Pedido, pedido_id)
                if not pedido:
                    return jsonify({'ok': False, 'error': 'Pedido no encontrado'}), 404
                import json as _json
                pedido.analisis_json = _json.dumps(body, ensure_ascii=False)
                pedido.analisis_guardado_en = now_ar()
                session.commit()
                return jsonify({'ok': True,
                                'guardado_en': pedido.analisis_guardado_en.strftime('%d/%m/%Y %H:%M')})
            except Exception as e:
                session.rollback()
                return jsonify({'ok': False, 'error': str(e)}), 500

    @app.route('/order/<int:pedido_id>/clear-state', methods=['POST'])
    def order_clear_state(pedido_id):
        """Borra el snapshot del análisis para arrancar de cero."""
        with database.get_db() as session:
            try:
                pedido = session.get(Pedido, pedido_id)
                if not pedido:
                    return jsonify({'ok': False, 'error': 'Pedido no encontrado'}), 404
                pedido.analisis_json = None
                pedido.analisis_guardado_en = None
                session.commit()
                return jsonify({'ok': True})
            except Exception as e:
                session.rollback()
                return jsonify({'ok': False, 'error': str(e)}), 500

    @app.route('/order/<int:pedido_id>/save-module-matches', methods=['POST'])
    def order_save_module_matches(pedido_id):
        """Guarda equivalencias EAN-módulo → barcode-pedido en tabla productos."""
        body    = request.get_json(silent=True) or {}
        matches = body.get('matches', []) if isinstance(body, dict) else body
        with database.get_db() as session:
            try:
                pedido = session.get(Pedido, pedido_id)
                lab_id = None
                if pedido and pedido.laboratorio:
                    from helpers import get_or_create_laboratorio
                    lab = get_or_create_laboratorio(session, pedido.laboratorio.strip())
                    lab_id = lab.id if lab else None

                saved = 0
                for m in matches:
                    module_ean  = str(m.get('module_ean', '')).strip()
                    pedido_bc   = str(m.get('pedido_barcode', '')).strip()
                    pedido_nom  = m.get('pedido_nombre', '')
                    if not module_ean or not pedido_bc or module_ean == pedido_bc:
                        continue
                    _upsert_producto(session, pedido_bc, pedido_nom, laboratorio_id=lab_id)
                    _add_alt_barcode(session, pedido_bc, module_ean)
                    saved += 1
                session.commit()
                # equiv usa el helper _all_eans para combinar principal + 1-a-N
                alts_por_pid = {}
                for pid, ean in (session.query(database.ProductoCodigoBarra.producto_id,
                                                database.ProductoCodigoBarra.codigo_barra)
                                  .filter(database.ProductoCodigoBarra.es_principal.is_(False))
                                  .all()):
                    alts_por_pid.setdefault(pid, []).append(ean)
                equiv = [
                    {'barcodes': [b for b in [p.codigo_barra] + alts_por_pid.get(p.id, []) if b]}
                    for p in session.query(Producto).all()
                ]
                return jsonify({'ok': True, 'saved': saved, 'equiv': equiv})
            except Exception as e:
                session.rollback()
                return jsonify({'ok': False, 'error': str(e)}), 500

    @app.route('/order/<int:pedido_id>/modules-template', methods=['GET'])
    def order_modules_template(pedido_id):
        """Descarga una plantilla XLSX lista para completar con módulos."""
        import io

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        with database.get_db() as session:
            pedido = session.query(database.Pedido).get(pedido_id)
            lab = pedido.laboratorio if pedido else 'Laboratorio'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Módulos'

        hdr_fill = PatternFill('solid', fgColor='1C1C1E')
        hdr_font = Font(bold=True, color='EAB308')
        mod_fill = PatternFill('solid', fgColor='FEF9C3')
        mod_font = Font(bold=True, color='92400E')
        thin     = Side(style='thin', color='D0D0D0')
        border   = Border(bottom=Side(style='thin', color='D0D0D0'))
        gray     = Font(color='999999', italic=True)

        ws.append([f'MÓDULOS {lab.upper()}'])
        ws['A1'].font = Font(bold=True, size=13)
        ws.append([])

        headers = ['NOMBRE MÓDULO', 'CÓDIGO EAN', 'DESCRIPCIÓN', 'CANT.', 'DESC. %']
        ws.append(headers)
        for ci, _ in enumerate(headers, 1):
            c = ws.cell(row=3, column=ci)
            c.fill = hdr_fill
            c.font = hdr_font
            c.border = border

        ws.append(['MOD. EJEMPLO A'])
        c = ws.cell(row=4, column=1)
        c.fill = mod_fill; c.font = mod_font

        for ean, desc, cant, pct in [
            ('7793450000001', 'PRODUCTO EJEMPLO 1', 2, 7),
            ('7793450000002', 'PRODUCTO EJEMPLO 2', 1, 7),
        ]:
            ws.append(['MOD. EJEMPLO A', ean, desc, cant, pct])

        ws.append([])

        ws.append(['MOD. EJEMPLO B'])
        c = ws.cell(row=ws.max_row, column=1)
        c.fill = mod_fill; c.font = mod_font
        ws.append(['MOD. EJEMPLO B', '7793450000003', 'PRODUCTO EJEMPLO 3', 3, 10])

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 10

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f'plantilla_modulos_{lab.lower().replace(" ", "_")}.xlsx'
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    @app.route('/order/<int:pedido_id>/parse-modules', methods=['POST'])
    def order_parse_modules(pedido_id):
        from parsers.modulos_xlsx import parse_modulos_xlsx
        f = request.files.get('modules_file')
        if not f or not f.filename:
            return jsonify({'error': 'No se recibió archivo'}), 400
        tmp = os.path.join(UPLOAD_FOLDER, f'mod_{pedido_id}_{secure_filename(f.filename)}')
        f.save(tmp)
        try:
            modules = parse_modulos_xlsx(tmp)
            destacados_count = sum(
                1 for m in modules for it in m.get('items', []) if it.get('destacado')
            )
            with database.get_db() as session:
                pack_candidates = _detectar_packs_en_modulos(modules, session)
            return jsonify({
                'modules': modules,
                'pack_candidates': pack_candidates,
                'destacados_count': destacados_count,
                'total_items': sum(len(m.get('items', [])) for m in modules),
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500
        finally:
            try: os.remove(tmp)
            except OSError: pass

    @app.route('/order/<int:pedido_id>/save-packs', methods=['POST'])
    def order_save_packs(pedido_id):
        """Guarda en modulo_packs los packs confirmados por el usuario.
        Body: {packs: [{ean_pack, ean_unidad, cantidad, descripcion}]}"""
        body = request.get_json(silent=True) or {}
        packs = body.get('packs', [])
        guardados = actualizados = 0
        with database.get_db() as session:
            for p in packs:
                ean_pack = (p.get('ean_pack') or '').strip()
                ean_unidad = (p.get('ean_unidad') or '').strip()
                try:
                    cantidad = int(p.get('cantidad') or 0)
                except (ValueError, TypeError):
                    cantidad = 0
                if not ean_pack or not ean_unidad or cantidad <= 0:
                    continue
                existing = session.query(ModuloPack).filter_by(ean_pack=ean_pack).first()
                if existing:
                    existing.ean_unidad = ean_unidad
                    existing.cantidad = cantidad
                    existing.descripcion = (p.get('descripcion') or '')[:255] or existing.descripcion
                    actualizados += 1
                else:
                    session.add(ModuloPack(
                        ean_pack=ean_pack, ean_unidad=ean_unidad,
                        cantidad=cantidad,
                        descripcion=(p.get('descripcion') or '')[:255],
                    ))
                    guardados += 1
            session.commit()
        return jsonify({'guardados': guardados, 'actualizados': actualizados})

    @app.route('/order/<int:pedido_id>/parse-offers', methods=['POST'])
    def order_parse_offers(pedido_id):
        from parsers.ofertas_xlsx import parse_ofertas_xlsx
        f = request.files.get('offers_file')
        if not f or not f.filename:
            return jsonify({'error': 'No se recibió archivo'}), 400
        tmp = os.path.join(UPLOAD_FOLDER, f'off_{pedido_id}_{secure_filename(f.filename)}')
        f.save(tmp)
        try:
            ofertas = parse_ofertas_xlsx(tmp)
            return jsonify({'ofertas': ofertas})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
        finally:
            try: os.remove(tmp)
            except OSError: pass

    @app.route('/order/<int:pedido_id>/export/plantilla', methods=['POST'])
    def order_export_plantilla(pedido_id):
        """Exporta el resumen usando la plantilla configurada para el laboratorio."""
        import json as _json
        from io import BytesIO

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        from database import ExportTemplate, Laboratorio, Pedido

        raw = request.form.get('data', '[]')
        try:
            rows = _json.loads(raw)
        except Exception:
            return 'Datos inválidos', 400

        with database.get_db() as session:
            pedido = session.get(Pedido, pedido_id)
            if not pedido:
                return 'Pedido no encontrado', 404
            lab = session.query(Laboratorio).filter_by(nombre=pedido.laboratorio).first()
            tpl = session.get(ExportTemplate, lab.id) if lab else None
            if not tpl or not tpl.columns_json:
                return 'El laboratorio no tiene plantilla configurada', 400
            cols    = [c for c in _json.loads(tpl.columns_json) if c.get('enabled')]
            hdr_txt = tpl.custom_header

        if not cols:
            return 'La plantilla no tiene columnas activas', 400

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pedido'

        row_offset = 1
        if hdr_txt:
            ws.cell(row=1, column=1, value=hdr_txt).font = Font(bold=True, size=12)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
            row_offset = 2

        hdr_fill = PatternFill('solid', fgColor='1e1e1e')
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=row_offset, column=ci, value=col['label'])
            cell.font      = Font(bold=True, color='FFFFFF', size=10)
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        for ri, row in enumerate(rows, row_offset + 1):
            for ci, col in enumerate(cols, 1):
                val = row.get(col['field'])
                if val is None or val == '':
                    val = None
                ws.cell(row=ri, column=ci, value=val)

        for ci, col in enumerate(cols, 1):
            field = col['field']
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = (
                20 if field in ('nombre',) else 15 if field == 'ean' else 12
            )

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"Pedido_{pedido.laboratorio}_{pedido.periodo or ''}.xlsx".replace(' ', '_')
        from flask import send_file
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/order/<int:pedido_id>/export-prov-plantilla', methods=['POST'])
    def order_export_prov_plantilla(pedido_id):
        """Genera archivo de ancho fijo con el resumen según plantilla del proveedor."""
        from flask import Response
        data = request.get_json(silent=True) or {}
        rows = data.get('rows', [])
        proveedor_id = data.get('proveedor_id')
        if not proveedor_id:
            return jsonify({'error': 'proveedor_id requerido'}), 400

        with database.get_db() as session:
            plantilla = session.query(database.PlantillaExportacion).filter_by(
                proveedor_id=int(proveedor_id)).first()
            if not plantilla:
                return jsonify({'error': 'Este proveedor no tiene plantilla configurada.'}), 404
            campos = sorted(plantilla.campos, key=lambda c: c.col_inicio)
            if not campos:
                return jsonify({'error': 'La plantilla no tiene campos definidos.'}), 400
            line_len = max(c.col_inicio + c.longitud for c in campos)
            ext = plantilla.extension or 'txt'
            pedido = session.get(Pedido, pedido_id)
            lab = (pedido.laboratorio or 'pedido').replace(' ', '_') if pedido else 'pedido'
            periodo = (pedido.periodo or '').replace(' ', '_') if pedido else ''

            lines = []
            for row in rows:
                line = bytearray(b' ' * line_len)
                for c in campos:
                    cs = c.campo_sistema
                    if cs == 'fijo':
                        val = c.valor_fijo or ''
                    elif cs == 'codigo_barra':
                        val = str(row.get('ean', '') or '')
                    elif cs == 'descripcion':
                        val = str(row.get('nombre', '') or '')
                    elif cs == 'cantidad':
                        val = str(int(row.get('cantidad', 0) or 0))
                    elif cs == 'cant_modulo':
                        val = str(int(row.get('cant_modulo', 0) or 0))
                    elif cs == 'cant_oferta':
                        val = str(int(row.get('cant_oferta', 0) or 0))
                    elif cs == 'cant_oferta_min':
                        val = str(int(row.get('cant_oferta_min', 0) or 0))
                    elif cs == 'cant_nodeal':
                        val = str(int(row.get('cant_nodeal', 0) or 0))
                    elif cs == 'precio':
                        val = str(row.get('precio_pvp', '') or '')
                    elif cs == 'erp_qty':
                        val = str(row.get('erp_qty', '') or '')
                    elif cs == 'rotacion':
                        val = str(row.get('rotacion', '') or '')
                    elif cs == 'avg_monthly':
                        _v = row.get('avg_monthly', 0) or 0
                        val = str(int(round(float(_v)))) if _v else ''
                    elif cs == 'espacio':
                        val = ''
                    else:
                        val = ''
                    pad = (c.relleno or ' ')[0]
                    lng = c.longitud
                    val = val[-lng:].rjust(lng, pad) if c.alineacion == 'R' else val[:lng].ljust(lng, pad)
                    start = c.col_inicio
                    end = min(start + lng, line_len)
                    encoded = val.encode('latin-1', errors='replace')[:end - start]
                    line[start:start + len(encoded)] = encoded
                lines.append(bytes(line).decode('latin-1'))

        content = '\r\n'.join(lines) + '\r\n'
        filename = f'pedido_{lab}_{periodo}.{ext}'
        return Response(
            content.encode('latin-1'),
            mimetype='text/plain',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )

    @app.route('/order/<int:pedido_id>/canal', methods=['POST'])
    def order_set_canal(pedido_id):
        """Setea el canal de compra del pedido: laboratorio (directo) o droguería.

        Body JSON:
          {canal: 'laboratorio'|'drogueria', partner_id: int|null}
        """
        body = request.get_json(silent=True) or {}
        canal = (body.get('canal') or '').strip()
        if canal not in ('laboratorio', 'drogueria', ''):
            return jsonify({'error': 'canal inválido'}), 400
        partner_id = body.get('partner_id')
        try:
            partner_id = int(partner_id) if partner_id else None
        except (ValueError, TypeError):
            return jsonify({'error': 'partner_id inválido'}), 400

        with database.get_db() as session:
            pedido = session.get(Pedido, pedido_id)
            if not pedido:
                return jsonify({'error': 'Pedido no encontrado'}), 404
            if canal == 'drogueria' and not partner_id:
                return jsonify({'error': 'Para canal=drogueria hace falta partner_id'}), 400
            # Resolver partner si corresponde
            partner_nombre = None
            if canal == 'drogueria':
                prov = session.get(database.Provider, partner_id)
                if not prov:
                    return jsonify({'error': 'Droguería no encontrada'}), 404
                partner_nombre = prov.razon_social
            elif canal == 'laboratorio':
                # El partner es el laboratorio fabricante.
                # Si existe en Laboratorio, resolvemos; si no, queda solo el string.
                lab = (session.query(Laboratorio)
                       .filter_by(nombre=pedido.laboratorio).first())
                if lab:
                    partner_id = lab.id
                    partner_nombre = lab.nombre
                else:
                    partner_id = None
                    partner_nombre = pedido.laboratorio
            pedido.canal = canal or None
            pedido.partner_id = partner_id
            pedido.canal_elegido_en = now_ar() if canal else None
            session.commit()
            return jsonify({
                'ok': True,
                'canal': pedido.canal,
                'partner_id': pedido.partner_id,
                'partner_nombre': partner_nombre,
                'canal_elegido_en': pedido.canal_elegido_en.isoformat()
                                    if pedido.canal_elegido_en else None,
            })

    @app.route('/order/<int:pedido_id>/export-plantilla/<int:plantilla_id>', methods=['POST'])
    def order_export_plantilla_unified(pedido_id, plantilla_id):
        """Exporta usando la Plantilla unificada (xlsx | csv | txt_fijo)."""
        import json as _json
        from io import BytesIO, StringIO

        from flask import Response, send_file

        body = request.get_json(silent=True) or {}
        rows = body.get('rows') or []

        with database.get_db() as session:
            pedido = session.get(database.Pedido, pedido_id)
            if not pedido:
                return jsonify({'error': 'Pedido no encontrado'}), 404
            plant = session.get(database.Plantilla, plantilla_id)
            if not plant:
                return jsonify({'error': 'Plantilla no encontrada'}), 404
            try:
                cfg = _json.loads(plant.config_json or '{}')
            except Exception:
                cfg = {}
            formato = plant.formato
            nombre_plant = plant.nombre
            periodo = (pedido.periodo or '').replace(' ', '_')
            lab = (pedido.laboratorio or 'pedido').replace(' ', '_')

        def _norm_cols(raw):
            """Normaliza columnas: acepta strings o dicts. Devuelve lista de dicts
            con keys `field`, `label`, `enabled`."""
            out = []
            for c in (raw or []):
                if isinstance(c, str):
                    out.append({'field': c, 'label': c, 'enabled': True})
                elif isinstance(c, dict):
                    if not c.get('enabled', True):
                        continue
                    out.append({
                        'field': c.get('field') or c.get('campo') or c.get('campo_sistema') or '',
                        'label': c.get('label') or c.get('field') or c.get('campo') or '',
                        'enabled': True,
                    })
            return [c for c in out if c['field']]

        def _val(row, field):
            # Mapeo de campo_sistema → key en rows construido por el front
            if field == 'fijo':            return ''
            if field in ('codigo_barra',): return str(row.get('ean', '') or '')
            if field in ('descripcion',):  return str(row.get('nombre', '') or '')
            if field == 'cantidad':        return str(int(row.get('cantidad', row.get('total', 0)) or 0))
            if field == 'cant_modulo':     return str(int(row.get('cant_modulo', 0) or 0))
            if field == 'cant_oferta':     return str(int(row.get('cant_oferta', 0) or 0))
            if field == 'cant_oferta_min': return str(int(row.get('cant_oferta_min', 0) or 0))
            if field == 'cant_nodeal':     return str(int(row.get('cant_nodeal', 0) or 0))
            if field == 'precio':          return str(row.get('precio_pvp', '') or '')
            if field == 'erp_qty':         return str(row.get('erp_qty', '') or '')
            if field == 'rotacion':        return str(row.get('rotacion', '') or '')
            if field == 'avg_monthly':
                v = row.get('avg_monthly', 0) or 0
                return str(int(round(float(v)))) if v else ''
            if field == 'espacio':         return ''
            return str(row.get(field, '') or '')

        if formato == 'txt_fijo':
            campos = sorted(cfg.get('campos', []), key=lambda c: c.get('col_inicio', 0))
            if not campos:
                return jsonify({'error': 'Plantilla sin campos'}), 400
            line_len = max(c.get('col_inicio', 0) + c.get('longitud', 0) for c in campos)
            lines = []
            for row in rows:
                line = bytearray(b' ' * line_len)
                for c in campos:
                    cs = c.get('campo', c.get('campo_sistema', ''))
                    val = c.get('valor_fijo', '') if cs == 'fijo' else _val(row, cs)
                    pad = (c.get('relleno') or ' ')[0]
                    lng = c.get('longitud', 0)
                    align = c.get('alineacion', 'L')
                    val = val[-lng:].rjust(lng, pad) if align == 'R' else val[:lng].ljust(lng, pad)
                    start = c.get('col_inicio', 0)
                    end = min(start + lng, line_len)
                    encoded = val.encode('latin-1', errors='replace')[:end - start]
                    line[start:start + len(encoded)] = encoded
                lines.append(bytes(line).decode('latin-1'))
            content = '\r\n'.join(lines) + '\r\n'
            fname = f'{nombre_plant}_{lab}_{periodo}.txt'.replace(' ', '_')
            return Response(content.encode(cfg.get('encoding', 'latin-1'), errors='replace'),
                            mimetype='text/plain',
                            headers={'Content-Disposition': f'attachment; filename="{fname}"'})

        if formato == 'csv':
            import csv as _csv
            cols = _norm_cols(cfg.get('columnas'))
            if not cols:
                return jsonify({'error': 'Plantilla sin columnas activas'}), 400
            buf = StringIO()
            delim = cfg.get('separador') or cfg.get('delimiter') or ','
            w = _csv.writer(buf, delimiter=delim)
            w.writerow([c['label'] for c in cols])
            for row in rows:
                w.writerow([_val(row, c['field']) for c in cols])
            fname = f'{nombre_plant}_{lab}_{periodo}.csv'.replace(' ', '_')
            return Response(buf.getvalue(), mimetype='text/csv',
                            headers={'Content-Disposition': f'attachment; filename="{fname}"'})

        # xlsx (default)
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        cols = _norm_cols(cfg.get('columnas'))
        if not cols:
            return jsonify({'error': 'Plantilla sin columnas activas'}), 400
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Pedido'
        row_offset = 1
        hdr_txt = cfg.get('custom_header')
        if hdr_txt:
            ws.cell(row=1, column=1, value=hdr_txt).font = Font(bold=True, size=12)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
            row_offset = 2
        hdr_fill = PatternFill('solid', fgColor='1e1e1e')
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=row_offset, column=ci, value=col.get('label') or col['field'])
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
        for ri, row in enumerate(rows, row_offset + 1):
            for ci, col in enumerate(cols, 1):
                ws.cell(row=ri, column=ci, value=_val(row, col['field']) or None)
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        fname = f'{nombre_plant}_{lab}_{periodo}.xlsx'.replace(' ', '_')
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/order/<int:pedido_id>/export/<step>/<fmt>', methods=['POST'])
    def order_export(pedido_id, step, fmt):
        """step: modules | offers | nodeal | summary.  fmt: xlsx | pdf"""
        from io import BytesIO

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        raw = request.form.get('data')
        if not raw:
            return 'Sin datos', 400
        data = json.loads(raw)
        lab  = request.form.get('laboratorio', 'Pedido')
        periodo = request.form.get('periodo', '')

        wb = openpyxl.Workbook()
        ws = wb.active

        hdr_fill  = PatternFill('solid', fgColor='1C1C1E')
        hdr_font  = Font(bold=True, color='EAB308')
        mod_fill  = PatternFill('solid', fgColor='FEF9C3')
        mod_font  = Font(bold=True, color='92400E')
        bold      = Font(bold=True)
        thin      = Side(style='thin', color='D0D0D0')
        border    = Border(bottom=thin)
        center    = Alignment(horizontal='center')
        right_al  = Alignment(horizontal='right')

        def hrow(ws, values, fill=hdr_fill, font=hdr_font):
            r = ws.max_row + 1
            for ci, v in enumerate(values, 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.fill = fill; c.font = font; c.border = border

        ws.append([f'{lab} — {periodo}'])
        ws['A1'].font = Font(bold=True, size=13)
        ws.append([])

        if step == 'modules':
            hrow(ws, ['Módulo', 'EAN', 'Descripción', 'Cant/Módulo', 'Cant.Pedida',
                      'Cant.Calculada', 'Propuesta', 'Saldo', 'Desc%'])
            ws.column_dimensions['A'].width = 22
            ws.column_dimensions['C'].width = 38
            for mod in data:
                r = ws.max_row + 1
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
                c = ws.cell(row=r, column=1, value=f"{mod['nombre']}  ·  {mod.get('modulos_sugeridos', '')} módulos sugeridos")
                c.fill = mod_fill; c.font = mod_font
                for it in mod.get('items', []):
                    ws.append([
                        mod['nombre'],
                        it.get('ean', ''),
                        it.get('descripcion', ''),
                        it.get('cant', ''),
                        it.get('cant_pedida', ''),
                        it.get('cant_calculada', ''),
                        it.get('propuesta', ''),
                        it.get('saldo', ''),
                        it.get('desc_pct', ''),
                    ])

        elif step == 'offers':
            hrow(ws, ['EAN', 'Descripción', 'Cant. a pedir'])
            ws.column_dimensions['B'].width = 42
            for it in data:
                ws.append([it.get('ean', ''), it.get('nombre', ''), it.get('cantidad', '')])

        elif step == 'nodeal':
            hrow(ws, ['EAN', 'Descripción', 'Cant. a pedir'])
            ws.column_dimensions['B'].width = 42
            for it in data:
                ws.append([it.get('ean', ''), it.get('nombre', ''), it.get('cantidad', '')])

        elif step == 'summary':
            rows = data if isinstance(data, list) else []

            hrow(ws, ['EAN', 'Producto', 'Stock ERP', 'Rot.', 'Prom.mes',
                      'Precio PVP', 'Cant. módulo', 'Cant. oferta', 'Oferta c/mín', 'Sin Deal',
                      'Total', 'Cant. pedida', 'Saldo'])
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 6
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 10
            ws.column_dimensions['K'].width = 10
            ws.column_dimensions['L'].width = 12
            ws.column_dimensions['M'].width = 10

            for row in rows:
                saldo = row.get('saldo', '')
                ws.append([
                    row.get('ean', ''),
                    row.get('nombre', ''),
                    row.get('erp_qty', '') if row.get('erp_qty') is not None else '',
                    row.get('rotacion', ''),
                    row.get('avg_monthly', '') if row.get('avg_monthly') is not None else '',
                    row.get('precio_pvp', '') if row.get('precio_pvp') else '',
                    row.get('cant_modulo', '') if row.get('cant_modulo') else '',
                    row.get('cant_oferta', '') if row.get('cant_oferta') else '',
                    row.get('cant_oferta_min', '') if row.get('cant_oferta_min') else '',
                    row.get('cant_nodeal', '') if row.get('cant_nodeal') else '',
                    row.get('total', ''),
                    row.get('cant_pedida', ''),
                    saldo if saldo != '' else '',
                ])
                saldo_val = row.get('saldo')
                if saldo_val is not None:
                    from openpyxl.styles import PatternFill as _PF
                    if saldo_val > 0:
                        ws.cell(row=ws.max_row, column=13).fill = _PF(fill_type='solid', fgColor='FEE2E2')
                    elif saldo_val < 0:
                        ws.cell(row=ws.max_row, column=13).fill = _PF(fill_type='solid', fgColor='D1FAE5')

        if fmt == 'xlsx':
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            step_names = {'modules': 'Modulos', 'offers': 'Ofertas',
                          'nodeal': 'SinDeal', 'summary': 'Resumen'}
            fname = f"{lab}_{step_names.get(step, step)}.xlsx"
            resp = make_response(buf.read())
            resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
            return resp

        return 'Formato no soportado', 400
