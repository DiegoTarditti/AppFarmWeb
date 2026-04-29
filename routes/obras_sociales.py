"""Obras Sociales — dashboard y listado de dispensas.

⚠ AÚN MOCK DATA. Plan de migración a datos reales:
1. Sincronizar `DW.ProductosVendidos` detalle → tabla local `obs_ventas_detalle`
   (ya está confirmado el schema, ver `c:/AppSeguimiento/10-obras-sociales.md`).
2. Crear vista local `v_dispensas_os` que junte ventas + cliente + OS + plan + médico.
3. Reemplazar `_mock_dispensas()` con queries reales contra esa vista.
La UI no debería cambiar — solo cambia la fuente de datos.
"""

import random
from datetime import date, datetime, timedelta

from flask import render_template, request

# ── Mock data ─────────────────────────────────────────────────────────

_MEDICOS = [
    ('12345', 'Dr. Juan Pérez'),
    ('23456', 'Dra. María González'),
    ('34567', 'Dr. Carlos Rodríguez'),
    ('45678', 'Dra. Laura Fernández'),
    ('56789', 'Dr. Roberto Martínez'),
    ('67890', 'Dra. Ana Sánchez'),
    ('78901', 'Dr. Diego López'),
]

_MEDICAMENTOS = [
    ('7791234567890', 'ATORVASTATINA 20MG x 30 COMP', 2450.00),
    ('7791234567891', 'METFORMINA 850MG x 30 COMP', 890.00),
    ('7791234567892', 'ENALAPRIL 10MG x 30 COMP', 560.00),
    ('7791234567893', 'LOSARTAN 50MG x 30 COMP', 1200.00),
    ('7791234567894', 'OMEPRAZOL 20MG x 14 CAP', 780.00),
    ('7791234567895', 'LEVOTIROXINA 100MCG x 50 COMP', 1850.00),
    ('7791234567896', 'AMLODIPINA 5MG x 30 COMP', 920.00),
    ('7791234567897', 'INSULINA GLARGINA 100UI/ML FCO', 48500.00),
    ('7791234567898', 'ACIDO ACETILSALICILICO 100MG x 30', 340.00),
    ('7791234567899', 'CLOPIDOGREL 75MG x 28 COMP', 3200.00),
    ('7791234567900', 'SALBUTAMOL AEROSOL 200 DOSIS', 2100.00),
    ('7791234567901', 'PARACETAMOL 500MG x 20 COMP', 280.00),
]

_AFILIADOS_PAMI = [
    ('150234567/00', '20-12345678-9', 'GÓMEZ, MARIA ELENA'),
    ('150234568/01', '27-23456789-0', 'LÓPEZ, ROBERTO'),
    ('150234569/00', '20-34567890-1', 'FERNÁNDEZ, CARLOS'),
    ('150234570/02', '27-45678901-2', 'MARTÍNEZ, SUSANA'),
    ('150234571/00', '20-56789012-3', 'GONZÁLEZ, JUAN'),
    ('150234572/01', '27-67890123-4', 'RODRÍGUEZ, ANA'),
]

_AFILIADOS_IAPOS = [
    ('A-1234567', '20-11223344-5', 'PÉREZ, LUIS'),
    ('A-2345678', '27-22334455-6', 'SÁNCHEZ, MARTA'),
    ('A-3456789', '20-33445566-7', 'DÍAZ, PABLO'),
    ('A-4567890', '27-44556677-8', 'RAMÍREZ, SILVIA'),
    ('A-5678901', '20-55667788-9', 'TORRES, MIGUEL'),
]

_PLANES = {
    'PAMI': ['Básico', 'PROFE', 'Convenio', 'Oncológico'],
    'IAPOS': ['Ambulatorio', 'Insulina', 'Crónicos', 'PMO'],
}

_MOCK_CACHE = None


def _mock_dispensas():
    """Genera 350 dispensas mockeadas distribuidas en los últimos 45 días."""
    global _MOCK_CACHE
    if _MOCK_CACHE is not None:
        return _MOCK_CACHE

    rng = random.Random(42)
    rows = []
    hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    for i in range(1, 351):
        os_code = rng.choices(['PAMI', 'IAPOS'], weights=[0.65, 0.35])[0]
        afil = rng.choice(_AFILIADOS_PAMI if os_code == 'PAMI' else _AFILIADOS_IAPOS)
        med = rng.choice(_MEDICOS)
        plan = rng.choice(_PLANES[os_code])
        tipo_receta = rng.choices(['electronica', 'manual', 'OPF'], weights=[0.80, 0.10, 0.10])[0]
        n_items = rng.choices([1, 2, 3, 4], weights=[0.55, 0.28, 0.12, 0.05])[0]
        dias_atras = rng.randint(0, 44)
        hora = rng.randint(8, 20)
        minuto = rng.randint(0, 59)
        fecha_venta = hoy - timedelta(days=dias_atras) + timedelta(hours=hora, minutes=minuto)
        fecha_emision = fecha_venta - timedelta(days=rng.randint(0, 25))
        ticket = f'{os_code[0]}{rng.randint(100000, 999999)}'
        items = []
        importe_os_total = 0.0
        importe_pac_total = 0.0
        for _ in range(n_items):
            troquel, desc, pvp = rng.choice(_MEDICAMENTOS)
            cant = rng.choices([1, 2, 3], weights=[0.75, 0.20, 0.05])[0]
            cobertura = rng.choice([40, 50, 60, 70, 100])
            importe_total = pvp * cant
            importe_os = round(importe_total * cobertura / 100, 2)
            importe_pac = round(importe_total - importe_os, 2)
            importe_os_total += importe_os
            importe_pac_total += importe_pac
            items.append({
                'troquel': troquel,
                'descripcion': desc,
                'cantidad': cant,
                'pvp_unitario': pvp,
                'cobertura_pct': cobertura,
                'importe_os': importe_os,
                'importe_paciente': importe_pac,
            })
        rows.append({
            'dispensa_id': i,
            'receta_id': f'R-{i:05d}',
            'fecha_venta': fecha_venta,
            'os_codigo': os_code,
            'os_plan': plan,
            'ticket_validacion': ticket,
            'afiliado_nro': afil[0],
            'afiliado_dni': afil[1],
            'afiliado_nombre': afil[2],
            'medico_matricula': med[0],
            'medico_nombre': med[1],
            'receta_fecha_emision': fecha_emision,
            'receta_tipo': tipo_receta,
            'items': items,
            'importe_os': round(importe_os_total, 2),
            'importe_paciente': round(importe_pac_total, 2),
            'importe_total': round(importe_os_total + importe_pac_total, 2),
            'lote_nro': None,
            'estado': 'pendiente',
        })
    _MOCK_CACHE = rows
    return rows


# ── Rutas ─────────────────────────────────────────────────────────────

def init_app(app):

    @app.route('/obras-sociales')
    def os_index():
        """Hub del módulo OS con cards estilo /informes, agrupadas por
        dimensión de análisis (macro / médico / medicamento / paciente /
        convenio / operativo / temporal / estratégico).
        """
        return render_template('os_index.html')

    @app.route('/obras-sociales/rentabilidad')
    def os_rentabilidad():
        """Rentabilidad real por OS: PVP - costo estimado - descuento OS.

        El "costo estimado" se calcula cruzando obs_ventas_detalle.producto_observer
        contra el último precio_unitario de factura_items (vía producto_codigos_barra
        + productos.observer_id). Productos sin compra registrada quedan fuera del
        cálculo de costo (se reportan como 'sin costo conocido').
        """
        from datetime import date as _date
        from datetime import datetime as _dt

        from sqlalchemy import text as _text

        import database

        hoy = _dt.now().date()
        desde_str = (request.args.get('desde') or '').strip()
        hasta_str = (request.args.get('hasta') or '').strip()
        try:
            desde = _date.fromisoformat(desde_str) if desde_str else hoy.replace(day=1)
        except ValueError:
            desde = hoy.replace(day=1)
        try:
            hasta = _date.fromisoformat(hasta_str) if hasta_str else hoy
        except ValueError:
            hasta = hoy
        if hasta < desde:
            desde, hasta = hasta, desde

        return render_template('os_rentabilidad.html',
                               desde=desde.isoformat(), hasta=hasta.isoformat())

    @app.route('/api/obras-sociales/rentabilidad')
    def api_os_rentabilidad():
        """JSON con KPIs + ranking + alertas + serie temporal.

        Query params: desde, hasta (YYYY-MM-DD).
        """
        from datetime import date as _date
        from datetime import datetime as _dt

        from sqlalchemy import text as _text

        import database

        hoy = _dt.now().date()
        try:
            desde = _date.fromisoformat(request.args.get('desde', '')) or hoy.replace(day=1)
        except ValueError:
            desde = hoy.replace(day=1)
        try:
            hasta = _date.fromisoformat(request.args.get('hasta', '')) or hoy
        except ValueError:
            hasta = hoy

        with database.get_db() as session:
            # Query principal: agrega por OS, calcula costo cruzando productos
            sql = _text('''
                WITH ultima_compra AS (
                    SELECT codigo_barra, precio_unitario,
                           ROW_NUMBER() OVER (PARTITION BY codigo_barra
                                              ORDER BY id DESC) AS rn
                    FROM factura_items
                    WHERE precio_unitario IS NOT NULL AND precio_unitario > 0
                      AND codigo_barra IS NOT NULL
                ),
                costo_por_observer AS (
                    SELECT p.observer_id, MAX(uc.precio_unitario) AS costo_unitario
                    FROM ultima_compra uc
                    JOIN producto_codigos_barra pcb ON pcb.codigo_barra = uc.codigo_barra
                    JOIN productos p ON p.id = pcb.producto_id
                    WHERE uc.rn = 1 AND p.observer_id IS NOT NULL
                    GROUP BY p.observer_id
                )
                SELECT
                    ovd.obra_social_observer AS os_id,
                    COALESCE(oos.descripcion, 'Sin OS') AS os_nombre,
                    SUM(ovd.importe) AS facturado,
                    SUM(COALESCE(ovd.importe_a_cargo_os, 0)) AS a_cargo_os,
                    SUM(CASE WHEN cpo.costo_unitario IS NOT NULL
                             THEN ovd.cantidad * cpo.costo_unitario ELSE 0 END) AS costo_estimado,
                    SUM(CASE WHEN cpo.costo_unitario IS NOT NULL
                             THEN ovd.importe ELSE 0 END) AS facturado_con_costo,
                    COUNT(*) AS dispensas,
                    SUM(CASE WHEN cpo.costo_unitario IS NULL THEN 1 ELSE 0 END) AS dispensas_sin_costo,
                    COUNT(DISTINCT ovd.id_operacion) AS recetas
                FROM obs_ventas_detalle ovd
                LEFT JOIN costo_por_observer cpo ON cpo.observer_id = ovd.producto_observer
                LEFT JOIN obs_obras_sociales oos ON oos.observer_id = ovd.obra_social_observer
                WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                  AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                  AND ovd.obra_social_observer IS NOT NULL
                GROUP BY ovd.obra_social_observer, oos.descripcion
                ORDER BY (SUM(ovd.importe) - SUM(CASE WHEN cpo.costo_unitario IS NOT NULL
                                                     THEN ovd.cantidad * cpo.costo_unitario ELSE 0 END)) DESC
            ''')
            rows = session.execute(sql, {'desde': desde, 'hasta': hasta}).fetchall()

            ranking = []
            tot_fact = 0.0
            tot_costo = 0.0
            tot_a_cargo = 0.0
            tot_dispensas = 0
            tot_dispensas_sin_costo = 0
            tot_facturado_con_costo = 0.0
            for r in rows:
                fact = float(r.facturado or 0)
                costo = float(r.costo_estimado or 0)
                fact_cc = float(r.facturado_con_costo or 0)
                a_cargo = float(r.a_cargo_os or 0)
                ganancia = fact_cc - costo  # solo de líneas con costo conocido
                margen_pct = (ganancia / fact_cc * 100) if fact_cc > 0 else None
                cobertura_costo_pct = (fact_cc / fact * 100) if fact > 0 else 0
                ranking.append({
                    'os_id': r.os_id,
                    'os_nombre': r.os_nombre,
                    'facturado': round(fact, 2),
                    'a_cargo_os': round(a_cargo, 2),
                    'costo_estimado': round(costo, 2),
                    'facturado_con_costo': round(fact_cc, 2),
                    'ganancia': round(ganancia, 2),
                    'margen_pct': round(margen_pct, 2) if margen_pct is not None else None,
                    'cobertura_costo_pct': round(cobertura_costo_pct, 1),
                    'dispensas': int(r.dispensas or 0),
                    'dispensas_sin_costo': int(r.dispensas_sin_costo or 0),
                    'recetas': int(r.recetas or 0),
                })
                tot_fact += fact
                tot_costo += costo
                tot_a_cargo += a_cargo
                tot_facturado_con_costo += fact_cc
                tot_dispensas += int(r.dispensas or 0)
                tot_dispensas_sin_costo += int(r.dispensas_sin_costo or 0)

            tot_ganancia = tot_facturado_con_costo - tot_costo
            tot_margen_pct = (tot_ganancia / tot_facturado_con_costo * 100) if tot_facturado_con_costo > 0 else None
            tot_cobertura = (tot_facturado_con_costo / tot_fact * 100) if tot_fact > 0 else 0

            # Alertas: OS con margen negativo o muy bajo
            alertas = []
            for r in ranking:
                if r['margen_pct'] is not None:
                    if r['margen_pct'] < 0:
                        alertas.append({
                            'tipo': 'critico',
                            'os_nombre': r['os_nombre'],
                            'mensaje': f"Margen NEGATIVO ({r['margen_pct']}%) — perdés ${abs(r['ganancia']):,.0f} en {r['dispensas']} dispensas",
                        })
                    elif r['margen_pct'] < 15:
                        alertas.append({
                            'tipo': 'advertencia',
                            'os_nombre': r['os_nombre'],
                            'mensaje': f"Margen bajo ({r['margen_pct']}%). Revisar descuento o reemplazar.",
                        })

            # Serie temporal: facturado y ganancia por día
            sql_dia = _text('''
                WITH ultima_compra AS (
                    SELECT codigo_barra, precio_unitario,
                           ROW_NUMBER() OVER (PARTITION BY codigo_barra ORDER BY id DESC) AS rn
                    FROM factura_items
                    WHERE precio_unitario IS NOT NULL AND precio_unitario > 0
                      AND codigo_barra IS NOT NULL
                ),
                costo_por_observer AS (
                    SELECT p.observer_id, MAX(uc.precio_unitario) AS costo_unitario
                    FROM ultima_compra uc
                    JOIN producto_codigos_barra pcb ON pcb.codigo_barra = uc.codigo_barra
                    JOIN productos p ON p.id = pcb.producto_id
                    WHERE uc.rn = 1 AND p.observer_id IS NOT NULL
                    GROUP BY p.observer_id
                )
                SELECT
                    ovd.fecha_estadistica AS fecha,
                    SUM(ovd.importe) AS facturado,
                    SUM(CASE WHEN cpo.costo_unitario IS NOT NULL
                             THEN ovd.cantidad * cpo.costo_unitario ELSE 0 END) AS costo,
                    SUM(CASE WHEN cpo.costo_unitario IS NOT NULL
                             THEN ovd.importe ELSE 0 END) AS facturado_con_costo
                FROM obs_ventas_detalle ovd
                LEFT JOIN costo_por_observer cpo ON cpo.observer_id = ovd.producto_observer
                WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                  AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                  AND ovd.obra_social_observer IS NOT NULL
                GROUP BY ovd.fecha_estadistica
                ORDER BY ovd.fecha_estadistica
            ''')
            rows_dia = session.execute(sql_dia, {'desde': desde, 'hasta': hasta}).fetchall()
            serie = [{
                'fecha': r.fecha.isoformat() if r.fecha else None,
                'facturado': float(r.facturado or 0),
                'ganancia': float(r.facturado_con_costo or 0) - float(r.costo or 0),
                'costo': float(r.costo or 0),
            } for r in rows_dia]

        return jsonify({
            'desde': desde.isoformat(),
            'hasta': hasta.isoformat(),
            'kpis': {
                'facturado_total': round(tot_fact, 2),
                'costo_estimado': round(tot_costo, 2),
                'ganancia_bruta': round(tot_ganancia, 2),
                'margen_pct': round(tot_margen_pct, 2) if tot_margen_pct is not None else None,
                'a_cargo_os': round(tot_a_cargo, 2),
                'a_cargo_os_pct': round(tot_a_cargo / tot_fact * 100, 1) if tot_fact > 0 else 0,
                'dispensas': tot_dispensas,
                'dispensas_sin_costo': tot_dispensas_sin_costo,
                'cobertura_costo_pct': round(tot_cobertura, 1),
            },
            'ranking': ranking,
            'alertas': alertas,
            'serie_temporal': serie,
        })

    @app.route('/obras-sociales/pacientes')
    def os_pacientes():
        """Análisis por paciente prescripto: Pareto, inactivos, drogas frecuentes."""
        from datetime import date as _date
        from datetime import datetime as _dt

        hoy = _dt.now().date()
        try:
            desde = _date.fromisoformat(request.args.get('desde', '')) or hoy.replace(day=1)
        except ValueError:
            desde = hoy.replace(day=1)
        try:
            hasta = _date.fromisoformat(request.args.get('hasta', '')) or hoy
        except ValueError:
            hasta = hoy
        return render_template('os_pacientes.html',
                               desde=desde.isoformat(), hasta=hasta.isoformat())

    @app.route('/api/obras-sociales/pacientes')
    def api_os_pacientes():
        """JSON: ranking + inactivos + drogas top."""
        from datetime import date as _date
        from datetime import datetime as _dt
        from datetime import timedelta as _td

        from sqlalchemy import text as _text

        import database

        hoy = _dt.now().date()
        try:
            desde = _date.fromisoformat(request.args.get('desde', '')) or hoy.replace(day=1)
        except ValueError:
            desde = hoy.replace(day=1)
        try:
            hasta = _date.fromisoformat(request.args.get('hasta', '')) or hoy
        except ValueError:
            hasta = hoy
        try:
            os_id = int(request.args.get('os_id', '')) or None
        except ValueError:
            os_id = None

        os_filter_sql = ''
        params = {'desde': desde, 'hasta': hasta, 'hoy': hoy}
        if os_id:
            os_filter_sql = 'AND ovd.obra_social_observer = :os_id'
            params['os_id'] = os_id

        with database.get_db() as session:
            # Ranking principal
            sql_ranking = _text(f'''
                WITH actual AS (
                    SELECT
                        ovd.cliente_observer AS cli_id,
                        SUM(ovd.importe) AS importe,
                        SUM(COALESCE(ovd.importe_a_cargo_os, 0)) AS importe_os,
                        COUNT(DISTINCT ovd.id_operacion) AS recetas,
                        COUNT(DISTINCT ovd.medico_observer) AS medicos,
                        COUNT(DISTINCT ovd.producto_observer) AS productos,
                        COUNT(DISTINCT ovd.obra_social_observer) AS n_os,
                        MAX(ovd.fecha_estadistica) AS ultima_visita,
                        MIN(ovd.fecha_estadistica) AS primera_visita
                    FROM obs_ventas_detalle ovd
                    WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                      AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                      AND ovd.cliente_observer IS NOT NULL
                      {os_filter_sql}
                    GROUP BY ovd.cliente_observer
                ),
                os_principal AS (
                    SELECT cli_id, os_nombre, recetas_os
                    FROM (
                        SELECT
                            ovd.cliente_observer AS cli_id,
                            COALESCE(oos.descripcion, 'Sin OS') AS os_nombre,
                            COUNT(DISTINCT ovd.id_operacion) AS recetas_os,
                            ROW_NUMBER() OVER (PARTITION BY ovd.cliente_observer
                                               ORDER BY COUNT(DISTINCT ovd.id_operacion) DESC) AS rn
                        FROM obs_ventas_detalle ovd
                        LEFT JOIN obs_obras_sociales oos ON oos.observer_id = ovd.obra_social_observer
                        WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                          AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                          AND ovd.cliente_observer IS NOT NULL
                        GROUP BY ovd.cliente_observer, oos.descripcion
                    ) sub
                    WHERE rn = 1
                )
                SELECT
                    a.cli_id,
                    oc.apellido_nombre AS nombre,
                    oc.documento_numero AS dni,
                    oc.localidad AS localidad,
                    a.importe,
                    a.importe_os,
                    a.recetas,
                    a.medicos,
                    a.productos,
                    a.n_os,
                    a.ultima_visita,
                    a.primera_visita,
                    osp.os_nombre AS os_principal,
                    osp.recetas_os
                FROM actual a
                LEFT JOIN obs_clientes oc ON oc.observer_id = a.cli_id
                LEFT JOIN os_principal osp ON osp.cli_id = a.cli_id
                ORDER BY a.importe DESC
                LIMIT 200
            ''')
            rows = session.execute(sql_ranking, params).fetchall()
            ranking = []
            tot_imp = tot_imp_os = tot_recetas = 0
            for r in rows:
                imp = float(r.importe or 0)
                imp_os = float(r.importe_os or 0)
                pct_top_os = (r.recetas_os / r.recetas * 100) if r.recetas else 0
                dias_desde_visita = (hoy - r.ultima_visita).days if r.ultima_visita else None
                ranking.append({
                    'cli_id': r.cli_id,
                    'nombre': r.nombre or f'Cliente #{r.cli_id}',
                    'dni': str(r.dni) if r.dni else '',
                    'localidad': r.localidad or '',
                    'importe': round(imp, 2),
                    'importe_os': round(imp_os, 2),
                    'pct_os_paga': round(imp_os / imp * 100, 1) if imp > 0 else 0,
                    'recetas': int(r.recetas or 0),
                    'medicos': int(r.medicos or 0),
                    'productos': int(r.productos or 0),
                    'n_os': int(r.n_os or 0),
                    'ticket_prom': round(imp / r.recetas, 2) if r.recetas else 0,
                    'os_principal': r.os_principal or '',
                    'os_principal_pct': round(pct_top_os, 1),
                    'dias_desde_visita': dias_desde_visita,
                    'ultima_visita': r.ultima_visita.isoformat() if r.ultima_visita else None,
                    'primera_visita': r.primera_visita.isoformat() if r.primera_visita else None,
                })
                tot_imp += imp
                tot_imp_os += imp_os
                tot_recetas += int(r.recetas or 0)

            # Pareto top 5%
            n_top5pct = max(1, len(ranking) // 20)
            top5pct_imp = sum(r['importe'] for r in ranking[:n_top5pct])
            top5pct_pct = (top5pct_imp / tot_imp * 100) if tot_imp > 0 else 0

            # Pacientes inactivos (recurrentes históricos que no volvieron)
            # Definición: tenían >=4 recetas en últimos 12m hasta hace 90d, y 0 recetas
            # en los últimos 60d. Probables crónicos que dejaron de venir.
            sql_inact = _text('''
                WITH base AS (
                    SELECT
                        ovd.cliente_observer AS cli_id,
                        COUNT(DISTINCT ovd.id_operacion) AS recetas_hist,
                        SUM(ovd.importe) AS importe_hist,
                        MAX(ovd.fecha_estadistica) AS ultima_visita
                    FROM obs_ventas_detalle ovd
                    WHERE ovd.fecha_estadistica BETWEEN (:hoy - INTERVAL '12 months') AND :hoy
                      AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                      AND ovd.cliente_observer IS NOT NULL
                    GROUP BY ovd.cliente_observer
                    HAVING COUNT(DISTINCT ovd.id_operacion) >= 4
                       AND MAX(ovd.fecha_estadistica) < (:hoy - INTERVAL '60 days')
                )
                SELECT
                    b.cli_id,
                    oc.apellido_nombre AS nombre,
                    oc.documento_numero AS dni,
                    oc.telefono,
                    b.recetas_hist,
                    b.importe_hist,
                    b.ultima_visita
                FROM base b
                LEFT JOIN obs_clientes oc ON oc.observer_id = b.cli_id
                ORDER BY b.importe_hist DESC
                LIMIT 50
            ''')
            rows_inact = session.execute(sql_inact, {'hoy': hoy}).fetchall()
            inactivos = []
            for r in rows_inact:
                dias = (hoy - r.ultima_visita).days if r.ultima_visita else None
                inactivos.append({
                    'cli_id': r.cli_id,
                    'nombre': r.nombre or f'Cliente #{r.cli_id}',
                    'dni': str(r.dni) if r.dni else '',
                    'telefono': r.telefono or '',
                    'recetas_hist': int(r.recetas_hist or 0),
                    'importe_hist': round(float(r.importe_hist or 0), 2),
                    'ultima_visita': r.ultima_visita.isoformat() if r.ultima_visita else None,
                    'dias': dias,
                })

            # OS list
            sql_os_list = _text('''
                SELECT DISTINCT oos.observer_id, oos.descripcion
                FROM obs_ventas_detalle ovd
                JOIN obs_obras_sociales oos ON oos.observer_id = ovd.obra_social_observer
                WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                  AND oos.fecha_baja IS NULL
                ORDER BY oos.descripcion
            ''')
            obras_sociales = [{'os_id': r[0], 'nombre': r[1]}
                              for r in session.execute(sql_os_list, {'desde': desde, 'hasta': hasta}).fetchall()]

        return jsonify({
            'desde': desde.isoformat(),
            'hasta': hasta.isoformat(),
            'os_id': os_id,
            'kpis': {
                'n_pacientes': len(ranking),
                'recetas': tot_recetas,
                'importe': round(tot_imp, 2),
                'importe_os': round(tot_imp_os, 2),
                'pct_os_paga': round(tot_imp_os / tot_imp * 100, 1) if tot_imp > 0 else 0,
                'top5pct_n': n_top5pct,
                'top5pct_imp': round(top5pct_imp, 2),
                'top5pct_pct': round(top5pct_pct, 1),
                'n_inactivos': len(inactivos),
                'importe_inactivos': round(sum(i['importe_hist'] for i in inactivos), 2),
                'ticket_promedio': round(tot_imp / tot_recetas, 2) if tot_recetas else 0,
            },
            'ranking': ranking,
            'inactivos': inactivos,
            'obras_sociales': obras_sociales,
        })

    @app.route('/obras-sociales/medicos')
    def os_medicos():
        """Análisis por médico prescriptor: ranking, fugas, heatmap OS."""
        from datetime import date as _date
        from datetime import datetime as _dt

        hoy = _dt.now().date()
        try:
            desde = _date.fromisoformat(request.args.get('desde', '')) or hoy.replace(day=1)
        except ValueError:
            desde = hoy.replace(day=1)
        try:
            hasta = _date.fromisoformat(request.args.get('hasta', '')) or hoy
        except ValueError:
            hasta = hoy
        return render_template('os_medicos.html',
                               desde=desde.isoformat(), hasta=hasta.isoformat())

    @app.route('/api/obras-sociales/medicos')
    def api_os_medicos():
        """JSON: ranking + fugas + heatmap por médico."""
        from datetime import date as _date
        from datetime import datetime as _dt
        from datetime import timedelta as _td

        from sqlalchemy import text as _text

        import database

        hoy = _dt.now().date()
        try:
            desde = _date.fromisoformat(request.args.get('desde', '')) or hoy.replace(day=1)
        except ValueError:
            desde = hoy.replace(day=1)
        try:
            hasta = _date.fromisoformat(request.args.get('hasta', '')) or hoy
        except ValueError:
            hasta = hoy
        try:
            os_id = int(request.args.get('os_id', '')) or None
        except ValueError:
            os_id = None

        # Período anterior del mismo largo (para detectar fugas)
        delta = (hasta - desde).days + 1
        desde_ant = desde - _td(days=delta)
        hasta_ant = desde - _td(days=1)

        os_filter_sql = ''
        params = {'desde': desde, 'hasta': hasta,
                  'desde_ant': desde_ant, 'hasta_ant': hasta_ant}
        if os_id:
            os_filter_sql = 'AND ovd.obra_social_observer = :os_id'
            params['os_id'] = os_id

        with database.get_db() as session:
            # Ranking principal: médico con sus métricas
            sql_ranking = _text(f'''
                WITH actual AS (
                    SELECT
                        ovd.medico_observer AS med_id,
                        SUM(ovd.importe) AS importe,
                        COUNT(DISTINCT ovd.id_operacion) AS recetas,
                        COUNT(DISTINCT ovd.cliente_observer) AS pacientes,
                        COUNT(DISTINCT ovd.producto_observer) AS productos,
                        COUNT(DISTINCT ovd.obra_social_observer) AS n_os,
                        MAX(ovd.fecha_estadistica) AS ultima_receta
                    FROM obs_ventas_detalle ovd
                    WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                      AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                      AND ovd.medico_observer IS NOT NULL
                      {os_filter_sql}
                    GROUP BY ovd.medico_observer
                ),
                anterior AS (
                    SELECT
                        ovd.medico_observer AS med_id,
                        SUM(ovd.importe) AS importe_ant,
                        COUNT(DISTINCT ovd.id_operacion) AS recetas_ant
                    FROM obs_ventas_detalle ovd
                    WHERE ovd.fecha_estadistica BETWEEN :desde_ant AND :hasta_ant
                      AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                      AND ovd.medico_observer IS NOT NULL
                      {os_filter_sql}
                    GROUP BY ovd.medico_observer
                ),
                os_principal AS (
                    SELECT med_id, os_nombre, os_id, recetas_os
                    FROM (
                        SELECT
                            ovd.medico_observer AS med_id,
                            ovd.obra_social_observer AS os_id,
                            COALESCE(oos.descripcion, 'Sin OS') AS os_nombre,
                            COUNT(DISTINCT ovd.id_operacion) AS recetas_os,
                            ROW_NUMBER() OVER (PARTITION BY ovd.medico_observer
                                               ORDER BY COUNT(DISTINCT ovd.id_operacion) DESC) AS rn
                        FROM obs_ventas_detalle ovd
                        LEFT JOIN obs_obras_sociales oos ON oos.observer_id = ovd.obra_social_observer
                        WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                          AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                          AND ovd.medico_observer IS NOT NULL
                        GROUP BY ovd.medico_observer, ovd.obra_social_observer, oos.descripcion
                    ) sub
                    WHERE rn = 1
                )
                SELECT
                    a.med_id,
                    om.nombre AS medico_nombre,
                    (SELECT MIN(matricula) FROM obs_medicos_matriculas mm
                     WHERE mm.medico_observer = a.med_id) AS matricula,
                    a.importe,
                    a.recetas,
                    a.pacientes,
                    a.productos,
                    a.n_os,
                    a.ultima_receta,
                    osp.os_nombre AS os_principal,
                    osp.recetas_os,
                    COALESCE(an.importe_ant, 0) AS importe_ant,
                    COALESCE(an.recetas_ant, 0) AS recetas_ant
                FROM actual a
                LEFT JOIN obs_medicos om ON om.observer_id = a.med_id
                LEFT JOIN anterior an ON an.med_id = a.med_id
                LEFT JOIN os_principal osp ON osp.med_id = a.med_id
                ORDER BY a.importe DESC
                LIMIT 200
            ''')
            rows = session.execute(sql_ranking, params).fetchall()

            ranking = []
            tot_imp = tot_recetas = tot_pac = 0
            for r in rows:
                imp = float(r.importe or 0)
                imp_ant = float(r.importe_ant or 0)
                var_pct = ((imp - imp_ant) / imp_ant * 100) if imp_ant > 0 else None
                pct_top_os = (r.recetas_os / r.recetas * 100) if r.recetas else 0
                ranking.append({
                    'med_id': r.med_id,
                    'nombre': r.medico_nombre or f'Médico #{r.med_id}',
                    'matricula': r.matricula or '',
                    'importe': round(imp, 2),
                    'recetas': int(r.recetas or 0),
                    'pacientes': int(r.pacientes or 0),
                    'productos': int(r.productos or 0),
                    'n_os': int(r.n_os or 0),
                    'ticket_prom': round(imp / r.recetas, 2) if r.recetas else 0,
                    'os_principal': r.os_principal or '',
                    'os_principal_pct': round(pct_top_os, 1),
                    'importe_ant': round(imp_ant, 2),
                    'recetas_ant': int(r.recetas_ant or 0),
                    'variacion_pct': round(var_pct, 1) if var_pct is not None else None,
                    'ultima_receta': r.ultima_receta.isoformat() if r.ultima_receta else None,
                })
                tot_imp += imp
                tot_recetas += int(r.recetas or 0)
                tot_pac += int(r.pacientes or 0)  # NB: doble-conteo si paciente atendido por 2 médicos

            # Pareto top 5
            top5_imp = sum(r['importe'] for r in ranking[:5])
            top5_pct = (top5_imp / tot_imp * 100) if tot_imp > 0 else 0

            # Médicos con FUGA (estaban activos antes y bajaron mucho o desaparecieron)
            sql_fugas = _text('''
                WITH ant AS (
                    SELECT
                        ovd.medico_observer AS med_id,
                        SUM(ovd.importe) AS importe_ant,
                        COUNT(DISTINCT ovd.id_operacion) AS recetas_ant
                    FROM obs_ventas_detalle ovd
                    WHERE ovd.fecha_estadistica BETWEEN :desde_ant AND :hasta_ant
                      AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                      AND ovd.medico_observer IS NOT NULL
                    GROUP BY ovd.medico_observer
                    HAVING COUNT(DISTINCT ovd.id_operacion) >= 5
                ),
                act AS (
                    SELECT
                        ovd.medico_observer AS med_id,
                        SUM(ovd.importe) AS importe_act,
                        COUNT(DISTINCT ovd.id_operacion) AS recetas_act
                    FROM obs_ventas_detalle ovd
                    WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                      AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                      AND ovd.medico_observer IS NOT NULL
                    GROUP BY ovd.medico_observer
                )
                SELECT
                    ant.med_id,
                    om.nombre AS medico_nombre,
                    ant.importe_ant,
                    ant.recetas_ant,
                    COALESCE(act.importe_act, 0) AS importe_act,
                    COALESCE(act.recetas_act, 0) AS recetas_act
                FROM ant
                LEFT JOIN act ON act.med_id = ant.med_id
                LEFT JOIN obs_medicos om ON om.observer_id = ant.med_id
                WHERE COALESCE(act.importe_act, 0) < ant.importe_ant * 0.5
                ORDER BY (ant.importe_ant - COALESCE(act.importe_act, 0)) DESC
                LIMIT 30
            ''')
            rows_fugas = session.execute(sql_fugas, {'desde': desde, 'hasta': hasta,
                                                     'desde_ant': desde_ant, 'hasta_ant': hasta_ant}).fetchall()
            fugas = []
            for r in rows_fugas:
                imp_a = float(r.importe_ant or 0)
                imp_n = float(r.importe_act or 0)
                caida_pct = ((imp_a - imp_n) / imp_a * 100) if imp_a > 0 else 0
                fugas.append({
                    'med_id': r.med_id,
                    'nombre': r.medico_nombre or f'Médico #{r.med_id}',
                    'importe_ant': round(imp_a, 2),
                    'importe_act': round(imp_n, 2),
                    'recetas_ant': int(r.recetas_ant or 0),
                    'recetas_act': int(r.recetas_act or 0),
                    'caida_pct': round(caida_pct, 1),
                    'estado': 'fuga_total' if r.recetas_act == 0 else 'fuga_parcial',
                })

            # Heatmap top 20 médicos × top 10 OS
            top_med_ids = [r['med_id'] for r in ranking[:20]]
            sql_top_os = _text('''
                SELECT
                    ovd.obra_social_observer AS os_id,
                    COALESCE(oos.descripcion, 'Sin OS') AS os_nombre,
                    SUM(ovd.importe) AS importe
                FROM obs_ventas_detalle ovd
                LEFT JOIN obs_obras_sociales oos ON oos.observer_id = ovd.obra_social_observer
                WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                  AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                  AND ovd.obra_social_observer IS NOT NULL
                GROUP BY ovd.obra_social_observer, oos.descripcion
                ORDER BY SUM(ovd.importe) DESC
                LIMIT 10
            ''')
            top_os = session.execute(sql_top_os, {'desde': desde, 'hasta': hasta}).fetchall()
            top_os_ids = [r.os_id for r in top_os]
            top_os_nombres = [{'os_id': r.os_id, 'nombre': r.os_nombre} for r in top_os]

            heatmap = []
            if top_med_ids and top_os_ids:
                from sqlalchemy import bindparam
                sql_heat = _text(f'''
                    SELECT
                        ovd.medico_observer AS med_id,
                        ovd.obra_social_observer AS os_id,
                        SUM(ovd.importe) AS importe
                    FROM obs_ventas_detalle ovd
                    WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                      AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                      AND ovd.medico_observer = ANY(:med_ids)
                      AND ovd.obra_social_observer = ANY(:os_ids)
                    GROUP BY ovd.medico_observer, ovd.obra_social_observer
                ''')
                rows_heat = session.execute(sql_heat, {
                    'desde': desde, 'hasta': hasta,
                    'med_ids': top_med_ids, 'os_ids': top_os_ids
                }).fetchall()
                for r in rows_heat:
                    heatmap.append({
                        'med_id': r.med_id,
                        'os_id': r.os_id,
                        'importe': float(r.importe or 0),
                    })

            # OS list para dropdown filtro
            sql_os_list = _text('''
                SELECT DISTINCT oos.observer_id, oos.descripcion
                FROM obs_ventas_detalle ovd
                JOIN obs_obras_sociales oos ON oos.observer_id = ovd.obra_social_observer
                WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                  AND oos.fecha_baja IS NULL
                ORDER BY oos.descripcion
            ''')
            obras_sociales = [{'os_id': r[0], 'nombre': r[1]}
                              for r in session.execute(sql_os_list, {'desde': desde, 'hasta': hasta}).fetchall()]

        return jsonify({
            'desde': desde.isoformat(),
            'hasta': hasta.isoformat(),
            'desde_ant': desde_ant.isoformat(),
            'hasta_ant': hasta_ant.isoformat(),
            'os_id': os_id,
            'kpis': {
                'n_medicos': len(ranking),
                'recetas': tot_recetas,
                'importe': round(tot_imp, 2),
                'top5_imp': round(top5_imp, 2),
                'top5_pct': round(top5_pct, 1),
                'n_fugas': len(fugas),
                'fugas_importe_perdido': round(sum(f['importe_ant'] - f['importe_act'] for f in fugas), 2),
            },
            'ranking': ranking,
            'fugas': fugas,
            'heatmap': {
                'medicos': [{'med_id': r['med_id'], 'nombre': r['nombre'], 'matricula': r['matricula']} for r in ranking[:20]],
                'obras_sociales': top_os_nombres,
                'celdas': heatmap,
            },
            'obras_sociales': obras_sociales,
        })

    @app.route('/obras-sociales/antiguedad')
    def os_antiguedad():
        """Aging report: cuánto tenés a cobrar de cada OS, agrupado por buckets
        de antigüedad. Sirve para detectar OS con deuda vieja sin liquidar.
        """
        return render_template('os_antiguedad.html')

    @app.route('/api/obras-sociales/antiguedad')
    def api_os_antiguedad():
        """JSON con aging por OS.

        Buckets: 0-30, 31-60, 61-90, 91-180, >180 días.
        Asume que importe_a_cargo_os representa deuda pendiente (no cobrada).
        Si en el futuro se suma una tabla de cobros/liquidaciones, restamos.
        """
        from datetime import date as _date
        from datetime import datetime as _dt
        from datetime import timedelta as _td

        from sqlalchemy import text as _text

        import database

        hoy = _dt.now().date()

        with database.get_db() as session:
            # Aging por OS — un solo query con buckets
            sql = _text('''
                SELECT
                    ovd.obra_social_observer AS os_id,
                    COALESCE(oos.descripcion, 'Sin OS') AS os_nombre,
                    SUM(COALESCE(ovd.importe_a_cargo_os, 0)) AS total_pendiente,
                    SUM(CASE WHEN (:hoy - ovd.fecha_estadistica) <= 30
                             THEN COALESCE(ovd.importe_a_cargo_os, 0) ELSE 0 END) AS b1_30,
                    SUM(CASE WHEN (:hoy - ovd.fecha_estadistica) > 30
                              AND (:hoy - ovd.fecha_estadistica) <= 60
                             THEN COALESCE(ovd.importe_a_cargo_os, 0) ELSE 0 END) AS b31_60,
                    SUM(CASE WHEN (:hoy - ovd.fecha_estadistica) > 60
                              AND (:hoy - ovd.fecha_estadistica) <= 90
                             THEN COALESCE(ovd.importe_a_cargo_os, 0) ELSE 0 END) AS b61_90,
                    SUM(CASE WHEN (:hoy - ovd.fecha_estadistica) > 90
                              AND (:hoy - ovd.fecha_estadistica) <= 180
                             THEN COALESCE(ovd.importe_a_cargo_os, 0) ELSE 0 END) AS b91_180,
                    SUM(CASE WHEN (:hoy - ovd.fecha_estadistica) > 180
                             THEN COALESCE(ovd.importe_a_cargo_os, 0) ELSE 0 END) AS b181_mas,
                    COUNT(*) AS dispensas,
                    COUNT(DISTINCT ovd.id_operacion) AS recetas,
                    MIN(ovd.fecha_estadistica) AS fecha_mas_vieja,
                    MAX(ovd.fecha_estadistica) AS fecha_mas_nueva
                FROM obs_ventas_detalle ovd
                LEFT JOIN obs_obras_sociales oos ON oos.observer_id = ovd.obra_social_observer
                WHERE COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                  AND ovd.obra_social_observer IS NOT NULL
                  AND COALESCE(ovd.importe_a_cargo_os, 0) > 0
                GROUP BY ovd.obra_social_observer, oos.descripcion
                HAVING SUM(COALESCE(ovd.importe_a_cargo_os, 0)) > 0
                ORDER BY SUM(COALESCE(ovd.importe_a_cargo_os, 0)) DESC
            ''')
            rows = session.execute(sql, {'hoy': hoy}).fetchall()

            ranking = []
            tot = {'pendiente': 0.0, 'b1_30': 0.0, 'b31_60': 0.0, 'b61_90': 0.0,
                   'b91_180': 0.0, 'b181_mas': 0.0}
            tot_dispensas = tot_recetas = 0
            for r in rows:
                b1 = float(r.b1_30 or 0)
                b2 = float(r.b31_60 or 0)
                b3 = float(r.b61_90 or 0)
                b4 = float(r.b91_180 or 0)
                b5 = float(r.b181_mas or 0)
                pendiente = float(r.total_pendiente or 0)
                viejo_pct = ((b3 + b4 + b5) / pendiente * 100) if pendiente > 0 else 0
                ranking.append({
                    'os_id': r.os_id,
                    'os_nombre': r.os_nombre,
                    'total_pendiente': round(pendiente, 2),
                    'b1_30': round(b1, 2),
                    'b31_60': round(b2, 2),
                    'b61_90': round(b3, 2),
                    'b91_180': round(b4, 2),
                    'b181_mas': round(b5, 2),
                    'viejo_pct': round(viejo_pct, 1),  # % >60 días
                    'dispensas': int(r.dispensas or 0),
                    'recetas': int(r.recetas or 0),
                    'fecha_mas_vieja': r.fecha_mas_vieja.isoformat() if r.fecha_mas_vieja else None,
                    'fecha_mas_nueva': r.fecha_mas_nueva.isoformat() if r.fecha_mas_nueva else None,
                })
                tot['pendiente'] += pendiente
                tot['b1_30'] += b1
                tot['b31_60'] += b2
                tot['b61_90'] += b3
                tot['b91_180'] += b4
                tot['b181_mas'] += b5
                tot_dispensas += int(r.dispensas or 0)
                tot_recetas += int(r.recetas or 0)

            # Alertas
            alertas = []
            for r in ranking:
                viejo = r['b91_180'] + r['b181_mas']
                if r['b181_mas'] > 0 and r['b181_mas'] / r['total_pendiente'] > 0.2:
                    alertas.append({
                        'tipo': 'critico',
                        'os_nombre': r['os_nombre'],
                        'mensaje': f"${r['b181_mas']:,.0f} con MÁS DE 180 DÍAS sin liquidar ({r['b181_mas']/r['total_pendiente']*100:.0f}% del total). Reclamar urgente.",
                    })
                elif viejo > 0 and viejo / r['total_pendiente'] > 0.4:
                    alertas.append({
                        'tipo': 'advertencia',
                        'os_nombre': r['os_nombre'],
                        'mensaje': f"${viejo:,.0f} con más de 90 días pendiente ({viejo/r['total_pendiente']*100:.0f}% del total).",
                    })

        return jsonify({
            'hoy': hoy.isoformat(),
            'kpis': {
                'total_pendiente': round(tot['pendiente'], 2),
                'b1_30': round(tot['b1_30'], 2),
                'b31_60': round(tot['b31_60'], 2),
                'b61_90': round(tot['b61_90'], 2),
                'b91_180': round(tot['b91_180'], 2),
                'b181_mas': round(tot['b181_mas'], 2),
                'pct_b1_30': round(tot['b1_30'] / tot['pendiente'] * 100, 1) if tot['pendiente'] > 0 else 0,
                'pct_viejo': round((tot['b61_90'] + tot['b91_180'] + tot['b181_mas']) / tot['pendiente'] * 100, 1) if tot['pendiente'] > 0 else 0,
                'n_os': len(ranking),
                'dispensas': tot_dispensas,
                'recetas': tot_recetas,
            },
            'ranking': ranking,
            'alertas': alertas,
        })

    @app.route('/obras-sociales/productos-rentabilidad')
    def os_productos_rentabilidad():
        """Ranking de productos por margen, cruzado con OS asociadas.

        Útil para: ver qué productos generan más ganancia neta y con qué
        OS están asociados. Detectar productos con margen negativo.
        """
        from datetime import date as _date
        from datetime import datetime as _dt

        hoy = _dt.now().date()
        try:
            desde = _date.fromisoformat(request.args.get('desde', '')) or hoy.replace(day=1)
        except ValueError:
            desde = hoy.replace(day=1)
        try:
            hasta = _date.fromisoformat(request.args.get('hasta', '')) or hoy
        except ValueError:
            hasta = hoy

        return render_template('os_productos_rentabilidad.html',
                               desde=desde.isoformat(), hasta=hasta.isoformat())

    @app.route('/api/obras-sociales/productos-rentabilidad')
    def api_os_productos_rentabilidad():
        """JSON: ranking de productos por margen + KPIs.

        Filtros: desde, hasta, os_id (opcional), limit (top N).
        """
        from datetime import date as _date
        from datetime import datetime as _dt

        from sqlalchemy import text as _text

        import database

        hoy = _dt.now().date()
        try:
            desde = _date.fromisoformat(request.args.get('desde', '')) or hoy.replace(day=1)
        except ValueError:
            desde = hoy.replace(day=1)
        try:
            hasta = _date.fromisoformat(request.args.get('hasta', '')) or hoy
        except ValueError:
            hasta = hoy
        try:
            limit = max(10, min(int(request.args.get('limit', '100')), 500))
        except ValueError:
            limit = 100
        try:
            os_id = int(request.args.get('os_id', '')) or None
        except ValueError:
            os_id = None

        with database.get_db() as session:
            params = {'desde': desde, 'hasta': hasta, 'limit': limit}
            os_filter_sql = ''
            if os_id is not None:
                os_filter_sql = 'AND ovd.obra_social_observer = :os_id'
                params['os_id'] = os_id

            sql = _text(f'''
                WITH ultima_compra AS (
                    SELECT codigo_barra, precio_unitario,
                           ROW_NUMBER() OVER (PARTITION BY codigo_barra ORDER BY id DESC) AS rn
                    FROM factura_items
                    WHERE precio_unitario IS NOT NULL AND precio_unitario > 0
                      AND codigo_barra IS NOT NULL
                ),
                costo_por_observer AS (
                    SELECT p.observer_id, MAX(uc.precio_unitario) AS costo_unitario
                    FROM ultima_compra uc
                    JOIN producto_codigos_barra pcb ON pcb.codigo_barra = uc.codigo_barra
                    JOIN productos p ON p.id = pcb.producto_id
                    WHERE uc.rn = 1 AND p.observer_id IS NOT NULL
                    GROUP BY p.observer_id
                )
                SELECT
                    op.observer_id AS prod_observer,
                    op.descripcion AS prod_nombre,
                    cpo.costo_unitario AS costo_unitario,
                    SUM(ovd.cantidad) AS unidades,
                    SUM(ovd.importe) AS facturado,
                    SUM(COALESCE(ovd.importe_a_cargo_os, 0)) AS a_cargo_os,
                    SUM(CASE WHEN cpo.costo_unitario IS NOT NULL
                             THEN ovd.cantidad * cpo.costo_unitario ELSE 0 END) AS costo_estimado,
                    SUM(CASE WHEN cpo.costo_unitario IS NOT NULL
                             THEN ovd.importe ELSE 0 END) AS facturado_con_costo,
                    COUNT(*) AS dispensas,
                    COUNT(DISTINCT ovd.id_operacion) AS recetas,
                    COUNT(DISTINCT ovd.obra_social_observer) AS n_os
                FROM obs_ventas_detalle ovd
                JOIN obs_productos op ON op.observer_id = ovd.producto_observer
                LEFT JOIN costo_por_observer cpo ON cpo.observer_id = ovd.producto_observer
                WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                  AND COALESCE(ovd.es_venta_particular, FALSE) = FALSE
                  AND ovd.obra_social_observer IS NOT NULL
                  {os_filter_sql}
                GROUP BY op.observer_id, op.descripcion, cpo.costo_unitario
                HAVING SUM(ovd.importe) > 0
                ORDER BY (SUM(CASE WHEN cpo.costo_unitario IS NOT NULL THEN ovd.importe ELSE 0 END)
                          - SUM(CASE WHEN cpo.costo_unitario IS NOT NULL THEN ovd.cantidad * cpo.costo_unitario ELSE 0 END)) DESC
                LIMIT :limit
            ''')
            rows = session.execute(sql, params).fetchall()

            ranking = []
            tot_fact = tot_costo = tot_facturado_con_costo = tot_a_cargo = 0.0
            tot_unidades = tot_dispensas = 0
            n_negativos = n_sin_costo = 0
            for r in rows:
                fact = float(r.facturado or 0)
                costo = float(r.costo_estimado or 0)
                fact_cc = float(r.facturado_con_costo or 0)
                a_cargo = float(r.a_cargo_os or 0)
                ganancia = fact_cc - costo
                margen_pct = (ganancia / fact_cc * 100) if fact_cc > 0 else None
                tiene_costo = r.costo_unitario is not None
                if margen_pct is not None and margen_pct < 0:
                    n_negativos += 1
                if not tiene_costo:
                    n_sin_costo += 1
                ranking.append({
                    'prod_observer': r.prod_observer,
                    'descripcion': r.prod_nombre,
                    'costo_unitario': float(r.costo_unitario) if r.costo_unitario else None,
                    'unidades': float(r.unidades or 0),
                    'facturado': round(fact, 2),
                    'a_cargo_os': round(a_cargo, 2),
                    'costo_estimado': round(costo, 2),
                    'ganancia': round(ganancia, 2),
                    'margen_pct': round(margen_pct, 2) if margen_pct is not None else None,
                    'dispensas': int(r.dispensas or 0),
                    'recetas': int(r.recetas or 0),
                    'n_os': int(r.n_os or 0),
                    'tiene_costo': tiene_costo,
                })
                tot_fact += fact
                tot_costo += costo
                tot_facturado_con_costo += fact_cc
                tot_a_cargo += a_cargo
                tot_unidades += float(r.unidades or 0)
                tot_dispensas += int(r.dispensas or 0)

            tot_ganancia = tot_facturado_con_costo - tot_costo
            tot_margen_pct = (tot_ganancia / tot_facturado_con_costo * 100) if tot_facturado_con_costo > 0 else None

            # Lista de OS para el dropdown filtro
            sql_os = _text('''
                SELECT DISTINCT oos.observer_id, oos.descripcion
                FROM obs_ventas_detalle ovd
                JOIN obs_obras_sociales oos ON oos.observer_id = ovd.obra_social_observer
                WHERE ovd.fecha_estadistica BETWEEN :desde AND :hasta
                  AND oos.fecha_baja IS NULL
                ORDER BY oos.descripcion
            ''')
            obras_sociales = [
                {'os_id': r[0], 'nombre': r[1]}
                for r in session.execute(sql_os, {'desde': desde, 'hasta': hasta}).fetchall()
            ]

        return jsonify({
            'desde': desde.isoformat(),
            'hasta': hasta.isoformat(),
            'os_id': os_id,
            'kpis': {
                'productos_en_top': len(ranking),
                'facturado_total': round(tot_fact, 2),
                'costo_estimado': round(tot_costo, 2),
                'ganancia_bruta': round(tot_ganancia, 2),
                'margen_pct': round(tot_margen_pct, 2) if tot_margen_pct is not None else None,
                'a_cargo_os': round(tot_a_cargo, 2),
                'unidades_total': int(tot_unidades),
                'dispensas_total': tot_dispensas,
                'productos_negativos': n_negativos,
                'productos_sin_costo': n_sin_costo,
            },
            'ranking': ranking,
            'obras_sociales': obras_sociales,
        })

    @app.route('/obras-sociales/dashboard')
    def os_dashboard():
        """Dashboard de OS — datos reales desde obs_ventas_detalle.

        Usa categorías 'PAMI'/'IAPOS'/'OTROS' detectando por la descripción de
        la OS. Una "receta" = grupo de líneas con el mismo (id_operacion,
        cliente, medico) en el mismo día (DW.Recetas no expone esto, lo
        reconstruimos).
        """
        from sqlalchemy import case, distinct
        from sqlalchemy import func as _f

        import database

        hoy = datetime.now().date()
        # Rango configurable por querystring. Default: mes corriente.
        desde_str = (request.args.get('desde') or '').strip()
        hasta_str = (request.args.get('hasta') or '').strip()
        try:
            desde = date.fromisoformat(desde_str) if desde_str else hoy.replace(day=1)
        except ValueError:
            desde = hoy.replace(day=1)
        try:
            hasta = date.fromisoformat(hasta_str) if hasta_str else hoy
        except ValueError:
            hasta = hoy
        if hasta < desde:
            desde, hasta = hasta, desde
        primer_dia = desde   # se sigue llamando así abajo, refactor mínimo
        # El chart muestra los últimos 30 días HASTA "hasta" (no a partir de hoy fijo).
        desde_30d = hasta - timedelta(days=29)

        os_filter = (request.args.get('os') or '').upper()
        if os_filter not in ('PAMI', 'IAPOS', 'OTROS'):
            os_filter = ''
        # Filtro adicional por una OS individual (id de obs_obras_sociales).
        try:
            os_id_filter = int(request.args.get('os_id') or 0) or None
        except (ValueError, TypeError):
            os_id_filter = None
        # Métrica del chart: 'importe' (default) o 'cantidad' (recetas).
        chart_metric = request.args.get('metric') or 'importe'
        if chart_metric not in ('importe', 'cantidad'):
            chart_metric = 'importe'

        with database.get_db() as session:
            ObsVD = database.ObsVentaDetalle

            # Mapear cada OS observada a categoría (PAMI/IAPOS/OTROS).
            os_obs_ids = [r[0] for r in session.query(distinct(ObsVD.obra_social_observer))
                          .filter(ObsVD.obra_social_observer.isnot(None)).all()]
            os_descrs = dict(session.query(database.ObsObraSocial.observer_id,
                                            database.ObsObraSocial.descripcion).all())

            def categoria(os_id):
                if os_id is None:
                    return None
                desc = (os_descrs.get(os_id) or '').upper()
                if 'PAMI' in desc:
                    return 'PAMI'
                if 'IAPOS' in desc:
                    return 'IAPOS'
                return 'OTROS'

            os_to_cat = {oid: categoria(oid) for oid in os_obs_ids}
            pami_ids = {oid for oid, c in os_to_cat.items() if c == 'PAMI'}
            iapos_ids = {oid for oid, c in os_to_cat.items() if c == 'IAPOS'}

            # CASE expr para SQL
            cat_expr = case(
                (ObsVD.obra_social_observer.in_(list(pami_ids) or [-1]), 'PAMI'),
                (ObsVD.obra_social_observer.in_(list(iapos_ids) or [-1]), 'IAPOS'),
                else_='OTROS',
            )

            # Filtro base: solo ventas con OS (no particulares), dentro del rango desde-hasta.
            base_filters = [
                ObsVD.obra_social_observer.isnot(None),
                ObsVD.fecha_estadistica >= desde,
                ObsVD.fecha_estadistica <= hasta,
            ]
            otros_ids_full = {oid for oid, c in os_to_cat.items() if c == 'OTROS'}
            if os_id_filter:
                base_filters.append(ObsVD.obra_social_observer == os_id_filter)
            elif os_filter == 'PAMI':
                base_filters.append(ObsVD.obra_social_observer.in_(list(pami_ids) or [-1]))
            elif os_filter == 'IAPOS':
                base_filters.append(ObsVD.obra_social_observer.in_(list(iapos_ids) or [-1]))
            elif os_filter == 'OTROS':
                base_filters.append(ObsVD.obra_social_observer.in_(list(otros_ids_full) or [-1]))

            # Card "del mes corriente": agrupar por receta (id_operacion + medico + cliente + fecha).
            # Para contar recetas únicas y sumar importes.
            recetas_q = (session.query(
                            _f.count(distinct(_f.concat(
                                ObsVD.id_operacion, '|',
                                _f.coalesce(ObsVD.cliente_observer, 0), '|',
                                _f.coalesce(ObsVD.medico_observer, 0), '|',
                                ObsVD.fecha_estadistica,
                            ))).label('recetas'),
                            _f.coalesce(_f.sum(ObsVD.importe_a_cargo_os), 0).label('importe_os'),
                            _f.coalesce(_f.sum(_f.coalesce(ObsVD.importe, 0)
                                                - _f.coalesce(ObsVD.importe_a_cargo_os, 0)), 0).label('importe_pac'),
                         )
                         .filter(*base_filters))
            row = recetas_q.first()
            mes_recetas = int(row[0] or 0) if row else 0
            mes_os_total = float(row[1] or 0) if row else 0
            mes_pac_total = float(row[2] or 0) if row else 0
            mes_ticket_prom = ((mes_os_total + mes_pac_total) / mes_recetas) if mes_recetas else 0

            # Breakdown por categoría: PAMI, IAPOS y OTROS.
            # Si filtraste por una OS individual, NO mostramos los breakdowns
            # generales (modo zoom: ves solo lo que filtraste).
            otros_ids = {oid for oid, c in os_to_cat.items() if c == 'OTROS'}
            por_os = {}
            if os_id_filter:
                # En modo zoom no se calculan los cards de categoría.
                pass
            else:
                for cat_code, ids in (('PAMI', pami_ids), ('IAPOS', iapos_ids), ('OTROS', otros_ids)):
                    if not ids:
                        por_os[cat_code] = {'recetas': 0, 'importe_os': 0, 'importe_paciente': 0}
                        continue
                    r = (session.query(
                            _f.count(distinct(_f.concat(
                                ObsVD.id_operacion, '|',
                                _f.coalesce(ObsVD.cliente_observer, 0), '|',
                                _f.coalesce(ObsVD.medico_observer, 0), '|',
                                ObsVD.fecha_estadistica,
                            ))),
                            _f.coalesce(_f.sum(ObsVD.importe_a_cargo_os), 0),
                            _f.coalesce(_f.sum(_f.coalesce(ObsVD.importe, 0)
                                                - _f.coalesce(ObsVD.importe_a_cargo_os, 0)), 0),
                         )
                         .filter(ObsVD.obra_social_observer.in_(list(ids)),
                                 ObsVD.fecha_estadistica >= desde,
                                 ObsVD.fecha_estadistica <= hasta)
                         .first())
                    por_os[cat_code] = {
                        'recetas': int(r[0] or 0),
                        'importe_os': float(r[1] or 0),
                        'importe_paciente': float(r[2] or 0),
                    }

            # Distribución por cobertura — porcentaje a cargo OS por receta.
            # Buckets: 0%, 1-39%, 40-69%, 70-99%, 100%.
            # Calculado a nivel receta (group by id_operacion + cliente + medico + fecha).
            cob_buckets_q = (session.query(
                                _f.coalesce(_f.sum(ObsVD.importe_a_cargo_os), 0).label('imp_os'),
                                _f.coalesce(_f.sum(ObsVD.importe), 0).label('imp_total'),
                            )
                            .filter(*base_filters)
                            .group_by(ObsVD.id_operacion,
                                      ObsVD.cliente_observer,
                                      ObsVD.medico_observer,
                                      ObsVD.fecha_estadistica)).all()
            cob_dist = {
                '0%':       {'recetas': 0, 'importe_os': 0.0, 'importe_total': 0.0},
                '1-39%':    {'recetas': 0, 'importe_os': 0.0, 'importe_total': 0.0},
                '40-69%':   {'recetas': 0, 'importe_os': 0.0, 'importe_total': 0.0},
                '70-99%':   {'recetas': 0, 'importe_os': 0.0, 'importe_total': 0.0},
                '100%':     {'recetas': 0, 'importe_os': 0.0, 'importe_total': 0.0},
            }
            for imp_os, imp_total in cob_buckets_q:
                imp_os = float(imp_os or 0)
                imp_total = float(imp_total or 0)
                if imp_total <= 0:
                    pct = 0.0
                else:
                    pct = imp_os / imp_total * 100
                if pct <= 0:        bucket = '0%'
                elif pct < 40:      bucket = '1-39%'
                elif pct < 70:      bucket = '40-69%'
                elif pct < 100:     bucket = '70-99%'
                else:               bucket = '100%'
                cob_dist[bucket]['recetas']        += 1
                cob_dist[bucket]['importe_os']     += imp_os
                cob_dist[bucket]['importe_total']  += imp_total
            total_recetas_cob = sum(b['recetas'] for b in cob_dist.values())
            cobertura_buckets = [
                {'rango': k, 'recetas': v['recetas'],
                 'importe_os': round(v['importe_os'], 2),
                 'importe_total': round(v['importe_total'], 2),
                 'pct_recetas': round((v['recetas'] / total_recetas_cob * 100), 1) if total_recetas_cob else 0}
                for k, v in cob_dist.items()
            ]

            # Top 10 OS individuales del mes — solo cuando NO hay filtro individual.
            top_os = []
            if not os_id_filter:
                top_os_q = (session.query(
                                ObsVD.obra_social_observer,
                                _f.count(distinct(_f.concat(
                                    ObsVD.id_operacion, '|',
                                    _f.coalesce(ObsVD.cliente_observer, 0), '|',
                                    _f.coalesce(ObsVD.medico_observer, 0), '|',
                                    ObsVD.fecha_estadistica,
                                ))).label('recetas'),
                                _f.coalesce(_f.sum(ObsVD.importe_a_cargo_os), 0).label('importe_os'),
                                _f.coalesce(_f.sum(_f.coalesce(ObsVD.importe, 0)
                                                    - _f.coalesce(ObsVD.importe_a_cargo_os, 0)), 0).label('importe_pac'),
                            )
                            .filter(ObsVD.obra_social_observer.isnot(None),
                                    ObsVD.fecha_estadistica >= desde,
                                 ObsVD.fecha_estadistica <= hasta)
                            .group_by(ObsVD.obra_social_observer)
                            .order_by(_f.sum(ObsVD.importe_a_cargo_os).desc())
                            .limit(10)).all()
                top_os = [{
                    'os_id': oid,
                    'nombre': os_descrs.get(oid) or f'OS #{oid}',
                    'categoria': os_to_cat.get(oid) or 'OTROS',
                    'recetas': int(rec or 0),
                    'importe_os': float(imp_os or 0),
                    'importe_paciente': float(imp_pac or 0),
                } for (oid, rec, imp_os, imp_pac) in top_os_q]

            # Serie diaria últimos 30 días.
            # Modo normal: 3 series (PAMI/IAPOS/OTROS). Modo zoom: 1 sola serie con la OS filtrada.
            # Métrica: 'importe' suma importe_a_cargo_os, 'cantidad' cuenta recetas únicas.
            if chart_metric == 'cantidad':
                metric_expr = _f.count(distinct(_f.concat(
                    ObsVD.id_operacion, '|',
                    _f.coalesce(ObsVD.cliente_observer, 0), '|',
                    _f.coalesce(ObsVD.medico_observer, 0), '|',
                    ObsVD.fecha_estadistica,
                )))
            else:
                metric_expr = _f.sum(ObsVD.importe_a_cargo_os)
            serie_dict = {}
            if os_id_filter:
                for i in range(30):
                    d = hoy - timedelta(days=29 - i)
                    serie_dict[d.isoformat()] = {'OS': 0.0}
                serie_q = (session.query(
                                ObsVD.fecha_estadistica,
                                metric_expr,
                           )
                           .filter(ObsVD.obra_social_observer == os_id_filter,
                                   ObsVD.fecha_estadistica >= desde_30d,
                                   ObsVD.fecha_estadistica <= hasta)
                           .group_by(ObsVD.fecha_estadistica))
                for fecha, val in serie_q.all():
                    if not fecha:
                        continue
                    key = fecha.isoformat() if hasattr(fecha, 'isoformat') else str(fecha)
                    if key in serie_dict:
                        serie_dict[key]['OS'] = float(val or 0)
                serie_list = [{'fecha': k, 'OS': round(v['OS'], 2)}
                              for k, v in serie_dict.items()]
            else:
                for i in range(30):
                    d = hoy - timedelta(days=29 - i)
                    serie_dict[d.isoformat()] = {'PAMI': 0.0, 'IAPOS': 0.0, 'OTROS': 0.0}
                serie_q = (session.query(
                                ObsVD.fecha_estadistica,
                                cat_expr.label('cat'),
                                metric_expr,
                           )
                           .filter(ObsVD.obra_social_observer.isnot(None),
                                   ObsVD.fecha_estadistica >= desde_30d,
                                   ObsVD.fecha_estadistica <= hasta)
                           .group_by(ObsVD.fecha_estadistica, cat_expr))
                for fecha, cat, val in serie_q.all():
                    if not fecha:
                        continue
                    key = fecha.isoformat() if hasattr(fecha, 'isoformat') else str(fecha)
                    if key in serie_dict and cat in ('PAMI', 'IAPOS', 'OTROS'):
                        serie_dict[key][cat] += float(val or 0)
                serie_list = [{'fecha': k,
                               'PAMI': round(v['PAMI'], 2),
                               'IAPOS': round(v['IAPOS'], 2),
                               'OTROS': round(v['OTROS'], 2)}
                              for k, v in serie_dict.items()]

            # Top médicos del mes (filtrado por OS si aplica).
            med_q = (session.query(
                        ObsVD.medico_observer,
                        _f.count(distinct(ObsVD.id_operacion)).label('recetas'),
                        _f.coalesce(_f.sum(ObsVD.importe_a_cargo_os), 0).label('importe'),
                     )
                     .filter(*base_filters)
                     .filter(ObsVD.medico_observer.isnot(None))
                     .group_by(ObsVD.medico_observer)
                     .order_by(_f.sum(ObsVD.importe_a_cargo_os).desc())
                     .limit(10)).all()
            med_ids = [r[0] for r in med_q if r[0]]
            med_nombres = dict(session.query(database.ObsMedico.observer_id,
                                              database.ObsMedico.nombre)
                               .filter(database.ObsMedico.observer_id.in_(med_ids)).all()) if med_ids else {}
            top_medicos = [{
                'matricula': str(m_id),
                'nombre': med_nombres.get(m_id, f'Médico #{m_id}'),
                'recetas': int(rec or 0),
                'importe': float(imp or 0),
            } for (m_id, rec, imp) in med_q]

            # Top productos del mes (filtrado por OS si aplica).
            prod_q = (session.query(
                        ObsVD.producto_observer,
                        _f.coalesce(_f.sum(ObsVD.cantidad), 0).label('unidades'),
                        _f.coalesce(_f.sum(ObsVD.importe_a_cargo_os), 0).label('importe'),
                      )
                      .filter(*base_filters)
                      .group_by(ObsVD.producto_observer)
                      .order_by(_f.sum(ObsVD.importe_a_cargo_os).desc())
                      .limit(10)).all()
            prod_ids = [r[0] for r in prod_q]
            prod_descrs = dict(session.query(database.ObsProducto.observer_id,
                                              database.ObsProducto.descripcion)
                               .filter(database.ObsProducto.observer_id.in_(prod_ids)).all()) if prod_ids else {}
            top_productos = [{
                'descripcion': prod_descrs.get(pid, f'Producto #{pid}'),
                'unidades': int(uds or 0),
                'importe': float(imp or 0),
            } for (pid, uds, imp) in prod_q]

            # "Lotes abiertos": no tenemos lote real en DW. Como aproximación:
            # recetas del mes que aún no fueron presentadas (asumimos todas
            # están "abiertas" hasta que tengamos el dato de lote).
            lotes = []
            for cat_code in ('PAMI', 'IAPOS'):
                info = por_os.get(cat_code, {})
                if info.get('recetas', 0) > 0:
                    lotes.append({
                        'os_codigo': cat_code,
                        'cantidad_recetas': info['recetas'],
                        'importe_os': info['importe_os'],
                        'fecha_apertura': primer_dia,
                    })

        # Lista de OS para el dropdown de filtro individual (solo las que tienen
        # actividad en el período visible — evita 879 opciones inertes).
        os_options = sorted(
            [{'os_id': oid, 'nombre': os_descrs.get(oid) or f'OS #{oid}'}
             for oid in os_obs_ids],
            key=lambda x: x['nombre'].lower()
        )
        os_id_filter_nombre = os_descrs.get(os_id_filter) if os_id_filter else None

        return render_template('os_dashboard.html',
                               os_filter=os_filter,
                               os_id_filter=os_id_filter,
                               os_id_filter_nombre=os_id_filter_nombre,
                               os_options=os_options,
                               chart_metric=chart_metric,
                               desde=desde.isoformat(),
                               hasta=hasta.isoformat(),
                               mes_recetas=mes_recetas,
                               mes_os_total=mes_os_total,
                               mes_pac_total=mes_pac_total,
                               mes_ticket_prom=mes_ticket_prom,
                               por_os=por_os,
                               serie=serie_list,
                               top_medicos=top_medicos,
                               top_productos=top_productos,
                               top_os=top_os,
                               cobertura_buckets=cobertura_buckets,
                               lotes=lotes)

    def _query_dispensas(filtros, session, id_farmacia):
        """Helper reusable: dado los filtros parseados, devuelve
        (dispensas, total_recetas, total_os, total_pac, os_options).

        Lo usan tanto /obras-sociales/dispensas (HTML) como
        /obras-sociales/dispensas/export.xlsx (Excel).

        filtros: dict con keys os_filter_id, desde, hasta, tipo_filter, q.
        """
        from sqlalchemy import case, func

        from database import (
            ObsCliente, ObsConvenio, ObsObraSocial, ObsPlan, ObsProducto, ObsVentaDetalle,
        )

        os_filter_id = filtros.get('os_filter_id')
        med_filter_id = filtros.get('med_filter_id')
        prod_filter_id = filtros.get('prod_filter_id')
        plan_filter_id = filtros.get('plan_filter_id')
        cob_filter = filtros.get('cob_filter') or ''
        importe_min = filtros.get('importe_min')
        importe_max = filtros.get('importe_max')
        desde = filtros['desde']
        hasta = filtros['hasta']
        tipo_filter = filtros.get('tipo_filter') or ''
        q = (filtros.get('q') or '').lower()

        particular_agg = func.max(
            case((ObsVentaDetalle.es_venta_particular.is_(True), 1), else_=0)
        )

        recetas_q = (session.query(
            ObsVentaDetalle.cliente_observer.label('cli_id'),
            ObsVentaDetalle.medico_observer.label('med_id'),
            ObsVentaDetalle.id_operacion.label('op_id'),
            func.min(ObsVentaDetalle.fecha_operacion).label('fecha'),
            func.min(ObsVentaDetalle.obra_social_observer).label('os_id'),
            func.min(ObsVentaDetalle.plan_principal_observer).label('plan_id'),
            particular_agg.label('particular'),
            func.sum(ObsVentaDetalle.importe).label('total'),
            func.sum(ObsVentaDetalle.importe_a_cargo_os).label('total_os'),
            func.count(ObsVentaDetalle.id_producto_vendido).label('n_items'),
        )
            .filter(
                ObsVentaDetalle.id_farmacia == id_farmacia,
                ObsVentaDetalle.fecha_estadistica >= desde,
                ObsVentaDetalle.fecha_estadistica <= hasta,
            )
            .group_by(
                ObsVentaDetalle.cliente_observer,
                ObsVentaDetalle.medico_observer,
                ObsVentaDetalle.id_operacion,
            )
            .order_by(func.min(ObsVentaDetalle.fecha_operacion).desc()))

        if os_filter_id:
            recetas_q = recetas_q.having(
                func.min(ObsVentaDetalle.obra_social_observer) == os_filter_id)
        if med_filter_id:
            recetas_q = recetas_q.filter(ObsVentaDetalle.medico_observer == med_filter_id)
        if prod_filter_id:
            # Solo recetas que tengan AL MENOS UN item con ese producto.
            sub_op = (session.query(ObsVentaDetalle.id_operacion)
                      .filter(ObsVentaDetalle.id_farmacia == id_farmacia,
                              ObsVentaDetalle.fecha_estadistica >= desde,
                              ObsVentaDetalle.fecha_estadistica <= hasta,
                              ObsVentaDetalle.producto_observer == prod_filter_id)
                      .distinct().subquery())
            recetas_q = recetas_q.filter(ObsVentaDetalle.id_operacion.in_(sub_op))
        if plan_filter_id:
            recetas_q = recetas_q.having(
                func.min(ObsVentaDetalle.plan_principal_observer) == plan_filter_id)
        if cob_filter:
            # Cobertura % por receta = sum(importe_a_cargo_os) / sum(importe).
            # Aplicamos como HAVING sobre el grupo de líneas de la receta.
            sum_os = func.sum(ObsVentaDetalle.importe_a_cargo_os)
            sum_total = func.sum(ObsVentaDetalle.importe)
            # En SQL no podemos dividir si total=0; protegemos con CASE.
            pct_expr = case(
                (sum_total > 0, sum_os * 100.0 / sum_total),
                else_=0.0,
            )
            if cob_filter == '0%':
                recetas_q = recetas_q.having(pct_expr <= 0)
            elif cob_filter == '1-39%':
                recetas_q = recetas_q.having(pct_expr > 0).having(pct_expr < 40)
            elif cob_filter == '40-69%':
                recetas_q = recetas_q.having(pct_expr >= 40).having(pct_expr < 70)
            elif cob_filter == '70-99%':
                recetas_q = recetas_q.having(pct_expr >= 70).having(pct_expr < 100)
            elif cob_filter == '100%':
                recetas_q = recetas_q.having(pct_expr >= 100)
        if tipo_filter == 'particular':
            recetas_q = recetas_q.having(particular_agg == 1)
        elif tipo_filter == 'os':
            recetas_q = recetas_q.having(particular_agg == 0)
        # Filtro por importe total de la receta (rango).
        if importe_min is not None:
            recetas_q = recetas_q.having(func.sum(ObsVentaDetalle.importe) >= importe_min)
        if importe_max is not None:
            recetas_q = recetas_q.having(func.sum(ObsVentaDetalle.importe) <= importe_max)

        recetas_full = recetas_q.limit(500).all()

        cli_ids = {r.cli_id for r in recetas_full if r.cli_id}
        os_ids = {r.os_id for r in recetas_full if r.os_id}
        plan_ids = {r.plan_id for r in recetas_full if r.plan_id}
        med_ids = {r.med_id for r in recetas_full if r.med_id}
        op_ids = [r.op_id for r in recetas_full if r.op_id]

        cli_map, os_map, plan_map, med_map = {}, {}, {}, {}
        if cli_ids:
            for c in session.query(ObsCliente).filter(
                    ObsCliente.observer_id.in_(list(cli_ids))).all():
                cli_map[c.observer_id] = c
        if os_ids:
            for o in session.query(ObsObraSocial).filter(
                    ObsObraSocial.observer_id.in_(list(os_ids))).all():
                os_map[o.observer_id] = o
        if plan_ids:
            for p in session.query(ObsPlan).filter(
                    ObsPlan.observer_id.in_(list(plan_ids))).all():
                plan_map[p.observer_id] = p
        if med_ids:
            from database import ObsMedico
            for m in session.query(ObsMedico).filter(
                    ObsMedico.observer_id.in_(list(med_ids))).all():
                med_map[m.observer_id] = m

        items_por_receta = {}
        if op_ids:
            items_q = (session.query(
                            ObsVentaDetalle.cliente_observer,
                            ObsVentaDetalle.medico_observer,
                            ObsVentaDetalle.id_operacion,
                            ObsVentaDetalle.cantidad,
                            ObsVentaDetalle.importe,
                            ObsVentaDetalle.importe_a_cargo_os,
                            ObsProducto.descripcion,
                            ObsProducto.observer_id.label('prod_id'),
                        )
                        .outerjoin(ObsProducto,
                                   ObsProducto.observer_id == ObsVentaDetalle.producto_observer)
                        .filter(
                            ObsVentaDetalle.id_farmacia == id_farmacia,
                            ObsVentaDetalle.id_operacion.in_(op_ids[:500]),
                            ObsVentaDetalle.fecha_estadistica >= desde,
                            ObsVentaDetalle.fecha_estadistica <= hasta,
                        ))
            for row in items_q.all():
                key = (row.cliente_observer, row.medico_observer, row.id_operacion)
                items_por_receta.setdefault(key, []).append({
                    'descripcion': row.descripcion or '(sin descripción)',
                    'producto_id': row.prod_id,
                    'cantidad': float(row.cantidad or 0),
                    'importe': float(row.importe or 0),
                    'importe_os': float(row.importe_a_cargo_os or 0),
                    'cobertura_pct': (
                        round(float(row.importe_a_cargo_os or 0) /
                              float(row.importe) * 100, 0)
                        if row.importe and float(row.importe) > 0 else 0
                    ),
                })

        dispensas = []
        for r in recetas_full:
            cli = cli_map.get(r.cli_id)
            os_obj = os_map.get(r.os_id)
            plan_obj = plan_map.get(r.plan_id)
            afiliado_nombre = cli.apellido_nombre if cli else ''
            afiliado_dni = (
                f'{cli.documento_tipo or "DNI"} {cli.documento_numero}'
                if cli and cli.documento_numero else ''
            )
            os_nombre = os_obj.descripcion if os_obj else '—'
            plan_nombre = plan_obj.descripcion if plan_obj else ''
            particular = bool(r.particular)
            items = items_por_receta.get((r.cli_id, r.med_id, r.op_id), [])

            if q:
                hay = (
                    f'{afiliado_nombre} {afiliado_dni} {os_nombre} '
                    f'{plan_nombre} {r.op_id or ""}'
                ).lower()
                if q not in hay:
                    continue

            dispensas.append({
                'fecha_venta': r.fecha,
                'os_codigo': '—' if particular else (os_nombre[:15] if os_nombre else '—'),
                'os_nombre': os_nombre,
                'os_plan': plan_nombre,
                'particular': particular,
                'ticket_validacion': str(r.op_id or '—'),
                'afiliado_nombre': afiliado_nombre,
                'afiliado_nro': afiliado_dni,
                'afiliado_dni': afiliado_dni,
                'medico_id': r.med_id,
                'medico_nombre': (med_map.get(r.med_id).nombre if r.med_id and med_map.get(r.med_id) and med_map.get(r.med_id).nombre else (f'Médico #{r.med_id}' if r.med_id else '—')),
                'medico_matricula': r.med_id or '',
                'receta_tipo': '—',
                'receta_fecha_emision': r.fecha,
                'items': items,
                'importe_total': float(r.total or 0),
                'importe_os': float(r.total_os or 0),
                'importe_paciente': float(r.total or 0) - float(r.total_os or 0),
            })

        total_recetas = len(dispensas)
        total_os = sum(d['importe_os'] for d in dispensas)
        total_pac = sum(d['importe_paciente'] for d in dispensas)

        top_os = (session.query(
                        ObsObraSocial.observer_id,
                        ObsObraSocial.descripcion,
                        func.count(ObsVentaDetalle.id_producto_vendido).label('n'),
                    )
                    .join(ObsVentaDetalle,
                          ObsVentaDetalle.obra_social_observer == ObsObraSocial.observer_id)
                    .filter(
                        ObsVentaDetalle.id_farmacia == id_farmacia,
                        ObsVentaDetalle.fecha_estadistica >= desde,
                        ObsVentaDetalle.fecha_estadistica <= hasta,
                    )
                    .group_by(ObsObraSocial.observer_id, ObsObraSocial.descripcion)
                    .order_by(func.count(ObsVentaDetalle.id_producto_vendido).desc())
                    .limit(30).all())
        os_options = [{'id': o.observer_id, 'nombre': o.descripcion, 'n': o.n}
                      for o in top_os]

        # Top 30 médicos del período (cantidad de recetas).
        from database import ObsMedico
        top_med = (session.query(
                        ObsMedico.observer_id,
                        ObsMedico.nombre,
                        func.count(func.distinct(ObsVentaDetalle.id_operacion)).label('n'),
                    )
                    .join(ObsVentaDetalle,
                          ObsVentaDetalle.medico_observer == ObsMedico.observer_id)
                    .filter(
                        ObsVentaDetalle.id_farmacia == id_farmacia,
                        ObsVentaDetalle.fecha_estadistica >= desde,
                        ObsVentaDetalle.fecha_estadistica <= hasta,
                    )
                    .group_by(ObsMedico.observer_id, ObsMedico.nombre)
                    .order_by(func.count(func.distinct(ObsVentaDetalle.id_operacion)).desc())
                    .limit(30).all())
        med_options = [{'id': m.observer_id, 'nombre': m.nombre or f'Médico #{m.observer_id}', 'n': m.n}
                       for m in top_med]

        # Top 30 productos del período (cantidad de líneas).
        top_prod = (session.query(
                        ObsProducto.observer_id,
                        ObsProducto.descripcion,
                        func.count(ObsVentaDetalle.id_producto_vendido).label('n'),
                    )
                    .join(ObsVentaDetalle,
                          ObsVentaDetalle.producto_observer == ObsProducto.observer_id)
                    .filter(
                        ObsVentaDetalle.id_farmacia == id_farmacia,
                        ObsVentaDetalle.fecha_estadistica >= desde,
                        ObsVentaDetalle.fecha_estadistica <= hasta,
                    )
                    .group_by(ObsProducto.observer_id, ObsProducto.descripcion)
                    .order_by(func.count(ObsVentaDetalle.id_producto_vendido).desc())
                    .limit(30).all())
        prod_options = [{'id': p.observer_id, 'nombre': p.descripcion, 'n': p.n}
                        for p in top_prod]

        # Planes — filtrados por OS si se seleccionó una; sino los más usados del período.
        if os_filter_id:
            plan_q = (session.query(ObsPlan.observer_id, ObsPlan.descripcion,
                                    func.count(ObsVentaDetalle.id_producto_vendido).label('n'))
                      .join(ObsVentaDetalle,
                            ObsVentaDetalle.plan_principal_observer == ObsPlan.observer_id)
                      .join(ObsConvenio, ObsConvenio.observer_id == ObsPlan.convenio_observer)
                      .filter(
                          ObsVentaDetalle.id_farmacia == id_farmacia,
                          ObsVentaDetalle.fecha_estadistica >= desde,
                          ObsVentaDetalle.fecha_estadistica <= hasta,
                          ObsConvenio.obra_social_observer == os_filter_id,
                      )
                      .group_by(ObsPlan.observer_id, ObsPlan.descripcion)
                      .order_by(func.count(ObsVentaDetalle.id_producto_vendido).desc())
                      .limit(50).all())
        else:
            plan_q = (session.query(ObsPlan.observer_id, ObsPlan.descripcion,
                                    func.count(ObsVentaDetalle.id_producto_vendido).label('n'))
                      .join(ObsVentaDetalle,
                            ObsVentaDetalle.plan_principal_observer == ObsPlan.observer_id)
                      .filter(
                          ObsVentaDetalle.id_farmacia == id_farmacia,
                          ObsVentaDetalle.fecha_estadistica >= desde,
                          ObsVentaDetalle.fecha_estadistica <= hasta,
                      )
                      .group_by(ObsPlan.observer_id, ObsPlan.descripcion)
                      .order_by(func.count(ObsVentaDetalle.id_producto_vendido).desc())
                      .limit(30).all())
        plan_options = [{'id': p.observer_id, 'nombre': p.descripcion, 'n': p.n}
                        for p in plan_q]

        return dispensas, total_recetas, total_os, total_pac, os_options, med_options, prod_options, plan_options

    def _parse_filtros_dispensas():
        """Parsea querystring → dict con desde/hasta/os/tipo/q/medico/producto."""
        os_filter_id = request.args.get('os_id', type=int)
        med_filter_id = request.args.get('medico_id', type=int)
        prod_filter_id = request.args.get('producto_id', type=int)
        plan_filter_id = request.args.get('plan_id', type=int)
        cob_filter = (request.args.get('cobertura') or '').strip()
        if cob_filter not in ('0%', '1-39%', '40-69%', '70-99%', '100%'):
            cob_filter = ''
        # Filtro por importe total de la receta (rango). Útil para detectar
        # recetas grandes (cobranza alta) o muy chicas (filtro de ruido).
        try:
            importe_min = float((request.args.get('importe_min') or '').strip()) \
                if (request.args.get('importe_min') or '').strip() else None
        except ValueError:
            importe_min = None
        try:
            importe_max = float((request.args.get('importe_max') or '').strip()) \
                if (request.args.get('importe_max') or '').strip() else None
        except ValueError:
            importe_max = None
        desde_str = (request.args.get('desde') or '').strip()
        hasta_str = (request.args.get('hasta') or '').strip()
        tipo_filter = (request.args.get('tipo') or '').strip()
        q = (request.args.get('q') or '').strip()
        hoy = date.today()
        try:
            desde = date.fromisoformat(desde_str) if desde_str else (hoy - timedelta(days=30))
        except ValueError:
            desde = hoy - timedelta(days=30)
        try:
            hasta = date.fromisoformat(hasta_str) if hasta_str else hoy
        except ValueError:
            hasta = hoy
        return {
            'os_filter_id': os_filter_id,
            'med_filter_id': med_filter_id,
            'prod_filter_id': prod_filter_id,
            'plan_filter_id': plan_filter_id,
            'cob_filter': cob_filter,
            'importe_min': importe_min,
            'importe_max': importe_max,
            'desde': desde,
            'hasta': hasta,
            'tipo_filter': tipo_filter,
            'q': q,
        }

    @app.route('/obras-sociales/dispensas')
    def os_dispensas():
        """Listado de dispensas OS — datos reales de obs_ventas_detalle.

        Cada 'receta' = grupo de líneas con misma (cliente, medico, operacion).
        Las líneas son items individuales del carrito asociado a esa receta.
        """
        import os as _os

        from database import get_db

        id_farmacia = int(_os.environ.get('OBSERVER_ID_FARMACIA', '10525'))
        filtros = _parse_filtros_dispensas()

        with get_db() as session:
            dispensas, total_recetas, total_os, total_pac, os_options, med_options, prod_options, plan_options = (
                _query_dispensas(filtros, session, id_farmacia))
            # Resolver nombre del médico/producto/plan seleccionado para mostrar en el chip.
            med_nombre = None
            if filtros.get('med_filter_id'):
                from database import ObsMedico
                m = session.get(ObsMedico, filtros['med_filter_id'])
                med_nombre = m.nombre if m and m.nombre else f"Médico #{filtros['med_filter_id']}"
            prod_nombre = None
            if filtros.get('prod_filter_id'):
                from database import ObsProducto as _ObsProducto
                p = session.get(_ObsProducto, filtros['prod_filter_id'])
                prod_nombre = p.descripcion if p else f"Producto #{filtros['prod_filter_id']}"
            plan_nombre = None
            if filtros.get('plan_filter_id'):
                from database import ObsPlan as _ObsPlan
                pl = session.get(_ObsPlan, filtros['plan_filter_id'])
                plan_nombre = pl.descripcion if pl else f"Plan #{filtros['plan_filter_id']}"

        return render_template(
            'os_dispensas.html',
            dispensas=dispensas[:200],
            total_recetas=total_recetas,
            total_os=total_os,
            total_pac=total_pac,
            os_filter_id=filtros['os_filter_id'],
            med_filter_id=filtros.get('med_filter_id'),
            prod_filter_id=filtros.get('prod_filter_id'),
            plan_filter_id=filtros.get('plan_filter_id'),
            cob_filter=filtros.get('cob_filter') or '',
            importe_min=filtros.get('importe_min'),
            importe_max=filtros.get('importe_max'),
            med_nombre=med_nombre,
            prod_nombre=prod_nombre,
            plan_nombre=plan_nombre,
            os_options=os_options,
            med_options=med_options,
            prod_options=prod_options,
            plan_options=plan_options,
            tipo_filter=filtros['tipo_filter'],
            q_text=filtros['q'],
            desde=filtros['desde'].isoformat(),
            hasta=filtros['hasta'].isoformat(),
            es_real=True,
        )

    @app.route('/obras-sociales/dispensas/export.xlsx')
    def os_dispensas_export():
        """Export Excel del listado de dispensas con los mismos filtros que la
        pantalla. 1 fila por item (no por receta) — formato plano para Excel.
        """
        import io
        import os as _os

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from flask import send_file

        from database import get_db

        id_farmacia = int(_os.environ.get('OBSERVER_ID_FARMACIA', '10525'))
        filtros = _parse_filtros_dispensas()

        with get_db() as session:
            dispensas, total_recetas, total_os, total_pac, _os_options, _med_options, _prod_options, _plan_options = (
                _query_dispensas(filtros, session, id_farmacia))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Dispensas'

        # Header
        headers = [
            'Fecha', 'N° Operación', 'Cliente', 'DNI', 'Particular',
            'Obra Social', 'Plan', 'Médico ID',
            'Producto', 'Cantidad', 'Importe Total',
            'Importe OS', 'Importe Paciente', 'Cobertura %',
        ]
        ws.append(headers)
        bold = Font(bold=True)
        fill = PatternFill('solid', fgColor='1c1c1e')
        white = Font(bold=True, color='ffffff')
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = white
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')

        # 1 fila por item (no por receta)
        for d in dispensas:
            for it in d['items']:
                ws.append([
                    d['fecha_venta'].strftime('%d/%m/%Y %H:%M') if d['fecha_venta'] else '',
                    d['ticket_validacion'],
                    d['afiliado_nombre'],
                    d['afiliado_dni'],
                    'SÍ' if d['particular'] else 'NO',
                    d['os_nombre'] if not d['particular'] else '—',
                    d['os_plan'] or '',
                    d['medico_id'] or '',
                    it['descripcion'],
                    it['cantidad'],
                    it['importe'],
                    it['importe_os'],
                    it['importe'] - it['importe_os'],
                    it['cobertura_pct'],
                ])

        # Totales abajo
        ws.append([])
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=10, value='TOTAL').font = bold
        ws.cell(row=total_row, column=11, value=total_recetas).font = bold
        ws.cell(row=total_row, column=12, value=total_os).font = bold
        ws.cell(row=total_row, column=13, value=total_pac).font = bold

        # Anchos de columna
        for col, w in zip('ABCDEFGHIJKLMN',
                          [16, 12, 35, 14, 8, 25, 22, 10, 40, 8, 12, 12, 12, 10]):
            ws.column_dimensions[col].width = w

        # Send
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        nombre = (
            f'dispensas_{filtros["desde"].isoformat()}_'
            f'{filtros["hasta"].isoformat()}.xlsx'
        )
        return send_file(buf, as_attachment=True, download_name=nombre,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/medico/<int:medico_id>')
    def medico_detalle(medico_id):
        """Análisis de un médico: donut por laboratorio + detalle por producto."""
        from sqlalchemy import distinct, func
        import os as _os
        from database import (
            ObsLaboratorio, ObsMedico, ObsObraSocial, ObsProducto,
            ObsVentaDetalle, get_db,
        )

        id_farmacia = int(_os.environ.get('OBSERVER_ID_FARMACIA', '10525'))

        hoy = date.today()
        # Default: últimos 6 meses (180 días).
        desde_str = (request.args.get('desde') or '').strip()
        hasta_str = (request.args.get('hasta') or '').strip()
        try:
            desde = date.fromisoformat(desde_str) if desde_str else (hoy - timedelta(days=180))
        except ValueError:
            desde = hoy - timedelta(days=180)
        try:
            hasta = date.fromisoformat(hasta_str) if hasta_str else hoy
        except ValueError:
            hasta = hoy
        if hasta < desde:
            desde, hasta = hasta, desde

        os_filter_id = request.args.get('os_id', type=int)
        lab_filter_id = request.args.get('lab_id', type=int)

        with get_db() as session:
            medico = session.get(ObsMedico, medico_id)
            if not medico:
                flash('Médico no encontrado.', 'error')
                return redirect(url_for('os_dashboard'))

            base_filters = [
                ObsVentaDetalle.id_farmacia == id_farmacia,
                ObsVentaDetalle.medico_observer == medico_id,
                ObsVentaDetalle.fecha_estadistica >= desde,
                ObsVentaDetalle.fecha_estadistica <= hasta,
            ]
            if os_filter_id:
                base_filters.append(ObsVentaDetalle.obra_social_observer == os_filter_id)
            if lab_filter_id:
                # Subquery: productos de ese lab.
                sub_prod = (session.query(ObsProducto.observer_id)
                            .filter(ObsProducto.laboratorio_observer == lab_filter_id)
                            .subquery())
                base_filters.append(ObsVentaDetalle.producto_observer.in_(sub_prod))

            # KPIs cabecera
            kpi_q = (session.query(
                        func.count(distinct(func.concat(
                            ObsVentaDetalle.id_operacion, '|',
                            func.coalesce(ObsVentaDetalle.cliente_observer, 0), '|',
                            ObsVentaDetalle.fecha_estadistica,
                        ))).label('recetas'),
                        func.coalesce(func.sum(ObsVentaDetalle.cantidad), 0).label('unidades'),
                        func.coalesce(func.sum(ObsVentaDetalle.importe), 0).label('total'),
                        func.coalesce(func.sum(ObsVentaDetalle.importe_a_cargo_os), 0).label('os'),
                    )
                    .filter(*base_filters)).first()
            kpi = {
                'recetas':  int(kpi_q[0] or 0),
                'unidades': float(kpi_q[1] or 0),
                'total':    float(kpi_q[2] or 0),
                'os':       float(kpi_q[3] or 0),
                'paciente': float((kpi_q[2] or 0) - (kpi_q[3] or 0)),
            }

            # Donut: ventas por laboratorio (de los productos recetados).
            lab_q = (session.query(
                        ObsLaboratorio.observer_id,
                        ObsLaboratorio.descripcion,
                        func.count(distinct(func.concat(
                            ObsVentaDetalle.id_operacion, '|',
                            func.coalesce(ObsVentaDetalle.cliente_observer, 0),
                        ))).label('recetas'),
                        func.coalesce(func.sum(ObsVentaDetalle.importe), 0).label('importe'),
                    )
                    .join(ObsProducto, ObsProducto.observer_id == ObsVentaDetalle.producto_observer)
                    .join(ObsLaboratorio, ObsLaboratorio.observer_id == ObsProducto.laboratorio_observer)
                    .filter(*base_filters)
                    .group_by(ObsLaboratorio.observer_id, ObsLaboratorio.descripcion)
                    .order_by(func.sum(ObsVentaDetalle.importe).desc())
                    .limit(20)).all()
            donut_labs = [{
                'lab_id': r[0], 'lab': r[1], 'recetas': int(r[2] or 0), 'importe': float(r[3] or 0)
            } for r in lab_q]

            # Detalle por producto.
            prod_q = (session.query(
                        ObsProducto.observer_id.label('prod_id'),
                        ObsProducto.descripcion.label('prod'),
                        ObsLaboratorio.descripcion.label('lab'),
                        func.count(distinct(func.concat(
                            ObsVentaDetalle.id_operacion, '|',
                            func.coalesce(ObsVentaDetalle.cliente_observer, 0),
                        ))).label('recetas'),
                        func.coalesce(func.sum(ObsVentaDetalle.cantidad), 0).label('unidades'),
                        func.coalesce(func.sum(ObsVentaDetalle.importe), 0).label('importe'),
                        func.coalesce(func.sum(ObsVentaDetalle.importe_a_cargo_os), 0).label('os'),
                    )
                    .join(ObsProducto, ObsProducto.observer_id == ObsVentaDetalle.producto_observer)
                    .outerjoin(ObsLaboratorio,
                                ObsLaboratorio.observer_id == ObsProducto.laboratorio_observer)
                    .filter(*base_filters)
                    .group_by(ObsProducto.observer_id, ObsProducto.descripcion,
                              ObsLaboratorio.descripcion)
                    .order_by(func.sum(ObsVentaDetalle.importe).desc())
                    .limit(500)).all()
            detalle = [{
                'prod_id':  r[0],
                'producto': r[1],
                'lab':      r[2] or '—',
                'recetas':  int(r[3] or 0),
                'unidades': float(r[4] or 0),
                'importe':  float(r[5] or 0),
                'os':       float(r[6] or 0),
                'paciente': float(r[5] or 0) - float(r[6] or 0),
            } for r in prod_q]
            # Totales para la fila al pie
            totales = {
                'recetas':  sum(d['recetas']  for d in detalle),
                'unidades': sum(d['unidades'] for d in detalle),
                'importe':  sum(d['importe']  for d in detalle),
                'os':       sum(d['os']       for d in detalle),
                'paciente': sum(d['paciente'] for d in detalle),
            }

            # Opciones para los selectores OS y Lab (las usadas por este médico).
            os_opts = (session.query(ObsObraSocial.observer_id, ObsObraSocial.descripcion,
                                      func.count(distinct(ObsVentaDetalle.id_operacion)).label('n'))
                       .join(ObsVentaDetalle,
                             ObsVentaDetalle.obra_social_observer == ObsObraSocial.observer_id)
                       .filter(ObsVentaDetalle.id_farmacia == id_farmacia,
                               ObsVentaDetalle.medico_observer == medico_id,
                               ObsVentaDetalle.fecha_estadistica >= desde,
                               ObsVentaDetalle.fecha_estadistica <= hasta)
                       .group_by(ObsObraSocial.observer_id, ObsObraSocial.descripcion)
                       .order_by(func.count(distinct(ObsVentaDetalle.id_operacion)).desc())
                       .limit(30).all())
            os_options = [{'id': r[0], 'nombre': r[1], 'n': r[2]} for r in os_opts]
            lab_options = sorted(
                [{'id': l['lab_id'], 'nombre': l['lab']} for l in donut_labs if l['lab_id']],
                key=lambda x: x['nombre'].lower()
            )
            os_filter_nombre = next((o['nombre'] for o in os_options if o['id'] == os_filter_id), None)
            lab_filter_nombre = next((l['nombre'] for l in lab_options if l['id'] == lab_filter_id), None)

        return render_template(
            'medico_detalle.html',
            medico=medico,
            kpi=kpi,
            donut_labs=donut_labs,
            detalle=detalle,
            totales=totales,
            os_options=os_options,
            lab_options=lab_options,
            os_filter_id=os_filter_id,
            lab_filter_id=lab_filter_id,
            os_filter_nombre=os_filter_nombre,
            lab_filter_nombre=lab_filter_nombre,
            desde=desde.isoformat(),
            hasta=hasta.isoformat(),
        )
