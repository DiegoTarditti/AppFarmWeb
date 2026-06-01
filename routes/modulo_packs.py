"""Módulo packs routes: CRUD, import, vista, toggle activo."""

import os

from flask import flash, jsonify, make_response, redirect, render_template, request, url_for
from werkzeug.utils import secure_filename

import database
from database import Laboratorio, Modulo, ModuloPack, Producto
from helpers import UPLOAD_FOLDER


def init_app(app):

    @app.route('/modulo-packs')
    def modulo_packs_list():
        with database.get_db() as session:
            all_prods = session.query(Producto).order_by(Producto.descripcion).all()
            prod_map = {p.codigo_barra: p for p in all_prods}
            # Solo labs que tienen al menos un módulo
            labs_con_modulos = {
                lid for (lid,) in session.query(Modulo.laboratorio_id)
                .filter(Modulo.laboratorio_id.isnot(None)).distinct().all()
            }
            labs = (session.query(Laboratorio)
                    .filter(Laboratorio.id.in_(labs_con_modulos))
                    .order_by(Laboratorio.nombre).all()) if labs_con_modulos else []

            # Fallback para descripciones desde obs_productos (cuando la unidad
            # sugerida viene del catálogo general, ean_unidad = str(observer_id))
            ean_unidades = {mp.ean_unidad for mp in session.query(database.ModuloPack.ean_unidad).all()}
            # ean_unidad con prefijo 'OBS:' = observer_id directo al catálogo ObServer
            obs_candidates = [e for e in ean_unidades if e and e.startswith('OBS:') and e not in prod_map]
            obs_desc_map = {}
            if obs_candidates:
                obs_ids = []
                for e in obs_candidates:
                    try:
                        obs_ids.append(int(e[4:]))
                    except (ValueError, TypeError):
                        pass
                if obs_ids:
                    obs_rows = session.query(database.ObsProducto.observer_id, database.ObsProducto.descripcion).filter(
                        database.ObsProducto.observer_id.in_(obs_ids)
                    ).all()
                    obs_desc_map = {f'OBS:{oid}': desc for oid, desc in obs_rows}

            def _desc_unidad(mp):
                if mp.ean_unidad in prod_map:
                    return prod_map[mp.ean_unidad].descripcion or ''
                return obs_desc_map.get(mp.ean_unidad, '')

            # ── Stock + presentación por ean_unidad (para columna "Stock") ──
            # Mapeamos ean_unidad → observer_id (vía 'OBS:N' o Producto.observer_id).
            from services.farmacia import farmacia_operativa
            _id_farm = farmacia_operativa()
            ean_to_obs = {}
            for mp in session.query(database.ModuloPack.ean_unidad).distinct():
                e = mp[0]
                if not e:
                    continue
                if e.startswith('OBS:'):
                    try:
                        ean_to_obs[e] = int(e[4:])
                    except (ValueError, TypeError):
                        pass
                elif e in prod_map and prod_map[e].observer_id:
                    ean_to_obs[e] = prod_map[e].observer_id
            stock_actual_by_obs = {}
            if ean_to_obs and _id_farm:
                _obs_ids = list({oid for oid in ean_to_obs.values()})
                for po, sa in (session.query(
                        database.ObsStock.producto_observer,
                        database.ObsStock.stock_actual)
                        .filter(database.ObsStock.id_farmacia == _id_farm,
                                database.ObsStock.producto_observer.in_(_obs_ids))):
                    stock_actual_by_obs[po] = int(sa or 0)
            # Fraccionado + cantidad_envase del producto unidad (cuando exista
            # en el catálogo master). Si ean_unidad es 'OBS:N' o no está en
            # productos, no aplica fraccionable display.
            _prod_ids_unidad = [p.id for p in all_prods if p.codigo_barra in ean_to_obs and p.codigo_barra in prod_map]
            envase_by_pid = {}
            if _prod_ids_unidad:
                for pid, ce in (session.query(database.ProductoAtributo.producto_id,
                                              database.ProductoAtributo.cantidad_envase)
                                .filter(database.ProductoAtributo.producto_id.in_(_prod_ids_unidad),
                                        database.ProductoAtributo.cantidad_envase.isnot(None))):
                    envase_by_pid[pid] = int(ce)

            def _stock_info_unidad(ean):
                """Devuelve dict con stock + flags para mostrar la celda. None si no hay dato."""
                if not ean:
                    return None
                oid = ean_to_obs.get(ean)
                if oid is None or oid not in stock_actual_by_obs:
                    return None
                stock = stock_actual_by_obs[oid]
                p = prod_map.get(ean)
                fraccionado = bool(p.fraccionado) if p else False
                envase = envase_by_pid.get(p.id) if p else None
                return {'stock': stock, 'fraccionado': fraccionado, 'envase': envase}

            def _pack_dict(mp):
                return {'id': mp.id, 'ean_pack': mp.ean_pack, 'ean_unidad': mp.ean_unidad,
                        'cantidad': mp.cantidad,
                        'cant_modulo': mp.cant_modulo,
                        'desc_pct': float(mp.desc_pct) if mp.desc_pct is not None else None,
                        'desc_pack':   mp.descripcion or '',
                        'desc_unidad': _desc_unidad(mp),
                        'prod_unidad_id': prod_map[mp.ean_unidad].id if mp.ean_unidad in prod_map else None,
                        'modulo_id': mp.modulo_id,
                        'stock_unidad': _stock_info_unidad(mp.ean_unidad)}

            modulos_raw = (session.query(Modulo)
                           .outerjoin(Laboratorio)
                           .order_by(Modulo.lista_nombre, Laboratorio.nombre, Modulo.nombre).all())
            def _lista_nombre(m):
                if m.lista_nombre:
                    return m.lista_nombre
                return m.laboratorio.nombre if m.laboratorio else '—'

            modulos = [{'id': m.id, 'nombre': m.nombre,
                        'lab_nombre': m.laboratorio.nombre if m.laboratorio else '—',
                        'lab_id': m.laboratorio_id or 0,
                        'lista_nombre': _lista_nombre(m),
                        'is_lista_marker': bool(m.lista_nombre and m.nombre == m.lista_nombre),
                        'creado_en': m.creado_en.strftime('%d/%m/%Y') if m.creado_en else '',
                        'activo': m.activo,
                        'packs': [_pack_dict(mp) for mp in m.packs]}
                       for m in modulos_raw]

            lista_activo_map = {}
            lista_toggle_map = {}
            for md in modulos:
                ln = md['lista_nombre']
                if md['activo']:
                    lista_activo_map[ln] = True
                if ln not in lista_toggle_map or md['is_lista_marker']:
                    lista_toggle_map[ln] = md['id']
            for md in modulos:
                ln = md['lista_nombre']
                md['lista_activo']    = lista_activo_map.get(ln, False)
                md['lista_toggle_id'] = lista_toggle_map.get(ln, md['id'])

            orphan_packs = [_pack_dict(mp) for mp in
                            session.query(ModuloPack).filter(ModuloPack.modulo_id.is_(None))
                            .order_by(ModuloPack.ean_pack).all()]

            # Primer EAN alternativo (de producto_codigos_barra) por producto.
            # Antes era `alt1 or alt2 or alt3`; ahora la 1-a-N es la fuente.
            from database import ProductoCodigoBarra
            prod_ids_visibles = [p.id for p in all_prods]
            primer_alt = {}
            if prod_ids_visibles:
                for pid, ean in (session.query(ProductoCodigoBarra.producto_id,
                                                ProductoCodigoBarra.codigo_barra)
                                  .filter(ProductoCodigoBarra.producto_id.in_(prod_ids_visibles))
                                  .filter(ProductoCodigoBarra.es_principal.is_(False))
                                  .all()):
                    if pid not in primer_alt and ean:
                        primer_alt[pid] = ean

            def _first_alt(p):
                return primer_alt.get(p.id, '')
            prods_pack = [{'ean': p.codigo_barra, 'desc': p.descripcion or '',
                           'alt': _first_alt(p), 'is_pack': bool(p.es_pack)}
                          for p in all_prods if p.es_pack]
            prods_all  = [{'ean': p.codigo_barra, 'desc': p.descripcion or '',
                           'alt': _first_alt(p), 'is_pack': bool(p.es_pack)}
                          for p in all_prods]
            return render_template('modulo_packs.html',
                                   modulos=modulos, orphan_packs=orphan_packs,
                                   labs=[{'id': l.id, 'nombre': l.nombre} for l in labs],
                                   prods_pack=prods_pack, prods_all=prods_all)

    @app.route('/modulo-packs/vista')
    def modulo_packs_vista():
        # /modulo-packs/vista fue consolidado en /modulo-packs (las dos vistas
        # hacían lo mismo). Mantenemos la ruta como redirect para no romper
        # bookmarks / links externos.
        lab = request.args.get('lab', '').strip()
        return redirect(
            url_for('modulo_packs_list') + (f'?lab={lab}' if lab else ''))

    @app.route('/modulo-packs/plantilla')
    def modulo_packs_plantilla():
        """Descarga plantilla XLSX para importar módulos (Formato A)."""
        import io

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.styles.numbers import FORMAT_NUMBER
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Módulos'

        hdr_fill  = PatternFill('solid', fgColor='1C1C1E')
        hdr_font  = Font(bold=True, color='EAB308', size=10)
        mod_fill  = PatternFill('solid', fgColor='FEF9C3')
        mod_font  = Font(bold=True, color='92400E')
        item_fill = PatternFill('solid', fgColor='FAFAFA')
        border_b  = Border(bottom=Side(style='thin', color='D0D0D0'))

        # Fila 1: encabezados
        headers = ['NOMBRE MÓDULO', 'EAN', 'DESCRIPCIÓN', 'CANT. MÓDULO', 'DESC. %']
        ws.append(headers)
        for ci in range(1, 6):
            c = ws.cell(row=1, column=ci)
            c.fill = hdr_fill; c.font = hdr_font; c.border = border_b
            c.alignment = Alignment(horizontal='center')

        # Módulo de ejemplo 1
        ws.append(['MOD. EJEMPLO AMOXICILINA', None, None, None, None])
        c = ws.cell(row=2, column=1); c.fill = mod_fill; c.font = mod_font
        for ean, desc, cant, dto in [
            ('7790001000001', 'AMOXICILINA 500 mg COM x 16', 10, 7.0),
            ('7790001000003', 'AMOXICILINA 500 mg COM x 32', 6,  7.0),
        ]:
            r = ws.max_row + 1
            ws.append([None, ean, desc, cant, dto])
            ws.cell(row=r, column=1).fill = item_fill
            ws.cell(row=r, column=2).number_format = '@'

        # Fila vacía separadora
        ws.append([])

        # Módulo de ejemplo 2
        ws.append(['MOD. EJEMPLO IBUPROFENO', None, None, None, None])
        c = ws.cell(row=ws.max_row, column=1); c.fill = mod_fill; c.font = mod_font
        for ean, desc, cant, dto in [
            ('7790002000001', 'IBUPROFENO 400 mg COM x 20', 12, 10.0),
        ]:
            r = ws.max_row + 1
            ws.append([None, ean, desc, cant, dto])
            ws.cell(row=r, column=1).fill = item_fill
            ws.cell(row=r, column=2).number_format = '@'

        # Anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 10

        # Congelar encabezados
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = 'attachment; filename="plantilla_modulos.xlsx"'
        return resp

    @app.route('/api/packs/buscar-unidad')
    def api_packs_buscar_unidad():
        """Búsqueda combinada productos locales + obs_productos para el dropdown
        del Buscar equivalente. Excluye candidatos que parecen pack (descripción
        con 'PACK X N').

        Query: ?q=amoxidal  (mínimo 2 chars)
        Respuesta: {results: [{ean, desc, fuente: 'local'|'observer'}]}
        """
        import re as _re
        q = (request.args.get('q') or '').strip()
        if len(q) < 2:
            return jsonify({'results': []})
        q_low = q.lower()
        PACK_RE = _re.compile(r'\bPACK\s*X\s*\d+\b', _re.IGNORECASE)
        out = []
        with database.get_db() as session:
            # Locales (tienen EAN real)
            locales = (session.query(database.Producto)
                       .filter(database.Producto.codigo_barra.ilike(f'%{q}%') |
                               database.Producto.descripcion.ilike(f'%{q}%'))
                       .limit(30).all())
            for p in locales:
                if not p.descripcion or PACK_RE.search(p.descripcion):
                    continue
                out.append({'ean': p.codigo_barra, 'desc': p.descripcion or '',
                            'fuente': 'local'})
            # Obs_productos (usamos observer_id como pseudo-EAN)
            eans_vistos = {r['ean'] for r in out}
            obs = (session.query(database.ObsProducto)
                   .filter(database.ObsProducto.descripcion.ilike(f'%{q}%'),
                           database.ObsProducto.fecha_baja.is_(None))
                   .limit(30).all())
            for o in obs:
                if not o.descripcion or PACK_RE.search(o.descripcion):
                    continue
                ean = f'OBS:{o.observer_id}'
                if ean in eans_vistos:
                    continue
                out.append({'ean': ean, 'desc': o.descripcion,
                            'fuente': 'observer'})
        # Orden: primero los que el q matchea en desc al inicio
        out.sort(key=lambda r: (0 if r['desc'].lower().startswith(q_low) else 1,
                                r['fuente'] == 'observer',
                                len(r['desc'])))
        return jsonify({'results': out[:40]})

    @app.route('/modulo-packs/importar', methods=['POST'])
    def modulo_packs_importar():
        """Importa módulos desde un XLSX (formato Roemmers o plantilla propia).

        Detecta automáticamente los packs usando:
          - filas destacadas en amarillo en el Excel
          - patrón 'PACK X N' en la descripción
          - ausencia de ventas históricas por ese EAN
        Los detectados se guardan con ean_unidad + cantidad sugeridos.
        Los no-pack se guardan con ean_unidad=ean_pack y cantidad=1 (como antes)."""
        from pack_detector import detectar_packs
        from parsers.modulos_xlsx import parse_modulos_xlsx
        f = request.files.get('file')
        lab_id = request.form.get('lab_id') or None
        lista_nombre = (request.form.get('lista_nombre') or '').strip() or None
        if not f:
            return jsonify({'error': 'No se recibió archivo'}), 400
        tmp = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
        f.save(tmp)
        try:
            with database.get_db() as session:
                try:
                    modules = parse_modulos_xlsx(tmp)
                    if not modules:
                        return jsonify({'error': 'No se encontraron módulos en el archivo'}), 400

                    # Detectar packs. Criterios:
                    #  - alta/media con cantidad del regex → auto con cant real
                    #  - destacado en amarillo pero sin cantidad → cant=2 placeholder
                    #    (el user edita después). El amarillo lo marcó el vendedor,
                    #    es la señal más fuerte.
                    #  - baja (solo sin_ventas) → revisión manual, no auto
                    packs_detectados = detectar_packs(modules, session, saltear_registrados=False)
                    pack_map = {}
                    for p in packs_detectados:
                        if p['confianza'] == 'baja':
                            continue
                        if not p.get('cantidad'):
                            if p.get('destacado'):
                                # Placeholder: está marcado como pack pero sin
                                # cantidad definida. cant=2 para que supere 1
                                # y se trate como pack en el front.
                                p = dict(p, cantidad=2)
                            else:
                                continue
                        pack_map[p['ean_pack']] = p

                    creados = 0
                    packs_agregados = 0
                    packs_auto = 0
                    for mod in modules:
                        nombre_mod = mod['nombre']
                        modulo_actual = session.query(Modulo).filter_by(nombre=nombre_mod, lista_nombre=lista_nombre).first()
                        if not modulo_actual:
                            modulo_actual = Modulo(nombre=nombre_mod,
                                                   laboratorio_id=int(lab_id) if lab_id else None,
                                                   lista_nombre=lista_nombre)
                            session.add(modulo_actual)
                            session.flush()
                            creados += 1
                        for item in mod['items']:
                            ean_pack = item['ean']
                            if not ean_pack:
                                continue
                            existe = session.query(ModuloPack).filter_by(
                                ean_pack=ean_pack, modulo_id=modulo_actual.id).first()
                            if existe:
                                continue

                            det = pack_map.get(ean_pack)
                            if det:
                                # Si no hay unidad sugerida, ponemos ean_pack
                                # (marca el pack pero queda pendiente de
                                # completar la equivalencia a mano)
                                ean_u = det['ean_unidad_sug'] or ean_pack
                                cant = det['cantidad']
                                packs_auto += 1
                            else:
                                ean_u = ean_pack   # no es pack → relación 1:1
                                cant = 1

                            session.add(ModuloPack(
                                ean_pack=ean_pack,
                                ean_unidad=ean_u,
                                cantidad=cant,
                                descripcion=item.get('descripcion', ''),
                                cant_modulo=item.get('cant'),
                                desc_pct=item.get('desc_pct'),
                                modulo_id=modulo_actual.id,
                            ))
                            packs_agregados += 1
                    session.commit()
                    return jsonify({'ok': True,
                                    'modulos_creados': creados,
                                    'packs_agregados': packs_agregados,
                                    'packs_auto_detectados': packs_auto,
                                    'packs_pendientes_revision': len(packs_detectados) - packs_auto})
                except Exception as e:
                    session.rollback()
                    return jsonify({'error': str(e)}), 500
        finally:
            try: os.remove(tmp)
            except OSError: pass

    # ── Nuevo flujo: preview + confirmar (no ensucia las tablas hasta confirmar) ──

    def _parsear_y_armar_preview(modules, lab_id, lista_nombre):
        """Toma la lista de módulos (output del parser regex o IA) y devuelve
        el JSON de preview con detección de packs + validación EANs contra
        catálogo + detección de conflicto (módulo activo del mismo lab)."""
        from pack_detector import detectar_packs
        with database.get_db() as session:
            # Detección packs (igual que /importar).
            detect = detectar_packs(modules, session, saltear_registrados=False)
            pack_map = {}
            for p in detect:
                if p['confianza'] == 'baja':
                    continue
                if not p.get('cantidad'):
                    if p.get('destacado'):
                        p = dict(p, cantidad=2)
                    else:
                        continue
                pack_map[p['ean_pack']] = p

            # Lookup en pack_equivalencias (aprendidas o cargadas manualmente
            # vía Excel desde el ABM de laboratorios). Si un ean_pack ya tiene
            # equivalencia persistida, lo aplicamos sin pedirle al user.
            # Same shape que /importar (modulos_import.py:295-320).
            todos_ean_pack = [it.get('ean') for mod in modules
                              for it in (mod.get('items') or []) if it.get('ean')]
            if todos_ean_pack:
                q = session.query(database.PackEquivalencia).filter(
                    database.PackEquivalencia.ean_pack.in_(todos_ean_pack[:1000]))
                # Si vino lab_id en el form, priorizar equivalencias del lab
                # (caen primero — luego globales como fallback).
                aprendidas = q.all()
                if lab_id:
                    aprendidas.sort(key=lambda x: 0 if x.laboratorio_id == lab_id else 1)
                seen = set()
                for pe in aprendidas:
                    if pe.ean_pack in seen:
                        continue
                    seen.add(pe.ean_pack)
                    cur = pack_map.get(pe.ean_pack)
                    if cur is None:
                        # No detectado como pack — promoverlo igual porque la
                        # tabla dice que es uno.
                        pack_map[pe.ean_pack] = {
                            'ean_pack': pe.ean_pack,
                            'ean_unidad_sug': pe.ean_unidad,
                            'desc_unidad_sug': pe.desc_unidad or '',
                            'cantidad': pe.cantidad or 1,
                            'confianza': 'alta',
                            'razon': 'aprendido',
                        }
                    elif not cur.get('ean_unidad_sug'):
                        cur['ean_unidad_sug'] = pe.ean_unidad
                        if not cur.get('desc_unidad_sug') and pe.desc_unidad:
                            cur['desc_unidad_sug'] = pe.desc_unidad
                        if pe.cantidad and pe.cantidad >= 2:
                            cur['cantidad'] = pe.cantidad
                        cur['razon'] = (cur.get('razon', '') + '+aprendido').lstrip('+')

            # Validación EANs contra catálogo (master + alts + obs_productos).
            from database import ProductoCodigoBarra
            todos_eans = set()
            for mod in modules:
                for it in mod.get('items') or []:
                    if it.get('ean'):
                        todos_eans.add(str(it['ean']).strip())
            for p in pack_map.values():
                if p.get('ean_unidad_sug'):
                    todos_eans.add(p['ean_unidad_sug'])
            eans_en_catalogo = set()
            desc_unidad_map = {}
            if todos_eans:
                # EANs reales (numericos): productos master + alts.
                eans_reales = [e for e in todos_eans if not e.startswith('OBS:')]
                if eans_reales:
                    eans_en_catalogo = {row[0] for row in
                        session.query(Producto.codigo_barra)
                        .filter(Producto.codigo_barra.in_(eans_reales)).all()}
                    eans_en_catalogo |= {row[0] for row in
                        session.query(ProductoCodigoBarra.codigo_barra)
                        .filter(ProductoCodigoBarra.codigo_barra.in_(eans_reales)).all()}
                # 'OBS:N' = referencia directa a obs_productos.observer_id.
                # Si el observer_id existe en obs_productos, lo damos por valido.
                obs_refs = {e for e in todos_eans if e.startswith('OBS:')}
                obs_ids = []
                if obs_refs:
                    for e in obs_refs:
                        try:
                            obs_ids.append(int(e[4:]))
                        except (ValueError, TypeError):
                            pass
                    if obs_ids:
                        ids_vigentes = {oid for (oid,) in
                            session.query(database.ObsProducto.observer_id)
                            .filter(database.ObsProducto.observer_id.in_(obs_ids)).all()}
                        for e in obs_refs:
                            try:
                                if int(e[4:]) in ids_vigentes:
                                    eans_en_catalogo.add(e)
                            except (ValueError, TypeError):
                                pass

                # Mapa EAN → descripcion (para mostrar al lado del EAN unidad
                # del preview, "descripcion equivalencia hija").
                desc_unidad_map = {}
                if eans_reales:
                    for cb, dsc in (session.query(Producto.codigo_barra, Producto.descripcion)
                                    .filter(Producto.codigo_barra.in_(eans_reales)).all()):
                        desc_unidad_map[cb] = (dsc or '').strip()
                    # Alts: si el EAN aparece como alt de un Producto, usar su desc.
                    for cb, pid in (session.query(ProductoCodigoBarra.codigo_barra,
                                                  ProductoCodigoBarra.producto_id)
                                    .filter(ProductoCodigoBarra.codigo_barra.in_(eans_reales)).all()):
                        if cb not in desc_unidad_map:
                            p = session.get(Producto, pid)
                            if p:
                                desc_unidad_map[cb] = (p.descripcion or '').strip()
                if obs_ids:
                    for oid, dsc in (session.query(database.ObsProducto.observer_id,
                                                   database.ObsProducto.descripcion)
                                     .filter(database.ObsProducto.observer_id.in_(obs_ids)).all()):
                        desc_unidad_map[f'OBS:{oid}'] = (dsc or '').strip()

            # Conflicto: ¿hay un módulo activo de este lab ya?
            conflicto = None
            if lab_id:
                activo = (session.query(Modulo)
                          .filter(Modulo.laboratorio_id == lab_id,
                                  Modulo.activo.is_(True))
                          .order_by(Modulo.id.desc()).first())
                if activo:
                    n_packs = session.query(ModuloPack).filter_by(modulo_id=activo.id).count()
                    conflicto = {'modulo_id': activo.id, 'nombre': activo.nombre,
                                 'lista_nombre': activo.lista_nombre or '',
                                 'n_packs': n_packs}

            # Armar payload por módulo + item.
            modulos_preview = []
            n_packs_auto = n_packs_pendientes = 0
            n_eans_pack_no_cat = n_eans_unidad_no_cat = 0
            for mod in modules:
                items_preview = []
                for it in mod.get('items') or []:
                    ean_pack = (str(it.get('ean') or '').strip()) or None
                    if not ean_pack:
                        continue
                    det = pack_map.get(ean_pack)
                    es_pack = bool(det)
                    if det:
                        ean_unidad = det.get('ean_unidad_sug') or ean_pack
                        cantidad = int(det.get('cantidad') or 1)
                        confianza = det.get('confianza', '—')
                        n_packs_auto += 1
                    else:
                        ean_unidad = ean_pack
                        cantidad = 1
                        confianza = None
                    cat_pack = 'ok' if ean_pack in eans_en_catalogo else 'no_en_catalogo'
                    if cat_pack == 'no_en_catalogo':
                        n_eans_pack_no_cat += 1
                    if es_pack and ean_unidad != ean_pack:
                        cat_unidad = 'ok' if ean_unidad in eans_en_catalogo else 'no_en_catalogo'
                        if cat_unidad == 'no_en_catalogo':
                            n_eans_unidad_no_cat += 1
                    else:
                        cat_unidad = 'igual_pack'
                    if es_pack and not det.get('ean_unidad_sug'):
                        n_packs_pendientes += 1
                    items_preview.append({
                        'ean_pack': ean_pack,
                        'descripcion': it.get('descripcion', ''),
                        'cant_modulo': it.get('cant'),
                        'desc_pct': float(it['desc_pct']) if it.get('desc_pct') is not None else None,
                        'destacado': bool(it.get('destacado')),
                        'es_pack': es_pack,
                        'ean_unidad': ean_unidad,
                        'desc_unidad': desc_unidad_map.get(ean_unidad, '') if ean_unidad != ean_pack else '',
                        'cantidad_pack': cantidad,
                        'confianza': confianza,
                        'catalogo_pack': cat_pack,
                        'catalogo_unidad': cat_unidad,
                    })
                if items_preview:
                    modulos_preview.append({'nombre': mod['nombre'], 'items': items_preview})

            stats = {
                'n_modulos': len(modulos_preview),
                'n_items': sum(len(m['items']) for m in modulos_preview),
                'n_packs_auto': n_packs_auto,
                'n_packs_pendientes': n_packs_pendientes,
                'n_eans_pack_no_catalogo': n_eans_pack_no_cat,
                'n_eans_unidad_no_catalogo': n_eans_unidad_no_cat,
            }
            return {
                'ok': True,
                'modulos': modulos_preview,
                'lista_nombre': lista_nombre,
                'lab_id': lab_id,
                'conflicto_activo': conflicto,
                'stats': stats,
            }

    @app.route('/modulo-packs/import-preview', methods=['POST'])
    def modulo_packs_import_preview():
        """Sube el archivo, lo parsea (regex o IA) y devuelve el JSON de
        preview con equivalencias + validación. NO toca la DB.
        Form: file (XLSX/PDF/imagen), lab_id, lista_nombre, usar_ia (0|1).
        """
        f = request.files.get('file')
        if not f:
            return jsonify({'error': 'No se recibió archivo'}), 400
        lab_id = request.form.get('lab_id') or None
        try:
            lab_id = int(lab_id) if lab_id else None
        except (TypeError, ValueError):
            lab_id = None
        lista_nombre = (request.form.get('lista_nombre') or '').strip() or None
        usar_ia = request.form.get('usar_ia') in ('1', 'true', 'on')

        ext = os.path.splitext(f.filename)[1].lower()
        IMG_EXTS = ('.jpg', '.jpeg', '.png', '.webp', '.bmp', '.tiff', '.tif')
        if usar_ia:
            if ext not in ('.xlsx', '.xls', '.pdf', *IMG_EXTS):
                return jsonify({'error': f'Formato {ext} no soportado por IA. Aceptamos XLSX, PDF y fotos.'}), 400
            if not os.environ.get('ANTHROPIC_API_KEY', '').strip():
                return jsonify({'error': 'IA no disponible: falta ANTHROPIC_API_KEY.'}), 503
        else:
            if ext not in ('.xlsx',):
                return jsonify({'error': 'El parser regex solo acepta XLSX. Para PDF/foto tildá "Usar IA".'}), 400

        tmp = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
        f.save(tmp)
        try:
            if usar_ia:
                from services import modulos_ia
                try:
                    modules = modulos_ia.extraer(tmp, ext)
                except RuntimeError as e:
                    return jsonify({'error': str(e)}), 503
                except Exception as e:   # noqa: BLE001
                    return jsonify({'error': f'IA: {e}'}), 500
            else:
                from parsers.modulos_xlsx import parse_modulos_xlsx
                try:
                    modules = parse_modulos_xlsx(tmp)
                except Exception as e:   # noqa: BLE001
                    return jsonify({'error': f'Parser: {e}'}), 500

            if not modules:
                return jsonify({'error': 'No se encontraron módulos en el archivo'}), 400

            preview = _parsear_y_armar_preview(modules, lab_id, lista_nombre)
            return jsonify(preview)
        finally:
            try: os.remove(tmp)
            except OSError: pass

    @app.route('/modulo-packs/import-confirmar', methods=['POST'])
    def modulo_packs_import_confirmar():
        """Recibe el JSON del preview (posiblemente editado) y persiste:
          - Si hay módulo activo del mismo lab: aplica accion_anterior:
              'pisar' → borra el viejo (cascade ModuloPacks);
              'conservar' → setea su activo=False (queda histórico);
              ausente y hay conflicto → 409.
          - Crea los Modulo + ModuloPack nuevos con activo=True.
        Body JSON: {lab_id, lista_nombre, modulos, accion_anterior?}.
        """
        from helpers import now_ar
        data = request.get_json(silent=True) or {}
        try:
            lab_id = int(data.get('lab_id')) if data.get('lab_id') else None
        except (TypeError, ValueError):
            lab_id = None
        lista_nombre = (data.get('lista_nombre') or '').strip() or None
        modulos_in = data.get('modulos') or []
        accion = (data.get('accion_anterior') or '').strip().lower()  # '' | 'pisar' | 'conservar'

        if not modulos_in:
            return jsonify({'error': 'No hay módulos para confirmar'}), 400

        with database.get_db() as session:
            try:
                # Conflicto con el activo previo (si hay lab).
                viejo = None
                if lab_id:
                    viejo = (session.query(Modulo)
                             .filter(Modulo.laboratorio_id == lab_id,
                                     Modulo.activo.is_(True))
                             .order_by(Modulo.id.desc()).first())
                if viejo and not accion:
                    return jsonify({
                        'error': 'conflicto_activo',
                        'conflicto': {'modulo_id': viejo.id, 'nombre': viejo.nombre},
                    }), 409
                if viejo:
                    if accion == 'pisar':
                        session.delete(viejo)
                        session.flush()
                    elif accion == 'conservar':
                        viejo.activo = False
                    else:
                        return jsonify({'error': f'accion_anterior inválida: {accion}'}), 400

                creados = 0
                packs_agregados = 0
                primer_modulo_id = None
                for mod in modulos_in:
                    nombre_mod = (mod.get('nombre') or '').strip()
                    if not nombre_mod:
                        continue
                    nuevo = Modulo(
                        nombre=nombre_mod, laboratorio_id=lab_id,
                        lista_nombre=lista_nombre, activo=True,
                    )
                    session.add(nuevo)
                    session.flush()
                    if primer_modulo_id is None:
                        primer_modulo_id = nuevo.id
                    creados += 1
                    for it in mod.get('items') or []:
                        ean_pack = (str(it.get('ean_pack') or '').strip())
                        if not ean_pack:
                            continue
                        ean_unidad = (str(it.get('ean_unidad') or '').strip()) or ean_pack
                        try:
                            cantidad = max(1, int(it.get('cantidad_pack') or 1))
                        except (TypeError, ValueError):
                            cantidad = 1
                        desc_pct = it.get('desc_pct')
                        try:
                            desc_pct = float(desc_pct) if desc_pct not in (None, '') else None
                        except (TypeError, ValueError):
                            desc_pct = None
                        try:
                            cant_modulo = int(it.get('cant_modulo')) if it.get('cant_modulo') not in (None, '') else None
                        except (TypeError, ValueError):
                            cant_modulo = None
                        session.add(ModuloPack(
                            ean_pack=ean_pack, ean_unidad=ean_unidad,
                            cantidad=cantidad,
                            descripcion=(it.get('descripcion') or '')[:255],
                            cant_modulo=cant_modulo, desc_pct=desc_pct,
                            modulo_id=nuevo.id,
                            creado_en=now_ar(),
                        ))
                        packs_agregados += 1
                session.commit()
                return jsonify({
                    'ok': True, 'modulos_creados': creados,
                    'packs_agregados': packs_agregados,
                    'primer_modulo_id': primer_modulo_id,
                })
            except Exception as e:   # noqa: BLE001
                session.rollback()
                return jsonify({'error': str(e)}), 500

    @app.route('/modulos/delete-by-lista', methods=['POST'])
    def modulos_delete_by_lista():
        """Elimina todos los módulos (y sus packs en cascada) de una lista/importación."""
        data = request.get_json(silent=True) or {}
        lista_nombre = data.get('lista_nombre', '').strip()
        if not lista_nombre:
            return jsonify({'error': 'lista_nombre requerido'}), 400
        with database.get_db() as session:
            try:
                modulos = session.query(Modulo).filter(
                    (Modulo.lista_nombre == lista_nombre) |
                    ((Modulo.lista_nombre.is_(None)) &
                     (Modulo.laboratorio.has(database.Laboratorio.nombre == lista_nombre)))
                ).all()
                count = len(modulos)
                for m in modulos:
                    session.delete(m)
                session.commit()
                return jsonify({'ok': True, 'eliminados': count})
            except Exception as e:
                session.rollback()
                return jsonify({'error': str(e)}), 500

    @app.route('/modulo-packs/activos')
    def modulo_packs_activos():
        """Devuelve las listas activas agrupadas por lista_nombre. ?lab=Nombre filtra por laboratorio."""
        lab_nombre = request.args.get('lab', '').strip()
        with database.get_db() as session:
            q = session.query(Modulo).filter_by(activo=True).outerjoin(Laboratorio)
            if lab_nombre:
                q = q.filter(Laboratorio.nombre == lab_nombre)
            raw = q.order_by(Modulo.lista_nombre, Modulo.nombre).all()
            prod_map = {p.codigo_barra: p for p in session.query(Producto).all()}

            # ── Resolver desc + categoría del EAN unidad por pack ─────────
            # cat_unidad: 'local' (en master/alts) | 'observer' (OBS:N vigente) | 'none'.
            # Permite mostrar un badge inline en /order/<id> sin n+1 queries.
            from database import ProductoCodigoBarra
            eans_pack_local = set()
            obs_ids_pack = []
            for m in raw:
                if m.lista_nombre and m.nombre == m.lista_nombre:
                    continue
                for mp in m.packs:
                    eu = mp.ean_unidad
                    if not eu:
                        continue
                    if eu.startswith('OBS:'):
                        try:
                            obs_ids_pack.append(int(eu[4:]))
                        except (ValueError, TypeError):
                            pass
                    else:
                        eans_pack_local.add(eu)
            # EANs reconocidos como locales (master + alts).
            eans_locales_ok = set(prod_map.keys()) & eans_pack_local
            if eans_pack_local:
                for (cb,) in (session.query(ProductoCodigoBarra.codigo_barra)
                              .filter(ProductoCodigoBarra.codigo_barra.in_(list(eans_pack_local)))):
                    eans_locales_ok.add(cb)
            obs_local = {}
            if obs_ids_pack:
                for oid, dsc in (session.query(database.ObsProducto.observer_id,
                                                database.ObsProducto.descripcion)
                                 .filter(database.ObsProducto.observer_id.in_(obs_ids_pack))):
                    obs_local[oid] = (dsc or '').strip()

            def _resolve_unit(eu):
                if not eu:
                    return ('', 'none')
                if eu.startswith('OBS:'):
                    try:
                        oid = int(eu[4:])
                        if oid in obs_local:
                            return (obs_local[oid], 'observer')
                    except (ValueError, TypeError):
                        pass
                    return ('', 'none')
                if eu in prod_map:
                    return (prod_map[eu].descripcion or '', 'local')
                if eu in eans_locales_ok:
                    return ('', 'local')
                return ('', 'none')

            from collections import OrderedDict
            listas = OrderedDict()
            for m in raw:
                ln = m.lista_nombre or m.nombre
                if ln not in listas:
                    listas[ln] = {'lista_nombre': ln,
                                   'lab_nombre': m.laboratorio.nombre if m.laboratorio else '',
                                   'modulos': []}
                if m.lista_nombre and m.nombre == m.lista_nombre:
                    continue
                packs = []
                for mp in m.packs:
                    desc_u, cat_u = _resolve_unit(mp.ean_unidad)
                    packs.append({
                        'ean_pack':    mp.ean_pack,
                        'desc_pack':   mp.descripcion or '',
                        'ean_unidad':  mp.ean_unidad,
                        'desc_unidad': desc_u,
                        'cat_unidad':  cat_u,
                        'cant_modulo': mp.cant_modulo if mp.cant_modulo is not None else mp.cantidad,
                        'desc_pct':    float(mp.desc_pct) if mp.desc_pct is not None else 0.0,
                    })
                listas[ln]['modulos'].append({'id': m.id, 'nombre': m.nombre, 'packs': packs})

            return jsonify({'listas': list(listas.values())})

    @app.route('/modulo/<int:modulo_id>/toggle-activo', methods=['POST'])
    def modulo_toggle_activo(modulo_id):
        with database.get_db() as session:
            try:
                m = session.get(Modulo, modulo_id)
                if not m:
                    return jsonify({'error': 'No encontrado'}), 404
                nuevo_estado = not bool(m.activo)
                if nuevo_estado:
                    session.query(Modulo).filter(
                        Modulo.laboratorio_id == m.laboratorio_id
                    ).update({'activo': False})
                    session.flush()
                if m.lista_nombre:
                    session.query(Modulo).filter(
                        Modulo.lista_nombre == m.lista_nombre
                    ).update({'activo': nuevo_estado})
                else:
                    m.activo = nuevo_estado
                session.commit()
                return jsonify({'activo': nuevo_estado})
            except Exception as e:
                session.rollback()
                return jsonify({'error': str(e)}), 500

    @app.route('/modulo/add', methods=['POST'])
    def modulo_add():
        data = request.get_json(silent=True) or {}
        nombre = (data.get('nombre') or '').strip()
        lab_id = data.get('laboratorio_id')
        lista_nombre = (data.get('lista_nombre') or nombre or '').strip() or None
        if not nombre:
            return jsonify({'error': 'Nombre requerido'}), 400
        with database.get_db() as session:
            try:
                m = Modulo(nombre=nombre,
                           laboratorio_id=int(lab_id) if lab_id else None,
                           lista_nombre=lista_nombre)
                session.add(m)
                session.commit()
                lab_nombre = m.laboratorio.nombre if m.laboratorio else '—'
                return jsonify({'ok': True, 'id': m.id, 'nombre': m.nombre,
                                'lab_nombre': lab_nombre,
                                'creado_en': m.creado_en.strftime('%d/%m/%Y') if m.creado_en else ''})
            except Exception as e:
                session.rollback()
                return jsonify({'error': str(e)}), 500

    @app.route('/modulo/<int:modulo_id>/delete', methods=['POST'])
    def modulo_delete(modulo_id):
        with database.get_db() as session:
            try:
                m = session.get(Modulo, modulo_id)
                if m:
                    session.delete(m)
                    session.commit()
                return jsonify({'ok': True})
            except Exception as e:
                session.rollback()
                return jsonify({'error': str(e)}), 500

    @app.route('/modulo-pack/<int:pack_id>/assign', methods=['POST'])
    def modulo_pack_assign(pack_id):
        data = request.get_json(silent=True) or {}
        modulo_id = data.get('modulo_id')
        with database.get_db() as session:
            try:
                mp = session.get(ModuloPack, pack_id)
                if not mp:
                    return jsonify({'error': 'Pack no encontrado'}), 404
                mp.modulo_id = int(modulo_id) if modulo_id else None
                session.commit()
                return jsonify({'ok': True})
            except Exception as e:
                session.rollback()
                return jsonify({'error': str(e)}), 500

    @app.route('/modulo-pack/add', methods=['POST'])
    def modulo_pack_add():
        data = request.get_json(silent=True) or {}
        ean_pack   = (data.get('ean_pack') or '').strip()
        ean_unidad = (data.get('ean_unidad') or '').strip()
        cantidad   = int(data.get('cantidad') or 1)
        descripcion = (data.get('descripcion') or '').strip()
        modulo_id  = data.get('modulo_id')
        if not ean_pack or not ean_unidad or cantidad < 1:
            return {'error': 'Datos incompletos'}, 400
        with database.get_db() as session:
            try:
                # Dedup por (modulo_id, ean_pack): el mismo EAN puede estar
                # en distintos módulos.
                q = session.query(ModuloPack).filter_by(ean_pack=ean_pack)
                if modulo_id is not None:
                    q = q.filter_by(modulo_id=int(modulo_id) if modulo_id else None)
                existing = q.first()
                if existing:
                    existing.ean_unidad = ean_unidad
                    existing.cantidad = cantidad
                    existing.descripcion = descripcion or existing.descripcion
                    if modulo_id is not None:
                        existing.modulo_id = int(modulo_id) if modulo_id else None
                else:
                    session.add(ModuloPack(ean_pack=ean_pack, ean_unidad=ean_unidad,
                                           cantidad=cantidad, descripcion=descripcion,
                                           modulo_id=int(modulo_id) if modulo_id else None))
                session.commit()
                return {'ok': True}
            except Exception as e:
                session.rollback()
                return {'error': str(e)}, 500

    @app.route('/modulo-pack/<int:pack_id>/update', methods=['POST'])
    def modulo_pack_update(pack_id):
        with database.get_db() as session:
            try:
                data = request.get_json(silent=True) or {}
                mp = session.get(ModuloPack, pack_id)
                if not mp:
                    return jsonify({'error': 'No encontrado'}), 404
                if 'ean_pack' in data:
                    mp.ean_pack = str(data['ean_pack']).strip()
                if 'descripcion' in data:
                    mp.descripcion = str(data['descripcion']).strip() or None
                if 'ean_unidad' in data:
                    mp.ean_unidad = str(data['ean_unidad']).strip()
                if 'cantidad' in data:
                    mp.cantidad = int(data['cantidad'])
                # Si quedó una equivalencia válida (pack distinto a unidad),
                # persistirla en pack_equivalencias para auto-aplicar en
                # imports futuros de cualquier módulo.
                if mp.ean_pack and mp.ean_unidad and mp.ean_pack != mp.ean_unidad:
                    pe = (session.query(database.PackEquivalencia)
                          .filter_by(ean_pack=mp.ean_pack[:30]).first())
                    if not pe:
                        session.add(database.PackEquivalencia(
                            ean_pack=mp.ean_pack[:30],
                            ean_unidad=mp.ean_unidad[:30],
                            cantidad=mp.cantidad or 1,
                            desc_pack=(mp.descripcion or '')[:255] or None,
                            aprendido_de=mp.modulo_id,
                        ))
                    else:
                        pe.ean_unidad = mp.ean_unidad[:30]
                        if mp.cantidad and mp.cantidad >= 1:
                            pe.cantidad = mp.cantidad
                if 'cant_modulo' in data:
                    mp.cant_modulo = int(data['cant_modulo']) if data['cant_modulo'] is not None else None
                if 'desc_pct' in data:
                    mp.desc_pct = float(data['desc_pct']) if data['desc_pct'] is not None else None
                session.commit()
                return jsonify({'ok': True})
            except Exception as e:
                session.rollback()
                return jsonify({'error': str(e)}), 500

    @app.route('/modulo-pack/<int:pack_id>/delete', methods=['POST'])
    def modulo_pack_delete(pack_id):
        with database.get_db() as session:
            mp = session.get(ModuloPack, pack_id)
            if mp:
                session.delete(mp)
                session.commit()
            return {'ok': True}
