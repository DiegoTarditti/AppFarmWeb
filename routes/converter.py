"""Converter routes: herramienta de aprendizaje de parsers de facturas.

Flujo:
  1. Usuario sube PDF
  2. Se detecta proveedor por CUIT/razón social
  3. Si existe parser del proveedor → se prueba. Si OK, se ofrece usarlo ya.
  4. Si falla o no hay parser → modo aprendizaje: seleccionás fragmentos de una fila
     y el sistema infiere una regex. Se ofrece guardar como parser del proveedor.
"""

import os
import json
import re
from flask import render_template, request, redirect, url_for, flash, jsonify, send_file
from werkzeug.utils import secure_filename
from helpers import CONVERTER_DIR, _build_item_pattern, PARSERS_FOLDER
import database
from data_extract import extract_provider_info_from_pdf, parse_invoice_pdf


def _converter_meta_path(fname):
    return os.path.join(CONVERTER_DIR, fname + '.meta.json')


def _converter_read_meta(fname):
    p = _converter_meta_path(fname)
    if os.path.exists(p):
        try:
            with open(p, 'r', encoding='utf-8') as fh:
                return json.load(fh) or {}
        except Exception:
            return {}
    return {}


def _converter_write_meta(fname, data):
    with open(_converter_meta_path(fname), 'w', encoding='utf-8') as fh:
        json.dump(data, fh)


def _detectar_proveedor(info):
    """Intenta encontrar un proveedor en DB por CUIT primero, luego por razón social."""
    from database import Provider
    cuit = (info.get('cuit') or '').strip()
    razon = (info.get('razon_social') or '').strip()
    with database.get_db() as session:
        prov = None
        if cuit:
            prov = session.query(Provider).filter_by(cuit=cuit).first()
        if not prov and razon:
            prov = session.query(Provider).filter(
                Provider.razon_social.ilike(f'%{razon}%')
            ).first()
        if not prov:
            return None
        return {
            'id': prov.id,
            'razon_social': prov.razon_social,
            'cuit': prov.cuit,
            'parser_file': prov.parser_file,
        }


def _probar_parser(parser_file, pdf_path):
    """Prueba un parser sobre un PDF. Devuelve dict con éxito o error."""
    if not parser_file:
        return {'ok': False, 'error': 'Sin parser configurado'}
    try:
        data = parse_invoice_pdf(pdf_path, parser_file)
        items = data.get('items') or []
        def _f(v):
            if v is None: return None
            try: return float(v)
            except Exception: return None
        return {
            'ok': len(items) > 0,
            'n_items': len(items),
            'primeros': [
                {'codigo_barra': i.get('codigo_barra'), 'descripcion': i.get('descripcion'),
                 'cantidad': i.get('cantidad'), 'importe': i.get('importe')}
                for i in items[:5]
            ],
            'items_full': [{
                'codigo_barra':    i.get('codigo_barra'),
                'descripcion':     i.get('descripcion'),
                'cantidad':        i.get('cantidad'),
                'precio_publico':  _f(i.get('precio_publico')),
                'dto':             _f(i.get('dto')),
                'precio_unitario': _f(i.get('precio_unitario')),
                'importe':         _f(i.get('importe')),
                'lote':            i.get('lote'),
                'vencimiento':     i.get('vencimiento'),
            } for i in items],
            'numero_factura': data.get('numero_factura'),
            'fecha': str(data.get('fecha') or ''),
            'total': _f(data.get('total')),
            'monto_exento':  _f(data.get('monto_exento')),
            'monto_gravado': _f(data.get('monto_gravado')),
            'iva':           _f(data.get('iva')),
            'percepciones':  _f(data.get('percepciones')),
        }
    except Exception as e:
        return {'ok': False, 'error': str(e)}


def init_app(app):

    @app.route('/converter', methods=['GET'])
    def converter_index():
        files = []
        try:
            for fn in sorted(os.listdir(CONVERTER_DIR), reverse=True):
                if fn.endswith('.meta.json'):
                    continue
                p = os.path.join(CONVERTER_DIR, fn)
                if os.path.isfile(p):
                    meta = _converter_read_meta(fn)
                    files.append({
                        'name': fn,
                        'size': os.path.getsize(p),
                        'proveedor': meta.get('proveedor_razon', ''),
                        'cuit': meta.get('proveedor_cuit', ''),
                    })
        except Exception:
            pass
        return render_template('converter_index.html', files=files[:20])

    @app.route('/converter/upload', methods=['POST'])
    def converter_upload():
        """Guarda el PDF y devuelve el token. El análisis se dispara desde JS
        en pasos (para mostrar animaciones)."""
        import uuid as _uuid
        f = request.files.get('document')
        if not f or not f.filename:
            if request.headers.get('Accept', '').startswith('application/json'):
                return jsonify({'error': 'Elegí un archivo PDF.'}), 400
            flash('Elegí un archivo PDF.')
            return redirect(url_for('converter_index'))
        ext = os.path.splitext(f.filename)[1].lower() or '.pdf'
        if ext != '.pdf':
            return jsonify({'error': 'Por ahora sólo PDF.'}), 400
        token = _uuid.uuid4().hex[:12]
        fname = token + '_' + secure_filename(f.filename)
        path = os.path.join(CONVERTER_DIR, fname)
        f.save(path)
        return jsonify({'ok': True, 'token': fname, 'size': os.path.getsize(path)})

    @app.route('/converter/<token>/analizar', methods=['POST'])
    def converter_analizar(token):
        """Ejecuta el análisis completo (encabezado + proveedor + parser test) y devuelve JSON.

        Si se llama varias veces, el cliente puede animar los pasos uno por uno.
        Acá ejecutamos todo junto porque en local es rápido."""
        safe = secure_filename(token)
        path = os.path.join(CONVERTER_DIR, safe)
        if not os.path.exists(path):
            return jsonify({'error': 'Documento no encontrado'}), 404

        # Paso 1: extraer info del encabezado
        try:
            info = extract_provider_info_from_pdf(path)
        except Exception as e:
            info = {'razon_social': None, 'cuit': None, 'error_header': str(e)}

        # Paso 2: matchear con DB de proveedores
        proveedor = _detectar_proveedor(info) if (info.get('cuit') or info.get('razon_social')) else None

        # Paso 3: probar parser si existe
        prueba = None
        if proveedor and proveedor.get('parser_file'):
            prueba = _probar_parser(proveedor['parser_file'], path)

        # Paso 4: extraer primeras líneas del PDF para mostrar
        import pdfplumber
        n_lineas = 0
        try:
            with pdfplumber.open(path) as pdf:
                for page in pdf.pages:
                    n_lineas += len((page.extract_text() or '').split('\n'))
        except Exception:
            pass

        meta = {
            'proveedor_razon': info.get('razon_social') or (proveedor or {}).get('razon_social', ''),
            'proveedor_cuit': info.get('cuit') or (proveedor or {}).get('cuit', ''),
            'proveedor_id': proveedor['id'] if proveedor else None,
            'parser_file': (proveedor or {}).get('parser_file'),
            'fecha_detectada': info.get('fecha') or '',
            'numero_detectado': info.get('numero') or '',
        }
        _converter_write_meta(safe, meta)

        return jsonify({
            'ok': True,
            'token': safe,
            'info': info,
            'proveedor': proveedor,
            'prueba': prueba,
            'n_lineas': n_lineas,
            'redirect': url_for('converter_detectar', token=safe),
        })

    @app.route('/converter/<token>/detectar', methods=['GET'])
    def converter_detectar(token):
        """Pantalla de decisión: mostrar qué se detectó y ofrecer caminos."""
        safe = secure_filename(token)
        path = os.path.join(CONVERTER_DIR, safe)
        if not os.path.exists(path):
            flash('Documento no encontrado.')
            return redirect(url_for('converter_index'))

        meta = _converter_read_meta(safe)
        proveedor = None
        if meta.get('proveedor_id'):
            with database.get_db() as session:
                prov = session.get(database.Provider, meta['proveedor_id'])
                if prov:
                    proveedor = {
                        'id': prov.id, 'razon_social': prov.razon_social,
                        'cuit': prov.cuit, 'parser_file': prov.parser_file,
                    }

        prueba = None
        if proveedor and proveedor.get('parser_file'):
            prueba = _probar_parser(proveedor['parser_file'], path)

        # Texto del PDF para referencia (primeras 40 líneas)
        import pdfplumber
        pdf_lines = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                for line in (page.extract_text() or '').split('\n'):
                    pdf_lines.append(line)
                    if len(pdf_lines) >= 60:
                        break
                if len(pdf_lines) >= 60:
                    break

        return render_template('converter_detectar.html',
                               token=safe, filename=safe,
                               info={'razon_social': meta.get('proveedor_razon'),
                                     'cuit': meta.get('proveedor_cuit')},
                               proveedor=proveedor, prueba=prueba,
                               pdf_lines=pdf_lines)

    @app.route('/converter/<token>/verify', methods=['GET'])
    def converter_verify(token):
        """Pantalla de verificación: muestra todos los items parseados con
        validación matemática por fila (cant × unit = importe, pub × (1-dto%) = unit)
        y permite editarlos antes de confirmar el import."""
        safe = secure_filename(token)
        path = os.path.join(CONVERTER_DIR, safe)
        if not os.path.exists(path):
            flash('Documento no encontrado.')
            return redirect(url_for('converter_index'))

        meta = _converter_read_meta(safe)
        proveedor = None
        if meta.get('proveedor_id'):
            with database.get_db() as session:
                prov = session.get(database.Provider, meta['proveedor_id'])
                if prov:
                    proveedor = {'id': prov.id, 'razon_social': prov.razon_social,
                                 'cuit': prov.cuit, 'parser_file': prov.parser_file}

        if not proveedor or not proveedor.get('parser_file'):
            flash('Este documento no tiene parser configurado — no hay nada para verificar.')
            return redirect(url_for('converter_detectar', token=safe))

        prueba = _probar_parser(proveedor['parser_file'], path)
        if not prueba.get('ok'):
            flash('El parser no devolvió ítems para verificar.')
            return redirect(url_for('converter_detectar', token=safe))

        return render_template('converter_verify.html',
                               token=safe, filename=safe,
                               proveedor=proveedor, prueba=prueba)

    @app.route('/converter/<token>/verify/import', methods=['POST'])
    def converter_verify_import(token):
        """Recibe los items (posiblemente editados) y crea la factura real."""
        body = request.get_json(silent=True) or {}
        header = body.get('header', {}) or {}
        rows   = body.get('rows', [])
        tipo   = body.get('tipo_comprobante', 'FAC')
        try:
            inv_id, mensaje = _guardar_factura_desde_aprendizaje(token, header, rows, tipo)
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        except Exception as e:
            return jsonify({'error': f'Error al guardar: {e}'}), 500
        return jsonify({'ok': True, 'invoice_id': inv_id,
                        'url_factura': url_for('invoice_items', invoice_id=inv_id)})

    @app.route('/converter/<token>/auto', methods=['GET'])
    def converter_auto(token):
        """Intenta detectar una tabla con pdfplumber.extract_tables()."""
        import pdfplumber as _plumber
        from collections import defaultdict as _dd
        safe = secure_filename(token)
        path = os.path.join(CONVERTER_DIR, safe)
        if not os.path.exists(path):
            flash('Documento no encontrado.')
            return redirect(url_for('converter_index'))

        tables = []
        with _plumber.open(path) as pdf:
            for page in pdf.pages:
                for tbl in (page.extract_tables() or []):
                    if tbl and len(tbl) > 2:
                        tables.append(tbl)

        if tables:
            best = max(tables, key=lambda t: len(t))
            rows = best
            source = 'Tabla detectada automáticamente'
        else:
            all_words = []
            with _plumber.open(path) as pdf:
                for page in pdf.pages:
                    all_words.extend(page.extract_words(x_tolerance=3, y_tolerance=3) or [])
            rows_dict = _dd(list)
            for w in all_words:
                y_key = round(w['top'] / 4) * 4
                rows_dict[y_key].append(w)
            rows = []
            for y in sorted(rows_dict.keys()):
                row = sorted(rows_dict[y], key=lambda w: w['x0'])
                rows.append([w['text'] for w in row])
            source = 'Palabras agrupadas por posición Y'

        meta = _converter_read_meta(safe)
        return render_template('converter_auto.html',
                               token=safe, filename=safe,
                               rows=rows, source=source,
                               proveedor_razon=meta.get('proveedor_razon', ''))

    @app.route('/converter/<token>/delete', methods=['POST'])
    def converter_delete(token):
        safe = secure_filename(token)
        path = os.path.join(CONVERTER_DIR, safe)
        if os.path.exists(path):
            try:
                os.remove(path)
                mp = _converter_meta_path(safe)
                if os.path.exists(mp):
                    os.remove(mp)
                flash('Documento eliminado.')
            except Exception as e:
                flash(f'Error al eliminar: {e}')
        return redirect(url_for('converter_index'))

    @app.route('/converter/<token>/pick', methods=['GET'])
    def converter_pick(token):
        """Modo aprendizaje: seleccionás partes de una fila ejemplo y el sistema infiere regex."""
        import pdfplumber as _plumber
        safe = secure_filename(token)
        path = os.path.join(CONVERTER_DIR, safe)
        if not os.path.exists(path):
            flash('Documento no encontrado.')
            return redirect(url_for('converter_index'))
        pdf_text = ''
        with _plumber.open(path) as pdf:
            for page in pdf.pages:
                pdf_text += (page.extract_text() or '') + '\n\n'
        meta = _converter_read_meta(safe)
        # Reejecutar la detección siempre, para tener datos frescos
        # (evita depender del meta.json que puede ser de una versión anterior del extractor)
        try:
            info_fresh = extract_provider_info_from_pdf(path)
        except Exception:
            info_fresh = {}

        header_precargado = {}
        razon = info_fresh.get('razon_social') or meta.get('proveedor_razon')
        cuit = info_fresh.get('cuit') or meta.get('proveedor_cuit')
        fecha = info_fresh.get('fecha') or meta.get('fecha_detectada')
        numero = info_fresh.get('numero') or meta.get('numero_detectado')
        if razon:  header_precargado['razon_social'] = razon
        if cuit:   header_precargado['cuit']         = cuit
        if fecha:  header_precargado['fecha']        = fecha
        if numero: header_precargado['numero']       = numero

        # Actualizar meta con lo fresco por si acaso
        meta.update({
            'proveedor_razon': razon or '',
            'proveedor_cuit': cuit or '',
            'fecha_detectada': fecha or '',
            'numero_detectado': numero or '',
        })
        _converter_write_meta(safe, meta)

        return render_template('converter_pick.html',
                               token=safe, pdf_text=pdf_text, filename=safe,
                               proveedor_razon=razon or '',
                               proveedor_id=meta.get('proveedor_id'),
                               header_precargado=header_precargado)

    @app.route('/converter/<token>/infer', methods=['POST'])
    def converter_infer(token):
        """Dado un ejemplo de línea + selecciones, infiere el patrón y aplica a todo el doc."""
        body = request.get_json(silent=True) or {}
        example_line = body.get('example_line', '')
        selections = body.get('selections', [])

        safe = secure_filename(token)
        path = os.path.join(CONVERTER_DIR, safe)
        if not os.path.exists(path):
            return jsonify({'error': 'Documento no encontrado'}), 404
        if not example_line or not selections:
            return jsonify({'error': 'Faltan datos'}), 400

        pattern, fields, base_fields, _base = _build_item_pattern(example_line, selections)

        import pdfplumber as _plumber
        pdf_text = ''
        with _plumber.open(path) as pdf:
            for page in pdf.pages:
                pdf_text += (page.extract_text() or '') + '\n'

        # Cortar el texto en marcadores de sección "sin stock" para no parsear faltantes
        items_text = pdf_text
        for marker in (r'\*\*\*\s*PRODUCTOS\s+EN\s+FALTA',):
            m = re.search(marker, items_text)
            if m:
                items_text = items_text[:m.start()]

        rx = re.compile(pattern, re.MULTILINE)
        rows = []
        matched_spans = set()
        for m in rx.finditer(items_text):
            matched_spans.add(m.start())
            row = {b: [] for b in base_fields}
            for i, f in enumerate(fields):
                val = m.group(i + 1)
                if val:
                    row[_base(f)].append(val)
            rows.append({b: re.sub(r'\s+', ' ', ' '.join(row[b]).strip()) for b in base_fields})

        # Fallback: filas "gravadas" con solo 5 columnas (ean cant desc unit importe)
        # Sólo si el pattern primario tenía al menos los campos mínimos.
        min_ok = all(b in base_fields for b in ('codigo_barra', 'cantidad', 'descripcion', 'precio_unitario', 'importe'))
        if min_ok:
            GRAV = re.compile(
                r'^(\d{7,14})\s+(\d+)\s+(.+?)\s+([\d.]+,\d{2})\s+([\d.]+,\d{2})\s*$',
                re.MULTILINE
            )
            for m in GRAV.finditer(items_text):
                if m.start() in matched_spans:
                    continue
                extra = {b: '' for b in base_fields}
                extra['codigo_barra']    = m.group(1)
                extra['cantidad']        = m.group(2)
                extra['descripcion']     = re.sub(r'\s+(WEB|TRZ)\s*$', '', m.group(3)).strip()
                extra['precio_unitario'] = m.group(4)
                extra['importe']         = m.group(5)
                rows.append(extra)

        return jsonify({'pattern': pattern, 'fields': base_fields, 'rows': rows})

    @app.route('/converter/<token>/guardar-parser', methods=['POST'])
    def converter_guardar_parser(token):
        """Guarda el patrón aprendido como parser persistente del proveedor.

        Body JSON: {
            pattern: regex inferida,
            fields: ['codigo_barra', 'cantidad', ...],
            provider_id: id del proveedor (opcional, si no se crea uno)
            provider_name: nombre si es nuevo proveedor
        }
        """
        body = request.get_json(silent=True) or {}
        pattern = body.get('pattern', '').strip()
        fields = body.get('fields', [])
        provider_id = body.get('provider_id')
        provider_name = (body.get('provider_name') or '').strip()
        cuit = (body.get('cuit') or '').strip()

        if not pattern or not fields:
            return jsonify({'error': 'pattern y fields son obligatorios'}), 400

        with database.get_db() as session:
            prov = None
            if provider_id:
                prov = session.get(database.Provider, int(provider_id))
            if not prov and provider_name:
                prov = session.query(database.Provider).filter(
                    database.Provider.razon_social.ilike(f'%{provider_name}%')
                ).first()
                if not prov:
                    prov = database.Provider(razon_social=provider_name, cuit=cuit or None)
                    session.add(prov)
                    session.commit()
                    session.refresh(prov)
            if not prov:
                return jsonify({'error': 'Falta proveedor (id o nombre)'}), 400

            from helpers import _make_parser_slug, _ensure_parser_file
            parser_slug = prov.parser_file
            if not parser_slug:
                parser_slug = _make_parser_slug(prov.razon_social)
                _ensure_parser_file(parser_slug, prov.razon_social)
                prov.parser_file = parser_slug

            # Escribir el parser aprendido
            parser_path = os.path.join(PARSERS_FOLDER, parser_slug + '.py')
            codigo = _generar_codigo_parser(pattern, fields, prov.razon_social, prov.cuit or '')
            with open(parser_path, 'w', encoding='utf-8') as fh:
                fh.write(codigo)
            session.commit()
            return jsonify({
                'ok': True,
                'provider_id': prov.id,
                'parser_file': parser_slug,
                'mensaje': f'Parser guardado en parsers/{parser_slug}.py',
            })

    @app.route('/converter/<token>/guardar-factura', methods=['POST'])
    def converter_guardar_factura(token):
        """Guarda los datos aprendidos como una Invoice (más ítems)."""
        body = request.get_json(silent=True) or {}
        header = body.get('header', {}) or {}
        rows = body.get('rows', [])
        tipo_comprobante = body.get('tipo_comprobante', 'FAC')
        try:
            inv_id, mensaje = _guardar_factura_desde_aprendizaje(token, header, rows, tipo_comprobante)
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        except Exception as e:
            return jsonify({'error': f'Error al guardar: {e}'}), 500
        return jsonify({
            'ok': True, 'invoice_id': inv_id, 'mensaje': mensaje,
            'url_factura': url_for('invoice_items', invoice_id=inv_id),
        })

    @app.route('/converter/<token>/enviar-a-proceso', methods=['POST'])
    def converter_enviar_a_proceso(token):
        """Guarda la factura y crea un ProcesoCompra tipo droguería asociado."""
        from datetime import datetime
        from helpers import now_ar
        body = request.get_json(silent=True) or {}
        header = body.get('header', {}) or {}
        rows = body.get('rows', [])
        tipo_comprobante = body.get('tipo_comprobante', 'FAC')
        try:
            inv_id, _ = _guardar_factura_desde_aprendizaje(token, header, rows, tipo_comprobante)
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        except Exception as e:
            return jsonify({'error': f'Error al guardar factura: {e}'}), 500

        # Crear el proceso
        with database.get_db() as session:
            inv = session.get(database.Invoice, inv_id)
            partner_nombre = inv.proveedor_razon or 'Proveedor sin identificar'
            # Buscar si ya existe proveedor registrado
            prov = None
            if inv.proveedor_cuit:
                prov = session.query(database.Provider).filter_by(cuit=inv.proveedor_cuit).first()
            if not prov and inv.proveedor_razon:
                prov = session.query(database.Provider).filter(
                    database.Provider.razon_social.ilike(f'%{inv.proveedor_razon}%')
                ).first()

            proc = database.ProcesoCompra(
                tipo='drogueria',
                partner_id=prov.id if prov else None,
                partner_nombre=partner_nombre,
                estado='FACTURADO',
                factura_id=inv.id,
                factura_hecha_en=now_ar(),
                notas=f'Creado desde aprendizaje de parser ({token[:12]})',
            )
            session.add(proc)
            session.commit()
            proc_id = proc.id

        return jsonify({
            'ok': True, 'invoice_id': inv_id, 'proceso_id': proc_id,
            'url_proceso': url_for('proceso_detail', proceso_id=proc_id),
            'mensaje': f'Factura guardada y asociada al proceso #{proc_id}.',
        })

    @app.route('/converter/<token>/export', methods=['POST'])
    def converter_export(token):
        """Exporta las filas extraídas a XLSX."""
        import io as _io
        import openpyxl as _ox
        body = request.get_json(silent=True) or {}
        rows = body.get('rows', [])
        header = body.get('header', {}) or {}
        fields = body.get('fields', [])
        if not rows:
            return jsonify({'error': 'No hay filas para exportar'}), 400

        wb = _ox.Workbook()
        ws = wb.active
        ws.title = 'Datos'

        r = 1
        if header:
            for k, v in header.items():
                ws.cell(row=r, column=1, value=k)
                ws.cell(row=r, column=2, value=v)
                r += 1
            r += 1

        cols = fields or (list(rows[0].keys()) if rows else [])
        for ci, c in enumerate(cols, start=1):
            ws.cell(row=r, column=ci, value=c)
        r += 1
        for row in rows:
            for ci, c in enumerate(cols, start=1):
                ws.cell(row=r, column=ci, value=row.get(c, ''))
            r += 1

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        out_name = os.path.splitext(token)[0] + '.xlsx'
        return send_file(buf, as_attachment=True, download_name=out_name,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _guardar_factura_desde_aprendizaje(token, header, rows, tipo_comprobante='FAC'):
    """Crea Invoice + InvoiceItems a partir de los datos aprendidos en el converter.

    Retorna (invoice_id, mensaje) o lanza ValueError con el detalle.
    """
    from datetime import datetime, date
    from helpers import UPLOAD_FOLDER
    import shutil
    from werkzeug.utils import secure_filename as _sf

    safe = _sf(token)
    path = os.path.join(CONVERTER_DIR, safe)
    if not os.path.exists(path):
        raise ValueError('Documento no encontrado')
    if not rows:
        raise ValueError('No hay ítems para guardar')

    tipo_comprobante = (tipo_comprobante or 'FAC').upper()
    if tipo_comprobante not in ('FAC', 'NCR'):
        tipo_comprobante = 'FAC'

    def parse_num(s):
        if s is None:
            return None
        t = str(s).strip()
        if not t:
            return None
        try:
            # Formato arg: 1.234,56 o 1234,56
            if ',' in t and re.search(r',\d{1,3}$', t):
                return float(t.replace('.', '').replace(',', '.'))
            return float(t.replace(',', '.'))
        except ValueError:
            return None

    def parse_int(s):
        n = parse_num(s)
        return int(n) if n is not None else 0

    # Parsear fecha
    fecha = date.today()
    if header.get('fecha'):
        for fmt in ('%d/%m/%Y', '%d/%m/%y', '%d-%m-%Y'):
            try:
                fecha = datetime.strptime(str(header['fecha']).strip(), fmt).date()
                break
            except ValueError:
                continue

    # Límites del DB: DECIMAL(14,2) → max ≈ 10^12. Rechazamos filas con
    # valores absurdos antes de que Postgres explote.
    MAX_VAL = 10**11  # 100 mil millones, suficiente para cualquier precio real
    items_data = []
    total_items = 0.0
    rows_malas = []
    for i, r in enumerate(rows, 1):
        importe = parse_num(r.get('importe')) or 0
        unit    = parse_num(r.get('precio_unitario'))
        cant_n  = parse_int(r.get('cantidad'))
        dto_n   = parse_num(r.get('dto'))

        problemas = []
        if unit is not None and abs(unit) > MAX_VAL:
            problemas.append(f'precio_unitario fuera de rango ({unit:.2f})')
        if abs(importe) > MAX_VAL:
            problemas.append(f'importe fuera de rango ({importe:.2f})')
        if dto_n is not None and (dto_n < -100 or dto_n > 100):
            problemas.append(f'dto fuera de rango ({dto_n})')
        # sanity: cant × unit vs importe (tolerancia generosa 2%)
        if cant_n and unit is not None and importe:
            calc = cant_n * unit
            if abs(calc - importe) > max(1.0, abs(importe) * 0.02):
                problemas.append(f'cant × unit = {calc:.2f} ≠ importe {importe:.2f}')

        if problemas:
            rows_malas.append((i, (r.get('codigo_barra') or '').strip(),
                               (r.get('descripcion') or '').strip()[:50], problemas))
            continue

        total_items += importe
        items_data.append({
            'codigo_barra': (r.get('codigo_barra') or '').strip()[:20],
            'cantidad': cant_n,
            'descripcion': (r.get('descripcion') or '').strip()[:150],
            'precio_unitario': unit,
            'dto': dto_n,
            'importe': importe,
            'lote': (r.get('lote') or '').strip()[:30],
            'vencimiento': (r.get('vencimiento') or '').strip()[:20],
        })

    if rows_malas:
        detalle = '; '.join(
            f'fila {i} ({cb or "s/cod"} · {desc}): {", ".join(p)}'
            for i, cb, desc, p in rows_malas[:5]
        )
        more = f' y {len(rows_malas) - 5} más' if len(rows_malas) > 5 else ''
        raise ValueError(
            f'{len(rows_malas)} fila(s) con datos inválidos — '
            f'corregilas o descartalas antes de guardar. Ej: {detalle}{more}'
        )

    # Desglose fiscal del header (exento/gravado/iva/percepciones/otros)
    h_exento  = parse_num(header.get('monto_exento'))
    h_gravado = parse_num(header.get('monto_gravado'))
    h_iva105  = parse_num(header.get('iva_105'))
    h_iva21   = parse_num(header.get('iva_21'))
    h_percep  = parse_num(header.get('percepciones'))
    h_otros   = parse_num(header.get('otros'))

    total_header = parse_num(header.get('total'))
    # Prioridad para el total: suma desglose (si hay al menos 2 campos) > header.total > suma items
    desglose_sum = None
    desglose_campos = [h_exento, h_gravado, h_iva105, h_iva21, h_percep, h_otros]
    if sum(1 for v in desglose_campos if v is not None) >= 2:
        desglose_sum = sum((v or 0) for v in desglose_campos)
    total = desglose_sum if desglose_sum else (total_header if total_header else total_items)
    sign = -1 if tipo_comprobante == 'NCR' else 1

    # Copiar PDF al upload folder para que sea accesible desde invoice_items
    pdf_dest_name = safe
    pdf_dest = os.path.join(UPLOAD_FOLDER, pdf_dest_name)
    if not os.path.exists(pdf_dest):
        try:
            shutil.copy(path, pdf_dest)
        except Exception:
            pass

    with database.get_db() as session:
        inv = database.Invoice(
            numero_factura=((header.get('numero') or 'SIN_NUMERO').strip())[:20],
            fecha=fecha,
            proveedor_razon=((header.get('razon_social') or '').strip() or None),
            proveedor_cuit=((header.get('cuit') or '').strip() or None),
            tipo_comprobante=tipo_comprobante,
            total=(total or 0) * sign,
            total_articulos=len(items_data),
            total_unidades=sum(i['cantidad'] for i in items_data),
            pdf_filename=pdf_dest_name,
            monto_exento  = (h_exento  * sign) if h_exento  is not None else None,
            monto_gravado = (h_gravado * sign) if h_gravado is not None else None,
            iva_105       = (h_iva105  * sign) if h_iva105  is not None else None,
            iva_21        = (h_iva21   * sign) if h_iva21   is not None else None,
            percepciones  = (h_percep  * sign) if h_percep  is not None else None,
            otros         = (h_otros   * sign) if h_otros   is not None else None,
        )
        session.add(inv)
        session.flush()
        for it in items_data:
            session.add(database.InvoiceItem(
                factura_id=inv.id,
                codigo_barra=it['codigo_barra'] or None,
                cantidad=it['cantidad'],
                descripcion=it['descripcion'],
                precio_unitario=it['precio_unitario'],
                dto=it['dto'],
                importe=(it['importe'] or 0) * sign,
                lote=it['lote'] or None,
                vencimiento=it['vencimiento'] or None,
            ))
        session.commit()
        return inv.id, f'Factura {inv.numero_factura} guardada con {len(items_data)} ítems.'


def _generar_codigo_parser(pattern, fields, razon_social, cuit):
    """Genera el código Python del parser a partir del patrón aprendido."""
    fields_repr = ', '.join(repr(f) for f in fields)
    return f'''"""Parser auto-generado para: {razon_social}
CUIT: {cuit or "—"}

Creado desde el modo aprendizaje del conversor.
Si el layout del proveedor cambia, reentrenar el patrón desde /converter.
"""
import re
import pdfplumber
from datetime import datetime


PATTERN = r"""{pattern}"""
FIELDS = [{fields_repr}]


def _to_float(s):
    """Convierte formato argentino '1.234,56' a float 1234.56."""
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    try:
        return float(s.replace('.', '').replace(',', '.'))
    except Exception:
        return None


def _to_int(s):
    try:
        return int(float(str(s).replace('.', '').replace(',', '.')))
    except Exception:
        return 0


def parse_invoice_pdf(pdf_path):
    pages_text = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            pages_text.append(page.extract_text() or '')
    full_text = '\\n'.join(pages_text)

    # Encabezado genérico
    numero_m = re.search(r'(?:FACTURA|REMITO|N[º°])\\s*[:\\s]*(\\S+)', full_text)
    fecha_m = re.search(r'(?:FECHA|Fecha)[:\\s]*(\\d{{2}}/\\d{{2}}/\\d{{4}})', full_text)
    numero_factura = numero_m.group(1) if numero_m else 'SIN_NUMERO'
    fecha = (datetime.strptime(fecha_m.group(1), '%d/%m/%Y').date()
             if fecha_m else datetime.today().date())

    # Ítems desde el patrón aprendido
    rx = re.compile(PATTERN, re.MULTILINE)
    items = []
    for m in rx.finditer(full_text):
        row = {{}}
        for i, f in enumerate(FIELDS):
            base = f.rstrip('0123456789_')
            val = m.group(i + 1) or ''
            row.setdefault(base, []).append(val)
        joined = {{b: re.sub(r'\\s+', ' ', ' '.join(v).strip()) for b, v in row.items()}}
        items.append({{
            'codigo_barra': joined.get('codigo_barra', ''),
            'cantidad': _to_int(joined.get('cantidad', 0)),
            'descripcion': joined.get('descripcion', ''),
            'precio_unitario': _to_float(joined.get('precio_unitario')),
            'importe': _to_float(joined.get('importe')) or 0,
        }})

    total = sum((it.get('importe') or 0) for it in items)

    return {{
        'numero_factura': numero_factura,
        'fecha': fecha,
        'proveedor_razon': {repr(razon_social)},
        'proveedor_cuit': {repr(cuit) if cuit else 'None'},
        'proveedor_domicilio': None,
        'cliente_codigo': None,
        'cliente_razon': None,
        'total': total,
        'total_articulos': len(items),
        'items': items,
    }}
'''
