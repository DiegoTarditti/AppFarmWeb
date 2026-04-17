"""Parser para Excel de ofertas Bernabó (Venta Directa / Venta Directa Enero)."""

import openpyxl


def _cell(ws, row, col):
    """Valor de celda resolviendo celdas fusionadas."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return None


def _pct(val):
    """'20%' / 0.20 / 20 → 20.0  (porcentaje como float)."""
    if val is None:
        return None
    if isinstance(val, str):
        try:
            return float(val.strip().rstrip('%').replace(',', '.'))
        except ValueError:
            return None
    f = float(val)
    return round(f * 100, 2) if 0 < abs(f) <= 1 else f


def _int(val):
    if val is None:
        return None
    try:
        return int(float(str(val).strip().replace(',', '.')))
    except (ValueError, TypeError):
        return None


def _is_ean(val):
    s = str(val).strip().replace('.0', '')
    return s.isdigit() and 7 <= len(s) <= 14


# ─────────────────────────────────────────────────────────────────────────────

def parse_bernabo_ofertas(path):
    """
    Parsea un Excel de Venta Directa Bernabó.

    Retorna lista de dicts:
      ean, codigo, descripcion,
      unidades_minima (int|None),
      descuento_psl   (float %|None),
      rentabilidad    (float %|None),
      plazo_pago      (str|None),
      grupo_id        (int|None)   ← None = mínimo individual
                                      N    = mínimo compartido del grupo N
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # ── 1. Detectar fila de cabecera ────────────────────────────────────────
    col_grupo = col_ean = col_cod = col_prod = col_min = col_dto = col_rent = col_plazo = None
    header_row = None

    for ri in range(1, min(20, ws.max_row + 1)):
        row_texts = {}
        for ci in range(1, ws.max_column + 1):
            v = _cell(ws, ri, ci)
            if v:
                row_texts[ci] = str(v).lower().replace('\n', ' ').strip()

        combined = ' '.join(row_texts.values())
        if 'barra' in combined or ('codigo' in combined and 'producto' in combined):
            header_row = ri
            for ci, t in row_texts.items():
                if 'grupo' in t and not col_grupo:
                    col_grupo = ci
                elif 'barra' in t and not col_ean:
                    col_ean = ci
                elif t.startswith('cod') and col_ean and ci == col_ean + 1 and not col_cod:
                    col_cod = ci
                elif 'producto' in t and not col_prod:
                    col_prod = ci
                elif ('unidad' in t or 'minim' in t) and not col_min:
                    col_min = ci
                elif ('descuento' in t or 'psl' in t) and not col_dto:
                    col_dto = ci
                elif ('rentabilidad' in t or 'farmacia' in t) and not col_rent:
                    col_rent = ci
                elif 'plazo' in t and not col_plazo:
                    col_plazo = ci
            break

    if header_row is None or col_ean is None:
        raise ValueError('No se encontró la cabecera (columna "Código de barra")')

    if col_cod is None and col_prod and col_ean + 1 < col_prod:
        col_cod = col_ean + 1

    # ── 2. Leer filas crudas ────────────────────────────────────────────────
    # Tupla: (grupo_val, ean, cod, prod, vmin, vdto, vren, vpla)
    raw = []
    for ri in range(header_row + 1, ws.max_row + 1):
        grupo_val = _cell(ws, ri, col_grupo) if col_grupo else None
        ean  = str(_cell(ws, ri, col_ean)  or '').strip()
        cod  = str(_cell(ws, ri, col_cod)  or '').strip() if col_cod  else ''
        prod = str(_cell(ws, ri, col_prod) or '').strip() if col_prod else ''
        vmin = _cell(ws, ri, col_min)  if col_min  else None
        vdto = _cell(ws, ri, col_dto)  if col_dto  else None
        vren = _cell(ws, ri, col_rent) if col_rent else None
        vpla = str(_cell(ws, ri, col_plazo) or '').strip() if col_plazo else None
        raw.append((grupo_val, ean, cod, prod, vmin, vdto, vren, vpla or None))

    # ── 3. Procesar según si hay columna Grupo explícita ───────────────────
    if col_grupo is not None:
        return _parse_with_grupo_col(raw)
    else:
        return _parse_by_blank_rows(raw)


def _parse_with_grupo_col(raw):
    """Excel con columna Grupo explícita (col A con número de grupo o None)."""
    result = []

    # Recopilar el mínimo compartido de cada grupo (primer fila con datos)
    group_shared = {}
    for grupo_val, ean, cod, prod, vmin, vdto, vren, vpla in raw:
        if not _is_ean(ean):
            continue
        gid = _int(grupo_val)
        if gid is not None and gid not in group_shared:
            group_shared[gid] = {
                'min':   _int(vmin),
                'dto':   _pct(vdto),
                'rent':  _pct(vren),
                'plazo': vpla,
            }

    for grupo_val, ean, cod, prod, vmin, vdto, vren, vpla in raw:
        if not _is_ean(ean):
            continue
        gid = _int(grupo_val)
        if gid is not None:
            shared = group_shared.get(gid, {})
            result.append({
                'ean':             ean,
                'codigo':          cod,
                'descripcion':     prod,
                'unidades_minima': shared.get('min'),
                'descuento_psl':   shared.get('dto'),
                'rentabilidad':    shared.get('rent'),
                'plazo_pago':      shared.get('plazo'),
                'grupo_id':        gid,
            })
        else:
            result.append({
                'ean':             ean,
                'codigo':          cod,
                'descripcion':     prod,
                'unidades_minima': _int(vmin),
                'descuento_psl':   _pct(vdto),
                'rentabilidad':    _pct(vren),
                'plazo_pago':      vpla,
                'grupo_id':        None,
            })

    return result


def _parse_by_blank_rows(raw):
    """Fallback: Excel sin columna Grupo — detecta grupos por filas vacías."""
    # Agrupar por filas vacías (sin EAN ni producto)
    groups, cur = [], []
    for row in raw:
        blank = not row[1] and not row[3]
        if blank:
            if cur:
                groups.append(cur)
                cur = []
        else:
            cur.append(row)
    if cur:
        groups.append(cur)

    result = []
    grupo_counter = 0

    for group in groups:
        valid = [r for r in group if _is_ean(r[1])]
        if not valid:
            continue

        rows_with_min = [r for r in valid if r[4] is not None]
        # Mínimo compartido = múltiples productos y el mínimo aparece en menos filas
        is_grouped = len(valid) > 1 and len(rows_with_min) < len(valid)

        if is_grouped:
            grupo_counter += 1
            gid = grupo_counter
            ref = rows_with_min[0] if rows_with_min else valid[0]
            shared = {
                'min':   _int(ref[4]),
                'dto':   _pct(ref[5]),
                'rent':  _pct(ref[6]),
                'plazo': ref[7],
            }
        else:
            gid = None
            shared = None

        for _, ean, cod, prod, vmin, vdto, vren, vpla in valid:
            if shared:
                u, d, r, p = shared['min'], shared['dto'], shared['rent'], shared['plazo']
            else:
                u, d, r, p = _int(vmin), _pct(vdto), _pct(vren), vpla

            result.append({
                'ean':             ean,
                'codigo':          cod,
                'descripcion':     prod,
                'unidades_minima': u,
                'descuento_psl':   d,
                'rentabilidad':    r,
                'plazo_pago':      p,
                'grupo_id':        gid,
            })

    return result
