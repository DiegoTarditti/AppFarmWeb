"""
Parser para archivos de módulos de descuento.
Soporta: Excel (.xlsx / .xls), PDF, imágenes (.jpg, .png, etc.)

Estructura esperada (columnas):
  NOMBRE MODULO | CODIGO EAN | DESCRIPCION | CANT | DESC(%)

Devuelve:
  list of dict {
    'nombre': str,
    'items': list of {
        'ean': str, 'descripcion': str,
        'cantidad': int, 'descuento': float,
        'es_principal': bool
    }
  }
"""

import openpyxl
import pdfplumber

# Tonos amarillos que usa Excel para resaltar la fila principal
_YELLOW_RGBS = {
    'FFFFFF00', 'FF FFFF00', 'FFFF00', 'FFFFC000',
    'FFFFFF99', 'FFFFEB3B', 'FFF9A825',
}


def _is_yellow(cell) -> bool:
    try:
        fg = cell.fill.fgColor
        if fg.type == 'rgb':
            rgb = fg.rgb.upper().replace(' ', '')
            return rgb in _YELLOW_RGBS or rgb.endswith('FFFF00') or rgb.endswith('FFC000')
        if fg.type == 'theme':
            # theme 7 u 8 suele ser amarillo
            return fg.theme in (7, 8)
    except Exception:
        pass
    return False


def _safe_str(v) -> str:
    return str(v).strip() if v is not None else ''


def _safe_float(v) -> float:
    if v is None:
        return 0.0
    try:
        s = str(v).replace('%', '').replace(',', '.').strip()
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _safe_int(v) -> int:
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return 1


def _detect_columns(ws, header_row: int) -> dict:
    """Detecta índices (0-based) de cada columna por su encabezado."""
    col_map = {}
    for i, cell in enumerate(ws[header_row]):
        h = _safe_str(cell.value).upper()
        if not h:
            continue
        if 'MODULO' in h or ('NOMBRE' in h and 'MODULO' in h):
            col_map.setdefault('modulo', i)
        elif 'EAN' in h or ('CODIGO' in h and 'modulo' not in col_map.get('modulo', -1).__class__.__name__):
            col_map.setdefault('ean', i)
        elif 'CODIGO' in h:
            col_map.setdefault('ean', i)
        elif 'DESCRIPCION' in h or 'DESCRIPCIÓN' in h:
            col_map.setdefault('descripcion', i)
        elif 'CANT' in h:
            col_map.setdefault('cantidad', i)
        # DESC al final (evitar confundir con DESCRIPCION)
    # segunda pasada para DESC (descuento) — último campo con 'DESC' que no sea DESCRIPCION
    for i, cell in enumerate(ws[header_row]):
        h = _safe_str(cell.value).upper()
        if h == 'DESC' or (h.startswith('DESC') and 'RIPCION' not in h and 'RIPCIÓN' not in h):
            col_map['descuento'] = i
    return col_map


def _find_header_row(ws) -> int:
    """Busca la fila de encabezado (que contenga CODIGO o EAN)."""
    for i, row in enumerate(ws.iter_rows(max_row=10), 1):
        for cell in row:
            v = _safe_str(cell.value).upper()
            if 'CODIGO' in v or 'EAN' in v or 'MODULO' in v:
                return i
    return 1


def _parse_table_rows(rows: list) -> list:
    """
    Recibe filas como listas de strings (ya sea de Excel o de PDF).
    Detecta encabezado y arma la lista de módulos.
    """
    if not rows:
        return []

    # Detectar fila de encabezado (contiene CODIGO, EAN o MODULO)
    header_idx = 0
    for i, row in enumerate(rows):
        joined = ' '.join(str(c) for c in row if c).upper()
        if any(k in joined for k in ('CODIGO', 'EAN', 'MODULO')):
            header_idx = i
            break

    header = [str(c).upper().strip() if c else '' for c in rows[header_idx]]

    # Mapear columnas por nombre
    col = {}
    for i, h in enumerate(header):
        if not h:
            continue
        if 'MODULO' in h and 'modulo' not in col:
            col['modulo'] = i
        elif ('EAN' in h or ('CODIGO' in h and 'modulo' not in col)) and 'ean' not in col:
            col['ean'] = i
        elif 'CODIGO' in h and 'ean' not in col:
            col['ean'] = i
        elif ('DESCRIPCION' in h or 'DESCRIPCIÓN' in h) and 'descripcion' not in col:
            col['descripcion'] = i
        elif 'CANT' in h and 'cantidad' not in col:
            col['cantidad'] = i
    # DESC (descuento) — segunda pasada
    for i, h in enumerate(header):
        if h.startswith('DESC') and 'RIPCION' not in h and 'RIPCIÓN' not in h:
            col['descuento'] = i

    if 'ean' not in col and 'descripcion' not in col:
        raise ValueError('No se detectaron columnas EAN ni DESCRIPCION.')

    modulos = []
    current_nombre = None
    current_items = []

    for row in rows[header_idx + 1:]:
        # Normalizar longitud
        while len(row) <= max(col.values(), default=0):
            row.append(None)

        nombre_val = _safe_str(row[col['modulo']]) if 'modulo' in col else ''
        ean_val    = _safe_str(row[col['ean']]) if 'ean' in col else ''
        desc_val   = _safe_str(row[col['descripcion']]) if 'descripcion' in col else ''
        cant_val   = _safe_int(row[col['cantidad']]) if 'cantidad' in col else 1
        dto_val    = _safe_float(row[col['descuento']]) if 'descuento' in col else 0.0

        if not ean_val and not desc_val:
            continue

        if nombre_val:
            if current_nombre is not None and current_items:
                modulos.append({'nombre': current_nombre, 'items': current_items})
            current_nombre = nombre_val
            current_items = []

        if current_nombre is None:
            current_nombre = 'SIN NOMBRE'

        es_principal = len(current_items) == 0
        current_items.append({
            'ean': ean_val,
            'descripcion': desc_val,
            'cantidad': cant_val,
            'descuento': dto_val,
            'es_principal': es_principal,
        })

    if current_nombre and current_items:
        modulos.append({'nombre': current_nombre, 'items': current_items})

    return modulos


def parse_descuento_pdf(path: str) -> list:
    """
    Extrae tablas de un PDF usando pdfplumber y las parsea con la misma lógica.
    """
    all_rows = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    cleaned = [str(c).strip() if c else '' for c in row]
                    if any(cleaned):
                        all_rows.append(cleaned)

    if not all_rows:
        raise ValueError('No se encontraron tablas en el PDF.')

    return _parse_table_rows(all_rows)


def parse_descuento_xlsx(path: str) -> list:
    """
    Lee el Excel y devuelve la lista de módulos con sus artículos.
    Lanza ValueError si no puede detectar columnas mínimas.
    Mantiene detección de fila amarilla (es_principal) que no tiene el parser genérico.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    col = _detect_columns(ws, header_row)

    if 'ean' not in col and 'descripcion' not in col:
        raise ValueError('No se detectaron columnas EAN ni DESCRIPCION en el archivo.')

    modulos = []
    current_nombre = None
    current_items = []

    for row in ws.iter_rows(min_row=header_row + 1):
        vals = [cell.value for cell in row]

        if all(v is None or _safe_str(v) == '' for v in vals):
            continue

        nombre_cell = row[col['modulo']] if 'modulo' in col else None
        nombre_val  = _safe_str(nombre_cell.value if nombre_cell else None)
        ean_val     = _safe_str(vals[col['ean']]) if 'ean' in col else ''
        desc_val    = _safe_str(vals[col['descripcion']]) if 'descripcion' in col else ''
        cant_val    = _safe_int(vals[col['cantidad']]) if 'cantidad' in col else 1
        dto_val     = _safe_float(vals[col['descuento']]) if 'descuento' in col else 0.0

        if not ean_val and not desc_val:
            continue

        if nombre_val:
            if current_nombre is not None and current_items:
                modulos.append({'nombre': current_nombre, 'items': current_items})
            current_nombre = nombre_val
            current_items = []

        if current_nombre is None:
            current_nombre = 'SIN NOMBRE'

        desc_cell = row[col['descripcion']] if 'descripcion' in col else None
        es_principal = len(current_items) == 0 or bool(desc_cell and _is_yellow(desc_cell))

        current_items.append({
            'ean': ean_val,
            'descripcion': desc_val,
            'cantidad': cant_val,
            'descuento': dto_val,
            'es_principal': es_principal,
        })

    if current_nombre and current_items:
        modulos.append({'nombre': current_nombre, 'items': current_items})

    return modulos


def parse_descuento(path: str) -> list:
    """
    Despachador: elige el parser según la extensión del archivo.
    Soporta .xlsx, .xls (Excel) y .pdf.
    """
    ext = path.rsplit('.', 1)[-1].lower()
    if ext in ('xlsx', 'xls'):
        return parse_descuento_xlsx(path)
    elif ext == 'pdf':
        return parse_descuento_pdf(path)
    else:
        raise ValueError(f'Formato no soportado: .{ext}. Usá Excel (.xlsx/.xls) o PDF.')
