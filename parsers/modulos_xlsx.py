"""
Parser para Excel de módulos de laboratorio.

Detecta automáticamente 2 formatos:

  FORMATO A (5 cols — nuestra plantilla):
    Fila 1: Título
    Fila 2: NOMBRE MÓDULO | CÓDIGO EAN | DESCRIPCIÓN | CANT. | DESC. %
    Fila módulo: (nombre, None, None, None, None)
    Fila ítem:   (nombre, ean, descripcion, cant, desc_pct)

  FORMATO B (6 cols — Excel real de laboratorio, ej. Roemmers):
    Fila 1: (None, Título, ...)
    Fila 2: COD MOD. | NOMBRE MODULO | CODIGO EAN | DESCRIPCION | CANT, | DESC, %
    Fila módulo: (None, nombre, None, None, None, None)
    Fila ítem:   (cod, nombre, ean, descripcion, cant, desc_pct)
"""


def _detect_format(ws):
    """
    Devuelve 'A' (5 cols, plantilla) o 'B' (6 cols, lab real).
    Busca la fila de encabezados y analiza las columnas.
    """
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if not row or row[0] is None:
            continue
        vals = [str(v).upper().strip() if v else '' for v in row]
        # Formato B: primera columna es "COD MOD" o similar
        if vals[0] in ('COD MOD.', 'COD MOD', 'CODIGO MODULO', 'CÓD. MOD.'):
            return 'B'
        # Formato A: primera columna es el nombre del módulo
        if vals[0] in ('NOMBRE MÓDULO', 'NOMBRE MODULO', 'MÓDULO', 'MODULO'):
            return 'A'
        # Formato C: layout "combo" — encabezados nombrados, la DESCRIPCIÓN del
        # producto va en la primera columna y los módulos se separan con filas
        # "Combo N" (sin EAN). El orden de columnas NO es fijo: se mapea por
        # nombre de encabezado. (ej. export de combos Pharmadorf)
        if vals[0] in ('DESCRIPCION', 'DESCRIPCIÓN') and 'EAN' in vals:
            return 'C'
    return 'A'  # default


def _norm_ean(val):
    """Normaliza un valor de celda a string EAN. Retorna None si inválido."""
    if val is None:
        return None
    try:
        return str(int(float(str(val).strip())))
    except (ValueError, TypeError):
        s = str(val).strip()
        return s if s else None


def _parse_int(val, default=1):
    try:
        return int(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _parse_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _es_amarillo(fg_hex):
    """Detecta amarillo-ish por componentes RGB.
    Un amarillo tiene R y G altos, B bajo."""
    if not fg_hex or len(fg_hex) < 6:
        return False
    fg_hex = str(fg_hex).upper()
    # ARGB (8 chars) o RGB (6 chars). Nos quedamos con los últimos 6.
    rgb = fg_hex[-6:]
    try:
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
    except ValueError:
        return False
    # Amarillos: R y G altos (>= 200), B claramente menor (<=180)
    # y no totalmente blanco (evitar FFFFFF)
    return r >= 200 and g >= 200 and b <= 180 and not (r == 255 and g == 255 and b >= 230)


def _row_destacada(row_cells):
    """True si alguna celda con contenido tiene fondo amarillo.
    row_cells es una tupla de openpyxl.Cell."""
    for c in row_cells:
        if c.value is None:
            continue
        fill = c.fill
        if not fill or fill.patternType in (None, 'none'):
            continue
        fg = getattr(fill.fgColor, 'rgb', None)
        if not fg:
            continue
        if _es_amarillo(str(fg)):
            return True
    return False


def _parse_formato_c(ws):
    """Formato C — layout 'combo' con encabezados nombrados.

    - Mapea columnas por NOMBRE de encabezado (orden no fijo).
    - Agrupa por filas separadoras 'Combo N' (filas con texto pero sin EAN).
    - La descripción del producto va en la primera columna; el nombre del
      módulo sale de la fila separadora.
    Tolera columnas extra (ID, PLAZO, etc.) que simplemente se ignoran, y
    columnas mal ubicadas (ej. el combo bajo 'CANT. PEDIDA' y 'COMBO' vacía):
    el agrupado se hace por las filas separadoras, no por esas columnas.
    """
    # 1. Localizar fila de encabezados (la primera que tenga 'EAN').
    header_idx = None
    headers = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        up = [str(v).strip().upper() if v is not None else '' for v in row]
        if 'EAN' in up:
            header_idx, headers = i, up
            break
    if header_idx is None:
        return []

    def _find(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    idx_ean  = _find('EAN', 'CODIGO EAN', 'CÓDIGO EAN', 'CODIGO DE BARRAS', 'CÓDIGO DE BARRAS')
    idx_desc = _find('DESCRIPCION', 'DESCRIPCIÓN')
    # Descuento: NO matchear 'CANT. PEDIDA' (en estos archivos trae el combo).
    idx_dto  = _find('DESC.', 'DESC', 'DESCUENTO', 'DESC. %', 'DTO', 'DTO.', '% DESC', 'DESCUENTO %')
    # Cantidad/mínimo: preferimos 'MIN O FIJO'/'CANT. MODULO'. Evitamos 'CANT. PEDIDA'.
    idx_cant = _find('MIN O FIJO', 'CANT. MODULO', 'CANT. MÓDULO', 'CANTIDAD', 'MINIMO', 'MÍNIMO', 'MIN')
    idx_combo = _find('COMBO', 'NOMBRE MODULO', 'NOMBRE MÓDULO', 'MODULO', 'MÓDULO')

    if idx_ean is None or idx_desc is None:
        return []

    modules, current = [], None
    for row_cells in ws.iter_rows(min_row=header_idx + 1):
        row = [c.value for c in row_cells]
        if not row or all(v is None for v in row):
            continue
        destacado = _row_destacada(row_cells)
        ean = _norm_ean(row[idx_ean]) if idx_ean < len(row) else None

        if not ean:
            # Fila separadora → abre módulo nuevo. Nombre = texto en col 0 / desc.
            texto = ''
            for j in (idx_desc, 0):
                if j is not None and j < len(row) and row[j]:
                    texto = str(row[j]).strip()
                    break
            if texto:
                if current is not None:
                    modules.append(current)
                current = {'nombre': texto, 'items': []}
            continue

        # Fila de ítem.
        if current is None:
            nombre = ''
            if idx_combo is not None and idx_combo < len(row) and row[idx_combo]:
                nombre = str(row[idx_combo]).strip()
            current = {'nombre': nombre or 'SIN NOMBRE', 'items': []}
        desc = (str(row[idx_desc]).strip()
                if idx_desc < len(row) and row[idx_desc] else '')
        cant = row[idx_cant] if idx_cant is not None and idx_cant < len(row) else None
        dto  = row[idx_dto] if idx_dto is not None and idx_dto < len(row) else 0
        current['items'].append({
            'ean':         ean,
            'descripcion': desc,
            'cant':        _parse_int(cant),
            'desc_pct':    _parse_float(dto),
            'destacado':   destacado,
        })

    if current is not None:
        modules.append(current)
    modules = [m for m in modules if m['items']]
    for m in modules:
        m['items'].sort(key=lambda it: (it.get('descripcion') or '').strip().upper())
    return modules


def parse_modulos_xlsx(path):
    """
    Retorna lista de módulos:
    [
        {
            'nombre': 'MOD. OPTAMOX DUO',
            'items': [
                {'ean': '7793450121123', 'descripcion': '...', 'cant': 10, 'desc_pct': 7.0,
                 'destacado': True/False},
                ...
            ]
        },
        ...
    ]

    'destacado' = True si la fila venía resaltada en amarillo en el Excel
    (indicador visual que los laboratorios usan para marcar packs).
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    fmt = _detect_format(ws)

    if fmt == 'C':
        return _parse_formato_c(ws)

    modules = []
    current = None

    for row_cells in ws.iter_rows(min_row=1):
        row = tuple(c.value for c in row_cells)
        destacado = _row_destacada(row_cells)
        if not row or all(v is None for v in row):
            continue

        # ── FORMATO B ────────────────────────────────────────────────
        if fmt == 'B':
            cod      = row[0] if len(row) > 0 else None
            nombre   = str(row[1]).strip() if len(row) > 1 and row[1] else None
            ean_raw  = row[2] if len(row) > 2 else None
            desc     = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            cant     = row[4] if len(row) > 4 else None
            desc_pct = row[5] if len(row) > 5 else 0

            # Saltar fila de encabezados de columna
            if nombre and str(nombre).upper() in ('NOMBRE MODULO', 'NOMBRE MÓDULO'):
                continue

            # Cabecera de módulo: col[0] vacío y cols[2..5] todas vacías
            # Usamos "not x" en vez de "is None" para cubrir también strings vacíos
            tail_empty = all(
                not row[i]
                for i in range(2, min(6, len(row)))
            )
            if not cod and tail_empty and nombre:
                if current is not None:
                    modules.append(current)
                current = {'nombre': nombre, 'items': []}
                continue

            # Fila de ítem: necesita al menos un nombre en col[1]
            if not nombre:
                continue

            ean = _norm_ean(ean_raw)
            if not ean:
                # Fallback: EAN puede estar en col[1] (layout real Roemmers: COD MOD | EAN | vacío | DESC | CANT | %)
                ean = _norm_ean(nombre)
            if not ean:
                # EAN ausente → saltar ítem pero NO cortar el módulo
                continue

            if current is None:
                current = {'nombre': nombre, 'items': []}

            current['items'].append({
                'ean':         ean,
                'descripcion': desc,
                'cant':        _parse_int(cant),
                'desc_pct':    _parse_float(desc_pct),
                'destacado':   destacado,
            })
            continue   # ← FORMAT B completamente manejado aquí

        # ── FORMATO A ────────────────────────────────────────────────
        nombre   = str(row[0]).strip() if row[0] else None
        ean_raw  = row[1] if len(row) > 1 else None
        desc     = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        cant     = row[3] if len(row) > 3 else None
        desc_pct = row[4] if len(row) > 4 else 0

        # Saltar fila de encabezados de columna
        if nombre and str(nombre).upper() in ('NOMBRE MÓDULO', 'NOMBRE MODULO', 'MÓDULO', 'MODULO'):
            continue

        ean = _norm_ean(ean_raw)

        if ean is None:
            # Fila de cabecera de módulo (sin EAN)
            if nombre and (current is None or nombre != current['nombre']):
                if current is not None:
                    modules.append(current)
                current = {'nombre': nombre, 'items': []}
        else:
            # Fila de ítem — si el nombre del módulo en col[0] cambió respecto al
            # current, abrir módulo nuevo (formato Siegfried donde el nombre se
            # repite por cada fila, en vez de aparecer como cabecera separada).
            if current is None or (nombre and nombre != current['nombre']):
                if current is not None:
                    modules.append(current)
                current = {'nombre': nombre or 'SIN NOMBRE', 'items': []}
            current['items'].append({
                'ean':         ean,
                'descripcion': desc,
                'cant':        _parse_int(cant),
                'desc_pct':    _parse_float(desc_pct),
                'destacado':   destacado,
            })

    if current is not None:
        modules.append(current)

    # Filtrar módulos sin ítems (ej. fila de título global)
    modules = [m for m in modules if m['items']]

    # Productos por orden alfabético dentro de cada módulo (regla global del proyecto).
    for m in modules:
        m['items'].sort(key=lambda it: (it.get('descripcion') or '').strip().upper())

    return modules
