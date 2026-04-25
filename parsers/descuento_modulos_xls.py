"""
Parser para Excel de módulos de descuento (ej: MODULOS ROEMMERS).
Formato esperado:
  - Fila de encabezado con columnas: NOMBRE MODULO (o similar), CODIGO EAN, DESCRIPCION, CANT, DESC
  - El nombre del módulo puede estar en celdas fusionadas (se propaga hacia abajo)
  - El descuento puede variar por ítem

Retorna lista de módulos:
[
  {
    'nombre_modulo': 'MOD. OPTAMOX DUO',
    'items': [
      {'codigo_ean': '7798012345678', 'descripcion': '...', 'cantidad': 2, 'descuento': 7.0},
      ...
    ]
  },
  ...
]
"""
import re

import openpyxl


def _clean(v):
    if v is None:
        return ''
    return str(v).strip()


def _to_int(v, default=1):
    try:
        return max(1, int(float(str(v).strip())))
    except Exception:
        return default


def _to_float(v, default=0.0):
    try:
        s = str(v).strip().replace('%', '').replace(',', '.')
        return float(s)
    except Exception:
        return default


def _find_header_row(rows):
    """
    Busca la fila de encabezado inspeccionando palabras clave.
    Devuelve (row_idx, col_map) donde col_map = {'modulo': idx, 'ean': idx, 'desc': idx, 'cant': idx, 'dto': idx}
    """
    keywords = {
        'modulo': ['modulo', 'módulo', 'nombre', 'mod'],
        'ean':    ['ean', 'codigo', 'código', 'barras', 'barra'],
        'desc':   ['descripcion', 'descripción', 'producto'],
        'cant':   ['cant', 'cantidad', 'unid'],
        'dto':    ['desc', 'dto', 'descuento', '%'],
    }
    for i, row in enumerate(rows):
        texts = [_clean(c).lower() for c in row]
        if not any(texts):
            continue
        col_map = {}
        for field, kws in keywords.items():
            for j, t in enumerate(texts):
                if any(kw in t for kw in kws):
                    if field not in col_map:
                        col_map[field] = j
        # Necesitamos al menos EAN y módulo
        if 'ean' in col_map and 'modulo' in col_map:
            return i, col_map
    return None, {}


def parse_descuento_modulos_xls(path):
    """
    Parsea un Excel de módulos de descuento.
    Soporta .xlsx. Para .xls convertir antes a .xlsx o usar xlrd.
    Devuelve lista de módulos con sus ítems.
    """
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb.active

    # Expandir celdas fusionadas: copiar el valor de la celda origen a todas las que cubre
    merged_values = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = top_left

    rows = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        cells = []
        for c_idx, cell in enumerate(row, start=1):
            val = merged_values.get((r_idx, c_idx), cell.value)
            cells.append(val)
        rows.append(cells)

    wb.close()

    header_idx, col_map = _find_header_row(rows)
    if header_idx is None:
        # Intentar formato simple: 5 columnas, sin encabezado explícito
        # columna 0=modulo, 1=ean, 2=desc, 3=cant, 4=dto
        col_map = {'modulo': 0, 'ean': 1, 'desc': 2, 'cant': 3, 'dto': 4}
        header_idx = 0

    modulos = []
    current_mod = None
    last_mod_name = ''

    for row in rows[header_idx + 1:]:
        if not any(c for c in row if c is not None):
            continue

        mod_val   = _clean(row[col_map['modulo']])  if col_map.get('modulo') is not None and col_map['modulo'] < len(row) else ''
        ean_val   = _clean(row[col_map['ean']])      if col_map.get('ean')    is not None and col_map['ean']    < len(row) else ''
        desc_val  = _clean(row[col_map.get('desc', 2)]) if col_map.get('desc') is not None and col_map['desc'] < len(row) else ''
        cant_val  = row[col_map['cant']]             if col_map.get('cant')   is not None and col_map['cant']  < len(row) else 1
        dto_val   = row[col_map.get('dto', 4)]       if col_map.get('dto')    is not None and col_map['dto']   < len(row) else 0

        # Normalizar EAN: solo dígitos, mínimo 7
        ean_clean = re.sub(r'\D', '', ean_val)
        if len(ean_clean) < 7:
            continue

        cantidad  = _to_int(cant_val)
        descuento = _to_float(dto_val)

        # Detectar cambio de módulo
        if mod_val and mod_val != last_mod_name:
            last_mod_name = mod_val
            current_mod = {'nombre_modulo': mod_val, 'items': []}
            modulos.append(current_mod)

        if current_mod is None:
            # Sin nombre de módulo todavía, crear uno genérico
            current_mod = {'nombre_modulo': 'SIN NOMBRE', 'items': []}
            modulos.append(current_mod)

        current_mod['items'].append({
            'codigo_ean': ean_clean,
            'descripcion': desc_val,
            'cantidad': cantidad,
            'descuento': descuento,
        })

    return modulos
