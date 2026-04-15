"""
Parser para importación libre de descuentos.

Formato esperado (columnas detectadas por nombre):
  LAB | CODIGO DE BARRAS / EAN | DESCRIPCION | CANTIDAD / CANT | DESCUENTO / DESC

Devuelve:
  list of { 'lab': str, 'ean': str, 'descripcion': str,
             'cantidad': int, 'descuento': float }
"""
import openpyxl


def _safe_str(v) -> str:
    return str(v).strip() if v is not None else ''


def _safe_float(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(str(v).replace('%', '').replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0


def _safe_int(v) -> int:
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return 0


def _find_header(ws):
    """Busca la fila de encabezado (contiene LAB, CODIGO o EAN)."""
    for i, row in enumerate(ws.iter_rows(max_row=10), 1):
        joined = ' '.join(_safe_str(c.value).upper() for c in row if c.value)
        if any(k in joined for k in ('LAB', 'CODIGO', 'EAN', 'DESCRIPCION')):
            return i
    return 1


def _detect_columns(ws, header_row: int) -> dict:
    col = {}
    for i, cell in enumerate(ws[header_row]):
        h = _safe_str(cell.value).upper()
        if not h:
            continue
        if h == 'LAB' or h.startswith('LAB'):
            col.setdefault('lab', i)
        elif 'CODIGO' in h or 'EAN' in h or 'BARRA' in h:
            col.setdefault('ean', i)
        elif 'DESCRIPCION' in h or 'DESCRIPCIÓN' in h:
            col.setdefault('descripcion', i)
        elif 'CANTIDAD' in h or h == 'CANT':
            col.setdefault('cantidad', i)
        elif 'DESCUENTO' in h or h == 'DESC':
            col.setdefault('descuento', i)
    return col


def parse_descuento_libre(path: str) -> list:
    """
    Lee el Excel y devuelve lista plana de artículos con su lab.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row = _find_header(ws)
    col = _detect_columns(ws, header_row)

    if 'ean' not in col and 'descripcion' not in col:
        raise ValueError('No se detectaron columnas EAN/CODIGO ni DESCRIPCION.')

    items = []
    current_lab = ''

    for row in ws.iter_rows(min_row=header_row + 1):
        vals = [cell.value for cell in row]
        if all(v is None or _safe_str(v) == '' for v in vals):
            continue

        lab_val  = _safe_str(vals[col['lab']]) if 'lab' in col else ''
        ean_val  = _safe_str(vals[col['ean']]) if 'ean' in col else ''
        desc_val = _safe_str(vals[col['descripcion']]) if 'descripcion' in col else ''
        cant_val = _safe_int(vals[col['cantidad']]) if 'cantidad' in col else 0
        dto_val  = _safe_float(vals[col['descuento']]) if 'descuento' in col else 0.0

        if not ean_val and not desc_val:
            continue

        # Lab puede repetirse en cada fila o venir solo en la primera
        if lab_val:
            current_lab = lab_val

        items.append({
            'lab': current_lab,
            'ean': ean_val,
            'descripcion': desc_val,
            'cantidad': cant_val,
            'descuento': dto_val,
        })

    return items
