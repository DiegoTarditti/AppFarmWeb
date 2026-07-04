"""Informe XLSX de ventas para una lista de medicamentos.

Para cada producto solicitado:
- Identifica su monodroga y laboratorio.
- Calcula ventas mes a mes (últimos 12 meses) desde obs_ventas_mensuales.
- Agrupa por monodroga y suma TODOS los productos de esa monodroga
  (lo que llamamos "competencia": misma droga, otros laboratorios).
- Genera un XLSX con:
    - Hoja 'Resumen' (totales por producto vs competencia, % cuota).
    - Una hoja por monodroga con todos los productos y ventas mensuales.

Uso (desde DENTRO del container web):
    docker compose exec web python scripts/informe_medicamentos.py

Salida: informes/informe_medicamentos_YYYY-MM-DD.xlsx en el working dir.
"""
import os
import sys
from collections import defaultdict
from datetime import date

# Permitir importar database/ desde /app cuando se corre con docker compose exec.
_repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _repo_root not in sys.path:
    sys.path.insert(0, _repo_root)

# Productos a investigar (búsqueda por descripción ILIKE)
PRODUCTOS_BUSCADOS = ['ALPERTA', 'NABILA', 'REDUPROST', 'ATENIX', 'QUETIAZIC']

# Cuántos meses hacia atrás incluir
MESES_VENTANA = 12


def openpyxl_strref(ws, row, col):
    """Construye una SeriesLabel con StrRef a una celda (para títulos de serie)."""
    from openpyxl.chart.data_source import StrRef
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils import get_column_letter
    ref = f"'{ws.title}'!${get_column_letter(col)}${row}"
    return SeriesLabel(strRef=StrRef(ref))


def main():
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from database import (
        ObsLaboratorio,
        ObsNombreDroga,
        ObsProducto,
        ObsVentaMensual,
        get_db,
        init_engine,
    )

    # Inicializar engine — el script no pasa por app.py
    init_engine(os.environ.get('DATABASE_URL', 'sqlite:///farmacia.db'))

    id_farmacia = int(os.environ.get('OBSERVER_ID_FARMACIA', '10525'))
    hoy = date.today()

    # Ventana de meses: lista [(anio, mes), ...] ordenada del más viejo al más nuevo
    meses_ventana = []
    y, m = hoy.year, hoy.month
    for _ in range(MESES_VENTANA):
        meses_ventana.append((y, m))
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    meses_ventana.reverse()

    with get_db() as s:
        # 1) Resolver los productos buscados: por cada nombre traemos las filas
        #    de obs_productos que matchean. Si un nombre matchea varias presentaciones
        #    (ej. ALPERTA 5mg y ALPERTA 10mg), las traemos todas — el informe las separa.
        print('Resolviendo productos...')
        productos_target = []  # [(observer_id, descripcion, droga_id, lab_id)]
        for n in PRODUCTOS_BUSCADOS:
            rows = (s.query(ObsProducto.observer_id, ObsProducto.descripcion,
                            ObsProducto.nombre_droga_observer,
                            ObsProducto.laboratorio_observer)
                    .filter(ObsProducto.descripcion.ilike(f'%{n}%'),
                            ObsProducto.fecha_baja.is_(None))
                    .order_by(ObsProducto.descripcion)
                    .all())
            if not rows:
                print(f'  ⚠ Sin resultados para {n}')
                continue
            for r in rows:
                productos_target.append((r.observer_id, r.descripcion,
                                          r.nombre_droga_observer,
                                          r.laboratorio_observer))
                print(f'  ✓ {n}: {r.descripcion} (obs_id={r.observer_id})')

        if not productos_target:
            print('No se encontró ningún producto. Abortando.')
            return

        # 2) Monodrogas únicas que vamos a analizar
        drogas_ids = {p[2] for p in productos_target if p[2]}
        if not drogas_ids:
            print('Ninguno de los productos tiene monodroga vinculada. Abortando.')
            return

        # Nombres de monodroga + laboratorios
        droga_nombre = dict(s.query(ObsNombreDroga.observer_id, ObsNombreDroga.descripcion)
                            .filter(ObsNombreDroga.observer_id.in_(list(drogas_ids))))
        lab_nombre = dict(s.query(ObsLaboratorio.observer_id, ObsLaboratorio.descripcion))

        # 3) TODOS los productos de esas monodrogas (competencia)
        print(f'\nDescubriendo competencia para {len(drogas_ids)} monodrogas...')
        productos_competencia = (s.query(ObsProducto.observer_id, ObsProducto.descripcion,
                                          ObsProducto.nombre_droga_observer,
                                          ObsProducto.laboratorio_observer)
                                 .filter(ObsProducto.nombre_droga_observer.in_(list(drogas_ids)),
                                         ObsProducto.fecha_baja.is_(None))
                                 .all())
        print(f'  Total productos a analizar (incl. competencia): {len(productos_competencia)}')

        # 4) Ventas mes a mes para todos esos productos
        obs_ids = [p.observer_id for p in productos_competencia]
        anio_min, mes_min = meses_ventana[0]
        anio_max, mes_max = meses_ventana[-1]
        ventas = (s.query(ObsVentaMensual.producto_observer,
                          ObsVentaMensual.anio, ObsVentaMensual.mes,
                          ObsVentaMensual.unidades, ObsVentaMensual.monto)
                  .filter(ObsVentaMensual.id_farmacia == id_farmacia,
                          ObsVentaMensual.producto_observer.in_(obs_ids),
                          (ObsVentaMensual.anio * 100 + ObsVentaMensual.mes) >= (anio_min * 100 + mes_min),
                          (ObsVentaMensual.anio * 100 + ObsVentaMensual.mes) <= (anio_max * 100 + mes_max))
                  .all())

        # ventas_por[obs_id][(anio,mes)] = (unidades, monto)
        ventas_por = defaultdict(dict)
        for v in ventas:
            ventas_por[v.producto_observer][(v.anio, v.mes)] = (int(v.unidades or 0),
                                                                  float(v.monto or 0))

    # 5) Construir Excel
    wb = Workbook()
    wb.remove(wb.active)

    # Paleta
    fill_header = PatternFill(start_color='1D9E75', end_color='1D9E75', fill_type='solid')
    fill_target = PatternFill(start_color='FFF3C4', end_color='FFF3C4', fill_type='solid')
    fill_alt = PatternFill(start_color='F5F5F7', end_color='F5F5F7', fill_type='solid')
    fill_title = PatternFill(start_color='22222D', end_color='22222D', fill_type='solid')
    font_header = Font(bold=True, color='FFFFFF', size=11)
    font_title_white = Font(bold=True, color='FFFFFF', size=14)
    font_bold = Font(bold=True)
    thin = Side(border_style='thin', color='DDDDDD')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Agrupar productos por monodroga (necesario antes del Resumen)
    por_droga = defaultdict(list)
    for p in productos_competencia:
        por_droga[p.nombre_droga_observer].append(p)
    target_obs_ids = {p[0] for p in productos_target}
    drogas_ordenadas = sorted(por_droga.keys(),
                              key=lambda x: droga_nombre.get(x, ''))

    # ── Hoja RESUMEN ─────────────────────────────────────────────────────
    ws = wb.create_sheet('Resumen')

    # Título tipo banner
    ws.merge_cells('A1:G2')
    ws['A1'] = 'INFORME DE VENTAS — Productos vs Competencia'
    ws['A1'].font = font_title_white
    ws['A1'].fill = fill_title
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    ws['A4'] = 'Productos buscados:'
    ws['A4'].font = font_bold
    ws['B4'] = ', '.join(PRODUCTOS_BUSCADOS)
    ws['A5'] = 'Ventana:'
    ws['A5'].font = font_bold
    ws['B5'] = f'últimos {MESES_VENTANA} meses ({meses_ventana[0][0]}-{meses_ventana[0][1]:02d} → {meses_ventana[-1][0]}-{meses_ventana[-1][1]:02d})'
    ws['A6'] = 'Generado:'
    ws['A6'].font = font_bold
    ws['B6'] = hoy.isoformat()
    ws['A7'] = 'Farmacia ID:'
    ws['A7'].font = font_bold
    ws['B7'] = id_farmacia

    headers_resumen = ['Monodroga', 'Producto', 'Laboratorio',
                       'Tipo', 'U totales', '% cuota U', 'Promedio U/mes']
    HEADER_ROW = 9
    for i, h in enumerate(headers_resumen, start=1):
        c = ws.cell(row=HEADER_ROW, column=i, value=h)
        c.fill = fill_header
        c.font = font_header
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border_all
    ws.row_dimensions[HEADER_ROW].height = 22

    row_idx = HEADER_ROW + 1
    alt_flag = False
    # Para el chart por monodroga (resumen)
    monodroga_totales = []  # [(nombre_droga, total_buscado, total_competencia)]

    for droga_id in drogas_ordenadas:
        droga_n = droga_nombre.get(droga_id, f'(monodroga #{droga_id})')
        total_u_droga = 0
        for p in por_droga[droga_id]:
            v = ventas_por.get(p.observer_id, {})
            total_u_droga += sum(u for u, _m in v.values())
        total_u_buscado = 0
        # Ordenar: BUSCADO primero, después por U totales desc
        prods_ordenados = sorted(
            por_droga[droga_id],
            key=lambda p: (
                0 if p.observer_id in target_obs_ids else 1,
                -sum(u for u, _m in ventas_por.get(p.observer_id, {}).values())
            )
        )
        for p in prods_ordenados:
            v = ventas_por.get(p.observer_id, {})
            tot_u = sum(u for u, _m in v.values())
            tipo = 'BUSCADO' if p.observer_id in target_obs_ids else 'competencia'
            if tipo == 'BUSCADO':
                total_u_buscado += tot_u
            cuota = (tot_u / total_u_droga * 100) if total_u_droga else 0
            avg = tot_u / MESES_VENTANA
            data = [droga_n, p.descripcion,
                    lab_nombre.get(p.laboratorio_observer, '—'),
                    tipo, tot_u, round(cuota, 1), round(avg, 1)]
            for i, val in enumerate(data, start=1):
                c = ws.cell(row=row_idx, column=i, value=val)
                c.border = border_all
                if tipo == 'BUSCADO':
                    c.fill = fill_target
                    c.font = font_bold
                elif alt_flag:
                    c.fill = fill_alt
                if i == 5:
                    c.number_format = '#,##0'
                elif i == 6:
                    c.number_format = '0.0"%"'
                elif i == 7:
                    c.number_format = '#,##0.0'
            row_idx += 1
            alt_flag = not alt_flag
        monodroga_totales.append((droga_n, total_u_buscado,
                                   total_u_droga - total_u_buscado))

    # Anchos
    for i, w in enumerate([30, 55, 32, 14, 13, 12, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze headers
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    # ── Gráfico: BUSCADO vs COMPETENCIA por monodroga ────────────────────
    # Vamos a poner una tabla auxiliar a la derecha + un BarChart agrupado
    aux_col = 9
    ws.cell(row=HEADER_ROW, column=aux_col, value='Monodroga').font = font_bold
    ws.cell(row=HEADER_ROW, column=aux_col + 1, value='Buscado U').font = font_bold
    ws.cell(row=HEADER_ROW, column=aux_col + 2, value='Competencia U').font = font_bold
    for i, (n, b, c) in enumerate(monodroga_totales, start=1):
        ws.cell(row=HEADER_ROW + i, column=aux_col, value=n)
        ws.cell(row=HEADER_ROW + i, column=aux_col + 1, value=b).number_format = '#,##0'
        ws.cell(row=HEADER_ROW + i, column=aux_col + 2, value=c).number_format = '#,##0'
    ws.column_dimensions[get_column_letter(aux_col)].width = 28
    ws.column_dimensions[get_column_letter(aux_col + 1)].width = 14
    ws.column_dimensions[get_column_letter(aux_col + 2)].width = 14

    chart = BarChart()
    chart.type = 'bar'
    chart.style = 11
    chart.title = 'Unidades totales por monodroga — Buscado vs Competencia'
    chart.y_axis.title = 'Monodroga'
    chart.x_axis.title = 'Unidades'
    chart.height = max(10, 1 + len(monodroga_totales) * 1.5)
    chart.width = 22
    data_ref = Reference(ws,
                          min_col=aux_col + 1, max_col=aux_col + 2,
                          min_row=HEADER_ROW,
                          max_row=HEADER_ROW + len(monodroga_totales))
    cats_ref = Reference(ws, min_col=aux_col,
                          min_row=HEADER_ROW + 1,
                          max_row=HEADER_ROW + len(monodroga_totales))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    # Anclar el chart debajo de la tabla
    anchor_row = row_idx + 2
    ws.add_chart(chart, f'A{anchor_row}')

    # ── Hojas POR MONODROGA ──────────────────────────────────────────────
    for droga_id in drogas_ordenadas:
        droga_n = droga_nombre.get(droga_id, f'monodroga_{droga_id}')
        sheet_n = droga_n[:28] + '...' if len(droga_n) > 31 else droga_n
        for ch in '[]:*?/\\':
            sheet_n = sheet_n.replace(ch, '_')
        ws = wb.create_sheet(sheet_n)

        # Banner
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=4 + len(meses_ventana))
        ws.cell(row=1, column=1, value=f'  {droga_n}')
        ws.cell(row=1, column=1).font = font_title_white
        ws.cell(row=1, column=1).fill = fill_title
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 24

        HEADER_ROW = 3
        headers = ['Producto', 'Laboratorio', 'Tipo']
        for y, m in meses_ventana:
            headers.append(f'{y}-{m:02d}')
        headers.append('Total U')
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=HEADER_ROW, column=i, value=h)
            c.fill = fill_header
            c.font = font_header
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border_all
        ws.row_dimensions[HEADER_ROW].height = 22

        # Ordenar: BUSCADO primero, después por total desc
        prods_ordenados = sorted(
            por_droga[droga_id],
            key=lambda p: (
                0 if p.observer_id in target_obs_ids else 1,
                -sum(u for u, _m in ventas_por.get(p.observer_id, {}).values())
            )
        )

        # Limitar competencia a top 8 para que el gráfico no sea ilegible
        TOP_N = 8
        buscados = [p for p in prods_ordenados if p.observer_id in target_obs_ids]
        competencia = [p for p in prods_ordenados if p.observer_id not in target_obs_ids][:TOP_N]
        prods_a_pintar = buscados + competencia

        row_idx = HEADER_ROW + 1
        first_data_row = row_idx
        alt_flag = False
        for p in prods_a_pintar:
            tipo = 'BUSCADO' if p.observer_id in target_obs_ids else 'competencia'
            v = ventas_por.get(p.observer_id, {})
            row = [p.descripcion,
                   lab_nombre.get(p.laboratorio_observer, '—'),
                   tipo]
            tot_u = 0
            for y, m in meses_ventana:
                u, _mm = v.get((y, m), (0, 0))
                row.append(u)
                tot_u += u
            row.append(tot_u)
            for i, val in enumerate(row, start=1):
                c = ws.cell(row=row_idx, column=i, value=val)
                c.border = border_all
                if tipo == 'BUSCADO':
                    c.fill = fill_target
                    c.font = font_bold
                elif alt_flag:
                    c.fill = fill_alt
                if i >= 4:
                    c.number_format = '#,##0'
            row_idx += 1
            alt_flag = not alt_flag
        last_data_row = row_idx - 1

        # Anchos
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 14
        for i in range(4, 4 + len(meses_ventana) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 10

        # Freeze first 3 cols + header
        ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=4)

        # ── Gráfico de líneas: evolución mensual por producto ────────────
        if last_data_row >= first_data_row:
            chart = LineChart()
            chart.title = f'Evolución mensual — {droga_n[:60]}'
            chart.style = 12
            chart.y_axis.title = 'Unidades'
            chart.x_axis.title = 'Mes'
            chart.height = 12
            chart.width = 28
            # Cada fila es una serie. Datos en columnas 4..(4+len(meses)-1)
            for r in range(first_data_row, last_data_row + 1):
                serie = Reference(ws,
                                   min_col=4, max_col=3 + len(meses_ventana),
                                   min_row=r, max_row=r)
                # Título de la serie = celda con la descripción del producto
                title_ref = Reference(ws, min_col=1, max_col=1, min_row=r, max_row=r)
                chart.add_data(serie, titles_from_data=False)
                chart.series[-1].tx = openpyxl_strref(ws, r, 1)
            # Eje X: encabezados de meses
            cats = Reference(ws,
                              min_col=4, max_col=3 + len(meses_ventana),
                              min_row=HEADER_ROW, max_row=HEADER_ROW)
            chart.set_categories(cats)
            ws.add_chart(chart, f'A{last_data_row + 3}')

    # 6) Guardar
    from datetime import datetime
    out_dir = 'informes'
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    out_path = os.path.join(out_dir, f'informe_medicamentos_{stamp}.xlsx')
    wb.save(out_path)
    print(f'\n✓ Informe generado: {out_path}')
    print(f'  Hojas: Resumen + {len(por_droga)} monodrogas (con gráficos)')


if __name__ == '__main__':
    sys.exit(main() or 0)
