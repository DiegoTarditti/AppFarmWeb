"""Informe XLSX tipo IQVIA por monodroga.

Para cada monodroga (de los productos buscados), genera una "tarjeta" con:
  - 3 indicadores: Unidades, Valores $, Precios (PVP promedio)
  - 3 cortes: M Total / ETICO / OTC
  - 5 variaciones %: MoM, YoY, YTD, MAT, Incremento año (precios)
  - PVP promedio al cierre del mes de referencia

Clasificación ETICO vs OTC (id_tipo_venta_control):
  - L = venta libre → OTC
  - R, A = receta / receta archivada → ETICO
  - 1..8 = controlados / psicotrópicos → ETICO

Uso:
    docker compose exec web python scripts/informe_mercado_droga.py
"""
import os
import sys
from collections import defaultdict
from datetime import date

_repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _repo_root not in sys.path:
    sys.path.insert(0, _repo_root)

PRODUCTOS_BUSCADOS = ['ALPERTA', 'NABILA', 'REDUPROST', 'ATENIX', 'QUETIAZIC']

# id_tipo_venta_control → 'OTC' | 'ETICO'
def _segmento(tvc):
    t = (tvc or '').strip()
    if t == 'L':
        return 'OTC'
    return 'ETICO'  # R/A/1..8 → recetario


def _mes_anterior(y, m):
    if m == 1:
        return y - 1, 12
    return y, m - 1


def _restar_meses(y, m, n):
    for _ in range(n):
        y, m = _mes_anterior(y, m)
    return y, m


def _pct(actual, base):
    if not base:
        return None
    return round((actual - base) / base * 100, 1)


def main():
    from datetime import datetime

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from database import (
        ObsLaboratorio,
        ObsNombreDroga,
        ObsProducto,
        ObsVentaMensual,
        get_db,
        init_engine,
    )

    init_engine(os.environ.get('DATABASE_URL', 'sqlite:///farmacia.db'))
    id_farmacia = int(os.environ.get('OBSERVER_ID_FARMACIA', '10525'))

    with get_db() as s:
        # 1) Encontrar todas las monodrogas de los productos buscados
        print('Buscando productos y monodrogas...')
        drogas_set = set()
        productos_buscados_obs_ids = set()
        for n in PRODUCTOS_BUSCADOS:
            rows = (s.query(ObsProducto.observer_id,
                            ObsProducto.nombre_droga_observer)
                    .filter(ObsProducto.descripcion.ilike(f'%{n}%'),
                            ObsProducto.fecha_baja.is_(None),
                            ObsProducto.nombre_droga_observer.isnot(None))
                    .all())
            for r in rows:
                drogas_set.add(r.nombre_droga_observer)
                productos_buscados_obs_ids.add(r.observer_id)
        lab_nombre_global = dict(s.query(ObsLaboratorio.observer_id,
                                          ObsLaboratorio.descripcion))

        droga_nombre = dict(s.query(ObsNombreDroga.observer_id, ObsNombreDroga.descripcion)
                            .filter(ObsNombreDroga.observer_id.in_(list(drogas_set))))
        print(f'  Monodrogas detectadas: {len(drogas_set)}')

        # 2) Todos los productos por monodroga + su id_tipo_venta_control + lab
        prods = (s.query(ObsProducto.observer_id,
                         ObsProducto.descripcion,
                         ObsProducto.nombre_droga_observer,
                         ObsProducto.id_tipo_venta_control,
                         ObsProducto.laboratorio_observer)
                 .filter(ObsProducto.nombre_droga_observer.in_(list(drogas_set)),
                         ObsProducto.fecha_baja.is_(None))
                 .all())
        # droga → segmento → set(obs_id)
        droga_seg_ids = defaultdict(lambda: defaultdict(set))
        # obs_id → (descripcion, lab_id)
        producto_meta = {}
        # droga → list(obs_id) para iterar al armar TOP
        droga_prods = defaultdict(list)
        for p in prods:
            seg = _segmento(p.id_tipo_venta_control)
            droga_seg_ids[p.nombre_droga_observer][seg].add(p.observer_id)
            droga_seg_ids[p.nombre_droga_observer]['TOTAL'].add(p.observer_id)
            producto_meta[p.observer_id] = (p.descripcion, p.laboratorio_observer)
            droga_prods[p.nombre_droga_observer].append(p.observer_id)
        print(f'  Productos totales (incl. competencia): {sum(len(d["TOTAL"]) for d in droga_seg_ids.values())}')

        # 3) Mes de referencia: último mes CERRADO (mes anterior al actual).
        #    Diego 2026-06-23: junio está en curso → usamos mayo para que
        #    todas las comparativas sean contra meses completos.
        hoy = date.today()
        ref_y, ref_m = _mes_anterior(hoy.year, hoy.month)
        es_parcial = False
        print(f'  Mes de referencia: {ref_y}-{ref_m:02d} (último cerrado)')

        # 4) Traer TODAS las ventas mensuales necesarias en una sola query.
        #    Ventana: 24 meses hacia atrás desde el mes de referencia (para MAT).
        ventana_min_y, ventana_min_m = _restar_meses(ref_y, ref_m, 24)
        todos_obs_ids = set()
        for d in droga_seg_ids.values():
            todos_obs_ids.update(d['TOTAL'])
        ventas = (s.query(ObsVentaMensual.producto_observer,
                          ObsVentaMensual.anio, ObsVentaMensual.mes,
                          ObsVentaMensual.unidades, ObsVentaMensual.monto)
                  .filter(ObsVentaMensual.id_farmacia == id_farmacia,
                          ObsVentaMensual.producto_observer.in_(list(todos_obs_ids)),
                          (ObsVentaMensual.anio * 100 + ObsVentaMensual.mes) >= (ventana_min_y * 100 + ventana_min_m),
                          (ObsVentaMensual.anio * 100 + ObsVentaMensual.mes) <= (ref_y * 100 + ref_m))
                  .all())
        # ventas_por[(obs_id, anio, mes)] = (u, m)
        ventas_por = {}
        for v in ventas:
            ventas_por[(v.producto_observer, v.anio, v.mes)] = (
                int(v.unidades or 0), float(v.monto or 0))

    # Helper: agregar (u, m) en un set de obs_ids para un rango de meses
    def _agg(obs_ids, meses):
        u_tot = m_tot = 0
        for oid in obs_ids:
            for (y, m) in meses:
                u, mm = ventas_por.get((oid, y, m), (0, 0))
                u_tot += u
                m_tot += mm
        return u_tot, m_tot

    # Helper: contar cuántos meses del período tienen ventas > 0 (a nivel
    # monodroga). Diego 2026-06-23: el sync Observer arranca en feb 2025
    # → períodos anteriores a esa fecha pueden tener gaps que inflan las
    # variaciones. Si el período base tiene <10/12 meses con datos, el
    # MAT se marca como 'n/d'.
    def _meses_con_datos(obs_ids, meses):
        n = 0
        for (y, m) in meses:
            u = sum(ventas_por.get((oid, y, m), (0, 0))[0] for oid in obs_ids)
            if u > 0:
                n += 1
        return n

    def _meses_rango(y_end, m_end, n):
        """Lista de los últimos n meses terminando en (y_end, m_end)."""
        out = []
        y, m = y_end, m_end
        for _ in range(n):
            out.append((y, m))
            y, m = _mes_anterior(y, m)
        return out

    def _meses_ytd(y, m):
        return [(y, mm) for mm in range(1, m + 1)]

    # 5) Construir Excel
    wb = Workbook()
    wb.remove(wb.active)

    fill_title = PatternFill(start_color='B8E0C7', end_color='B8E0C7', fill_type='solid')
    fill_header = PatternFill(start_color='F7F77B', end_color='F7F77B', fill_type='solid')
    fill_section = PatternFill(start_color='B8E0C7', end_color='B8E0C7', fill_type='solid')
    fill_data = PatternFill(start_color='FFFCEC', end_color='FFFCEC', fill_type='solid')
    fill_footer = PatternFill(start_color='B8E0C7', end_color='B8E0C7', fill_type='solid')
    font_title = Font(bold=True, size=14)
    font_header = Font(bold=True, italic=True)
    font_section = Font(bold=True)
    font_bold = Font(bold=True)
    thin = Side(border_style='thin', color='999999')
    medium = Side(border_style='medium', color='666666')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_strong = Border(left=medium, right=medium, top=medium, bottom=medium)

    MESES_ES = ['', 'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
                'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']

    drogas_ordenadas = sorted(drogas_set,
                              key=lambda x: droga_nombre.get(x, ''))

    # Hoja índice/resumen
    ws_idx = wb.create_sheet('Índice')
    ws_idx.merge_cells('A1:B2')
    ws_idx['A1'] = 'INFORMES MERCADO POR MONODROGA'
    ws_idx['A1'].font = Font(bold=True, size=14)
    ws_idx['A1'].fill = fill_title
    ws_idx['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_idx['A4'] = f'Mes de referencia: {MESES_ES[ref_m]} {ref_y}'
    ws_idx['A4'].font = font_bold
    ws_idx['A5'] = f'Productos buscados: {", ".join(PRODUCTOS_BUSCADOS)}'
    ws_idx['A6'] = f'Farmacia ID: {id_farmacia}'
    ws_idx['A7'] = 'Fuente: ObServer (obs_ventas_mensuales) + datos propios'
    ws_idx['A9'] = 'Monodroga'
    ws_idx['B9'] = '# Productos'
    ws_idx['A9'].font = font_bold
    ws_idx['B9'].font = font_bold

    for i, droga_id in enumerate(drogas_ordenadas, start=1):
        droga_n = droga_nombre.get(droga_id, f'monodroga #{droga_id}')
        n_prods = len(droga_seg_ids[droga_id]['TOTAL'])
        ws_idx.cell(row=9 + i, column=1, value=droga_n)
        ws_idx.cell(row=9 + i, column=2, value=n_prods)

    ws_idx.column_dimensions['A'].width = 55
    ws_idx.column_dimensions['B'].width = 14

    # 6) Una hoja por monodroga con el layout tipo IQVIA
    for droga_id in drogas_ordenadas:
        droga_n = droga_nombre.get(droga_id, f'monodroga_{droga_id}')
        sheet_n = droga_n[:28] + '...' if len(droga_n) > 31 else droga_n
        for ch in '[]:*?/\\':
            sheet_n = sheet_n.replace(ch, '_')
        ws = wb.create_sheet(sheet_n)

        # Calcular todos los números para los 3 segmentos
        segmentos = ['TOTAL', 'ETICO', 'OTC']
        # Períodos
        meses_actual = [(ref_y, ref_m)]
        meses_anterior = [_mes_anterior(ref_y, ref_m)]
        meses_mismo_ant = [(ref_y - 1, ref_m)]
        meses_ytd_actual = _meses_ytd(ref_y, ref_m)
        meses_ytd_anterior = _meses_ytd(ref_y - 1, ref_m)
        meses_mat_actual = _meses_rango(ref_y, ref_m, 12)
        # MAT anterior = los 12 meses inmediatamente anteriores al MAT actual
        y_prev, m_prev = _mes_anterior(*meses_mat_actual[-1])  # último mes - 1
        meses_mat_anterior = _meses_rango(y_prev, m_prev, 12)
        # Incremento año = YTD vs YTD pero solo en precios (mismo cálculo)

        # Resultados: dict[segmento] = dict[indicador] = dict[periodo] = valor
        # indicadores: 'U' (unidades), 'V' (valores $), 'P' (precio = V/U)
        res = defaultdict(lambda: defaultdict(dict))
        # mat_confiable_por_seg[seg] = bool — si el MAT anterior tiene
        # los 12 meses con datos (>=10 de 12 para ser tolerante).
        mat_confiable_por_seg = {}
        for seg in segmentos:
            obs_ids = droga_seg_ids[droga_id][seg]
            if not obs_ids:
                res[seg] = None
                mat_confiable_por_seg[seg] = False
                continue
            for periodo_n, meses in [
                ('act', meses_actual),
                ('mes_ant', meses_anterior),
                ('mismo_ant', meses_mismo_ant),
                ('ytd_act', meses_ytd_actual),
                ('ytd_ant', meses_ytd_anterior),
                ('mat_act', meses_mat_actual),
                ('mat_ant', meses_mat_anterior),
            ]:
                u, m_ = _agg(obs_ids, meses)
                res[seg]['U'][periodo_n] = u
                res[seg]['V'][periodo_n] = m_
                res[seg]['P'][periodo_n] = (m_ / u) if u else 0
            # MAT confiable solo si AMBOS períodos tienen al menos 10 de 12
            # meses con ventas. Si el sync arrancó mid-2024 o después, el
            # mat_ant tendrá gaps → marcamos n/d.
            mat_confiable_por_seg[seg] = (
                _meses_con_datos(obs_ids, meses_mat_actual) >= 10 and
                _meses_con_datos(obs_ids, meses_mat_anterior) >= 10
            )

        # ── Construir hoja ────────────────────────────────────────────
        # Título
        ws.merge_cells('A1:F1')
        parcial_lbl = ' (parcial)' if es_parcial else ''
        ws['A1'] = f'  Información Mercado     {MESES_ES[ref_m]}{parcial_lbl}   {ref_y}-'
        ws['A1'].font = font_title
        ws['A1'].fill = fill_title
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 28

        ws.merge_cells('A2:F2')
        ws['A2'] = droga_n
        ws['A2'].font = font_bold
        ws['A2'].alignment = Alignment(horizontal='center')

        # Header tabla
        HEAD = 4
        ws.cell(row=HEAD, column=1, value='Evolución  en %').fill = fill_header
        ws.cell(row=HEAD, column=1).font = font_header
        ws.cell(row=HEAD, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        headers_periodos = [
            'Mes Actual\nvs Mes\nAnterior',
            'Mismo\nMes año\nanterior',
            f'Acumulado\nAño {ref_y}/{ref_y-1}\n(ENERO-{MESES_ES[ref_m]})\n({ref_m}meses ) YTD',
            f'Acumulado\n12 Meses Móviles\n({MESES_ES[ref_m]} {ref_y}/ {MESES_ES[meses_mat_anterior[0][1]]} {meses_mat_anterior[0][0]}) MAT',
            f'Incremento\naño {ref_y} ({ref_m} meses)',
        ]
        for i, h in enumerate(headers_periodos, start=2):
            c = ws.cell(row=HEAD, column=i, value=h)
            c.fill = fill_header
            c.font = font_header
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border_all

        ws.row_dimensions[HEAD].height = 70
        ws.cell(row=HEAD, column=1).border = border_all
        for i in range(2, 7):
            ws.cell(row=HEAD, column=i).border = border_all

        # Función para escribir un bloque (Sección + 3 filas)
        def _escribir_bloque(row_inicio, titulo_seccion, indicador, incluir_incremento):
            # Fila titulo (sección)
            ws.cell(row=row_inicio, column=1, value=titulo_seccion).fill = fill_section
            ws.cell(row=row_inicio, column=1).font = font_section
            for i in range(2, 7):
                ws.cell(row=row_inicio, column=i).fill = fill_section
            for i in range(1, 7):
                ws.cell(row=row_inicio, column=i).border = border_all

            etiquetas = [('M Total', 'TOTAL'), ('ETICO', 'ETICO'), ('OTC', 'OTC')]
            for offset, (label, seg) in enumerate(etiquetas, start=1):
                r = row_inicio + offset
                ws.cell(row=r, column=1, value=label).font = font_bold
                ws.cell(row=r, column=1).fill = fill_data
                ws.cell(row=r, column=1).border = border_all

                seg_res = res.get(seg)
                if not seg_res:
                    for col in range(2, 7):
                        c = ws.cell(row=r, column=col, value='—')
                        c.fill = fill_data
                        c.border = border_all
                        c.alignment = Alignment(horizontal='center')
                    continue

                vals = seg_res[indicador]
                # 1) MoM
                v_mom = _pct(vals['act'], vals['mes_ant'])
                # 2) YoY mismo mes
                v_yoy = _pct(vals['act'], vals['mismo_ant'])
                # 3) YTD
                v_ytd = _pct(vals['ytd_act'], vals['ytd_ant'])
                # 4) MAT — solo si el período es confiable (sync completo
                #    en ambos años). Sino marcamos 'n/d' para no engañar.
                if mat_confiable_por_seg.get(seg):
                    v_mat = _pct(vals['mat_act'], vals['mat_ant'])
                else:
                    v_mat = 'n/d'
                # 5) Incremento año (solo precios) — usa YTD
                v_inc = None
                if incluir_incremento:
                    v_inc = _pct(vals['ytd_act'], vals['ytd_ant'])

                celdas = [v_mom, v_yoy, v_ytd, v_mat, v_inc]
                for col, val in enumerate(celdas, start=2):
                    if val is None:
                        c = ws.cell(row=r, column=col, value='—')
                        c.alignment = Alignment(horizontal='center')
                    elif val == 'n/d':
                        c = ws.cell(row=r, column=col, value='n/d')
                        c.alignment = Alignment(horizontal='center')
                        c.font = Font(italic=True, color='888888')
                    else:
                        c = ws.cell(row=r, column=col, value=val)
                        c.number_format = '0.0'
                        c.alignment = Alignment(horizontal='center')
                        if val < 0:
                            c.font = Font(color='C00000', bold=True)
                    c.fill = fill_data
                    c.border = border_all

        # Bloque 1: Venta Unidades
        _escribir_bloque(HEAD + 1, 'Venta Unidades', 'U', incluir_incremento=False)
        # Bloque 2: Venta Valores
        _escribir_bloque(HEAD + 5, 'Venta Valores $', 'V', incluir_incremento=False)
        # Bloque 3: Precios
        _escribir_bloque(HEAD + 9, 'Precios', 'P', incluir_incremento=True)

        # PVP promedio al cierre del mes de referencia
        pvp_start = HEAD + 14
        ws.cell(row=pvp_start, column=1,
                value=f'PVP promedio al cierre\nde {MESES_ES[ref_m]} {ref_y} (*)').font = font_bold
        ws.cell(row=pvp_start, column=1).fill = fill_section
        ws.cell(row=pvp_start, column=1).alignment = Alignment(wrap_text=True, vertical='center')
        ws.cell(row=pvp_start, column=1).border = border_all
        ws.cell(row=pvp_start, column=2,
                value='(*) No considera\nefecto PVP Pami').alignment = Alignment(wrap_text=True)
        ws.cell(row=pvp_start, column=2).fill = fill_data
        ws.cell(row=pvp_start, column=2).border = border_all
        ws.row_dimensions[pvp_start].height = 40

        for offset, (label, seg) in enumerate([('M Total', 'TOTAL'),
                                                ('ETICO', 'ETICO'),
                                                ('VTA LIBRE', 'OTC')], start=1):
            r = pvp_start + offset
            ws.cell(row=r, column=1, value=label).font = font_bold
            ws.cell(row=r, column=1).fill = fill_data
            ws.cell(row=r, column=1).border = border_all
            seg_res = res.get(seg)
            if seg_res:
                pvp = seg_res['P']['act']
                c = ws.cell(row=r, column=2, value=pvp)
                c.number_format = '#,##0.00'
            else:
                c = ws.cell(row=r, column=2, value='—')
                c.alignment = Alignment(horizontal='center')
            c.fill = fill_data
            c.border = border_all

        # Fila Fuente
        fuente_row = pvp_start + 4
        ws.cell(row=fuente_row, column=1, value='Fuente:').font = font_bold
        ws.cell(row=fuente_row, column=1).fill = fill_section
        ws.cell(row=fuente_row, column=2,
                value='ObServer y Datos propios').fill = fill_section
        ws.cell(row=fuente_row, column=1).border = border_all
        ws.cell(row=fuente_row, column=2).border = border_all

        # Nota sobre 'n/d' si aplica
        nd_aplica = any(not v for v in mat_confiable_por_seg.values()
                        if v is not None)
        next_row = fuente_row + 2
        if nd_aplica:
            ws.merge_cells(start_row=next_row, start_column=1,
                           end_row=next_row, end_column=6)
            ws.cell(row=next_row, column=1,
                    value='n/d = no disponible. El MAT compara 12 meses '
                    'móviles; el período anterior tiene gaps por sync '
                    'incompleto (Observer sincronizó desde feb 2025).')
            ws.cell(row=next_row, column=1).font = Font(italic=True, color='666666', size=9)
            ws.cell(row=next_row, column=1).alignment = Alignment(wrap_text=True)
            next_row += 2

        # ── TOP 5 productos por unidades YTD ──────────────────────────
        # Ranking con: producto | lab | U mes ref | U YTD | % share YTD
        prods_de_droga = droga_prods[droga_id]
        ranking = []
        for oid in prods_de_droga:
            u_act = ventas_por.get((oid, ref_y, ref_m), (0, 0))[0]
            u_ytd = sum(ventas_por.get((oid, y, m), (0, 0))[0]
                        for (y, m) in meses_ytd_actual)
            ranking.append((oid, u_act, u_ytd))
        ranking.sort(key=lambda x: -x[2])  # por YTD desc
        total_u_ytd = sum(r[2] for r in ranking) or 1

        top_titulo_row = next_row + 1
        ws.merge_cells(start_row=top_titulo_row, start_column=1,
                       end_row=top_titulo_row, end_column=6)
        ws.cell(row=top_titulo_row, column=1,
                value=f'  TOP 5 productos (YTD {MESES_ES[1]}-{MESES_ES[ref_m]} {ref_y})')
        ws.cell(row=top_titulo_row, column=1).font = font_title
        ws.cell(row=top_titulo_row, column=1).fill = fill_title
        ws.cell(row=top_titulo_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[top_titulo_row].height = 22

        head_top_row = top_titulo_row + 1
        top_headers = ['Producto', 'Laboratorio',
                       f'U {MESES_ES[ref_m]} {ref_y}',
                       'U YTD', '% share', '']
        for i, h in enumerate(top_headers[:-1], start=1):
            c = ws.cell(row=head_top_row, column=i, value=h)
            c.fill = fill_header
            c.font = font_header
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border_all

        for i, (oid, u_act, u_ytd) in enumerate(ranking[:5], start=1):
            r = head_top_row + i
            desc, lab_id = producto_meta.get(oid, ('?', None))
            lab = lab_nombre_global.get(lab_id, '—')
            is_buscado = oid in productos_buscados_obs_ids
            share = u_ytd / total_u_ytd * 100
            row_vals = [desc[:55], lab[:30], u_act, u_ytd, round(share, 1)]
            for col, val in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = border_all
                if is_buscado:
                    c.fill = PatternFill(start_color='FFF3C4', end_color='FFF3C4', fill_type='solid')
                    c.font = font_bold
                else:
                    c.fill = fill_data
                if col in (3, 4):
                    c.number_format = '#,##0'
                elif col == 5:
                    c.number_format = '0.0"%"'
                    c.alignment = Alignment(horizontal='center')

        # Si la monodroga tiene >5 productos, marcamos 'Otros' como resumen
        if len(ranking) > 5:
            otros_u_act = sum(r[1] for r in ranking[5:])
            otros_u_ytd = sum(r[2] for r in ranking[5:])
            otros_share = otros_u_ytd / total_u_ytd * 100
            r = head_top_row + 6
            row_vals = [f'... otros {len(ranking) - 5} productos', '',
                        otros_u_act, otros_u_ytd, round(otros_share, 1)]
            for col, val in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = border_all
                c.fill = fill_data
                c.font = Font(italic=True, color='666666')
                if col in (3, 4):
                    c.number_format = '#,##0'
                elif col == 5:
                    c.number_format = '0.0"%"'
                    c.alignment = Alignment(horizontal='center')

        # Anchos
        ws.column_dimensions['A'].width = 28
        for col in ('B', 'C', 'D', 'E', 'F'):
            ws.column_dimensions[col].width = 14

    # 7) Guardar
    out_dir = 'informes'
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    out_path = os.path.join(out_dir, f'informe_mercado_droga_{stamp}.xlsx')
    wb.save(out_path)
    print(f'\n✓ Informe generado: {out_path}')
    print(f'  Hojas: Índice + {len(drogas_set)} monodrogas (tarjeta tipo IQVIA)')


if __name__ == '__main__':
    sys.exit(main() or 0)
