"""Exportaciones de la lista de carga del robot (/rowa/carga) a XLSX y PDF.

La lista se recorre en el depósito laboratorio por laboratorio, así que las dos
exportaciones van **agrupadas por laboratorio con un título por grupo**, no como
una tabla plana: en papel, una columna "Laboratorio" repetida 300 veces no separa
nada. El laboratorio deja de ser columna y pasa a ser encabezado de sección.

Los items llegan ya ordenados y filtrados desde `routes/rowa.py`; acá no se
reordena ni se decide qué entra.
"""
from __future__ import annotations

import io
from datetime import datetime
from itertools import groupby

# Columnas, en orden. (clave del item, titulo, ancho_xlsx, ancho_pdf_cm, alineacion)
COLUMNAS = [
    ("nombre",         "Producto",  46, 6.6, "left"),
    ("ean",            "EAN",       16, 2.7, "left"),
    ("cantidad",       "Robot",      8, 1.5, "right"),
    ("stock_deposito", "Depósito",  10, 1.8, "right"),
    ("stock_total",    "Total",      8, 1.5, "right"),
    ("salidas_dia",    "Sal/día",    9, 1.6, "right"),
    ("cobertura",      "Cobertura", 10, 1.9, "right"),
    ("sug_cargar",     "Cargar",     9, 1.6, "right"),
]

# Cobertura 999 = sin salidas conocidas. Mostrar el numero seria mentir.
SIN_DATO = "—"


def _valor(item, clave):
    v = item.get(clave)
    if v is None:
        return SIN_DATO
    if clave == "cobertura":
        return SIN_DATO if v >= 999 else f"{v} d"
    if clave == "sug_cargar" and not v:
        return ""
    return v


def _agrupar(items):
    """[(laboratorio, [items...]), ...] respetando el orden que ya traen."""
    return [(lab or "Sin laboratorio", list(grupo))
            for lab, grupo in groupby(items, key=lambda i: i["laboratorio"])]


# ── XLSX ──────────────────────────────────────────────────────────────────
def construir_xlsx(items, generado: datetime | None = None) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    generado = generado or datetime.now()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carga robot"

    titulo = Font(bold=True, size=13)
    cab = Font(bold=True, color="FFFFFF")
    cab_fill = PatternFill("solid", start_color="1C1C1E", end_color="1C1C1E")
    lab_font = Font(bold=True, size=11, color="0B5D46")
    lab_fill = PatternFill("solid", start_color="D6F0E6", end_color="D6F0E6")
    der = Alignment(horizontal="right")

    ws.append([f"Carga del robot — {generado:%d/%m/%Y %H:%M}"])
    ws["A1"].font = titulo
    ws.append([f"{len(items)} artículos"])
    ws.append([])

    fila = ws.max_row + 1
    ws.append([c[1] for c in COLUMNAS])
    for cell in ws[fila]:
        cell.font = cab
        cell.fill = cab_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{fila + 1}"

    for lab, grupo in _agrupar(items):
        # Renglon de laboratorio: fusionado a lo ancho, para que se lea como
        # titulo de seccion y no como una fila mas de datos.
        fila = ws.max_row + 1
        ws.append([f"{lab}  ({len(grupo)})"])
        ws.merge_cells(start_row=fila, start_column=1,
                       end_row=fila, end_column=len(COLUMNAS))
        c = ws.cell(row=fila, column=1)
        c.font = lab_font
        c.fill = lab_fill

        for it in grupo:
            ws.append([_valor(it, clave) for clave, *_ in COLUMNAS])
            r = ws.max_row
            for col, (clave, _t, _w, _pw, alin) in enumerate(COLUMNAS, start=1):
                if alin == "right":
                    ws.cell(row=r, column=col).alignment = der

    for i, (_c, _t, ancho, *_r) in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[chr(64 + i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ───────────────────────────────────────────────────────────────────
def construir_pdf(items, generado: datetime | None = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        KeepTogether, Paragraph, SimpleDocTemplate, Table, TableStyle,
    )

    generado = generado or datetime.now()
    estilos = getSampleStyleSheet()
    st_titulo = ParagraphStyle("t", parent=estilos["Title"], fontSize=15, spaceAfter=2)
    st_sub = ParagraphStyle("s", parent=estilos["Normal"], fontSize=8,
                            textColor=colors.HexColor("#666666"), spaceAfter=10)
    st_lab = ParagraphStyle("l", parent=estilos["Heading2"], fontSize=10.5,
                            textColor=colors.HexColor("#0B5D46"),
                            spaceBefore=8, spaceAfter=3)
    st_celda = ParagraphStyle("c", parent=estilos["Normal"], fontSize=7.4, leading=9)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.2 * cm, bottomMargin=1.2 * cm,
        title="Carga del robot",
    )

    anchos = [c[3] * cm for c in COLUMNAS]
    encabezado = [c[1] for c in COLUMNAS]

    estilo_tabla = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C1C1E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.4),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7F5")]),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ])

    hist = [Paragraph("Carga del robot", st_titulo),
            Paragraph(f"{generado:%d/%m/%Y %H:%M} · {len(items)} artículos", st_sub)]

    for lab, grupo in _agrupar(items):
        datos = [encabezado]
        for it in grupo:
            fila = []
            for clave, _t, _w, _pw, _a in COLUMNAS:
                v = _valor(it, clave)
                # El nombre va como Paragraph para que corte de linea; el resto
                # es corto y en texto plano se alinea mejor.
                fila.append(Paragraph(str(v), st_celda) if clave == "nombre" else str(v))
            datos.append(fila)

        tabla = Table(datos, colWidths=anchos, repeatRows=1)
        tabla.setStyle(estilo_tabla)
        titulo = Paragraph(f"{lab} ({len(grupo)})", st_lab)
        if len(grupo) <= 12:
            # Grupo corto: titulo y tabla juntos, para no dejar el encabezado
            # de laboratorio huerfano al pie de una pagina.
            hist.append(KeepTogether([titulo, tabla]))
        else:
            # Grupo largo: no tiene sentido forzarlo a una pagina. reportlab lo
            # parte y `repeatRows=1` repite la fila de columnas en cada hoja.
            hist.append(titulo)
            hist.append(tabla)

    doc.build(hist)
    return buf.getvalue()
