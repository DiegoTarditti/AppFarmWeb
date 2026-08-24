"""Impresión de la planilla de carga del robot: PDF para caminar y XLSX.

Esto no es un informe: es una **orden de trabajo**. Alguien la imprime, agarra el
changuito y recorre el depósito. Todo lo de acá sale de esa premisa:

- **Agrupado por laboratorio, con el laboratorio como título de sección.** En
  papel, una columna "Laboratorio" repetida 300 veces no separa nada, y el
  depósito se camina un laboratorio a la vez.
- **Una casilla al principio de cada fila** para ir tildando. Sin eso, con 300
  renglones se pierde el hilo apenas suena el teléfono.
- **La cantidad sugerida grande y en negrita**, la del máximo chica al lado.
  Cuando hay un solo número en el papel, ese número se ejecuta: tiene que ser el
  que decidimos, no el que quedó de referencia.

Los avisos se abrevian porque en papel no hay tooltip: "VACÍO", "parcial",
"máx→N", "DIF". La explicación de cada uno va al pie, una sola vez.

Recibe las filas ya armadas por `services/rowa_planilla`; acá no se recalcula ni
se reordena nada.
"""
from __future__ import annotations

import io
from datetime import datetime
from itertools import groupby

# (clave, título, ancho_xlsx, ancho_pdf_cm, alineación)
COLUMNAS = [
    ("_check",     "",          3,  0.7, "center"),
    ("nombre",     "Producto",  44, 6.2, "left"),
    ("en_robot",   "Robot",      8, 1.3, "right"),
    ("maximo",     "Máx",        8, 1.2, "right"),
    ("deposito",   "Depós.",     9, 1.4, "right"),
    ("a_mover_sug", "A MOVER",  10, 1.7, "right"),
    ("a_mover_max", "Hasta máx",  8, 1.6, "right"),
    ("_avisos",    "Avisos",    26, 3.6, "left"),
]

SIN_DATO = "—"


def _avisos(f) -> str:
    """Avisos abreviados. En papel no hay tooltip: la explicación va al pie."""
    partes = []
    if f.get("vacio"):
        partes.append("VACÍO")
    if f.get("parcial"):
        partes.append("parcial")
    if f.get("corregir_a"):
        partes.append("máx→%s" % f["corregir_a"])
    if f.get("diferencia"):
        partes.append("DIF")
    if f.get("sin_maximo"):
        partes.append("sin máx")
    return " ".join(partes)


def _valor(f, clave):
    if clave == "_check":
        return ""
    if clave == "_avisos":
        return _avisos(f)
    v = f.get(clave)
    if v is None:
        return SIN_DATO
    if clave == "a_mover_max" and not v:
        return ""
    return v


def _agrupar(filas):
    """[(laboratorio, [filas...]), ...] respetando el orden que ya traen."""
    return [(lab or "Sin laboratorio", list(g))
            for lab, g in groupby(filas, key=lambda f: f.get("laboratorio") or "")]


def _encabezado(totales, generado):
    return "%s · %s artículos · %s packs a mover (si se llena al máximo: %s)" % (
        generado.strftime("%d/%m/%Y %H:%M"),
        totales.get("articulos", 0),
        totales.get("packs_sug", 0),
        totales.get("packs_max", 0),
    )


# ── XLSX ──────────────────────────────────────────────────────────────────
def construir_xlsx(filas, totales, generado: datetime | None = None) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    generado = generado or datetime.now()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carga robot"

    f_titulo = Font(bold=True, size=13)
    f_cab = Font(bold=True, color="FFFFFF")
    fill_cab = PatternFill("solid", start_color="1C1C1E", end_color="1C1C1E")
    f_lab = Font(bold=True, size=11, color="0B5D46")
    fill_lab = PatternFill("solid", start_color="D6F0E6", end_color="D6F0E6")
    f_mover = Font(bold=True, size=12)
    fill_mover = PatternFill("solid", start_color="E8F5EF", end_color="E8F5EF")
    der = Alignment(horizontal="right")
    centro = Alignment(horizontal="center")
    caja = Border(*[Side(style="thin", color="BBBBBB")] * 4)

    ws.append(["Planilla de carga del robot"])
    ws["A1"].font = f_titulo
    ws.append([_encabezado(totales, generado)])
    ws.append([])

    fila = ws.max_row + 1
    ws.append([c[1] for c in COLUMNAS])
    for celda in ws[fila]:
        celda.font = f_cab
        celda.fill = fill_cab
        celda.alignment = centro
    ws.freeze_panes = "A%d" % (fila + 1)
    col_mover = [c[0] for c in COLUMNAS].index("a_mover_sug") + 1

    for lab, grupo in _agrupar(filas):
        fila = ws.max_row + 1
        ws.append(["%s  (%d)" % (lab, len(grupo))])
        ws.merge_cells(start_row=fila, start_column=1,
                       end_row=fila, end_column=len(COLUMNAS))
        c = ws.cell(row=fila, column=1)
        c.font, c.fill = f_lab, fill_lab

        for f in grupo:
            ws.append([_valor(f, clave) for clave, *_ in COLUMNAS])
            r = ws.max_row
            for col, (_clave, _t, _w, _pw, alin) in enumerate(COLUMNAS, start=1):
                if alin == "right":
                    ws.cell(row=r, column=col).alignment = der
                elif alin == "center":
                    ws.cell(row=r, column=col).alignment = centro
            # La casilla para tildar: borde en una celda vacía.
            ws.cell(row=r, column=1).border = caja
            # La cifra que se ejecuta, destacada también en Excel.
            m = ws.cell(row=r, column=col_mover)
            m.font, m.fill = f_mover, fill_mover

    for i, (_c, _t, ancho, *_r) in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[chr(64 + i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ───────────────────────────────────────────────────────────────────
def construir_pdf(filas, totales, generado: datetime | None = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        KeepTogether,
        Paragraph,
        SimpleDocTemplate,
        Table,
        TableStyle,
    )

    generado = generado or datetime.now()
    est = getSampleStyleSheet()
    st_titulo = ParagraphStyle("t", parent=est["Title"], fontSize=15, spaceAfter=2)
    st_sub = ParagraphStyle("s", parent=est["Normal"], fontSize=8,
                            textColor=colors.HexColor("#666666"), spaceAfter=10)
    st_lab = ParagraphStyle("l", parent=est["Heading2"], fontSize=10.5,
                            textColor=colors.HexColor("#0B5D46"),
                            spaceBefore=8, spaceAfter=3)
    st_celda = ParagraphStyle("c", parent=est["Normal"], fontSize=7.4, leading=9)
    st_pie = ParagraphStyle("p", parent=est["Normal"], fontSize=7,
                            textColor=colors.HexColor("#666666"), spaceBefore=12)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.2 * cm, bottomMargin=1.2 * cm,
        title="Planilla de carga del robot",
    )

    anchos = [c[3] * cm for c in COLUMNAS]
    cab = [c[1] for c in COLUMNAS]
    i_check = [c[0] for c in COLUMNAS].index("_check")
    i_mover = [c[0] for c in COLUMNAS].index("a_mover_sug")

    estilo = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C1C1E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.4),
        ("ALIGN", (2, 0), (-2, -1), "RIGHT"),
        ("ALIGN", (i_check, 0), (i_check, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7F5")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        # La cifra a ejecutar: más grande, en negrita y sobre fondo propio. Es lo
        # único que el operador tiene que leer rápido mientras camina.
        ("FONTNAME", (i_mover, 1), (i_mover, -1), "Helvetica-Bold"),
        ("FONTSIZE", (i_mover, 1), (i_mover, -1), 10),
        ("BACKGROUND", (i_mover, 1), (i_mover, -1), colors.HexColor("#E8F5EF")),
    ])

    hist = [Paragraph("Planilla de carga del robot", st_titulo),
            Paragraph(_encabezado(totales, generado), st_sub)]

    for lab, grupo in _agrupar(filas):
        datos = [cab]
        for f in grupo:
            fila = []
            for clave, _t, _w, _pw, _a in COLUMNAS:
                v = _valor(f, clave)
                # Nombre y avisos como Paragraph para que corten de línea; el
                # resto es corto y en texto plano se alinea mejor.
                fila.append(Paragraph(str(v), st_celda)
                            if clave in ("nombre", "_avisos") else str(v))
            datos.append(fila)

        tabla = Table(datos, colWidths=anchos, repeatRows=1)
        tabla.setStyle(estilo)
        titulo = Paragraph("%s (%d)" % (lab, len(grupo)), st_lab)
        if len(grupo) <= 12:
            # Grupo corto: título y tabla juntos, para no dejar el encabezado de
            # laboratorio huérfano al pie de una página.
            hist.append(KeepTogether([titulo, tabla]))
        else:
            # Grupo largo: no tiene sentido forzarlo a una página. reportlab lo
            # parte y `repeatRows=1` repite la fila de columnas en cada hoja.
            hist.append(titulo)
            hist.append(tabla)

    hist.append(Paragraph(
        "<b>A MOVER</b> es la cantidad sugerida por la demanda — es la que hay que "
        "ejecutar. <b>Hasta máx</b> es lo que pediría llenar hasta la Cantidad Máxima de "
        "ObServer, y va sólo de referencia. Las dos están limitadas por lo que haya en "
        "el depósito.<br/>"
        "<b>VACÍO</b>: no hay nada de ese artículo en el robot. &nbsp;"
        "<b>parcial</b>: el depósito no alcanza para llenar el hueco; se mueve lo que "
        "hay. Si se repite, el mínimo de compra quedó corto. &nbsp;"
        "<b>máx→N</b>: el máximo de ObServer no se justifica con la demanda; conviene "
        "bajarlo a N. &nbsp;"
        "<b>DIF</b>: el robot tiene más packs que el stock de ObServer, así que el "
        "depósito de esa fila no es confiable.", st_pie))

    doc.build(hist)
    return buf.getvalue()
