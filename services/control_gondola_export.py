"""Export de la planilla de Control de stock por laboratorio (PDF y XLSX).

Columnas: Producto · En robot · En depósito · Total · Contado (en blanco, para
anotar a mano al ir a contar). Mismo estilo que services/rowa_planilla_export.
"""
from __future__ import annotations

import io
from datetime import datetime

_FILTRO_LABEL = {
    'con_stock': 'Con stock',
    'solo_robot': 'Solo en robot',
    'solo_deposito': 'Solo en depósito',
}

# (clave, encabezado, ancho_xlsx, ancho_pdf_cm, alineación)
COLUMNAS = [
    ('nombre',   'Producto',     46, 8.6, 'left'),
    ('en_robot', 'En robot',     10, 2.0, 'right'),
    ('deposito', 'En depósito',  11, 2.2, 'right'),
    ('total',    'Total',         9, 1.8, 'right'),
    ('_contado', 'Contado',      11, 2.4, 'center'),
]


def _valor(f, clave):
    if clave == '_contado':
        return ''
    v = f.get(clave)
    if clave in ('en_robot', 'deposito'):
        return v if v else '—'
    return v if v is not None else ''


def _titulo(lab, filtro):
    return f'Control de stock — {lab} ({_FILTRO_LABEL.get(filtro, filtro)})'


def construir_xlsx(filas, lab, filtro, generado: datetime | None = None) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    generado = generado or datetime.now()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Control stock'

    ws.append([_titulo(lab, filtro)])
    ws.append([f'Generado {generado.strftime("%d/%m/%Y %H:%M")} · {len(filas)} artículo(s)'])
    ws.append([])
    hdr_row = ws.max_row + 1
    ws.append([c[1] for c in COLUMNAS])

    thin = Side(style='thin', color='CCCCCC')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, len(COLUMNAS) + 1):
        c = ws.cell(row=hdr_row, column=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1C1C1E')
        c.border = borde
        c.alignment = Alignment(horizontal='center')

    for f in filas:
        ws.append([_valor(f, clave) for clave, *_ in COLUMNAS])
        r = ws.max_row
        for col, (_clave, _t, _w, _pw, alin) in enumerate(COLUMNAS, start=1):
            cell = ws.cell(row=r, column=col)
            cell.border = borde
            cell.alignment = Alignment(horizontal=alin)

    for i, (_c, _t, ancho, *_r) in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def construir_pdf(filas, lab, filtro, generado: datetime | None = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

    generado = generado or datetime.now()
    est = getSampleStyleSheet()
    st_titulo = ParagraphStyle('t', parent=est['Title'], fontSize=14, spaceAfter=2)
    st_sub = ParagraphStyle('s', parent=est['Normal'], fontSize=8,
                            textColor=colors.HexColor('#666666'), spaceAfter=10)
    st_celda = ParagraphStyle('c', parent=est['Normal'], fontSize=7.6, leading=9)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.2 * cm, rightMargin=1.2 * cm,
                            topMargin=1.2 * cm, bottomMargin=1.2 * cm,
                            title='Control de stock por laboratorio')

    anchos = [c[3] * cm for c in COLUMNAS]
    cab = [c[1] for c in COLUMNAS]
    datos = [cab]
    for f in filas:
        fila = []
        for clave, _t, _w, _pw, _a in COLUMNAS:
            v = _valor(f, clave)
            fila.append(Paragraph(str(v), st_celda) if clave == 'nombre' else str(v))
        datos.append(fila)

    tabla = Table(datos, colWidths=anchos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1C1C1E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.6),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (-1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F7F5')]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    story = [Paragraph(_titulo(lab, filtro), st_titulo),
             Paragraph(f'Generado {generado.strftime("%d/%m/%Y %H:%M")} · '
                       f'{len(filas)} artículo(s)', st_sub),
             tabla]
    doc.build(story)
    return buf.getvalue()
