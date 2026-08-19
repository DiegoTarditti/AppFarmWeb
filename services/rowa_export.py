"""Genera las 3 planillas del robot Rowa (Stock, Optimización, Capacidad) en vivo.

Reemplaza el export manual de Lisandro: se conecta al robot, analiza y escribe un
.xlsx con el mismo formato. Se puede correr suelto o llamar desde el blueprint web.

    PYTHONPATH=. python -m services.rowa_export --out Rowa-en-vivo.xlsx
"""

from __future__ import annotations

import argparse
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.rowa_analisis import analizar_alturas, analizar_stock
from services.rowa_client import RowaClient

_AZUL = "FF1F4E79"
_HDR_FONT = Font(bold=True, color="FFFFFFFF")
_HDR_FILL = PatternFill("solid", fgColor=_AZUL)
_TITULO = Font(bold=True, size=13)


def _encabezar(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _autoancho(ws, anchos: list[int]) -> None:
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def construir_workbook(filas, diag, alturas) -> Workbook:
    wb = Workbook()

    # -- Hoja 1: Diagnóstico -------------------------------------------
    ws = wb.active
    ws.title = "Diagnostico"
    ws["A1"] = "OPTIMIZACION DEL ROBOT ROWA"
    ws["A1"].font = _TITULO
    orden_rot = ["Alta", "Media", "Baja", "Durmiente"]
    resumen = [
        ("Generado", diag.generado.strftime("%Y-%m-%d")),
        ("Articulos en robot", diag.articulos),
        ("Unidades (packs) totales", diag.packs),
        ("Volumen ocupado (L)", diag.vol_ocupado_l),
        ("", ""),
        ("POR ROTACION (proxy por antiguedad de packs)", ""),
    ]
    for rot in orden_rot:
        resumen.append((f"  {rot}", diag.por_rotacion.get(rot, 0)))
    resumen += [
        ("", ""),
        ("OPORTUNIDAD", ""),
        ("  Articulos con accion sugerida", diag.con_accion),
        ("  Espacio recuperable al deposito (L)", diag.espacio_recuperable_l),
        ("", ""),
        ("NOTA", "Rotacion y unid/mes son proxy por antiguedad."),
        ("", "Se afinan al cruzar con ventas reales de ObServer."),
    ]
    for i, (k, v) in enumerate(resumen, start=3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    _autoancho(ws, [46, 40])

    # -- Hoja 2: Recomendaciones ---------------------------------------
    ws = wb.create_sheet("Recomendaciones")
    cols = ["ArticleId", "EAN", "Nombre", "Cantidad", "RotacionProxy", "UnidMesEst",
            "AntigProm_d", "AntigMax_d", "VolUnit_cm3", "VolTotal_cm3", "ProxVenc",
            "DiasProxVenc", "Recomendacion", "SugEnRobot", "AlDeposito"]
    ws.append(cols)
    for f in sorted(filas, key=lambda x: (x.nombre or "").lower()):
        ws.append([
            f.article_id, f.ean, f.nombre, f.cantidad, f.rotacion, f.unid_mes_est,
            f.antig_prom_d, f.antig_max_d, f.vol_unit_cm3, f.vol_total_cm3,
            f.prox_venc.strftime("%Y-%m-%d") if f.prox_venc else None,
            f.dias_prox_venc, f.recomendacion, f.sug_en_robot, f.al_deposito,
        ])
    _encabezar(ws, len(cols))
    _autoancho(ws, [10, 15, 30, 9, 13, 10, 11, 10, 11, 12, 12, 12, 30, 11, 11])

    # -- Hoja 3: Sacar/Reducir (solo los con acción) -------------------
    ws = wb.create_sheet("Sacar-Reducir del robot")
    cols3 = ["ArticleId", "EAN", "Nombre", "Cantidad", "RotacionProxy", "AntigMax_d",
             "ProxVenc", "VolUnit_cm3", "EspacioLibera_cm3", "SugEnRobot",
             "AlDeposito", "Recomendacion"]
    ws.append(cols3)
    accion = [f for f in filas if f.al_deposito > 0]
    for f in sorted(accion, key=lambda x: x.al_deposito * x.vol_unit_cm3, reverse=True):
        ws.append([
            f.article_id, f.ean, f.nombre, f.cantidad, f.rotacion, f.antig_max_d,
            f.prox_venc.strftime("%Y-%m-%d") if f.prox_venc else None,
            f.vol_unit_cm3, round(f.al_deposito * f.vol_unit_cm3, 1),
            f.sug_en_robot, f.al_deposito, f.recomendacion,
        ])
    _encabezar(ws, len(cols3))
    _autoancho(ws, [10, 15, 30, 9, 13, 10, 12, 11, 16, 11, 11, 30])

    # -- Hoja 4: Vencimientos próximos ---------------------------------
    ws = wb.create_sheet("Vencimientos")
    cols4 = ["ArticleId", "EAN", "Nombre", "Cantidad", "ProxVenc", "DiasProxVenc",
             "PacksEnAlerta", "Recomendacion"]
    ws.append(cols4)
    venc = [f for f in filas if f.dias_prox_venc is not None and f.packs_venc_alerta]
    for f in sorted(venc, key=lambda x: x.dias_prox_venc):
        ws.append([f.article_id, f.ean, f.nombre, f.cantidad,
                   f.prox_venc.strftime("%Y-%m-%d"), f.dias_prox_venc,
                   f.packs_venc_alerta, f.recomendacion])
    _encabezar(ws, len(cols4))
    _autoancho(ws, [10, 15, 30, 9, 12, 12, 13, 30])

    # -- Hoja 5: Capacidad-Alturas -------------------------------------
    ws = wb.create_sheet("Capacidad-Alturas")
    r = 1
    ws.cell(row=r, column=1, value="ROWA - Ganar capacidad densificando estantes por altura").font = _TITULO
    r += 1
    ws.cell(row=r, column=1,
            value=(f"Packs {alturas['packs']} | altura prom {alturas['altura_prom']}mm / "
                   f"mediana {alturas['altura_mediana']}mm / max {alturas['altura_max']}mm | "
                   f"margen (grosor estante) {alturas['margen_mm']}mm")); r += 2
    ws.cell(row=r, column=1, value="DISTRIBUCION DE ALTURA").font = Font(bold=True); r += 1
    ws.append(["Rango", "Packs", "%"])
    for d in alturas["distribucion"]:
        ws.append([d["rango"], d["packs"], d["pct"]])
    r += 1 + len(alturas["distribucion"]) + 1
    ws.cell(row=r, column=1, value="RENDIMIENTO SEGUN SEPARACION DE ESTANTES").font = Font(bold=True); r += 1
    ws.append(["Canal (mm)", "Cubre stock %", "Filas x metro", "vs 80mm"])
    for rr in alturas["rendimiento"]:
        ws.append([rr["canal_mm"], rr["cubre_pct"], rr["filas_x_metro"], rr.get("vs_80mm")])
    _autoancho(ws, [16, 16, 16, 12])

    return wb


def generar(out_path: str, host: str | None = None, port: int | None = None) -> dict:
    kwargs = {}
    if host:
        kwargs["host"] = host
    if port:
        kwargs["port"] = port
    with RowaClient(**kwargs) as r:
        info = r.robot_info
        arts = r.stock_info(include_packs=True)
    filas, diag = analizar_stock(arts)
    alturas = analizar_alturas(arts)
    wb = construir_workbook(filas, diag, alturas)
    wb.save(out_path)
    return {
        "out": out_path,
        "robot": f"{info.get('ProductInfo')} v{info.get('VersionInfo')}",
        "articulos": diag.articulos,
        "packs": diag.packs,
        "con_accion": diag.con_accion,
        "recuperable_l": diag.espacio_recuperable_l,
    }


def _main() -> None:
    ap = argparse.ArgumentParser(description="Genera las planillas del robot Rowa en vivo")
    ap.add_argument("--out", default=f"Rowa-en-vivo-{datetime.now():%Y-%m-%d}.xlsx")
    ap.add_argument("--host")
    ap.add_argument("--port", type=int)
    args = ap.parse_args()
    res = generar(args.out, host=args.host, port=args.port)
    print(f"OK -> {res['out']}")
    for k, v in res.items():
        if k != "out":
            print(f"  {k}: {v}")


if __name__ == "__main__":
    _main()
