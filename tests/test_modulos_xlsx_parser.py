"""Tests del parser de Excel de módulos (`parsers/modulos_xlsx.py`).

Foco: el "Formato C" (layout combo con encabezados nombrados + filas
separadoras 'Combo N'), agregado el 2026-05-27 para el export de combos de
Pharmadorf. El importador viejo (columnas fijas Formato A) leía mal estos
archivos: tomaba la descripción como nombre de módulo, el descuento como
cantidad y el mínimo como descuento, y metía la fila de títulos como un
módulo basura. Estos tests fijan el parseo correcto y cuidan que A/B no se
rompan.
"""
import openpyxl

from parsers.modulos_xlsx import _detect_format, parse_modulos_xlsx


def _mkxlsx(tmp_path, rows, name='m.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    p = tmp_path / name
    wb.save(p)
    return str(p)


# ── Formato C: layout combo Pharmadorf ───────────────────────────────────────

# Encabezado real del archivo: la descripción va en col 0, el combo aparece mal
# ubicado bajo 'CANT. PEDIDA' y la columna 'COMBO' queda vacía. El agrupado se
# hace por las filas separadoras 'Combo N' (sin EAN).
_COMBO_ROWS = [
    ['DESCRIPCION', 'EAN', 'ID', 'DESC.', 'MIN O FIJO', 'PLAZO', 'CANT. PEDIDA', 'COMBO'],
    ['Combo 1'],
    ['EDICTUM 0,5 X 60 COMP', '7798103391264', '194646', 20, 2, 'HABITUAL', 'Combo 1', None],
    ['IRIX COLIRIO 15ML',     '7798103391165', '194650', 18, 5, 'HABITUAL', 'Combo 1', None],
    ['Combo 2'],
    ['IRIX COLIRIO 15ML',     '7798103391165', '194652', 18, 5, 'HABITUAL', 'Combo 2', None],
    ['Combo 3'],
    ['CREMOQUINONA CICATRICES 30G', '7702195930676', '276360', 15, 2, 'HABITUAL', 'Combo 3', None],
    ['IRIX COLIRIO 15ML',     '7798103391165', '276358', 20, 5, 'HABITUAL', 'Combo 3', None],
]


class TestFormatoCombo:

    def test_detecta_formato_c(self, tmp_path):
        p = _mkxlsx(tmp_path, _COMBO_ROWS)
        ws = openpyxl.load_workbook(p, data_only=True).active
        assert _detect_format(ws) == 'C'

    def test_agrupa_por_combo(self, tmp_path):
        mods = parse_modulos_xlsx(_mkxlsx(tmp_path, _COMBO_ROWS))
        nombres = {m['nombre']: len(m['items']) for m in mods}
        assert nombres == {'Combo 1': 2, 'Combo 2': 1, 'Combo 3': 2}

    def test_mapea_cantidad_y_descuento_por_nombre(self, tmp_path):
        """cant <- MIN O FIJO, desc_pct <- DESC. (no al revés, no CANT. PEDIDA)."""
        mods = parse_modulos_xlsx(_mkxlsx(tmp_path, _COMBO_ROWS))
        c1 = next(m for m in mods if m['nombre'] == 'Combo 1')
        edictum = next(it for it in c1['items'] if it['ean'] == '7798103391264')
        assert edictum['cant'] == 2          # MIN O FIJO
        assert edictum['desc_pct'] == 20.0   # DESC.
        assert edictum['descripcion'] == 'EDICTUM 0,5 X 60 COMP'

    def test_no_crea_modulo_basura_del_header(self, tmp_path):
        mods = parse_modulos_xlsx(_mkxlsx(tmp_path, _COMBO_ROWS))
        assert all(m['nombre'] not in ('DESCRIPCION', 'DESCRIPCIÓN') for m in mods)

    def test_mismo_ean_en_varios_combos(self, tmp_path):
        """IRIX repetido en los 3 combos: una entrada por combo, valor propio."""
        mods = parse_modulos_xlsx(_mkxlsx(tmp_path, _COMBO_ROWS))
        c3 = next(m for m in mods if m['nombre'] == 'Combo 3')
        irix = next(it for it in c3['items'] if it['ean'] == '7798103391165')
        assert irix['desc_pct'] == 20.0  # en Combo 1/2 es 18, en Combo 3 es 20


# ── Regresión: A y B siguen detectándose ─────────────────────────────────────

class TestNoRegresion:

    def test_formato_a(self, tmp_path):
        p = _mkxlsx(tmp_path, [
            ['NOMBRE MÓDULO', 'EAN', 'DESCRIPCIÓN', 'CANT. MÓDULO', 'DESC. %'],
            ['MOD EJ', None, None, None, None],
            [None, '7790001000001', 'AMOXI 500 x16', 10, 7.0],
        ])
        ws = openpyxl.load_workbook(p, data_only=True).active
        assert _detect_format(ws) == 'A'
        mods = parse_modulos_xlsx(p)
        assert len(mods) == 1 and len(mods[0]['items']) == 1

    def test_formato_b(self, tmp_path):
        p = _mkxlsx(tmp_path, [
            [None, 'Titulo', None, None, None, None],
            ['COD MOD.', 'NOMBRE MODULO', 'CODIGO EAN', 'DESCRIPCION', 'CANT', 'DESC %'],
            ['M1', 'MOD ROEM', '7790002000001', 'IBU 400', 12, 10.0],
        ])
        ws = openpyxl.load_workbook(p, data_only=True).active
        assert _detect_format(ws) == 'B'
        mods = parse_modulos_xlsx(p)
        assert len(mods) == 1 and len(mods[0]['items']) == 1
